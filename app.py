import os
import json
import re
import uuid
import csv
from io import StringIO, BytesIO
import xml.etree.ElementTree as ET
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation
from functools import wraps

from danfse_parser import parse_danfse_xml

import mysql.connector
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, Response, send_file
from werkzeug.security import generate_password_hash, check_password_hash

from config import Config
from database import obter_conexao

try:
    from google_drive_storage import (
        GoogleDriveStorageError,
        categoria_por_origem,
        google_drive_habilitado,
        registrar_arquivo_sistema,
        upload_arquivo_path_google_drive,
        baixar_arquivo_google_drive,
        obter_metadados_arquivo_google_drive,
    )
except Exception as _gdrive_import_error:
    GoogleDriveStorageError = Exception

    def google_drive_habilitado():
        return False

    def categoria_por_origem(origem):
        return "Outros"

    def registrar_arquivo_sistema(*args, **kwargs):
        return None

    def upload_arquivo_path_google_drive(*args, **kwargs):
        raise RuntimeError(f"Google Drive indisponível: {_gdrive_import_error}")

    def baixar_arquivo_google_drive(*args, **kwargs):
        raise RuntimeError(f"Google Drive indisponível: {_gdrive_import_error}")

    def obter_metadados_arquivo_google_drive(*args, **kwargs):
        raise RuntimeError(f"Google Drive indisponível: {_gdrive_import_error}")

app = Flask(__name__)
app.config.from_object(Config)

# Permite envio da selfie comprimida no formulário de confirmação de chegada.
# A imagem é reduzida no navegador antes do envio.
app.config['MAX_CONTENT_LENGTH'] = 4 * 1024 * 1024


# ==========================================================
# FILTROS E FUNÇÕES AUXILIARES
# ==========================================================
@app.template_filter('formatar_data_br')
def formatar_data_br(valor):
    if not valor:
        return '-'

    try:
        if isinstance(valor, datetime):
            return valor.strftime('%d/%m/%Y')

        if isinstance(valor, date):
            return valor.strftime('%d/%m/%Y')

        if hasattr(valor, 'strftime'):
            return valor.strftime('%d/%m/%Y')

        valor_str = str(valor).strip()
        if not valor_str:
            return '-'

        formatos = [
            '%Y-%m-%d',
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y-%m-%dT%H:%M:%S'
        ]

        for formato in formatos:
            try:
                data = datetime.strptime(valor_str[:19], formato)
                return data.strftime('%d/%m/%Y')
            except Exception:
                pass

        return valor_str

    except Exception:
        return str(valor)


@app.template_filter('formatar_data_hora_br')
def formatar_data_hora_br(valor):
    if not valor:
        return '-'

    try:
        # Se vier como datetime do MySQL
        if hasattr(valor, 'strftime'):
            return valor.strftime('%d/%m/%Y %H:%M')

        # Se vier como string: 2026-06-28 21:41:47
        valor_str = str(valor).strip()

        from datetime import datetime

        formatos = [
            '%Y-%m-%d %H:%M:%S',
            '%Y-%m-%d %H:%M',
            '%Y-%m-%dT%H:%M:%S',
            '%Y-%m-%d'
        ]

        for formato in formatos:
            try:
                data = datetime.strptime(valor_str[:19], formato)
                return data.strftime('%d/%m/%Y %H:%M')
            except Exception:
                pass

        return valor_str

    except Exception:
        return str(valor)


@app.template_filter('moeda_br')
def moeda_br(valor):
    try:
        numero = Decimal(str(valor or 0)).quantize(Decimal('0.01'))
    except Exception:
        numero = Decimal('0.00')

    texto = f"{numero:,.2f}"
    texto = texto.replace(',', 'X').replace('.', ',').replace('X', '.')
    return f"R$ {texto}"




def normalizar_horario_input(valor):
    """
    Normaliza valores de horário vindos do MySQL para uso em <input type="time">.
    O MySQL Connector pode retornar TIME como datetime.timedelta, e o HTML exige HH:MM.
    """
    if valor is None or valor == "":
        return ""

    try:
        if isinstance(valor, timedelta):
            total_segundos = int(valor.total_seconds())
            if total_segundos < 0:
                total_segundos = 0
            horas = (total_segundos // 3600) % 24
            minutos = (total_segundos % 3600) // 60
            return f"{horas:02d}:{minutos:02d}"

        if hasattr(valor, 'strftime'):
            return valor.strftime('%H:%M')

        texto = str(valor).strip()
        if not texto:
            return ""

        partes = texto.split(':')
        if len(partes) >= 2:
            horas = int(partes[0])
            minutos = int(partes[1])
            return f"{horas:02d}:{minutos:02d}"

        return texto
    except Exception:
        return ""


def somente_digitos(valor):
    return ''.join(filter(str.isdigit, str(valor or '')))


def limitar_texto(valor, limite):
    valor = str(valor or '').strip()
    return valor[:limite]


def converter_decimal(valor):
    try:
        if valor is None or str(valor).strip() == '':
            return Decimal('0.00')

        valor_limpo = str(valor).strip().replace('R$', '').replace(' ', '')

        # Padrão brasileiro: 1.234,56
        if ',' in valor_limpo:
            valor_limpo = valor_limpo.replace('.', '').replace(',', '.')

        return Decimal(valor_limpo).quantize(Decimal('0.01'))
    except (InvalidOperation, ValueError):
        return Decimal('0.00')


def valor_decimal_form(valor):
    return converter_decimal(valor)



def arquivo_url(caminho):
    """
    Retorna URL de visualização para arquivos do SGR.

    Regras:
    - Links http/https continuam iguais.
    - /arquivos/visualizar/<id> continua sendo usado para Google Drive/arquivos_sistema.
    - uploads/... usa uma rota local protegida por login. Isso evita links relativos como
      /financeiro/titulos/uploads/... e permite abrir comprovantes salvos no fallback local.
    - static/... continua usando /static.
    """
    caminho = str(caminho or '').strip()
    if not caminho:
        return ''

    if caminho.startswith(('http://', 'https://')):
        return caminho

    if caminho.startswith('/arquivos/visualizar/') or caminho.startswith('/arquivos/local/'):
        return caminho

    if caminho.startswith('/uploads/'):
        caminho = caminho.lstrip('/')
        try:
            return url_for('visualizar_upload_local', caminho=caminho)
        except Exception:
            return '/' + caminho

    if caminho.startswith('uploads/'):
        try:
            return url_for('visualizar_upload_local', caminho=caminho)
        except Exception:
            return '/' + caminho

    if caminho.startswith('/static/'):
        return caminho
    if caminho.startswith('static/'):
        return '/' + caminho

    if caminho.startswith('/'):
        return caminho

    try:
        return url_for('static', filename=caminho)
    except Exception:
        return caminho


@app.template_filter('arquivo_url')
def arquivo_url_filter(caminho):
    return arquivo_url(caminho)


def _buscar_contexto_drive_upload(cur, empresa_id, motorista_id=None):
    """Busca nomes amigáveis para organizar pastas no Google Drive."""
    empresa_nome = f"Empresa_{empresa_id}"
    motorista_nome = None

    try:
        cur.execute("""
            SELECT COALESCE(NULLIF(nome_fantasia, ''), NULLIF(razao_social, ''), CONCAT('Empresa_', id)) AS nome_empresa
            FROM empresas
            WHERE id = %s
            LIMIT 1
        """, (empresa_id,))
        row = cur.fetchone() or {}
        empresa_nome = row.get('nome_empresa') or empresa_nome
    except Exception:
        pass

    if motorista_id:
        try:
            cur.execute("""
                SELECT COALESCE(NULLIF(apelido, ''), NULLIF(nome_completo, ''), CONCAT('Motorista_', id)) AS nome_motorista
                FROM pessoas
                WHERE id = %s
                  AND empresa_id = %s
                LIMIT 1
            """, (motorista_id, empresa_id))
            row = cur.fetchone() or {}
            motorista_nome = row.get('nome_motorista') or f"Motorista_{motorista_id}"
        except Exception:
            motorista_nome = f"Motorista_{motorista_id}"

    return empresa_nome, motorista_nome


def tentar_enviar_arquivo_google_drive(
    cur,
    caminho_absoluto,
    caminho_relativo,
    *,
    empresa_id,
    motorista_id=None,
    origem,
    origem_id=None,
    tipo_arquivo,
    nome_original=None,
    mime_type=None,
    criado_por_usuario_id=None,
):
    """
    Envia arquivo para o Google Drive quando habilitado.
    Se falhar, mantém o caminho local para não interromper o fluxo operacional.
    """
    if not google_drive_habilitado():
        print(f"[Google Drive] Upload não realizado para {origem}/{origem_id}: Google Drive desabilitado ou credenciais não localizadas. Usando fallback local: {caminho_relativo}")
        return caminho_relativo

    try:
        empresa_nome, motorista_nome = _buscar_contexto_drive_upload(cur, empresa_id, motorista_id)
        upload_info = upload_arquivo_path_google_drive(
            caminho_local=caminho_absoluto,
            empresa_id=empresa_id,
            empresa_nome=empresa_nome,
            categoria=categoria_por_origem(origem),
            origem=origem,
            origem_id=origem_id,
            motorista_id=motorista_id,
            motorista_nome=motorista_nome,
            nome_original=nome_original or os.path.basename(caminho_absoluto),
            mime_type=mime_type,
        )
        arquivo_id = registrar_arquivo_sistema(
            cur,
            empresa_id=empresa_id,
            pessoa_id=motorista_id,
            motorista_id=motorista_id,
            origem=origem,
            origem_id=origem_id,
            tipo_arquivo=tipo_arquivo,
            upload_info=upload_info,
            caminho_local=caminho_relativo,
            status_arquivo='ATIVO',
            criado_por_usuario_id=criado_por_usuario_id or session.get('usuario_id'),
        )

        # Retornamos uma rota interna protegida do SGR, não o link direto do Drive.
        # Assim o arquivo continua privado no Google Drive e só usuários logados,
        # com permissão na empresa/motorista, conseguem visualizar.
        if arquivo_id:
            try:
                return url_for('visualizar_arquivo_sistema', arquivo_id=arquivo_id)
            except Exception:
                return f"/arquivos/visualizar/{arquivo_id}"

        return caminho_relativo
    except Exception as exc:
        print(f"[Google Drive] Falha no upload de {origem}/{origem_id}: {exc}")
        return caminho_relativo



def validar_data_iso(valor):
    if not valor:
        return False
    try:
        datetime.strptime(valor, '%Y-%m-%d')
        return True
    except ValueError:
        return False


def documento_valido(cpf_cnpj_limpo, tipo_documento=None):
    tamanho = len(cpf_cnpj_limpo)

    if tipo_documento == 'PF':
        return tamanho == 11
    if tipo_documento == 'PJ':
        return tamanho == 14

    return tamanho in (11, 14)


def get_db_or_redirect(destino='dashboard', mensagem='Erro de conexão com o banco de dados.'):
    con = obter_conexao()
    if con is None:
        flash(mensagem, 'danger')
    return con


def fechar_cursor_conexao(cur=None, con=None):
    try:
        if cur:
            cur.close()
    except Exception:
        pass

    try:
        if con:
            con.close()
    except Exception:
        pass


# ==========================================================
# BLOCO 6.2 — PESSOAS: CATEGORIAS, PRESTADORES E BUSCA
# ==========================================================
CATEGORIAS_CADASTRO_PESSOA = [
    'Prestador de Serviço',
    'Fornecedor',
    'Funcionário',
    'Cliente / Tomador',
    'Transportadora',
    'Órgão Público',
    'Outros',
]

TIPOS_PRESTADOR_SERVICO = [
    'Motorista',
    'Ajudante',
    'Motorista e Ajudante',
    'Outros serviços',
]

CATEGORIAS_LEGADAS_PESSOA = {
    'Motorista': ('Prestador de Serviço', 'Motorista'),
    'Ajudante': ('Prestador de Serviço', 'Ajudante'),
    'Prestador': ('Prestador de Serviço', 'Outros serviços'),
    'Prestador de Serviço': ('Prestador de Serviço', None),
    'Funcionario': ('Funcionário', None),
}

def normalizar_categoria_pessoa(tipo_cadastro, tipo_prestador=None):
    tipo_cadastro = (tipo_cadastro or '').strip()
    tipo_prestador = (tipo_prestador or '').strip()

    if tipo_cadastro in CATEGORIAS_LEGADAS_PESSOA:
        categoria, prestador_padrao = CATEGORIAS_LEGADAS_PESSOA[tipo_cadastro]
        return categoria, (tipo_prestador or prestador_padrao)

    if tipo_cadastro not in CATEGORIAS_CADASTRO_PESSOA:
        return tipo_cadastro, tipo_prestador or None

    if tipo_cadastro != 'Prestador de Serviço':
        return tipo_cadastro, None

    if tipo_prestador not in TIPOS_PRESTADOR_SERVICO:
        tipo_prestador = 'Outros serviços'

    return tipo_cadastro, tipo_prestador

def pessoa_label_busca(pessoa):
    if not pessoa:
        return ''
    doc = pessoa.get('cpf_cnpj') or 'sem documento'
    categoria = pessoa.get('tipo_cadastro') or 'Cadastro'
    tipo_prestador = pessoa.get('tipo_prestador')
    complemento = f" / {tipo_prestador}" if tipo_prestador else ''
    return f"#{pessoa.get('id')} — {pessoa.get('nome_completo')} — {doc} — {categoria}{complemento}"

def condicao_sql_motorista_prestador(alias='p'):
    return f"""
        (
            {alias}.tipo_cadastro = 'Motorista'
            OR (
                {alias}.tipo_cadastro = 'Prestador de Serviço'
                AND COALESCE({alias}.tipo_prestador, '') IN ('Motorista', 'Motorista e Ajudante')
            )
        )
    """

def condicao_sql_ajudante_prestador(alias='p'):
    return f"""
        (
            {alias}.tipo_cadastro = 'Ajudante'
            OR (
                {alias}.tipo_cadastro = 'Prestador de Serviço'
                AND COALESCE({alias}.tipo_prestador, '') IN ('Ajudante', 'Motorista e Ajudante')
            )
        )
    """


def local_name(tag):
    return str(tag).split('}', 1)[-1]


def texto_primeira_tag(root, nomes):
    nomes = set(nomes)
    for el in root.iter():
        if local_name(el.tag) in nomes and el.text and el.text.strip():
            return el.text.strip()
    return None


def texto_filho_de_pai(root, nomes_pai, nomes_filho):
    nomes_pai = set(nomes_pai)
    nomes_filho = set(nomes_filho)

    for pai in root.iter():
        if local_name(pai.tag) in nomes_pai:
            for filho in pai.iter():
                if local_name(filho.tag) in nomes_filho and filho.text and filho.text.strip():
                    return filho.text.strip()
    return None


def extrair_dados_nfse(arquivo_xml):
    tree = ET.parse(arquivo_xml)
    root = tree.getroot()

    chave_acesso = None
    for el in root.iter():
        for attr_value in el.attrib.values():
            numeros = somente_digitos(attr_value)
            if len(numeros) >= 20:
                chave_acesso = numeros[:44]
                break
        if chave_acesso:
            break

    numero_nf = texto_primeira_tag(root, ['nNFSe', 'numero_nfse', 'Numero', 'NumeroNfse', 'NumeroNFe'])
    data_emissao = texto_primeira_tag(root, ['dhEmi', 'DataEmissao', 'data_emissao'])
    valor_total = texto_primeira_tag(root, ['vServ', 'ValorServicos', 'valor_total', 'ValorTotal'])

    emitente_cnpj = (
            texto_filho_de_pai(root, ['emit', 'prest', 'prestador', 'PrestadorServico', 'Prestador'],
                               ['CNPJ', 'CpfCnpj'])
            or texto_filho_de_pai(root, ['PrestadorServico'], ['Cnpj'])
    )

    tomador_cnpj = (
            texto_filho_de_pai(root, ['toma', 'tomador', 'TomadorServico', 'Tomador'], ['CNPJ', 'CpfCnpj'])
            or texto_filho_de_pai(root, ['TomadorServico'], ['Cnpj'])
    )

    if not numero_nf:
        raise ValueError('Número da nota fiscal não encontrado no XML.')
    if not data_emissao:
        raise ValueError('Data de emissão não encontrada no XML.')
    if not valor_total:
        raise ValueError('Valor total não encontrado no XML.')

    data_emissao = data_emissao[:10]
    if not validar_data_iso(data_emissao):
        raise ValueError('Data de emissão inválida no XML.')

    valor_total = converter_decimal(valor_total)
    if valor_total <= 0:
        raise ValueError('Valor total da nota fiscal inválido no XML.')

    if not chave_acesso:
        chave_acesso = f"NFSE{somente_digitos(numero_nf)}{data_emissao.replace('-', '')}"
        chave_acesso = chave_acesso[:44]

    return {
        'numero_nf': limitar_texto(numero_nf, 20),
        'chave_acesso': limitar_texto(chave_acesso, 44),
        'data_emissao': data_emissao,
        'valor_total': valor_total,
        'emitente_cnpj': somente_digitos(emitente_cnpj),
        'tomador_cnpj': somente_digitos(tomador_cnpj),
    }


def gerar_slug_empresa(texto):
    texto = str(texto or '').strip().lower()
    texto = re.sub(r'[áàãâä]', 'a', texto)
    texto = re.sub(r'[éèêë]', 'e', texto)
    texto = re.sub(r'[íìîï]', 'i', texto)
    texto = re.sub(r'[óòõôö]', 'o', texto)
    texto = re.sub(r'[úùûü]', 'u', texto)
    texto = re.sub(r'[ç]', 'c', texto)
    texto = re.sub(r'[^a-z0-9]+', '-', texto)
    texto = texto.strip('-')
    return texto or 'empresa'


# ==========================================================
# EMPRESA RAIZ / PADRÃO DO SISTEMA
# ==========================================================
EMPRESA_PADRAO_SISTEMA_ID = 1
EMPRESA_PADRAO_SISTEMA_CNPJ = '51953119000110'
EMPRESA_PADRAO_SISTEMA_NOME = 'J Tecnologia'
EMPRESA_PADRAO_SISTEMA_SLUG = 'j-tecnologia'
EMPRESA_PADRAO_SISTEMA_PLANO = 'Ilimitado'


def empresa_eh_padrao_sistema(empresa):
    """Identifica a empresa raiz do SGR.

    A J Tecnologia é a empresa padrão do ambiente: permanece ativa,
    com plano ilimitado e não pode ser inativada/excluída por telas.
    """
    if not empresa:
        return False
    try:
        if int(empresa.get('id') or 0) == EMPRESA_PADRAO_SISTEMA_ID:
            return True
    except Exception:
        pass
    cnpj = somente_digitos(empresa.get('cnpj') or '')
    return cnpj == EMPRESA_PADRAO_SISTEMA_CNPJ


def normalizar_empresa_padrao_sistema(cur):
    """Garante os campos críticos da empresa padrão.

    Essa função é leve e segura para ser chamada em telas administrativas.
    Não cria usuários nem apaga dados; apenas mantém a J Tecnologia protegida
    como empresa ativa/ilimitada.
    """
    try:
        cur.execute("""
            UPDATE empresas
            SET razao_social = CASE
                    WHEN COALESCE(razao_social, '') = '' THEN %s
                    ELSE razao_social
                END,
                nome_fantasia = %s,
                cnpj = %s,
                slug = %s,
                status_empresa = 'Ativa',
                plano = %s,
                limite_usuarios = NULL,
                observacao = CASE
                    WHEN COALESCE(observacao, '') LIKE '%%Empresa padrão raiz do SGR%%' THEN observacao
                    ELSE CONCAT(COALESCE(observacao, ''),
                                CASE WHEN COALESCE(observacao, '') = '' THEN '' ELSE '\n' END,
                                'Empresa padrão raiz do SGR. Registro protegido contra inativação, exclusão e limitação de plano.')
                END
            WHERE id = %s OR REPLACE(REPLACE(REPLACE(cnpj, '.', ''), '/', ''), '-', '') = %s
        """, (
            EMPRESA_PADRAO_SISTEMA_NOME,
            EMPRESA_PADRAO_SISTEMA_NOME,
            EMPRESA_PADRAO_SISTEMA_CNPJ,
            EMPRESA_PADRAO_SISTEMA_SLUG,
            EMPRESA_PADRAO_SISTEMA_PLANO,
            EMPRESA_PADRAO_SISTEMA_ID,
            EMPRESA_PADRAO_SISTEMA_CNPJ,
        ))
    except Exception as e:
        print(f"[Empresa Padrão] Aviso ao normalizar J Tecnologia: {e}")


def super_admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if int(session.get('is_super_admin') or 0) != 1:
            flash("Acesso restrito ao Super Admin.", "danger")
            return redirect(url_for('dashboard'))

        return f(*args, **kwargs)

    return decorated_function

# ============================================================
# SGR Web — Visualizador DANFSe para XML de NFS-e
# Requer: from danfse_parser import parse_danfse_xml
# ============================================================
def _buscar_nf_com_permissao(cur, nf_id):
    """
    Busca NF e aplica a mesma regra de permissão do XML atual.
    Ajuste nomes de campos se seu SELECT atual usar alias diferentes.
    """
    cur.execute("""
        SELECT
            nf.*,
            p.nome_completo AS motorista_nome,
            p.empresa_id AS motorista_empresa_id
        FROM motorista_notas_fiscais nf
        LEFT JOIN pessoas p ON p.id = nf.motorista_id
        WHERE nf.id = %s
        LIMIT 1
    """, (nf_id,))
    nf = cur.fetchone()
    if not nf:
        return None, "Nota fiscal não encontrada."

    perfil = session.get('perfil_de_acesso') or session.get('perfil')
    usuario_empresa_id = session.get('empresa_id')
    pessoa_id = session.get('pessoa_id')
    is_super_admin = session.get('is_super_admin') == 1 or session.get('is_super_admin') is True

    if perfil == 'Terminal Base':
        return None, "Terminal Base não tem permissão para acessar notas fiscais."

    if is_super_admin:
        return nf, None

    if perfil == 'Motorista':
        if int(nf.get('motorista_id') or 0) != int(pessoa_id or 0):
            return None, "Você não tem permissão para acessar esta nota fiscal."
        return nf, None

    if usuario_empresa_id and nf.get('empresa_id') and int(nf['empresa_id']) != int(usuario_empresa_id):
        return None, "Você não tem permissão para acessar nota fiscal de outra empresa."

    return nf, None


def _ler_xml_nf(nf):
    """
    Lê XML da NF a partir do caminho local, da rota interna /arquivos/visualizar/<id>
    ou do registro em arquivos_sistema/Google Drive.

    Esta versão não usa mysql.connection; segue o padrão do projeto com obter_conexao().
    """
    possiveis_campos = [
        nf.get('caminho_arquivo_xml') if isinstance(nf, dict) else None,
        nf.get('nome_arquivo_xml') if isinstance(nf, dict) else None,
        nf.get('arquivo_xml') if isinstance(nf, dict) else None,
    ]

    for caminho in possiveis_campos:
        if not caminho:
            continue

        caminho = str(caminho).strip()

        # Quando o campo aponta para a rota protegida do SGR, buscamos o arquivo_id.
        if caminho.startswith('/arquivos/visualizar/'):
            try:
                arquivo_id = int(caminho.rstrip('/').split('/')[-1])
            except Exception:
                arquivo_id = None

            if arquivo_id:
                con_arq = obter_conexao()
                if con_arq is None:
                    raise FileNotFoundError('Não foi possível conectar para localizar o XML.')

                cur_arq = con_arq.cursor(dictionary=True)
                try:
                    cur_arq.execute("""
                        SELECT *
                        FROM arquivos_sistema
                        WHERE id = %s
                          AND status_arquivo <> 'EXCLUIDO'
                        LIMIT 1
                    """, (arquivo_id,))
                    arq = cur_arq.fetchone()

                    if not arq:
                        raise FileNotFoundError('Registro do XML não localizado.')

                    caminho_local = arq.get('caminho_local') or arq.get('path_local') or arq.get('file_path')
                    if caminho_local:
                        caminho_local_norm = str(caminho_local).replace('\\', '/').lstrip('/')
                        candidatos = [
                            os.path.join(app.root_path, caminho_local_norm),
                            os.path.join(app.root_path, 'static', caminho_local_norm.replace('static/', '', 1)),
                        ]

                        for candidato in candidatos:
                            if os.path.exists(candidato):
                                with open(candidato, 'rb') as f:
                                    return f.read()

                    if arq.get('drive_file_id'):
                        return baixar_arquivo_google_drive(arq.get('drive_file_id'))

                finally:
                    fechar_cursor_conexao(cur_arq, con_arq)

        if caminho.startswith('/portal-motorista/'):
            continue

        if caminho.startswith(('http://', 'https://')):
            continue

        candidatos = []

        if os.path.isabs(caminho):
            candidatos.append(caminho)

        candidatos.append(os.path.join(app.root_path, caminho.lstrip('/')))
        candidatos.append(os.path.join(app.root_path, 'static', caminho.lstrip('/').replace('static/', '', 1)))

        # XML legado salvo apenas com nome do arquivo.
        candidatos.append(os.path.join(pasta_upload_xml_motorista(), os.path.basename(caminho)))

        for candidato in candidatos:
            if candidato and os.path.exists(candidato):
                with open(candidato, 'rb') as f:
                    return f.read()

    nf_id = nf.get('id') if isinstance(nf, dict) else None

    if not nf_id:
        raise FileNotFoundError('XML da nota não localizado.')

    con = obter_conexao()
    if con is None:
        raise FileNotFoundError('Não foi possível conectar para localizar o XML.')

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
            SELECT *
            FROM arquivos_sistema
            WHERE origem = 'XML_MOTORISTA'
              AND origem_id = %s
              AND status_arquivo <> 'EXCLUIDO'
            ORDER BY id DESC
            LIMIT 1
        """, (nf_id,))
        arq = cur.fetchone()

        if not arq:
            raise FileNotFoundError('XML da nota não localizado.')

        caminho_local = arq.get('caminho_local') or arq.get('path_local') or arq.get('file_path')
        if caminho_local:
            caminho_local_norm = str(caminho_local).replace('\\', '/').lstrip('/')
            candidatos = [
                os.path.join(app.root_path, caminho_local_norm),
                os.path.join(app.root_path, 'static', caminho_local_norm.replace('static/', '', 1)),
            ]

            for candidato in candidatos:
                if os.path.exists(candidato):
                    with open(candidato, 'rb') as f:
                        return f.read()

        if arq.get('drive_file_id'):
            return baixar_arquivo_google_drive(arq.get('drive_file_id'))

        raise FileNotFoundError('XML da nota não localizado.')

    finally:
        fechar_cursor_conexao(cur, con)

# ==========================================================
# HELPER - BLOQUEIO DE EDIÇÃO DE ROTAS EM FLUXO FISCAL/PAGAMENTO
# ==========================================================
def rota_pode_ser_editada(rota, possui_documento_ativo=False):
    """
    Regra:
    A rota só pode ser editada antes de entrar no fluxo fiscal/pagamento.

    Pode editar:
    - Aguardando liberação
    - Bloqueada, desde que não possua documento ativo vinculado

    Não pode editar:
    - Liberada para NF
    - NF enviada
    - Em análise
    - Aprovada para pagamento
    - Pagamento confirmado
    - Cancelada
    - Qualquer rota com documento ativo vinculado
    """
    if not rota:
        return False

    if possui_documento_ativo:
        return False

    status_motorista = rota.get('status_motorista') or 'Aguardando conferência'

    return status_motorista in ['Aguardando liberação', 'Bloqueada']


def rota_tem_documento_ativo(cur, rota_id, empresa_id):
    """
    Documento ativo = qualquer documento vinculado que não esteja Recusado.
    Mantemos documentos recusados para auditoria, mas eles não devem bloquear reenvio.
    """
    cur.execute("""
                SELECT COUNT(*) AS total
                FROM motorista_nf_rotas v
                         INNER JOIN motorista_notas_fiscais nf
                                    ON nf.id = v.motorista_nf_id
                                        AND nf.empresa_id = v.empresa_id
                WHERE v.rota_id = %s
                  AND v.empresa_id = %s
                  AND nf.status_nf <> 'Recusada'
                """, (rota_id, empresa_id))

    resultado = cur.fetchone() or {}
    return int(resultado.get('total') or 0) > 0


# ==========================================================
# PROTEÇÃO DE ROTAS
# ==========================================================
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            flash('Por favor, faça login para acessar esta página.', 'danger')
            return redirect(url_for('login'))
        return f(*args, **kwargs)

    return decorated_function


# ==========================================================
# FASE 4.0 - PERMISSÕES, SEGURANÇA E MULTIEMPRESA
# ==========================================================
def usuario_eh_super_admin_global():
    return int(session.get('is_super_admin') or 0) == 1


def usuario_eh_suporte_sistema():
    """Usuário técnico de suporte pode enxergar visão global de manutenção.

    Mantém o Super Admin como regra principal e permite um usuário de suporte
    sem abrir a visão global para administradores comuns de empresa.
    """
    if usuario_eh_super_admin_global():
        return True
    login = (session.get('usuario_login') or '').strip().lower()
    perfil = (session.get('perfil_de_acesso') or '').strip().lower()
    return login in {'suporte', 'suporte.sgr', 'sgr.suporte', 'admin.suporte'} or perfil in {'suporte', 'suporte do sistema'}


def usuario_pode_ver_escopo_global_sistema():
    return usuario_eh_super_admin_global() or usuario_eh_suporte_sistema()


def usuario_eh_super_admin_ou_suporte():
    """Alias de compatibilidade para rotinas novas/antigas."""
    return usuario_pode_ver_escopo_global_sistema()


def perfil_usuario_logado():
    return session.get('perfil_de_acesso')


def acesso_negado_por_perfil():
    """
    Redirecionamento seguro quando o usuário tenta acessar uma URL sem permissão.

    Ajuste v21:
    - Não manda usuário autenticado de volta para login.
    - Se a rota bloqueada for a página inicial, tenta redirecionar para o primeiro menu permitido.
    - Se o perfil não tiver nenhum módulo liberado, abre uma tela neutra interna do SGR Web.
    """
    perfil = perfil_usuario_logado()

    if perfil == 'Motorista':
        return redirect(url_for('portal_motorista'))

    if perfil == 'Terminal Base':
        return redirect(url_for('terminal_base_qrcode'))

    if request.endpoint == 'dashboard':
        destino = primeira_rota_permitida_usuario(excluir_endpoints={'dashboard'})
        if destino:
            return redirect(destino)
        return redirect(url_for('inicio_sem_modulos'))

    flash("Você não tem permissão para acessar esta área.", "danger")
    destino = primeira_rota_permitida_usuario(excluir_endpoints={request.endpoint, 'dashboard'})
    if destino:
        return redirect(destino)
    return redirect(url_for('inicio_sem_modulos'))


def perfis_permitidos(*perfis):
    """
    Decorator central de permissão por perfil.

    Regras:
    - Super Admin acessa tudo.
    - Usuário comum precisa ter perfil permitido.
    - Motorista só acessa rotas explicitamente liberadas para Motorista.
    """

    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            if 'usuario_id' not in session:
                flash('Por favor, faça login para acessar esta página.', 'danger')
                return redirect(url_for('login'))

            if usuario_eh_super_admin_global() or usuario_eh_suporte_sistema():
                return f(*args, **kwargs)

            permissao_avancada = permissao_avancada_endpoint(request.endpoint, 'visualizar')
            if permissao_avancada is True:
                return f(*args, **kwargs)
            if permissao_avancada is False:
                return acesso_negado_por_perfil()

            perfil = perfil_usuario_logado()

            if perfil in perfis:
                return f(*args, **kwargs)

            return acesso_negado_por_perfil()

        return decorated_function

    return decorator


def somente_perfis_internos(f):
    """
    Bloqueia Motorista em telas internas administrativas/operacionais.
    Útil para rotas de leitura geral do sistema.
    """

    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            flash('Por favor, faça login para acessar esta página.', 'danger')
            return redirect(url_for('login'))

        if usuario_eh_super_admin_global():
            return f(*args, **kwargs)

        if perfil_usuario_logado() == 'Motorista':
            return acesso_negado_por_perfil()

        return f(*args, **kwargs)

    return decorated_function



# ==========================================================
# BLOCO 12 - NÚCLEO DE PERMISSÕES E MENUS DINÂMICOS
# ==========================================================
_PERMISSOES_SCHEMA_OK = False

MENU_SISTEMA_PADRAO = [
    # grupo, parent_codigo, codigo, titulo, endpoint, icone, ordem, perfis
    ('VISÃO GERAL', None, 'dashboard', 'Dashboard', 'dashboard', 'fa-solid fa-chart-pie', 10, ['Administrador', 'Financeiro', 'Operacional', 'Supervisor', 'Terminal Base']),

    ('VISÃO GERAL', None, 'portal_motorista', 'Portal do Motorista', 'portal_motorista', 'fa-solid fa-truck-fast', 10, ['Motorista']),
    ('VISÃO GERAL', None, 'minhas_rotas_motorista', 'Minhas Rotas', 'minhas_rotas_motorista', 'fa-solid fa-route', 20, ['Motorista']),
    ('VISÃO GERAL', None, 'minha_semana_motorista', 'Minha Semana', 'minha_semana_motorista', 'fa-solid fa-calendar-week', 30, ['Motorista']),
    ('VISÃO GERAL', None, 'disponibilidade_motorista', 'Minha Disponibilidade', 'disponibilidade_motorista', 'fa-solid fa-calendar-check', 40, ['Motorista']),
    ('VISÃO GERAL', None, 'enviar_nf_motorista', 'Enviar NF/XML', 'enviar_nf_motorista', 'fa-solid fa-file-arrow-up', 50, ['Motorista']),
    ('VISÃO GERAL', None, 'solicitar_pagamento_sem_nf_motorista', 'Solicitar Pagamento sem NF', 'solicitar_pagamento_sem_nf_motorista', 'fa-solid fa-hand-holding-dollar', 60, ['Motorista']),
    ('VISÃO GERAL', None, 'minhas_nfs_motorista', 'Meus Documentos', 'minhas_nfs_motorista', 'fa-solid fa-file-invoice', 70, ['Motorista']),

    ('CADASTROS', None, 'pessoas', 'Pessoas', None, 'fa-solid fa-address-book', 10, ['Administrador', 'Operacional', 'Supervisor']),
    ('CADASTROS', 'pessoas', 'cadastro_pessoa', 'Novo Cadastro PF/PJ', 'cadastro_pessoa', 'fa-solid fa-user-plus', 10, ['Administrador', 'Operacional']),
    ('CADASTROS', 'pessoas', 'visualizar_pessoas', 'Visualizar Cadastros', 'visualizar_pessoas', 'fa-solid fa-users', 20, ['Administrador', 'Operacional', 'Supervisor']),

    ('OPERAÇÃO', None, 'rotas_ajudantes', 'Rotas e Ajudantes', None, 'fa-solid fa-truck-fast', 10, ['Administrador', 'Operacional', 'Supervisor']),
    ('OPERAÇÃO', 'rotas_ajudantes', 'lancar_rota', 'Lançar Nova Rota', 'lancar_rota', 'fa-solid fa-road', 10, ['Administrador', 'Operacional']),
    ('OPERAÇÃO', 'rotas_ajudantes', 'visualizar_rotas', 'Visualizar Rotas', 'visualizar_rotas', 'fa-solid fa-list-check', 20, ['Administrador', 'Operacional', 'Supervisor']),
    ('OPERAÇÃO', 'rotas_ajudantes', 'divergencias_rotas_motoristas', 'Divergências de Rotas', 'divergencias_rotas_motoristas', 'fa-solid fa-triangle-exclamation', 30, ['Administrador', 'Operacional', 'Supervisor']),
    ('OPERAÇÃO', 'rotas_ajudantes', 'lancamento_ajudante', 'Lançamento Ajudante', 'lancamento_ajudante', 'fa-solid fa-user-gear', 40, ['Administrador', 'Operacional']),
    ('OPERAÇÃO', 'rotas_ajudantes', 'escala_motoristas', 'Escala dos Motoristas', 'escala_motoristas', 'fa-solid fa-clipboard-list', 50, ['Administrador', 'Operacional', 'Supervisor']),
    ('OPERAÇÃO', 'rotas_ajudantes', 'visualizar_bases_operacionais', 'Bases Operacionais', 'visualizar_bases_operacionais', 'fa-solid fa-location-dot', 60, ['Administrador', 'Operacional', 'Supervisor']),
    ('OPERAÇÃO', 'rotas_ajudantes', 'central_pendencias_operacao', 'Central de Pendências', 'central_pendencias_operacao', 'fa-solid fa-list-check', 70, ['Administrador', 'Operacional', 'Supervisor']),
    ('OPERAÇÃO', 'rotas_ajudantes', 'score_motoristas', 'Score dos Motoristas', 'score_motoristas', 'fa-solid fa-ranking-star', 80, ['Administrador', 'Operacional', 'Supervisor']),
    ('OPERAÇÃO', 'rotas_ajudantes', 'historico_motoristas', 'Histórico de Motoristas', 'historico_motoristas', 'fa-solid fa-id-card-clip', 90, ['Administrador', 'Operacional', 'Supervisor']),

    ('FINANCEIRO', None, 'gestao_financeira', 'Gestão Financeira', None, 'fa-solid fa-sack-dollar', 10, ['Administrador', 'Financeiro']),
    ('FINANCEIRO', 'gestao_financeira', 'financeiro_dashboard', 'Dashboard Financeiro', 'financeiro_dashboard', 'fa-solid fa-chart-line', 10, ['Administrador', 'Financeiro']),
    ('FINANCEIRO', 'gestao_financeira', 'financeiro_titulos', 'Títulos Financeiros', 'financeiro_titulos', 'fa-solid fa-list-check', 20, ['Administrador', 'Financeiro']),
    ('FINANCEIRO', 'gestao_financeira', 'financeiro_contas_caixa', 'Contas Caixa', 'financeiro.financeiro_contas_caixa', 'fa-solid fa-building-columns', 30, ['Administrador', 'Financeiro']),
    ('FINANCEIRO', 'gestao_financeira', 'financeiro_movimentacoes_caixa', 'Movimentações Caixa', 'financeiro_movimentacoes_caixa', 'fa-solid fa-right-left', 40, ['Administrador', 'Financeiro']),
    ('FINANCEIRO', 'gestao_financeira', 'financeiro_conciliacao_caixa', 'Conciliação de Caixa', 'financeiro_conciliacao_caixa', 'fa-solid fa-scale-balanced', 50, ['Administrador', 'Financeiro']),
    ('FINANCEIRO', 'gestao_financeira', 'financeiro_configuracoes', 'Configurações Financeiras', 'financeiro_configuracoes', 'fa-solid fa-gear', 60, ['Administrador']),
    ('FINANCEIRO', 'gestao_financeira', 'financeiro_auditoria', 'Auditoria Financeira', 'financeiro_auditoria', 'fa-solid fa-shield-halved', 70, ['Administrador', 'Financeiro']),
    ('FINANCEIRO', 'gestao_financeira', 'financeiro_nfs_motoristas', 'Documentos Motoristas', 'financeiro_nfs_motoristas', 'fa-solid fa-file-signature', 80, ['Administrador', 'Financeiro']),
    ('FINANCEIRO', 'gestao_financeira', 'pagamentos_ajudante', 'Pagamentos Ajudante', 'pagamentos_ajudante', 'fa-solid fa-people-carry-box', 90, ['Administrador', 'Financeiro']),
    ('FINANCEIRO', 'gestao_financeira', 'faturamento', 'Faturamento XML', 'faturamento', 'fa-solid fa-file-invoice-dollar', 100, ['Administrador', 'Financeiro']),
    ('FINANCEIRO', 'gestao_financeira', 'recebimento', 'Recebimento de NF', 'recebimento', 'fa-solid fa-file-circle-check', 110, ['Administrador', 'Financeiro']),

    ('RELATÓRIOS', None, 'relatorios', 'Relatórios', None, 'fa-solid fa-chart-column', 10, ['Administrador', 'Financeiro']),
    ('RELATÓRIOS', 'relatorios', 'relatorios_central', 'Central Gerencial', 'relatorios_central', 'fa-solid fa-table-columns', 10, ['Administrador', 'Financeiro']),
    ('RELATÓRIOS', 'relatorios', 'relatorios_financeiro', 'Relatório Financeiro', 'relatorios_financeiro', 'fa-solid fa-file-invoice-dollar', 20, ['Administrador', 'Financeiro']),

    ('AUDITORIA', None, 'historicos', 'Históricos', None, 'fa-solid fa-shield-halved', 10, ['Administrador', 'Supervisor']),
    ('AUDITORIA', 'historicos', 'historico_estornos', 'Histórico de Estornos NF', 'historico_estornos', 'fa-solid fa-clock-rotate-left', 10, ['Administrador']),
    ('AUDITORIA', 'historicos', 'historico_pagamentos_ajudante', 'Histórico Pag. Ajudante', 'historico_pagamentos_ajudante', 'fa-solid fa-clipboard-list', 20, ['Administrador', 'Financeiro']),
    ('AUDITORIA', 'historicos', 'auditoria_checkin_base', 'Auditoria de Check-in', 'auditoria_checkin_base', 'fa-solid fa-shield-halved', 30, ['Administrador', 'Supervisor']),
    ('AUDITORIA', 'historicos', 'mapa_checkins', 'Mapa de Check-ins', 'mapa_checkins', 'fa-solid fa-map-location-dot', 40, ['Administrador', 'Supervisor']),
    ('AUDITORIA', 'historicos', 'visualizar_auditoria_supervisor', 'Auditoria do Supervisor', 'visualizar_auditoria_supervisor', 'fa-solid fa-user-shield', 50, ['Administrador', 'Supervisor']),
    ('AUDITORIA', 'historicos', 'relatorio_operacional_escala_base', 'Relatório Operacional', 'relatorio_operacional_escala_base', 'fa-solid fa-chart-line', 60, ['Administrador', 'Supervisor']),

    ('SISTEMA', None, 'usuarios_acessos', 'Usuários e Acessos', None, 'fa-solid fa-gear', 10, ['Administrador']),
    ('SISTEMA', 'usuarios_acessos', 'visualizar_empresas', 'Empresas', 'visualizar_empresas', 'fa-solid fa-building', 5, []),
    ('SISTEMA', 'usuarios_acessos', 'criar_usuario', 'Novo Usuário', 'criar_usuario', 'fa-solid fa-user-plus', 10, ['Administrador']),
    ('SISTEMA', 'usuarios_acessos', 'visualizar_usuarios', 'Visualizar Usuários', 'visualizar_usuarios', 'fa-solid fa-user-shield', 20, ['Administrador']),
    ('SISTEMA', 'usuarios_acessos', 'perfil_acesso', 'Permissões e Menus', 'perfil_acesso', 'fa-solid fa-lock', 30, ['Administrador']),
    ('SISTEMA', 'usuarios_acessos', 'configuracoes_operacionais_motorista', 'Config. Motorista', 'configuracoes_operacionais_motorista', 'fa-solid fa-sliders', 35, ['Administrador', 'Supervisor']),
    ('SISTEMA', 'usuarios_acessos', 'gerenciar_menus_modulos', 'Menus e Módulos', 'gerenciar_menus_modulos', 'fa-solid fa-sitemap', 40, []),
]

ACOES_SISTEMA_PADRAO = ['visualizar', 'criar', 'editar', 'excluir', 'aprovar', 'recusar', 'baixar', 'estornar', 'exportar', 'ver_auditoria', 'ver_documentos']

# Menus/endpoints que só podem aparecer para Super Admin global.
# Administrador de empresa continua sendo administrador apenas da própria empresa,
# mas não deve visualizar cadastros globais como Empresas.
MENUS_APENAS_SUPER_ADMIN = {'visualizar_empresas', 'gerenciar_menus_modulos'}
ENDPOINTS_APENAS_SUPER_ADMIN = {'visualizar_empresas', 'cadastro_empresa', 'editar_empresa', 'gerenciar_menus_modulos', 'novo_modulo_sistema', 'editar_modulo_sistema', 'alternar_modulo_sistema', 'novo_menu_sistema', 'editar_menu_sistema', 'alternar_menu_sistema'}
_PERMISSOES_CACHE = {}


def menu_ou_endpoint_restrito_super_admin(menu_codigo=None, endpoint=None):
    return (menu_codigo in MENUS_APENAS_SUPER_ADMIN) or (endpoint in ENDPOINTS_APENAS_SUPER_ADMIN)


def invalidar_cache_permissoes(usuario_id=None):
    global _PERMISSOES_CACHE
    if usuario_id is None:
        _PERMISSOES_CACHE = {}
        return
    prefixo = f"{usuario_id}:"
    _PERMISSOES_CACHE = {k: v for k, v in _PERMISSOES_CACHE.items() if not str(k).startswith(prefixo)}


def coluna_existe(cur, tabela, coluna):
    try:
        cur.execute("SHOW COLUMNS FROM {} LIKE %s".format(tabela), (coluna,))
        return cur.fetchone() is not None
    except Exception:
        return False



def executar_ddl_permissoes(cur, sql):
    """Executa DDL do núcleo de permissões ignorando objetos já existentes.
    Evita erro 1050/1060/1061 quando o banco já recebeu a migration SQL.
    """
    try:
        cur.execute(sql)
        return True
    except Exception as e:
        codigo = getattr(e, 'errno', None)
        if codigo is None and getattr(e, 'args', None):
            try:
                codigo = int(e.args[0])
            except Exception:
                codigo = None
        msg = str(e).lower()
        if codigo in (1050, 1060, 1061, 1068, 1091) or 'already exists' in msg or 'duplicate column' in msg or 'duplicate key name' in msg:
            return False
        raise

def nucleo_permissoes_ja_instalado(cur):
    """Verificação rápida para produção.

    No Render, o worker nasce com o cache em memória zerado. Sem essa checagem,
    o login tentava sincronizar todos os módulos/menus/perfis novamente e podia
    estourar o timeout do Gunicorn. Quando as tabelas básicas já existem e têm
    dados, marcamos o núcleo como pronto e evitamos a rotina pesada.
    """
    try:
        cur.execute("SELECT 1 FROM sistema_modulos LIMIT 1")
        if cur.fetchone() is None:
            return False
        cur.execute("SELECT 1 FROM sistema_menus LIMIT 1")
        if cur.fetchone() is None:
            return False
        cur.execute("SELECT 1 FROM sistema_acoes LIMIT 1")
        if cur.fetchone() is None:
            return False
        cur.execute("SELECT 1 FROM perfil_permissoes LIMIT 1")
        return True
    except Exception:
        return False


def inicializar_nucleo_permissoes(force=False):
    """Cria e sincroniza a base de módulos, menus e permissões sem quebrar instalações antigas."""
    global _PERMISSOES_SCHEMA_OK
    if _PERMISSOES_SCHEMA_OK and not force:
        return True

    con = obter_conexao()
    if con is None:
        return False
    cur = con.cursor(dictionary=True)
    try:
        if not force and nucleo_permissoes_ja_instalado(cur):
            _PERMISSOES_SCHEMA_OK = True
            return True

        executar_ddl_permissoes(cur, """
            CREATE TABLE IF NOT EXISTS sistema_modulos (
                id INT AUTO_INCREMENT PRIMARY KEY,
                codigo VARCHAR(80) NOT NULL UNIQUE,
                nome VARCHAR(120) NOT NULL,
                descricao VARCHAR(255) NULL,
                icone VARCHAR(80) NULL,
                ordem INT DEFAULT 0,
                ativo TINYINT(1) DEFAULT 1,
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        executar_ddl_permissoes(cur, """
            CREATE TABLE IF NOT EXISTS sistema_menus (
                id INT AUTO_INCREMENT PRIMARY KEY,
                modulo_id INT NULL,
                menu_pai_id INT NULL,
                grupo_menu VARCHAR(80) NOT NULL,
                codigo VARCHAR(120) NOT NULL UNIQUE,
                titulo VARCHAR(160) NOT NULL,
                endpoint VARCHAR(120) NULL,
                rota_url VARCHAR(255) NULL,
                icone VARCHAR(100) NULL,
                ordem INT DEFAULT 0,
                ativo TINYINT(1) DEFAULT 1,
                visivel_menu TINYINT(1) DEFAULT 1,
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_sistema_menus_pai (menu_pai_id),
                INDEX idx_sistema_menus_endpoint (endpoint),
                INDEX idx_sistema_menus_grupo (grupo_menu)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        executar_ddl_permissoes(cur, """
            CREATE TABLE IF NOT EXISTS sistema_acoes (
                id INT AUTO_INCREMENT PRIMARY KEY,
                codigo VARCHAR(60) NOT NULL UNIQUE,
                nome VARCHAR(120) NOT NULL,
                ativo TINYINT(1) DEFAULT 1
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        executar_ddl_permissoes(cur, """
            CREATE TABLE IF NOT EXISTS perfil_permissoes (
                id INT AUTO_INCREMENT PRIMARY KEY,
                perfil_de_acesso VARCHAR(80) NOT NULL,
                menu_codigo VARCHAR(120) NOT NULL,
                acao_codigo VARCHAR(60) NOT NULL DEFAULT 'visualizar',
                empresa_id INT NOT NULL DEFAULT 0,
                permitido TINYINT(1) DEFAULT 1,
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uk_perfil_menu_acao_empresa (perfil_de_acesso, menu_codigo, acao_codigo, empresa_id),
                INDEX idx_perfil_permissoes_menu (menu_codigo),
                INDEX idx_perfil_permissoes_empresa (empresa_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        executar_ddl_permissoes(cur, """
            CREATE TABLE IF NOT EXISTS usuario_permissoes (
                id INT AUTO_INCREMENT PRIMARY KEY,
                usuario_id INT NOT NULL,
                menu_codigo VARCHAR(120) NOT NULL,
                acao_codigo VARCHAR(60) NOT NULL DEFAULT 'visualizar',
                empresa_id INT NOT NULL DEFAULT 0,
                permitido TINYINT(1) DEFAULT 1,
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uk_usuario_menu_acao_empresa (usuario_id, menu_codigo, acao_codigo, empresa_id),
                INDEX idx_usuario_permissoes_menu (menu_codigo),
                INDEX idx_usuario_permissoes_empresa (empresa_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        executar_ddl_permissoes(cur, """
            CREATE TABLE IF NOT EXISTS usuario_empresas_acesso (
                id INT AUTO_INCREMENT PRIMARY KEY,
                usuario_id INT NOT NULL,
                empresa_id INT NOT NULL,
                ativo TINYINT(1) DEFAULT 1,
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                UNIQUE KEY uk_usuario_empresa_acesso (usuario_id, empresa_id),
                INDEX idx_usuario_empresas_empresa (empresa_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)
        executar_ddl_permissoes(cur, """
            CREATE TABLE IF NOT EXISTS auditoria_permissoes (
                id INT AUTO_INCREMENT PRIMARY KEY,
                usuario_executor_id INT NULL,
                usuario_afetado_id INT NULL,
                perfil_afetado VARCHAR(80) NULL,
                empresa_id INT NULL,
                acao VARCHAR(80) NOT NULL,
                detalhe TEXT NULL,
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
        """)

        for acao in ACOES_SISTEMA_PADRAO:
            nome = acao.replace('_', ' ').title()
            cur.execute("""
                INSERT INTO sistema_acoes (codigo, nome, ativo)
                VALUES (%s, %s, 1)
                ON DUPLICATE KEY UPDATE nome=VALUES(nome), ativo=1
            """, (acao, nome))

        grupos = {}
        for grupo, parent_codigo, codigo, titulo, endpoint, icone, ordem, perfis in MENU_SISTEMA_PADRAO:
            modulo_codigo = grupo.lower().replace('ç', 'c').replace('ã', 'a').replace('ó', 'o').replace(' ', '_')
            if modulo_codigo not in grupos:
                cur.execute("""
                    INSERT INTO sistema_modulos (codigo, nome, icone, ordem, ativo)
                    VALUES (%s, %s, %s, %s, 1)
                    ON DUPLICATE KEY UPDATE nome=VALUES(nome), icone=VALUES(icone), ordem=VALUES(ordem), ativo=1
                """, (modulo_codigo, grupo.title(), icone, len(grupos) * 10))
                cur.execute("SELECT id FROM sistema_modulos WHERE codigo=%s LIMIT 1", (modulo_codigo,))
                grupos[modulo_codigo] = cur.fetchone()['id']

        ids_menus = {}
        # Pais primeiro, filhos depois.
        for rodada in [0, 1]:
            for grupo, parent_codigo, codigo, titulo, endpoint, icone, ordem, perfis in MENU_SISTEMA_PADRAO:
                is_filho = parent_codigo is not None
                if rodada == 0 and is_filho:
                    continue
                if rodada == 1 and not is_filho:
                    continue
                modulo_codigo = grupo.lower().replace('ç', 'c').replace('ã', 'a').replace('ó', 'o').replace(' ', '_')
                parent_id = ids_menus.get(parent_codigo) if parent_codigo else None
                if endpoint and endpoint not in app.view_functions:
                    # Mantém cadastrado como inativo se a rota não existir nesta instalação.
                    ativo = 0
                else:
                    ativo = 1
                cur.execute("""
                    INSERT INTO sistema_menus (modulo_id, menu_pai_id, grupo_menu, codigo, titulo, endpoint, rota_url, icone, ordem, ativo, visivel_menu)
                    VALUES (%s, %s, %s, %s, %s, %s, NULL, %s, %s, %s, 1)
                    ON DUPLICATE KEY UPDATE
                        modulo_id=VALUES(modulo_id), menu_pai_id=VALUES(menu_pai_id), grupo_menu=VALUES(grupo_menu),
                        titulo=VALUES(titulo), endpoint=VALUES(endpoint), icone=VALUES(icone), ordem=VALUES(ordem), ativo=VALUES(ativo), visivel_menu=1
                """, (grupos[modulo_codigo], parent_id, grupo, codigo, titulo, endpoint, icone, ordem, ativo))
                cur.execute("SELECT id FROM sistema_menus WHERE codigo=%s LIMIT 1", (codigo,))
                ids_menus[codigo] = cur.fetchone()['id']

                for perfil in perfis:
                    if menu_ou_endpoint_restrito_super_admin(menu_codigo=codigo, endpoint=endpoint):
                        continue
                    cur.execute("""
                        INSERT INTO perfil_permissoes (perfil_de_acesso, menu_codigo, acao_codigo, empresa_id, permitido)
                        VALUES (%s, %s, 'visualizar', 0, 1)
                        ON DUPLICATE KEY UPDATE permitido=VALUES(permitido)
                    """, (perfil, codigo))

        # Garante vínculo usuário x empresa para usuários já existentes.
        if tabela_existe(cur, 'usuarios'):
            cur.execute("""
                INSERT IGNORE INTO usuario_empresas_acesso (usuario_id, empresa_id, ativo)
                SELECT id, empresa_id, 1
                FROM usuarios
                WHERE empresa_id IS NOT NULL
            """)

        # Bloco 12.3: perfis globais viram modelo e cada empresa recebe cópia própria.
        sincronizar_perfis_padrao_empresas(cur)

        con.commit()
        _PERMISSOES_SCHEMA_OK = True
        return True
    except Exception as e:
        con.rollback()
        print(f"[Permissões] Falha ao inicializar núcleo: {e}")
        return False
    finally:
        fechar_cursor_conexao(cur, con)


def registrar_auditoria_permissao(acao, detalhe='', usuario_afetado_id=None, perfil_afetado=None, empresa_id=None):
    con = obter_conexao()
    if con is None:
        return
    cur = con.cursor(dictionary=True)
    try:
        if not tabela_existe(cur, 'auditoria_permissoes'):
            return
        cur.execute("""
            INSERT INTO auditoria_permissoes (usuario_executor_id, usuario_afetado_id, perfil_afetado, empresa_id, acao, detalhe)
            VALUES (%s, %s, %s, %s, %s, %s)
        """, (session.get('usuario_id'), usuario_afetado_id, perfil_afetado, empresa_id, acao, detalhe))
        con.commit()
    except Exception as e:
        con.rollback()
        print(f"[Permissões] Falha ao auditar permissão: {e}")
    finally:
        fechar_cursor_conexao(cur, con)


def empresas_permitidas_usuario(usuario_id=None):
    usuario_id = usuario_id or session.get('usuario_id')
    empresa_sessao = session.get('empresa_id')
    if usuario_eh_super_admin_global():
        return None  # None = todas as empresas.
    if not usuario_id:
        return [int(empresa_sessao)] if empresa_sessao else []
    inicializar_nucleo_permissoes()
    con = obter_conexao()
    if con is None:
        return [int(empresa_sessao)] if empresa_sessao else []
    cur = con.cursor(dictionary=True)
    try:
        if tabela_existe(cur, 'usuario_empresas_acesso'):
            cur.execute("""
                SELECT empresa_id
                FROM usuario_empresas_acesso
                WHERE usuario_id=%s AND ativo=1
            """, (usuario_id,))
            ids = [int(r['empresa_id']) for r in cur.fetchall() if r.get('empresa_id')]
            if ids:
                return ids
        return [int(empresa_sessao)] if empresa_sessao else []
    except Exception as e:
        print(f"[Permissões] Falha ao buscar empresas do usuário: {e}")
        return [int(empresa_sessao)] if empresa_sessao else []
    finally:
        fechar_cursor_conexao(cur, con)


def usuario_pode_empresa(empresa_id, usuario_id=None):
    if usuario_eh_super_admin_global():
        return True
    if not empresa_id:
        return False
    permitidas = empresas_permitidas_usuario(usuario_id)
    return permitidas is None or int(empresa_id) in [int(x) for x in permitidas]


def buscar_menu_por_endpoint(cur, endpoint):
    if not endpoint:
        return None
    cur.execute("""
        SELECT codigo, endpoint, titulo
        FROM sistema_menus
        WHERE endpoint=%s AND ativo=1
        LIMIT 1
    """, (endpoint,))
    return cur.fetchone()


def permissao_avancada_endpoint(endpoint=None, acao='visualizar', empresa_id=None):
    """
    Retorna True/False quando há regra explícita no núcleo novo.
    Retorna None quando não há regra e o sistema deve cair no decorator antigo.
    Prioridade: Super Admin > usuário > perfil.
    """
    if 'usuario_id' not in session:
        return False
    endpoint_validado = endpoint or request.endpoint
    if usuario_eh_super_admin_global():
        return True
    if menu_ou_endpoint_restrito_super_admin(endpoint=endpoint_validado):
        return False

    inicializar_nucleo_permissoes()
    con = obter_conexao()
    if con is None:
        return None
    cur = con.cursor(dictionary=True)
    try:
        empresa_ref = empresa_id if empresa_id is not None else session.get('empresa_id')
        if empresa_ref and not usuario_pode_empresa(empresa_ref):
            return False
        menu = buscar_menu_por_endpoint(cur, endpoint or request.endpoint)
        if not menu:
            return None
        menu_codigo = menu['codigo']
        usuario_id = session.get('usuario_id')
        perfil = perfil_codigo_permissao_atual()

        # Regra manual por usuário tem prioridade absoluta.
        cur.execute("""
            SELECT permitido
            FROM usuario_permissoes
            WHERE usuario_id=%s AND menu_codigo=%s AND acao_codigo=%s
              AND empresa_id = COALESCE(%s, 0)
            ORDER BY empresa_id DESC, id DESC
            LIMIT 1
        """, (usuario_id, menu_codigo, acao, empresa_ref))
        row = cur.fetchone()
        if row is not None:
            return bool(int(row.get('permitido') or 0))

        # Regra por perfil.
        cur.execute("""
            SELECT permitido
            FROM perfil_permissoes
            WHERE perfil_de_acesso=%s AND menu_codigo=%s AND acao_codigo=%s
              AND empresa_id = COALESCE(%s, 0)
            ORDER BY empresa_id DESC, id DESC
            LIMIT 1
        """, (perfil, menu_codigo, acao, empresa_ref))
        row = cur.fetchone()
        if row is not None:
            return bool(int(row.get('permitido') or 0))
        return None
    except Exception as e:
        print(f"[Permissões] Falha ao validar endpoint {endpoint}: {e}")
        return None
    finally:
        fechar_cursor_conexao(cur, con)


def usuario_tem_permissao_menu(menu_codigo, acao='visualizar', empresa_id=None):
    if usuario_eh_super_admin_global():
        return True
    if not menu_codigo:
        return False
    if menu_ou_endpoint_restrito_super_admin(menu_codigo=menu_codigo):
        return False
    inicializar_nucleo_permissoes()
    con = obter_conexao()
    if con is None:
        return False
    cur = con.cursor(dictionary=True)
    try:
        usuario_id = session.get('usuario_id')
        perfil = perfil_codigo_permissao_atual()
        empresa_ref = empresa_id if empresa_id is not None else session.get('empresa_id')
        cur.execute("""
            SELECT permitido
            FROM usuario_permissoes
            WHERE usuario_id=%s AND menu_codigo=%s AND acao_codigo=%s
              AND empresa_id = COALESCE(%s, 0)
            ORDER BY empresa_id DESC, id DESC
            LIMIT 1
        """, (usuario_id, menu_codigo, acao, empresa_ref))
        row = cur.fetchone()
        if row is not None:
            return bool(int(row.get('permitido') or 0))
        cur.execute("""
            SELECT permitido
            FROM perfil_permissoes
            WHERE perfil_de_acesso=%s AND menu_codigo=%s AND acao_codigo=%s
              AND empresa_id = COALESCE(%s, 0)
            ORDER BY empresa_id DESC, id DESC
            LIMIT 1
        """, (perfil, menu_codigo, acao, empresa_ref))
        row = cur.fetchone()
        return bool(int(row.get('permitido') or 0)) if row is not None else False
    except Exception as e:
        print(f"[Permissões] Falha ao validar menu {menu_codigo}: {e}")
        return False
    finally:
        fechar_cursor_conexao(cur, con)


def carregar_menu_lateral_usuario():
    """Monta menu lateral a partir do banco, com cache por usuário/sessão.
    Regras importantes:
    - Super Admin vê tudo.
    - Administrador de empresa NÃO vê menus globais restritos, como Empresas.
    - Permissões são consultadas em lote para reduzir lentidão.
    """
    if 'usuario_id' not in session:
        return []

    usuario_id = session.get('usuario_id')
    perfil = perfil_codigo_permissao_atual()
    empresa_id = session.get('empresa_id') or 0
    is_super = 1 if usuario_pode_ver_escopo_global_sistema() else 0
    cache_key = f"{usuario_id}:{perfil}:{empresa_id}:{is_super}:{request.endpoint}"
    if cache_key in _PERMISSOES_CACHE:
        return _PERMISSOES_CACHE[cache_key]

    inicializar_nucleo_permissoes()
    con = obter_conexao()
    if con is None:
        return []
    cur = con.cursor(dictionary=True)
    try:
        cur.execute("""
            SELECT id, menu_pai_id, grupo_menu, codigo, titulo, endpoint, rota_url, icone, ordem
            FROM sistema_menus
            WHERE ativo=1 AND visivel_menu=1
            ORDER BY FIELD(grupo_menu, 'VISÃO GERAL','CADASTROS','OPERAÇÃO','FINANCEIRO','RELATÓRIOS','AUDITORIA','SISTEMA'), grupo_menu, menu_pai_id IS NOT NULL, ordem, titulo
        """)
        rows = cur.fetchall()
        if not rows:
            return []

        permissoes = {}
        if is_super:
            permissoes = {r.get('codigo'): True for r in rows}
        else:
            codigos = [r.get('codigo') for r in rows if r.get('codigo')]
            placeholders = ','.join(['%s'] * len(codigos))
            # Permissões por perfil.
            cur.execute(f"""
                SELECT menu_codigo, permitido, empresa_id
                FROM perfil_permissoes
                WHERE perfil_de_acesso=%s
                  AND acao_codigo='visualizar'
                  AND menu_codigo IN ({placeholders})
                  AND empresa_id = COALESCE(%s, 0)
                ORDER BY empresa_id ASC, id ASC
            """, [perfil] + codigos + [empresa_id])
            for r in cur.fetchall():
                permissoes[r.get('menu_codigo')] = bool(int(r.get('permitido') or 0))

            # Permissões manuais por usuário sobrescrevem perfil.
            cur.execute(f"""
                SELECT menu_codigo, permitido, empresa_id
                FROM usuario_permissoes
                WHERE usuario_id=%s
                  AND acao_codigo='visualizar'
                  AND menu_codigo IN ({placeholders})
                  AND empresa_id = COALESCE(%s, 0)
                ORDER BY empresa_id ASC, id ASC
            """, [usuario_id] + codigos + [empresa_id])
            for r in cur.fetchall():
                permissoes[r.get('menu_codigo')] = bool(int(r.get('permitido') or 0))

            # Bloqueio absoluto de menus globais de Super Admin.
            for r in rows:
                if menu_ou_endpoint_restrito_super_admin(r.get('codigo'), r.get('endpoint')):
                    permissoes[r.get('codigo')] = False

        by_id = {r['id']: dict(r, filhos=[]) for r in rows}
        for r in rows:
            if r.get('menu_pai_id') and r.get('menu_pai_id') in by_id:
                by_id[r['menu_pai_id']]['filhos'].append(by_id[r['id']])
        raiz = [by_id[r['id']] for r in rows if not r.get('menu_pai_id')]

        grupos = []
        grupo_atual = None
        endpoint_atual = request.endpoint
        for item in raiz:
            filhos_permitidos = []
            for filho in item.get('filhos', []):
                if not permissoes.get(filho.get('codigo'), False):
                    continue
                try:
                    href = url_for(filho.get('endpoint')) if filho.get('endpoint') else (filho.get('rota_url') or '#')
                except Exception:
                    continue
                filho['href'] = href
                filho['active'] = (endpoint_atual == filho.get('endpoint'))
                filhos_permitidos.append(filho)

            permitido_item = permissoes.get(item.get('codigo'), False)
            if not permitido_item and not filhos_permitidos:
                continue
            try:
                href_item = url_for(item.get('endpoint')) if item.get('endpoint') else (item.get('rota_url') or '#')
            except Exception:
                href_item = '#'
            item['href'] = href_item
            item['filhos'] = filhos_permitidos
            item['active'] = endpoint_atual == item.get('endpoint') or any(f.get('active') for f in filhos_permitidos)

            if grupo_atual is None or grupo_atual['nome'] != item.get('grupo_menu'):
                grupo_atual = {'nome': item.get('grupo_menu'), 'itens': []}
                grupos.append(grupo_atual)
            grupo_atual['itens'].append(item)

        # Cache simples em memória para reduzir consultas repetidas no render.
        if len(_PERMISSOES_CACHE) > 500:
            _PERMISSOES_CACHE.clear()
        _PERMISSOES_CACHE[cache_key] = grupos
        return grupos
    except Exception as e:
        print(f"[Permissões] Falha ao carregar menu dinâmico: {e}")
        return []
    finally:
        fechar_cursor_conexao(cur, con)


@app.context_processor
def contexto_permissoes_menus():
    return {
        'menu_lateral_dinamico': carregar_menu_lateral_usuario(),
        'pode_menu': usuario_tem_permissao_menu,
        'empresas_permitidas_usuario': empresas_permitidas_usuario,
    }


@app.before_request
def preparar_nucleo_permissoes_request():
    if 'usuario_id' in session:
        inicializar_nucleo_permissoes()


# ==========================================================
# CONSISTÊNCIA MULTIEMPRESA: PESSOA, USUÁRIO E MOVIMENTAÇÕES
# ==========================================================
def tabela_existe(cur, nome_tabela):
    """Retorna True se a tabela existir no banco atual."""
    try:
        cur.execute("SHOW TABLES LIKE %s", (nome_tabela,))
        return cur.fetchone() is not None
    except Exception as e:
        print(f"Erro ao verificar tabela {nome_tabela}: {e}")
        return False


def contar_movimentos(cur, tabela, condicao, params):
    """Conta registros de uma tabela somente quando ela existe."""
    try:
        if not tabela_existe(cur, tabela):
            return 0
        cur.execute(f"SELECT COUNT(*) AS total FROM {tabela} WHERE {condicao}", params)
        row = cur.fetchone() or {}
        return int(row.get('total') or 0)
    except Exception as e:
        print(f"Erro ao contar movimentações em {tabela}: {e}")
        return 0


def pessoa_tem_movimentacao_empresa(cur, pessoa_id, empresa_id):
    """
    Verifica se uma pessoa já possui movimentação operacional/financeira em uma empresa.

    Regra de negócio:
    - A mesma pessoa pode existir em várias empresas, mas como cadastros separados.
    - Não é seguro mover o MESMO cadastro de pessoa para outra empresa quando já existe histórico.
    - Usuário/login sozinho não é considerado movimentação; ele pode ser sincronizado quando não há histórico.
    """
    if not pessoa_id or not empresa_id:
        return {'total': 0, 'detalhes': []}

    pessoa_id = int(pessoa_id)
    empresa_id = int(empresa_id)

    verificacoes = [
        ('rotas', 'motorista_id = %s AND empresa_id = %s', (pessoa_id, empresa_id), 'rotas como motorista'),
        ('rotas', 'transportadora_id = %s AND empresa_id = %s', (pessoa_id, empresa_id), 'rotas como transportadora'),
        ('motorista_notas_fiscais', 'motorista_id = %s AND empresa_id = %s', (pessoa_id, empresa_id), 'documentos/NFs de motorista'),
        ('disponibilidade_motorista', 'motorista_id = %s AND empresa_id = %s', (pessoa_id, empresa_id), 'disponibilidade semanal'),
        ('escala_motorista', 'motorista_id = %s AND empresa_id = %s', (pessoa_id, empresa_id), 'escala de motoristas'),
        ('ciencia_escala_motorista', 'motorista_id = %s AND empresa_id = %s', (pessoa_id, empresa_id), 'ciência de escala'),
        ('fila_cancelados_base', 'motorista_id = %s AND empresa_id = %s', (pessoa_id, empresa_id), 'fila de cancelados'),
        ('auditoria_checkin_base', 'motorista_id = %s AND empresa_id = %s', (pessoa_id, empresa_id), 'auditoria de check-in'),
        ('justificativas_ausencia_motorista', 'motorista_id = %s AND empresa_id = %s', (pessoa_id, empresa_id), 'justificativas de ausência'),
        ('arquivos_sistema', '(motorista_id = %s OR pessoa_id = %s) AND empresa_id = %s', (pessoa_id, pessoa_id, empresa_id), 'arquivos do sistema'),
        ('lancamentos_ajudantes', 'ajudante_id = %s AND empresa_id = %s', (pessoa_id, empresa_id), 'lançamentos de ajudante'),
        ('titulos_financeiros', 'pessoa_id = %s AND empresa_id = %s', (pessoa_id, empresa_id), 'títulos financeiros'),
    ]

    detalhes = []
    total_geral = 0

    for tabela, condicao, params, descricao in verificacoes:
        total = contar_movimentos(cur, tabela, condicao, params)
        if total > 0:
            total_geral += total
            detalhes.append(f"{descricao}: {total}")

    return {'total': total_geral, 'detalhes': detalhes}


def buscar_pessoa_para_vinculo_usuario(cur, pessoa_id):
    if not pessoa_id:
        return None
    try:
        cur.execute("""
                    SELECT id, empresa_id, nome_completo, tipo_cadastro, status_cadastro
                    FROM pessoas
                    WHERE id = %s
                    LIMIT 1
                    """, (pessoa_id,))
        return cur.fetchone()
    except Exception as e:
        print(f"Erro ao buscar pessoa para vínculo de usuário: {e}")
        return None


def validar_base_operacional_usuario(cur, base_operacional_id, empresa_id):
    if not base_operacional_id or not empresa_id:
        return None
    try:
        cur.execute("""
                    SELECT id, nome_base, empresa_id, status_base
                    FROM bases_operacionais
                    WHERE id = %s
                      AND empresa_id = %s
                      AND status_base = 'Ativa'
                    LIMIT 1
                    """, (base_operacional_id, empresa_id))
        return cur.fetchone()
    except Exception as e:
        print(f"Erro ao validar base operacional do usuário: {e}")
        return None


# ==========================================================
# LOGIN / LOGOUT
# ==========================================================
@app.route('/arquivos/visualizar/<int:arquivo_id>')
@login_required
def visualizar_arquivo_sistema(arquivo_id):
    """
    Visualização protegida de arquivos do SGR.

    O arquivo permanece privado no Google Drive. O usuário acessa pelo sistema,
    o Flask valida permissão e entrega o conteúdo usando as credenciais do servidor.
    """
    con = obter_conexao()
    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('dashboard'))

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
            SELECT id, empresa_id, pessoa_id, motorista_id, origem, origem_id, tipo_arquivo,
                   nome_original, nome_armazenado, mime_type, tamanho_bytes,
                   storage_provider, caminho_local, drive_file_id, drive_view_url,
                   drive_download_url, status_arquivo
            FROM arquivos_sistema
            WHERE id = %s
              AND status_arquivo <> 'EXCLUIDO'
            LIMIT 1
        """, (arquivo_id,))

        arquivo = cur.fetchone()

        if not arquivo:
            flash("Arquivo não encontrado.", "danger")
            return redirect(url_for('dashboard'))

        perfil = session.get('perfil_de_acesso')
        empresa_sessao = int(session.get('empresa_id') or 0)
        pessoa_sessao = session.get('pessoa_id')

        if not usuario_eh_super_admin_global():
            if int(arquivo.get('empresa_id') or 0) != empresa_sessao:
                flash("Você não tem permissão para acessar este arquivo.", "danger")
                return redirect(url_for('dashboard'))

            # Terminal Base nunca visualiza arquivos/anexos.
            if perfil == 'Terminal Base':
                flash("Usuário Terminal Base não tem permissão para visualizar arquivos.", "danger")
                return redirect(url_for('terminal_base_qrcode'))

            # Motorista só visualiza arquivos explicitamente vinculados a ele.
            # Regra segura: se o arquivo não tiver motorista_id/pessoa_id preenchido, bloqueia.
            if perfil == 'Motorista':
                pessoa_sessao_int = int(pessoa_sessao or 0)
                ids_permitidos = set()
                if arquivo.get('motorista_id'):
                    ids_permitidos.add(int(arquivo.get('motorista_id')))
                if arquivo.get('pessoa_id'):
                    ids_permitidos.add(int(arquivo.get('pessoa_id')))

                if not pessoa_sessao_int or pessoa_sessao_int not in ids_permitidos:
                    flash("Você não tem permissão para acessar este arquivo.", "danger")
                    return redirect(url_for('portal_motorista'))

        nome_download = arquivo.get('nome_original') or arquivo.get('nome_armazenado') or f"arquivo_{arquivo_id}"
        mime_type = arquivo.get('mime_type') or 'application/octet-stream'

        if arquivo.get('drive_file_id'):
            dados = baixar_arquivo_google_drive(arquivo.get('drive_file_id'))
            headers = {
                'Content-Disposition': f'inline; filename="{nome_download}"',
                'Cache-Control': 'private, max-age=300',
            }
            return Response(dados, mimetype=mime_type, headers=headers)

        # Fallback para arquivos antigos ainda armazenados localmente.
        caminho_local = str(arquivo.get('caminho_local') or '').strip()
        if caminho_local:
            caminho_local = caminho_local.replace('\\', '/').lstrip('/')
            if caminho_local.startswith('static/'):
                caminho_local = caminho_local[len('static/'): ]
            caminho_absoluto = os.path.abspath(os.path.join(app.root_path, 'static', caminho_local))
            pasta_static = os.path.abspath(os.path.join(app.root_path, 'static'))
            if caminho_absoluto.startswith(pasta_static) and os.path.exists(caminho_absoluto):
                with open(caminho_absoluto, 'rb') as f:
                    dados = f.read()
                headers = {
                    'Content-Disposition': f'inline; filename="{nome_download}"',
                    'Cache-Control': 'private, max-age=300',
                }
                return Response(dados, mimetype=mime_type, headers=headers)

        flash("Arquivo não está disponível para visualização.", "danger")
        return redirect(url_for('dashboard'))

    except Exception as e:
        print(f"Erro ao visualizar arquivo {arquivo_id}: {e}")
        flash("Erro ao visualizar o arquivo.", "danger")
        return redirect(url_for('dashboard'))

    finally:
        cur.close()
        con.close()


@app.route('/arquivos/local/<path:caminho>')
@login_required
def visualizar_upload_local(caminho):
    """
    Visualização protegida para arquivos locais salvos em /uploads.

    Usado principalmente como fallback dos comprovantes financeiros quando o Google Drive
    não está habilitado ou quando o upload para o Drive falha.
    """
    perfil = session.get('perfil_de_acesso')
    if perfil == 'Terminal Base':
        flash('Usuário Terminal Base não tem permissão para visualizar arquivos.', 'danger')
        return redirect(url_for('dashboard'))

    caminho_limpo = str(caminho or '').replace('\\', '/').lstrip('/')

    if not caminho_limpo.startswith('uploads/'):
        flash('Arquivo inválido.', 'danger')
        return redirect(url_for('dashboard'))

    base_uploads = os.path.abspath(os.path.join(app.root_path, 'uploads'))
    caminho_absoluto = os.path.abspath(os.path.join(app.root_path, caminho_limpo))

    if not caminho_absoluto.startswith(base_uploads + os.sep) and caminho_absoluto != base_uploads:
        flash('Arquivo inválido.', 'danger')
        return redirect(url_for('dashboard'))

    if not os.path.exists(caminho_absoluto) or not os.path.isfile(caminho_absoluto):
        flash('Arquivo não encontrado no armazenamento local.', 'warning')
        return redirect(url_for('dashboard'))

    return send_file(
        caminho_absoluto,
        as_attachment=False,
        download_name=os.path.basename(caminho_absoluto),
        max_age=300,
    )


def primeira_rota_permitida_usuario(excluir_endpoints=None):
    """Retorna o primeiro href permitido no menu lateral do usuário atual.

    Usado no login e em bloqueios de rota para evitar loop/erro quando o
    dashboard não está liberado para o perfil.
    """
    excluir_endpoints = set(excluir_endpoints or [])
    try:
        grupos = carregar_menu_lateral_usuario() or []
        for grupo in grupos:
            for item in grupo.get('itens', []):
                candidatos = [item] + list(item.get('filhos') or [])
                for candidato in candidatos:
                    endpoint = candidato.get('endpoint')
                    href = candidato.get('href')
                    if endpoint in excluir_endpoints:
                        continue
                    if not href or href == '#':
                        continue
                    return href
    except Exception as e:
        print(f"[Permissões] Falha ao resolver primeira rota permitida: {e}")
    return None


def redirecionamento_inicial_usuario():
    """Define a página inicial real do usuário após login.

    Regra v21:
    - Super Admin/Suporte vai para o Dashboard.
    - Motorista e Terminal Base mantêm seus portais.
    - Usuário com Dashboard liberado abre o Dashboard.
    - Usuário sem Dashboard, mas com algum menu liberado, vai para o primeiro menu.
    - Usuário sem nenhum módulo liberado cai na tela neutra interna.
    """
    perfil = session.get('perfil_de_acesso')

    if perfil == 'Motorista':
        return redirect(url_for('portal_motorista'))

    if perfil == 'Terminal Base':
        return redirect(url_for('terminal_base_qrcode'))

    if usuario_pode_ver_escopo_global_sistema():
        return redirect(url_for('dashboard'))

    if usuario_tem_permissao_menu('dashboard', 'visualizar'):
        return redirect(url_for('dashboard'))

    destino = primeira_rota_permitida_usuario(excluir_endpoints={'dashboard'})
    if destino:
        return redirect(destino)

    return redirect(url_for('inicio_sem_modulos'))


@app.route('/inicio')
@login_required
def inicio_sem_modulos():
    """Tela inicial neutra para usuário autenticado sem módulos liberados."""
    nome = session.get('usuario_nome') or session.get('usuario_login') or 'Usuário'
    empresa = session.get('empresa_nome') or 'Empresa atual'
    perfil = session.get('perfil_exibicao') or session.get('perfil_de_acesso') or 'Perfil sem nome'
    html = f"""
    <!doctype html>
    <html lang="pt-br">
    <head>
        <meta charset="utf-8">
        <meta name="viewport" content="width=device-width, initial-scale=1">
        <title>SGR Web</title>
        <style>
            :root {{ --bg:#f5f7fb; --card:#ffffff; --text:#111827; --muted:#64748b; --brand:#0891b2; --line:#e5e7eb; }}
            * {{ box-sizing:border-box; }}
            body {{ margin:0; font-family: Arial, Helvetica, sans-serif; background:linear-gradient(135deg,#f8fafc,#eef6fb); color:var(--text); }}
            .shell {{ min-height:100vh; display:flex; align-items:center; justify-content:center; padding:32px; }}
            .card {{ width:100%; max-width:860px; background:var(--card); border:1px solid var(--line); border-radius:24px; padding:38px; box-shadow:0 24px 70px rgba(15,23,42,.12); }}
            .brand {{ display:flex; align-items:center; gap:14px; margin-bottom:26px; }}
            .logo {{ width:54px; height:54px; border-radius:16px; background:linear-gradient(135deg,#06b6d4,#0f172a); color:#fff; display:flex; align-items:center; justify-content:center; font-weight:900; letter-spacing:.5px; }}
            .brand small {{ color:var(--muted); display:block; margin-top:3px; }}
            h1 {{ margin:0 0 10px; font-size:30px; }}
            p {{ color:var(--muted); font-size:16px; line-height:1.6; margin:0 0 18px; }}
            .info {{ margin:24px 0; display:grid; grid-template-columns:repeat(auto-fit,minmax(180px,1fr)); gap:12px; }}
            .pill {{ border:1px solid var(--line); background:#f8fafc; border-radius:14px; padding:13px 15px; }}
            .pill span {{ display:block; color:var(--muted); font-size:12px; font-weight:700; text-transform:uppercase; letter-spacing:.04em; margin-bottom:4px; }}
            .pill strong {{ font-size:14px; }}
            .notice {{ background:#ecfeff; border:1px solid #a5f3fc; color:#155e75; padding:15px 16px; border-radius:16px; margin-top:10px; }}
            .actions {{ margin-top:26px; display:flex; gap:12px; flex-wrap:wrap; }}
            a {{ text-decoration:none; border-radius:12px; padding:12px 16px; font-weight:800; }}
            .primary {{ background:#0f172a; color:#fff; }}
            .secondary {{ background:#e5e7eb; color:#111827; }}
        </style>
    </head>
    <body>
        <main class="shell">
            <section class="card">
                <div class="brand">
                    <div class="logo">SGR</div>
                    <div>
                        <strong>SGR Web</strong>
                        <small>Sistema de Gerenciamento de Rotas</small>
                    </div>
                </div>
                <h1>Bem-vindo ao SGR Web</h1>
                <p>Seu acesso está ativo, mas nenhum módulo inicial foi liberado para este perfil. Assim que o administrador liberar permissões, os menus aparecerão automaticamente.</p>
                <div class="info">
                    <div class="pill"><span>Usuário</span><strong>{nome}</strong></div>
                    <div class="pill"><span>Empresa</span><strong>{empresa}</strong></div>
                    <div class="pill"><span>Perfil</span><strong>{perfil}</strong></div>
                </div>
                <div class="notice">Aguardando liberação de módulos pelo administrador da empresa.</div>
                <div class="actions">
                    <a class="primary" href="/logout">Sair</a>
                    <a class="secondary" href="/inicio">Atualizar</a>
                </div>
            </section>
        </main>
    </body>
    </html>
    """
    return Response(html, status=200, mimetype='text/html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if 'usuario_id' in session:
        return redirecionamento_inicial_usuario()

    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')

        if not username or not password:
            flash("Preencha usuário e senha para acessar.", "danger")
            return redirect(url_for('login'))

        con = obter_conexao()

        if con is None:
            flash("Erro de conexão com o banco de dados.", "danger")
            return redirect(url_for('login'))

        cur = con.cursor(dictionary=True)

        try:
            cur.execute("""
                        SELECT u.id                          AS usuario_id,
                               u.login,
                               u.senha_hash,
                               u.status_usuario,
                               u.perfil_de_acesso,
                               u.perfil_id,
                               u.empresa_id,
                               u.base_operacional_id,
                               COALESCE(u.is_super_admin, 0) AS is_super_admin,

                               p.id                          AS pessoa_id,
                               p.empresa_id                  AS empresa_pessoa_id,
                               p.nome_completo,
                               p.status_cadastro,

                               e.id                          AS empresa_id_confirmado,
                               e.razao_social                AS empresa_razao_social,
                               e.nome_fantasia               AS empresa_nome_fantasia,
                               e.slug                        AS empresa_slug,
                               e.status_empresa,
                               e.plano                       AS empresa_plano,
                               pa.codigo                     AS perfil_codigo_banco,
                               pa.nome                       AS perfil_nome_banco,
                               pa.empresa_id                 AS perfil_empresa_id
                        FROM usuarios u
                                 LEFT JOIN pessoas p ON p.id = u.pessoa_id
                                 LEFT JOIN empresas e ON e.id = u.empresa_id
                                 LEFT JOIN perfis_acesso pa ON pa.id = u.perfil_id
                        WHERE u.login = %s LIMIT 1
                        """, (username,))

            usuario = cur.fetchone()

            if not usuario:
                flash("Usuário ou senha incorretos.", "danger")
                return redirect(url_for('login'))

            if usuario['status_usuario'] != 'Ativo':
                flash("Este usuário está inativo. Procure o administrador do sistema.", "danger")
                return redirect(url_for('login'))

            # Usuários Terminal Base podem não ter pessoa vinculada.
            # Quando houver pessoa vinculada, validamos o status do cadastro normalmente.
            if usuario.get('pessoa_id') and usuario.get('status_cadastro') != 'Ativo':
                flash("O cadastro vinculado a este usuário está inativo.", "danger")
                return redirect(url_for('login'))

            if not check_password_hash(usuario['senha_hash'], password):
                flash("Usuário ou senha incorretos.", "danger")
                return redirect(url_for('login'))

            is_super_admin = int(usuario.get('is_super_admin') or 0)

            if not usuario.get('empresa_id'):
                flash(
                    "Este usuário não está vinculado a nenhuma empresa. "
                    "Procure o administrador do sistema.",
                    "danger"
                )
                return redirect(url_for('login'))

            if not usuario.get('empresa_id_confirmado'):
                flash(
                    "A empresa vinculada a este usuário não foi encontrada. "
                    "Procure o administrador do sistema.",
                    "danger"
                )
                return redirect(url_for('login'))

            if usuario['status_empresa'] != 'Ativa' and is_super_admin != 1:
                flash(
                    "A empresa vinculada a este usuário está inativa. "
                    "Procure o administrador do sistema.",
                    "danger"
                )
                return redirect(url_for('login'))

            if usuario.get('perfil_de_acesso') == 'Motorista':
                if not usuario.get('pessoa_id'):
                    flash(
                        "Usuário motorista sem pessoa vinculada. Procure o administrador do sistema.",
                        "danger"
                    )
                    return redirect(url_for('login'))

                if str(usuario.get('empresa_id')) != str(usuario.get('empresa_pessoa_id')):
                    flash(
                        "Cadastro inconsistente: a empresa do usuário motorista é diferente da empresa do cadastro da pessoa. "
                        "Acesse como administrador e ajuste o usuário/pessoa antes de continuar.",
                        "danger"
                    )
                    return redirect(url_for('login'))

            # Garante que a empresa do usuário tenha perfis próprios antes de montar a sessão.
            try:
                inicializar_nucleo_permissoes()
                sincronizar_perfis_padrao_empresas(cur, usuario.get('empresa_id'))
                perfil_resolvido = resolver_perfil_permissao_usuario(cur, usuario)
                con.commit()
            except Exception as e:
                print(f"[Permissões] Falha ao preparar perfil da empresa no login: {e}")
                perfil_resolvido = {'id': usuario.get('perfil_id'), 'codigo': usuario.get('perfil_de_acesso'), 'nome': usuario.get('perfil_de_acesso')}

            session.clear()

            session['usuario_id'] = usuario['usuario_id']
            session['usuario_login'] = usuario['login']
            session['usuario_nome'] = usuario.get('nome_completo') or usuario.get('login') or 'Terminal Base'
            session['perfil_de_acesso'] = usuario['perfil_de_acesso']
            session['perfil_id'] = perfil_resolvido.get('id') or perfil_resolvido.get('perfil_id') or usuario.get('perfil_id')
            session['perfil_permissao_codigo'] = perfil_resolvido.get('codigo') or usuario.get('perfil_de_acesso')
            session['perfil_exibicao'] = perfil_resolvido.get('nome') or usuario.get('perfil_de_acesso')
            session['pessoa_id'] = usuario.get('pessoa_id')
            session['base_operacional_id'] = usuario.get('base_operacional_id')

            session['empresa_id'] = usuario['empresa_id']
            session['empresa_nome'] = usuario['empresa_nome_fantasia'] or usuario['empresa_razao_social']
            session['empresa_slug'] = usuario['empresa_slug']
            session['empresa_plano'] = usuario['empresa_plano']
            session['is_super_admin'] = is_super_admin

            nome_boas_vindas = usuario.get('nome_completo') or usuario.get('login') or 'Terminal Base'
            flash(
                f"Bem-vindo, {nome_boas_vindas}! "
                f"Empresa: {session['empresa_nome']}.",
                "success"
            )

            if usuario.get('perfil_de_acesso') == 'Motorista':
                return redirect(url_for('portal_motorista'))

            if usuario.get('perfil_de_acesso') == 'Terminal Base':
                if not usuario.get('base_operacional_id'):
                    flash("Usuário Terminal Base sem base vinculada. Procure o administrador.", "danger")
                    return redirect(url_for('logout'))
                return redirect(url_for('terminal_base_qrcode'))

            return redirecionamento_inicial_usuario()

        except Exception as e:
            print(f"Erro no login: {e}")
            flash("Erro técnico ao realizar login.", "danger")
            return redirect(url_for('login'))

        finally:
            cur.close()
            con.close()

    return render_template('login.html')


def garantir_pessoa_id_sessao():
    """
    Garante que session['pessoa_id'] exista.
    Ajuda caso o login ainda não tenha sido ajustado.
    """
    pessoa_id = session.get('pessoa_id')

    if pessoa_id:
        return pessoa_id

    usuario_id = session.get('usuario_id')

    if not usuario_id:
        return None

    con = obter_conexao()

    if con is None:
        return None

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT pessoa_id
                    FROM usuarios
                    WHERE id = %s LIMIT 1
                    """, (usuario_id,))

        usuario = cur.fetchone()

        if usuario and usuario.get('pessoa_id'):
            session['pessoa_id'] = usuario['pessoa_id']
            return usuario['pessoa_id']

    except Exception as e:
        print(f"Erro ao garantir pessoa_id na sessão: {e}")

    finally:
        cur.close()
        con.close()

    return None


def motorista_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        perfil = session.get('perfil_de_acesso')
        is_super_admin = int(session.get('is_super_admin') or 0) == 1

        if perfil != 'Motorista' and not is_super_admin:
            flash("Acesso restrito ao Portal do Motorista.", "danger")
            return redirect(url_for('dashboard'))

        return f(*args, **kwargs)

    return decorated_function


@app.route('/logout')
def logout():
    session.clear()
    flash('Sessão encerrada com sucesso.', 'info')
    return redirect(url_for('login'))


# ==========================================================
# DASHBOARD EXECUTIVO GERAL - BLOCO 11
# ==========================================================
@app.route('/')
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro', 'Consulta')
def dashboard():
    if session.get('perfil_de_acesso') == 'Terminal Base':
        return redirect(url_for('terminal_base_qrcode'))

    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()
    perfil_logado = perfil_usuario_logado()

    # Bloco 11.1: base inicial para Dashboard controlado por perfil.
    # A etapa futura deve evoluir isso para permissões granulares por usuário/módulo/empresa,
    # mas este mapa já evita exposição visual desnecessária no dashboard executivo.
    dashboard_acl = {
        'financeiro': is_super_admin or perfil_logado in ['Administrador', 'Financeiro'],
        'operacao': is_super_admin or perfil_logado in ['Administrador', 'Operacional'],
        'documentos': is_super_admin or perfil_logado in ['Administrador', 'Operacional', 'Financeiro'],
        'prestadores': is_super_admin or perfil_logado in ['Administrador', 'Operacional', 'Financeiro'],
        'auditoria': is_super_admin or perfil_logado in ['Administrador'],
        'relatorios': is_super_admin or perfil_logado in ['Administrador', 'Financeiro', 'Operacional', 'Consulta'],
    }

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    # Filtros globais do dashboard. Reaproveita o mesmo padrão da Central de Relatórios.
    try:
        periodo, data_inicio_dt, data_fim_dt = _relatorios_datas_request()
    except Exception:
        hoje = date.today()
        periodo, data_inicio_dt, data_fim_dt = 'mes_atual', hoje.replace(day=1), hoje

    empresa_id_filtro = (request.args.get('empresa_id') or '').strip()
    if is_super_admin and empresa_id_filtro and empresa_id_filtro.isdigit():
        empresa_consulta_id = int(empresa_id_filtro)
    elif is_super_admin and not empresa_id_filtro:
        empresa_consulta_id = None
    else:
        empresa_consulta_id = int(empresa_logada_id)
        empresa_id_filtro = str(empresa_logada_id)

    filtro_empresa_sql = 'empresa_id = %s' if empresa_consulta_id else 'empresa_id IS NOT NULL'
    filtro_empresa_params = [empresa_consulta_id] if empresa_consulta_id else []

    filtros = {
        'periodo': periodo,
        'data_inicio': data_inicio_dt.strftime('%Y-%m-%d'),
        'data_fim': data_fim_dt.strftime('%Y-%m-%d'),
        'empresa_id': empresa_id_filtro,
    }
    session['dashboard_executivo_filtros'] = {k: v for k, v in filtros.items() if v not in [None, '']}

    # Estrutura padrão. Mantém a página funcionando mesmo se alguma tabela ainda não existir no ambiente.
    dashboard_data = {
        'financeiro': {
            'total_pagar': 0, 'total_receber': 0, 'total_baixado': 0, 'saldo_liquido': 0,
            'total_aberto': 0, 'total_vencido': 0, 'total_estornado': 0, 'total_cancelado': 0,
            'qtd_titulos': 0, 'qtd_vencidos': 0,
        },
        'operacao': {
            'total_rotas': 0, 'valor_rotas': 0, 'rotas_pendentes': 0, 'rotas_faturadas': 0,
            'rotas_quitadas': 0, 'rotas_divergencia': 0, 'rotas_bloqueadas': 0, 'rotas_liberadas_nf': 0,
            'rotas_aprovadas_pagamento': 0,
        },
        'prestadores': {
            'ativos': 0, 'pagamentos_pendentes': 0, 'pagamentos_quitados': 0, 'pagamentos_estornados': 0,
            'valor_pendente': 0, 'valor_quitado': 0, 'valor_estornado': 0,
        },
        'documentos': {
            'uploads': 0, 'nfs_enviadas': 0, 'documentos_aprovados': 0, 'documentos_recusados': 0,
            'documentos_pendentes': 0, 'sem_nf': 0,
        },
        'auditoria': {
            'acoes': 0, 'baixas': 0, 'estornos': 0, 'cancelamentos': 0, 'alteracoes_criticas': 0,
        },
        'rankings': {
            'prestadores_valor': [],
            'origens_financeiras': [],
            'status_rotas': [],
            'usuarios_auditoria': [],
        },
        'alertas': [],
    }

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão ao carregar o dashboard executivo.", "danger")
        return render_template(
            'dashboard.html',
            cards=dashboard_data,
            filtros=filtros,
            empresas=[],
            is_super_admin=is_super_admin,
            usuario_logado=usuario_logado,
            dashboard_acl=dashboard_acl
        )

    cur = con.cursor(dictionary=True)

    def safe_fetchone(sql, params=None, fallback=None, label='dashboard'):
        try:
            cur.execute(sql, params or [])
            return cur.fetchone() or (fallback or {})
        except Exception as exc:
            print(f"[Dashboard Executivo] Falha ao carregar {label}: {exc}")
            return fallback or {}

    def safe_fetchall(sql, params=None, label='dashboard'):
        try:
            cur.execute(sql, params or [])
            return cur.fetchall() or []
        except Exception as exc:
            print(f"[Dashboard Executivo] Falha ao carregar {label}: {exc}")
            return []

    try:
        if dashboard_acl.get('financeiro'):
            # ==========================================================
            # FINANCEIRO - títulos do período por vencimento
            # ==========================================================
            financeiro = safe_fetchone(f"""
                SELECT
                    COUNT(*) AS qtd_titulos,
                    COALESCE(SUM(CASE WHEN tipo_titulo = 'PAGAR' THEN valor_liquido ELSE 0 END), 0) AS total_pagar,
                    COALESCE(SUM(CASE WHEN tipo_titulo = 'RECEBER' THEN valor_liquido ELSE 0 END), 0) AS total_receber,
                    COALESCE(SUM(CASE WHEN status_titulo IN ('Pago','Recebido') THEN COALESCE(valor_baixado, valor_liquido) ELSE 0 END), 0) AS total_baixado,
                    COALESCE(SUM(CASE WHEN status_titulo IN ('Aberto','Solicitado') THEN valor_liquido ELSE 0 END), 0) AS total_aberto,
                    COALESCE(SUM(CASE WHEN status_titulo NOT IN ('Pago','Recebido','Cancelado','Estornado') AND data_vencimento < CURDATE() THEN valor_liquido ELSE 0 END), 0) AS total_vencido,
                    COALESCE(SUM(CASE WHEN status_titulo = 'Estornado' THEN valor_liquido ELSE 0 END), 0) AS total_estornado,
                    COALESCE(SUM(CASE WHEN status_titulo = 'Cancelado' THEN valor_liquido ELSE 0 END), 0) AS total_cancelado,
                    SUM(CASE WHEN status_titulo NOT IN ('Pago','Recebido','Cancelado','Estornado') AND data_vencimento < CURDATE() THEN 1 ELSE 0 END) AS qtd_vencidos,
                    COALESCE(SUM(CASE WHEN tipo_titulo = 'RECEBER' THEN valor_liquido ELSE -valor_liquido END), 0) AS saldo_liquido
                FROM titulos_financeiros
                WHERE {filtro_empresa_sql}
                  AND data_vencimento BETWEEN %s AND %s
            """, filtro_empresa_params + [data_inicio_dt, data_fim_dt], dashboard_data['financeiro'], 'financeiro')
            dashboard_data['financeiro'].update(financeiro)

        if dashboard_acl.get('operacao'):
            # ==========================================================
            # OPERAÇÃO - rotas do período por data_lancamento
            # ==========================================================
            operacao = safe_fetchone(f"""
                SELECT
                    COUNT(*) AS total_rotas,
                    COALESCE(SUM(COALESCE(valor_rota,0) + COALESCE(valor_km,0) + COALESCE(outras_despesas,0)), 0) AS valor_rotas,
                    SUM(CASE WHEN situacao_rota = 'Pendente' THEN 1 ELSE 0 END) AS rotas_pendentes,
                    SUM(CASE WHEN situacao_rota = 'Faturada' THEN 1 ELSE 0 END) AS rotas_faturadas,
                    SUM(CASE WHEN situacao_rota = 'Quitada' THEN 1 ELSE 0 END) AS rotas_quitadas,
                    SUM(CASE WHEN status_motorista = 'Divergência apontada' THEN 1 ELSE 0 END) AS rotas_divergencia,
                    SUM(CASE WHEN status_motorista LIKE '%bloquead%' OR status_motorista LIKE '%Bloquead%' THEN 1 ELSE 0 END) AS rotas_bloqueadas,
                    SUM(CASE WHEN status_motorista = 'Liberada para NF' THEN 1 ELSE 0 END) AS rotas_liberadas_nf,
                    SUM(CASE WHEN status_motorista = 'Aprovada para pagamento' THEN 1 ELSE 0 END) AS rotas_aprovadas_pagamento
                FROM rotas
                WHERE {filtro_empresa_sql}
                  AND data_lancamento BETWEEN %s AND %s
                  AND COALESCE(situacao_rota, '') <> 'Cancelada'
            """, filtro_empresa_params + [data_inicio_dt, data_fim_dt], dashboard_data['operacao'], 'operação')
            dashboard_data['operacao'].update(operacao)

        if dashboard_acl.get('prestadores'):
            # ==========================================================
            # PRESTADORES/AJUDANTES
            # ==========================================================
            prestadores = safe_fetchone(f"""
                SELECT
                    COUNT(DISTINCT CASE WHEN p.tipo_cadastro IN ('Motorista','Prestador','Ajudante') THEN p.id ELSE NULL END) AS ativos
                FROM pessoas p
                WHERE {'p.empresa_id = %s' if empresa_consulta_id else 'p.empresa_id IS NOT NULL'}
                  AND COALESCE(p.status_cadastro, 'Ativo') = 'Ativo'
            """, filtro_empresa_params, {'ativos': 0}, 'prestadores ativos')
            dashboard_data['prestadores'].update(prestadores)

            ajudantes = safe_fetchone(f"""
                SELECT
                    SUM(CASE WHEN status_pagamento = 'Pendente' THEN 1 ELSE 0 END) AS pagamentos_pendentes,
                    SUM(CASE WHEN status_pagamento = 'Quitado' THEN 1 ELSE 0 END) AS pagamentos_quitados,
                    SUM(CASE WHEN status_pagamento = 'Estornado' THEN 1 ELSE 0 END) AS pagamentos_estornados,
                    COALESCE(SUM(CASE WHEN status_pagamento = 'Pendente' THEN valor_total ELSE 0 END), 0) AS valor_pendente,
                    COALESCE(SUM(CASE WHEN status_pagamento = 'Quitado' THEN valor_total ELSE 0 END), 0) AS valor_quitado,
                    COALESCE(SUM(CASE WHEN status_pagamento = 'Estornado' THEN valor_total ELSE 0 END), 0) AS valor_estornado
                FROM lancamentos_ajudantes
                WHERE {filtro_empresa_sql}
                  AND data_lancamento BETWEEN %s AND %s
            """, filtro_empresa_params + [data_inicio_dt, data_fim_dt], dashboard_data['prestadores'], 'ajudantes')
            dashboard_data['prestadores'].update(ajudantes)

        if dashboard_acl.get('documentos'):
            # ==========================================================
            # DOCUMENTOS / DRIVE
            # ==========================================================
            uploads = safe_fetchone(f"""
                SELECT COUNT(*) AS uploads
                FROM arquivos_sistema
                WHERE {filtro_empresa_sql}
                  AND DATE(created_at) BETWEEN %s AND %s
            """, filtro_empresa_params + [data_inicio_dt, data_fim_dt], {'uploads': 0}, 'uploads')
            dashboard_data['documentos'].update(uploads)

            docs = safe_fetchone(f"""
                SELECT
                    SUM(CASE WHEN status_nf IN ('Enviada','Em análise') THEN 1 ELSE 0 END) AS documentos_pendentes,
                    SUM(CASE WHEN status_nf = 'Aprovada' THEN 1 ELSE 0 END) AS documentos_aprovados,
                    SUM(CASE WHEN status_nf = 'Recusada' THEN 1 ELSE 0 END) AS documentos_recusados,
                    SUM(CASE WHEN status_nf = 'Enviada' THEN 1 ELSE 0 END) AS nfs_enviadas,
                    SUM(CASE WHEN tipo_documento_pagamento = 'SEM_NF' THEN 1 ELSE 0 END) AS sem_nf
                FROM motorista_notas_fiscais
                WHERE {filtro_empresa_sql}
                  AND DATE(data_envio) BETWEEN %s AND %s
            """, filtro_empresa_params + [data_inicio_dt, data_fim_dt], dashboard_data['documentos'], 'documentos')
            dashboard_data['documentos'].update(docs)

        if dashboard_acl.get('auditoria'):
            # ==========================================================
            # AUDITORIA
            # ==========================================================
            auditoria = safe_fetchone(f"""
                SELECT
                    COUNT(*) AS acoes,
                    SUM(CASE WHEN acao LIKE '%baixa%' OR acao LIKE '%BAIXA%' THEN 1 ELSE 0 END) AS baixas,
                    SUM(CASE WHEN acao LIKE '%estorno%' OR acao LIKE '%ESTORNO%' THEN 1 ELSE 0 END) AS estornos,
                    SUM(CASE WHEN acao LIKE '%cancel%' OR acao LIKE '%CANCEL%' THEN 1 ELSE 0 END) AS cancelamentos,
                    SUM(CASE WHEN modulo IN ('CONFIGURACOES_FINANCEIRAS','FINANCEIRO_CONFIGURACOES','DOCUMENTOS_MOTORISTAS') THEN 1 ELSE 0 END) AS alteracoes_criticas
                FROM auditoria_financeira
                WHERE {filtro_empresa_sql}
                  AND DATE(created_at) BETWEEN %s AND %s
            """, filtro_empresa_params + [data_inicio_dt, data_fim_dt], dashboard_data['auditoria'], 'auditoria')
            dashboard_data['auditoria'].update(auditoria)

        # ==========================================================
        # RANKINGS E LISTAS GERENCIAIS
        # ==========================================================
        if dashboard_acl.get('prestadores') or dashboard_acl.get('financeiro'):
            dashboard_data['rankings']['prestadores_valor'] = safe_fetchall(f"""
                SELECT COALESCE(p.nome_completo, 'Sem pessoa vinculada') AS nome,
                       COALESCE(p.cpf_cnpj, '') AS documento,
                       COUNT(*) AS quantidade,
                       COALESCE(SUM(t.valor_liquido), 0) AS valor
                FROM titulos_financeiros t
                LEFT JOIN pessoas p ON p.id = t.pessoa_id
                WHERE {'t.empresa_id = %s' if empresa_consulta_id else 't.empresa_id IS NOT NULL'}
                  AND t.data_vencimento BETWEEN %s AND %s
                GROUP BY p.id, p.nome_completo, p.cpf_cnpj
                ORDER BY valor DESC
                LIMIT 5
            """, filtro_empresa_params + [data_inicio_dt, data_fim_dt], 'ranking prestadores')

        if dashboard_acl.get('financeiro'):
            dashboard_data['rankings']['origens_financeiras'] = safe_fetchall(f"""
                SELECT COALESCE(t.origem, 'Sem origem') AS origem,
                       COUNT(*) AS quantidade,
                       COALESCE(SUM(t.valor_liquido), 0) AS valor
                FROM titulos_financeiros t
                WHERE {'t.empresa_id = %s' if empresa_consulta_id else 't.empresa_id IS NOT NULL'}
                  AND t.data_vencimento BETWEEN %s AND %s
                GROUP BY t.origem
                ORDER BY valor DESC
                LIMIT 5
            """, filtro_empresa_params + [data_inicio_dt, data_fim_dt], 'origens financeiras')

        if dashboard_acl.get('operacao'):
            dashboard_data['rankings']['status_rotas'] = safe_fetchall(f"""
                SELECT COALESCE(situacao_rota, 'Sem status') AS status,
                       COUNT(*) AS quantidade,
                       COALESCE(SUM(COALESCE(valor_rota,0) + COALESCE(valor_km,0) + COALESCE(outras_despesas,0)), 0) AS valor
                FROM rotas
                WHERE {filtro_empresa_sql}
                  AND data_lancamento BETWEEN %s AND %s
                GROUP BY situacao_rota
                ORDER BY quantidade DESC
                LIMIT 5
            """, filtro_empresa_params + [data_inicio_dt, data_fim_dt], 'status rotas')

        if dashboard_acl.get('auditoria'):
            dashboard_data['rankings']['usuarios_auditoria'] = safe_fetchall(f"""
                SELECT COALESCE(p.nome_completo, u.login, CONCAT('Usuário #', a.usuario_id), 'Sem usuário') AS nome,
                       COUNT(*) AS quantidade
                FROM auditoria_financeira a
                LEFT JOIN usuarios u ON u.id = a.usuario_id
                LEFT JOIN pessoas p ON p.id = u.pessoa_id
                WHERE {'a.empresa_id = %s' if empresa_consulta_id else 'a.empresa_id IS NOT NULL'}
                  AND DATE(a.created_at) BETWEEN %s AND %s
                GROUP BY a.usuario_id, p.nome_completo, u.login
                ORDER BY quantidade DESC
                LIMIT 5
            """, filtro_empresa_params + [data_inicio_dt, data_fim_dt], 'usuários auditoria')

        # Alertas críticos do período
        if dashboard_acl.get('financeiro') and int(dashboard_data['financeiro'].get('qtd_vencidos') or 0) > 0:
            dashboard_data['alertas'].append({
                'tipo': 'danger',
                'titulo': 'Títulos vencidos',
                'texto': f"{int(dashboard_data['financeiro'].get('qtd_vencidos') or 0)} título(s) vencido(s) no período.",
                'url': url_for('relatorios_financeiro', status_titulo='', periodo=periodo, data_inicio=filtros['data_inicio'], data_fim=filtros['data_fim'])
            })
        if dashboard_acl.get('operacao') and int(dashboard_data['operacao'].get('rotas_divergencia') or 0) > 0:
            dashboard_data['alertas'].append({
                'tipo': 'warning',
                'titulo': 'Rotas com divergência',
                'texto': f"{int(dashboard_data['operacao'].get('rotas_divergencia') or 0)} rota(s) com divergência para acompanhar.",
                'url': url_for('rotas_divergencias')
            })
        if dashboard_acl.get('documentos') and int(dashboard_data['documentos'].get('documentos_recusados') or 0) > 0:
            dashboard_data['alertas'].append({
                'tipo': 'warning',
                'titulo': 'Documentos recusados',
                'texto': f"{int(dashboard_data['documentos'].get('documentos_recusados') or 0)} documento(s) recusado(s) no período.",
                'url': url_for('financeiro_nfs_motoristas')
            })

        empresas = []
        if is_super_admin:
            empresas = safe_fetchall('SELECT id, nome_fantasia, razao_social FROM empresas ORDER BY nome_fantasia ASC', [], 'empresas')

        return render_template(
            'dashboard.html',
            cards=dashboard_data,
            filtros=filtros,
            empresas=empresas,
            is_super_admin=is_super_admin,
            usuario_logado=usuario_logado,
            dashboard_acl=dashboard_acl
        )

    except Exception as e:
        print(f"Erro ao carregar Dashboard Executivo: {e}")
        flash("Erro ao carregar algumas métricas do dashboard executivo.", "warning")
        return render_template(
            'dashboard.html',
            cards=dashboard_data,
            filtros=filtros,
            empresas=[],
            is_super_admin=is_super_admin,
            usuario_logado=usuario_logado,
            dashboard_acl=dashboard_acl
        )
    finally:
        fechar_cursor_conexao(cur, con)


@app.route('/portal-motorista' , methods=['GET'])
@login_required
@motorista_required
def portal_motorista():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_id = session.get('empresa_id')
    perfil = session.get('perfil_de_acesso')
    is_super_admin = int(session.get('is_super_admin') or 0) == 1

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    pessoa_id = garantir_pessoa_id_sessao()

    if not pessoa_id:
        flash(
            "Seu usuário não está vinculado a uma pessoa/motorista. "
            "Peça ao administrador para vincular seu usuário ao cadastro de pessoa.",
            "danger"
        )
        return redirect(url_for('logout'))

    con_validacao = obter_conexao()
    if con_validacao is None:
        flash("Erro de conexão ao validar o cadastro do motorista.", "danger")
        return redirect(url_for('logout'))

    cur_validacao = con_validacao.cursor(dictionary=True)
    try:
        cur_validacao.execute("""
                              SELECT id, empresa_id, status_cadastro, tipo_cadastro
                              FROM pessoas
                              WHERE id = %s
                              LIMIT 1
                              """, (pessoa_id,))
        pessoa_validada = cur_validacao.fetchone()

        if not pessoa_validada or str(pessoa_validada.get('empresa_id')) != str(empresa_id):
            flash(
                "Cadastro inconsistente: o motorista não pertence à empresa do usuário logado. "
                "A sessão foi encerrada por segurança.",
                "danger"
            )
            return redirect(url_for('logout'))

        if pessoa_validada.get('status_cadastro') != 'Ativo':
            flash("Cadastro do motorista está inativo. Procure o administrador.", "danger")
            return redirect(url_for('logout'))

    except Exception as e:
        print(f"Erro ao validar vínculo do motorista no portal: {e}")
        flash("Erro ao validar vínculo do motorista. Faça login novamente.", "danger")
        return redirect(url_for('logout'))
    finally:
        fechar_cursor_conexao(cur_validacao, con_validacao)

    identi_rota = request.args.get('identi_rota', '').strip()
    status_motorista = request.args.get('status_motorista', '').strip()
    data_inicio = request.args.get('data_inicio', '').strip()
    data_fim = request.args.get('data_fim', '').strip()

    status_visiveis = [
        'Liberada para NF',
        'NF enviada',
        'Em análise',
        'Aprovada para pagamento',
        'Pagamento confirmado',
        'Bloqueada'
    ]

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('dashboard'))

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT id,
                           nome_completo,
                           cpf_cnpj,
                           tipo_cadastro,
                           tipo_prestador,
                           status_cadastro
                    FROM pessoas
                    WHERE id = %s
                      AND empresa_id = %s
                      AND (tipo_cadastro = 'Motorista' OR (tipo_cadastro = 'Prestador de Serviço' AND COALESCE(tipo_prestador, '') IN ('Motorista', 'Motorista e Ajudante'))) LIMIT 1
                    """, (pessoa_id, empresa_id))

        motorista = cur.fetchone()

        if not motorista and not is_super_admin:
            flash(
                "Seu usuário não está vinculado a um cadastro do tipo Motorista nesta empresa.",
                "danger"
            )
            return redirect(url_for('dashboard'))

        query = """
                SELECT r.id,
                       r.empresa_id,
                       r.data_lancamento,
                       r.identi_rota,
                       r.tipo_rota,
                       r.valor_rota,
                       r.valor_km,
                       r.outras_despesas,
                       (COALESCE(r.valor_rota, 0) + COALESCE(r.valor_km, 0) +
                        COALESCE(r.outras_despesas, 0))                     AS valor_total_rota,
                       r.transportadora_id,
                       transp.nome_completo                                 AS transportadora_nome,
                       transp.cpf_cnpj                                      AS transportadora_cpf_cnpj,
                       r.motorista_id,
                       mot.nome_completo                                    AS motorista_nome,
                       r.situacao_rota,
                       COALESCE(r.status_motorista, 'Aguardando conferência') AS status_motorista
                FROM rotas r
                         LEFT JOIN pessoas transp
                                   ON transp.id = r.transportadora_id
                                       AND transp.empresa_id = r.empresa_id
                         LEFT JOIN pessoas mot
                                   ON mot.id = r.motorista_id
                                       AND mot.empresa_id = r.empresa_id
                WHERE r.empresa_id = %s
                  AND r.motorista_id = %s
                  AND COALESCE(r.status_motorista, 'Aguardando conferência') IN (
                                                                               'Liberada para NF',
                                                                               'NF enviada',
                                                                               'Em análise',
                                                                               'Aprovada para pagamento',
                                                                               'Pagamento confirmado',
                                                                               'Bloqueada'
                    ) \
                """

        params = [empresa_id, pessoa_id]

        if identi_rota:
            query += " AND r.identi_rota LIKE %s"
            params.append(f"%{identi_rota}%")

        if status_motorista in status_visiveis:
            query += " AND COALESCE(r.status_motorista, 'Aguardando conferência') = %s"
            params.append(status_motorista)

        if data_inicio:
            query += " AND r.data_lancamento >= %s"
            params.append(data_inicio)

        if data_fim:
            query += " AND r.data_lancamento <= %s"
            params.append(data_fim)

        query += " ORDER BY r.data_lancamento DESC, r.id DESC"

        cur.execute(query, params)
        rotas = cur.fetchall()

        resumo = {
            'liberadas': 0,
            'nf_enviada': 0,
            'em_analise': 0,
            'aprovadas': 0,
            'pagas': 0,
            'valor_liberado': 0
        }

        for rota in rotas:
            sm = rota.get('status_motorista') or ''

            if sm == 'Liberada para NF':
                resumo['liberadas'] += 1
                resumo['valor_liberado'] += float(rota.get('valor_total_rota') or 0)
            elif sm == 'NF enviada':
                resumo['nf_enviada'] += 1
            elif sm == 'Em análise':
                resumo['em_analise'] += 1
            elif sm == 'Aprovada para pagamento':
                resumo['aprovadas'] += 1
            elif sm == 'Pagamento confirmado':
                resumo['pagas'] += 1

    except Exception as e:
        print(f"Erro ao carregar Portal do Motorista: {e}")
        flash("Erro técnico ao carregar Portal do Motorista.", "danger")
        rotas = []
        motorista = None
        resumo = {
            'liberadas': 0,
            'nf_enviada': 0,
            'em_analise': 0,
            'aprovadas': 0,
            'pagas': 0,
            'valor_liberado': 0
        }

    finally:
        cur.close()
        con.close()

    filtros = {
        'identi_rota': identi_rota,
        'status_motorista': status_motorista,
        'data_inicio': data_inicio,
        'data_fim': data_fim
    }

    return render_template(
        'portal_motorista.html',
        usuario_logado=usuario_logado,
        motorista=motorista,
        rotas=rotas,
        resumo=resumo,
        filtros=filtros
    )


def nome_arquivo_seguro_xml_motorista(nome_arquivo):
    nome_arquivo = str(nome_arquivo or 'nota_motorista.xml').strip()
    nome_arquivo = nome_arquivo.replace('\\', '_').replace('/', '_')
    nome_arquivo = re.sub(r'[^a-zA-Z0-9_.-]+', '_', nome_arquivo)

    if not nome_arquivo.lower().endswith('.xml'):
        nome_arquivo += '.xml'

    return nome_arquivo


def pasta_upload_xml_motorista():
    pasta = os.path.join(app.root_path, 'uploads', 'motorista_xmls')
    os.makedirs(pasta, exist_ok=True)
    return pasta


def buscar_motorista_logado():
    empresa_id = session.get('empresa_id')
    pessoa_id = garantir_pessoa_id_sessao()

    if not empresa_id or not pessoa_id:
        return None

    con = obter_conexao()

    if con is None:
        return None

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT id,
                           empresa_id,
                           nome_completo,
                           cpf_cnpj,
                           tipo_cadastro,
                           tipo_prestador,
                           status_cadastro
                    FROM pessoas
                    WHERE id = %s
                      AND empresa_id = %s
                      AND status_cadastro = 'Ativo'
                      AND (tipo_cadastro = 'Motorista' OR (tipo_cadastro = 'Prestador de Serviço' AND COALESCE(tipo_prestador, '') IN ('Motorista', 'Motorista e Ajudante'))) LIMIT 1
                    """, (pessoa_id, empresa_id))

        return cur.fetchone()

    except Exception as e:
        print(f"Erro ao buscar motorista logado: {e}")
        return None

    finally:
        cur.close()
        con.close()



@app.route('/portal-motorista/minhas-rotas')
@login_required
@motorista_required
def minhas_rotas_motorista():
    usuario_logado = session.get('usuario_nome', 'Motorista')
    empresa_id = session.get('empresa_id')
    motorista = buscar_motorista_logado()

    if not empresa_id or not motorista:
        flash("Não foi possível identificar o motorista logado.", "danger")
        return redirect(url_for('portal_motorista'))

    status_filtro = request.args.get('status_motorista', '').strip()
    identi_rota = request.args.get('identi_rota', '').strip()
    data_inicio = request.args.get('data_inicio', '').strip()
    data_fim = request.args.get('data_fim', '').strip()

    status_disponiveis = [
        'Aguardando conferência',
        'Conferida pelo motorista',
        'Divergência apontada',
        'Liberada para NF',
        'NF enviada',
        'Em análise',
        'Aprovada para pagamento',
        'Pagamento confirmado',
        'Bloqueada para correção',
        'Cancelada'
    ]

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('portal_motorista'))

    cur = con.cursor(dictionary=True)

    try:
        query = """
            SELECT
                r.id,
                r.empresa_id,
                r.data_lancamento,
                r.identi_rota,
                r.tipo_rota,
                r.valor_rota,
                r.valor_km,
                r.outras_despesas,
                (COALESCE(r.valor_rota, 0) + COALESCE(r.valor_km, 0) + COALESCE(r.outras_despesas, 0)) AS valor_total_rota,
                r.transportadora_id,
                COALESCE(transp.nome_completo, emp.razao_social, emp.nome_fantasia) AS tomador_nome,
                COALESCE(transp.cpf_cnpj, emp.cnpj) AS tomador_cpf_cnpj,
                r.motorista_id,
                COALESCE(r.status_motorista, 'Aguardando conferência') AS status_motorista,
                r.situacao_rota,
                r.data_aprovacao_motorista,
                r.data_inclusao
            FROM rotas r
            LEFT JOIN pessoas transp
                ON transp.id = r.transportadora_id
               AND transp.empresa_id = r.empresa_id
            LEFT JOIN empresas emp
                ON emp.id = r.empresa_id
            WHERE r.empresa_id = %s
              AND r.motorista_id = %s
        """
        params = [empresa_id, motorista['id']]

        if identi_rota:
            query += " AND r.identi_rota LIKE %s"
            params.append(f"%{identi_rota}%")

        if status_filtro in status_disponiveis:
            query += " AND COALESCE(r.status_motorista, 'Aguardando conferência') = %s"
            params.append(status_filtro)

        if data_inicio:
            query += " AND r.data_lancamento >= %s"
            params.append(data_inicio)

        if data_fim:
            query += " AND r.data_lancamento <= %s"
            params.append(data_fim)

        query += " ORDER BY r.data_lancamento DESC, r.id DESC"

        cur.execute(query, params)
        rotas = cur.fetchall()

        resumo = {
            'aguardando': 0,
            'conferidas': 0,
            'liberadas': 0,
            'nf_enviada': 0,
            'pagas': 0,
            'canceladas': 0,
            'valor_total': 0
        }

        for rota in rotas:
            status = rota.get('status_motorista') or 'Aguardando conferência'
            situacao_rota = rota.get('situacao_rota') or ''

            # Se a operação cancelou a rota, o portal precisa refletir Cancelada,
            # mesmo que algum status antigo tenha ficado gravado por histórico.
            if situacao_rota == 'Cancelada' or status == 'Cancelada':
                status = 'Cancelada'
                rota['status_motorista'] = 'Cancelada'

            if status == 'Aguardando conferência':
                resumo['aguardando'] += 1
            elif status == 'Conferida pelo motorista':
                resumo['conferidas'] += 1
            elif status == 'Liberada para NF':
                resumo['liberadas'] += 1
            elif status == 'NF enviada':
                resumo['nf_enviada'] += 1
            elif status == 'Pagamento confirmado':
                resumo['pagas'] += 1
            elif status == 'Cancelada':
                resumo['canceladas'] += 1

            # Rotas canceladas não entram no valor ativo do motorista.
            if status != 'Cancelada':
                resumo['valor_total'] += float(rota.get('valor_total_rota') or 0)

    except Exception as e:
        print(f"Erro ao carregar Minhas Rotas do motorista: {e}")
        flash("Erro técnico ao carregar suas rotas.", "danger")
        rotas = []
        resumo = {'aguardando': 0, 'conferidas': 0, 'liberadas': 0, 'nf_enviada': 0, 'pagas': 0, 'canceladas': 0, 'valor_total': 0}

    finally:
        cur.close()
        con.close()

    filtros = {
        'status_motorista': status_filtro,
        'identi_rota': identi_rota,
        'data_inicio': data_inicio,
        'data_fim': data_fim
    }

    return render_template(
        'minhas_rotas_motorista.html',
        usuario_logado=usuario_logado,
        motorista=motorista,
        rotas=rotas,
        resumo=resumo,
        filtros=filtros,
        status_disponiveis=status_disponiveis
    )


@app.route('/portal-motorista/minhas-rotas/<int:rota_id>/confirmar', methods=['POST'])
@login_required
@motorista_required
def confirmar_rota_motorista(rota_id):
    empresa_id = session.get('empresa_id')
    motorista = buscar_motorista_logado()
    usuario_id = session.get('usuario_id')

    if not empresa_id or not motorista:
        flash("Não foi possível identificar o motorista logado.", "danger")
        return redirect(url_for('portal_motorista'))

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('minhas_rotas_motorista'))

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
            SELECT
                id,
                empresa_id,
                identi_rota,
                motorista_id,
                COALESCE(status_motorista, 'Aguardando conferência') AS status_motorista,
                situacao_rota
            FROM rotas
            WHERE id = %s
              AND empresa_id = %s
              AND motorista_id = %s
            LIMIT 1
        """, (rota_id, empresa_id, motorista['id']))
        rota = cur.fetchone()

        if not rota:
            flash("Rota não encontrada ou não pertence ao seu cadastro.", "danger")
            return redirect(url_for('minhas_rotas_motorista'))

        status_atual = rota.get('status_motorista') or 'Aguardando conferência'

        if rota.get('situacao_rota') == 'Cancelada':
            flash("Esta rota está cancelada e não pode ser conferida.", "warning")
            return redirect(url_for('minhas_rotas_motorista'))

        if status_atual != 'Aguardando conferência':
            flash(f"Esta rota não está aguardando conferência. Status atual: {status_atual}.", "warning")
            return redirect(url_for('minhas_rotas_motorista'))

        cur.execute("""
            UPDATE rotas
            SET status_motorista = 'Conferida pelo motorista',
                data_aprovacao_motorista = NOW(),
                usuario_aprovacao_motorista_id = %s
            WHERE id = %s
              AND empresa_id = %s
              AND motorista_id = %s
        """, (usuario_id, rota_id, empresa_id, motorista['id']))

        con.commit()

        registrar_historico_rota_motorista(
            empresa_id=empresa_id,
            rota_id=rota_id,
            usuario_id=usuario_id,
            status_anterior=status_atual,
            status_novo='Conferida pelo motorista',
            motivo='Conferência do motorista',
            observacao='Motorista confirmou que os dados da rota estão corretos pelo Portal do Motorista.'
        )

        flash(f"Rota '{rota['identi_rota']}' conferida com sucesso. Aguarde a liberação para envio da NF.", "success")

    except Exception as e:
        con.rollback()
        print(f"Erro ao confirmar rota pelo motorista: {e}")
        flash("Erro técnico ao confirmar rota.", "danger")

    finally:
        cur.close()
        con.close()


    return redirect(url_for('minhas_rotas_motorista'))


@app.route('/portal-motorista/minhas-rotas/<int:rota_id>/divergencia', methods=['POST'])
@login_required
@motorista_required
def apontar_divergencia_rota_motorista(rota_id):
    empresa_id = session.get('empresa_id')
    motorista = buscar_motorista_logado()
    usuario_id = session.get('usuario_id')

    tipo_divergencia = (request.form.get('tipo_divergencia') or '').strip()
    descricao = (request.form.get('descricao') or '').strip()

    if not empresa_id or not motorista:
        flash("Não foi possível identificar o motorista logado.", "danger")
        return redirect(url_for('portal_motorista'))

    if not tipo_divergencia:
        flash("Selecione o tipo da divergência.", "warning")
        return redirect(url_for('minhas_rotas_motorista'))

    if not descricao or len(descricao) < 8:
        flash("Descreva a divergência com mais detalhes.", "warning")
        return redirect(url_for('minhas_rotas_motorista'))

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('minhas_rotas_motorista'))

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
            SELECT id,
                   empresa_id,
                   identi_rota,
                   motorista_id,
                   COALESCE(status_motorista, 'Aguardando conferência') AS status_motorista,
                   situacao_rota
            FROM rotas
            WHERE id = %s
              AND empresa_id = %s
              AND motorista_id = %s
            LIMIT 1
        """, (rota_id, empresa_id, motorista['id']))
        rota = cur.fetchone()

        if not rota:
            flash("Rota não encontrada ou não pertence ao seu cadastro.", "danger")
            return redirect(url_for('minhas_rotas_motorista'))

        status_atual = rota.get('status_motorista') or 'Aguardando conferência'

        if rota.get('situacao_rota') == 'Cancelada':
            flash("Esta rota está cancelada e não pode receber divergência.", "warning")
            return redirect(url_for('minhas_rotas_motorista'))

        if status_atual not in ['Aguardando conferência', 'Conferida pelo motorista']:
            flash(f"Esta rota não aceita divergência neste momento. Status atual: {status_atual}.", "warning")
            return redirect(url_for('minhas_rotas_motorista'))

        cur.execute("""
            SELECT id
            FROM rotas_divergencias_motorista
            WHERE rota_id = %s
              AND empresa_id = %s
              AND motorista_id = %s
              AND status_divergencia IN ('Aberta', 'Em análise')
            LIMIT 1
        """, (rota_id, empresa_id, motorista['id']))
        divergencia_aberta = cur.fetchone()

        if divergencia_aberta:
            flash("Esta rota já possui uma divergência aberta. Aguarde a análise da operação.", "warning")
            return redirect(url_for('minhas_rotas_motorista'))

        cur.execute("""
            INSERT INTO rotas_divergencias_motorista
                (empresa_id, rota_id, motorista_id, usuario_motorista_id, tipo_divergencia, descricao, status_divergencia)
            VALUES (%s, %s, %s, %s, %s, %s, 'Aberta')
        """, (empresa_id, rota_id, motorista['id'], usuario_id, tipo_divergencia, descricao))

        cur.execute("""
            UPDATE rotas
            SET status_motorista = 'Divergência apontada'
            WHERE id = %s
              AND empresa_id = %s
              AND motorista_id = %s
        """, (rota_id, empresa_id, motorista['id']))

        con.commit()

        registrar_historico_rota_motorista(
            empresa_id=empresa_id,
            rota_id=rota_id,
            usuario_id=usuario_id,
            status_anterior=status_atual,
            status_novo='Divergência apontada',
            motivo=tipo_divergencia,
            observacao=descricao
        )

        flash(f"Divergência registrada para a rota '{rota['identi_rota']}'. A operação irá analisar.", "success")

    except Exception as e:
        con.rollback()
        print(f"Erro ao apontar divergência da rota: {e}")
        flash("Erro técnico ao registrar divergência.", "danger")

    finally:
        cur.close()
        con.close()

    return redirect(url_for('minhas_rotas_motorista'))


def carregar_rotas_liberadas_motorista(empresa_id, motorista_id):
    con = obter_conexao()

    if con is None:
        return []

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT r.id,
                           r.empresa_id,
                           r.data_lancamento,
                           r.identi_rota,
                           r.tipo_rota,
                           r.valor_rota,
                           r.valor_km,
                           r.outras_despesas,
                           (COALESCE(r.valor_rota, 0) + COALESCE(r.valor_km, 0) +
                            COALESCE(r.outras_despesas, 0))                     AS valor_total_rota,
                           r.transportadora_id,
                           COALESCE(transp.nome_completo, emp.razao_social, emp.nome_fantasia) AS transportadora_nome,
                           COALESCE(transp.cpf_cnpj, emp.cnpj)                                  AS transportadora_cpf_cnpj,
                           r.motorista_id,
                           COALESCE(r.status_motorista, 'Aguardando conferência') AS status_motorista
                    FROM rotas r
                             LEFT JOIN pessoas transp
                                       ON transp.id = r.transportadora_id
                                           AND transp.empresa_id = r.empresa_id
                             LEFT JOIN empresas emp
                                       ON emp.id = r.empresa_id
                    WHERE r.empresa_id = %s
                      AND r.motorista_id = %s
                      AND COALESCE(r.status_motorista, 'Aguardando conferência') = 'Liberada para NF'
                    ORDER BY r.data_lancamento DESC, r.id DESC
                    """, (empresa_id, motorista_id))

        return cur.fetchall()

    except Exception as e:
        print(f"Erro ao carregar rotas liberadas do motorista: {e}")
        return []

    finally:
        cur.close()
        con.close()


@app.route('/portal-motorista/enviar-nf', methods=['GET', 'POST'])
@login_required
@motorista_required
def enviar_nf_motorista():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_id = session.get('empresa_id')
    motorista = buscar_motorista_logado()

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    if not motorista:
        flash("Seu usuário não está vinculado a um motorista ativo nesta empresa.", "danger")
        return redirect(url_for('portal_motorista'))

    motorista_id = motorista['id']
    rota_id_preselecionada = request.args.get('rota_id', '').strip()

    if request.method == 'POST':
        rota_ids_raw = request.form.getlist('rota_ids')
        observacao = request.form.get('observacao', '').strip()
        arquivo_xml = request.files.get('xml_nf')

        rota_ids = []

        for item in rota_ids_raw:
            if str(item).isdigit():
                rota_ids.append(int(item))

        rota_ids = list(dict.fromkeys(rota_ids))

        if not rota_ids:
            flash("Selecione pelo menos uma rota liberada para envio da NF.", "warning")
            return redirect(url_for('enviar_nf_motorista'))

        if not arquivo_xml or not arquivo_xml.filename:
            flash("Envie o arquivo XML da nota fiscal.", "danger")
            return redirect(url_for('enviar_nf_motorista'))

        if not arquivo_xml.filename.lower().endswith('.xml'):
            flash("Arquivo inválido. Envie apenas XML.", "danger")
            return redirect(url_for('enviar_nf_motorista'))

        con = obter_conexao()

        if con is None:
            flash("Erro de conexão com o banco de dados.", "danger")
            return redirect(url_for('enviar_nf_motorista'))

        cur = con.cursor(dictionary=True)

        try:
            placeholders = ", ".join(["%s"] * len(rota_ids))

            cur.execute(f"""
                SELECT
                    r.id,
                    r.empresa_id,
                    r.identi_rota,
                    r.motorista_id,
                    r.transportadora_id,
                    r.situacao_rota,
                    COALESCE(r.status_motorista, 'Aguardando conferência') AS status_motorista,
                    (COALESCE(r.valor_rota, 0) + COALESCE(r.valor_km, 0) + COALESCE(r.outras_despesas, 0)) AS valor_total_rota,
                    COALESCE(transp.nome_completo, emp.razao_social, emp.nome_fantasia) AS transportadora_nome,
                    COALESCE(transp.cpf_cnpj, emp.cnpj) AS transportadora_cpf_cnpj
                FROM rotas r
                LEFT JOIN pessoas transp
                    ON transp.id = r.transportadora_id
                   AND transp.empresa_id = r.empresa_id
                LEFT JOIN empresas emp
                    ON emp.id = r.empresa_id
                WHERE r.id IN ({placeholders})
                  AND r.empresa_id = %s
                  AND r.motorista_id = %s
                  AND COALESCE(r.status_motorista, 'Aguardando conferência') = 'Liberada para NF'
            """, rota_ids + [empresa_id, motorista_id])

            rotas = cur.fetchall()

            if len(rotas) != len(rota_ids):
                flash(
                    "Uma ou mais rotas selecionadas não estão disponíveis para envio de NF. "
                    "Atualize a página e tente novamente.",
                    "danger"
                )
                return redirect(url_for('enviar_nf_motorista'))

            # Garante que nenhuma rota já esteja vinculada a outra NF
            cur.execute(f"""
                SELECT v.rota_id
                FROM motorista_nf_rotas v
                INNER JOIN motorista_notas_fiscais nf
                    ON nf.id = v.motorista_nf_id
                   AND nf.empresa_id = v.empresa_id
                WHERE v.rota_id IN ({placeholders})
                  AND nf.status_nf <> 'Recusada'
            """, rota_ids)

            ja_vinculadas = cur.fetchall()

            if ja_vinculadas:
                flash("Uma ou mais rotas selecionadas já possuem documento ativo vinculado.", "danger")
                return redirect(url_for('enviar_nf_motorista'))

            documentos_tomador = set()

            for rota in rotas:
                doc_tomador = somente_digitos(rota.get('transportadora_cpf_cnpj'))
                if doc_tomador:
                    documentos_tomador.add(doc_tomador)

            if not documentos_tomador:
                flash("Não foi possível validar o tomador: a transportadora/tomador da rota não possui CPF/CNPJ.",
                      "danger")
                return redirect(url_for('enviar_nf_motorista'))

            if len(documentos_tomador) > 1:
                flash(
                    "As rotas selecionadas possuem tomadores diferentes. "
                    "Envie uma NF por tomador/transportadora.",
                    "danger"
                )
                return redirect(url_for('enviar_nf_motorista'))

            doc_motorista = somente_digitos(motorista.get('cpf_cnpj'))
            doc_tomador_rota = list(documentos_tomador)[0]

            if not doc_motorista:
                flash("Seu cadastro de motorista não possui CPF/CNPJ para validação da NF.", "danger")
                return redirect(url_for('enviar_nf_motorista'))

            arquivo_xml.stream.seek(0)
            dados_xml = extrair_dados_nfse(arquivo_xml.stream)
            arquivo_xml.stream.seek(0)

            prestador_xml = somente_digitos(dados_xml.get('emitente_cnpj'))
            tomador_xml = somente_digitos(dados_xml.get('tomador_cnpj'))
            valor_xml = converter_decimal(dados_xml.get('valor_total'))

            if not prestador_xml:
                flash("Não foi possível identificar o prestador no XML.", "danger")
                return redirect(url_for('enviar_nf_motorista'))

            if prestador_xml != doc_motorista:
                flash(
                    "NF recusada: o prestador do XML não corresponde ao CPF/CNPJ do motorista logado.",
                    "danger"
                )
                return redirect(url_for('enviar_nf_motorista'))

            if not tomador_xml:
                flash("Não foi possível identificar o tomador no XML.", "danger")
                return redirect(url_for('enviar_nf_motorista'))

            if tomador_xml != doc_tomador_rota:
                flash(
                    "NF recusada: o tomador do XML não corresponde à transportadora/tomador das rotas selecionadas.",
                    "danger"
                )
                return redirect(url_for('enviar_nf_motorista'))

            soma_rotas = Decimal('0.00')

            for rota in rotas:
                soma_rotas += converter_decimal(rota.get('valor_total_rota'))

            soma_rotas = soma_rotas.quantize(Decimal('0.01'))

            diferenca = abs(valor_xml - soma_rotas)

            if diferenca > Decimal('0.01'):
                flash(
                    f"NF recusada: o XML tem valor de {moeda_br(valor_xml)}, "
                    f"mas as rotas selecionadas somam {moeda_br(soma_rotas)}.",
                    "danger"
                )
                return redirect(url_for('enviar_nf_motorista'))

            cur.execute("""
                        SELECT id, status_nf
                        FROM motorista_notas_fiscais
                        WHERE empresa_id = %s
                          AND chave_acesso = %s
                        ORDER BY id DESC LIMIT 1
                        """, (empresa_id, dados_xml['chave_acesso']))

            nf_existente = cur.fetchone()
            nf_reenvio_recusada_id = None

            if nf_existente:
                status_xml_existente = nf_existente.get('status_nf')

                if status_xml_existente == 'Recusada':
                    # Documento recusado pode ser reativado para reenvio/correção.
                    nf_reenvio_recusada_id = nf_existente.get('id')

                elif status_xml_existente == 'Estornada':
                    # Documento estornado faz parte de processo encerrado/cancelado.
                    # Neste caso, o processo será feito do zero com nova rota e novo vínculo,
                    # então permitimos criar um novo registro com a mesma chave XML.
                    nf_reenvio_recusada_id = None

                else:
                    flash("Este XML já possui um documento ativo no sistema.", "danger")
                    return redirect(url_for('enviar_nf_motorista'))

            nome_seguro = nome_arquivo_seguro_xml_motorista(arquivo_xml.filename)
            timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
            nome_final = f"empresa_{empresa_id}_motorista_{motorista_id}_{timestamp}_{nome_seguro}"
            caminho_final = os.path.join(pasta_upload_xml_motorista(), nome_final)

            arquivo_xml.save(caminho_final)
            caminho_relativo_xml = f"uploads/motorista_xmls/{nome_final}"

            if nf_reenvio_recusada_id:
                cur.execute("""
                            UPDATE motorista_notas_fiscais
                            SET motorista_id             = %s,
                                tipo_documento_pagamento = 'XML',
                                numero_nf                = %s,
                                data_emissao             = %s,
                                valor_total              = %s,
                                valor_bruto              = %s,
                                valor_liquido            = %s,
                                prestador_cpf_cnpj       = %s,
                                tomador_cpf_cnpj         = %s,
                                status_nf                = 'Enviada',
                                nome_arquivo_xml         = %s,
                                data_envio               = NOW(),
                                data_recusa              = NULL,
                                motivo_recusa            = NULL,
                                usuario_recusa_id        = NULL,
                                observacao               = CONCAT(
                                        COALESCE(observacao, ''),
                                        CASE WHEN COALESCE(observacao, '') = '' THEN '' ELSE '\\n' END,
                                        'XML reenviado pelo motorista em ',
                                        DATE_FORMAT(NOW(), '%d/%m/%Y %H:%i'),
                                        CASE WHEN %s IS NULL OR %s = '' THEN '' ELSE CONCAT('. Observação: ', %s) END
                                                           )
                            WHERE id = %s
                              AND empresa_id = %s
                              AND status_nf = 'Recusada'
                            """, (
                                motorista_id,
                                dados_xml['numero_nf'],
                                dados_xml['data_emissao'],
                                valor_xml,
                                valor_xml,
                                valor_xml,
                                prestador_xml,
                                tomador_xml,
                                nome_final,
                                observacao or None,
                                observacao or None,
                                observacao or None,
                                nf_reenvio_recusada_id,
                                empresa_id
                            ))

                motorista_nf_id = nf_reenvio_recusada_id
            else:
                cur.execute("""
                            INSERT INTO motorista_notas_fiscais (empresa_id,
                                                                 motorista_id,
                                                                 tipo_documento_pagamento,
                                                                 numero_nf,
                                                                 chave_acesso,
                                                                 data_emissao,
                                                                 valor_total,
                                                                 valor_bruto,
                                                                 valor_liquido,
                                                                 prestador_cpf_cnpj,
                                                                 tomador_cpf_cnpj,
                                                                 status_nf,
                                                                 nome_arquivo_xml,
                                                                 observacao)
                            VALUES (%s, %s, 'XML', %s, %s, %s, %s, %s, %s, %s, %s, 'Enviada', %s, %s)
                            """, (
                                empresa_id,
                                motorista_id,
                                dados_xml['numero_nf'],
                                dados_xml['chave_acesso'],
                                dados_xml['data_emissao'],
                                valor_xml,
                                valor_xml,
                                valor_xml,
                                prestador_xml,
                                tomador_xml,
                                nome_final,
                                observacao or None
                            ))

                motorista_nf_id = cur.lastrowid

            # Fase Google Drive: envia o XML/NF do motorista para o Drive, mantendo fallback local.
            # O campo nome_arquivo_xml passa a guardar a rota interna protegida quando o upload der certo.
            nome_xml_para_sistema = tentar_enviar_arquivo_google_drive(
                cur,
                caminho_final,
                caminho_relativo_xml,
                empresa_id=empresa_id,
                motorista_id=motorista_id,
                origem='XML_MOTORISTA',
                origem_id=motorista_nf_id,
                tipo_arquivo='XML_NF_MOTORISTA',
                nome_original=arquivo_xml.filename,
                mime_type=getattr(arquivo_xml, 'mimetype', None) or 'application/xml',
                criado_por_usuario_id=session.get('usuario_id'),
            )

            if nome_xml_para_sistema and nome_xml_para_sistema != nome_final:
                cur.execute("""
                            UPDATE motorista_notas_fiscais
                            SET nome_arquivo_xml = %s
                            WHERE id = %s
                              AND empresa_id = %s
                            """, (nome_xml_para_sistema, motorista_nf_id, empresa_id))

            # Se estamos reativando um documento recusado com a mesma chave XML,
            # removemos os vínculos antigos deste mesmo documento antes de recriar.
            # Isso evita rotas duplicadas na tela de detalhes.
            cur.execute("""
                        DELETE
                        FROM motorista_nf_rotas
                        WHERE motorista_nf_id = %s
                          AND empresa_id = %s
                        """, (motorista_nf_id, empresa_id))

            # Remove vínculos antigos de OUTROS documentos recusados para permitir novo envio da mesma rota.
            # O histórico do documento recusado permanece em motorista_notas_fiscais/historico_operacoes.
            cur.execute(f"""
                DELETE v
                FROM motorista_nf_rotas v
                INNER JOIN motorista_notas_fiscais nf
                    ON nf.id = v.motorista_nf_id
                   AND nf.empresa_id = v.empresa_id
                WHERE v.rota_id IN ({placeholders})
                  AND v.empresa_id = %s
                  AND nf.status_nf = 'Recusada'
            """, rota_ids + [empresa_id])

            for rota in rotas:
                cur.execute("""
                            INSERT INTO motorista_nf_rotas (empresa_id,
                                                            motorista_nf_id,
                                                            rota_id,
                                                            valor_rota)
                            VALUES (%s, %s, %s, %s)
                            """, (
                                empresa_id,
                                motorista_nf_id,
                                rota['id'],
                                converter_decimal(rota.get('valor_total_rota'))
                            ))

                cur.execute("""
                            UPDATE rotas
                            SET status_motorista = 'NF enviada'
                            WHERE id = %s
                              AND empresa_id = %s
                              AND motorista_id = %s
                              AND COALESCE(status_motorista, 'Aguardando conferência') = 'Liberada para NF'
                            """, (rota['id'], empresa_id, motorista_id))

            con.commit()

            for rota in rotas:
                registrar_historico_rota_motorista(
                    empresa_id=empresa_id,
                    rota_id=rota['id'],
                    usuario_id=session.get('usuario_id'),
                    status_anterior='Liberada para NF',
                    status_novo='NF enviada',
                    motivo='Envio de XML pelo motorista',
                    observacao=f"NF {dados_xml['numero_nf']} enviada pelo Portal do Motorista."
                )

            flash(
                f"NF {dados_xml['numero_nf']} enviada com sucesso. "
                f"{len(rotas)} rota(s) vinculada(s) e aguardando análise.",
                "success"
            )
            return redirect(url_for('portal_motorista'))

        except Exception as e:
            con.rollback()
            print(f"Erro no envio de NF do motorista: {e}")
            flash(f"Erro técnico ao enviar XML: {e}", "danger")
            return redirect(url_for('enviar_nf_motorista'))

        finally:
            cur.close()
            con.close()

    rotas_liberadas = carregar_rotas_liberadas_motorista(empresa_id, motorista_id)

    return render_template(
        'enviar_nf_motorista.html',
        usuario_logado=usuario_logado,
        motorista=motorista,
        rotas=rotas_liberadas,
        rota_id_preselecionada=rota_id_preselecionada
    )




def caminho_local_xml_motorista(nome_arquivo_xml):
    """Resolve XML legado/local salvo antes da integração com o Google Drive."""
    if not nome_arquivo_xml:
        return None

    valor = str(nome_arquivo_xml).strip().replace('\\', '/')

    # Links internos do SGR e links externos não são caminhos locais.
    if valor.startswith('/arquivos/visualizar/') or valor.startswith('http://') or valor.startswith('https://'):
        return None

    if valor.startswith('uploads/motorista_xmls/'):
        caminho = os.path.join(app.root_path, valor.replace('/', os.sep))
    elif valor.startswith('motorista_xmls/'):
        caminho = os.path.join(app.root_path, 'uploads', valor.replace('/', os.sep))
    else:
        caminho = os.path.join(pasta_upload_xml_motorista(), os.path.basename(valor))

    return caminho if os.path.exists(caminho) else None


@app.route('/portal-motorista/nfs/<int:id>/xml', methods=['GET'])
@login_required
def visualizar_xml_nf_motorista(id):
    """Visualiza o XML/NF do motorista com proteção de sessão, empresa e perfil."""
    empresa_logada_id = session.get('empresa_id')
    perfil = session.get('perfil_de_acesso')
    pessoa_id = garantir_pessoa_id_sessao()
    is_super_admin = int(session.get('is_super_admin') or 0) == 1

    if perfil == 'Terminal Base':
        flash('Terminal Base não possui permissão para visualizar arquivos.', 'danger')
        return redirect(url_for('terminal_base_qrcode'))

    con = obter_conexao()
    if con is None:
        flash('Erro de conexão ao abrir XML.', 'danger')
        return redirect(url_for('portal_motorista'))

    cur = con.cursor(dictionary=True)
    try:
        cur.execute("""
                    SELECT nf.id,
                           nf.empresa_id,
                           nf.motorista_id,
                           nf.numero_nf,
                           nf.nome_arquivo_xml,
                           mot.nome_completo AS motorista_nome
                    FROM motorista_notas_fiscais nf
                    LEFT JOIN pessoas mot
                           ON mot.id = nf.motorista_id
                          AND mot.empresa_id = nf.empresa_id
                    WHERE nf.id = %s
                    LIMIT 1
                    """, (id,))
        nf = cur.fetchone()

        if not nf:
            flash('XML/NF não encontrado.', 'warning')
            return redirect(url_for('portal_motorista'))

        permitido = False
        if is_super_admin:
            permitido = True
        elif perfil in ['Administrador', 'Financeiro', 'Operacional', 'Consulta']:
            permitido = str(nf.get('empresa_id')) == str(empresa_logada_id)
        elif perfil == 'Motorista':
            permitido = (
                str(nf.get('empresa_id')) == str(empresa_logada_id)
                and pessoa_id
                and str(nf.get('motorista_id')) == str(pessoa_id)
            )

        if not permitido:
            flash('Você não tem permissão para acessar este XML/NF.', 'danger')
            if perfil == 'Motorista':
                return redirect(url_for('portal_motorista'))
            return redirect(url_for('dashboard'))

        caminho_xml = nf.get('nome_arquivo_xml')
        if not caminho_xml:
            flash('Este documento não possui XML vinculado.', 'warning')
            return redirect(url_for('detalhes_nf_motorista', id=id) if perfil != 'Motorista' else url_for('minhas_nfs_motorista'))

        caminho_xml = str(caminho_xml).strip()

        if caminho_xml.startswith('/arquivos/visualizar/'):
            return redirect(caminho_xml)

        if caminho_xml.startswith('http://') or caminho_xml.startswith('https://'):
            # Mantido apenas para registros antigos; novos arquivos devem usar /arquivos/visualizar/<id>.
            return redirect(caminho_xml)

        caminho_local = caminho_local_xml_motorista(caminho_xml)
        if caminho_local:
            nome_download = os.path.basename(caminho_local)
            return send_file(
                caminho_local,
                mimetype='application/xml',
                as_attachment=False,
                download_name=nome_download,
            )

        flash('Arquivo XML não localizado no armazenamento.', 'warning')
        return redirect(url_for('detalhes_nf_motorista', id=id) if perfil != 'Motorista' else url_for('minhas_nfs_motorista'))

    except Exception as e:
        print(f'Erro ao visualizar XML/NF do motorista {id}: {e}')
        flash('Erro técnico ao abrir XML/NF.', 'danger')
        return redirect(url_for('dashboard'))

    finally:
        fechar_cursor_conexao(cur, con)


@app.route('/portal-motorista/minhas-nfs', methods=['GET'])
@login_required
@motorista_required
def minhas_nfs_motorista():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_id = session.get('empresa_id')
    motorista = buscar_motorista_logado()

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    if not motorista:
        flash("Seu usuário não está vinculado a um motorista ativo nesta empresa.", "danger")
        return redirect(url_for('portal_motorista'))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('portal_motorista'))

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT nf.id,
                           nf.tipo_documento_pagamento,
                           nf.numero_nf,
                           nf.chave_acesso,
                           nf.data_emissao,
                           nf.valor_total,
                           nf.valor_bruto,
                           nf.valor_liquido,
                           nf.status_nf,
                           nf.data_envio,
                           nf.motivo_recusa,
                           nf.observacao,
                           COUNT(v.id)                                              AS qtd_rotas,
                           GROUP_CONCAT(r.identi_rota ORDER BY r.id SEPARATOR ', ') AS rotas_vinculadas
                    FROM motorista_notas_fiscais nf
                             LEFT JOIN motorista_nf_rotas v
                                       ON v.motorista_nf_id = nf.id
                                           AND v.empresa_id = nf.empresa_id
                             LEFT JOIN rotas r
                                       ON r.id = v.rota_id
                                           AND r.empresa_id = nf.empresa_id
                    WHERE nf.empresa_id = %s
                      AND nf.motorista_id = %s
                    GROUP BY nf.id,
                             nf.tipo_documento_pagamento,
                             nf.numero_nf,
                             nf.chave_acesso,
                             nf.data_emissao,
                             nf.valor_total,
                             nf.valor_bruto,
                             nf.valor_liquido,
                             nf.status_nf,
                             nf.data_envio,
                             nf.motivo_recusa,
                             nf.observacao
                    ORDER BY nf.data_envio DESC, nf.id DESC
                    """, (empresa_id, motorista['id']))

        notas = cur.fetchall()

    except Exception as e:
        print(f"Erro ao carregar NFs do motorista: {e}")
        flash("Erro técnico ao carregar suas notas fiscais.", "danger")
        notas = []

    finally:
        cur.close()
        con.close()

    return render_template(
        'minhas_nfs_motorista.html',
        usuario_logado=usuario_logado,
        motorista=motorista,
        notas=notas
    )

@app.route('/portal-motorista/nfs/<int:nf_id>/danfse')
@login_required
def visualizar_danfse_nf(nf_id):
    con = obter_conexao()
    if con is None:
        flash('Erro de conexão ao abrir a DANFSe.', 'danger')
        return redirect(url_for('portal_motorista'))

    cur = con.cursor(dictionary=True)

    try:
        nf, erro = _buscar_nf_com_permissao(cur, nf_id)

        if erro:
            flash(erro, 'danger')
            return redirect(url_for('portal_motorista'))

        xml_bytes = _ler_xml_nf(nf)
        danfse = parse_danfse_xml(xml_bytes)

        return render_template(
            'visualizar_danfse_nf.html',
            nf=nf,
            danfse=danfse,
            xml_download_url=url_for('baixar_xml_nf_original', nf_id=nf_id)
        )

    except Exception as e:
        print(f'Erro ao gerar DANFSe da NF {nf_id}: {e}')
        flash(f'Não foi possível gerar a visualização da DANFSe: {e}', 'danger')
        return redirect(url_for('portal_motorista'))

    finally:
        fechar_cursor_conexao(cur, con)


@app.route('/portal-motorista/nfs/<int:nf_id>/xml-original')
@login_required
def baixar_xml_nf_original(nf_id):
    con = obter_conexao()
    if con is None:
        flash('Erro de conexão ao baixar XML original.', 'danger')
        return redirect(url_for('portal_motorista'))

    cur = con.cursor(dictionary=True)

    try:
        nf, erro = _buscar_nf_com_permissao(cur, nf_id)

        if erro:
            flash(erro, 'danger')
            return redirect(url_for('portal_motorista'))

        xml_bytes = _ler_xml_nf(nf)

        filename = f"nfse_{nf_id}.xml"
        return Response(
            xml_bytes,
            mimetype='application/xml',
            headers={'Content-Disposition': f'attachment; filename={filename}'}
        )

    except Exception as e:
        print(f'Erro ao baixar XML original da NF {nf_id}: {e}')
        flash(f'XML original não localizado: {e}', 'danger')
        return redirect(url_for('portal_motorista'))

    finally:
        fechar_cursor_conexao(cur, con)

# ==========================================================
# HELPERS FINANCEIRO NF MOTORISTA
# Necessário antes das rotas da Fase 3.5
# ==========================================================
def usuario_pode_analisar_nf_motorista():
    """
    Controle simples de acesso para o painel financeiro das NFs/documentos dos motoristas.
    Super Admin sempre pode.
    Perfis internos permitidos também podem.
    Motorista não pode acessar este painel.
    """
    is_super_admin = int(session.get('is_super_admin') or 0) == 1
    perfil = session.get('perfil_de_acesso')

    if is_super_admin:
        return True

    perfis_permitidos = [
        'Administrador',
        'Financeiro',
        'Operacional'
    ]

    return perfil in perfis_permitidos


def financeiro_nf_motorista_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not usuario_pode_analisar_nf_motorista():
            flash("Acesso restrito ao painel financeiro de documentos dos motoristas.", "danger")
            return redirect(url_for('dashboard'))

        return f(*args, **kwargs)

    return decorated_function


# ==========================================================
# FASE 3.5 - APROVAR / RECUSAR DOCUMENTO DO MOTORISTA
# ==========================================================
@app.route('/financeiro/nfs-motoristas/<int:id>/aprovar', methods=['POST'])
@login_required
@financeiro_nf_motorista_required
def aprovar_documento_motorista(id):
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = int(session.get('is_super_admin') or 0) == 1

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('financeiro_nfs_motoristas'))

    cur = con.cursor(dictionary=True)

    try:
        query = """
                SELECT id,
                       empresa_id,
                       motorista_id,
                       tipo_documento_pagamento,
                       numero_nf,
                       status_nf,
                       valor_total
                FROM motorista_notas_fiscais
                WHERE id = %s \
                """

        params = [id]

        if not is_super_admin:
            query += " AND empresa_id = %s"
            params.append(empresa_logada_id)

        query += " LIMIT 1"

        cur.execute(query, params)
        documento = cur.fetchone()

        if not documento:
            flash("Documento não encontrado ou não pertence à empresa logada.", "danger")
            return redirect(url_for('financeiro_nfs_motoristas'))

        status_atual = documento.get('status_nf')

        if status_atual not in ['Enviada', 'Em análise']:
            flash(
                f"Este documento não pode ser aprovado agora. Status atual: {status_atual}.",
                "warning"
            )
            return redirect(url_for('detalhes_nf_motorista', id=id))

        cur.execute("""
                    SELECT v.rota_id,
                           r.status_motorista,
                           r.identi_rota
                    FROM motorista_nf_rotas v
                             INNER JOIN rotas r
                                        ON r.id = v.rota_id
                                            AND r.empresa_id = v.empresa_id
                    WHERE v.motorista_nf_id = %s
                      AND v.empresa_id = %s
                    """, (id, documento['empresa_id']))

        rotas_vinculadas = cur.fetchall()

        if not rotas_vinculadas:
            flash("Este documento não possui rotas vinculadas.", "danger")
            return redirect(url_for('detalhes_nf_motorista', id=id))

        cur.execute("""
                    UPDATE motorista_notas_fiscais
                    SET status_nf            = 'Aprovada',
                        data_aprovacao       = NOW(),
                        usuario_aprovacao_id = %s
                    WHERE id = %s
                      AND empresa_id = %s
                    """, (usuario_id, id, documento['empresa_id']))

        for rota in rotas_vinculadas:
            cur.execute("""
                        UPDATE rotas
                        SET status_motorista = 'Aprovada para pagamento'
                        WHERE id = %s
                          AND empresa_id = %s
                          AND COALESCE(status_motorista, 'Aguardando liberação') IN ('NF enviada', 'Em análise')
                        """, (rota['rota_id'], documento['empresa_id']))

        con.commit()

        registrar_historico_nf_motorista(
            empresa_id=documento['empresa_id'],
            motorista_nf_id=id,
            usuario_id=usuario_id,
            status_anterior=status_atual,
            status_novo='Aprovada',
            motivo='Documento aprovado pelo financeiro',
            observacao=f"Documento {documento['numero_nf']} aprovado para pagamento."
        )

        for rota in rotas_vinculadas:
            registrar_historico_rota_motorista(
                empresa_id=documento['empresa_id'],
                rota_id=rota['rota_id'],
                usuario_id=usuario_id,
                status_anterior=rota.get('status_motorista'),
                status_novo='Aprovada para pagamento',
                motivo='Documento do motorista aprovado',
                observacao=f"Documento {documento['numero_nf']} aprovado para pagamento."
            )

        flash(f"Documento {documento['numero_nf']} aprovado com sucesso.", "success")
        return redirect(url_for('detalhes_nf_motorista', id=id))

    except Exception as e:
        con.rollback()
        print(f"Erro ao aprovar documento do motorista: {e}")
        flash(f"Erro técnico ao aprovar documento: {e}", "danger")
        return redirect(url_for('detalhes_nf_motorista', id=id))

    finally:
        cur.close()
        con.close()


@app.route('/financeiro/nfs-motoristas/<int:id>/recusar', methods=['POST'])
@login_required
@financeiro_nf_motorista_required
def recusar_documento_motorista(id):
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = int(session.get('is_super_admin') or 0) == 1

    motivo_recusa = request.form.get('motivo_recusa', '').strip()
    destino_rotas_recusa = request.form.get('destino_rotas_recusa', '').strip()

    destinos_validos = [
        'LIBERAR_REENVIO',
        'BLOQUEAR_CORRECAO'
    ]

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    if len(motivo_recusa) < 3:
        flash("Informe um motivo válido para recusar o documento.", "warning")
        return redirect(url_for('detalhes_nf_motorista', id=id))

    if destino_rotas_recusa not in destinos_validos:
        flash("Selecione o destino das rotas após a recusa.", "warning")
        return redirect(url_for('detalhes_nf_motorista', id=id))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('financeiro_nfs_motoristas'))

    cur = con.cursor(dictionary=True)

    try:
        query = """
                SELECT id,
                       empresa_id,
                       motorista_id,
                       tipo_documento_pagamento,
                       numero_nf,
                       status_nf,
                       valor_total
                FROM motorista_notas_fiscais
                WHERE id = %s \
                """

        params = [id]

        if not is_super_admin:
            query += " AND empresa_id = %s"
            params.append(empresa_logada_id)

        query += " LIMIT 1"

        cur.execute(query, params)
        documento = cur.fetchone()

        if not documento:
            flash("Documento não encontrado ou não pertence à empresa logada.", "danger")
            return redirect(url_for('financeiro_nfs_motoristas'))

        status_atual = documento.get('status_nf')

        if status_atual not in ['Enviada', 'Em análise']:
            flash(
                f"Este documento não pode ser recusado agora. Status atual: {status_atual}.",
                "warning"
            )
            return redirect(url_for('detalhes_nf_motorista', id=id))

        cur.execute("""
                    SELECT v.rota_id,
                           r.status_motorista,
                           r.identi_rota
                    FROM motorista_nf_rotas v
                             INNER JOIN rotas r
                                        ON r.id = v.rota_id
                                            AND r.empresa_id = v.empresa_id
                    WHERE v.motorista_nf_id = %s
                      AND v.empresa_id = %s
                    """, (id, documento['empresa_id']))

        rotas_vinculadas = cur.fetchall()

        if not rotas_vinculadas:
            flash("Este documento não possui rotas vinculadas.", "danger")
            return redirect(url_for('detalhes_nf_motorista', id=id))

        if destino_rotas_recusa == 'LIBERAR_REENVIO':
            novo_status_rota = 'Liberada para NF'
            descricao_destino = 'rotas liberadas para reenvio pelo motorista'
        else:
            novo_status_rota = 'Bloqueada'
            descricao_destino = 'rotas bloqueadas para correção operacional'

        observacao_recusa = (
            f"Documento recusado em {datetime.now().strftime('%d/%m/%Y %H:%M')}. "
            f"Destino das rotas: {descricao_destino}. Motivo: {motivo_recusa}"
        )

        cur.execute("""
                    UPDATE motorista_notas_fiscais
                    SET status_nf         = 'Recusada',
                        data_recusa       = NOW(),
                        usuario_recusa_id = %s,
                        motivo_recusa     = %s,
                        observacao        = CONCAT(
                                COALESCE(observacao, ''),
                                CASE WHEN COALESCE(observacao, '') = '' THEN '' ELSE '\n' END,
                                %s
                                            )
                    WHERE id = %s
                      AND empresa_id = %s
                    """, (
                        usuario_id,
                        motivo_recusa,
                        observacao_recusa,
                        id,
                        documento['empresa_id']
                    ))

        # Destino da rota após recusa:
        # LIBERAR_REENVIO   -> rota volta para Liberada para NF, permitindo novo envio pelo motorista.
        # BLOQUEAR_CORRECAO -> rota vai para Bloqueada, impedindo reenvio até correção operacional.
        for rota in rotas_vinculadas:
            cur.execute("""
                        UPDATE rotas
                        SET status_motorista = %s
                        WHERE id = %s
                          AND empresa_id = %s
                          AND COALESCE(status_motorista, 'Aguardando liberação') IN ('NF enviada', 'Em análise')
                        """, (novo_status_rota, rota['rota_id'], documento['empresa_id']))

        con.commit()

        registrar_historico_nf_motorista(
            empresa_id=documento['empresa_id'],
            motorista_nf_id=id,
            usuario_id=usuario_id,
            status_anterior=status_atual,
            status_novo='Recusada',
            motivo=motivo_recusa,
            observacao=f"Documento {documento['numero_nf']} recusado. Destino das rotas: {descricao_destino}."
        )

        for rota in rotas_vinculadas:
            registrar_historico_rota_motorista(
                empresa_id=documento['empresa_id'],
                rota_id=rota['rota_id'],
                usuario_id=usuario_id,
                status_anterior=rota.get('status_motorista'),
                status_novo=novo_status_rota,
                motivo='Documento recusado pelo financeiro',
                observacao=(
                    f"Documento {documento['numero_nf']} recusado. "
                    f"Destino: {descricao_destino}. Motivo: {motivo_recusa}"
                )
            )

        if destino_rotas_recusa == 'LIBERAR_REENVIO':
            flash(
                f"Documento {documento['numero_nf']} recusado. "
                "As rotas foram liberadas novamente para o motorista reenviar.",
                "success"
            )
        else:
            flash(
                f"Documento {documento['numero_nf']} recusado. "
                "As rotas foram bloqueadas para correção operacional.",
                "success"
            )

        return redirect(url_for('detalhes_nf_motorista', id=id))

    except Exception as e:
        con.rollback()
        print(f"Erro ao recusar documento do motorista: {e}")
        flash(f"Erro técnico ao recusar documento: {e}", "danger")
        return redirect(url_for('detalhes_nf_motorista', id=id))

    finally:
        cur.close()
        con.close()


@app.route('/pessoas/cadastro', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional')
def cadastro_pessoa():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = int(session.get('is_super_admin') or 0) == 1

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    if request.method == 'POST':
        if is_super_admin:
            empresa_id_destino = request.form.get('empresa_id')
        else:
            empresa_id_destino = empresa_logada_id

        if not empresa_id_destino or not str(empresa_id_destino).isdigit():
            flash("Selecione uma empresa válida para este cadastro.", "danger")
            return redirect(url_for('cadastro_pessoa'))

        empresa_id_destino = int(empresa_id_destino)

        nome_completo = request.form.get('nome_completo', '').strip()
        apelido = request.form.get('apelido', '').strip()
        cpf_cnpj = request.form.get('cpf_cnpj', '').strip()
        email = request.form.get('email', '').strip()
        telefone = request.form.get('telefone', '').strip()
        tipo_cadastro = request.form.get('tipo_cadastro', '').strip()
        tipo_prestador = request.form.get('tipo_prestador', '').strip()
        permite_acesso_portal = request.form.get('permite_acesso_portal', 'N').strip()
        tipo_cadastro, tipo_prestador = normalizar_categoria_pessoa(tipo_cadastro, tipo_prestador)
        permite_acesso_portal = 'S' if permite_acesso_portal == 'S' else 'N'
        status_cadastro = request.form.get('status_cadastro', 'Ativo').strip()

        cep = request.form.get('cep', '').strip()
        rua = request.form.get('rua', '').strip()
        numero = request.form.get('numero', '').strip()
        bairro = request.form.get('bairro', '').strip()
        cidade = request.form.get('cidade', '').strip()
        uf = request.form.get('uf', '').strip()
        observacao = request.form.get('observacao', '').strip()

        if not nome_completo:
            flash("Informe o nome completo.", "danger")
            return redirect(url_for('cadastro_pessoa'))

        if not cpf_cnpj:
            flash("Informe o CPF/CNPJ.", "danger")
            return redirect(url_for('cadastro_pessoa'))

        if not tipo_cadastro:
            flash("Selecione o tipo de cadastro.", "danger")
            return redirect(url_for('cadastro_pessoa'))

        if status_cadastro not in ['Ativo', 'Inativo']:
            flash("Status de cadastro inválido.", "danger")
            return redirect(url_for('cadastro_pessoa'))

        con = obter_conexao()

        if con is None:
            flash("Erro de conexão com o banco de dados.", "danger")
            return redirect(url_for('cadastro_pessoa'))

        cur = con.cursor(dictionary=True)

        try:
            cur.execute("""
                        SELECT id, status_empresa
                        FROM empresas
                        WHERE id = %s LIMIT 1
                        """, (empresa_id_destino,))

            empresa = cur.fetchone()

            if not empresa:
                flash("Empresa selecionada não encontrada.", "danger")
                return redirect(url_for('cadastro_pessoa'))

            if empresa['status_empresa'] != 'Ativa':
                flash("Não é possível cadastrar pessoa em empresa inativa.", "danger")
                return redirect(url_for('cadastro_pessoa'))

            cpf_cnpj_limpo = ''.join(filter(str.isdigit, cpf_cnpj))

            cur.execute("""
                        SELECT id
                        FROM pessoas
                        WHERE empresa_id = %s
                          AND REPLACE(REPLACE(REPLACE(REPLACE(cpf_cnpj, '.', ''), '-', ''), '/', ''), ' ', '') =
                              %s
                            LIMIT 1
                        """, (empresa_id_destino, cpf_cnpj_limpo))

            pessoa_existente = cur.fetchone()

            if pessoa_existente:
                flash("Já existe uma pessoa com este CPF/CNPJ cadastrada nesta empresa.", "danger")
                return redirect(url_for('cadastro_pessoa'))

            cur.execute("""
                        INSERT INTO pessoas (empresa_id,
                                             data_cadastro,
                                             nome_completo,
                                             apelido,
                                             cpf_cnpj,
                                             email,
                                             telefone,
                                             tipo_cadastro,
                                             tipo_prestador,
                                             permite_acesso_portal,
                                             status_cadastro,
                                             cep,
                                             rua,
                                             numero,
                                             bairro,
                                             cidade,
                                             uf,
                                             observacao)
                        VALUES (%s,
                                CURDATE(),
                                %s,
                                %s,
                                %s,
                                %s,
                                %s,
                                %s,
                                %s,
                                %s,
                                %s,
                                %s,
                                %s,
                                %s,
                                %s,
                                %s,
                                %s,
                                %s)
                        """, (
                            empresa_id_destino,
                            nome_completo,
                            apelido or None,
                            cpf_cnpj,
                            email or None,
                            telefone or None,
                            tipo_cadastro,
                            tipo_prestador or None,
                            permite_acesso_portal,
                            status_cadastro,
                            cep or None,
                            rua or None,
                            numero or None,
                            bairro or None,
                            cidade or None,
                            uf or None,
                            observacao or None
                        ))

            con.commit()

            flash("Cadastro realizado com sucesso!", "success")
            return redirect(url_for('visualizar_pessoas'))

        except Exception as e:
            con.rollback()
            print(f"Erro ao cadastrar pessoa: {e}")
            flash("Erro técnico ao cadastrar pessoa.", "danger")
            return redirect(url_for('cadastro_pessoa'))

        finally:
            cur.close()
            con.close()

    empresas = []

    if is_super_admin:
        con = obter_conexao()

        if con:
            cur = con.cursor(dictionary=True)

            try:
                cur.execute("""
                            SELECT id,
                                   razao_social,
                                   nome_fantasia
                            FROM empresas
                            WHERE status_empresa = 'Ativa'
                            ORDER BY nome_fantasia ASC, razao_social ASC
                            """)
                empresas = cur.fetchall()

            except Exception as e:
                print(f"Erro ao carregar empresas no cadastro de pessoa: {e}")
                flash("Erro ao carregar empresas.", "warning")

            finally:
                fechar_cursor_conexao(cur, con)

    return render_template(
        'cadastro_pessoa.html',
        usuario_logado=usuario_logado,
        empresas=empresas,
        is_super_admin=is_super_admin,
        empresa_logada_id=empresa_logada_id,
        categorias_cadastro=CATEGORIAS_CADASTRO_PESSOA,
        tipos_prestador=TIPOS_PRESTADOR_SERVICO
    )


@app.route('/pessoas/visualizar')
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro', 'Consulta')
def visualizar_pessoas():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = int(session.get('is_super_admin') or 0) == 1

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    pesquisa = request.args.get('pesquisa', '').strip()
    tipo_cadastro = request.args.get('tipo_cadastro', '').strip()
    status_cadastro = request.args.get('status_cadastro', '').strip()
    empresa_id_filtro = request.args.get('empresa_id', '').strip()

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão ao carregar cadastros.", "danger")
        return render_template(
            'visualizar_pessoas.html',
            pessoas=[],
            empresas=[],
            filtros={},
            is_super_admin=is_super_admin,
            usuario_logado=usuario_logado
        )

    cur = con.cursor(dictionary=True)

    try:
        query = """
                SELECT p.id,
                       p.empresa_id,
                       e.nome_fantasia AS empresa_nome,
                       e.razao_social  AS empresa_razao_social,
                       p.nome_completo,
                       p.apelido,
                       p.cpf_cnpj,
                       p.tipo_cadastro,
                       p.tipo_prestador,
                       p.permite_acesso_portal,
                       p.telefone,
                       p.cidade,
                       p.uf,
                       p.status_cadastro
                FROM pessoas p
                         INNER JOIN empresas e ON e.id = p.empresa_id
                WHERE 1 = 1 \
                """

        params = []

        if is_super_admin:
            if empresa_id_filtro and empresa_id_filtro.isdigit():
                query += " AND p.empresa_id = %s"
                params.append(int(empresa_id_filtro))
        else:
            query += " AND p.empresa_id = %s"
            params.append(empresa_logada_id)

        if pesquisa:
            query += """
                AND (
                    p.nome_completo LIKE %s
                    OR p.apelido LIKE %s
                    OR p.cpf_cnpj LIKE %s
                    OR p.telefone LIKE %s
                    OR p.email LIKE %s
                )
            """
            like = f"%{pesquisa}%"
            params.extend([like, like, like, like, like])

        if tipo_cadastro:
            query += " AND p.tipo_cadastro = %s"
            params.append(tipo_cadastro)

        if status_cadastro in ['Ativo', 'Inativo']:
            query += " AND p.status_cadastro = %s"
            params.append(status_cadastro)

        query += " ORDER BY e.nome_fantasia ASC, e.razao_social ASC, p.nome_completo ASC"

        cur.execute(query, params)
        pessoas = cur.fetchall()

        empresas = []

        if is_super_admin:
            cur.execute("""
                        SELECT id,
                               razao_social,
                               nome_fantasia
                        FROM empresas
                        ORDER BY nome_fantasia ASC, razao_social ASC
                        """)
            empresas = cur.fetchall()

    except Exception as e:
        print(f"Erro ao buscar pessoas: {e}")
        flash("Erro técnico ao listar cadastros.", "danger")
        pessoas = []
        empresas = []

    finally:
        fechar_cursor_conexao(cur, con)

    filtros = {
        'pesquisa': pesquisa,
        'tipo_cadastro': tipo_cadastro,
        'status_cadastro': status_cadastro,
        'empresa_id': empresa_id_filtro
    }

    return render_template(
        'visualizar_pessoas.html',
        pessoas=pessoas,
        empresas=empresas,
        filtros=filtros,
        is_super_admin=is_super_admin,
        usuario_logado=usuario_logado,
        categorias_cadastro=CATEGORIAS_CADASTRO_PESSOA,
        tipos_prestador=TIPOS_PRESTADOR_SERVICO
    )


@app.route('/pessoas/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional')
def editar_pessoa(id):
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = int(session.get('is_super_admin') or 0) == 1

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('visualizar_pessoas'))

    cur = con.cursor(dictionary=True)

    def buscar_pessoa():
        query = """
                SELECT p.id,
                       p.empresa_id,
                       e.nome_fantasia AS empresa_nome,
                       e.razao_social  AS empresa_razao_social,
                       p.nome_completo,
                       p.apelido,
                       p.cpf_cnpj,
                       p.email,
                       p.telefone,
                       p.tipo_cadastro,
                       p.tipo_prestador,
                       p.permite_acesso_portal,
                       p.status_cadastro,
                       p.cep,
                       p.rua,
                       p.numero,
                       p.bairro,
                       p.cidade,
                       p.uf,
                       p.observacao
                FROM pessoas p
                         INNER JOIN empresas e ON e.id = p.empresa_id
                WHERE p.id = %s \
                """

        params = [id]

        if not is_super_admin:
            query += " AND p.empresa_id = %s"
            params.append(empresa_logada_id)

        query += " LIMIT 1"

        cur.execute(query, params)
        return cur.fetchone()

    if request.method == 'POST':
        if is_super_admin:
            empresa_id_destino = request.form.get('empresa_id')
        else:
            empresa_id_destino = empresa_logada_id

        if not empresa_id_destino or not str(empresa_id_destino).isdigit():
            flash("Selecione uma empresa válida.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('editar_pessoa', id=id))

        empresa_id_destino = int(empresa_id_destino)

        nome_completo = request.form.get('nome_completo', '').strip()
        apelido = request.form.get('apelido', '').strip()
        cpf_cnpj = request.form.get('cpf_cnpj', '').strip()
        email = request.form.get('email', '').strip()
        telefone = request.form.get('telefone', '').strip()
        tipo_cadastro = request.form.get('tipo_cadastro', '').strip()
        tipo_prestador = request.form.get('tipo_prestador', '').strip()
        permite_acesso_portal = request.form.get('permite_acesso_portal', 'N').strip()
        tipo_cadastro, tipo_prestador = normalizar_categoria_pessoa(tipo_cadastro, tipo_prestador)
        permite_acesso_portal = 'S' if permite_acesso_portal == 'S' else 'N'
        status_cadastro = request.form.get('status_cadastro', 'Ativo').strip()

        cep = request.form.get('cep', '').strip()
        rua = request.form.get('rua', '').strip()
        numero = request.form.get('numero', '').strip()
        bairro = request.form.get('bairro', '').strip()
        cidade = request.form.get('cidade', '').strip()
        uf = request.form.get('uf', '').strip()
        observacao = request.form.get('observacao', '').strip()

        if not nome_completo:
            flash("Informe o nome completo.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('editar_pessoa', id=id))

        if not cpf_cnpj:
            flash("Informe o CPF/CNPJ.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('editar_pessoa', id=id))

        if not tipo_cadastro:
            flash("Selecione o tipo de cadastro.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('editar_pessoa', id=id))

        if status_cadastro not in ['Ativo', 'Inativo']:
            flash("Status de cadastro inválido.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('editar_pessoa', id=id))

        try:
            pessoa_atual = buscar_pessoa()

            if not pessoa_atual:
                flash("Cadastro não encontrado ou não pertence à empresa logada.", "warning")
                return redirect(url_for('visualizar_pessoas'))

            if not is_super_admin and int(pessoa_atual['empresa_id']) != int(empresa_logada_id):
                flash("Você não pode editar cadastros de outra empresa.", "danger")
                return redirect(url_for('visualizar_pessoas'))

            cur.execute("""
                        SELECT id, status_empresa
                        FROM empresas
                        WHERE id = %s LIMIT 1
                        """, (empresa_id_destino,))

            empresa = cur.fetchone()

            if not empresa:
                flash("Empresa selecionada não encontrada.", "danger")
                return redirect(url_for('editar_pessoa', id=id))

            if empresa['status_empresa'] != 'Ativa':
                flash("Não é possível vincular pessoa a uma empresa inativa.", "danger")
                return redirect(url_for('editar_pessoa', id=id))

            empresa_original = int(pessoa_atual['empresa_id'])

            # Proteção multiempresa:
            # A pessoa pode existir em várias empresas, mas como cadastros separados.
            # Se este cadastro já possui movimentações na empresa original, a empresa não pode ser alterada.
            if empresa_original != empresa_id_destino:
                movimentos = pessoa_tem_movimentacao_empresa(cur, id, empresa_original)
                total_movimentos = int(movimentos.get('total') or 0)

                if total_movimentos > 0:
                    detalhes = '; '.join(movimentos.get('detalhes') or [])
                    flash(
                        "Não é permitido trocar a empresa deste cadastro, pois ele já possui movimentações "
                        "na empresa atual. Para manter o histórico correto, crie um novo cadastro desta pessoa "
                        "na empresa de destino." + (f" Movimentações encontradas: {detalhes}." if detalhes else ""),
                        "danger"
                    )
                    return redirect(url_for('editar_pessoa', id=id))

            cpf_cnpj_limpo = ''.join(filter(str.isdigit, cpf_cnpj))

            cur.execute("""
                        SELECT id
                        FROM pessoas
                        WHERE empresa_id = %s
                          AND id <> %s
                          AND REPLACE(REPLACE(REPLACE(REPLACE(cpf_cnpj, '.', ''), '-', ''), '/', ''), ' ', '') =
                              %s
                            LIMIT 1
                        """, (empresa_id_destino, id, cpf_cnpj_limpo))

            pessoa_existente = cur.fetchone()

            if pessoa_existente:
                flash("Já existe outra pessoa com este CPF/CNPJ cadastrada nesta empresa.", "danger")
                return redirect(url_for('editar_pessoa', id=id))

            query_update = """
                           UPDATE pessoas
                           SET empresa_id      = %s,
                               nome_completo   = %s,
                               apelido         = %s,
                               cpf_cnpj        = %s,
                               email           = %s,
                               telefone        = %s,
                               tipo_cadastro   = %s,
                               tipo_prestador  = %s,
                               permite_acesso_portal = %s,
                               status_cadastro = %s,
                               cep             = %s,
                               rua             = %s,
                               numero          = %s,
                               bairro          = %s,
                               cidade          = %s,
                               uf              = %s,
                               observacao      = %s
                           WHERE id = %s \
                           """

            params_update = [
                empresa_id_destino,
                nome_completo,
                apelido or None,
                cpf_cnpj,
                email or None,
                telefone or None,
                tipo_cadastro,
                tipo_prestador or None,
                permite_acesso_portal,
                status_cadastro,
                cep or None,
                rua or None,
                numero or None,
                bairro or None,
                cidade or None,
                uf or None,
                observacao or None,
                id
            ]

            if not is_super_admin:
                query_update += " AND empresa_id = %s"
                params_update.append(empresa_logada_id)

            cur.execute(query_update, params_update)

            if cur.rowcount == 0:
                con.rollback()
                flash("Nenhuma alteração realizada ou cadastro não pertence à empresa logada.", "warning")
                return redirect(url_for('visualizar_pessoas'))

            # Se a troca de empresa foi permitida por não haver movimentação, sincroniza usuário vinculado.
            if empresa_original != empresa_id_destino:
                cur.execute("UPDATE usuarios SET empresa_id = %s WHERE pessoa_id = %s", (empresa_id_destino, id))

            con.commit()

            flash("Cadastro atualizado com sucesso!", "success")
            return redirect(url_for('visualizar_pessoas'))

        except Exception as e:
            con.rollback()
            print(f"Erro ao editar pessoa: {e}")
            flash("Erro técnico ao editar cadastro.", "danger")
            return redirect(url_for('editar_pessoa', id=id))

        finally:
            fechar_cursor_conexao(cur, con)

    try:
        pessoa = buscar_pessoa()

        if not pessoa:
            flash("Cadastro não encontrado ou não pertence à empresa logada.", "warning")
            return redirect(url_for('visualizar_pessoas'))

        empresas = []

        if is_super_admin:
            cur.execute("""
                        SELECT id,
                               razao_social,
                               nome_fantasia
                        FROM empresas
                        WHERE status_empresa = 'Ativa'
                        ORDER BY nome_fantasia ASC, razao_social ASC
                        """)
            empresas = cur.fetchall()

    except Exception as e:
        print(f"Erro ao carregar pessoa: {e}")
        flash("Erro técnico ao carregar cadastro.", "danger")
        return redirect(url_for('visualizar_pessoas'))

    finally:
        fechar_cursor_conexao(cur, con)

    return render_template(
        'editar_pessoa.html',
        pessoa=pessoa,
        empresas=empresas,
        is_super_admin=is_super_admin,
        usuario_logado=usuario_logado,
        categorias_cadastro=CATEGORIAS_CADASTRO_PESSOA,
        tipos_prestador=TIPOS_PRESTADOR_SERVICO
    )


@app.route('/pessoas/excluir/<int:id>', methods=['POST'])
@login_required
@perfis_permitidos('Administrador')
def excluir_pessoa(id):
    empresa_id = session.get('empresa_id')

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('visualizar_pessoas'))

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT id, nome_completo
                    FROM pessoas
                    WHERE id = %s
                      AND empresa_id = %s LIMIT 1
                    """, (id, empresa_id))

        pessoa = cur.fetchone()

        if not pessoa:
            flash("Cadastro não encontrado ou não pertence à empresa logada.", "danger")
            return redirect(url_for('visualizar_pessoas'))

        cur.execute("""
                    DELETE
                    FROM pessoas
                    WHERE id = %s
                      AND empresa_id = %s
                    """, (id, empresa_id))

        con.commit()

        flash(f"Cadastro de {pessoa['nome_completo']} excluído com sucesso.", "success")

    except Exception as e:
        con.rollback()
        print(f"Erro ao excluir pessoa: {e}")
        flash(
            "Não foi possível excluir este cadastro. Verifique se ele está vinculado a rotas, usuários ou lançamentos.",
            "danger")

    finally:
        cur.close()
        con.close()

    return redirect(url_for('visualizar_pessoas'))


@app.route('/api/pessoas/buscar')
@login_required
def api_buscar_pessoas():
    """Busca inteligente de pessoas por ID, nome, apelido ou CPF/CNPJ.

    Parâmetros:
    - q: texto digitado
    - uso: financeiro | motorista | ajudante | prestador | todos
    """
    empresa_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()
    q = (request.args.get('q') or '').strip()
    uso = (request.args.get('uso') or 'todos').strip().lower()
    empresa_id_filtro = (request.args.get('empresa_id') or '').strip()

    if not empresa_id:
        return jsonify({'results': []})

    if len(q) < 1:
        return jsonify({'results': []})

    con = obter_conexao()
    if con is None:
        return jsonify({'results': []})

    cur = con.cursor(dictionary=True)
    try:
        query = """
            SELECT
                p.id,
                p.empresa_id,
                p.nome_completo,
                p.apelido,
                p.cpf_cnpj,
                p.tipo_cadastro,
                p.tipo_prestador,
                p.permite_acesso_portal,
                e.nome_fantasia AS empresa_nome,
                e.razao_social AS empresa_razao_social
            FROM pessoas p
            INNER JOIN empresas e ON e.id = p.empresa_id
            WHERE p.status_cadastro = 'Ativo'
        """
        params = []

        if is_super_admin and empresa_id_filtro.isdigit():
            query += " AND p.empresa_id = %s"
            params.append(int(empresa_id_filtro))
        elif not is_super_admin:
            query += " AND p.empresa_id = %s"
            params.append(empresa_id)

        if uso == 'motorista':
            query += f" AND {condicao_sql_motorista_prestador('p')}"
        elif uso == 'ajudante':
            query += f" AND {condicao_sql_ajudante_prestador('p')}"
        elif uso == 'prestador':
            query += " AND p.tipo_cadastro = 'Prestador de Serviço'"
        elif uso == 'financeiro':
            query += """
                AND p.tipo_cadastro IN (
                    'Prestador de Serviço', 'Fornecedor', 'Funcionário', 'Funcionario',
                    'Cliente / Tomador', 'Transportadora', 'Órgão Público', 'Outros',
                    'Motorista', 'Ajudante', 'Prestador'
                )
            """

        q_limpo = ''.join(filter(str.isdigit, q))
        query += """
            AND (
                p.nome_completo LIKE %s
                OR p.apelido LIKE %s
                OR p.cpf_cnpj LIKE %s
                OR p.id = %s
        """
        like = f"%{q}%"
        params.extend([like, like, like, int(q) if q.isdigit() else -1])
        if q_limpo:
            query += " OR REPLACE(REPLACE(REPLACE(REPLACE(p.cpf_cnpj, '.', ''), '-', ''), '/', ''), ' ', '') LIKE %s"
            params.append(f"%{q_limpo}%")
        query += ") ORDER BY p.nome_completo ASC LIMIT 20"

        cur.execute(query, params)
        rows = cur.fetchall()
        results = []
        for row in rows:
            results.append({
                'id': row.get('id'),
                'text': pessoa_label_busca(row),
                'nome': row.get('nome_completo'),
                'cpf_cnpj': row.get('cpf_cnpj'),
                'tipo_cadastro': row.get('tipo_cadastro'),
                'tipo_prestador': row.get('tipo_prestador'),
                'empresa': row.get('empresa_nome') or row.get('empresa_razao_social')
            })
        return jsonify({'results': results})
    except Exception as e:
        print(f"Erro na busca inteligente de pessoas: {e}")
        return jsonify({'results': []})
    finally:
        fechar_cursor_conexao(cur, con)


# ==========================================================
# ROTAS OPERACIONAIS
# ==========================================================
@app.route('/movimentacao/lancar', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional')
def lancar_rota():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_id = session.get('empresa_id')

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('dashboard'))

    cur = con.cursor(dictionary=True)

    if request.method == 'POST':
        data_lancamento = request.form.get('data_lancamento') or request.form.get('data_rota')
        identi_rota = request.form.get('identi_rota', '').strip()
        tipo_rota = request.form.get('tipo_rota', '').strip()

        # Bloco Rotas 1: a transportadora/tomador deixou de ser obrigatório no lançamento.
        # A empresa logada passa a ser a responsável pelo lançamento e o motorista é o vínculo principal.
        # Transportadora/Tomador deixou de ser obrigatório no fluxo novo.
        # A rota pertence à empresa logada; apenas o motorista responsável é selecionado.
        transportadora_id = None
        motorista_id = request.form.get('motorista_id')

        valor_rota = request.form.get('valor_rota', '0')
        valor_km = request.form.get('valor_km', '0')
        outras_despesas = request.form.get('outras_despesas', '0')

        situacao_rota = 'Pendente'
        # A rota nasce visível para conferência do motorista antes da liberação fiscal.
        status_motorista = 'Aguardando conferência'

        if not data_lancamento:
            flash("Informe a data da rota.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('lancar_rota'))

        if not identi_rota:
            flash("Informe a identificação da rota.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('lancar_rota'))

        if not tipo_rota:
            flash("Selecione o tipo da rota.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('lancar_rota'))

        if transportadora_id and not str(transportadora_id).isdigit():
            flash("Transportadora inválida.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('lancar_rota'))

        if not motorista_id or not str(motorista_id).isdigit():
            flash("Selecione um motorista válido.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('lancar_rota'))

        try:
            transportadora_id = int(transportadora_id) if transportadora_id else None
            motorista_id = int(motorista_id)

            valor_rota = converter_decimal(valor_rota)
            valor_km = converter_decimal(valor_km)
            outras_despesas = converter_decimal(outras_despesas)

            if transportadora_id:
                cur.execute("""
                            SELECT id, nome_completo
                            FROM pessoas
                            WHERE id = %s
                              AND empresa_id = %s
                              AND tipo_cadastro = 'Transportadora'
                              AND status_cadastro = 'Ativo' LIMIT 1
                            """, (transportadora_id, empresa_id))

                transportadora = cur.fetchone()

                if not transportadora:
                    flash("Transportadora inválida ou não pertence à empresa logada.", "danger")
                    return redirect(url_for('lancar_rota'))

            cur.execute("""
                        SELECT id, nome_completo
                        FROM pessoas
                        WHERE id = %s
                          AND empresa_id = %s
                          AND status_cadastro = 'Ativo'
                          AND (tipo_cadastro = 'Motorista' OR (tipo_cadastro = 'Prestador de Serviço' AND COALESCE(tipo_prestador, '') IN ('Motorista', 'Motorista e Ajudante'))) LIMIT 1
                        """, (motorista_id, empresa_id))

            motorista = cur.fetchone()

            if not motorista:
                flash("Motorista inválido ou não pertence à empresa logada.", "danger")
                return redirect(url_for('lancar_rota'))

            cur.execute("""
                        INSERT INTO rotas (empresa_id,
                                           data_lancamento,
                                           identi_rota,
                                           valor_rota,
                                           valor_km,
                                           outras_despesas,
                                           tipo_rota,
                                           transportadora_id,
                                           motorista_id,
                                           situacao_rota,
                                           status_motorista)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            empresa_id,
                            data_lancamento,
                            identi_rota,
                            valor_rota,
                            valor_km,
                            outras_despesas,
                            tipo_rota,
                            transportadora_id,
                            motorista_id,
                            situacao_rota,
                            status_motorista
                        ))

            con.commit()

            flash(
                f"Rota '{identi_rota}' cadastrada com sucesso para o motorista {motorista['nome_completo']}.",
                "success"
            )
            return redirect(url_for('visualizar_rotas'))

        except Exception as e:
            con.rollback()
            print(f"Erro ao lançar rota: {e}")
            flash("Erro técnico ao lançar rota.", "danger")
            return redirect(url_for('lancar_rota'))

        finally:
            fechar_cursor_conexao(cur, con)

    transportadoras = []

    try:

        cur.execute("""
                    SELECT id, nome_completo
                    FROM pessoas
                    WHERE empresa_id = %s
                      AND status_cadastro = 'Ativo'
                      AND (tipo_cadastro = 'Motorista' OR (tipo_cadastro = 'Prestador de Serviço' AND COALESCE(tipo_prestador, '') IN ('Motorista', 'Motorista e Ajudante')))
                    ORDER BY nome_completo ASC
                    """, (empresa_id,))
        motoristas = cur.fetchall()

    except Exception as e:
        print(f"Erro ao carregar listas da rota: {e}")
        flash("Erro ao carregar transportadoras e motoristas.", "danger")
        transportadoras = []
        motoristas = []

    finally:
        fechar_cursor_conexao(cur, con)

    return render_template(
        'lancar_rota.html',
        usuario_logado=usuario_logado,
        transportadoras=transportadoras,
        motoristas=motoristas
    )


def carregar_pessoas_por_tipo(tipo_cadastro):
    con = obter_conexao()
    if con is None:
        return []

    cur = con.cursor(dictionary=True)
    try:
        cur.execute("""
                    SELECT id, nome_completo
                    FROM pessoas
                    WHERE tipo_cadastro = %s
                      AND status_cadastro = 'Ativo'
                    ORDER BY nome_completo ASC
                    """, (tipo_cadastro,))
        return cur.fetchall()
    except Exception as e:
        print(f'Erro ao carregar pessoas do tipo {tipo_cadastro}: {e}')
        return []
    finally:
        fechar_cursor_conexao(cur, con)


def usuario_eh_super_admin_rotas():
    return int(session.get('is_super_admin') or 0) == 1


def registrar_historico_rota_motorista(
        empresa_id,
        rota_id,
        usuario_id,
        status_anterior,
        status_novo,
        motivo=None,
        observacao=None
):
    """
    Registro defensivo de histórico.
    Se a tabela historico_operacoes não existir ou tiver estrutura diferente,
    o fluxo principal não quebra.
    """
    con_hist = obter_conexao()

    if con_hist is None:
        return

    cur_hist = con_hist.cursor()

    try:
        cur_hist.execute("""
                         INSERT INTO historico_operacoes (empresa_id,
                                                          tipo_operacao,
                                                          rota_id,
                                                          usuario_id,
                                                          status_anterior,
                                                          status_novo,
                                                          motivo,
                                                          observacao)
                         VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                         """, (
                             empresa_id,
                             'STATUS_MOTORISTA_ROTA',
                             rota_id,
                             usuario_id,
                             status_anterior,
                             status_novo,
                             motivo,
                             observacao
                         ))

        con_hist.commit()

    except Exception as e:
        con_hist.rollback()
        print(f"Aviso: não foi possível registrar histórico da rota {rota_id}: {e}")

    finally:
        cur_hist.close()
        con_hist.close()


@app.route('/movimentacao/visualizar', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro', 'Consulta')
def visualizar_rotas():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_rotas()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    identi_rota = request.args.get('identi_rota', '').strip()
    situacao_rota = request.args.get('situacao_rota', '').strip()
    status_motorista = request.args.get('status_motorista', '').strip()
    motorista_id = request.args.get('motorista_id', '').strip()
    transportadora_id = request.args.get('transportadora_id', '').strip()
    data_inicio = request.args.get('data_inicio', '').strip()
    data_fim = request.args.get('data_fim', '').strip()
    empresa_id_filtro = request.args.get('empresa_id', '').strip()

    status_motorista_validos = [
        'Aguardando conferência',
        'Conferida pelo motorista',
        'Divergência apontada',
        'Aguardando liberação',
        'Liberada para NF',
        'NF enviada',
        'Em análise',
        'Aprovada para pagamento',
        'Pagamento confirmado',
        'Bloqueada',
        'Cancelada'
    ]

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('dashboard'))

    cur = con.cursor(dictionary=True)

    try:
        query = """
                SELECT r.id,
                       r.empresa_id,
                       e.nome_fantasia                                      AS empresa_nome,
                       e.razao_social                                       AS empresa_razao_social,
                       r.data_lancamento,
                       r.identi_rota,
                       r.tipo_rota,
                       r.valor_rota,
                       r.valor_km,
                       r.outras_despesas,
                       (COALESCE(r.valor_rota, 0) + COALESCE(r.valor_km, 0) +
                        COALESCE(r.outras_despesas, 0))                     AS valor_total_rota,
                       r.transportadora_id,
                       transp.nome_completo                                 AS transportadora_nome,
                       transp.cpf_cnpj                                      AS transportadora_cpf_cnpj,
                       r.motorista_id,
                       mot.nome_completo                                    AS motorista_nome,
                       mot.cpf_cnpj                                         AS motorista_cpf_cnpj,
                       r.situacao_rota,
                       COALESCE(r.status_motorista, 'Aguardando conferência') AS status_motorista
                FROM rotas r
                         INNER JOIN empresas e ON e.id = r.empresa_id
                         LEFT JOIN pessoas transp
                                   ON transp.id = r.transportadora_id
                                       AND transp.empresa_id = r.empresa_id
                         LEFT JOIN pessoas mot
                                   ON mot.id = r.motorista_id
                                       AND mot.empresa_id = r.empresa_id
                WHERE 1 = 1 \
                """

        params = []

        if is_super_admin:
            if empresa_id_filtro and empresa_id_filtro.isdigit():
                query += " AND r.empresa_id = %s"
                params.append(int(empresa_id_filtro))
        else:
            query += " AND r.empresa_id = %s"
            params.append(empresa_logada_id)

        if identi_rota:
            query += " AND r.identi_rota LIKE %s"
            params.append(f"%{identi_rota}%")

        if situacao_rota:
            query += " AND r.situacao_rota = %s"
            params.append(situacao_rota)

        if status_motorista in status_motorista_validos:
            query += " AND COALESCE(r.status_motorista, 'Aguardando conferência') = %s"
            params.append(status_motorista)

        if motorista_id and motorista_id.isdigit():
            query += " AND r.motorista_id = %s"
            params.append(int(motorista_id))

        if transportadora_id and transportadora_id.isdigit():
            query += " AND r.transportadora_id = %s"
            params.append(int(transportadora_id))

        if data_inicio:
            query += " AND r.data_lancamento >= %s"
            params.append(data_inicio)

        if data_fim:
            query += " AND r.data_lancamento <= %s"
            params.append(data_fim)

        query += " ORDER BY r.data_lancamento DESC, r.id DESC"

        cur.execute(query, params)
        rotas = cur.fetchall()

        # Listas de filtros
        lista_params = []

        filtro_empresa_lista = ""

        if is_super_admin:
            if empresa_id_filtro and empresa_id_filtro.isdigit():
                filtro_empresa_lista = " AND empresa_id = %s"
                lista_params.append(int(empresa_id_filtro))
        else:
            filtro_empresa_lista = " AND empresa_id = %s"
            lista_params.append(empresa_logada_id)

        cur.execute(f"""
            SELECT id, nome_completo
            FROM pessoas
            WHERE tipo_cadastro = 'Motorista'
              AND status_cadastro = 'Ativo'
              {filtro_empresa_lista}
            ORDER BY nome_completo ASC
        """, lista_params)

        motoristas = cur.fetchall()

        cur.execute(f"""
            SELECT id, nome_completo
            FROM pessoas
            WHERE tipo_cadastro = 'Transportadora'
              AND status_cadastro = 'Ativo'
              {filtro_empresa_lista}
            ORDER BY nome_completo ASC
        """, lista_params)

        transportadoras = cur.fetchall()

        empresas = []

        if is_super_admin:
            cur.execute("""
                        SELECT id, razao_social, nome_fantasia
                        FROM empresas
                        ORDER BY nome_fantasia ASC, razao_social ASC
                        """)
            empresas = cur.fetchall()

        resumo = {
            'total_rotas': len(rotas),
            'aguardando': 0,
            'liberadas': 0,
            'com_nf': 0,
            'aprovadas': 0,
            'quitadas': 0,
            'canceladas': 0
        }

        for rota in rotas:
            sm = rota.get('status_motorista') or 'Aguardando conferência'
            situacao = rota.get('situacao_rota') or ''

            if sm == 'Aguardando liberação':
                resumo['aguardando'] += 1
            elif sm == 'Liberada para NF':
                resumo['liberadas'] += 1
            elif sm in ['NF enviada', 'Em análise']:
                resumo['com_nf'] += 1
            elif sm == 'Aprovada para pagamento':
                resumo['aprovadas'] += 1
            elif sm == 'Pagamento confirmado' or situacao == 'Quitado':
                resumo['quitadas'] += 1
            elif sm == 'Cancelada' or situacao == 'Cancelada':
                resumo['canceladas'] += 1

    except Exception as e:
        print(f"Erro ao buscar rotas com motorista: {e}")
        flash(f"Erro técnico ao listar rotas: {e}", "danger")
        rotas = []
        motoristas = []
        transportadoras = []
        empresas = []
        resumo = {
            'total_rotas': 0,
            'aguardando': 0,
            'liberadas': 0,
            'com_nf': 0,
            'aprovadas': 0,
            'quitadas': 0,
            'canceladas': 0
        }

    finally:
        cur.close()
        con.close()

    filtros = {
        'identi_rota': identi_rota,
        'situacao_rota': situacao_rota,
        'status_motorista': status_motorista,
        'motorista_id': motorista_id,
        'transportadora_id': transportadora_id,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'empresa_id': empresa_id_filtro
    }

    return render_template(
        'visualizar_rotas.html',
        usuario_logado=usuario_logado,
        rotas=rotas,
        motoristas=motoristas,
        transportadoras=transportadoras,
        empresas=empresas,
        filtros=filtros,
        resumo=resumo,
        is_super_admin=is_super_admin
    )



@app.route('/movimentacao/rotas/divergencias', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional')
def divergencias_rotas_motoristas():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_rotas()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    status_filtro = (request.args.get('status_divergencia') or '').strip()
    rota_filtro = (request.args.get('identi_rota') or '').strip()
    data_inicio = (request.args.get('data_inicio') or '').strip()
    data_fim = (request.args.get('data_fim') or '').strip()

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('visualizar_rotas'))

    cur = con.cursor(dictionary=True)

    try:
        query = """
            SELECT
                d.id,
                d.empresa_id,
                d.rota_id,
                d.motorista_id,
                d.tipo_divergencia,
                d.descricao,
                d.status_divergencia,
                d.resultado_operacao,
                d.observacao_operacao,
                d.created_at,
                d.resolvido_em,
                r.identi_rota,
                r.data_lancamento,
                r.tipo_rota,
                r.valor_rota,
                r.valor_km,
                r.outras_despesas,
                (COALESCE(r.valor_rota, 0) + COALESCE(r.valor_km, 0) + COALESCE(r.outras_despesas, 0)) AS valor_total_rota,
                COALESCE(r.status_motorista, 'Aguardando conferência') AS status_motorista,
                r.situacao_rota,
                mot.nome_completo AS motorista_nome,
                emp.nome_fantasia AS empresa_nome,
                emp.razao_social AS empresa_razao_social
            FROM rotas_divergencias_motorista d
            INNER JOIN rotas r ON r.id = d.rota_id AND r.empresa_id = d.empresa_id
            INNER JOIN pessoas mot ON mot.id = d.motorista_id AND mot.empresa_id = d.empresa_id
            LEFT JOIN empresas emp ON emp.id = d.empresa_id
            WHERE 1 = 1
        """
        params = []

        if not is_super_admin:
            query += " AND d.empresa_id = %s"
            params.append(empresa_logada_id)

        if status_filtro:
            query += " AND d.status_divergencia = %s"
            params.append(status_filtro)

        if rota_filtro:
            query += " AND r.identi_rota LIKE %s"
            params.append(f"%{rota_filtro}%")

        if data_inicio:
            query += " AND DATE(d.created_at) >= %s"
            params.append(data_inicio)

        if data_fim:
            query += " AND DATE(d.created_at) <= %s"
            params.append(data_fim)

        query += " ORDER BY FIELD(d.status_divergencia, 'Aberta', 'Em análise', 'Resolvida', 'Recusada'), d.created_at DESC"

        cur.execute(query, params)
        divergencias = cur.fetchall()

        resumo = {
            'abertas': 0,
            'em_analise': 0,
            'resolvidas': 0,
            'recusadas': 0,
            'total': len(divergencias)
        }

        for item in divergencias:
            st = item.get('status_divergencia')
            if st == 'Aberta':
                resumo['abertas'] += 1
            elif st == 'Em análise':
                resumo['em_analise'] += 1
            elif st == 'Resolvida':
                resumo['resolvidas'] += 1
            elif st == 'Recusada':
                resumo['recusadas'] += 1

    except Exception as e:
        print(f"Erro ao carregar divergências de rotas: {e}")
        flash("Erro técnico ao carregar divergências.", "danger")
        divergencias = []
        resumo = {'abertas': 0, 'em_analise': 0, 'resolvidas': 0, 'recusadas': 0, 'total': 0}

    finally:
        cur.close()
        con.close()

    filtros = {
        'status_divergencia': status_filtro,
        'identi_rota': rota_filtro,
        'data_inicio': data_inicio,
        'data_fim': data_fim
    }

    return render_template(
        'rotas_divergencias.html',
        usuario_logado=usuario_logado,
        divergencias=divergencias,
        resumo=resumo,
        filtros=filtros,
        is_super_admin=is_super_admin
    )


@app.route('/movimentacao/rotas/divergencias/<int:divergencia_id>/tratar', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional')
def tratar_divergencia_rota_motorista(divergencia_id):
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = usuario_eh_super_admin_rotas()

    acao = (request.form.get('acao') or '').strip()
    observacao = (request.form.get('observacao_operacao') or '').strip()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    if acao not in ['marcar_analise', 'retornar_conferencia', 'liberar_nf', 'bloquear']:
        flash("Ação inválida para tratar divergência.", "danger")
        return redirect(url_for('divergencias_rotas_motoristas'))

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('divergencias_rotas_motoristas'))

    cur = con.cursor(dictionary=True)

    try:
        query = """
            SELECT d.*,
                   r.identi_rota,
                   COALESCE(r.status_motorista, 'Divergência apontada') AS status_motorista
            FROM rotas_divergencias_motorista d
            INNER JOIN rotas r ON r.id = d.rota_id AND r.empresa_id = d.empresa_id
            WHERE d.id = %s
        """
        params = [divergencia_id]

        if not is_super_admin:
            query += " AND d.empresa_id = %s"
            params.append(empresa_logada_id)

        query += " LIMIT 1"

        cur.execute(query, params)
        divergencia = cur.fetchone()

        if not divergencia:
            flash("Divergência não encontrada ou não pertence à empresa logada.", "danger")
            return redirect(url_for('divergencias_rotas_motoristas'))

        status_atual_rota = divergencia.get('status_motorista') or 'Divergência apontada'

        if acao == 'marcar_analise':
            novo_status_divergencia = 'Em análise'
            novo_status_rota = 'Divergência apontada'
            resultado = 'Em análise pela operação'
            msg = 'Divergência marcada como em análise.'

        elif acao == 'retornar_conferencia':
            novo_status_divergencia = 'Resolvida'
            novo_status_rota = 'Aguardando conferência'
            resultado = 'Rota corrigida e devolvida para conferência do motorista'
            msg = 'Rota devolvida para nova conferência do motorista.'

        elif acao == 'liberar_nf':
            novo_status_divergencia = 'Resolvida'
            novo_status_rota = 'Liberada para NF'
            resultado = 'Liberada para NF pela operação'
            msg = 'Rota liberada para envio de NF.'

        else:
            novo_status_divergencia = 'Recusada'
            novo_status_rota = 'Bloqueada'
            resultado = 'Divergência recusada / rota bloqueada'
            msg = 'Divergência recusada e rota bloqueada.'

        cur.execute("""
            UPDATE rotas_divergencias_motorista
            SET status_divergencia = %s,
                resultado_operacao = %s,
                observacao_operacao = %s,
                usuario_operacao_id = %s,
                resolvido_em = CASE WHEN %s IN ('Resolvida', 'Recusada') THEN NOW() ELSE resolvido_em END,
                updated_at = NOW()
            WHERE id = %s
        """, (novo_status_divergencia, resultado, observacao, usuario_id, novo_status_divergencia, divergencia_id))

        cur.execute("""
            UPDATE rotas
            SET status_motorista = %s
            WHERE id = %s
              AND empresa_id = %s
        """, (novo_status_rota, divergencia['rota_id'], divergencia['empresa_id']))

        con.commit()

        registrar_historico_rota_motorista(
            empresa_id=divergencia['empresa_id'],
            rota_id=divergencia['rota_id'],
            usuario_id=usuario_id,
            status_anterior=status_atual_rota,
            status_novo=novo_status_rota,
            motivo=resultado,
            observacao=observacao or f"Tratativa da divergência #{divergencia_id}."
        )

        flash(msg, "success")

    except Exception as e:
        con.rollback()
        print(f"Erro ao tratar divergência da rota: {e}")
        flash("Erro técnico ao tratar divergência.", "danger")

    finally:
        cur.close()
        con.close()

    return redirect(url_for('divergencias_rotas_motoristas'))


@app.route('/movimentacao/rotas/<int:id>/liberar-nf', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional')
def liberar_rota_para_nf(id):
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = usuario_eh_super_admin_rotas()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('visualizar_rotas'))

    cur = con.cursor(dictionary=True)

    try:
        query = """
                SELECT id,
                       empresa_id,
                       identi_rota,
                       motorista_id,
                       COALESCE(status_motorista, 'Aguardando liberação') AS status_motorista
                FROM rotas
                WHERE id = %s \
                """

        params = [id]

        if not is_super_admin:
            query += " AND empresa_id = %s"
            params.append(empresa_logada_id)

        query += " LIMIT 1"

        cur.execute(query, params)
        rota = cur.fetchone()

        if not rota:
            flash("Rota não encontrada ou não pertence à empresa logada.", "danger")
            return redirect(url_for('visualizar_rotas'))

        if not rota.get('motorista_id'):
            flash("Não é possível liberar para NF: esta rota não possui motorista vinculado.", "danger")
            return redirect(url_for('visualizar_rotas'))

        status_atual = rota.get('status_motorista') or 'Aguardando conferência'

        if status_atual not in ['Aguardando liberação', 'Aguardando conferência', 'Conferida pelo motorista', 'Bloqueada']:
            flash(f"Esta rota não pode ser liberada agora. Status atual: {status_atual}.", "warning")
            return redirect(url_for('visualizar_rotas'))

        cur.execute("""
                    UPDATE rotas
                    SET status_motorista = 'Liberada para NF'
                    WHERE id = %s
                    """, (id,))

        con.commit()

        registrar_historico_rota_motorista(
            empresa_id=rota['empresa_id'],
            rota_id=id,
            usuario_id=usuario_id,
            status_anterior=status_atual,
            status_novo='Liberada para NF',
            motivo='Liberação manual',
            observacao='Rota liberada para envio de NF pelo motorista.'
        )

        flash(f"Rota '{rota['identi_rota']}' liberada para envio de NF.", "success")

    except Exception as e:
        con.rollback()
        print(f"Erro ao liberar rota para NF: {e}")
        flash("Erro técnico ao liberar rota para NF.", "danger")

    finally:
        cur.close()
        con.close()

    return redirect(url_for('visualizar_rotas'))


@app.route('/movimentacao/rotas/liberar-em-massa-nf', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional')
def liberar_rotas_em_massa_para_nf():
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = usuario_eh_super_admin_rotas()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    rota_ids_raw = request.form.getlist('rota_ids')

    rota_ids = []

    for item in rota_ids_raw:
        if str(item).isdigit():
            rota_ids.append(int(item))

    rota_ids = list(dict.fromkeys(rota_ids))

    if not rota_ids:
        flash("Selecione pelo menos uma rota para liberar.", "warning")
        return redirect(url_for('visualizar_rotas'))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('visualizar_rotas'))

    cur = con.cursor(dictionary=True)

    liberadas = 0
    ignoradas_sem_motorista = 0
    ignoradas_canceladas = 0
    ignoradas_status = 0
    ignoradas_empresa = 0

    try:
        placeholders = ", ".join(["%s"] * len(rota_ids))

        query = f"""
            SELECT
                id,
                empresa_id,
                identi_rota,
                motorista_id,
                situacao_rota,
                COALESCE(status_motorista, 'Aguardando liberação') AS status_motorista
            FROM rotas
            WHERE id IN ({placeholders})
        """

        params = rota_ids[:]

        if not is_super_admin:
            query += " AND empresa_id = %s"
            params.append(empresa_logada_id)

        cur.execute(query, params)
        rotas = cur.fetchall()

        ids_encontrados = {int(r['id']) for r in rotas}
        ignoradas_empresa = len([rid for rid in rota_ids if rid not in ids_encontrados])

        ids_liberar = []

        for rota in rotas:
            status_atual = rota.get('status_motorista') or 'Aguardando conferência'
            situacao_atual = rota.get('situacao_rota') or ''

            if not rota.get('motorista_id'):
                ignoradas_sem_motorista += 1
                continue

            if situacao_atual == 'Cancelada':
                ignoradas_canceladas += 1
                continue

            if status_atual not in ['Aguardando liberação', 'Aguardando conferência', 'Conferida pelo motorista', 'Bloqueada']:
                ignoradas_status += 1
                continue

            ids_liberar.append(int(rota['id']))

        if ids_liberar:
            placeholders_update = ", ".join(["%s"] * len(ids_liberar))

            cur.execute(f"""
                UPDATE rotas
                SET status_motorista = 'Liberada para NF'
                WHERE id IN ({placeholders_update})
            """, ids_liberar)

            liberadas = cur.rowcount

            con.commit()

            # Histórico por rota liberada
            rotas_liberadas = [r for r in rotas if int(r['id']) in ids_liberar]

            for rota in rotas_liberadas:
                registrar_historico_rota_motorista(
                    empresa_id=rota['empresa_id'],
                    rota_id=rota['id'],
                    usuario_id=usuario_id,
                    status_anterior=rota.get('status_motorista') or 'Aguardando conferência',
                    status_novo='Liberada para NF',
                    motivo='Liberação em massa',
                    observacao='Rota liberada em massa para envio de NF pelo motorista.'
                )
        else:
            con.rollback()

        partes_msg = []

        if liberadas > 0:
            partes_msg.append(f"{liberadas} rota(s) liberada(s) para NF")

        total_ignoradas = (
                ignoradas_sem_motorista +
                ignoradas_canceladas +
                ignoradas_status +
                ignoradas_empresa
        )

        if total_ignoradas > 0:
            detalhes = []

            if ignoradas_sem_motorista:
                detalhes.append(f"{ignoradas_sem_motorista} sem motorista")

            if ignoradas_canceladas:
                detalhes.append(f"{ignoradas_canceladas} cancelada(s)")

            if ignoradas_status:
                detalhes.append(f"{ignoradas_status} com status não permitido")

            if ignoradas_empresa:
                detalhes.append(f"{ignoradas_empresa} não encontrada(s) ou de outra empresa")

            partes_msg.append("ignoradas: " + ", ".join(detalhes))

        if partes_msg:
            categoria = "success" if liberadas > 0 else "warning"
            flash(" | ".join(partes_msg), categoria)
        else:
            flash("Nenhuma rota foi liberada.", "warning")

    except Exception as e:
        con.rollback()
        print(f"Erro ao liberar rotas em massa para NF: {e}")
        flash("Erro técnico ao liberar rotas em massa para NF.", "danger")

    finally:
        cur.close()
        con.close()

    return redirect(url_for('visualizar_rotas'))


@app.route('/movimentacao/rotas/<int:id>/bloquear-motorista', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional')
def bloquear_rota_motorista(id):
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = usuario_eh_super_admin_rotas()
    motivo = request.form.get('motivo', '').strip()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    if len(motivo) < 3:
        flash("Informe um motivo válido para bloquear a rota.", "danger")
        return redirect(url_for('visualizar_rotas'))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('visualizar_rotas'))

    cur = con.cursor(dictionary=True)

    try:
        query = """
                SELECT id,
                       empresa_id,
                       identi_rota,
                       COALESCE(status_motorista, 'Aguardando liberação') AS status_motorista
                FROM rotas
                WHERE id = %s \
                """

        params = [id]

        if not is_super_admin:
            query += " AND empresa_id = %s"
            params.append(empresa_logada_id)

        query += " LIMIT 1"

        cur.execute(query, params)
        rota = cur.fetchone()

        if not rota:
            flash("Rota não encontrada ou não pertence à empresa logada.", "danger")
            return redirect(url_for('visualizar_rotas'))

        status_atual = rota.get('status_motorista') or 'Aguardando conferência'

        if status_atual not in ['Aguardando liberação', 'Liberada para NF']:
            flash(f"Esta rota não pode ser bloqueada agora. Status atual: {status_atual}.", "warning")
            return redirect(url_for('visualizar_rotas'))

        cur.execute("""
                    UPDATE rotas
                    SET status_motorista = 'Bloqueada'
                    WHERE id = %s
                    """, (id,))

        con.commit()

        registrar_historico_rota_motorista(
            empresa_id=rota['empresa_id'],
            rota_id=id,
            usuario_id=usuario_id,
            status_anterior=status_atual,
            status_novo='Bloqueada',
            motivo=motivo,
            observacao='Rota bloqueada para o motorista.'
        )

        flash(f"Rota '{rota['identi_rota']}' bloqueada para o motorista.", "success")

    except Exception as e:
        con.rollback()
        print(f"Erro ao bloquear rota para motorista: {e}")
        flash("Erro técnico ao bloquear rota para motorista.", "danger")

    finally:
        cur.close()
        con.close()

    return redirect(url_for('visualizar_rotas'))


@app.route('/movimentacao/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional')
def editar_rota(id):
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_id = session.get('empresa_id')

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('visualizar_rotas'))

    cur = con.cursor(dictionary=True)

    if request.method == 'POST':
        data_lancamento = request.form.get('data_lancamento') or request.form.get('data_rota')
        identi_rota = request.form.get('identi_rota', '').strip()
        tipo_rota = request.form.get('tipo_rota', '').strip()

        # Transportadora/Tomador deixou de ser obrigatório no fluxo novo.
        # A rota pertence à empresa logada; apenas o motorista responsável é selecionado.
        motorista_id = request.form.get('motorista_id')

        situacao_rota = request.form.get('situacao_rota', 'Pendente').strip()
        status_motorista = request.form.get('status_motorista', 'Aguardando liberação').strip()

        valor_rota = request.form.get('valor_rota', '0')
        valor_km = request.form.get('valor_km', '0')
        outras_despesas = request.form.get('outras_despesas', '0')

        status_motorista_validos = [
            'Aguardando conferência',
            'Conferida pelo motorista',
            'Divergência apontada',
            'Aguardando liberação',
            'Liberada para NF',
            'NF enviada',
            'Em análise',
            'Aprovada para pagamento',
            'Pagamento confirmado',
            'Bloqueada',
            'Cancelada'
        ]

        if not data_lancamento:
            flash("Informe a data da rota.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('editar_rota', id=id))

        if not identi_rota:
            flash("Informe a identificação da rota.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('editar_rota', id=id))

        if not tipo_rota:
            flash("Selecione o tipo da rota.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('editar_rota', id=id))

        if not motorista_id or not str(motorista_id).isdigit():
            flash("Selecione um motorista válido.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('editar_rota', id=id))

        if situacao_rota not in ['Pendente', 'Faturada', 'Quitada', 'Cancelada']:
            flash("Situação da rota inválida.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('editar_rota', id=id))

        if status_motorista not in status_motorista_validos:
            flash("Status do motorista inválido.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('editar_rota', id=id))

        try:
            motorista_id = int(motorista_id)

            valor_rota = converter_decimal(valor_rota)
            valor_km = converter_decimal(valor_km)
            outras_despesas = converter_decimal(outras_despesas)

            cur.execute("""
                        SELECT id,
                               COALESCE(status_motorista, 'Aguardando liberação') AS status_motorista
                        FROM rotas
                        WHERE id = %s
                          AND empresa_id = %s LIMIT 1
                        """, (id, empresa_id))

            rota_atual = cur.fetchone()

            if not rota_atual:
                flash("Rota não encontrada ou não pertence à empresa logada.", "danger")
                return redirect(url_for('visualizar_rotas'))

            possui_documento_ativo = rota_tem_documento_ativo(cur, id, empresa_id)

            if not rota_pode_ser_editada(rota_atual, possui_documento_ativo):
                flash(
                    "Esta rota não pode ser editada, pois já entrou no fluxo fiscal/pagamento. "
                    "Para preservar a auditoria, reverta o documento ou utilize um fluxo de reabertura/correção.",
                    "danger"
                )
                return redirect(url_for('visualizar_rotas'))


            cur.execute("""
                        SELECT id
                        FROM pessoas
                        WHERE id = %s
                          AND empresa_id = %s
                          AND status_cadastro = 'Ativo'
                          AND (tipo_cadastro = 'Motorista' OR (tipo_cadastro = 'Prestador de Serviço' AND COALESCE(tipo_prestador, '') IN ('Motorista', 'Motorista e Ajudante'))) LIMIT 1
                        """, (motorista_id, empresa_id))

            if not cur.fetchone():
                flash("Motorista inválido ou não pertence à empresa logada.", "danger")
                return redirect(url_for('editar_rota', id=id))

            cur.execute("""
                        UPDATE rotas
                        SET data_lancamento   = %s,
                            identi_rota       = %s,
                            valor_rota        = %s,
                            valor_km          = %s,
                            outras_despesas   = %s,
                            tipo_rota         = %s,
                            motorista_id      = %s,
                            situacao_rota     = %s,
                            status_motorista  = %s
                        WHERE id = %s
                          AND empresa_id = %s
                        """, (
                            data_lancamento,
                            identi_rota,
                            valor_rota,
                            valor_km,
                            outras_despesas,
                            tipo_rota,
                            motorista_id,
                            situacao_rota,
                            status_motorista,
                            id,
                            empresa_id
                        ))

            con.commit()

            flash("Rota atualizada com sucesso!", "success")
            return redirect(url_for('visualizar_rotas'))

        except Exception as e:
            con.rollback()
            print(f"Erro ao editar rota: {e}")
            flash("Erro técnico ao editar rota.", "danger")
            return redirect(url_for('editar_rota', id=id))

        finally:
            fechar_cursor_conexao(cur, con)

    try:
        cur.execute("""
                    SELECT id,
                           empresa_id,
                           data_lancamento,
                           identi_rota,
                           tipo_rota,
                           valor_rota,
                           valor_km,
                           outras_despesas,
                           transportadora_id,
                           motorista_id,
                           situacao_rota,
                           COALESCE(status_motorista, 'Aguardando liberação') AS status_motorista
                    FROM rotas
                    WHERE id = %s
                      AND empresa_id = %s LIMIT 1
                    """, (id, empresa_id))

        rota = cur.fetchone()

        if not rota:
            flash("Rota não encontrada ou não pertence à empresa logada.", "warning")
            return redirect(url_for('visualizar_rotas'))

        possui_documento_ativo = rota_tem_documento_ativo(cur, id, empresa_id)

        if not rota_pode_ser_editada(rota, possui_documento_ativo):
            flash(
                "Esta rota não pode ser editada, pois já entrou no fluxo fiscal/pagamento.",
                "warning"
            )
            return redirect(url_for('visualizar_rotas'))


        cur.execute("""
                    SELECT id, nome_completo
                    FROM pessoas
                    WHERE empresa_id = %s
                      AND tipo_cadastro = 'Motorista'
                      AND status_cadastro = 'Ativo'
                    ORDER BY nome_completo ASC
                    """, (empresa_id,))
        motoristas = cur.fetchall()

    except Exception as e:
        print(f"Erro ao carregar rota: {e}")
        flash("Erro técnico ao carregar rota.", "danger")
        return redirect(url_for('visualizar_rotas'))

    finally:
        fechar_cursor_conexao(cur, con)

    return render_template(
        'editar_rota.html',
        rota=rota,
        motoristas=motoristas,
        usuario_logado=usuario_logado
    )


@app.route('/movimentacao/excluir/<int:id>', methods=['POST'])
@login_required
@perfis_permitidos('Administrador')
def excluir_rota(id):
    empresa_id = session.get('empresa_id')

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('visualizar_rotas'))

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT id, identi_rota, situacao_rota
                    FROM rotas
                    WHERE id = %s
                      AND empresa_id = %s LIMIT 1
                    """, (id, empresa_id))

        rota = cur.fetchone()

        if not rota:
            flash("Rota não encontrada ou não pertence à empresa logada.", "danger")
            return redirect(url_for('visualizar_rotas'))

        cur.execute("""
                    SELECT id
                    FROM nota_fiscal_rotas
                    WHERE rota_id = %s
                      AND empresa_id = %s LIMIT 1
                    """, (id, empresa_id))

        vinculada_nf = cur.fetchone()

        if vinculada_nf:
            flash("Não é possível excluir esta rota, pois ela está vinculada a uma Nota Fiscal.", "danger")
            return redirect(url_for('visualizar_rotas'))

        cur.execute("""
                    SELECT id
                    FROM lancamento_ajudante_rotas
                    WHERE rota_id = %s
                      AND empresa_id = %s LIMIT 1
                    """, (id, empresa_id))

        vinculada_ajudante = cur.fetchone()

        if vinculada_ajudante:
            flash("Não é possível excluir esta rota, pois ela está vinculada a um lançamento de ajudante.", "danger")
            return redirect(url_for('visualizar_rotas'))

        cur.execute("""
                    DELETE
                    FROM rotas
                    WHERE id = %s
                      AND empresa_id = %s
                    """, (id, empresa_id))

        con.commit()

        flash(f"Rota {rota['identi_rota']} excluída com sucesso.", "success")

    except Exception as e:
        con.rollback()
        print(f"Erro ao excluir rota: {e}")
        flash("Erro técnico ao excluir rota.", "danger")

    finally:
        cur.close()
        con.close()

    return redirect(url_for('visualizar_rotas'))


# ==========================================================
# LANÇAMENTO DE AJUDANTE
# ==========================================================
@app.route('/rotas/lancamento-ajudante', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional')
def lancamento_ajudante():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_id = session.get('empresa_id')

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    con = obter_conexao()

    if con is None:
        flash('Erro de conexão com o banco de dados.', 'danger')
        return redirect(url_for('dashboard'))

    cur = con.cursor(dictionary=True)

    if request.method == 'POST':
        ajudante_id = request.form.get('ajudante_id')
        rotas_selecionadas = request.form.getlist('rotas_selecionadas')
        observacao = limitar_texto(request.form.get('observacao'), 150)

        if not ajudante_id or not str(ajudante_id).isdigit():
            flash('Selecione um ajudante válido.', 'danger')
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('lancamento_ajudante'))

        if not rotas_selecionadas:
            flash('Selecione pelo menos uma rota.', 'danger')
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('lancamento_ajudante'))

        try:
            ajudante_id = int(ajudante_id)
            rota_ids = [int(r) for r in rotas_selecionadas if str(r).isdigit()]
        except ValueError:
            flash('Dados inválidos no envio do formulário.', 'danger')
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('lancamento_ajudante'))

        if len(rota_ids) != len(rotas_selecionadas):
            flash('Uma ou mais rotas selecionadas são inválidas.', 'danger')
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('lancamento_ajudante'))

        valores_por_rota = {}
        valor_total = Decimal('0.00')

        for rota_id in rota_ids:
            valor_ajudante = converter_decimal(request.form.get(f'valor_ajudante_{rota_id}', '0'))

            if valor_ajudante <= 0:
                flash('Informe valor maior que zero para todas as rotas selecionadas.', 'danger')
                fechar_cursor_conexao(cur, con)
                return redirect(url_for('lancamento_ajudante'))

            valores_por_rota[rota_id] = valor_ajudante
            valor_total += valor_ajudante

        try:
            cur.execute("""
                        SELECT id, nome_completo
                        FROM pessoas
                        WHERE id = %s
                          AND empresa_id = %s
                          AND status_cadastro = 'Ativo'
                          AND (tipo_cadastro = 'Ajudante' OR (tipo_cadastro = 'Prestador de Serviço' AND COALESCE(tipo_prestador, '') IN ('Ajudante', 'Motorista e Ajudante'))) LIMIT 1
                        """, (ajudante_id, empresa_id))

            ajudante = cur.fetchone()

            if not ajudante:
                flash('Ajudante inválido, inativo ou não pertence à empresa logada.', 'danger')
                return redirect(url_for('lancamento_ajudante'))

            placeholders = ','.join(['%s'] * len(rota_ids))

            cur.execute(f"""
                SELECT
                    r.id,
                    r.identi_rota,
                    r.situacao_rota
                FROM rotas r
                LEFT JOIN lancamento_ajudante_rotas lar
                       ON lar.rota_id = r.id
                      AND lar.empresa_id = r.empresa_id
                WHERE r.empresa_id = %s
                  AND r.id IN ({placeholders})
                  AND lar.id IS NULL
                  AND r.situacao_rota <> 'Cancelada'
            """, tuple([empresa_id] + rota_ids))

            rotas_validas = cur.fetchall()

            if len(rotas_validas) != len(rota_ids):
                flash(
                    'Uma ou mais rotas já possuem ajudante vinculado, estão canceladas ou não pertencem à empresa logada.',
                    'danger')
                return redirect(url_for('lancamento_ajudante'))

            cur.execute("""
                        INSERT INTO lancamentos_ajudantes (empresa_id,
                                                           ajudante_id,
                                                           valor_total,
                                                           status_pagamento,
                                                           observacao)
                        VALUES (%s, %s, %s, 'Pendente', %s)
                        """, (
                            empresa_id,
                            ajudante_id,
                            valor_total,
                            observacao or None
                        ))

            lancamento_id = cur.lastrowid

            for rota in rotas_validas:
                rota_id = rota['id']

                cur.execute("""
                            INSERT INTO lancamento_ajudante_rotas (empresa_id,
                                                                   lancamento_ajudante_id,
                                                                   rota_id,
                                                                   identi_rota,
                                                                   valor_ajudante)
                            VALUES (%s, %s, %s, %s, %s)
                            """, (
                                empresa_id,
                                lancamento_id,
                                rota_id,
                                rota['identi_rota'],
                                valores_por_rota[rota_id],
                            ))

            con.commit()

            flash(f"Lançamento do ajudante {ajudante['nome_completo']} salvo com sucesso!", 'success')
            return redirect(url_for('lancamento_ajudante'))

        except mysql.connector.Error as err:
            con.rollback()

            if err.errno == 1062:
                flash('Uma das rotas selecionadas já possui ajudante vinculado.', 'danger')
            else:
                print(f'Erro ao salvar lançamento de ajudante: {err}')
                flash('Erro técnico ao salvar o lançamento do ajudante.', 'danger')

            return redirect(url_for('lancamento_ajudante'))

        except Exception as e:
            con.rollback()
            print(f'Erro ao salvar lançamento de ajudante: {e}')
            flash('Erro técnico ao salvar o lançamento do ajudante.', 'danger')
            return redirect(url_for('lancamento_ajudante'))

        finally:
            fechar_cursor_conexao(cur, con)

    try:
        cur.execute("""
                    SELECT id, nome_completo
                    FROM pessoas
                    WHERE empresa_id = %s
                      AND tipo_cadastro = 'Ajudante'
                      AND status_cadastro = 'Ativo'
                    ORDER BY nome_completo ASC
                    """, (empresa_id,))

        ajudantes = cur.fetchall()

        cur.execute("""
                    SELECT r.id,
                           r.data_lancamento,
                           r.identi_rota,
                           r.tipo_rota,
                           r.valor_rota,
                           r.valor_km,
                           r.outras_despesas,
                           r.situacao_rota,
                           p.nome_completo AS transportadora,
                           (
                               COALESCE(r.valor_rota, 0) +
                               COALESCE(r.valor_km, 0) +
                               COALESCE(r.outras_despesas, 0)
                               )           AS total_rota
                    FROM rotas r
                             LEFT JOIN pessoas p
                                       ON p.id = r.transportadora_id
                                           AND p.empresa_id = r.empresa_id
                             LEFT JOIN lancamento_ajudante_rotas lar
                                       ON lar.rota_id = r.id
                                           AND lar.empresa_id = r.empresa_id
                    WHERE r.empresa_id = %s
                      AND lar.id IS NULL
                      AND r.situacao_rota <> 'Cancelada'
                    ORDER BY r.data_lancamento DESC, r.id DESC
                    """, (empresa_id,))

        rotas_disponiveis = cur.fetchall()

    except Exception as e:
        print(f'Erro ao carregar lançamento de ajudante: {e}')
        flash('Erro ao carregar dados da tela de lançamento de ajudante.', 'danger')
        ajudantes = []
        rotas_disponiveis = []

    finally:
        fechar_cursor_conexao(cur, con)

    return render_template(
        'lancamento_ajudante.html',
        usuario_logado=usuario_logado,
        ajudantes=ajudantes,
        rotas_disponiveis=rotas_disponiveis,
    )


# ==========================================================
# FATURAMENTO - ROTAS PENDENTES + NOTAS FATURADAS
# ==========================================================
@app.route('/faturamento')
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro', 'Consulta')
def faturamento():
    usuario_logado = session.get('usuario_nome', 'Usuário')

    con = obter_conexao()
    if con is None:
        flash('Não foi possível conectar ao banco de dados.', 'danger')
        return render_template(
            'faturamento.html',
            rotas_pendentes=[],
            notas_fiscais=[],
            usuario_logado=usuario_logado
        )

    cur = con.cursor(dictionary=True)

    try:
        # Rotas livres para novo faturamento
        cur.execute("""
                    SELECT r.id,
                           r.data_lancamento,
                           r.identi_rota,
                           r.tipo_rota,
                           r.valor_rota,
                           r.valor_km,
                           r.outras_despesas,
                           p.nome_completo AS transportadora,
                           (
                               COALESCE(r.valor_rota, 0) +
                               COALESCE(r.valor_km, 0) +
                               COALESCE(r.outras_despesas, 0)
                               )           AS total_rota
                    FROM rotas r
                             LEFT JOIN pessoas p ON p.id = r.transportadora_id
                             LEFT JOIN nota_fiscal_rotas nfr ON nfr.rota_id = r.id
                    WHERE r.situacao_rota = 'Pendente'
                      AND nfr.id IS NULL
                    ORDER BY r.data_lancamento DESC, r.id DESC
                    """)
        rotas_pendentes = cur.fetchall()

        # Notas para acompanhamento/estorno
        cur.execute("""
                    SELECT nf.id,
                           nf.numero_nf,
                           nf.chave_acesso,
                           nf.data_emissao,
                           nf.valor_total,
                           COALESCE(nf.status_nf, 'Faturada')                                         AS status_nf,
                           nf.data_estorno,
                           nf.motivo_estorno,
                           emit.nome_completo                                                         AS emitente_nome,
                           tom.nome_completo                                                          AS tomador_nome,
                           COUNT(nfr.rota_id)                                                         AS qtd_rotas,
                           COALESCE(SUM(CASE WHEN r.situacao_rota = 'Faturada' THEN 1 ELSE 0 END), 0) AS qtd_faturadas,
                           COALESCE(SUM(CASE WHEN r.situacao_rota = 'Quitada' THEN 1 ELSE 0 END), 0)  AS qtd_quitadas
                    FROM notas_fiscais nf
                             LEFT JOIN pessoas emit ON emit.id = nf.emitente_id
                             LEFT JOIN pessoas tom ON tom.id = nf.tomador_id
                             LEFT JOIN nota_fiscal_rotas nfr ON nfr.nota_fiscal_id = nf.id
                             LEFT JOIN rotas r ON r.id = nfr.rota_id
                    GROUP BY nf.id,
                             nf.numero_nf,
                             nf.chave_acesso,
                             nf.data_emissao,
                             nf.valor_total,
                             nf.status_nf,
                             nf.data_estorno,
                             nf.motivo_estorno,
                             emit.nome_completo,
                             tom.nome_completo
                    ORDER BY nf.data_importacao DESC, nf.id DESC LIMIT 100
                    """)
        notas_fiscais = cur.fetchall()

    except Exception as e:
        print(f'Erro ao carregar faturamento: {e}')
        flash('Erro ao buscar dados de faturamento.', 'danger')
        rotas_pendentes = []
        notas_fiscais = []

    finally:
        fechar_cursor_conexao(cur, con)

    return render_template(
        'faturamento.html',
        rotas_pendentes=rotas_pendentes,
        notas_fiscais=notas_fiscais,
        usuario_logado=usuario_logado
    )


@app.route('/processar-faturamento', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional')
def processar_faturamento():
    usuario_id = session.get('usuario_id')
    empresa_id = session.get('empresa_id')

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    ids_selecionados = request.form.getlist('ids_rotas')
    arquivo_xml = request.files.get('arquivo_xml')

    if not ids_selecionados or not arquivo_xml:
        flash("Selecione pelo menos uma rota e envie o arquivo XML.", "danger")
        return redirect(url_for('faturamento'))

    try:
        rota_ids = [int(x) for x in ids_selecionados]
    except Exception:
        flash("Rotas selecionadas inválidas.", "danger")
        return redirect(url_for('faturamento'))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('faturamento'))

    cur = con.cursor(dictionary=True)

    try:
        tree = ET.parse(arquivo_xml)
        root = tree.getroot()
        ns = {'ns': 'http://www.sped.fazenda.gov.br/nfse'}

        inf_nfse = root.find('.//ns:infNFSe', ns)

        if inf_nfse is None:
            flash("Formato inválido. O sistema esperava uma NFS-e Padrão Nacional.", "danger")
            return redirect(url_for('faturamento'))

        id_atributo = inf_nfse.attrib.get('Id', '')
        chave_acesso = somente_digitos(id_atributo)

        numero_nf = root.find('.//ns:nNFSe', ns).text
        data_emissao = root.find('.//ns:dhEmi', ns).text[:10]
        valor_total_nf = Decimal(root.find('.//ns:valores/ns:vServPrest/ns:vServ', ns).text).quantize(Decimal("0.01"))

        tomador_cnpj = root.find('.//ns:toma/ns:CNPJ', ns).text
        tomador_cnpj_limpo = somente_digitos(tomador_cnpj)

        prestador_cnpj_node = root.find('.//ns:prest/ns:CNPJ', ns)

        if prestador_cnpj_node is None:
            flash("Não foi possível identificar o CNPJ do prestador no XML.", "danger")
            return redirect(url_for('faturamento'))

        prestador_cnpj_limpo = somente_digitos(prestador_cnpj_node.text)

        if not chave_acesso:
            flash("Não foi possível identificar a chave de acesso no XML.", "danger")
            return redirect(url_for('faturamento'))

        # Tomador precisa ser transportadora da empresa logada
        cur.execute("""
                    SELECT id, nome_completo
                    FROM pessoas
                    WHERE empresa_id = %s
                      AND tipo_cadastro = 'Transportadora'
                      AND status_cadastro = 'Ativo'
                      AND REPLACE(REPLACE(REPLACE(REPLACE(cpf_cnpj, '.', ''), '-', ''), '/', ''), ' ', '') = %s LIMIT 1
                    """, (empresa_id, tomador_cnpj_limpo))

        tomador_sistema = cur.fetchone()

        if not tomador_sistema:
            flash(
                f"Faturamento recusado. O tomador {tomador_cnpj_limpo} não está cadastrado como Transportadora ativa nesta empresa.",
                "danger"
            )
            return redirect(url_for('faturamento'))

        # Prestador precisa estar cadastrado na empresa logada
        cur.execute("""
                    SELECT id, nome_completo
                    FROM pessoas
                    WHERE empresa_id = %s
                      AND tipo_cadastro = 'Prestador de Serviço'
                      AND status_cadastro = 'Ativo'
                      AND REPLACE(REPLACE(REPLACE(REPLACE(cpf_cnpj, '.', ''), '-', ''), '/', ''), ' ', '') = %s LIMIT 1
                    """, (empresa_id, prestador_cnpj_limpo))

        emitente_sistema = cur.fetchone()

        if not emitente_sistema:
            flash(
                f"Faturamento recusado. O prestador {prestador_cnpj_limpo} não está cadastrado como Prestador de Serviço ativo nesta empresa.",
                "danger"
            )
            return redirect(url_for('faturamento'))

        placeholders = ','.join(['%s'] * len(rota_ids))

        cur.execute(f"""
            SELECT
                id,
                identi_rota,
                transportadora_id,
                situacao_rota,
                (
                    COALESCE(valor_rota, 0) +
                    COALESCE(valor_km, 0) +
                    COALESCE(outras_despesas, 0)
                ) AS total_rota
            FROM rotas
            WHERE empresa_id = %s
              AND id IN ({placeholders})
        """, tuple([empresa_id] + rota_ids))

        rotas = cur.fetchall()

        if len(rotas) != len(rota_ids):
            flash("Uma ou mais rotas não foram encontradas ou não pertencem à empresa logada.", "danger")
            return redirect(url_for('faturamento'))

        for rota in rotas:
            if rota['situacao_rota'] != 'Pendente':
                flash(f"A rota {rota['identi_rota']} não está pendente.", "danger")
                return redirect(url_for('faturamento'))

            if int(rota['transportadora_id']) != int(tomador_sistema['id']):
                flash(
                    f"A rota {rota['identi_rota']} pertence a outra transportadora. O tomador do XML precisa bater com a transportadora da rota.",
                    "danger"
                )
                return redirect(url_for('faturamento'))

        soma_sistema = sum(
            Decimal(str(rota['total_rota'] or 0)) for rota in rotas
        ).quantize(Decimal("0.01"))

        if abs(soma_sistema - valor_total_nf) > Decimal("0.05"):
            flash(
                f"Divergência financeira. Rotas selecionadas: R$ {soma_sistema} | Valor no XML: R$ {valor_total_nf}.",
                "danger"
            )
            return redirect(url_for('faturamento'))

        cur.execute("""
                    SELECT id, status_nf
                    FROM notas_fiscais
                    WHERE empresa_id = %s
                      AND chave_acesso = %s LIMIT 1
                    """, (empresa_id, chave_acesso))

        nota_existente = cur.fetchone()

        if nota_existente and nota_existente['status_nf'] == 'Faturada':
            flash(f"A Nota Fiscal Nº {numero_nf} já está faturada nesta empresa.", "danger")
            return redirect(url_for('faturamento'))

        if nota_existente and nota_existente['status_nf'] == 'Estornada':
            nota_fiscal_id = nota_existente['id']

            cur.execute("""
                        UPDATE notas_fiscais
                        SET numero_nf          = %s,
                            data_emissao       = %s,
                            valor_total        = %s,
                            emitente_id        = %s,
                            tomador_id         = %s,
                            status_nf          = 'Faturada',
                            data_estorno       = NULL,
                            motivo_estorno     = NULL,
                            usuario_estorno_id = NULL
                        WHERE id = %s
                          AND empresa_id = %s
                        """, (
                            numero_nf,
                            data_emissao,
                            valor_total_nf,
                            emitente_sistema['id'],
                            tomador_sistema['id'],
                            nota_fiscal_id,
                            empresa_id
                        ))

            cur.execute("""
                        DELETE
                        FROM nota_fiscal_rotas
                        WHERE nota_fiscal_id = %s
                          AND empresa_id = %s
                        """, (nota_fiscal_id, empresa_id))

        else:
            cur.execute("""
                        INSERT INTO notas_fiscais (empresa_id,
                                                   numero_nf,
                                                   chave_acesso,
                                                   data_emissao,
                                                   valor_total,
                                                   emitente_id,
                                                   tomador_id,
                                                   status_nf)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, 'Faturada')
                        """, (
                            empresa_id,
                            numero_nf,
                            chave_acesso,
                            data_emissao,
                            valor_total_nf,
                            emitente_sistema['id'],
                            tomador_sistema['id']
                        ))

            nota_fiscal_id = cur.lastrowid

        for rota in rotas:
            cur.execute("""
                        INSERT INTO nota_fiscal_rotas (empresa_id,
                                                       nota_fiscal_id,
                                                       rota_id,
                                                       valor_rota_faturado)
                        VALUES (%s, %s, %s, %s)
                        """, (
                            empresa_id,
                            nota_fiscal_id,
                            rota['id'],
                            rota['total_rota']
                        ))

            cur.execute("""
                        INSERT INTO historico_operacoes (empresa_id,
                                                         tipo_operacao,
                                                         rota_id,
                                                         nota_fiscal_id,
                                                         usuario_id,
                                                         status_anterior,
                                                         status_novo,
                                                         motivo,
                                                         observacao)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            empresa_id,
                            'FATURAMENTO_XML',
                            rota['id'],
                            nota_fiscal_id,
                            usuario_id,
                            'Pendente',
                            'Faturada',
                            'Faturamento via XML',
                            f"NF {numero_nf} vinculada à rota {rota['identi_rota']}"
                        ))

        cur.execute(f"""
            UPDATE rotas
            SET situacao_rota = 'Faturada'
            WHERE empresa_id = %s
              AND id IN ({placeholders})
              AND situacao_rota = 'Pendente'
        """, tuple([empresa_id] + rota_ids))

        con.commit()

        flash(f"Nota Fiscal Nº {numero_nf} faturada e vinculada com sucesso.", "success")

    except Exception as e:
        con.rollback()
        print(f"Erro no faturamento XML: {e}")
        flash("Erro interno ao processar faturamento do XML.", "danger")

    finally:
        fechar_cursor_conexao(cur, con)

    return redirect(url_for('faturamento'))


@app.route('/faturamento/contingencia/<int:id>', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Financeiro')
def contingencia_quitar_rota(id):
    con = obter_conexao()
    if con is None:
        flash('Erro ao conectar ao banco de dados.', 'danger')
        return redirect(url_for('faturamento'))

    cur = con.cursor(dictionary=True)
    try:
        cur.execute('SELECT id, identi_rota, situacao_rota FROM rotas WHERE id = %s', (id,))
        rota = cur.fetchone()

        if not rota:
            flash('Rota não localizada.', 'warning')
            return redirect(url_for('faturamento'))

        if rota['situacao_rota'] != 'Pendente':
            flash(f"A rota {rota['identi_rota']} já não está pendente.", 'warning')
            return redirect(url_for('faturamento'))

        cur.execute("""
                    UPDATE rotas
                    SET situacao_rota = 'Quitada'
                    WHERE id = %s
                      AND situacao_rota = 'Pendente'
                    """, (id,))
        con.commit()
        flash(f"Contingência aplicada! Rota {rota['identi_rota']} movida para Quitada.", 'success')

    except Exception as e:
        con.rollback()
        print(f'Erro na contingência: {e}')
        flash('Erro técnico ao aplicar contingência.', 'danger')
    finally:
        fechar_cursor_conexao(cur, con)

    return redirect(url_for('faturamento'))


# ==========================================================
# RECEBIMENTO - NOTAS A RECEBER + RECEBIMENTOS REALIZADOS
# ==========================================================
@app.route('/financeiro/recebimento')
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro', 'Consulta')
def recebimento():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_id = session.get('empresa_id')

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão ao carregar recebimentos.", "danger")
        return render_template(
            'recebimento.html',
            faturamentos_pendentes=[],
            usuario_logado=usuario_logado
        )

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT nf.id,
                           nf.numero_nf,
                           nf.chave_acesso AS chave_nota,
                           nf.valor_total,
                           nf.status_nf,
                           p.nome_completo AS nome_parceiro,
                           p.cpf_cnpj,
                           COUNT(nfr.id)   AS qtd_rotas
                    FROM notas_fiscais nf
                             INNER JOIN pessoas p
                                        ON p.id = nf.tomador_id
                                            AND p.empresa_id = nf.empresa_id
                             INNER JOIN nota_fiscal_rotas nfr
                                        ON nfr.nota_fiscal_id = nf.id
                                            AND nfr.empresa_id = nf.empresa_id
                             INNER JOIN rotas r
                                        ON r.id = nfr.rota_id
                                            AND r.empresa_id = nfr.empresa_id
                    WHERE nf.empresa_id = %s
                      AND nf.status_nf = 'Faturada'
                      AND r.situacao_rota = 'Faturada'
                    GROUP BY nf.id,
                             nf.numero_nf,
                             nf.chave_acesso,
                             nf.valor_total,
                             nf.status_nf,
                             p.nome_completo,
                             p.cpf_cnpj
                    ORDER BY nf.data_emissao DESC, nf.id DESC
                    """, (empresa_id,))

        faturamentos_pendentes = cur.fetchall()

    except Exception as e:
        print(f"Erro ao carregar recebimentos: {e}")
        flash("Erro ao carregar recebimentos.", "danger")
        faturamentos_pendentes = []

    finally:
        fechar_cursor_conexao(cur, con)

    return render_template(
        'recebimento.html',
        faturamentos_pendentes=faturamentos_pendentes,
        usuario_logado=usuario_logado
    )


# ==========================================================
# CONFIRMAR RECEBIMENTO
# Rotas Faturadas -> Quitada
# ==========================================================
@app.route('/financeiro/recebimento/confirmar/<int:id>', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Financeiro')
def confirmar_recebimento(id):
    usuario_id = session.get('usuario_id')
    empresa_id = session.get('empresa_id')

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão ao confirmar recebimento.", "danger")
        return redirect(url_for('recebimento'))

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT id, numero_nf, status_nf
                    FROM notas_fiscais
                    WHERE id = %s
                      AND empresa_id = %s
                      AND status_nf = 'Faturada' LIMIT 1
                    """, (id, empresa_id))

        nota = cur.fetchone()

        if not nota:
            flash("Nota Fiscal não encontrada, já recebida ou não pertence à empresa logada.", "warning")
            return redirect(url_for('recebimento'))

        cur.execute("""
                    SELECT r.id, r.identi_rota, r.situacao_rota
                    FROM nota_fiscal_rotas nfr
                             INNER JOIN rotas r
                                        ON r.id = nfr.rota_id
                                            AND r.empresa_id = nfr.empresa_id
                    WHERE nfr.nota_fiscal_id = %s
                      AND nfr.empresa_id = %s
                      AND r.situacao_rota = 'Faturada'
                    """, (id, empresa_id))

        rotas = cur.fetchall()

        if not rotas:
            flash("Não há rotas faturadas para baixar nesta nota.", "warning")
            return redirect(url_for('recebimento'))

        rota_ids = [rota['id'] for rota in rotas]
        placeholders = ','.join(['%s'] * len(rota_ids))

        cur.execute(f"""
            UPDATE rotas
            SET situacao_rota = 'Quitada'
            WHERE empresa_id = %s
              AND id IN ({placeholders})
              AND situacao_rota = 'Faturada'
        """, tuple([empresa_id] + rota_ids))

        for rota in rotas:
            cur.execute("""
                        INSERT INTO historico_operacoes (empresa_id,
                                                         tipo_operacao,
                                                         rota_id,
                                                         nota_fiscal_id,
                                                         usuario_id,
                                                         status_anterior,
                                                         status_novo,
                                                         motivo,
                                                         observacao)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            empresa_id,
                            'BAIXA_RECEBIMENTO',
                            rota['id'],
                            id,
                            usuario_id,
                            'Faturada',
                            'Quitada',
                            'Confirmação de recebimento',
                            f"Recebimento confirmado da NF {nota['numero_nf']}."
                        ))

        con.commit()

        flash(f"Recebimento da NF {nota['numero_nf']} confirmado com sucesso.", "success")

    except Exception as e:
        con.rollback()
        print(f"Erro ao confirmar recebimento: {e}")
        flash("Erro técnico ao confirmar recebimento.", "danger")

    finally:
        fechar_cursor_conexao(cur, con)

    return redirect(url_for('recebimento'))


# ==========================================================
# USUÁRIOS MULTIEMPRESA
# # criar_usuario, visualizar_usuarios, editar_usuario, alternar_status_usuario
# ==========================================================

def usuario_eh_super_admin():
    return int(session.get('is_super_admin') or 0) == 1


def carregar_empresas_ativas():
    con = obter_conexao()
    if con is None:
        return []

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT id,
                           razao_social,
                           nome_fantasia,
                           slug,
                           status_empresa,
                           plano,
                           limite_usuarios
                    FROM empresas
                    WHERE status_empresa = 'Ativa'
                    ORDER BY nome_fantasia ASC, razao_social ASC
                    """)
        return cur.fetchall()

    except Exception as e:
        print(f"Erro ao carregar empresas ativas: {e}")
        return []

    finally:
        fechar_cursor_conexao(cur, con)


def carregar_pessoas_disponiveis_usuario(empresa_id=None, usuario_id_atual=None):
    con = obter_conexao()
    if con is None:
        return []

    cur = con.cursor(dictionary=True)

    try:
        query = """
                SELECT p.id,
                       p.empresa_id,
                       p.nome_completo,
                       p.tipo_cadastro,
                       e.nome_fantasia AS empresa_nome,
                       e.razao_social  AS empresa_razao_social
                FROM pessoas p
                         INNER JOIN empresas e
                                    ON e.id = p.empresa_id
                         LEFT JOIN usuarios u
                                   ON u.pessoa_id = p.id
                                       AND (%s IS NULL OR u.id <> %s)
                WHERE p.status_cadastro = 'Ativo'
                  AND u.id IS NULL \
                """

        params = [usuario_id_atual, usuario_id_atual]

        if empresa_id:
            query += " AND p.empresa_id = %s"
            params.append(empresa_id)

        query += " ORDER BY e.nome_fantasia ASC, p.nome_completo ASC"

        cur.execute(query, params)
        return cur.fetchall()

    except Exception as e:
        print(f"Erro ao carregar pessoas disponíveis para usuário: {e}")
        return []

    finally:
        fechar_cursor_conexao(cur, con)


def validar_empresa_para_usuario(empresa_id):
    con = obter_conexao()
    if con is None:
        return None

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT id,
                           razao_social,
                           nome_fantasia,
                           status_empresa,
                           limite_usuarios
                    FROM empresas
                    WHERE id = %s LIMIT 1
                    """, (empresa_id,))
        return cur.fetchone()

    except Exception as e:
        print(f"Erro ao validar empresa: {e}")
        return None

    finally:
        fechar_cursor_conexao(cur, con)


@app.route('/configuracao/usuarios/criar', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador')
def criar_usuario():
    is_super = usuario_eh_super_admin_global()
    empresa_logada_id = session.get('empresa_id')

    conexao = obter_conexao()
    cursor = conexao.cursor(dictionary=True)

    if request.method == 'POST':
        empresa_id = request.form.get('empresa_id') if is_super else empresa_logada_id
        perfil_de_acesso = request.form.get('perfil_de_acesso')
        login = request.form.get('login', '').strip()
        senha = request.form.get('senha')
        base_operacional_id = request.form.get('base_operacional_id')

        if perfil_de_acesso == 'Terminal Base':
            pessoa_id = None
            if not base_operacional_id:
                flash("Para o perfil Terminal Base, a seleção da base física é obrigatória.", "danger")
                return redirect(url_for('criar_usuario'))
        else:
            pessoa_id = request.form.get('pessoa_id')
            base_operacional_id = None
            if not pessoa_id:
                flash("Selecione uma pessoa física para vincular a este usuário.", "danger")
                return redirect(url_for('criar_usuario'))

        if not login or not senha or not perfil_de_acesso or not empresa_id:
            flash("Preencha todos os campos obrigatórios (*).", "danger")
            return redirect(url_for('criar_usuario'))

        try:
            cursor.execute("SELECT id FROM usuarios WHERE login = %s", (login,))
            if cursor.fetchone():
                flash(f"O login '{login}' já está em uso no sistema. Escolha outro.", "warning")
                return redirect(url_for('criar_usuario'))

            if perfil_de_acesso == 'Terminal Base':
                base_validada = validar_base_operacional_usuario(cursor, base_operacional_id, empresa_id)
                if not base_validada:
                    flash("A base operacional selecionada não pertence à empresa informada ou está inativa.", "danger")
                    return redirect(url_for('criar_usuario'))
            else:
                pessoa_vinculada = buscar_pessoa_para_vinculo_usuario(cursor, pessoa_id)
                if not pessoa_vinculada:
                    flash("Pessoa vinculada não encontrada.", "danger")
                    return redirect(url_for('criar_usuario'))

                if pessoa_vinculada.get('status_cadastro') != 'Ativo':
                    flash("Não é possível criar usuário para pessoa inativa.", "danger")
                    return redirect(url_for('criar_usuario'))

                if str(pessoa_vinculada.get('empresa_id')) != str(empresa_id):
                    flash(
                        "A empresa do usuário deve ser a mesma empresa do cadastro da pessoa vinculada. "
                        "Se a mesma pessoa atua em outra empresa, crie um cadastro separado nessa empresa.",
                        "danger"
                    )
                    return redirect(url_for('criar_usuario'))

                cursor.execute("SELECT id FROM usuarios WHERE pessoa_id = %s", (pessoa_id,))
                if cursor.fetchone():
                    flash("Esta pessoa já possui usuário vinculado.", "warning")
                    return redirect(url_for('criar_usuario'))

            senha_hash = generate_password_hash(senha)
            perfil_local = buscar_perfil_local_por_nome_codigo(cursor, perfil_de_acesso, empresa_id)
            perfil_id_resolvido = perfil_local.get('id') if perfil_local else None

            query = """
                    INSERT INTO usuarios
                    (empresa_id, pessoa_id, base_operacional_id, login, senha_hash, perfil_de_acesso, perfil_id, status_usuario,
                     is_super_admin, data_cadastro)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, 'Ativo', 0, NOW()) \
                    """
            cursor.execute(query, (empresa_id, pessoa_id, base_operacional_id, login, senha_hash, perfil_de_acesso, perfil_id_resolvido))
            conexao.commit()

            flash("Usuário criado com sucesso!", "success")
            return redirect(url_for('visualizar_usuarios'))

        except Exception as e:
            conexao.rollback()
            print(f"Erro ao criar usuário: {e}")
            flash(f"Erro técnico ao processar criação: {e}", "danger")
        finally:
            cursor.close()
            conexao.close()

    try:
        if is_super:
            cursor.execute("SELECT id, razao_social, nome_fantasia FROM empresas ORDER BY nome_fantasia")
            empresas = cursor.fetchall()
            cursor.execute(
                "SELECT id, nome_completo, empresa_id, tipo_cadastro, cpf_cnpj FROM pessoas WHERE status_cadastro='Ativo' ORDER BY nome_completo")
            pessoas = cursor.fetchall()
            cursor.execute(
                "SELECT id, nome_base, empresa_id, endereco FROM bases_operacionais WHERE status_base='Ativa' ORDER BY nome_base")
            bases_operacionais = cursor.fetchall()
        else:
            empresas = []
            cursor.execute(
                "SELECT id, nome_completo, empresa_id, tipo_cadastro, cpf_cnpj FROM pessoas WHERE empresa_id = %s AND status_cadastro='Ativo' ORDER BY nome_completo",
                (empresa_logada_id,))
            pessoas = cursor.fetchall()
            cursor.execute(
                "SELECT id, nome_base, empresa_id, endereco FROM bases_operacionais WHERE empresa_id = %s AND status_base='Ativa' ORDER BY nome_base",
                (empresa_logada_id,))
            bases_operacionais = cursor.fetchall()

        try:
            sincronizar_perfis_padrao_empresas(cursor, empresa_logada_id if not is_super else None)
            conexao.commit()
        except Exception:
            pass
        perfis_disponiveis = listar_perfis_disponiveis_para_empresa(cursor, empresa_logada_id, is_super)

        return render_template('criar_usuario.html',
                               is_super_admin=is_super,
                               empresas=empresas,
                               pessoas=pessoas,
                               bases_operacionais=bases_operacionais,
                               perfis_disponiveis=perfis_disponiveis)
    except Exception as e:
        print(f"Erro ao carregar formulário de criação de usuário: {e}")
        flash("Erro ao carregar dados complementares.", "danger")
        return redirect(url_for('visualizar_usuarios'))
    finally:
        if conexao and conexao.is_connected():
            cursor.close()
            conexao.close()


@app.route('/configuracao/usuarios/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador')
def editar_usuario(id):
    is_super = usuario_eh_super_admin_global()
    empresa_logada_id = session.get('empresa_id')
    usuario_sessao_id = session.get('usuario_id')

    conexao = obter_conexao()
    cursor = conexao.cursor(dictionary=True)

    try:
        cursor.execute("""
                       SELECT u.*, p.nome_completo
                       FROM usuarios u
                                LEFT JOIN pessoas p ON u.pessoa_id = p.id
                       WHERE u.id = %s
                       """, (id,))
        usuario = cursor.fetchone()
    except Exception as e:
        print(f"Erro ao buscar usuário para edição: {e}")
        flash("Erro ao conectar ao banco de dados.", "danger")
        cursor.close()
        conexao.close()
        return redirect(url_for('visualizar_usuarios'))

    if not usuario:
        flash("Usuário não encontrado.", "danger")
        cursor.close()
        conexao.close()
        return redirect(url_for('visualizar_usuarios'))

    if not is_super and str(usuario['empresa_id']) != str(empresa_logada_id):
        flash("Acesso negado. Este usuário pertence a outra organização.", "danger")
        cursor.close()
        conexao.close()
        return redirect(url_for('visualizar_usuarios'))

    if request.method == 'POST':
        empresa_id = request.form.get('empresa_id') if is_super else empresa_logada_id
        perfil_de_acesso = request.form.get('perfil_de_acesso')
        login = request.form.get('login', '').strip()
        nova_senha = request.form.get('senha')
        status_usuario = request.form.get('status_usuario', 'Ativo')
        base_operacional_id = request.form.get('base_operacional_id')

        if id == usuario_sessao_id and status_usuario == 'Bloqueado':
            flash("Ação abortada. Não é permitido bloquear o seu próprio usuário logado.", "warning")
            cursor.close()
            conexao.close()
            return redirect(url_for('editar_usuario', id=id))

        if perfil_de_acesso == 'Terminal Base':
            pessoa_id = None
            if not base_operacional_id:
                flash("Selecione uma Base Operacional física para este terminal.", "danger")
                cursor.close()
                conexao.close()
                return redirect(url_for('editar_usuario', id=id))
        else:
            pessoa_id = request.form.get('pessoa_id')
            base_operacional_id = None
            if not pessoa_id:
                flash("Este perfil exige o preenchimento de uma Pessoa Vinculada.", "danger")
                cursor.close()
                conexao.close()
                return redirect(url_for('editar_usuario', id=id))

        if not login or not perfil_de_acesso:
            flash("Campos obrigatórios em branco.", "danger")
            cursor.close()
            conexao.close()
            return redirect(url_for('editar_usuario', id=id))

        try:
            cursor.execute("SELECT id FROM usuarios WHERE login = %s AND id <> %s", (login, id))
            if cursor.fetchone():
                flash(f"O login '{login}' já está sendo utilizado por outro cadastro.", "warning")
                return redirect(url_for('editar_usuario', id=id))

            if perfil_de_acesso == 'Terminal Base':
                base_validada = validar_base_operacional_usuario(cursor, base_operacional_id, empresa_id)
                if not base_validada:
                    flash("A base operacional selecionada não pertence à empresa informada ou está inativa.", "danger")
                    return redirect(url_for('editar_usuario', id=id))
            else:
                pessoa_vinculada = buscar_pessoa_para_vinculo_usuario(cursor, pessoa_id)
                if not pessoa_vinculada:
                    flash("Pessoa vinculada não encontrada.", "danger")
                    return redirect(url_for('editar_usuario', id=id))

                if pessoa_vinculada.get('status_cadastro') != 'Ativo':
                    flash("Não é possível vincular usuário a uma pessoa inativa.", "danger")
                    return redirect(url_for('editar_usuario', id=id))

                if str(pessoa_vinculada.get('empresa_id')) != str(empresa_id):
                    flash(
                        "A empresa do usuário deve ser a mesma empresa do cadastro da pessoa vinculada. "
                        "Se a mesma pessoa atua em outra empresa, crie um cadastro separado nessa empresa.",
                        "danger"
                    )
                    return redirect(url_for('editar_usuario', id=id))

                cursor.execute("SELECT id FROM usuarios WHERE pessoa_id = %s AND id <> %s", (pessoa_id, id))
                if cursor.fetchone():
                    flash("Esta pessoa já possui outro usuário vinculado.", "warning")
                    return redirect(url_for('editar_usuario', id=id))

            mudanca_vinculo_sensivel = (
                str(usuario.get('pessoa_id') or '') != str(pessoa_id or '')
                or str(usuario.get('empresa_id') or '') != str(empresa_id or '')
                or (usuario.get('perfil_de_acesso') != perfil_de_acesso and usuario.get('perfil_de_acesso') in ['Motorista', 'Terminal Base'])
                or (usuario.get('perfil_de_acesso') != perfil_de_acesso and perfil_de_acesso in ['Motorista', 'Terminal Base'])
            )

            if mudanca_vinculo_sensivel and usuario.get('pessoa_id'):
                movimentos = pessoa_tem_movimentacao_empresa(cursor, usuario.get('pessoa_id'), usuario.get('empresa_id'))
                if int(movimentos.get('total') or 0) > 0:
                    detalhes = '; '.join(movimentos.get('detalhes') or [])
                    flash(
                        "Não é permitido trocar a pessoa, empresa ou perfil operacional deste usuário, "
                        "pois o cadastro vinculado já possui movimentações. Crie um novo cadastro/usuário "
                        "para a outra empresa, mantendo o histórico atual preservado."
                        + (f" Movimentações encontradas: {detalhes}." if detalhes else ""),
                        "danger"
                    )
                    return redirect(url_for('editar_usuario', id=id))

            perfil_local = buscar_perfil_local_por_nome_codigo(cursor, perfil_de_acesso, empresa_id)
            perfil_id_resolvido = perfil_local.get('id') if perfil_local else None

            query = """
                    UPDATE usuarios
                    SET empresa_id          = %s,
                        pessoa_id           = %s,
                        base_operacional_id = %s,
                        login               = %s,
                        perfil_de_acesso    = %s,
                        perfil_id           = %s,
                        status_usuario      = %s
                    WHERE id = %s \
                    """
            cursor.execute(query,
                           (empresa_id, pessoa_id, base_operacional_id, login, perfil_de_acesso, perfil_id_resolvido, status_usuario, id))

            if nova_senha and nova_senha.strip() != '':
                if len(nova_senha.strip()) < 6:
                    flash("A nova senha precisa ter no mínimo 6 dígitos.", "warning")
                    return redirect(url_for('editar_usuario', id=id))
                senha_hash = generate_password_hash(nova_senha)
                cursor.execute("UPDATE usuarios SET senha_hash = %s WHERE id = %s", (senha_hash, id))

            conexao.commit()
            flash("Usuário atualizado com sucesso!", "success")
            return redirect(url_for('visualizar_usuarios'))

        except Exception as e:
            conexao.rollback()
            print(f"Erro ao editar usuário: {e}")
            flash(f"Erro interno no processamento da edição: {e}", "danger")
            return redirect(url_for('editar_usuario', id=id))
        finally:
            cursor.close()
            conexao.close()

    try:
        if is_super:
            cursor.execute("SELECT id, razao_social, nome_fantasia FROM empresas ORDER BY nome_fantasia")
            empresas = cursor.fetchall()
            cursor.execute(
                "SELECT id, nome_completo, empresa_id, tipo_cadastro, cpf_cnpj FROM pessoas WHERE status_cadastro='Ativo' ORDER BY nome_completo")
            pessoas = cursor.fetchall()
            cursor.execute(
                "SELECT id, nome_base, empresa_id, endereco FROM bases_operacionais WHERE status_base='Ativa' ORDER BY nome_base")
            bases_operacionais = cursor.fetchall()
        else:
            empresas = []
            cursor.execute(
                "SELECT id, nome_completo, empresa_id, tipo_cadastro, cpf_cnpj FROM pessoas WHERE empresa_id = %s AND status_cadastro='Ativo' ORDER BY nome_completo",
                (usuario['empresa_id'],))
            pessoas = cursor.fetchall()
            cursor.execute(
                "SELECT id, nome_base, empresa_id, endereco FROM bases_operacionais WHERE empresa_id = %s AND status_base='Ativa' ORDER BY nome_base",
                (usuario['empresa_id'],))
            bases_operacionais = cursor.fetchall()

        try:
            sincronizar_perfis_padrao_empresas(cursor, usuario.get('empresa_id'))
            conexao.commit()
        except Exception:
            pass
        perfis_disponiveis = listar_perfis_disponiveis_para_empresa(cursor, usuario.get('empresa_id'), is_super)

        return render_template('editar_usuario.html',
                               usuario=usuario,
                               is_super_admin=is_super,
                               empresas=empresas,
                               pessoas=pessoas,
                               bases_operacionais=bases_operacionais,
                               perfis_disponiveis=perfis_disponiveis)
    except Exception as e:
        print(f"Erro ao carregar dados de edição: {e}")
        flash("Falha ao recuperar informações de apoio técnico.", "danger")
        return redirect(url_for('visualizar_usuarios'))
    finally:
        if conexao and conexao.is_connected():
            cursor.close()
            conexao.close()


@app.route('/configuracao/usuarios')
@login_required
@perfis_permitidos('Administrador')
def visualizar_usuarios():
    is_super = usuario_eh_super_admin_global()
    empresa_logada_id = session.get('empresa_id')

    pesquisa = request.args.get('pesquisa_usuario', '').strip()
    empresa_filtro = request.args.get('empresa_id', '').strip()
    perfil_filtro = request.args.get('perfil_de_acesso', '').strip()

    try:
        page = int(request.args.get('page', 1))
    except ValueError:
        page = 1
    per_page = 15
    offset = (page - 1) * per_page

    conexao = obter_conexao()
    cursor = conexao.cursor(dictionary=True)

    try:
        query_base = """
            FROM usuarios u
            LEFT JOIN pessoas p ON u.pessoa_id = p.id
            LEFT JOIN bases_operacionais b ON u.base_operacional_id = b.id
            LEFT JOIN empresas e ON u.empresa_id = e.id
            WHERE 1=1
        """
        params = []

        if not is_super:
            query_base += " AND u.empresa_id = %s"
            params.append(empresa_logada_id)
        elif empresa_filtro:
            query_base += " AND u.empresa_id = %s"
            params.append(empresa_filtro)

        if perfil_filtro:
            query_base += " AND u.perfil_de_acesso = %s"
            params.append(perfil_filtro)

        if pesquisa:
            query_base += " AND (u.login LIKE %s OR p.nome_completo LIKE %s OR b.nome_base LIKE %s)"
            termo = f"%{pesquisa}%"
            params.extend([termo, termo, termo])

        cursor.execute(f"SELECT COUNT(u.id) AS total {query_base}", params)
        total_records = cursor.fetchone()['total']
        total_pages = (total_records + per_page - 1) // per_page

        query_select = f"""
            SELECT 
                u.id, 
                u.login, 
                u.perfil_de_acesso, 
                u.status_usuario, 
                u.is_super_admin, 
                u.data_cadastro,
                u.empresa_id,
                p.nome_completo AS nome_pessoa,
                b.nome_base AS nome_base_fisica,
                e.nome_fantasia AS nome_empresa
            {query_base}
            ORDER BY u.id DESC
            LIMIT %s OFFSET %s
        """
        params_select = params + [per_page, offset]
        cursor.execute(query_select, params_select)
        usuarios_lista = cursor.fetchall()

        empresas_filtro_lista = []
        if is_super:
            cursor.execute("SELECT id, nome_fantasia, razao_social FROM empresas ORDER BY nome_fantasia")
            empresas_filtro_lista = cursor.fetchall()

        filtros_atuais = {
            'pesquisa_usuario': pesquisa,
            'empresa_id': empresa_filtro,
            'perfil_de_acesso': perfil_filtro
        }

        return render_template(
            'visualizar_usuarios.html',
            usuarios=usuarios_lista,
            empresas=empresas_filtro_lista,
            is_super_admin=is_super,
            filtros=filtros_atuais,
            page=page,
            total_pages=total_pages,
            total_records=total_records
        )

    except Exception as e:
        print(f"Erro ao carregar a listagem de usuários: {e}")
        flash(f"Erro interno ao processar a consulta de usuários: {e}", "danger")
        return redirect(url_for('dashboard'))
    finally:
        if conexao and conexao.is_connected():
            cursor.close()
            conexao.close()


@app.route('/usuarios/alternar_status/<int:id>', methods=['POST'])
@login_required
@perfis_permitidos('Administrador')
def alternar_status_usuario(id):
    usuario_logado_id = session.get('usuario_id')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    if int(id) == int(usuario_logado_id):
        flash("Você não pode bloquear ou reativar o próprio usuário logado.", "warning")
        return redirect(url_for('visualizar_usuarios'))

    con = obter_conexao()

    if con is None:
        flash("Erro ao conectar ao banco de dados.", "danger")
        return redirect(url_for('visualizar_usuarios'))

    cur = con.cursor(dictionary=True)

    try:
        query = """
                SELECT id,
                       login,
                       status_usuario,
                       empresa_id,
                       COALESCE(is_super_admin, 0) AS is_super_admin
                FROM usuarios
                WHERE id = %s \
                """

        params = [id]

        if not is_super_admin:
            query += " AND empresa_id = %s"
            params.append(empresa_logada_id)

        query += " LIMIT 1"

        cur.execute(query, params)
        usuario = cur.fetchone()

        if not usuario:
            flash("Usuário não localizado ou não pertence à empresa logada.", "warning")
            return redirect(url_for('visualizar_usuarios'))

        if not is_super_admin and int(usuario.get('is_super_admin') or 0) == 1:
            flash("Usuário comum não pode alterar status de Super Admin.", "danger")
            return redirect(url_for('visualizar_usuarios'))

        novo_status = 'Inativo' if usuario['status_usuario'] == 'Ativo' else 'Ativo'

        cur.execute("""
                    UPDATE usuarios
                    SET status_usuario = %s
                    WHERE id = %s
                    """, (novo_status, id))

        con.commit()

        flash(f"Status do usuário {usuario['login']} alterado para {novo_status}.", "success")

    except Exception as e:
        con.rollback()
        print(f"Erro ao alternar status do usuário: {e}")
        flash("Erro ao alterar status do usuário.", "danger")

    finally:
        fechar_cursor_conexao(cur, con)

    return redirect(url_for('visualizar_usuarios'))



# ==========================================================
# BLOCO 12.2 - EDITOR DE PERFIS E PERMISSÕES POR EMPRESA
# ==========================================================
def normalizar_codigo_perfil(nome):
    base = (nome or '').strip().lower()
    base = re.sub(r'[^a-z0-9]+', '_', base)
    base = re.sub(r'_+', '_', base).strip('_')
    return base[:70] or 'perfil'



# ==========================================================
# BLOCO 12.3 - ISOLAMENTO DE PERFIS POR EMPRESA
# ==========================================================
PERFIS_PADRAO_EMPRESA = [
    ('Administrador', 'Acesso administrativo da própria empresa.'),
    ('Financeiro', 'Acesso aos módulos financeiros da própria empresa.'),
    ('Operacional', 'Acesso aos módulos operacionais da própria empresa.'),
    ('Supervisor', 'Acesso de supervisão operacional da própria empresa.'),
    ('Motorista', 'Acesso ao portal do motorista/prestador.'),
    ('Terminal Base', 'Acesso operacional de terminal/base.'),
    ('Somente Consulta', 'Acesso somente para consulta.'),
]


def codigo_perfil_empresa(nome, empresa_id):
    return f"{normalizar_codigo_perfil(nome)}_{int(empresa_id or 0)}"[:80]


def perfil_codigo_permissao_atual():
    """Código usado nas tabelas de permissão.

    A partir do Bloco 12.3, usuários de empresa usam o código do perfil
    clonado para a própria empresa. O perfil global fica apenas como modelo.
    """
    return session.get('perfil_permissao_codigo') or session.get('perfil_de_acesso')


def sincronizar_perfis_padrao_empresas(cur, empresa_id_especifica=None):
    """Replica perfis globais/padrão para cada empresa e vincula usuários."""
    try:
        if not tabela_existe(cur, 'perfis_acesso') or not tabela_existe(cur, 'empresas'):
            return
        for nome, descricao in PERFIS_PADRAO_EMPRESA:
            codigo_global = nome
            cur.execute("""
                SELECT id FROM perfis_acesso
                WHERE empresa_id IS NULL AND (codigo=%s OR nome=%s)
                LIMIT 1
            """, (codigo_global, nome))
            if not cur.fetchone():
                cur.execute("""
                    INSERT INTO perfis_acesso (empresa_id, codigo, nome, descricao, perfil_sistema, ativo, created_at, updated_at)
                    VALUES (NULL, %s, %s, %s, 1, 1, NOW(), NOW())
                """, (codigo_global, nome, descricao))

        if empresa_id_especifica:
            cur.execute("SELECT id, nome_fantasia, razao_social FROM empresas WHERE id=%s", (empresa_id_especifica,))
        else:
            cur.execute("SELECT id, nome_fantasia, razao_social FROM empresas WHERE COALESCE(status_empresa, 'Ativa')='Ativa'")
        empresas = cur.fetchall() or []

        cur.execute("""
            SELECT id, codigo, nome, descricao, ativo
            FROM perfis_acesso
            WHERE empresa_id IS NULL AND ativo = 1
            ORDER BY id
        """)
        globais = cur.fetchall() or []

        for emp in empresas:
            eid = int(emp.get('id'))
            for g in globais:
                nome = g.get('nome') or g.get('codigo')
                codigo_local = codigo_perfil_empresa(nome, eid)
                cur.execute("""
                    SELECT id, codigo
                    FROM perfis_acesso
                    WHERE empresa_id=%s
                      AND (codigo=%s OR nome COLLATE utf8mb4_unicode_ci = %s COLLATE utf8mb4_unicode_ci)
                    LIMIT 1
                """, (eid, codigo_local, nome))
                local = cur.fetchone()
                if local:
                    perfil_local_id = local.get('id')
                    codigo_final = local.get('codigo') or codigo_local
                else:
                    cur.execute("""
                        INSERT INTO perfis_acesso (empresa_id, codigo, nome, descricao, perfil_sistema, ativo, created_at, updated_at)
                        VALUES (%s, %s, %s, %s, 0, 1, NOW(), NOW())
                    """, (eid, codigo_local, nome, g.get('descricao')))
                    perfil_local_id = cur.lastrowid
                    codigo_final = codigo_local

                cur.execute("""
                    SELECT COUNT(*) AS total
                    FROM perfil_permissoes
                    WHERE perfil_de_acesso=%s AND empresa_id=%s
                """, (codigo_final, eid))
                total_local = int((cur.fetchone() or {}).get('total') or 0)
                if total_local == 0 and tabela_existe(cur, 'perfil_permissoes'):
                    cur.execute("""
                        INSERT IGNORE INTO perfil_permissoes (perfil_de_acesso, menu_codigo, acao_codigo, empresa_id, permitido)
                        SELECT %s, pp.menu_codigo, pp.acao_codigo, %s, pp.permitido
                        FROM perfil_permissoes pp
                        WHERE pp.empresa_id = 0
                          AND pp.permitido = 1
                          AND pp.perfil_de_acesso IN (%s, %s, %s)
                    """, (codigo_final, eid, g.get('codigo'), g.get('nome'), nome))

                if tabela_existe(cur, 'usuarios') and perfil_local_id:
                    nomes_equivalentes = [nome, g.get('codigo')]
                    if nome == 'Somente Consulta':
                        nomes_equivalentes.extend(['Consulta', 'Somente Consulta'])
                    nomes_equivalentes = list(dict.fromkeys([x for x in nomes_equivalentes if x]))
                    placeholders = ','.join(['%s'] * len(nomes_equivalentes))
                    cur.execute(f"""
                        UPDATE usuarios
                        SET perfil_id=%s
                        WHERE empresa_id=%s
                          AND (perfil_id IS NULL OR perfil_id NOT IN (
                                SELECT id FROM perfis_acesso pa2 WHERE pa2.empresa_id=%s
                              ))
                          AND perfil_de_acesso IN ({placeholders})
                    """, [perfil_local_id, eid, eid] + nomes_equivalentes)
    except Exception as e:
        print(f"[Permissões] Falha ao sincronizar perfis por empresa: {e}")


def resolver_perfil_permissao_usuario(cur, usuario):
    """Resolve o perfil efetivo do usuário no escopo da própria empresa."""
    empresa_id = usuario.get('empresa_id') or usuario.get('empresa_id_confirmado')
    perfil_id = usuario.get('perfil_id')
    perfil_antigo = usuario.get('perfil_de_acesso')
    if not empresa_id:
        return {'perfil_id': perfil_id, 'codigo': perfil_antigo, 'nome': perfil_antigo, 'empresa_id': None}
    try:
        if perfil_id:
            cur.execute("""
                SELECT id, codigo, nome, empresa_id
                FROM perfis_acesso
                WHERE id=%s AND ativo=1
                LIMIT 1
            """, (perfil_id,))
            row = cur.fetchone()
            if row and row.get('empresa_id') is not None and str(row.get('empresa_id')) == str(empresa_id):
                return row
        cur.execute("""
            SELECT id, codigo, nome, empresa_id
            FROM perfis_acesso
            WHERE empresa_id=%s
              AND ativo=1
              AND (nome COLLATE utf8mb4_unicode_ci = %s COLLATE utf8mb4_unicode_ci
                   OR codigo COLLATE utf8mb4_unicode_ci = %s COLLATE utf8mb4_unicode_ci)
            ORDER BY id DESC
            LIMIT 1
        """, (empresa_id, perfil_antigo, perfil_antigo))
        row = cur.fetchone()
        if row:
            try:
                cur.execute("UPDATE usuarios SET perfil_id=%s WHERE id=%s", (row.get('id'), usuario.get('usuario_id') or usuario.get('id')))
            except Exception:
                pass
            return row
    except Exception as e:
        print(f"[Permissões] Falha ao resolver perfil efetivo: {e}")
    return {'perfil_id': perfil_id, 'codigo': perfil_antigo, 'nome': perfil_antigo, 'empresa_id': None}



def buscar_perfil_local_por_nome_codigo(cur, perfil_de_acesso, empresa_id):
    if not perfil_de_acesso or not empresa_id or not tabela_existe(cur, 'perfis_acesso'):
        return None
    cur.execute("""
        SELECT id, codigo, nome, empresa_id
        FROM perfis_acesso
        WHERE empresa_id=%s
          AND ativo=1
          AND (nome COLLATE utf8mb4_unicode_ci = %s COLLATE utf8mb4_unicode_ci
               OR codigo COLLATE utf8mb4_unicode_ci = %s COLLATE utf8mb4_unicode_ci)
        ORDER BY id DESC
        LIMIT 1
    """, (empresa_id, perfil_de_acesso, perfil_de_acesso))
    row = cur.fetchone()
    return row


def listar_perfis_disponiveis_para_empresa(cur, empresa_id, escopo_global=False):
    """Lista perfis para os formulários de usuário.

    Administrador de empresa vê apenas perfis da própria empresa. Super/Suporte
    vê perfis globais e por empresa para manutenção técnica.
    """
    try:
        if escopo_global:
            cur.execute("""
                SELECT pa.id, pa.empresa_id, pa.codigo, pa.nome,
                       COALESCE(e.nome_fantasia, 'Global/Sistema') AS empresa_nome
                FROM perfis_acesso pa
                LEFT JOIN empresas e ON e.id = pa.empresa_id
                WHERE pa.ativo=1
                ORDER BY pa.empresa_id IS NULL DESC, empresa_nome, pa.nome
            """)
        else:
            cur.execute("""
                SELECT id, empresa_id, codigo, nome, NULL AS empresa_nome
                FROM perfis_acesso
                WHERE empresa_id=%s AND ativo=1
                ORDER BY nome
            """, (empresa_id,))
        rows = cur.fetchall() or []
        if rows:
            return rows
    except Exception as e:
        print(f"[Permissões] Falha ao listar perfis disponíveis: {e}")
    return [
        {'id': None, 'empresa_id': empresa_id, 'codigo': 'Administrador', 'nome': 'Administrador'},
        {'id': None, 'empresa_id': empresa_id, 'codigo': 'Operacional', 'nome': 'Operacional'},
        {'id': None, 'empresa_id': empresa_id, 'codigo': 'Financeiro', 'nome': 'Financeiro'},
        {'id': None, 'empresa_id': empresa_id, 'codigo': 'Motorista', 'nome': 'Motorista'},
        {'id': None, 'empresa_id': empresa_id, 'codigo': 'Terminal Base', 'nome': 'Terminal Base'},
    ]

def escopo_empresa_perfil():
    """Retorna None para visão global e empresa_id para administrador comum."""
    return None if usuario_pode_ver_escopo_global_sistema() else session.get('empresa_id')


def buscar_perfil_acesso_por_id(cur, perfil_id):
    cur.execute("""
        SELECT id, empresa_id, codigo, nome, descricao, perfil_sistema, ativo
        FROM perfis_acesso
        WHERE id = %s
        LIMIT 1
    """, (perfil_id,))
    return cur.fetchone()


def usuario_pode_gerenciar_perfil(perfil):
    if not perfil:
        return False
    if usuario_pode_ver_escopo_global_sistema():
        return True
    empresa_id = session.get('empresa_id')
    return perfil.get('empresa_id') is not None and str(perfil.get('empresa_id')) == str(empresa_id)


def registrar_auditoria_permissao_segura(tipo_acao, entidade='perfil_permissoes', entidade_id=None, depois_json=None, empresa_id=None, perfil_afetado_id=None):
    """Auditoria tolerante às variações de schema criadas nas fases anteriores."""
    con = obter_conexao()
    if con is None:
        return
    cur = con.cursor(dictionary=True)
    try:
        if not tabela_existe(cur, 'auditoria_permissoes'):
            return
        campos = []
        valores = []
        def add(campo, valor):
            if coluna_existe(cur, 'auditoria_permissoes', campo):
                campos.append(campo)
                valores.append(valor)
        add('empresa_id', empresa_id)
        add('usuario_executor_id', session.get('usuario_id'))
        add('perfil_afetado_id', perfil_afetado_id)
        add('tipo_acao', tipo_acao)
        add('entidade', entidade)
        add('entidade_id', entidade_id)
        add('depois_json', depois_json)
        add('ip_origem', request.remote_addr if request else None)
        add('user_agent', request.headers.get('User-Agent') if request else None)
        add('created_at', datetime.now())
        if campos:
            ph = ', '.join(['%s'] * len(campos))
            cur.execute(f"INSERT INTO auditoria_permissoes ({', '.join(campos)}) VALUES ({ph})", valores)
            con.commit()
    except Exception as e:
        con.rollback()
        print(f"[Permissões] Auditoria segura falhou: {e}")
    finally:
        fechar_cursor_conexao(cur, con)


def carregar_menus_para_matriz(cur, escopo_global=False, perfil_sessao=None, usuario_id=None, empresa_id=None):
    ordem_sql = """
        ORDER BY FIELD(grupo_menu, 'VISÃO GERAL','CADASTROS','OPERAÇÃO','FINANCEIRO','RELATÓRIOS','AUDITORIA','SISTEMA'), grupo_menu, ordem, titulo
    """
    if escopo_global:
        cur.execute(f"""
            SELECT codigo, grupo_menu, titulo, endpoint, icone, ordem, ativo, visivel_menu
            FROM sistema_menus
            WHERE ativo = 1
            {ordem_sql}
        """)
        return cur.fetchall()
    cur.execute(f"""
        SELECT DISTINCT sm.codigo, sm.grupo_menu, sm.titulo, sm.endpoint, sm.icone, sm.ordem, sm.ativo, sm.visivel_menu
        FROM sistema_menus sm
        LEFT JOIN perfil_permissoes pp
               ON pp.menu_codigo = sm.codigo
              AND pp.perfil_de_acesso = %s
              AND pp.acao_codigo = 'visualizar'
              AND pp.permitido = 1
              AND pp.empresa_id = COALESCE(%s, 0)
        LEFT JOIN usuario_permissoes up
               ON up.menu_codigo = sm.codigo
              AND up.usuario_id = %s
              AND up.acao_codigo = 'visualizar'
              AND up.permitido = 1
              AND up.empresa_id = COALESCE(%s, 0)
        WHERE sm.ativo = 1
          AND sm.visivel_menu = 1
          AND sm.codigo NOT IN ('visualizar_empresas')
          AND sm.endpoint NOT IN ('visualizar_empresas', 'cadastro_empresa', 'editar_empresa')
          AND (pp.id IS NOT NULL OR up.id IS NOT NULL)
        {ordem_sql}
    """, (perfil_sessao, empresa_id, usuario_id, empresa_id))
    return cur.fetchall()

@app.route('/configuracoes/perfil-acesso')
@login_required
@perfis_permitidos('Administrador')
def perfil_acesso():
    # v20: não roda sincronização pesada em GET. Evita timeout no Render.
    inicializar_nucleo_permissoes()
    con = obter_conexao()
    if con is None:
        flash('Erro ao conectar ao banco de dados.', 'danger')
        return redirect(url_for('dashboard'))

    escopo_global = usuario_pode_ver_escopo_global_sistema()
    empresa_sessao = session.get('empresa_id')
    usuario_sessao = session.get('usuario_id')
    perfil_sessao = perfil_codigo_permissao_atual()

    cur = con.cursor(dictionary=True)
    try:
        ordem_sql = """
            ORDER BY FIELD(grupo_menu, 'VISÃO GERAL','CADASTROS','OPERAÇÃO','FINANCEIRO','RELATÓRIOS','AUDITORIA','SISTEMA'), grupo_menu, ordem, titulo
        """

        if escopo_global:
            cur.execute(f"""
                SELECT grupo_menu, codigo, titulo, endpoint, icone, ativo, visivel_menu, ordem
                FROM sistema_menus
                {ordem_sql}
            """)
            menus = cur.fetchall()

            cur.execute("""
                SELECT perfil_de_acesso, COUNT(*) AS total
                FROM perfil_permissoes
                WHERE permitido=1
                GROUP BY perfil_de_acesso
                ORDER BY perfil_de_acesso
            """)
            perfis = cur.fetchall()

            cur.execute("""
                SELECT u.id, u.login, u.perfil_de_acesso, e.nome_fantasia AS empresa_nome,
                       COUNT(ue.empresa_id) AS empresas_liberadas
                FROM usuarios u
                LEFT JOIN empresas e ON e.id = u.empresa_id
                LEFT JOIN usuario_empresas_acesso ue ON ue.usuario_id = u.id AND ue.ativo=1
                GROUP BY u.id, u.login, u.perfil_de_acesso, e.nome_fantasia
                ORDER BY u.id DESC
                LIMIT 100
            """)
            usuarios = cur.fetchall()

            cur.execute("SELECT COUNT(*) AS total FROM auditoria_permissoes")
            auditoria_total = (cur.fetchone() or {}).get('total') or 0
            escopo_titulo = 'Visão global do sistema'
            escopo_descricao = 'Super Admin/Suporte: visualiza usuários, empresas, menus e auditorias globais.'
        else:
            # Administrador comum de empresa só enxerga a própria empresa.
            cur.execute(f"""
                SELECT DISTINCT sm.grupo_menu, sm.codigo, sm.titulo, sm.endpoint, sm.icone, sm.ativo, sm.visivel_menu, sm.ordem
                FROM sistema_menus sm
                LEFT JOIN perfil_permissoes pp
                       ON pp.menu_codigo = sm.codigo
                      AND pp.perfil_de_acesso = %s
                      AND pp.acao_codigo = 'visualizar'
                      AND pp.permitido = 1
                      AND pp.empresa_id = COALESCE(%s, 0)
                LEFT JOIN usuario_permissoes up
                       ON up.menu_codigo = sm.codigo
                      AND up.usuario_id = %s
                      AND up.acao_codigo = 'visualizar'
                      AND up.permitido = 1
                      AND up.empresa_id = COALESCE(%s, 0)
                WHERE sm.ativo = 1
                  AND sm.visivel_menu = 1
                  AND sm.codigo NOT IN ('visualizar_empresas')
                  AND sm.endpoint NOT IN ('visualizar_empresas', 'cadastro_empresa', 'editar_empresa')
                  AND (pp.id IS NOT NULL OR up.id IS NOT NULL)
                {ordem_sql}
            """, (perfil_sessao, empresa_sessao, usuario_sessao, empresa_sessao))
            menus = cur.fetchall()

            cur.execute("""
                SELECT pa.nome AS perfil_de_acesso, COUNT(pp.id) AS total
                FROM perfis_acesso pa
                LEFT JOIN perfil_permissoes pp
                       ON pp.perfil_de_acesso = pa.codigo
                      AND pp.empresa_id = pa.empresa_id
                      AND pp.permitido = 1
                WHERE pa.empresa_id = %s
                  AND pa.ativo = 1
                GROUP BY pa.id, pa.nome
                ORDER BY pa.nome
            """, (empresa_sessao,))
            perfis = cur.fetchall()

            cur.execute("""
                SELECT u.id, u.login, u.perfil_de_acesso, e.nome_fantasia AS empresa_nome,
                       COUNT(ue.empresa_id) AS empresas_liberadas
                FROM usuarios u
                LEFT JOIN empresas e ON e.id = u.empresa_id
                LEFT JOIN usuario_empresas_acesso ue ON ue.usuario_id = u.id AND ue.ativo=1
                WHERE u.empresa_id = %s
                GROUP BY u.id, u.login, u.perfil_de_acesso, e.nome_fantasia
                ORDER BY u.id DESC
                LIMIT 100
            """, (empresa_sessao,))
            usuarios = cur.fetchall()

            cur.execute("""
                SELECT COUNT(*) AS total
                FROM auditoria_permissoes
                WHERE empresa_id = %s
            """, (empresa_sessao,))
            auditoria_total = (cur.fetchone() or {}).get('total') or 0
            escopo_titulo = 'Visão restrita à empresa atual'
            escopo_descricao = 'Administrador de empresa: visualiza somente usuários e acessos da própria empresa.'

        if escopo_global:
            cur.execute("""
                SELECT pa.id, pa.empresa_id, pa.codigo, pa.nome, pa.descricao, pa.perfil_sistema, pa.ativo,
                       COALESCE(e.nome_fantasia, 'Global/Sistema') AS empresa_nome,
                       COUNT(pp.id) AS total_permissoes
                FROM perfis_acesso pa
                LEFT JOIN empresas e ON e.id = pa.empresa_id
                LEFT JOIN perfil_permissoes pp
                       ON pp.perfil_de_acesso = pa.codigo
                      AND pp.empresa_id = COALESCE(pa.empresa_id, 0)
                      AND pp.permitido = 1
                GROUP BY pa.id, pa.empresa_id, pa.codigo, pa.nome, pa.descricao, pa.perfil_sistema, pa.ativo, e.nome_fantasia
                ORDER BY COALESCE(e.nome_fantasia, 'Global/Sistema'), pa.nome
            """)
        else:
            cur.execute("""
                SELECT pa.id, pa.empresa_id, pa.codigo, pa.nome, pa.descricao, pa.perfil_sistema, pa.ativo,
                       COALESCE(e.nome_fantasia, 'Global/Sistema') AS empresa_nome,
                       COUNT(pp.id) AS total_permissoes
                FROM perfis_acesso pa
                LEFT JOIN empresas e ON e.id = pa.empresa_id
                LEFT JOIN perfil_permissoes pp
                       ON pp.perfil_de_acesso = pa.codigo
                      AND pp.empresa_id = COALESCE(pa.empresa_id, 0)
                      AND pp.permitido = 1
                WHERE pa.empresa_id = %s
                GROUP BY pa.id, pa.empresa_id, pa.codigo, pa.nome, pa.descricao, pa.perfil_sistema, pa.ativo, e.nome_fantasia
                ORDER BY pa.nome
            """, (empresa_sessao,))
        perfis_acesso_lista = cur.fetchall()

        return render_template(
            'perfil_acesso.html',
            menus=menus,
            perfis=perfis,
            perfis_acesso_lista=perfis_acesso_lista,
            usuarios=usuarios,
            auditoria_total=auditoria_total,
            escopo_global=escopo_global,
            escopo_titulo=escopo_titulo,
            escopo_descricao=escopo_descricao,
            pode_sincronizar_menus=escopo_global,
        )
    except Exception as e:
        print(f"Erro ao carregar núcleo de permissões: {e}")
        flash(f'Erro ao carregar núcleo de permissões: {e}', 'danger')
        return redirect(url_for('dashboard'))
    finally:
        fechar_cursor_conexao(cur, con)


@app.post('/configuracoes/perfil-acesso/sincronizar')
@login_required
@perfis_permitidos('Administrador')
def sincronizar_permissoes_menus():
    if not usuario_pode_ver_escopo_global_sistema():
        flash('Somente Super Admin ou usuário de suporte pode sincronizar menus globais.', 'danger')
        return redirect(url_for('perfil_acesso'))

    ok = inicializar_nucleo_permissoes()
    if ok:
        registrar_auditoria_permissao('SINCRONIZAR_MENUS', 'Verificação segura do núcleo de menus e permissões.')
        invalidar_cache_permissoes()
        flash('Menus e permissões sincronizados com sucesso.', 'success')
    else:
        flash('Não foi possível sincronizar menus e permissões.', 'danger')
    return redirect(url_for('perfil_acesso'))



@app.route('/configuracoes/perfis/novo', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador')
def novo_perfil_acesso_empresa():
    inicializar_nucleo_permissoes()
    con = obter_conexao()
    if con is None:
        flash('Erro ao conectar ao banco de dados.', 'danger')
        return redirect(url_for('perfil_acesso'))
    cur = con.cursor(dictionary=True)
    escopo_global = usuario_pode_ver_escopo_global_sistema()
    empresa_sessao = session.get('empresa_id')
    try:
        empresas = []
        if escopo_global:
            cur.execute("SELECT id, nome_fantasia, razao_social FROM empresas WHERE status_empresa='Ativa' ORDER BY nome_fantasia, razao_social")
            empresas = cur.fetchall()
        if request.method == 'POST':
            nome = (request.form.get('nome') or '').strip()
            descricao = (request.form.get('descricao') or '').strip() or None
            ativo = 1 if request.form.get('ativo', '1') == '1' else 0
            if not nome:
                flash('Informe o nome do perfil.', 'warning')
                return redirect(url_for('novo_perfil_acesso_empresa'))
            if escopo_global:
                empresa_id = request.form.get('empresa_id') or None
                empresa_id = int(empresa_id) if empresa_id else None
            else:
                empresa_id = int(empresa_sessao)
            codigo_base = normalizar_codigo_perfil(nome)
            codigo = codigo_base if empresa_id is None else f"{codigo_base}_{empresa_id}"
            codigo = codigo[:80]
            cur.execute("""
                SELECT id FROM perfis_acesso
                WHERE ((empresa_id IS NULL AND %s IS NULL) OR empresa_id = %s)
                  AND codigo = %s
                LIMIT 1
            """, (empresa_id, empresa_id, codigo))
            if cur.fetchone():
                flash('Já existe um perfil com esse código/nome neste escopo.', 'warning')
                return redirect(url_for('novo_perfil_acesso_empresa'))
            cur.execute("""
                INSERT INTO perfis_acesso (empresa_id, codigo, nome, descricao, perfil_sistema, ativo, created_at, updated_at)
                VALUES (%s, %s, %s, %s, 0, %s, NOW(), NOW())
            """, (empresa_id, codigo, nome, descricao, ativo))
            perfil_id = cur.lastrowid
            con.commit()
            registrar_auditoria_permissao_segura('CRIAR_PERFIL_ACESSO', 'perfis_acesso', perfil_id, json.dumps({'codigo': codigo, 'nome': nome}, ensure_ascii=False), empresa_id, perfil_id)
            invalidar_cache_permissoes()
            flash('Perfil criado com sucesso. Agora configure as permissões dele.', 'success')
            return redirect(url_for('editar_permissoes_perfil_acesso', perfil_id=perfil_id))
        return render_template('perfil_acesso_form.html', empresas=empresas, escopo_global=escopo_global)
    except Exception as e:
        con.rollback()
        print(f"Erro ao criar perfil de acesso: {e}")
        flash(f'Erro ao criar perfil de acesso: {e}', 'danger')
        return redirect(url_for('perfil_acesso'))
    finally:
        fechar_cursor_conexao(cur, con)


@app.route('/configuracoes/perfis/<int:perfil_id>/permissoes', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador')
def editar_permissoes_perfil_acesso(perfil_id):
    inicializar_nucleo_permissoes()
    con = obter_conexao()
    if con is None:
        flash('Erro ao conectar ao banco de dados.', 'danger')
        return redirect(url_for('perfil_acesso'))
    cur = con.cursor(dictionary=True)
    escopo_global = usuario_pode_ver_escopo_global_sistema()
    empresa_sessao = session.get('empresa_id')
    try:
        perfil = buscar_perfil_acesso_por_id(cur, perfil_id)
        if not usuario_pode_gerenciar_perfil(perfil):
            flash('Você não tem permissão para gerenciar este perfil.', 'danger')
            return redirect(url_for('perfil_acesso'))
        empresa_permissao = int(perfil.get('empresa_id') or 0)
        menus = carregar_menus_para_matriz(
            cur,
            escopo_global=escopo_global,
            perfil_sessao=perfil_codigo_permissao_atual(),
            usuario_id=session.get('usuario_id'),
            empresa_id=empresa_sessao,
        )
        acoes = ACOES_SISTEMA_PADRAO
        if request.method == 'POST':
            if not escopo_global and perfil.get('empresa_id') is None:
                flash('Perfil do sistema não pode ser alterado por administrador de empresa. Crie um perfil próprio da empresa.', 'danger')
                return redirect(url_for('perfil_acesso'))
            permitidos = set()
            codigos_menus = [m['codigo'] for m in menus]
            for menu_codigo in codigos_menus:
                for acao in acoes:
                    if request.form.get(f"perm__{menu_codigo}__{acao}") == '1':
                        permitidos.add((menu_codigo, acao))
            if not permitidos:
                flash('Selecione ao menos uma permissão para este perfil.', 'warning')
                return redirect(url_for('editar_permissoes_perfil_acesso', perfil_id=perfil_id))
            cur.execute("DELETE FROM perfil_permissoes WHERE perfil_de_acesso=%s AND empresa_id=%s", (perfil['codigo'], empresa_permissao))
            for menu_codigo, acao in sorted(permitidos):
                cur.execute("""
                    INSERT INTO perfil_permissoes (perfil_de_acesso, menu_codigo, acao_codigo, empresa_id, permitido)
                    VALUES (%s, %s, %s, %s, 1)
                    ON DUPLICATE KEY UPDATE permitido=VALUES(permitido), atualizado_em=CURRENT_TIMESTAMP
                """, (perfil['codigo'], menu_codigo, acao, empresa_permissao))
            cur.execute("UPDATE perfis_acesso SET updated_at=NOW() WHERE id=%s", (perfil_id,))
            con.commit()
            registrar_auditoria_permissao_segura('EDITAR_PERMISSOES_PERFIL', 'perfil_permissoes', perfil_id, json.dumps({'perfil': perfil['codigo'], 'total': len(permitidos)}, ensure_ascii=False), perfil.get('empresa_id'), perfil_id)
            invalidar_cache_permissoes()
            flash('Permissões do perfil atualizadas com sucesso.', 'success')
            return redirect(url_for('perfil_acesso'))
        cur.execute("""
            SELECT menu_codigo, acao_codigo
            FROM perfil_permissoes
            WHERE perfil_de_acesso=%s AND empresa_id=%s AND permitido=1
        """, (perfil['codigo'], empresa_permissao))
        permissoes = {f"{r['menu_codigo']}__{r['acao_codigo']}" for r in cur.fetchall()}
        return render_template('perfil_permissoes_editar.html', perfil=perfil, menus=menus, acoes=acoes, permissoes=permissoes, escopo_global=escopo_global)
    except Exception as e:
        con.rollback()
        print(f"Erro ao editar permissões do perfil: {e}")
        flash(f'Erro ao editar permissões do perfil: {e}', 'danger')
        return redirect(url_for('perfil_acesso'))
    finally:
        fechar_cursor_conexao(cur, con)


@app.post('/configuracoes/perfis/<int:perfil_id>/alternar')
@login_required
@perfis_permitidos('Administrador')
def alternar_perfil_acesso_empresa(perfil_id):
    con = obter_conexao()
    if con is None:
        flash('Erro ao conectar ao banco de dados.', 'danger')
        return redirect(url_for('perfil_acesso'))
    cur = con.cursor(dictionary=True)
    try:
        perfil = buscar_perfil_acesso_por_id(cur, perfil_id)
        if not usuario_pode_gerenciar_perfil(perfil):
            flash('Você não tem permissão para alterar este perfil.', 'danger')
            return redirect(url_for('perfil_acesso'))
        if not usuario_pode_ver_escopo_global_sistema() and int(perfil.get('perfil_sistema') or 0) == 1:
            flash('Perfil padrão do sistema não pode ser inativado por administrador de empresa.', 'danger')
            return redirect(url_for('perfil_acesso'))
        novo = 0 if int(perfil.get('ativo') or 0) == 1 else 1
        cur.execute("UPDATE perfis_acesso SET ativo=%s, updated_at=NOW() WHERE id=%s", (novo, perfil_id))
        con.commit()
        registrar_auditoria_permissao_segura('ALTERAR_STATUS_PERFIL', 'perfis_acesso', perfil_id, json.dumps({'ativo': novo}, ensure_ascii=False), perfil.get('empresa_id'), perfil_id)
        invalidar_cache_permissoes()
        flash('Status do perfil alterado com sucesso.', 'success')
    except Exception as e:
        con.rollback()
        print(f"Erro ao alternar perfil: {e}")
        flash(f'Erro ao alterar perfil: {e}', 'danger')
    finally:
        fechar_cursor_conexao(cur, con)
    return redirect(url_for('perfil_acesso'))



# ==========================================================
# BLOCO 12.4 - GERENCIADOR GLOBAL DE MÓDULOS, MENUS E ROTAS
# ==========================================================
def usuario_pode_gerenciar_menus_globais():
    """Somente Super Admin ou Suporte do sistema podem alterar estrutura global."""
    return usuario_pode_ver_escopo_global_sistema()


def garantir_schema_menus_modulos(cur):
    """Completa o schema de módulos/menus sem quebrar bancos já migrados."""
    inicializar_nucleo_permissoes()
    # sistema_modulos
    if not coluna_existe(cur, 'sistema_modulos', 'descricao'):
        executar_ddl_permissoes(cur, "ALTER TABLE sistema_modulos ADD COLUMN descricao VARCHAR(255) NULL AFTER nome")
    if not coluna_existe(cur, 'sistema_modulos', 'visivel_menu'):
        executar_ddl_permissoes(cur, "ALTER TABLE sistema_modulos ADD COLUMN visivel_menu TINYINT(1) NOT NULL DEFAULT 1 AFTER ativo")
    if not coluna_existe(cur, 'sistema_modulos', 'somente_super_admin'):
        executar_ddl_permissoes(cur, "ALTER TABLE sistema_modulos ADD COLUMN somente_super_admin TINYINT(1) NOT NULL DEFAULT 0 AFTER visivel_menu")
    if not coluna_existe(cur, 'sistema_modulos', 'somente_suporte'):
        executar_ddl_permissoes(cur, "ALTER TABLE sistema_modulos ADD COLUMN somente_suporte TINYINT(1) NOT NULL DEFAULT 0 AFTER somente_super_admin")

    # sistema_menus
    if not coluna_existe(cur, 'sistema_menus', 'descricao'):
        executar_ddl_permissoes(cur, "ALTER TABLE sistema_menus ADD COLUMN descricao VARCHAR(255) NULL AFTER titulo")
    if not coluna_existe(cur, 'sistema_menus', 'exige_empresa'):
        executar_ddl_permissoes(cur, "ALTER TABLE sistema_menus ADD COLUMN exige_empresa TINYINT(1) NOT NULL DEFAULT 1 AFTER visivel_menu")
    if not coluna_existe(cur, 'sistema_menus', 'somente_super_admin'):
        executar_ddl_permissoes(cur, "ALTER TABLE sistema_menus ADD COLUMN somente_super_admin TINYINT(1) NOT NULL DEFAULT 0 AFTER exige_empresa")
    if not coluna_existe(cur, 'sistema_menus', 'somente_suporte'):
        executar_ddl_permissoes(cur, "ALTER TABLE sistema_menus ADD COLUMN somente_suporte TINYINT(1) NOT NULL DEFAULT 0 AFTER somente_super_admin")
    if not coluna_existe(cur, 'sistema_menus', 'liberar_admin_empresas'):
        executar_ddl_permissoes(cur, "ALTER TABLE sistema_menus ADD COLUMN liberar_admin_empresas TINYINT(1) NOT NULL DEFAULT 0 AFTER somente_suporte")

    # Marca menus globais críticos como restritos.
    try:
        cur.execute("""
            UPDATE sistema_menus
               SET somente_super_admin=1
             WHERE codigo IN ('visualizar_empresas', 'gerenciar_menus_modulos')
                OR endpoint IN ('visualizar_empresas', 'cadastro_empresa', 'editar_empresa',
                                'gerenciar_menus_modulos', 'novo_modulo_sistema', 'editar_modulo_sistema',
                                'alternar_modulo_sistema', 'novo_menu_sistema', 'editar_menu_sistema', 'alternar_menu_sistema')
        """)
    except Exception as e:
        print(f"[Menus] Falha ao marcar restrições globais: {e}")


def endpoint_existe_no_app(endpoint):
    if not endpoint:
        return False
    return endpoint in app.view_functions


def listar_modulos_sistema(cur):
    garantir_schema_menus_modulos(cur)
    cur.execute("""
        SELECT m.*,
               (SELECT COUNT(*) FROM sistema_menus sm WHERE sm.modulo_id = m.id) AS total_menus
          FROM sistema_modulos m
         ORDER BY m.ordem, m.nome
    """)
    return cur.fetchall()


def listar_menus_sistema(cur):
    garantir_schema_menus_modulos(cur)
    cur.execute("""
        SELECT sm.*, m.nome AS modulo_nome, pai.titulo AS menu_pai_titulo
          FROM sistema_menus sm
          LEFT JOIN sistema_modulos m ON m.id = sm.modulo_id
          LEFT JOIN sistema_menus pai ON pai.id = sm.menu_pai_id
         ORDER BY FIELD(sm.grupo_menu, 'VISÃO GERAL','CADASTROS','OPERAÇÃO','FINANCEIRO','RELATÓRIOS','AUDITORIA','SISTEMA'),
                  sm.grupo_menu, m.ordem, sm.menu_pai_id IS NOT NULL, sm.ordem, sm.titulo
    """)
    rows = cur.fetchall()
    for r in rows:
        r['endpoint_existe'] = endpoint_existe_no_app(r.get('endpoint'))
    return rows


def buscar_modulo_sistema(cur, modulo_id):
    garantir_schema_menus_modulos(cur)
    cur.execute("SELECT * FROM sistema_modulos WHERE id=%s LIMIT 1", (modulo_id,))
    return cur.fetchone()


def buscar_menu_sistema(cur, menu_id):
    garantir_schema_menus_modulos(cur)
    cur.execute("SELECT * FROM sistema_menus WHERE id=%s LIMIT 1", (menu_id,))
    row = cur.fetchone()
    if row:
        row['endpoint_existe'] = endpoint_existe_no_app(row.get('endpoint'))
    return row


def normalizar_codigo_menu_modulo(texto):
    import re
    texto = (texto or '').strip().lower()
    mapa = {'á':'a','à':'a','ã':'a','â':'a','é':'e','ê':'e','í':'i','ó':'o','ô':'o','õ':'o','ú':'u','ü':'u','ç':'c'}
    for a,b in mapa.items():
        texto = texto.replace(a,b)
    texto = re.sub(r'[^a-z0-9]+', '_', texto)
    texto = re.sub(r'_+', '_', texto).strip('_')
    return texto[:120] or 'novo_item'


@app.route('/configuracoes/menus-modulos')
@login_required
@perfis_permitidos('Administrador')
def gerenciar_menus_modulos():
    if not usuario_pode_gerenciar_menus_globais():
        flash('Somente Super Admin ou Suporte do sistema pode gerenciar módulos, menus e rotas globais.', 'danger')
        return redirect(url_for('perfil_acesso'))
    con = obter_conexao()
    if con is None:
        flash('Erro ao conectar ao banco de dados.', 'danger')
        return redirect(url_for('perfil_acesso'))
    cur = con.cursor(dictionary=True)
    try:
        modulos = listar_modulos_sistema(cur)
        menus = listar_menus_sistema(cur)
        con.commit()
        return render_template('gerenciar_menus_modulos.html', modulos=modulos, menus=menus)
    except Exception as e:
        con.rollback()
        print(f"Erro ao carregar gerenciador de menus: {e}")
        flash(f'Erro ao carregar gerenciador de menus: {e}', 'danger')
        return redirect(url_for('perfil_acesso'))
    finally:
        fechar_cursor_conexao(cur, con)


@app.route('/configuracoes/modulos/novo', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador')
def novo_modulo_sistema():
    if not usuario_pode_gerenciar_menus_globais():
        flash('Somente Super Admin ou Suporte do sistema pode criar módulos globais.', 'danger')
        return redirect(url_for('perfil_acesso'))
    con = obter_conexao(); cur = con.cursor(dictionary=True)
    try:
        garantir_schema_menus_modulos(cur)
        if request.method == 'POST':
            nome = (request.form.get('nome') or '').strip()
            codigo = (request.form.get('codigo') or normalizar_codigo_menu_modulo(nome)).strip()
            descricao = (request.form.get('descricao') or '').strip() or None
            icone = (request.form.get('icone') or 'fa-solid fa-folder').strip()
            ordem = int(request.form.get('ordem') or 0)
            ativo = 1 if request.form.get('ativo') == '1' else 0
            visivel_menu = 1 if request.form.get('visivel_menu') == '1' else 0
            somente_super_admin = 1 if request.form.get('somente_super_admin') == '1' else 0
            somente_suporte = 1 if request.form.get('somente_suporte') == '1' else 0
            if not nome or not codigo:
                flash('Informe nome e código do módulo.', 'warning')
                return redirect(url_for('novo_modulo_sistema'))
            cur.execute("""
                INSERT INTO sistema_modulos (codigo, nome, descricao, icone, ordem, ativo, visivel_menu, somente_super_admin, somente_suporte)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE nome=VALUES(nome), descricao=VALUES(descricao), icone=VALUES(icone), ordem=VALUES(ordem),
                                        ativo=VALUES(ativo), visivel_menu=VALUES(visivel_menu), somente_super_admin=VALUES(somente_super_admin), somente_suporte=VALUES(somente_suporte)
            """, (codigo, nome, descricao, icone, ordem, ativo, visivel_menu, somente_super_admin, somente_suporte))
            con.commit(); invalidar_cache_permissoes()
            registrar_auditoria_permissao_segura('CRIAR_MODULO_SISTEMA', 'sistema_modulos', None, json.dumps({'codigo':codigo,'nome':nome}, ensure_ascii=False), None, None)
            flash('Módulo salvo com sucesso.', 'success')
            return redirect(url_for('gerenciar_menus_modulos'))
        return render_template('modulo_sistema_form.html', modulo=None)
    except Exception as e:
        con.rollback(); flash(f'Erro ao salvar módulo: {e}', 'danger'); return redirect(url_for('gerenciar_menus_modulos'))
    finally:
        fechar_cursor_conexao(cur, con)


@app.route('/configuracoes/modulos/<int:modulo_id>/editar', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador')
def editar_modulo_sistema(modulo_id):
    if not usuario_pode_gerenciar_menus_globais():
        flash('Somente Super Admin ou Suporte do sistema pode editar módulos globais.', 'danger')
        return redirect(url_for('perfil_acesso'))
    con = obter_conexao(); cur = con.cursor(dictionary=True)
    try:
        modulo = buscar_modulo_sistema(cur, modulo_id)
        if not modulo:
            flash('Módulo não encontrado.', 'warning'); return redirect(url_for('gerenciar_menus_modulos'))
        if request.method == 'POST':
            nome = (request.form.get('nome') or '').strip()
            codigo = (request.form.get('codigo') or normalizar_codigo_menu_modulo(nome)).strip()
            descricao = (request.form.get('descricao') or '').strip() or None
            icone = (request.form.get('icone') or 'fa-solid fa-folder').strip()
            ordem = int(request.form.get('ordem') or 0)
            ativo = 1 if request.form.get('ativo') == '1' else 0
            visivel_menu = 1 if request.form.get('visivel_menu') == '1' else 0
            somente_super_admin = 1 if request.form.get('somente_super_admin') == '1' else 0
            somente_suporte = 1 if request.form.get('somente_suporte') == '1' else 0
            cur.execute("""
                UPDATE sistema_modulos
                   SET codigo=%s, nome=%s, descricao=%s, icone=%s, ordem=%s, ativo=%s, visivel_menu=%s,
                       somente_super_admin=%s, somente_suporte=%s
                 WHERE id=%s
            """, (codigo, nome, descricao, icone, ordem, ativo, visivel_menu, somente_super_admin, somente_suporte, modulo_id))
            con.commit(); invalidar_cache_permissoes()
            registrar_auditoria_permissao_segura('EDITAR_MODULO_SISTEMA', 'sistema_modulos', modulo_id, json.dumps({'codigo':codigo,'nome':nome}, ensure_ascii=False), None, None)
            flash('Módulo atualizado com sucesso.', 'success')
            return redirect(url_for('gerenciar_menus_modulos'))
        return render_template('modulo_sistema_form.html', modulo=modulo)
    except Exception as e:
        con.rollback(); flash(f'Erro ao editar módulo: {e}', 'danger'); return redirect(url_for('gerenciar_menus_modulos'))
    finally:
        fechar_cursor_conexao(cur, con)


@app.post('/configuracoes/modulos/<int:modulo_id>/alternar')
@login_required
@perfis_permitidos('Administrador')
def alternar_modulo_sistema(modulo_id):
    if not usuario_pode_gerenciar_menus_globais():
        flash('Somente Super Admin ou Suporte pode alterar módulos globais.', 'danger')
        return redirect(url_for('perfil_acesso'))
    con=obter_conexao(); cur=con.cursor(dictionary=True)
    try:
        modulo=buscar_modulo_sistema(cur, modulo_id)
        if modulo:
            novo=0 if int(modulo.get('ativo') or 0)==1 else 1
            cur.execute('UPDATE sistema_modulos SET ativo=%s WHERE id=%s', (novo, modulo_id))
            con.commit(); invalidar_cache_permissoes()
            registrar_auditoria_permissao_segura('ALTERAR_STATUS_MODULO', 'sistema_modulos', modulo_id, json.dumps({'ativo':novo}, ensure_ascii=False), None, None)
            flash('Status do módulo alterado.', 'success')
    except Exception as e:
        con.rollback(); flash(f'Erro ao alterar módulo: {e}', 'danger')
    finally:
        fechar_cursor_conexao(cur, con)
    return redirect(url_for('gerenciar_menus_modulos'))


@app.route('/configuracoes/menus/novo', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador')
def novo_menu_sistema():
    if not usuario_pode_gerenciar_menus_globais():
        flash('Somente Super Admin ou Suporte do sistema pode criar menus globais.', 'danger')
        return redirect(url_for('perfil_acesso'))
    con = obter_conexao(); cur = con.cursor(dictionary=True)
    try:
        modulos = listar_modulos_sistema(cur)
        menus_pai = listar_menus_sistema(cur)
        if request.method == 'POST':
            titulo = (request.form.get('titulo') or '').strip()
            codigo = (request.form.get('codigo') or normalizar_codigo_menu_modulo(titulo)).strip()
            grupo_menu = (request.form.get('grupo_menu') or '').strip().upper()
            modulo_id = request.form.get('modulo_id') or None
            menu_pai_id = request.form.get('menu_pai_id') or None
            endpoint = (request.form.get('endpoint') or '').strip() or None
            rota_url = (request.form.get('rota_url') or '').strip() or None
            descricao = (request.form.get('descricao') or '').strip() or None
            icone = (request.form.get('icone') or 'fa-solid fa-circle').strip()
            ordem = int(request.form.get('ordem') or 0)
            ativo = 1 if request.form.get('ativo') == '1' else 0
            visivel_menu = 1 if request.form.get('visivel_menu') == '1' else 0
            exige_empresa = 1 if request.form.get('exige_empresa') == '1' else 0
            somente_super_admin = 1 if request.form.get('somente_super_admin') == '1' else 0
            somente_suporte = 1 if request.form.get('somente_suporte') == '1' else 0
            if endpoint and endpoint not in app.view_functions:
                flash('Atenção: endpoint informado ainda não existe no app.py. O menu foi salvo, mas confira a rota antes de liberar.', 'warning')
            if not titulo or not codigo or not grupo_menu:
                flash('Informe título, código e grupo do menu.', 'warning')
                return redirect(url_for('novo_menu_sistema'))
            cur.execute("""
                INSERT INTO sistema_menus (modulo_id, menu_pai_id, grupo_menu, codigo, titulo, descricao, endpoint, rota_url, icone, ordem, ativo, visivel_menu, exige_empresa, somente_super_admin, somente_suporte)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                ON DUPLICATE KEY UPDATE modulo_id=VALUES(modulo_id), menu_pai_id=VALUES(menu_pai_id), grupo_menu=VALUES(grupo_menu), titulo=VALUES(titulo),
                    descricao=VALUES(descricao), endpoint=VALUES(endpoint), rota_url=VALUES(rota_url), icone=VALUES(icone), ordem=VALUES(ordem), ativo=VALUES(ativo),
                    visivel_menu=VALUES(visivel_menu), exige_empresa=VALUES(exige_empresa), somente_super_admin=VALUES(somente_super_admin), somente_suporte=VALUES(somente_suporte)
            """, (modulo_id, menu_pai_id, grupo_menu, codigo, titulo, descricao, endpoint, rota_url, icone, ordem, ativo, visivel_menu, exige_empresa, somente_super_admin, somente_suporte))
            menu_id = cur.lastrowid
            # Liberação opcional inicial para Administrador global apenas quando solicitado.
            if request.form.get('liberar_admin_global') == '1':
                cur.execute("""
                    INSERT INTO perfil_permissoes (perfil_de_acesso, menu_codigo, acao_codigo, empresa_id, permitido)
                    VALUES ('Administrador', %s, 'visualizar', 0, 1)
                    ON DUPLICATE KEY UPDATE permitido=1
                """, (codigo,))
            con.commit(); invalidar_cache_permissoes()
            registrar_auditoria_permissao_segura('CRIAR_MENU_SISTEMA', 'sistema_menus', menu_id, json.dumps({'codigo':codigo,'titulo':titulo,'endpoint':endpoint}, ensure_ascii=False), None, None)
            flash('Menu salvo com sucesso.', 'success')
            return redirect(url_for('gerenciar_menus_modulos'))
        return render_template('menu_sistema_form.html', menu=None, modulos=modulos, menus_pai=menus_pai, endpoint_existe=True)
    except Exception as e:
        con.rollback(); flash(f'Erro ao salvar menu: {e}', 'danger'); return redirect(url_for('gerenciar_menus_modulos'))
    finally:
        fechar_cursor_conexao(cur, con)


@app.route('/configuracoes/menus/<int:menu_id>/editar', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador')
def editar_menu_sistema(menu_id):
    if not usuario_pode_gerenciar_menus_globais():
        flash('Somente Super Admin ou Suporte do sistema pode editar menus globais.', 'danger')
        return redirect(url_for('perfil_acesso'))
    con = obter_conexao(); cur = con.cursor(dictionary=True)
    try:
        menu = buscar_menu_sistema(cur, menu_id)
        if not menu:
            flash('Menu não encontrado.', 'warning'); return redirect(url_for('gerenciar_menus_modulos'))
        modulos = listar_modulos_sistema(cur)
        menus_pai = [m for m in listar_menus_sistema(cur) if int(m.get('id')) != int(menu_id)]
        if request.method == 'POST':
            titulo = (request.form.get('titulo') or '').strip()
            codigo = (request.form.get('codigo') or normalizar_codigo_menu_modulo(titulo)).strip()
            grupo_menu = (request.form.get('grupo_menu') or '').strip().upper()
            modulo_id = request.form.get('modulo_id') or None
            menu_pai_id = request.form.get('menu_pai_id') or None
            endpoint = (request.form.get('endpoint') or '').strip() or None
            rota_url = (request.form.get('rota_url') or '').strip() or None
            descricao = (request.form.get('descricao') or '').strip() or None
            icone = (request.form.get('icone') or 'fa-solid fa-circle').strip()
            ordem = int(request.form.get('ordem') or 0)
            ativo = 1 if request.form.get('ativo') == '1' else 0
            visivel_menu = 1 if request.form.get('visivel_menu') == '1' else 0
            exige_empresa = 1 if request.form.get('exige_empresa') == '1' else 0
            somente_super_admin = 1 if request.form.get('somente_super_admin') == '1' else 0
            somente_suporte = 1 if request.form.get('somente_suporte') == '1' else 0
            if endpoint and endpoint not in app.view_functions:
                flash('Atenção: endpoint informado ainda não existe no app.py.', 'warning')
            cur.execute("""
                UPDATE sistema_menus
                   SET modulo_id=%s, menu_pai_id=%s, grupo_menu=%s, codigo=%s, titulo=%s, descricao=%s, endpoint=%s, rota_url=%s,
                       icone=%s, ordem=%s, ativo=%s, visivel_menu=%s, exige_empresa=%s, somente_super_admin=%s, somente_suporte=%s
                 WHERE id=%s
            """, (modulo_id, menu_pai_id, grupo_menu, codigo, titulo, descricao, endpoint, rota_url, icone, ordem, ativo, visivel_menu, exige_empresa, somente_super_admin, somente_suporte, menu_id))
            con.commit(); invalidar_cache_permissoes()
            registrar_auditoria_permissao_segura('EDITAR_MENU_SISTEMA', 'sistema_menus', menu_id, json.dumps({'codigo':codigo,'titulo':titulo,'endpoint':endpoint}, ensure_ascii=False), None, None)
            flash('Menu atualizado com sucesso.', 'success')
            return redirect(url_for('gerenciar_menus_modulos'))
        return render_template('menu_sistema_form.html', menu=menu, modulos=modulos, menus_pai=menus_pai, endpoint_existe=endpoint_existe_no_app(menu.get('endpoint')))
    except Exception as e:
        con.rollback(); flash(f'Erro ao editar menu: {e}', 'danger'); return redirect(url_for('gerenciar_menus_modulos'))
    finally:
        fechar_cursor_conexao(cur, con)


@app.post('/configuracoes/menus/<int:menu_id>/alternar')
@login_required
@perfis_permitidos('Administrador')
def alternar_menu_sistema(menu_id):
    if not usuario_pode_gerenciar_menus_globais():
        flash('Somente Super Admin ou Suporte pode alterar menus globais.', 'danger')
        return redirect(url_for('perfil_acesso'))
    con=obter_conexao(); cur=con.cursor(dictionary=True)
    try:
        menu=buscar_menu_sistema(cur, menu_id)
        if menu:
            novo=0 if int(menu.get('ativo') or 0)==1 else 1
            cur.execute('UPDATE sistema_menus SET ativo=%s WHERE id=%s', (novo, menu_id))
            con.commit(); invalidar_cache_permissoes()
            registrar_auditoria_permissao_segura('ALTERAR_STATUS_MENU', 'sistema_menus', menu_id, json.dumps({'ativo':novo}, ensure_ascii=False), None, None)
            flash('Status do menu alterado.', 'success')
    except Exception as e:
        con.rollback(); flash(f'Erro ao alterar menu: {e}', 'danger')
    finally:
        fechar_cursor_conexao(cur, con)
    return redirect(url_for('gerenciar_menus_modulos'))

# ==========================================================
# ESTORNO DE FATURAMENTO
# NF Faturada -> Estornada
# Rotas Faturadas -> Pendente
# Remove vínculos nota/rota para liberar novo faturamento
# ==========================================================
@app.route('/faturamento/estornar/<int:nota_id>', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Financeiro')
def estornar_faturamento(nota_id):
    motivo = request.form.get('motivo_estorno', '').strip()
    usuario_id = session.get('usuario_id')

    if not motivo:
        flash('Informe o motivo do estorno do faturamento.', 'danger')
        return redirect(url_for('faturamento'))

    con = obter_conexao()
    if con is None:
        flash('Erro de conexão com o banco de dados.', 'danger')
        return redirect(url_for('faturamento'))

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT id,
                           numero_nf,
                           COALESCE(status_nf, 'Faturada') AS status_nf
                    FROM notas_fiscais
                    WHERE id = %s LIMIT 1
                    """, (nota_id,))
        nota = cur.fetchone()

        if not nota:
            flash('Nota Fiscal não encontrada.', 'danger')
            return redirect(url_for('faturamento'))

        if nota['status_nf'] == 'Estornada':
            flash('Esta Nota Fiscal já está estornada.', 'warning')
            return redirect(url_for('faturamento'))

        cur.execute("""
                    SELECT r.id,
                           r.identi_rota,
                           r.situacao_rota
                    FROM nota_fiscal_rotas nfr
                             INNER JOIN rotas r ON r.id = nfr.rota_id
                    WHERE nfr.nota_fiscal_id = %s
                    """, (nota_id,))
        rotas = cur.fetchall()

        if not rotas:
            flash('Esta Nota Fiscal não possui rotas vinculadas para estornar.', 'danger')
            return redirect(url_for('faturamento'))

        for rota in rotas:
            if rota['situacao_rota'] == 'Quitada':
                flash(
                    f"A rota {rota['identi_rota']} já está Quitada. "
                    f"Primeiro estorne o recebimento e depois estorne o faturamento.",
                    'danger'
                )
                return redirect(url_for('faturamento'))

            if rota['situacao_rota'] != 'Faturada':
                flash(
                    f"A rota {rota['identi_rota']} não está Faturada. "
                    f"Operação cancelada por segurança.",
                    'danger'
                )
                return redirect(url_for('faturamento'))

        for rota in rotas:
            cur.execute("""
                        UPDATE rotas
                        SET situacao_rota = 'Pendente'
                        WHERE id = %s
                          AND situacao_rota = 'Faturada'
                        """, (rota['id'],))

            if cur.rowcount != 1:
                raise Exception(f"Falha ao devolver a rota {rota['identi_rota']} para Pendente.")

            cur.execute("""
                        INSERT INTO historico_operacoes (tipo_operacao,
                                                         rota_id,
                                                         nota_fiscal_id,
                                                         usuario_id,
                                                         status_anterior,
                                                         status_novo,
                                                         motivo,
                                                         observacao)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            'ESTORNO_FATURAMENTO',
                            rota['id'],
                            nota_id,
                            usuario_id,
                            'Faturada',
                            'Pendente',
                            motivo,
                            f"Estorno da NF nº {nota['numero_nf']}"
                        ))

        cur.execute("""
                    DELETE
                    FROM nota_fiscal_rotas
                    WHERE nota_fiscal_id = %s
                    """, (nota_id,))

        cur.execute("""
                    UPDATE notas_fiscais
                    SET status_nf          = 'Estornada',
                        data_estorno       = NOW(),
                        motivo_estorno     = %s,
                        usuario_estorno_id = %s
                    WHERE id = %s
                    """, (motivo, usuario_id, nota_id))

        con.commit()

        flash(
            f"Faturamento da NF nº {nota['numero_nf']} estornado com sucesso. "
            f"As rotas voltaram para Pendente.",
            'success'
        )

    except Exception as e:
        con.rollback()
        print(f'Erro ao estornar faturamento: {e}')
        flash('Erro técnico ao estornar faturamento.', 'danger')

    finally:
        fechar_cursor_conexao(cur, con)

    return redirect(url_for('faturamento'))


# ==========================================================
# ESTORNO DE RECEBIMENTO
# Rotas Quitadas -> Faturada
# Mantém vínculo da nota fiscal
# ==========================================================
@app.route('/financeiro/recebimento/estornar/<int:nota_id>', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Financeiro')
def estornar_recebimento(nota_id):
    motivo = request.form.get('motivo_estorno', '').strip()
    usuario_id = session.get('usuario_id')

    if not motivo:
        flash('Informe o motivo do estorno do recebimento.', 'danger')
        return redirect(url_for('recebimento'))

    con = obter_conexao()
    if con is None:
        flash('Erro de conexão ao estornar recebimento.', 'danger')
        return redirect(url_for('recebimento'))

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT id,
                           numero_nf,
                           COALESCE(status_nf, 'Faturada') AS status_nf
                    FROM notas_fiscais
                    WHERE id = %s LIMIT 1
                    """, (nota_id,))
        nota = cur.fetchone()

        if not nota:
            flash('Nota fiscal não encontrada.', 'warning')
            return redirect(url_for('recebimento'))

        if nota['status_nf'] != 'Faturada':
            flash('Não é possível estornar recebimento de uma nota estornada.', 'danger')
            return redirect(url_for('recebimento'))

        cur.execute("""
                    SELECT r.id,
                           r.identi_rota,
                           r.situacao_rota
                    FROM nota_fiscal_rotas nfr
                             INNER JOIN rotas r ON r.id = nfr.rota_id
                    WHERE nfr.nota_fiscal_id = %s
                      AND r.situacao_rota = 'Quitada'
                    """, (nota_id,))
        rotas = cur.fetchall()

        if not rotas:
            flash('Não existem rotas quitadas para estornar nesta nota.', 'warning')
            return redirect(url_for('recebimento'))

        for rota in rotas:
            cur.execute("""
                        UPDATE rotas
                        SET situacao_rota = 'Faturada'
                        WHERE id = %s
                          AND situacao_rota = 'Quitada'
                        """, (rota['id'],))

            if cur.rowcount != 1:
                raise Exception(f"Falha ao devolver a rota {rota['identi_rota']} para Faturada.")

            cur.execute("""
                        INSERT INTO historico_operacoes (tipo_operacao,
                                                         rota_id,
                                                         nota_fiscal_id,
                                                         usuario_id,
                                                         status_anterior,
                                                         status_novo,
                                                         motivo,
                                                         observacao)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            'ESTORNO_RECEBIMENTO',
                            rota['id'],
                            nota_id,
                            usuario_id,
                            'Quitada',
                            'Faturada',
                            motivo,
                            f"Estorno de recebimento da NF nº {nota['numero_nf']}"
                        ))

        con.commit()

        flash(
            f"Recebimento da NF nº {nota['numero_nf']} estornado com sucesso. "
            f"As rotas voltaram para Faturada.",
            'success'
        )

    except Exception as e:
        con.rollback()
        print(f'Erro ao estornar recebimento: {e}')
        flash('Erro técnico ao estornar recebimento.', 'danger')

    finally:
        fechar_cursor_conexao(cur, con)

    return redirect(url_for('recebimento'))


# ==========================================================
# HISTÓRICO DE ESTORNOS
# ==========================================================
@app.route('/financeiro/historico-estornos')
@login_required
@perfis_permitidos('Administrador', 'Financeiro', 'Consulta')
def historico_estornos():
    usuario_logado = session.get('usuario_nome', 'Usuário')

    con = obter_conexao()
    if con is None:
        flash('Erro de conexão ao carregar histórico.', 'danger')
        return render_template('historico_estornos.html', historico=[], usuario_logado=usuario_logado)

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT h.id,
                           h.data_operacao,
                           h.tipo_operacao,
                           h.status_anterior,
                           h.status_novo,
                           h.motivo,
                           h.observacao,
                           r.identi_rota,
                           nf.numero_nf,
                           p_user.nome_completo AS usuario_nome
                    FROM historico_operacoes h
                             LEFT JOIN rotas r ON r.id = h.rota_id
                             LEFT JOIN notas_fiscais nf ON nf.id = h.nota_fiscal_id
                             LEFT JOIN usuarios u ON u.id = h.usuario_id
                             LEFT JOIN pessoas p_user ON p_user.id = u.pessoa_id
                    WHERE h.tipo_operacao IN ('ESTORNO_FATURAMENTO', 'ESTORNO_RECEBIMENTO', 'BAIXA_RECEBIMENTO')
                    ORDER BY h.data_operacao DESC, h.id DESC LIMIT 300
                    """)
        historico = cur.fetchall()

    except Exception as e:
        print(f'Erro ao carregar histórico de estornos: {e}')
        flash('Erro ao carregar histórico de estornos.', 'danger')
        historico = []

    finally:
        fechar_cursor_conexao(cur, con)

    return render_template(
        'historico_estornos.html',
        historico=historico,
        usuario_logado=usuario_logado
    )


# ==========================================================
# PATCH - PAGAMENTOS DE AJUDANTE
# Requer: obter_conexao, login_required, fechar_cursor_conexao, limitar_texto.
# ==========================================================

@app.route('/financeiro/pagamentos-ajudante')
@login_required
@perfis_permitidos('Administrador', 'Financeiro', 'Consulta')
def pagamentos_ajudante():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_id = session.get('empresa_id')

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    con = obter_conexao()

    if con is None:
        flash('Erro de conexão ao carregar pagamentos de ajudante.', 'danger')
        return render_template(
            'pagamentos_ajudante.html',
            usuario_logado=usuario_logado,
            pendentes=[],
            quitados=[],
            estornados=[],
            cards={}
        )

    cur = con.cursor(dictionary=True)

    try:
        query_base = """
                     SELECT la.id,
                            la.data_lancamento,
                            la.valor_total,
                            la.status_pagamento,
                            la.observacao,
                            la.data_pagamento,
                            la.data_estorno_pagamento,
                            la.motivo_estorno_pagamento,
                            la.data_estorno_lancamento,
                            la.motivo_estorno_lancamento,
                            p.id                                                         AS ajudante_id,
                            p.nome_completo                                              AS ajudante_nome,
                            COUNT(lar.id)                                                AS qtd_rotas,
                            GROUP_CONCAT(lar.identi_rota ORDER BY lar.id SEPARATOR ', ') AS rotas
                     FROM lancamentos_ajudantes la
                              INNER JOIN pessoas p
                                         ON p.id = la.ajudante_id
                                             AND p.empresa_id = la.empresa_id
                              LEFT JOIN lancamento_ajudante_rotas lar
                                        ON lar.lancamento_ajudante_id = la.id
                                            AND lar.empresa_id = la.empresa_id
                     WHERE la.empresa_id = %s
                       AND la.status_pagamento = %s
                     GROUP BY la.id,
                              la.data_lancamento,
                              la.valor_total,
                              la.status_pagamento,
                              la.observacao,
                              la.data_pagamento,
                              la.data_estorno_pagamento,
                              la.motivo_estorno_pagamento,
                              la.data_estorno_lancamento,
                              la.motivo_estorno_lancamento,
                              p.id,
                              p.nome_completo
                     ORDER BY la.data_lancamento DESC, la.id DESC \
                     """

        cur.execute(query_base, (empresa_id, 'Pendente'))
        pendentes = cur.fetchall()

        cur.execute(query_base, (empresa_id, 'Quitado'))
        quitados = cur.fetchall()

        cur.execute(query_base, (empresa_id, 'Estornado'))
        estornados = cur.fetchall()

        cur.execute("""
                    SELECT COALESCE(SUM(CASE WHEN status_pagamento = 'Pendente' THEN valor_total ELSE 0 END),
                                    0) AS total_pendente,
                           COALESCE(SUM(CASE WHEN status_pagamento = 'Quitado' THEN valor_total ELSE 0 END),
                                    0) AS total_quitado,
                           COALESCE(SUM(CASE WHEN status_pagamento = 'Estornado' THEN valor_total ELSE 0 END),
                                    0) AS total_estornado,

                           COALESCE(SUM(CASE WHEN status_pagamento = 'Pendente' THEN 1 ELSE 0 END),
                                    0) AS qtd_pendente,
                           COALESCE(SUM(CASE WHEN status_pagamento = 'Quitado' THEN 1 ELSE 0 END),
                                    0) AS qtd_quitado,
                           COALESCE(SUM(CASE WHEN status_pagamento = 'Estornado' THEN 1 ELSE 0 END),
                                    0) AS qtd_estornado
                    FROM lancamentos_ajudantes
                    WHERE empresa_id = %s
                    """, (empresa_id,))

        cards = cur.fetchone() or {}

    except Exception as e:
        print(f'Erro ao carregar pagamentos de ajudante: {e}')
        flash('Erro ao carregar pagamentos de ajudante.', 'danger')
        pendentes, quitados, estornados, cards = [], [], [], {}

    finally:
        fechar_cursor_conexao(cur, con)

    return render_template(
        'pagamentos_ajudante.html',
        usuario_logado=usuario_logado,
        pendentes=pendentes,
        quitados=quitados,
        estornados=estornados,
        cards=cards
    )


@app.route('/financeiro/pagamentos-ajudante/baixar/<int:lancamento_id>', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Financeiro')
def baixar_pagamento_ajudante(lancamento_id):
    usuario_id = session.get('usuario_id')
    empresa_id = session.get('empresa_id')

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    con = obter_conexao()

    if con is None:
        flash('Erro de conexão ao baixar pagamento.', 'danger')
        return redirect(url_for('pagamentos_ajudante'))

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT la.id,
                           la.empresa_id,
                           la.ajudante_id,
                           la.valor_total,
                           la.status_pagamento,
                           p.nome_completo AS ajudante_nome
                    FROM lancamentos_ajudantes la
                             INNER JOIN pessoas p
                                        ON p.id = la.ajudante_id
                                            AND p.empresa_id = la.empresa_id
                    WHERE la.id = %s
                      AND la.empresa_id = %s LIMIT 1
                    """, (lancamento_id, empresa_id))

        lancamento = cur.fetchone()

        if not lancamento:
            flash('Lançamento não encontrado ou não pertence à empresa logada.', 'warning')
            return redirect(url_for('pagamentos_ajudante'))

        if lancamento['status_pagamento'] != 'Pendente':
            flash('Só é possível baixar pagamentos com status Pendente.', 'danger')
            return redirect(url_for('pagamentos_ajudante'))

        cur.execute("""
                    SELECT rota_id, identi_rota, valor_ajudante
                    FROM lancamento_ajudante_rotas
                    WHERE lancamento_ajudante_id = %s
                      AND empresa_id = %s
                    """, (lancamento_id, empresa_id))

        rotas = cur.fetchall()

        if not rotas:
            flash('Este lançamento não possui rotas vinculadas. Operação cancelada.', 'danger')
            return redirect(url_for('pagamentos_ajudante'))

        cur.execute("""
                    UPDATE lancamentos_ajudantes
                    SET status_pagamento         = 'Quitado',
                        data_pagamento           = NOW(),
                        usuario_pagamento_id     = %s,
                        data_estorno_pagamento   = NULL,
                        motivo_estorno_pagamento = NULL
                    WHERE id = %s
                      AND empresa_id = %s
                      AND status_pagamento = 'Pendente'
                    """, (usuario_id, lancamento_id, empresa_id))

        if cur.rowcount != 1:
            raise Exception('Falha ao atualizar status do pagamento para Quitado.')

        for rota in rotas:
            cur.execute("""
                        INSERT INTO historico_ajudante_pagamentos (empresa_id,
                                                                   tipo_operacao,
                                                                   lancamento_ajudante_id,
                                                                   ajudante_id,
                                                                   rota_id,
                                                                   identi_rota,
                                                                   usuario_id,
                                                                   status_anterior,
                                                                   status_novo,
                                                                   valor_operacao,
                                                                   motivo,
                                                                   observacao)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            empresa_id,
                            'BAIXA_PAGAMENTO_AJUDANTE',
                            lancamento_id,
                            lancamento['ajudante_id'],
                            rota['rota_id'],
                            rota['identi_rota'],
                            usuario_id,
                            'Pendente',
                            'Quitado',
                            rota['valor_ajudante'],
                            'Baixa de pagamento',
                            f"Baixa do pagamento do ajudante {lancamento['ajudante_nome']}"
                        ))

        con.commit()

        flash(f"Pagamento do ajudante {lancamento['ajudante_nome']} baixado com sucesso.", 'success')

    except Exception as e:
        con.rollback()
        print(f'Erro ao baixar pagamento de ajudante: {e}')
        flash('Erro técnico ao baixar pagamento de ajudante.', 'danger')

    finally:
        fechar_cursor_conexao(cur, con)

    return redirect(url_for('pagamentos_ajudante'))


@app.route('/financeiro/pagamentos-ajudante/estornar-baixa/<int:lancamento_id>', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Financeiro')
def estornar_baixa_pagamento_ajudante(lancamento_id):
    usuario_id = session.get('usuario_id')
    empresa_id = session.get('empresa_id')
    motivo = limitar_texto(request.form.get('motivo_estorno'), 255)

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    if not motivo:
        flash('Informe o motivo do estorno da baixa.', 'danger')
        return redirect(url_for('pagamentos_ajudante'))

    con = obter_conexao()

    if con is None:
        flash('Erro de conexão ao estornar baixa.', 'danger')
        return redirect(url_for('pagamentos_ajudante'))

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT la.id,
                           la.empresa_id,
                           la.ajudante_id,
                           la.valor_total,
                           la.status_pagamento,
                           p.nome_completo AS ajudante_nome
                    FROM lancamentos_ajudantes la
                             INNER JOIN pessoas p
                                        ON p.id = la.ajudante_id
                                            AND p.empresa_id = la.empresa_id
                    WHERE la.id = %s
                      AND la.empresa_id = %s LIMIT 1
                    """, (lancamento_id, empresa_id))

        lancamento = cur.fetchone()

        if not lancamento:
            flash('Lançamento não encontrado ou não pertence à empresa logada.', 'warning')
            return redirect(url_for('pagamentos_ajudante'))

        if lancamento['status_pagamento'] != 'Quitado':
            flash('Só é possível estornar baixa de pagamento com status Quitado.', 'danger')
            return redirect(url_for('pagamentos_ajudante'))

        cur.execute("""
                    SELECT rota_id, identi_rota, valor_ajudante
                    FROM lancamento_ajudante_rotas
                    WHERE lancamento_ajudante_id = %s
                      AND empresa_id = %s
                    """, (lancamento_id, empresa_id))

        rotas = cur.fetchall()

        cur.execute("""
                    UPDATE lancamentos_ajudantes
                    SET status_pagamento         = 'Pendente',
                        data_estorno_pagamento   = NOW(),
                        motivo_estorno_pagamento = %s,
                        usuario_estorno_id       = %s
                    WHERE id = %s
                      AND empresa_id = %s
                      AND status_pagamento = 'Quitado'
                    """, (motivo, usuario_id, lancamento_id, empresa_id))

        if cur.rowcount != 1:
            raise Exception('Falha ao devolver pagamento para Pendente.')

        for rota in rotas:
            cur.execute("""
                        INSERT INTO historico_ajudante_pagamentos (empresa_id,
                                                                   tipo_operacao,
                                                                   lancamento_ajudante_id,
                                                                   ajudante_id,
                                                                   rota_id,
                                                                   identi_rota,
                                                                   usuario_id,
                                                                   status_anterior,
                                                                   status_novo,
                                                                   valor_operacao,
                                                                   motivo,
                                                                   observacao)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            empresa_id,
                            'ESTORNO_BAIXA_PAGAMENTO_AJUDANTE',
                            lancamento_id,
                            lancamento['ajudante_id'],
                            rota['rota_id'],
                            rota['identi_rota'],
                            usuario_id,
                            'Quitado',
                            'Pendente',
                            rota['valor_ajudante'],
                            motivo,
                            f"Estorno da baixa do ajudante {lancamento['ajudante_nome']}"
                        ))

        con.commit()

        flash(f"Baixa do pagamento do ajudante {lancamento['ajudante_nome']} estornada com sucesso.", 'success')

    except Exception as e:
        con.rollback()
        print(f'Erro ao estornar baixa de pagamento: {e}')
        flash('Erro técnico ao estornar baixa de pagamento.', 'danger')

    finally:
        fechar_cursor_conexao(cur, con)

    return redirect(url_for('pagamentos_ajudante'))


@app.route('/financeiro/pagamentos-ajudante/estornar-lancamento/<int:lancamento_id>', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Financeiro')
def estornar_lancamento_ajudante(lancamento_id):
    usuario_id = session.get('usuario_id')
    empresa_id = session.get('empresa_id')
    motivo = limitar_texto(request.form.get('motivo_estorno'), 255)

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    if not motivo:
        flash('Informe o motivo do estorno do lançamento.', 'danger')
        return redirect(url_for('pagamentos_ajudante'))

    con = obter_conexao()

    if con is None:
        flash('Erro de conexão ao estornar lançamento.', 'danger')
        return redirect(url_for('pagamentos_ajudante'))

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT la.id,
                           la.empresa_id,
                           la.ajudante_id,
                           la.valor_total,
                           la.status_pagamento,
                           p.nome_completo AS ajudante_nome
                    FROM lancamentos_ajudantes la
                             INNER JOIN pessoas p
                                        ON p.id = la.ajudante_id
                                            AND p.empresa_id = la.empresa_id
                    WHERE la.id = %s
                      AND la.empresa_id = %s LIMIT 1
                    """, (lancamento_id, empresa_id))

        lancamento = cur.fetchone()

        if not lancamento:
            flash('Lançamento não encontrado ou não pertence à empresa logada.', 'warning')
            return redirect(url_for('pagamentos_ajudante'))

        if lancamento['status_pagamento'] == 'Estornado':
            flash('Este lançamento já está estornado.', 'warning')
            return redirect(url_for('pagamentos_ajudante'))

        if lancamento['status_pagamento'] == 'Quitado':
            flash('Este lançamento já está quitado. Primeiro estorne a baixa, depois estorne o lançamento.', 'danger')
            return redirect(url_for('pagamentos_ajudante'))

        if lancamento['status_pagamento'] != 'Pendente':
            flash('Só é possível estornar lançamento com status Pendente.', 'danger')
            return redirect(url_for('pagamentos_ajudante'))

        cur.execute("""
                    SELECT rota_id, identi_rota, valor_ajudante
                    FROM lancamento_ajudante_rotas
                    WHERE lancamento_ajudante_id = %s
                      AND empresa_id = %s
                    """, (lancamento_id, empresa_id))

        rotas = cur.fetchall()

        if not rotas:
            flash('Este lançamento não possui rotas vinculadas. Operação cancelada.', 'danger')
            return redirect(url_for('pagamentos_ajudante'))

        cur.execute("""
                    UPDATE lancamentos_ajudantes
                    SET status_pagamento          = 'Estornado',
                        data_estorno_lancamento   = NOW(),
                        motivo_estorno_lancamento = %s,
                        usuario_estorno_id        = %s
                    WHERE id = %s
                      AND empresa_id = %s
                      AND status_pagamento = 'Pendente'
                    """, (motivo, usuario_id, lancamento_id, empresa_id))

        if cur.rowcount != 1:
            raise Exception('Falha ao marcar lançamento como Estornado.')

        for rota in rotas:
            cur.execute("""
                        INSERT INTO historico_ajudante_pagamentos (empresa_id,
                                                                   tipo_operacao,
                                                                   lancamento_ajudante_id,
                                                                   ajudante_id,
                                                                   rota_id,
                                                                   identi_rota,
                                                                   usuario_id,
                                                                   status_anterior,
                                                                   status_novo,
                                                                   valor_operacao,
                                                                   motivo,
                                                                   observacao)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            empresa_id,
                            'ESTORNO_LANCAMENTO_AJUDANTE',
                            lancamento_id,
                            lancamento['ajudante_id'],
                            rota['rota_id'],
                            rota['identi_rota'],
                            usuario_id,
                            'Pendente',
                            'Estornado',
                            rota['valor_ajudante'],
                            motivo,
                            f"Estorno do lançamento do ajudante {lancamento['ajudante_nome']}. Rota liberada para novo vínculo."
                        ))

        cur.execute("""
                    DELETE
                    FROM lancamento_ajudante_rotas
                    WHERE lancamento_ajudante_id = %s
                      AND empresa_id = %s
                    """, (lancamento_id, empresa_id))

        con.commit()

        flash(
            f"Lançamento do ajudante {lancamento['ajudante_nome']} estornado com sucesso. "
            "As rotas foram liberadas para novo lançamento.",
            'success'
        )

    except Exception as e:
        con.rollback()
        print(f'Erro ao estornar lançamento de ajudante: {e}')
        flash('Erro técnico ao estornar lançamento de ajudante.', 'danger')

    finally:
        fechar_cursor_conexao(cur, con)

    return redirect(url_for('pagamentos_ajudante'))


@app.route('/financeiro/pagamentos-ajudante/historico')
@login_required
@perfis_permitidos('Administrador', 'Financeiro', 'Consulta')
def historico_pagamentos_ajudante():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_id = session.get('empresa_id')

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    con = obter_conexao()

    if con is None:
        flash('Erro de conexão ao carregar histórico.', 'danger')
        return render_template(
            'historico_pagamentos_ajudante.html',
            historico=[],
            usuario_logado=usuario_logado
        )

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT h.id,
                           h.data_operacao,
                           h.tipo_operacao,
                           h.lancamento_ajudante_id,
                           h.identi_rota,
                           h.status_anterior,
                           h.status_novo,
                           h.valor_operacao,
                           h.motivo,
                           h.observacao,
                           ajudante.nome_completo AS ajudante_nome,
                           operador.nome_completo AS usuario_nome
                    FROM historico_ajudante_pagamentos h
                             LEFT JOIN pessoas ajudante
                                       ON ajudante.id = h.ajudante_id
                                           AND ajudante.empresa_id = h.empresa_id
                             LEFT JOIN usuarios u
                                       ON u.id = h.usuario_id
                                           AND u.empresa_id = h.empresa_id
                             LEFT JOIN pessoas operador
                                       ON operador.id = u.pessoa_id
                                           AND operador.empresa_id = h.empresa_id
                    WHERE h.empresa_id = %s
                    ORDER BY h.data_operacao DESC, h.id DESC LIMIT 300
                    """, (empresa_id,))

        historico = cur.fetchall()

    except Exception as e:
        print(f'Erro ao carregar histórico de pagamentos de ajudante: {e}')
        flash('Erro ao carregar histórico de pagamentos de ajudante.', 'danger')
        historico = []

    finally:
        fechar_cursor_conexao(cur, con)

    return render_template(
        'historico_pagamentos_ajudante.html',
        historico=historico,
        usuario_logado=usuario_logado
    )


#
# Visualizar empresas
#
@app.route('/empresas')
@login_required
@super_admin_required
def visualizar_empresas():
    usuario_logado = session.get('usuario_nome', 'Usuário')

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão ao carregar empresas.", "danger")
        return render_template(
            'visualizar_empresas.html',
            empresas=[],
            usuario_logado=usuario_logado
        )

    cur = con.cursor(dictionary=True)

    try:
        normalizar_empresa_padrao_sistema(cur)
        con.commit()
        cur.execute("""
                    SELECT e.id,
                           e.data_cadastro,
                           e.razao_social,
                           e.nome_fantasia,
                           e.cnpj,
                           e.slug,
                           e.status_empresa,
                           e.plano,
                           e.limite_usuarios,
                           e.observacao,

                           (SELECT COUNT(*)
                            FROM usuarios u
                            WHERE u.empresa_id = e.id) AS total_usuarios,

                           (SELECT COUNT(*)
                            FROM pessoas p
                            WHERE p.empresa_id = e.id) AS total_pessoas,

                           (SELECT COUNT(*)
                            FROM rotas r
                            WHERE r.empresa_id = e.id) AS total_rotas

                    FROM empresas e
                    ORDER BY e.id ASC
                    """)

        empresas = cur.fetchall()

    except Exception as e:
        print(f"Erro ao listar empresas: {e}")
        flash("Erro técnico ao listar empresas.", "danger")
        empresas = []

    finally:
        fechar_cursor_conexao(cur, con)

    return render_template(
        'visualizar_empresas.html',
        empresas=empresas,
        usuario_logado=usuario_logado
    )


#
# Cadastrar empresa
#
@app.route('/empresas/cadastro', methods=['GET', 'POST'])
@login_required
@super_admin_required
def cadastro_empresa():
    usuario_logado = session.get('usuario_nome', 'Usuário')

    if request.method == 'POST':
        razao_social = request.form.get('razao_social', '').strip()
        nome_fantasia = request.form.get('nome_fantasia', '').strip()
        cnpj = request.form.get('cnpj', '').strip()
        slug = request.form.get('slug', '').strip()
        status_empresa = request.form.get('status_empresa', 'Ativa').strip()
        plano = request.form.get('plano', 'Profissional').strip()
        limite_usuarios = request.form.get('limite_usuarios', '').strip()
        observacao = request.form.get('observacao', '').strip()

        if not razao_social:
            flash("Informe a razão social da empresa.", "danger")
            return redirect(url_for('cadastro_empresa'))

        if status_empresa not in ['Ativa', 'Inativa']:
            flash("Status da empresa inválido.", "danger")
            return redirect(url_for('cadastro_empresa'))

        if plano not in ['Starter', 'Profissional', 'Enterprise', 'Ilimitado']:
            flash("Plano inválido.", "danger")
            return redirect(url_for('cadastro_empresa'))

        cnpj_limpo = somente_digitos(cnpj) if cnpj else None
        slug = gerar_slug_empresa(slug or nome_fantasia or razao_social)

        limite_usuarios_valor = None
        if limite_usuarios:
            try:
                limite_usuarios_valor = int(limite_usuarios)
            except Exception:
                flash("Limite de usuários inválido.", "danger")
                return redirect(url_for('cadastro_empresa'))

        con = obter_conexao()

        if con is None:
            flash("Erro de conexão com o banco de dados.", "danger")
            return redirect(url_for('cadastro_empresa'))

        cur = con.cursor(dictionary=True)

        try:
            cur.execute("""
                        SELECT id
                        FROM empresas
                        WHERE slug = %s LIMIT 1
                        """, (slug,))

            if cur.fetchone():
                flash("Já existe uma empresa com este slug.", "danger")
                return redirect(url_for('cadastro_empresa'))

            if cnpj_limpo:
                cur.execute("""
                            SELECT id
                            FROM empresas
                            WHERE cnpj = %s LIMIT 1
                            """, (cnpj_limpo,))

                if cur.fetchone():
                    flash("Já existe uma empresa com este CNPJ.", "danger")
                    return redirect(url_for('cadastro_empresa'))

            cur.execute("""
                        INSERT INTO empresas (razao_social,
                                              nome_fantasia,
                                              cnpj,
                                              slug,
                                              status_empresa,
                                              plano,
                                              limite_usuarios,
                                              observacao)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                        """, (
                            razao_social,
                            nome_fantasia or None,
                            cnpj_limpo,
                            slug,
                            status_empresa,
                            plano,
                            limite_usuarios_valor,
                            observacao or None
                        ))

            nova_empresa_id = cur.lastrowid
            try:
                sincronizar_perfis_padrao_empresas(cur, nova_empresa_id)
                semear_parametros_operacionais_motorista_empresa(cur, nova_empresa_id, usuario_id=session.get('usuario_id'))
            except Exception as e:
                print(f"[Empresa] Aviso ao criar estrutura padrão para nova empresa #{nova_empresa_id}: {e}")

            con.commit()

            flash("Empresa cadastrada com sucesso! Perfis e parâmetros padrão foram preparados.", "success")
            return redirect(url_for('visualizar_empresas'))

        except Exception as e:
            con.rollback()
            print(f"Erro ao cadastrar empresa: {e}")
            flash("Erro técnico ao cadastrar empresa.", "danger")
            return redirect(url_for('cadastro_empresa'))

        finally:
            fechar_cursor_conexao(cur, con)

    return render_template(
        'cadastro_empresa.html',
        usuario_logado=usuario_logado
    )


#
# Editar empresa
#
@app.route('/empresas/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@super_admin_required
def editar_empresa(id):
    usuario_logado = session.get('usuario_nome', 'Usuário')

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('visualizar_empresas'))

    cur = con.cursor(dictionary=True)

    if request.method == 'POST':
        razao_social = request.form.get('razao_social', '').strip()
        nome_fantasia = request.form.get('nome_fantasia', '').strip()
        cnpj = request.form.get('cnpj', '').strip()
        slug = request.form.get('slug', '').strip()
        status_empresa = request.form.get('status_empresa', 'Ativa').strip()
        plano = request.form.get('plano', 'Profissional').strip()
        limite_usuarios = request.form.get('limite_usuarios', '').strip()
        observacao = request.form.get('observacao', '').strip()

        if not razao_social:
            flash("Informe a razão social da empresa.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('editar_empresa', id=id))

        if status_empresa not in ['Ativa', 'Inativa']:
            flash("Status da empresa inválido.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('editar_empresa', id=id))

        if plano not in ['Starter', 'Profissional', 'Enterprise', 'Ilimitado']:
            flash("Plano inválido.", "danger")
            fechar_cursor_conexao(cur, con)
            return redirect(url_for('editar_empresa', id=id))

        cnpj_limpo = somente_digitos(cnpj) if cnpj else None
        slug = gerar_slug_empresa(slug or nome_fantasia or razao_social)

        limite_usuarios_valor = None
        if limite_usuarios:
            try:
                limite_usuarios_valor = int(limite_usuarios)
            except Exception:
                flash("Limite de usuários inválido.", "danger")
                fechar_cursor_conexao(cur, con)
                return redirect(url_for('editar_empresa', id=id))

        try:
            cur.execute("""
                        SELECT id, cnpj, status_empresa, plano, limite_usuarios
                        FROM empresas
                        WHERE id = %s LIMIT 1
                        """, (id,))

            empresa_atual = cur.fetchone()
            if not empresa_atual:
                flash("Empresa não encontrada.", "danger")
                return redirect(url_for('visualizar_empresas'))

            empresa_padrao_sistema = empresa_eh_padrao_sistema(empresa_atual)
            if empresa_padrao_sistema:
                cnpj_limpo = EMPRESA_PADRAO_SISTEMA_CNPJ
                slug = EMPRESA_PADRAO_SISTEMA_SLUG
                status_empresa = 'Ativa'
                plano = EMPRESA_PADRAO_SISTEMA_PLANO
                limite_usuarios_valor = None

            cur.execute("""
                        SELECT id
                        FROM empresas
                        WHERE slug = %s
                          AND id <> %s LIMIT 1
                        """, (slug, id))

            if cur.fetchone():
                flash("Já existe outra empresa com este slug.", "danger")
                return redirect(url_for('editar_empresa', id=id))

            if cnpj_limpo:
                cur.execute("""
                            SELECT id
                            FROM empresas
                            WHERE cnpj = %s
                              AND id <> %s LIMIT 1
                            """, (cnpj_limpo, id))

                if cur.fetchone():
                    flash("Já existe outra empresa com este CNPJ.", "danger")
                    return redirect(url_for('editar_empresa', id=id))

            cur.execute("""
                        UPDATE empresas
                        SET razao_social    = %s,
                            nome_fantasia   = %s,
                            cnpj            = %s,
                            slug            = %s,
                            status_empresa  = %s,
                            plano           = %s,
                            limite_usuarios = %s,
                            observacao      = %s
                        WHERE id = %s
                        """, (
                            razao_social,
                            nome_fantasia or None,
                            cnpj_limpo,
                            slug,
                            status_empresa,
                            plano,
                            limite_usuarios_valor,
                            observacao or None,
                            id
                        ))

            con.commit()

            flash("Empresa atualizada com sucesso!" + (" Campos críticos da empresa padrão foram mantidos protegidos." if id == EMPRESA_PADRAO_SISTEMA_ID else ""), "success")
            return redirect(url_for('visualizar_empresas'))

        except Exception as e:
            con.rollback()
            print(f"Erro ao editar empresa: {e}")
            flash("Erro técnico ao editar empresa.", "danger")
            return redirect(url_for('editar_empresa', id=id))

        finally:
            fechar_cursor_conexao(cur, con)

    try:
        normalizar_empresa_padrao_sistema(cur)
        con.commit()
        cur.execute("""
                    SELECT id,
                           razao_social,
                           nome_fantasia,
                           cnpj,
                           slug,
                           status_empresa,
                           plano,
                           limite_usuarios,
                           observacao
                    FROM empresas
                    WHERE id = %s LIMIT 1
                    """, (id,))

        empresa = cur.fetchone()

        if not empresa:
            flash("Empresa não encontrada.", "warning")
            return redirect(url_for('visualizar_empresas'))

        empresa_padrao_sistema = empresa_eh_padrao_sistema(empresa)

    except Exception as e:
        print(f"Erro ao carregar empresa: {e}")
        flash("Erro técnico ao carregar empresa.", "danger")
        return redirect(url_for('visualizar_empresas'))

    finally:
        fechar_cursor_conexao(cur, con)

    return render_template(
        'editar_empresa.html',
        empresa=empresa,
        empresa_padrao_sistema=empresa_padrao_sistema,
        usuario_logado=usuario_logado
    )


# ==========================================================
# FASE 3.4 - PAINEL FINANCEIRO DAS NFs DOS MOTORISTAS
# ==========================================================

def usuario_pode_analisar_nf_motorista():
    is_super_admin = int(session.get('is_super_admin') or 0) == 1
    perfil = session.get('perfil_de_acesso')

    if is_super_admin:
        return True

    perfis_permitidos = [
        'Administrador',
        'Financeiro',
        'Operacional'
    ]

    return perfil in perfis_permitidos


def financeiro_nf_motorista_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not usuario_pode_analisar_nf_motorista():
            flash('Acesso restrito ao painel financeiro de NFs dos motoristas.', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)

    return decorated_function


def registrar_historico_nf_motorista(empresa_id, motorista_nf_id, usuario_id, status_anterior, status_novo, motivo=None,
                                     observacao=None):
    con_hist = obter_conexao()
    if con_hist is None:
        return
    cur_hist = con_hist.cursor()
    try:
        obs_final = f'NF Motorista ID {motorista_nf_id}.'
        if observacao:
            obs_final += f' {observacao}'
        cur_hist.execute('''
                         INSERT INTO historico_operacoes (empresa_id,
                                                          tipo_operacao,
                                                          usuario_id,
                                                          status_anterior,
                                                          status_novo,
                                                          motivo,
                                                          observacao)
                         VALUES (%s, %s, %s, %s, %s, %s, %s)
                         ''', (
                             empresa_id,
                             'NF_MOTORISTA',
                             usuario_id,
                             status_anterior,
                             status_novo,
                             motivo,
                             obs_final
                         ))
        con_hist.commit()
    except Exception as e:
        con_hist.rollback()
        print(f'Aviso: não foi possível registrar histórico da NF do motorista {motorista_nf_id}: {e}')
    finally:
        cur_hist.close()
        con_hist.close()


@app.route('/financeiro/nfs-motoristas', methods=['GET'])
@login_required
@financeiro_nf_motorista_required
def financeiro_nfs_motoristas():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = int(session.get('is_super_admin') or 0) == 1

    if not empresa_logada_id:
        flash('Empresa não identificada na sessão. Faça login novamente.', 'danger')
        return redirect(url_for('logout'))

    status_nf = request.args.get('status_nf', '').strip()
    motorista_id = request.args.get('motorista_id', '').strip()
    numero_nf = request.args.get('numero_nf', '').strip()
    data_inicio = request.args.get('data_inicio', '').strip()
    data_fim = request.args.get('data_fim', '').strip()
    empresa_id_filtro = request.args.get('empresa_id', '').strip()

    status_validos = ['Enviada', 'Em análise', 'Aprovada', 'Pagamento solicitado', 'Recusada', 'Pagamento confirmado', 'Estornada']

    con = obter_conexao()
    if con is None:
        flash('Erro de conexão com o banco de dados.', 'danger')
        return redirect(url_for('dashboard'))

    cur = con.cursor(dictionary=True)

    try:
        query = '''
                SELECT nf.id,
                       nf.empresa_id,
                       e.nome_fantasia                                          AS empresa_nome,
                       e.razao_social                                           AS empresa_razao_social,
                       nf.motorista_id,
                       mot.nome_completo                                        AS motorista_nome,
                       mot.cpf_cnpj                                             AS motorista_cpf_cnpj,
                       nf.numero_nf,
                       COALESCE(nf.tipo_documento_pagamento, 'XML') AS tipo_documento_pagamento,
                       nf.chave_acesso,
                       nf.data_emissao,
                       nf.valor_total,
                       nf.prestador_cpf_cnpj,
                       nf.tomador_cpf_cnpj,
                       nf.status_nf,
                       nf.nome_arquivo_xml,
                       nf.data_envio,
                       nf.data_aprovacao,
                       nf.data_pagamento,
                       nf.data_recusa,
                       nf.motivo_recusa,
                       nf.observacao,
                       (SELECT t.id
                        FROM titulos_financeiros t
                        WHERE t.empresa_id = nf.empresa_id
                          AND t.origem = CASE
                              WHEN COALESCE(nf.tipo_documento_pagamento, 'XML') = 'SEM_NF' THEN 'SEM_NF_MOTORISTA'
                              ELSE 'NF_MOTORISTA'
                          END
                          AND t.origem_id = nf.id
                          AND t.status_titulo NOT IN ('Cancelado', 'Estornado')
                        ORDER BY t.id DESC
                        LIMIT 1) AS titulo_financeiro_id,
                       (SELECT t.status_titulo
                        FROM titulos_financeiros t
                        WHERE t.empresa_id = nf.empresa_id
                          AND t.origem = CASE
                              WHEN COALESCE(nf.tipo_documento_pagamento, 'XML') = 'SEM_NF' THEN 'SEM_NF_MOTORISTA'
                              ELSE 'NF_MOTORISTA'
                          END
                          AND t.origem_id = nf.id
                          AND t.status_titulo NOT IN ('Cancelado', 'Estornado')
                        ORDER BY t.id DESC
                        LIMIT 1) AS titulo_financeiro_status,
                       COUNT(v.id)                                              AS qtd_rotas,
                       GROUP_CONCAT(r.identi_rota ORDER BY r.id SEPARATOR ', ') AS rotas_vinculadas
                FROM motorista_notas_fiscais nf
                         INNER JOIN empresas e ON e.id = nf.empresa_id
                         INNER JOIN pessoas mot ON mot.id = nf.motorista_id AND mot.empresa_id = nf.empresa_id
                         LEFT JOIN motorista_nf_rotas v ON v.motorista_nf_id = nf.id AND v.empresa_id = nf.empresa_id
                         LEFT JOIN rotas r ON r.id = v.rota_id AND r.empresa_id = nf.empresa_id
                WHERE 1 = 1 \
                '''
        params = []

        if is_super_admin:
            if empresa_id_filtro and empresa_id_filtro.isdigit():
                query += ' AND nf.empresa_id = %s'
                params.append(int(empresa_id_filtro))
        else:
            query += ' AND nf.empresa_id = %s'
            params.append(empresa_logada_id)

        if status_nf in status_validos:
            query += ' AND nf.status_nf = %s'
            params.append(status_nf)

        if motorista_id and motorista_id.isdigit():
            query += ' AND nf.motorista_id = %s'
            params.append(int(motorista_id))

        if numero_nf:
            query += ' AND nf.numero_nf LIKE %s'
            params.append(f'%{numero_nf}%')

        if data_inicio:
            query += ' AND DATE(nf.data_envio) >= %s'
            params.append(data_inicio)

        if data_fim:
            query += ' AND DATE(nf.data_envio) <= %s'
            params.append(data_fim)

        query += '''
            GROUP BY
                nf.id, nf.empresa_id, e.nome_fantasia, e.razao_social,
                nf.motorista_id, mot.nome_completo, mot.cpf_cnpj,
                nf.numero_nf, nf.chave_acesso, nf.data_emissao, nf.valor_total,
                nf.valor_bruto,
                nf.valor_liquido,
                nf.prestador_cpf_cnpj, nf.tomador_cpf_cnpj, nf.status_nf,
                nf.nome_arquivo_xml, nf.data_envio, nf.data_aprovacao,
                nf.data_pagamento, nf.data_recusa, nf.motivo_recusa, nf.observacao
            ORDER BY nf.data_envio DESC, nf.id DESC
        '''

        cur.execute(query, params)
        notas = cur.fetchall()

        query_resumo = '''
                       SELECT COALESCE(SUM(CASE WHEN status_nf = 'Enviada' THEN 1 ELSE 0 END), 0)              AS enviadas,
                              COALESCE(SUM(CASE WHEN status_nf = 'Em análise' THEN 1 ELSE 0 END), 0)           AS em_analise,
                              COALESCE(SUM(CASE WHEN status_nf = 'Aprovada' THEN 1 ELSE 0 END), 0)             AS aprovadas,
                              COALESCE(SUM(CASE WHEN status_nf = 'Pagamento solicitado' THEN 1 ELSE 0 END), 0) AS pagamento_solicitado,
                              COALESCE(SUM(CASE WHEN status_nf = 'Recusada' THEN 1 ELSE 0 END), 0)             AS recusadas,
                              COALESCE(SUM(CASE WHEN status_nf = 'Pagamento confirmado' THEN 1 ELSE 0 END), 0) AS pagas,
                              COALESCE(SUM(valor_total), 0)                                                    AS valor_total
                       FROM motorista_notas_fiscais
                       WHERE 1 = 1 \
                       '''
        params_resumo = []

        if is_super_admin:
            if empresa_id_filtro and empresa_id_filtro.isdigit():
                query_resumo += ' AND empresa_id = %s'
                params_resumo.append(int(empresa_id_filtro))
        else:
            query_resumo += ' AND empresa_id = %s'
            params_resumo.append(empresa_logada_id)

        cur.execute(query_resumo, params_resumo)
        resumo = cur.fetchone() or {}

        query_motoristas = '''
                           SELECT id, nome_completo
                           FROM pessoas
                           WHERE tipo_cadastro = 'Motorista'
                             AND status_cadastro = 'Ativo' \
                           '''
        params_motoristas = []

        if is_super_admin:
            if empresa_id_filtro and empresa_id_filtro.isdigit():
                query_motoristas += ' AND empresa_id = %s'
                params_motoristas.append(int(empresa_id_filtro))
        else:
            query_motoristas += ' AND empresa_id = %s'
            params_motoristas.append(empresa_logada_id)

        query_motoristas += ' ORDER BY nome_completo ASC'
        cur.execute(query_motoristas, params_motoristas)
        motoristas = cur.fetchall()

        empresas = []
        if is_super_admin:
            cur.execute('''
                        SELECT id, razao_social, nome_fantasia
                        FROM empresas
                        ORDER BY nome_fantasia ASC, razao_social ASC
                        ''')
            empresas = cur.fetchall()

    except Exception as e:
        print(f'Erro ao carregar painel financeiro de NFs motoristas: {e}')
        flash('Erro técnico ao carregar painel financeiro das NFs dos motoristas.', 'danger')
        notas = []
        motoristas = []
        empresas = []
        resumo = {'enviadas': 0, 'em_analise': 0, 'aprovadas': 0, 'pagamento_solicitado': 0, 'recusadas': 0, 'pagas': 0, 'valor_total': 0}

    finally:
        cur.close()
        con.close()

    filtros = {
        'status_nf': status_nf,
        'motorista_id': motorista_id,
        'numero_nf': numero_nf,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'empresa_id': empresa_id_filtro
    }

    return render_template(
        'financeiro_nfs_motoristas.html',
        usuario_logado=usuario_logado,
        notas=notas,
        motoristas=motoristas,
        empresas=empresas,
        resumo=resumo,
        filtros=filtros,
        is_super_admin=is_super_admin
    )


@app.route('/financeiro/nfs-motoristas/<int:id>', methods=['GET'])
@login_required
@financeiro_nf_motorista_required
def detalhes_nf_motorista(id):
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = int(session.get('is_super_admin') or 0) == 1

    if not empresa_logada_id:
        flash('Empresa não identificada na sessão. Faça login novamente.', 'danger')
        return redirect(url_for('logout'))

    # Bloco 10.5: persistência dos últimos filtros do relatório financeiro.
    # Ao voltar para a tela sem parâmetros, reaplica automaticamente a última consulta usada.
    if not request.args and session.get('relatorios_financeiro_filtros'):
        return redirect(url_for('relatorios_financeiro', **session.get('relatorios_financeiro_filtros')))

    con = obter_conexao()
    if con is None:
        flash('Erro de conexão com o banco de dados.', 'danger')
        return redirect(url_for('financeiro_nfs_motoristas'))

    cur = con.cursor(dictionary=True)

    try:
        query_nf = '''
                   SELECT nf.*,
                          e.nome_fantasia         AS empresa_nome,
                          e.razao_social          AS empresa_razao_social,
                          mot.nome_completo       AS motorista_nome,
                          mot.cpf_cnpj            AS motorista_cpf_cnpj,
                          usu_aprov.nome_completo AS usuario_aprovacao_nome,
                          usu_pag.nome_completo   AS usuario_pagamento_nome,
                          usu_rec.nome_completo   AS usuario_recusa_nome,
                          (SELECT t.id
                           FROM titulos_financeiros t
                           WHERE t.empresa_id = nf.empresa_id
                             AND t.origem = CASE
                                 WHEN COALESCE(nf.tipo_documento_pagamento, 'XML') = 'SEM_NF' THEN 'SEM_NF_MOTORISTA'
                                 ELSE 'NF_MOTORISTA'
                             END
                             AND t.origem_id = nf.id
                             AND t.status_titulo NOT IN ('Cancelado', 'Estornado')
                           ORDER BY t.id DESC
                           LIMIT 1) AS titulo_financeiro_id,
                          (SELECT t.status_titulo
                           FROM titulos_financeiros t
                           WHERE t.empresa_id = nf.empresa_id
                             AND t.origem = CASE
                                 WHEN COALESCE(nf.tipo_documento_pagamento, 'XML') = 'SEM_NF' THEN 'SEM_NF_MOTORISTA'
                                 ELSE 'NF_MOTORISTA'
                             END
                             AND t.origem_id = nf.id
                             AND t.status_titulo NOT IN ('Cancelado', 'Estornado')
                           ORDER BY t.id DESC
                           LIMIT 1) AS titulo_financeiro_status
                   FROM motorista_notas_fiscais nf
                            INNER JOIN empresas e ON e.id = nf.empresa_id
                            INNER JOIN pessoas mot ON mot.id = nf.motorista_id AND mot.empresa_id = nf.empresa_id
                            LEFT JOIN usuarios u_aprov ON u_aprov.id = nf.usuario_aprovacao_id
                            LEFT JOIN pessoas usu_aprov ON usu_aprov.id = u_aprov.pessoa_id
                            LEFT JOIN usuarios u_pag ON u_pag.id = nf.usuario_pagamento_id
                            LEFT JOIN pessoas usu_pag ON usu_pag.id = u_pag.pessoa_id
                            LEFT JOIN usuarios u_rec ON u_rec.id = nf.usuario_recusa_id
                            LEFT JOIN pessoas usu_rec ON usu_rec.id = u_rec.pessoa_id
                   WHERE nf.id = %s \
                   '''
        params_nf = [id]

        if not is_super_admin:
            query_nf += ' AND nf.empresa_id = %s'
            params_nf.append(empresa_logada_id)

        query_nf += ' LIMIT 1'
        cur.execute(query_nf, params_nf)
        nf = cur.fetchone()

        if not nf:
            flash('Documento do motorista não encontrado ou não pertence à empresa logada.', 'danger')
            return redirect(url_for('financeiro_nfs_motoristas'))

        cur.execute('''
                    SELECT MIN(v.id)            AS id,
                           v.rota_id,
                           MAX(v.valor_rota)    AS valor_rota,
                           MAX(v.data_vinculo)  AS data_vinculo,
                           r.data_lancamento,
                           r.identi_rota,
                           r.tipo_rota,
                           r.situacao_rota,
                           r.status_motorista,
                           transp.nome_completo AS transportadora_nome,
                           transp.cpf_cnpj      AS transportadora_cpf_cnpj
                    FROM motorista_nf_rotas v
                             INNER JOIN rotas r ON r.id = v.rota_id AND r.empresa_id = v.empresa_id
                             LEFT JOIN pessoas transp
                                       ON transp.id = r.transportadora_id AND transp.empresa_id = r.empresa_id
                    WHERE v.motorista_nf_id = %s
                      AND v.empresa_id = %s
                    GROUP BY v.rota_id,
                             r.data_lancamento,
                             r.identi_rota,
                             r.tipo_rota,
                             r.situacao_rota,
                             r.status_motorista,
                             transp.nome_completo,
                             transp.cpf_cnpj
                    ORDER BY r.id ASC
                    ''', (id, nf['empresa_id']))
        rotas = cur.fetchall()

    except Exception as e:
        print(f'Erro ao carregar detalhes da NF do motorista: {e}')
        flash('Erro técnico ao carregar detalhes da NF.', 'danger')
        return redirect(url_for('financeiro_nfs_motoristas'))

    finally:
        cur.close()
        con.close()

    return render_template('detalhes_nf_motorista.html', usuario_logado=usuario_logado, nf=nf, rotas=rotas)


@app.route('/financeiro/nfs-motoristas/<int:id>/marcar-analise', methods=['POST'])
@login_required
@financeiro_nf_motorista_required
def marcar_nf_motorista_em_analise(id):
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = int(session.get('is_super_admin') or 0) == 1

    if not empresa_logada_id:
        flash('Empresa não identificada na sessão. Faça login novamente.', 'danger')
        return redirect(url_for('logout'))

    con = obter_conexao()
    if con is None:
        flash('Erro de conexão com o banco de dados.', 'danger')
        return redirect(url_for('financeiro_nfs_motoristas'))

    cur = con.cursor(dictionary=True)

    try:
        query = '''
                SELECT id, empresa_id, status_nf, numero_nf
                FROM motorista_notas_fiscais
                WHERE id = %s \
                '''
        params = [id]

        if not is_super_admin:
            query += ' AND empresa_id = %s'
            params.append(empresa_logada_id)

        query += ' LIMIT 1'
        cur.execute(query, params)
        nf = cur.fetchone()

        if not nf:
            flash('NF não encontrada.', 'danger')
            return redirect(url_for('financeiro_nfs_motoristas'))

        status_atual = nf.get('status_nf')

        if status_atual != 'Enviada':
            flash(f'Esta NF não pode ir para análise agora. Status atual: {status_atual}.', 'warning')
            return redirect(url_for('detalhes_nf_motorista', id=id))

        cur.execute('''
                    UPDATE motorista_notas_fiscais
                    SET status_nf = 'Em análise'
                    WHERE id = %s
                    ''', (id,))

        cur.execute('''
                    SELECT rota_id
                    FROM motorista_nf_rotas
                    WHERE motorista_nf_id = %s
                      AND empresa_id = %s
                    ''', (id, nf['empresa_id']))
        rotas = cur.fetchall()

        for item in rotas:
            cur.execute('''
                        UPDATE rotas
                        SET status_motorista = 'Em análise'
                        WHERE id = %s
                          AND empresa_id = %s
                          AND status_motorista = 'NF enviada'
                        ''', (item['rota_id'], nf['empresa_id']))

        con.commit()

        registrar_historico_nf_motorista(
            empresa_id=nf['empresa_id'],
            motorista_nf_id=id,
            usuario_id=usuario_id,
            status_anterior=status_atual,
            status_novo='Em análise',
            motivo='Análise iniciada',
            observacao=f"NF {nf['numero_nf']} marcada como Em análise."
        )

        for item in rotas:
            registrar_historico_rota_motorista(
                empresa_id=nf['empresa_id'],
                rota_id=item['rota_id'],
                usuario_id=usuario_id,
                status_anterior='NF enviada',
                status_novo='Em análise',
                motivo='Análise financeira iniciada',
                observacao=f"NF motorista {nf['numero_nf']} em análise."
            )

        flash(f"NF {nf['numero_nf']} marcada como Em análise.", 'success')

    except Exception as e:
        con.rollback()
        print(f'Erro ao marcar NF em análise: {e}')
        flash('Erro técnico ao marcar NF em análise.', 'danger')

    finally:
        cur.close()
        con.close()

    return redirect(url_for('detalhes_nf_motorista', id=id))


def gerar_identificador_pagamento_sem_nf(empresa_id, motorista_id):
    agora = datetime.now().strftime('%Y%m%d%H%M%S')
    return f"SEM-NF-{empresa_id}-{motorista_id}-{agora}"


@app.route('/portal-motorista/solicitar-pagamento-sem-nf', methods=['GET', 'POST'])
@login_required
@motorista_required
def solicitar_pagamento_sem_nf_motorista():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_id = session.get('empresa_id')
    motorista = buscar_motorista_logado()

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    if not motorista:
        flash("Seu usuário não está vinculado a um motorista ativo nesta empresa.", "danger")
        return redirect(url_for('portal_motorista'))

    motorista_id = motorista['id']
    rota_id_preselecionada = request.args.get('rota_id', '').strip()

    if request.method == 'POST':
        rota_ids_raw = request.form.getlist('rota_ids')
        observacao = request.form.get('observacao', '').strip()
        rota_ids = []

        for item in rota_ids_raw:
            if str(item).isdigit():
                rota_ids.append(int(item))

        rota_ids = list(dict.fromkeys(rota_ids))

        if not rota_ids:
            flash("Selecione pelo menos uma rota liberada para solicitar pagamento.", "warning")
            return redirect(url_for('solicitar_pagamento_sem_nf_motorista'))

        con = obter_conexao()
        if con is None:
            flash("Erro de conexão com o banco de dados.", "danger")
            return redirect(url_for('solicitar_pagamento_sem_nf_motorista'))

        cur = con.cursor(dictionary=True)

        try:
            placeholders = ", ".join(["%s"] * len(rota_ids))

            cur.execute(f"""
                SELECT
                    r.id,
                    r.empresa_id,
                    r.identi_rota,
                    r.motorista_id,
                    r.transportadora_id,
                    r.situacao_rota,
                    COALESCE(r.status_motorista, 'Aguardando conferência') AS status_motorista,
                    (COALESCE(r.valor_rota, 0) + COALESCE(r.valor_km, 0) + COALESCE(r.outras_despesas, 0)) AS valor_total_rota,
                    COALESCE(transp.nome_completo, emp.razao_social, emp.nome_fantasia) AS transportadora_nome,
                    COALESCE(transp.cpf_cnpj, emp.cnpj) AS transportadora_cpf_cnpj
                FROM rotas r
                LEFT JOIN pessoas transp
                    ON transp.id = r.transportadora_id
                   AND transp.empresa_id = r.empresa_id
                LEFT JOIN empresas emp
                    ON emp.id = r.empresa_id
                WHERE r.id IN ({placeholders})
                  AND r.empresa_id = %s
                  AND r.motorista_id = %s
                  AND COALESCE(r.status_motorista, 'Aguardando conferência') = 'Liberada para NF'
            """, rota_ids + [empresa_id, motorista_id])

            rotas = cur.fetchall()

            if len(rotas) != len(rota_ids):
                flash("Uma ou mais rotas selecionadas não estão disponíveis para solicitação de pagamento.", "danger")
                return redirect(url_for('solicitar_pagamento_sem_nf_motorista'))

            cur.execute(f"""
                SELECT rota_id
                FROM motorista_nf_rotas
                WHERE rota_id IN ({placeholders})
            """, rota_ids)

            if cur.fetchall():
                flash("Uma ou mais rotas selecionadas já possuem documento/solicitação vinculada.", "danger")
                return redirect(url_for('solicitar_pagamento_sem_nf_motorista'))

            soma_rotas = Decimal('0.00')
            for rota in rotas:
                soma_rotas += converter_decimal(rota.get('valor_total_rota'))
            soma_rotas = soma_rotas.quantize(Decimal('0.01'))

            identificador = gerar_identificador_pagamento_sem_nf(empresa_id, motorista_id)

            cur.execute("""
                        INSERT INTO motorista_notas_fiscais (empresa_id,
                                                             motorista_id,
                                                             tipo_documento_pagamento,
                                                             numero_nf,
                                                             chave_acesso,
                                                             data_emissao,
                                                             valor_total,
                                                             valor_bruto,
                                                             valor_liquido,
                                                             prestador_cpf_cnpj,
                                                             tomador_cpf_cnpj,
                                                             status_nf,
                                                             nome_arquivo_xml,
                                                             observacao)
                        VALUES (%s, %s, 'SEM_NF', %s, %s, CURDATE(), %s, %s, %s, %s, NULL, 'Enviada', NULL, %s)
                        """, (
                            empresa_id,
                            motorista_id,
                            identificador,
                            identificador,
                            soma_rotas,
                            soma_rotas,
                            soma_rotas,
                            somente_digitos(motorista.get('cpf_cnpj')),
                            observacao or 'Solicitação de pagamento sem NF/Pessoa Física.'
                        ))

            documento_id = cur.lastrowid

            # Remove vínculos antigos deste documento, caso ele esteja sendo reaproveitado futuramente.
            cur.execute("""
                        DELETE
                        FROM motorista_nf_rotas
                        WHERE motorista_nf_id = %s
                          AND empresa_id = %s
                        """, (documento_id, empresa_id))

            # Remove vínculos antigos de documentos recusados para permitir nova solicitação da mesma rota.
            cur.execute(f"""
                DELETE v
                FROM motorista_nf_rotas v
                INNER JOIN motorista_notas_fiscais nf
                    ON nf.id = v.motorista_nf_id
                   AND nf.empresa_id = v.empresa_id
                WHERE v.rota_id IN ({placeholders})
                  AND v.empresa_id = %s
                  AND nf.status_nf = 'Recusada'
            """, rota_ids + [empresa_id])

            for rota in rotas:
                cur.execute("""
                            INSERT INTO motorista_nf_rotas (empresa_id,
                                                            motorista_nf_id,
                                                            rota_id,
                                                            valor_rota)
                            VALUES (%s, %s, %s, %s)
                            """, (
                                empresa_id,
                                documento_id,
                                rota['id'],
                                converter_decimal(rota.get('valor_total_rota'))
                            ))

                cur.execute("""
                            UPDATE rotas
                            SET status_motorista = 'NF enviada'
                            WHERE id = %s
                              AND empresa_id = %s
                              AND motorista_id = %s
                              AND COALESCE(status_motorista, 'Aguardando liberação') = 'Liberada para NF'
                            """, (rota['id'], empresa_id, motorista_id))

            con.commit()

            for rota in rotas:
                registrar_historico_rota_motorista(
                    empresa_id=empresa_id,
                    rota_id=rota['id'],
                    usuario_id=session.get('usuario_id'),
                    status_anterior='Liberada para NF',
                    status_novo='NF enviada',
                    motivo='Solicitação de pagamento sem NF',
                    observacao=f"Solicitação sem NF {identificador} enviada pelo Portal do Motorista."
                )

            flash(
                f"Solicitação de pagamento enviada com sucesso. {len(rotas)} rota(s) vinculada(s), total {moeda_br(soma_rotas)}.",
                "success"
            )
            return redirect(url_for('portal_motorista'))

        except Exception as e:
            con.rollback()
            print(f"Erro na solicitação de pagamento sem NF: {e}")
            flash(f"Erro técnico ao solicitar pagamento sem NF: {e}", "danger")
            return redirect(url_for('solicitar_pagamento_sem_nf_motorista'))

        finally:
            cur.close()
            con.close()

    rotas_liberadas = carregar_rotas_liberadas_motorista(empresa_id, motorista_id)

    return render_template(
        'solicitar_pagamento_sem_nf_motorista.html',
        usuario_logado=usuario_logado,
        motorista=motorista,
        rotas=rotas_liberadas,
        rota_id_preselecionada=rota_id_preselecionada
    )


# ==========================================================
# EXECUÇÃO LOCAL
# ==========================================================

# ==========================================================
# FASE 3.5.1 - REVERTER APROVAÇÃO DO DOCUMENTO DO MOTORISTA
# ==========================================================
@app.route('/financeiro/nfs-motoristas/<int:id>/reverter-aprovacao', methods=['POST'])
@login_required
@financeiro_nf_motorista_required
def reverter_aprovacao_documento_motorista(id):
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = int(session.get('is_super_admin') or 0) == 1

    motivo_reversao = request.form.get('motivo_reversao', '').strip()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    if len(motivo_reversao) < 3:
        flash("Informe um motivo válido para reverter a aprovação.", "warning")
        return redirect(url_for('detalhes_nf_motorista', id=id))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('financeiro_nfs_motoristas'))

    cur = con.cursor(dictionary=True)

    try:
        query = """
                SELECT id,
                       empresa_id,
                       motorista_id,
                       tipo_documento_pagamento,
                       numero_nf,
                       status_nf,
                       data_pagamento,
                       valor_total
                FROM motorista_notas_fiscais
                WHERE id = %s \
                """

        params = [id]

        if not is_super_admin:
            query += " AND empresa_id = %s"
            params.append(empresa_logada_id)

        query += " LIMIT 1"

        cur.execute(query, params)
        documento = cur.fetchone()

        if not documento:
            flash("Documento não encontrado ou não pertence à empresa logada.", "danger")
            return redirect(url_for('financeiro_nfs_motoristas'))

        status_atual = documento.get('status_nf')

        if status_atual != 'Aprovada':
            flash(
                f"Somente documentos aprovados podem ter a aprovação revertida. Status atual: {status_atual}.",
                "warning"
            )
            return redirect(url_for('detalhes_nf_motorista', id=id))

        if documento.get('data_pagamento'):
            flash(
                "Este documento já possui pagamento registrado. Para corrigir, será necessário um fluxo de estorno/cancelamento de pagamento.",
                "danger"
            )
            return redirect(url_for('detalhes_nf_motorista', id=id))

        cur.execute("""
                    SELECT v.rota_id,
                           r.status_motorista,
                           r.identi_rota
                    FROM motorista_nf_rotas v
                             INNER JOIN rotas r
                                        ON r.id = v.rota_id
                                            AND r.empresa_id = v.empresa_id
                    WHERE v.motorista_nf_id = %s
                      AND v.empresa_id = %s
                    """, (id, documento['empresa_id']))

        rotas_vinculadas = cur.fetchall()

        if not rotas_vinculadas:
            flash("Este documento não possui rotas vinculadas.", "danger")
            return redirect(url_for('detalhes_nf_motorista', id=id))

        cur.execute("""
                    UPDATE motorista_notas_fiscais
                    SET status_nf            = 'Em análise',
                        data_aprovacao       = NULL,
                        usuario_aprovacao_id = NULL,
                        observacao           = CONCAT(
                                COALESCE(observacao, ''),
                                CASE WHEN COALESCE(observacao, '') = '' THEN '' ELSE '\\n' END,
                                'Aprovação revertida em ',
                                DATE_FORMAT(NOW(), '%d/%m/%Y %H:%i'),
                                '. Motivo: ',
                                %s
                                               )
                    WHERE id = %s
                      AND empresa_id = %s
                      AND status_nf = 'Aprovada'
                    """, (motivo_reversao, id, documento['empresa_id']))

        for rota in rotas_vinculadas:
            cur.execute("""
                        UPDATE rotas
                        SET status_motorista = 'Em análise'
                        WHERE id = %s
                          AND empresa_id = %s
                          AND COALESCE(status_motorista, 'Aguardando liberação') = 'Aprovada para pagamento'
                        """, (rota['rota_id'], documento['empresa_id']))

        con.commit()

        registrar_historico_nf_motorista(
            empresa_id=documento['empresa_id'],
            motorista_nf_id=id,
            usuario_id=usuario_id,
            status_anterior='Aprovada',
            status_novo='Em análise',
            motivo=motivo_reversao,
            observacao=f"Aprovação do documento {documento['numero_nf']} revertida."
        )

        for rota in rotas_vinculadas:
            registrar_historico_rota_motorista(
                empresa_id=documento['empresa_id'],
                rota_id=rota['rota_id'],
                usuario_id=usuario_id,
                status_anterior=rota.get('status_motorista'),
                status_novo='Em análise',
                motivo='Reversão de aprovação do documento',
                observacao=f"Documento {documento['numero_nf']} teve aprovação revertida. Motivo: {motivo_reversao}"
            )

        flash(
            f"Aprovação do documento {documento['numero_nf']} revertida com sucesso. "
            "O documento voltou para Em análise.",
            "success"
        )
        return redirect(url_for('detalhes_nf_motorista', id=id))

    except Exception as e:
        con.rollback()
        print(f"Erro ao reverter aprovação do documento do motorista: {e}")
        flash(f"Erro técnico ao reverter aprovação: {e}", "danger")
        return redirect(url_for('detalhes_nf_motorista', id=id))

    finally:
        cur.close()
        con.close()


# ==========================================================
# FASE 3.6 - CONFIRMAR / ESTORNAR PAGAMENTO DO DOCUMENTO DO MOTORISTA
# ==========================================================
@app.route('/financeiro/nfs-motoristas/<int:id>/confirmar-pagamento', methods=['POST'])
@login_required
@financeiro_nf_motorista_required
def confirmar_pagamento_documento_motorista(id):
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = int(session.get('is_super_admin') or 0) == 1

    data_pagamento_form = request.form.get('data_pagamento', '').strip()
    observacao_pagamento = request.form.get('observacao_pagamento', '').strip()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    if not data_pagamento_form:
        flash("Informe a data do pagamento.", "warning")
        return redirect(url_for('detalhes_nf_motorista', id=id))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('financeiro_nfs_motoristas'))

    cur = con.cursor(dictionary=True)

    try:
        query = """
                SELECT id,
                       empresa_id,
                       motorista_id,
                       tipo_documento_pagamento,
                       numero_nf,
                       status_nf,
                       valor_total
                FROM motorista_notas_fiscais
                WHERE id = %s \
                """

        params = [id]

        if not is_super_admin:
            query += " AND empresa_id = %s"
            params.append(empresa_logada_id)

        query += " LIMIT 1"

        cur.execute(query, params)
        documento = cur.fetchone()

        if not documento:
            flash("Documento não encontrado ou não pertence à empresa logada.", "danger")
            return redirect(url_for('financeiro_nfs_motoristas'))

        status_atual = documento.get('status_nf')

        if status_atual != 'Aprovada':
            flash(
                f"Somente documentos aprovados podem ter pagamento confirmado. Status atual: {status_atual}.",
                "warning"
            )
            return redirect(url_for('detalhes_nf_motorista', id=id))

        cur.execute("""
                    SELECT v.rota_id,
                           r.status_motorista,
                           r.situacao_rota,
                           r.identi_rota
                    FROM motorista_nf_rotas v
                             INNER JOIN rotas r
                                        ON r.id = v.rota_id
                                            AND r.empresa_id = v.empresa_id
                    WHERE v.motorista_nf_id = %s
                      AND v.empresa_id = %s
                    """, (id, documento['empresa_id']))

        rotas_vinculadas = cur.fetchall()

        if not rotas_vinculadas:
            flash("Este documento não possui rotas vinculadas.", "danger")
            return redirect(url_for('detalhes_nf_motorista', id=id))

        observacao_final = (
            f"Pagamento confirmado em {datetime.now().strftime('%d/%m/%Y %H:%M')}. "
            f"Data do pagamento: {data_pagamento_form}."
        )

        if observacao_pagamento:
            observacao_final += f" Observação: {observacao_pagamento}"

        cur.execute("""
                    UPDATE motorista_notas_fiscais
                    SET status_nf            = 'Pagamento confirmado',
                        data_pagamento       = %s,
                        usuario_pagamento_id = %s,
                        observacao           = CONCAT(
                                COALESCE(observacao, ''),
                                CASE WHEN COALESCE(observacao, '') = '' THEN '' ELSE '\n' END,
                                %s
                                               )
                    WHERE id = %s
                      AND empresa_id = %s
                      AND status_nf = 'Aprovada'
                    """, (
                        data_pagamento_form,
                        usuario_id,
                        observacao_final,
                        id,
                        documento['empresa_id']
                    ))

        for rota in rotas_vinculadas:
            cur.execute("""
                        UPDATE rotas
                        SET status_motorista = 'Pagamento confirmado',
                            situacao_rota    = 'Quitado'
                        WHERE id = %s
                          AND empresa_id = %s
                          AND COALESCE(status_motorista, 'Aguardando liberação') = 'Aprovada para pagamento'
                        """, (rota['rota_id'], documento['empresa_id']))

        con.commit()

        registrar_historico_nf_motorista(
            empresa_id=documento['empresa_id'],
            motorista_nf_id=id,
            usuario_id=usuario_id,
            status_anterior='Aprovada',
            status_novo='Pagamento confirmado',
            motivo='Pagamento confirmado pelo financeiro',
            observacao=f"Documento {documento['numero_nf']} pago em {data_pagamento_form}."
        )

        for rota in rotas_vinculadas:
            registrar_historico_rota_motorista(
                empresa_id=documento['empresa_id'],
                rota_id=rota['rota_id'],
                usuario_id=usuario_id,
                status_anterior=rota.get('status_motorista'),
                status_novo='Pagamento confirmado',
                motivo='Pagamento do documento confirmado',
                observacao=f"Pagamento confirmado para o documento {documento['numero_nf']}. Situação da rota alterada para Quitado."
            )

        flash(f"Pagamento do documento {documento['numero_nf']} confirmado com sucesso.", "success")
        return redirect(url_for('detalhes_nf_motorista', id=id))

    except Exception as e:
        con.rollback()
        print(f"Erro ao confirmar pagamento do documento do motorista: {e}")
        flash(f"Erro técnico ao confirmar pagamento: {e}", "danger")
        return redirect(url_for('detalhes_nf_motorista', id=id))

    finally:
        cur.close()
        con.close()


@app.route('/financeiro/nfs-motoristas/<int:id>/estornar-pagamento', methods=['POST'])
@login_required
@financeiro_nf_motorista_required
def estornar_pagamento_documento_motorista(id):
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = int(session.get('is_super_admin') or 0) == 1

    motivo_estorno = request.form.get('motivo_estorno_pagamento', '').strip()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    if len(motivo_estorno) < 3:
        flash("Informe um motivo válido para estornar o pagamento.", "warning")
        return redirect(url_for('detalhes_nf_motorista', id=id))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('financeiro_nfs_motoristas'))

    cur = con.cursor(dictionary=True)

    try:
        query = """
                SELECT id,
                       empresa_id,
                       motorista_id,
                       tipo_documento_pagamento,
                       numero_nf,
                       status_nf,
                       valor_total,
                       data_pagamento
                FROM motorista_notas_fiscais
                WHERE id = %s \
                """

        params = [id]

        if not is_super_admin:
            query += " AND empresa_id = %s"
            params.append(empresa_logada_id)

        query += " LIMIT 1"

        cur.execute(query, params)
        documento = cur.fetchone()

        if not documento:
            flash("Documento não encontrado ou não pertence à empresa logada.", "danger")
            return redirect(url_for('financeiro_nfs_motoristas'))

        status_atual = documento.get('status_nf')

        if status_atual != 'Pagamento confirmado':
            flash(
                f"Somente documentos com pagamento confirmado podem ser estornados. Status atual: {status_atual}.",
                "warning"
            )
            return redirect(url_for('detalhes_nf_motorista', id=id))

        cur.execute("""
                    SELECT v.rota_id,
                           r.status_motorista,
                           r.situacao_rota,
                           r.identi_rota
                    FROM motorista_nf_rotas v
                             INNER JOIN rotas r
                                        ON r.id = v.rota_id
                                            AND r.empresa_id = v.empresa_id
                    WHERE v.motorista_nf_id = %s
                      AND v.empresa_id = %s
                    """, (id, documento['empresa_id']))

        rotas_vinculadas = cur.fetchall()

        if not rotas_vinculadas:
            flash("Este documento não possui rotas vinculadas.", "danger")
            return redirect(url_for('detalhes_nf_motorista', id=id))

        observacao_estorno = (
            f"Pagamento estornado em {datetime.now().strftime('%d/%m/%Y %H:%M')}. "
            f"Motivo: {motivo_estorno}. "
            "Rotas vinculadas canceladas para novo processo operacional."
        )

        cur.execute("""
                    UPDATE motorista_notas_fiscais
                    SET status_nf                    = 'Estornada',
                        data_estorno_pagamento       = NOW(),
                        motivo_estorno_pagamento     = %s,
                        usuario_estorno_pagamento_id = %s,
                        observacao                   = CONCAT(
                                COALESCE(observacao, ''),
                                CASE WHEN COALESCE(observacao, '') = '' THEN '' ELSE '\n' END,
                                %s
                                                       )
                    WHERE id = %s
                      AND empresa_id = %s
                      AND status_nf = 'Pagamento confirmado'
                    """, (
                        motivo_estorno,
                        usuario_id,
                        observacao_estorno,
                        id,
                        documento['empresa_id']
                    ))

        for rota in rotas_vinculadas:
            cur.execute("""
                        UPDATE rotas
                        SET status_motorista = 'Cancelada',
                            situacao_rota    = 'Cancelada'
                        WHERE id = %s
                          AND empresa_id = %s
                          AND COALESCE(status_motorista, 'Aguardando liberação') = 'Pagamento confirmado'
                        """, (rota['rota_id'], documento['empresa_id']))

        con.commit()

        registrar_historico_nf_motorista(
            empresa_id=documento['empresa_id'],
            motorista_nf_id=id,
            usuario_id=usuario_id,
            status_anterior='Pagamento confirmado',
            status_novo='Estornada',
            motivo=motivo_estorno,
            observacao=f"Pagamento do documento {documento['numero_nf']} estornado. Rotas canceladas."
        )

        for rota in rotas_vinculadas:
            registrar_historico_rota_motorista(
                empresa_id=documento['empresa_id'],
                rota_id=rota['rota_id'],
                usuario_id=usuario_id,
                status_anterior=rota.get('status_motorista'),
                status_novo='Cancelada',
                motivo='Estorno de pagamento do documento',
                observacao=(
                    f"Pagamento do documento {documento['numero_nf']} estornado. "
                    f"Rota cancelada para novo processo. Motivo: {motivo_estorno}"
                )
            )

        flash(
            f"Pagamento do documento {documento['numero_nf']} estornado. "
            "As rotas vinculadas foram canceladas para o processo ser feito do zero.",
            "success"
        )
        return redirect(url_for('detalhes_nf_motorista', id=id))

    except Exception as e:
        con.rollback()
        print(f"Erro ao estornar pagamento do documento do motorista: {e}")
        flash(f"Erro técnico ao estornar pagamento: {e}", "danger")
        return redirect(url_for('detalhes_nf_motorista', id=id))

    finally:
        cur.close()
        con.close()


# ==========================================================
# FASE 4.1.1 - DISPONIBILIDADE SEMANAL DO MOTORISTA
# ==========================================================
def obter_config_disponibilidade_empresa(empresa_id):
    """
    Busca configurações de disponibilidade por empresa.
    Se não existir configuração, usa padrão seguro:
    - limite 6 dias disponíveis por semana
    - limite de edição até 11:00 do dia anterior
    - permite liberação excepcional futura pelo supervisor
    """
    config_padrao = {
        'horario_limite_edicao': '11:00',
        'limite_dias_disponiveis_semana': 6,
        'permite_liberacao_setimo_dia': 'S',
        'bloquear_disponibilidade_apos_limite': 'S',
        'permitir_liberacao_excepcional': 'S'
    }

    con = obter_conexao()

    if con is None:
        return config_padrao

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT horario_limite_edicao,
                           limite_dias_disponiveis_semana,
                           permite_liberacao_setimo_dia
                    FROM configuracoes_disponibilidade
                    WHERE empresa_id = %s LIMIT 1
                    """, (empresa_id,))

        row = cur.fetchone()

        if not row:
            try:
                params = carregar_parametros_operacionais_motorista_empresa(empresa_id, cur=cur)
                config_padrao['horario_limite_edicao'] = normalizar_time_param(params.get('motorista.horario_limite_disponibilidade', {}).get('valor'), '11:00')
                config_padrao['bloquear_disponibilidade_apos_limite'] = 'S' if parametro_bool(params.get('motorista.bloquear_disponibilidade_apos_limite', {}).get('valor')) else 'N'
                config_padrao['permitir_liberacao_excepcional'] = 'S' if parametro_bool(params.get('motorista.permitir_liberacao_excepcional', {}).get('valor')) else 'N'
            except Exception:
                pass
            return config_padrao

        retorno = {
            'horario_limite_edicao': str(row.get('horario_limite_edicao') or '11:00')[:5],
            'limite_dias_disponiveis_semana': int(row.get('limite_dias_disponiveis_semana') or 6),
            'permite_liberacao_setimo_dia': row.get('permite_liberacao_setimo_dia') or 'S',
            'bloquear_disponibilidade_apos_limite': 'S',
            'permitir_liberacao_excepcional': 'S'
        }
        try:
            params = carregar_parametros_operacionais_motorista_empresa(empresa_id, cur=cur)
            retorno['horario_limite_edicao'] = normalizar_time_param(
                params.get('motorista.horario_limite_disponibilidade', {}).get('valor'),
                retorno['horario_limite_edicao']
            )
            retorno['bloquear_disponibilidade_apos_limite'] = 'S' if parametro_bool(params.get('motorista.bloquear_disponibilidade_apos_limite', {}).get('valor')) else 'N'
            retorno['permitir_liberacao_excepcional'] = 'S' if parametro_bool(params.get('motorista.permitir_liberacao_excepcional', {}).get('valor')) else 'N'
        except Exception:
            pass
        return retorno

    except Exception as e:
        print(f"Aviso: erro ao buscar configuração de disponibilidade: {e}")
        return config_padrao

    finally:
        cur.close()
        con.close()


def semana_domingo_sabado(data_ref=None):
    """
    Retorna domingo e sábado da semana da data informada.
    Python: segunda=0 ... domingo=6
    """
    if data_ref is None:
        data_ref = date.today()

    if isinstance(data_ref, str):
        data_ref = datetime.strptime(data_ref[:10], '%Y-%m-%d').date()

    dias_desde_domingo = (data_ref.weekday() + 1) % 7
    domingo = data_ref - timedelta(days=dias_desde_domingo)
    sabado = domingo + timedelta(days=6)

    return domingo, sabado


def data_pode_ser_editada_pelo_motorista(data_disponibilidade, horario_limite='11:00'):
    """
    Regra:
    Motorista pode alterar a disponibilidade do dia alvo até o horário limite do dia anterior.
    Exemplo: disponibilidade de terça pode ser alterada até segunda às 11:00.
    Após isso, apenas supervisor/admin em fase futura.
    """
    if isinstance(data_disponibilidade, str):
        data_disponibilidade = datetime.strptime(data_disponibilidade[:10], '%Y-%m-%d').date()

    try:
        hora, minuto = str(horario_limite or '11:00')[:5].split(':')
        hora = int(hora)
        minuto = int(minuto)
    except Exception:
        hora, minuto = 11, 0

    limite = datetime.combine(data_disponibilidade - timedelta(days=1), datetime.min.time())
    limite = limite.replace(hour=hora, minute=minuto, second=0, microsecond=0)

    return datetime.now() <= limite


def carregar_disponibilidade_semana_motorista(empresa_id, motorista_id, data_inicio, data_fim):
    con = obter_conexao()

    if con is None:
        return {}

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT id,
                           data_disponibilidade,
                           status_disponibilidade,
                           observacao,
                           origem_lancamento,
                           liberado_excepcional,
                           data_atualizacao
                    FROM disponibilidade_motorista
                    WHERE empresa_id = %s
                      AND motorista_id = %s
                      AND data_disponibilidade BETWEEN %s AND %s
                    ORDER BY data_disponibilidade ASC
                    """, (empresa_id, motorista_id, data_inicio, data_fim))

        rows = cur.fetchall()
        retorno = {}

        for row in rows:
            chave = row['data_disponibilidade']
            if hasattr(chave, 'strftime'):
                chave = chave.strftime('%Y-%m-%d')
            else:
                chave = str(chave)[:10]

            retorno[chave] = row

        return retorno

    except Exception as e:
        print(f"Erro ao carregar disponibilidade do motorista: {e}")
        return {}

    finally:
        cur.close()
        con.close()


@app.route('/portal-motorista/disponibilidade', methods=['GET', 'POST'])
@login_required
@motorista_required
def disponibilidade_motorista():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_id = session.get('empresa_id')
    motorista_id = session.get('pessoa_id')

    if not empresa_id or not motorista_id:
        flash("Sessão do motorista incompleta. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    data_ref_str = request.args.get('semana', '').strip()

    if data_ref_str:
        try:
            data_ref = datetime.strptime(data_ref_str[:10], '%Y-%m-%d').date()
        except Exception:
            data_ref = date.today()
    else:
        data_ref = date.today()

    data_inicio, data_fim = semana_domingo_sabado(data_ref)
    config_disp = obter_config_disponibilidade_empresa(empresa_id)

    dias_semana_nomes = [
        'Domingo',
        'Segunda-feira',
        'Terça-feira',
        'Quarta-feira',
        'Quinta-feira',
        'Sexta-feira',
        'Sábado'
    ]

    disponibilidade_atual = carregar_disponibilidade_semana_motorista(
        empresa_id,
        motorista_id,
        data_inicio,
        data_fim
    )

    if request.method == 'POST':
        datas_semana = [data_inicio + timedelta(days=i) for i in range(7)]
        novos_status = {}
        novas_obs = {}
        bloqueadas_por_horario = []

        for data_dia in datas_semana:
            chave = data_dia.strftime('%Y-%m-%d')
            status = request.form.get(f'status_{chave}', 'Sem resposta').strip()
            obs = request.form.get(f'observacao_{chave}', '').strip()

            if status not in ['Disponível', 'Ausente', 'Sem resposta']:
                status = 'Sem resposta'

            if config_disp.get('bloquear_disponibilidade_apos_limite') != 'S':
                pode_editar = True
            else:
                pode_editar = data_pode_ser_editada_pelo_motorista(
                    data_dia,
                    config_disp.get('horario_limite_edicao')
                )

            registro_existente = disponibilidade_atual.get(chave)

            if not pode_editar:
                # Mantém o que já existia quando o prazo passou.
                if registro_existente:
                    novos_status[chave] = registro_existente.get('status_disponibilidade') or 'Sem resposta'
                    novas_obs[chave] = registro_existente.get('observacao') or ''
                else:
                    novos_status[chave] = 'Sem resposta'
                    novas_obs[chave] = ''

                bloqueadas_por_horario.append(chave)
                continue

            novos_status[chave] = status
            novas_obs[chave] = obs

        limite_dias = int(config_disp.get('limite_dias_disponiveis_semana') or 6)
        total_disponivel = sum(1 for status in novos_status.values() if status == 'Disponível')

        if total_disponivel > limite_dias:
            flash(
                f"Sua escala permite até {limite_dias} dias disponíveis na semana. "
                "Para trabalhar no 7º dia, solicite liberação ao supervisor.",
                "danger"
            )
            return redirect(url_for('disponibilidade_motorista', semana=data_inicio.strftime('%Y-%m-%d')))

        con = obter_conexao()

        if con is None:
            flash("Erro de conexão com o banco de dados.", "danger")
            return redirect(url_for('portal_motorista'))

        cur = con.cursor(dictionary=True)

        try:
            for data_dia in datas_semana:
                chave = data_dia.strftime('%Y-%m-%d')

                if chave in bloqueadas_por_horario:
                    continue

                cur.execute("""
                            INSERT INTO disponibilidade_motorista (empresa_id,
                                                                   motorista_id,
                                                                   data_disponibilidade,
                                                                   dia_semana,
                                                                   status_disponibilidade,
                                                                   observacao,
                                                                   origem_lancamento,
                                                                   usuario_lancamento_id,
                                                                   bloqueado_por_horario,
                                                                   liberado_excepcional,
                                                                   data_criacao,
                                                                   data_atualizacao)
                            VALUES (%s, %s, %s, %s, %s, %s, 'Motorista', %s, 'N', 'N', NOW(), NOW()) ON DUPLICATE KEY
                            UPDATE
                                status_disponibilidade =
                            VALUES (status_disponibilidade), observacao =
                            VALUES (observacao), origem_lancamento = 'Motorista', usuario_lancamento_id =
                            VALUES (usuario_lancamento_id), data_atualizacao = NOW()
                            """, (
                                empresa_id,
                                motorista_id,
                                chave,
                                dias_semana_nomes[(data_dia.weekday() + 1) % 7],
                                novos_status[chave],
                                novas_obs[chave] or None,
                                session.get('usuario_id')
                            ))

            con.commit()

            if bloqueadas_por_horario:
                flash(
                    "Disponibilidade salva. Alguns dias não foram alterados porque passaram do horário limite.",
                    "warning"
                )
            else:
                flash("Disponibilidade semanal salva com sucesso.", "success")

            return redirect(url_for('disponibilidade_motorista', semana=data_inicio.strftime('%Y-%m-%d')))

        except Exception as e:
            con.rollback()
            print(f"Erro ao salvar disponibilidade do motorista: {e}")
            flash(f"Erro técnico ao salvar disponibilidade: {e}", "danger")
            return redirect(url_for('disponibilidade_motorista', semana=data_inicio.strftime('%Y-%m-%d')))

        finally:
            cur.close()
            con.close()

    dias = []
    for i in range(7):
        data_dia = data_inicio + timedelta(days=i)
        chave = data_dia.strftime('%Y-%m-%d')
        registro = disponibilidade_atual.get(chave, {})

        dias.append({
            'data': data_dia,
            'data_str': chave,
            'dia_nome': dias_semana_nomes[i],
            'status': registro.get('status_disponibilidade') or 'Sem resposta',
            'observacao': registro.get('observacao') or '',
            'pode_editar': True if config_disp.get('bloquear_disponibilidade_apos_limite') != 'S' else data_pode_ser_editada_pelo_motorista(
                data_dia,
                config_disp.get('horario_limite_edicao')
            ),
            'origem': registro.get('origem_lancamento') or '-',
            'liberado_excepcional': registro.get('liberado_excepcional') or 'N'
        })

    total_disponivel = sum(1 for d in dias if d['status'] == 'Disponível')
    total_ausente = sum(1 for d in dias if d['status'] == 'Ausente')
    total_sem_resposta = sum(1 for d in dias if d['status'] == 'Sem resposta')

    semana_anterior = data_inicio - timedelta(days=7)
    proxima_semana = data_inicio + timedelta(days=7)

    return render_template(
        'disponibilidade_motorista.html',
        usuario_logado=usuario_logado,
        dias=dias,
        data_inicio=data_inicio,
        data_fim=data_fim,
        semana_anterior=semana_anterior,
        proxima_semana=proxima_semana,
        config_disp=config_disp,
        total_disponivel=total_disponivel,
        total_ausente=total_ausente,
        total_sem_resposta=total_sem_resposta
    )


def normalizar_coordenada(valor, casas=8):
    """
    Normaliza latitude/longitude para gravação no MySQL.
    Aceita vírgula decimal, remove espaços e arredonda para evitar erro:
    Data truncated for column 'latitude'
    """
    if valor is None:
        return None

    valor_str = str(valor).strip()

    if not valor_str:
        return None

    valor_str = valor_str.replace(',', '.')

    match = re.search(r'-?\d+(?:\.\d+)?', valor_str)

    if not match:
        return None

    return round(float(match.group(0)), casas)


def validar_coordenadas_base(latitude_val, longitude_val, raio_val):
    if latitude_val is not None and not (-90 <= latitude_val <= 90):
        return "Latitude inválida. Informe um valor entre -90 e 90."

    if longitude_val is not None and not (-180 <= longitude_val <= 180):
        return "Longitude inválida. Informe um valor entre -180 e 180."

    if raio_val < 30:
        return "O raio permitido deve ser de pelo menos 30 metros."

    return None


# ==========================================================
# FASE 4.1.2.2 - CADASTRO DE BASES OPERACIONAIS
# ==========================================================
def carregar_bases_operacionais(empresa_id, apenas_ativas=True):
    con = obter_conexao()

    if con is None:
        return []

    cur = con.cursor(dictionary=True)

    try:
        sql = """
              SELECT id,
                     empresa_id,
                     nome_base,
                     descricao,
                     endereco,
                     latitude,
                     longitude,
                     raio_permitido_metros,
                     codigo_qr_base,
                     status_base
              FROM bases_operacionais
              WHERE empresa_id = %s
              """
        params = [empresa_id]

        if apenas_ativas:
            sql += " AND status_base = 'Ativa'"

        sql += " ORDER BY nome_base ASC"

        cur.execute(sql, params)
        return cur.fetchall()

    except Exception as e:
        print(f"Erro ao carregar bases operacionais: {e}")
        return []

    finally:
        cur.close()
        con.close()


def buscar_base_operacional(empresa_id, base_id):
    if not base_id:
        return None

    con = obter_conexao()

    if con is None:
        return None

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT id,
                           empresa_id,
                           nome_base,
                           descricao,
                           endereco,
                           latitude,
                           longitude,
                           raio_permitido_metros,
                           codigo_qr_base,
                           status_base
                    FROM bases_operacionais
                    WHERE empresa_id = %s
                      AND id = %s LIMIT 1
                    """, (empresa_id, base_id))

        return cur.fetchone()

    except Exception as e:
        print(f"Erro ao buscar base operacional: {e}")
        return None

    finally:
        cur.close()
        con.close()


@app.route('/operacao/bases-operacionais', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor')
def visualizar_bases_operacionais():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    empresa_id_filtro = request.args.get('empresa_id', '').strip()
    status_filtro = request.args.get('status_base', '').strip()
    busca = request.args.get('busca', '').strip()

    if is_super_admin and empresa_id_filtro and empresa_id_filtro.isdigit():
        empresa_id = int(empresa_id_filtro)
    else:
        empresa_id = empresa_logada_id

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('dashboard'))

    cur = con.cursor(dictionary=True)

    try:
        query = """
                SELECT b.id,
                       b.empresa_id,
                       e.nome_fantasia AS empresa_nome,
                       e.razao_social  AS empresa_razao_social,
                       b.nome_base,
                       b.descricao,
                       b.endereco,
                       b.latitude,
                       b.longitude,
                       b.raio_permitido_metros,
                       b.codigo_qr_base,
                       b.qr_validade_minutos,
                       b.status_base,
                       b.created_at,
                       b.updated_at
                FROM bases_operacionais b
                         INNER JOIN empresas e ON e.id = b.empresa_id
                WHERE b.empresa_id = %s \
                """
        params = [empresa_id]

        if status_filtro in ['Ativa', 'Inativa']:
            query += " AND b.status_base = %s"
            params.append(status_filtro)

        if busca:
            query += " AND (b.nome_base LIKE %s OR b.endereco LIKE %s OR b.descricao LIKE %s)"
            like = f"%{busca}%"
            params.extend([like, like, like])

        query += " ORDER BY b.nome_base ASC"

        cur.execute(query, params)
        bases = cur.fetchall()

        empresas = []
        if is_super_admin:
            cur.execute("""
                        SELECT id, razao_social, nome_fantasia
                        FROM empresas
                        ORDER BY nome_fantasia ASC, razao_social ASC
                        """)
            empresas = cur.fetchall()

    except Exception as e:
        print(f"Erro ao listar bases operacionais: {e}")
        flash(f"Erro técnico ao listar bases: {e}", "danger")
        bases = []
        empresas = []

    finally:
        cur.close()
        con.close()

    return render_template(
        'visualizar_bases_operacionais.html',
        usuario_logado=usuario_logado,
        bases=bases,
        empresas=empresas,
        empresa_id=empresa_id,
        is_super_admin=is_super_admin,
        filtros={
            'empresa_id': empresa_id,
            'status_base': status_filtro,
            'busca': busca
        }
    )


@app.route('/operacao/bases-operacionais/cadastro', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor')
def cadastro_base_operacional():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    empresas = []

    if is_super_admin:
        con_emp = obter_conexao()
        if con_emp:
            cur_emp = con_emp.cursor(dictionary=True)
            try:
                cur_emp.execute("""
                                SELECT id, razao_social, nome_fantasia
                                FROM empresas
                                ORDER BY nome_fantasia ASC, razao_social ASC
                                """)
                empresas = cur_emp.fetchall()
            finally:
                cur_emp.close()
                con_emp.close()

    if request.method == 'POST':
        if is_super_admin:
            empresa_id_post = request.form.get('empresa_id', '').strip()
            empresa_id = int(empresa_id_post) if empresa_id_post.isdigit() else empresa_logada_id
        else:
            empresa_id = empresa_logada_id

        nome_base = request.form.get('nome_base', '').strip()
        descricao = request.form.get('descricao', '').strip()
        endereco = request.form.get('endereco', '').strip()
        latitude = request.form.get('latitude', '').strip()
        longitude = request.form.get('longitude', '').strip()
        raio_permitido_metros = request.form.get('raio_permitido_metros', '150').strip()
        codigo_qr_base = request.form.get('codigo_qr_base', '').strip()
        qr_validade_minutos = request.form.get('qr_validade_minutos', '5').strip()
        status_base = request.form.get('status_base', 'Ativa').strip()

        if not nome_base:
            flash("Informe o nome da base operacional.", "danger")
            return redirect(url_for('cadastro_base_operacional'))

        if status_base not in ['Ativa', 'Inativa']:
            status_base = 'Ativa'

        if not codigo_qr_base:
            codigo_qr_base = gerar_codigo_qr_base(nome_base, empresa_id)

        try:
            latitude_val = normalizar_coordenada(latitude)
            longitude_val = normalizar_coordenada(longitude)
            raio_val = int(raio_permitido_metros or 150)
            qr_validade_val = int(qr_validade_minutos or 5)
        except Exception:
            flash("Latitude, longitude ou raio informado em formato inválido.", "danger")
            return redirect(url_for('cadastro_base_operacional'))

        if qr_validade_val < 1:
            qr_validade_val = 5

        if qr_validade_val > 30:
            qr_validade_val = 30

        erro_coordenada = validar_coordenadas_base(latitude_val, longitude_val, raio_val)
        if erro_coordenada:
            flash(erro_coordenada, "danger")
            return redirect(url_for('cadastro_base_operacional'))

        con = obter_conexao()

        if con is None:
            flash("Erro de conexão com o banco de dados.", "danger")
            return redirect(url_for('visualizar_bases_operacionais'))

        cur = con.cursor()

        try:
            cur.execute("""
                        INSERT INTO bases_operacionais (empresa_id,
                                                        nome_base,
                                                        descricao,
                                                        endereco,
                                                        latitude,
                                                        longitude,
                                                        raio_permitido_metros,
                                                        codigo_qr_base,
                                                        qr_validade_minutos,
                                                        status_base,
                                                        created_at,
                                                        updated_at)
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                        """, (
                            empresa_id,
                            nome_base,
                            descricao or None,
                            endereco or None,
                            latitude_val,
                            longitude_val,
                            raio_val,
                            codigo_qr_base,
                            qr_validade_val,
                            status_base
                        ))

            con.commit()
            flash("Base operacional cadastrada com sucesso.", "success")
            return redirect(url_for('visualizar_bases_operacionais', empresa_id=empresa_id if is_super_admin else ''))

        except Exception as e:
            con.rollback()
            print(f"Erro ao cadastrar base operacional: {e}")
            flash(f"Erro técnico ao cadastrar base: {e}", "danger")

        finally:
            cur.close()
            con.close()

    return render_template(
        'cadastro_base_operacional.html',
        usuario_logado=usuario_logado,
        empresas=empresas,
        is_super_admin=is_super_admin
    )


@app.route('/operacao/bases-operacionais/editar/<int:id>', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor')
def editar_base_operacional(id):
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('visualizar_bases_operacionais'))

    cur = con.cursor(dictionary=True)

    try:
        if is_super_admin:
            cur.execute("""
                        SELECT *
                        FROM bases_operacionais
                        WHERE id = %s LIMIT 1
                        """, (id,))
        else:
            cur.execute("""
                        SELECT *
                        FROM bases_operacionais
                        WHERE id = %s
                          AND empresa_id = %s LIMIT 1
                        """, (id, empresa_logada_id))

        base = cur.fetchone()

        if not base:
            flash("Base operacional não encontrada ou sem permissão de acesso.", "danger")
            return redirect(url_for('visualizar_bases_operacionais'))

        if request.method == 'POST':
            nome_base = request.form.get('nome_base', '').strip()
            descricao = request.form.get('descricao', '').strip()
            endereco = request.form.get('endereco', '').strip()
            latitude = request.form.get('latitude', '').strip()
            longitude = request.form.get('longitude', '').strip()
            raio_permitido_metros = request.form.get('raio_permitido_metros', '150').strip()
            codigo_qr_base = request.form.get('codigo_qr_base', '').strip()
            qr_validade_minutos = request.form.get('qr_validade_minutos', '5').strip()
            status_base = request.form.get('status_base', 'Ativa').strip()

            if not nome_base:
                flash("Informe o nome da base operacional.", "danger")
                return redirect(url_for('editar_base_operacional', id=id))

            if status_base not in ['Ativa', 'Inativa']:
                status_base = 'Ativa'

            if not codigo_qr_base:
                codigo_qr_base = gerar_codigo_qr_base(nome_base, base.get('empresa_id'))

            try:
                latitude_val = normalizar_coordenada(latitude)
                longitude_val = normalizar_coordenada(longitude)
                raio_val = int(raio_permitido_metros or 150)
                qr_validade_val = int(qr_validade_minutos or 5)
            except Exception:
                flash("Latitude, longitude ou raio informado em formato inválido.", "danger")
                return redirect(url_for('editar_base_operacional', id=id))

            if qr_validade_val < 1:
                qr_validade_val = 5

            if qr_validade_val > 30:
                qr_validade_val = 30

            erro_coordenada = validar_coordenadas_base(latitude_val, longitude_val, raio_val)
            if erro_coordenada:
                flash(erro_coordenada, "danger")
                return redirect(url_for('editar_base_operacional', id=id))

            cur.execute("""
                        UPDATE bases_operacionais
                        SET nome_base             = %s,
                            descricao             = %s,
                            endereco              = %s,
                            latitude              = %s,
                            longitude             = %s,
                            raio_permitido_metros = %s,
                            codigo_qr_base        = %s,
                            qr_validade_minutos   = %s,
                            status_base           = %s,
                            updated_at            = NOW()
                        WHERE id = %s
                        """, (
                            nome_base,
                            descricao or None,
                            endereco or None,
                            latitude_val,
                            longitude_val,
                            raio_val,
                            codigo_qr_base,
                            qr_validade_val,
                            status_base,
                            id
                        ))

            con.commit()
            flash("Base operacional atualizada com sucesso.", "success")
            return redirect(
                url_for('visualizar_bases_operacionais', empresa_id=base.get('empresa_id') if is_super_admin else ''))

    except Exception as e:
        con.rollback()
        print(f"Erro ao editar base operacional: {e}")
        flash(f"Erro técnico ao editar base: {e}", "danger")
        return redirect(url_for('visualizar_bases_operacionais'))

    finally:
        cur.close()
        con.close()

    return render_template(
        'editar_base_operacional.html',
        usuario_logado=usuario_logado,
        base=base,
        is_super_admin=is_super_admin
    )


# ==========================================================
# FASE 4.1.2 - PAINEL DE ESCALA DO SUPERVISOR
# ==========================================================
def obter_config_escala_empresa(empresa_id):
    """
    Configuração operacional da escala.
    Padrão: motorista cancelado que deve comparecer precisa confirmar presença até 11:01.
    """
    config_padrao = {
        'horario_limite_presenca': '11:00',
        'aplicar_falta_automatica': 'S'
    }

    con = obter_conexao()

    if con is None:
        return config_padrao

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT horario_limite_presenca, aplicar_falta_automatica
                    FROM configuracoes_escala_motorista
                    WHERE empresa_id = %s LIMIT 1
                    """, (empresa_id,))

        row = cur.fetchone()

        retorno = dict(config_padrao)
        if row:
            retorno = {
                'horario_limite_presenca': str(row.get('horario_limite_presenca') or '11:00')[:5],
                'aplicar_falta_automatica': row.get('aplicar_falta_automatica') or 'S'
            }
        try:
            params = carregar_parametros_operacionais_motorista_empresa(empresa_id, cur=cur)
            retorno['horario_limite_presenca'] = normalizar_time_param(
                params.get('escala.horario_limite_checkin', {}).get('valor'),
                retorno.get('horario_limite_presenca') or '11:00'
            )
            retorno['aplicar_falta_automatica'] = 'S' if parametro_bool(params.get('escala.aplicar_falta_automatica_checkin', {}).get('valor')) else 'N'
        except Exception:
            pass
        return retorno

    except Exception as e:
        print(f"Aviso: erro ao buscar configuração de escala: {e}")
        return config_padrao

    finally:
        cur.close()
        con.close()


def obter_config_checkin_empresa(empresa_id, cur=None):
    """Parâmetros de check-in por empresa."""
    config = {
        'exigir_qrcode': 'S',
        'exigir_selfie': 'S',
        'exigir_gps_raio': 'S',
        'permitir_checkin_manual_supervisor': 'S'
    }
    try:
        params = carregar_parametros_operacionais_motorista_empresa(empresa_id, cur=cur)
        config['exigir_qrcode'] = 'S' if parametro_bool(params.get('checkin.exigir_qrcode', {}).get('valor')) else 'N'
        config['exigir_selfie'] = 'S' if parametro_bool(params.get('checkin.exigir_selfie', {}).get('valor')) else 'N'
        config['exigir_gps_raio'] = 'S' if parametro_bool(params.get('checkin.exigir_gps_raio', {}).get('valor')) else 'N'
        config['permitir_checkin_manual_supervisor'] = 'S' if parametro_bool(params.get('checkin.permitir_checkin_manual_supervisor', {}).get('valor')) else 'N'
    except Exception as e:
        print(f"Aviso: erro ao buscar configuração de check-in: {e}")
    return config


def aplicar_faltas_automaticas_escala(empresa_id, data_escala):
    """
    Regra inteligente:
    Se o motorista estiver como 'Cancelado, comparecer na base' e não confirmar presença
    até o horário limite do dia da escala, o sistema muda automaticamente para Falta.

    Essa rotina roda quando o supervisor abre a tela da escala.
    Depois poderemos evoluir para robô agendado.
    """
    config = obter_config_escala_empresa(empresa_id)

    if config.get('aplicar_falta_automatica') != 'S':
        return 0

    try:
        hora, minuto = str(config.get('horario_limite_presenca') or '11:00')[:5].split(':')
        hora = int(hora)
        minuto = int(minuto)
    except Exception:
        hora, minuto = 11, 1

    if isinstance(data_escala, str):
        data_ref = datetime.strptime(data_escala[:10], '%Y-%m-%d').date()
    else:
        data_ref = data_escala

    limite = datetime.combine(data_ref, datetime.min.time()).replace(
        hour=hora,
        minute=minuto,
        second=0,
        microsecond=0
    )

    if datetime.now() <= limite:
        return 0

    con = obter_conexao()

    if con is None:
        return 0

    cur = con.cursor()

    try:
        cur.execute("""
                    UPDATE escala_motorista
                    SET status_escala    = 'Falta',
                        status_presenca  = 'Não compareceu',
                        falta_automatica = 'S',
                        falta_marcada_em = NOW(),
                        falta_motivo     = 'Falta automática: motorista cancelado não confirmou presença na base até o horário limite.',
                        data_atualizacao = NOW()
                    WHERE empresa_id = %s
                      AND data_escala = %s
                      AND status_escala = 'Cancelado, comparecer na base'
                      AND status_presenca = 'Aguardando chegada'
                      AND (presenca_confirmada_em IS NULL)
                    """, (empresa_id, data_ref))

        total = cur.rowcount or 0
        con.commit()
        return total

    except Exception as e:
        con.rollback()
        print(f"Aviso: erro ao aplicar faltas automáticas da escala: {e}")
        return 0

    finally:
        cur.close()
        con.close()


def status_presenca_padrao_por_escala(status_escala, status_presenca_atual=None):
    """
    Define a presença de acordo com a regra de negócio:
    - Confirmado com rota: não precisa validar presença.
    - Cancelado, comparecer na base: precisa confirmar chegada.
    - Falta: não compareceu.
    - Demais status: não se aplica.
    """
    if status_escala == 'Cancelado, comparecer na base':
        if status_presenca_atual == 'Chegada confirmada':
            return 'Chegada confirmada'
        return 'Aguardando chegada'

    if status_escala == 'Falta':
        return 'Não compareceu'

    return 'Não se aplica'


def carregar_escala_supervisor(empresa_id, data_escala):
    """
    Lista todos os motoristas ativos da empresa, trazendo:
    - disponibilidade informada pelo motorista
    - escala definida pelo supervisor
    - status de presença
    """
    con = obter_conexao()

    if con is None:
        return []

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT mot.id                                              AS motorista_id,
                           mot.nome_completo                                   AS motorista_nome,
                           mot.cpf_cnpj                                        AS motorista_cpf_cnpj,
                           COALESCE(dm.status_disponibilidade, 'Sem resposta') AS status_disponibilidade,
                           dm.observacao                                       AS observacao_disponibilidade,
                           em.id                                               AS escala_id,
                           COALESCE(em.status_escala, 'Pendente')              AS status_escala,
                           COALESCE(em.status_presenca, 'Não se aplica')       AS status_presenca,
                           em.base_operacional_id,
                           em.base_operacao,
                           bo.nome_base                                        AS base_operacional_nome,
                           em.horario_apresentacao,
                           em.observacao_supervisor,
                           em.presenca_confirmada_em,
                           em.falta_automatica,
                           em.falta_marcada_em,
                           em.falta_motivo,
                           em.falta_revertida,
                           em.motivo_reversao,
                           em.data_reversao,
                           cem.data_ciencia,
                           cem.origem_ciencia
                    FROM pessoas mot
                             LEFT JOIN disponibilidade_motorista dm
                                       ON dm.empresa_id = mot.empresa_id
                                           AND dm.motorista_id = mot.id
                                           AND dm.data_disponibilidade = %s
                             LEFT JOIN escala_motorista em
                                       ON em.empresa_id = mot.empresa_id
                                           AND em.motorista_id = mot.id
                                           AND em.data_escala = %s
                             LEFT JOIN bases_operacionais bo
                                       ON bo.id = em.base_operacional_id
                                           AND bo.empresa_id = em.empresa_id
                             LEFT JOIN ciencia_escala_motorista cem
                                       ON cem.empresa_id = mot.empresa_id
                                           AND cem.motorista_id = mot.id
                                           AND cem.escala_id = em.id
                    WHERE mot.empresa_id = %s
                      AND mot.tipo_cadastro = 'Motorista'
                      AND mot.status_cadastro = 'Ativo'
                    ORDER BY mot.nome_completo ASC
                    """, (data_escala, data_escala, empresa_id))

        linhas = cur.fetchall()

        for linha in linhas:
            linha['horario_apresentacao'] = normalizar_horario_input(linha.get('horario_apresentacao'))

        return linhas

    except Exception as e:
        print(f"Erro ao carregar escala do supervisor: {e}")
        return []

    finally:
        cur.close()
        con.close()


def gerar_pendencias_escala(escala):
    pendencias = {
        'sem_resposta': [],
        'disponiveis_pendentes': [],
        'ciencia_pendente': [],
        'aguardando_chegada': [],
        'faltas_automaticas': []
    }

    for item in escala:
        status_disp = item.get('status_disponibilidade') or 'Sem resposta'
        status_escala = item.get('status_escala') or 'Pendente'
        status_presenca = item.get('status_presenca') or 'Não se aplica'

        if status_disp == 'Sem resposta':
            pendencias['sem_resposta'].append(item)

        if status_disp == 'Disponível' and status_escala == 'Pendente':
            pendencias['disponiveis_pendentes'].append(item)

        if status_escala != 'Pendente' and not item.get('data_ciencia'):
            pendencias['ciencia_pendente'].append(item)

        if status_escala == 'Cancelado, comparecer na base' and status_presenca == 'Aguardando chegada':
            pendencias['aguardando_chegada'].append(item)

        if status_escala == 'Falta' and item.get('falta_automatica') == 'S':
            pendencias['faltas_automaticas'].append(item)

    return pendencias


def gerar_mensagem_whatsapp_escala(escala, data_escala):
    data_br = data_escala.strftime('%d/%m/%Y') if hasattr(data_escala, 'strftime') else str(data_escala)

    grupos = {
        'Confirmado com rota': [],
        'Cancelado, comparecer na base': [],
        'Reserva / Avulso': [],
        'Cancelado definitivo': [],
        'Falta': [],
        'Sem resposta': []
    }

    for item in escala:
        status_escala = item.get('status_escala') or 'Pendente'
        status_disp = item.get('status_disponibilidade') or 'Sem resposta'

        if status_escala in grupos and status_escala != 'Pendente':
            grupos[status_escala].append(item)
        elif status_disp == 'Sem resposta':
            grupos['Sem resposta'].append(item)

    linhas = []
    linhas.append(f"📋 ESCALA DOS MOTORISTAS — {data_br}")
    linhas.append("")

    def add_grupo(titulo, itens, incluir_base=False, incluir_obs=False):
        linhas.append(titulo)

        if itens:
            for item in itens:
                nome = item.get('motorista_nome') or 'Motorista'
                detalhe = ""

                if incluir_base:
                    base = item.get('base_operacao') or ''
                    horario = item.get('horario_apresentacao') or ''

                    partes = []
                    if base:
                        partes.append(f"Base: {base}")
                    if horario:
                        partes.append(f"Horário: {horario}")

                    if partes:
                        detalhe = " — " + " | ".join(partes)

                if incluir_obs and item.get('observacao_supervisor'):
                    detalhe += f" — {item.get('observacao_supervisor')}"

                linhas.append(f"- {nome}{detalhe}")
        else:
            linhas.append("- Nenhum")

        linhas.append("")

    add_grupo("✅ CONFIRMADOS COM ROTA:", grupos['Confirmado com rota'], incluir_base=True)
    add_grupo("⚠️ CANCELADOS, COMPARECER NA BASE:", grupos['Cancelado, comparecer na base'], incluir_base=True,
              incluir_obs=True)
    add_grupo("🟡 RESERVA / AVULSO:", grupos['Reserva / Avulso'], incluir_base=True, incluir_obs=True)
    add_grupo("❌ CANCELADOS DEFINITIVO:", grupos['Cancelado definitivo'])
    add_grupo("🚫 FALTAS:", grupos['Falta'])
    add_grupo("🕓 SEM RESPOSTA:", grupos['Sem resposta'])

    linhas.append(
        "Atenção: motoristas orientados a comparecer na base deverão confirmar chegada conforme orientação operacional.")

    return "\n".join(linhas)


@app.route('/operacao/escala-motoristas', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor')
def escala_motoristas():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    data_escala_str = request.args.get('data_escala', '').strip()

    if not data_escala_str:
        data_escala = date.today() + timedelta(days=1)
    else:
        try:
            data_escala = datetime.strptime(data_escala_str[:10], '%Y-%m-%d').date()
        except Exception:
            data_escala = date.today() + timedelta(days=1)

    empresa_id_filtro = request.args.get('empresa_id', '').strip()

    if is_super_admin and empresa_id_filtro and empresa_id_filtro.isdigit():
        empresa_id = int(empresa_id_filtro)
    else:
        empresa_id = empresa_logada_id

    if request.method == 'POST':
        data_post = request.form.get('data_escala', '').strip()

        try:
            data_post_ref = datetime.strptime(data_post[:10], '%Y-%m-%d').date()
        except Exception:
            flash("Data da escala inválida.", "danger")
            return redirect(url_for('escala_motoristas'))

        if is_super_admin:
            empresa_post = request.form.get('empresa_id', '').strip()
            empresa_id_post = int(empresa_post) if empresa_post.isdigit() else empresa_logada_id
        else:
            empresa_id_post = empresa_logada_id

        motoristas_ids = request.form.getlist('motorista_id')
        status_validos = [
            'Pendente',
            'Confirmado com rota',
            'Cancelado, comparecer na base',
            'Cancelado definitivo',
            'Reserva / Avulso',
            'Falta'
        ]

        con = obter_conexao()

        if con is None:
            flash("Erro de conexão com o banco de dados.", "danger")
            return redirect(url_for('escala_motoristas', data_escala=data_post_ref.strftime('%Y-%m-%d')))

        cur = con.cursor(dictionary=True)

        try:
            for motorista_id_str in motoristas_ids:
                if not motorista_id_str.isdigit():
                    continue

                motorista_id = int(motorista_id_str)
                status_escala = request.form.get(f'status_escala_{motorista_id}', 'Pendente').strip()
                base_operacional_id_str = request.form.get(f'base_operacional_id_{motorista_id}', '').strip()
                base_operacional_id = int(base_operacional_id_str) if base_operacional_id_str.isdigit() else None
                base_operacao = None

                if base_operacional_id:
                    base_cad = buscar_base_operacional(empresa_id_post, base_operacional_id)
                    if base_cad:
                        base_operacao = base_cad.get('nome_base')

                horario_apresentacao = request.form.get(f'horario_apresentacao_{motorista_id}', '').strip()
                observacao_supervisor = request.form.get(f'observacao_supervisor_{motorista_id}', '').strip()

                if status_escala not in status_validos:
                    status_escala = 'Pendente'

                cur.execute("""
                            SELECT id, status_escala, status_presenca, falta_automatica
                            FROM escala_motorista
                            WHERE empresa_id = %s
                              AND motorista_id = %s
                              AND data_escala = %s LIMIT 1
                            """, (empresa_id_post, motorista_id, data_post_ref))

                escala_atual = cur.fetchone()
                status_presenca_atual = escala_atual.get('status_presenca') if escala_atual else None
                status_presenca = status_presenca_padrao_por_escala(status_escala, status_presenca_atual)

                falta_revertida = 'N'
                motivo_reversao = None
                data_reversao_sql = None

                # Se era falta automática e o supervisor alterou para outro status,
                # registramos reversão manual para auditoria.
                if escala_atual and escala_atual.get('status_escala') == 'Falta' and escala_atual.get(
                        'falta_automatica') == 'S' and status_escala != 'Falta':
                    falta_revertida = 'S'
                    motivo_reversao = observacao_supervisor or 'Falta automática revertida pelo supervisor.'
                    data_reversao_sql = datetime.now()

                cur.execute("""
                            INSERT INTO escala_motorista (empresa_id,
                                                          motorista_id,
                                                          data_escala,
                                                          status_escala,
                                                          status_presenca,
                                                          base_operacional_id,
                                                          base_operacao,
                                                          horario_apresentacao,
                                                          observacao_supervisor,
                                                          usuario_supervisor_id,
                                                          falta_revertida,
                                                          motivo_reversao,
                                                          data_reversao,
                                                          data_criacao,
                                                          data_atualizacao)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW()) ON DUPLICATE KEY
                            UPDATE
                                status_escala =
                            VALUES (status_escala), status_presenca =
                            VALUES (status_presenca), base_operacional_id =
                            VALUES (base_operacional_id), base_operacao =
                            VALUES (base_operacao), horario_apresentacao =
                            VALUES (horario_apresentacao), observacao_supervisor =
                            VALUES (observacao_supervisor), usuario_supervisor_id =
                            VALUES (usuario_supervisor_id), falta_revertida = IF(VALUES (falta_revertida) = 'S', 'S', falta_revertida), motivo_reversao = IF(VALUES (falta_revertida) = 'S', VALUES (motivo_reversao), motivo_reversao), data_reversao = IF(VALUES (falta_revertida) = 'S', VALUES (data_reversao), data_reversao), falta_automatica = IF(VALUES (status_escala) <> 'Falta', 'N', falta_automatica), falta_motivo = IF(VALUES (status_escala) <> 'Falta', NULL, falta_motivo), data_atualizacao = NOW()
                            """, (
                                empresa_id_post,
                                motorista_id,
                                data_post_ref,
                                status_escala,
                                status_presenca,
                                base_operacional_id,
                                base_operacao or None,
                                horario_apresentacao or None,
                                observacao_supervisor or None,
                                session.get('usuario_id'),
                                falta_revertida,
                                motivo_reversao,
                                data_reversao_sql
                            ))
            data_escala_str = request.form.get('data_escala')
            registrar_auditoria_supervisor(cur, "SALVAR_ESCALA",
                                           f"Salvou/Atualizou a escala operacional para a data {data_escala_str}.")
            con.commit()
            flash("Escala dos motoristas salva com sucesso.", "success")

            return redirect(url_for(
                'escala_motoristas',
                data_escala=data_post_ref.strftime('%Y-%m-%d'),
                empresa_id=empresa_id_post if is_super_admin else ''
            ))

        except Exception as e:
            con.rollback()
            print(f"Erro ao salvar escala dos motoristas: {e}")
            flash(f"Erro técnico ao salvar escala: {e}", "danger")

        finally:
            cur.close()
            con.close()

    faltas_aplicadas = aplicar_faltas_automaticas_escala(empresa_id, data_escala)

    if faltas_aplicadas:
        flash(f"{faltas_aplicadas} falta(s) automática(s) aplicada(s) por ausência de confirmação na base.", "warning")

    escala_completa = carregar_escala_supervisor(empresa_id, data_escala)

    resumo = {
        'total_motoristas': len(escala_completa),
        'disponiveis': 0,
        'ausentes': 0,
        'sem_resposta': 0,
        'confirmados': 0,
        'cancelados_comparecer': 0,
        'cancelados_definitivo': 0,
        'reservas': 0,
        'faltas': 0,
        'chegada_confirmada': 0
    }

    for item in escala_completa:
        status_disp = item.get('status_disponibilidade') or 'Sem resposta'
        status_escala = item.get('status_escala') or 'Pendente'
        status_presenca = item.get('status_presenca') or 'Não se aplica'

        if status_disp == 'Disponível':
            resumo['disponiveis'] += 1
        elif status_disp == 'Ausente':
            resumo['ausentes'] += 1
        else:
            resumo['sem_resposta'] += 1

        if status_escala == 'Confirmado com rota':
            resumo['confirmados'] += 1
        elif status_escala == 'Cancelado, comparecer na base':
            resumo['cancelados_comparecer'] += 1
        elif status_escala == 'Cancelado definitivo':
            resumo['cancelados_definitivo'] += 1
        elif status_escala == 'Reserva / Avulso':
            resumo['reservas'] += 1
        elif status_escala == 'Falta':
            resumo['faltas'] += 1

        if status_presenca == 'Chegada confirmada':
            resumo['chegada_confirmada'] += 1

    pendencias = gerar_pendencias_escala(escala_completa)
    mensagem_whatsapp = gerar_mensagem_whatsapp_escala(escala_completa, data_escala)

    filtro_disponibilidade = request.args.get('filtro_disponibilidade', '').strip()
    filtro_escala = request.args.get('filtro_escala', '').strip()
    filtro_presenca = request.args.get('filtro_presenca', '').strip()
    busca_motorista = request.args.get('busca_motorista', '').strip().lower()

    escala = []

    for item in escala_completa:
        if filtro_disponibilidade and (item.get('status_disponibilidade') or 'Sem resposta') != filtro_disponibilidade:
            continue

        if filtro_escala and (item.get('status_escala') or 'Pendente') != filtro_escala:
            continue

        if filtro_presenca and (item.get('status_presenca') or 'Não se aplica') != filtro_presenca:
            continue

        if busca_motorista and busca_motorista not in (item.get('motorista_nome') or '').lower():
            continue

        escala.append(item)

    filtros_escala = {
        'filtro_disponibilidade': filtro_disponibilidade,
        'filtro_escala': filtro_escala,
        'filtro_presenca': filtro_presenca,
        'busca_motorista': busca_motorista
    }

    empresas = []
    if is_super_admin:
        con_emp = obter_conexao()
        if con_emp:
            cur_emp = con_emp.cursor(dictionary=True)
            try:
                cur_emp.execute("""
                                SELECT id, razao_social, nome_fantasia
                                FROM empresas
                                ORDER BY nome_fantasia ASC, razao_social ASC
                                """)
                empresas = cur_emp.fetchall()
            except Exception as e:
                print(f"Erro ao carregar empresas para escala: {e}")
            finally:
                cur_emp.close()
                con_emp.close()

    config_escala = obter_config_escala_empresa(empresa_id)
    bases_operacionais = carregar_bases_operacionais(empresa_id, apenas_ativas=True)
    fila_cancelados = carregar_fila_cancelados_base(empresa_id, data_escala)
    data_anterior = data_escala - timedelta(days=1)
    proxima_data = data_escala + timedelta(days=1)

    return render_template(
        'escala_motoristas.html',
        usuario_logado=usuario_logado,
        escala=escala,
        resumo=resumo,
        data_escala=data_escala,
        data_anterior=data_anterior,
        proxima_data=proxima_data,
        is_super_admin=is_super_admin,
        empresas=empresas,
        empresa_id=empresa_id,
        config_escala=config_escala,
        pendencias=pendencias,
        mensagem_whatsapp=mensagem_whatsapp,
        filtros_escala=filtros_escala,
        total_filtrado=len(escala),
        bases_operacionais=bases_operacionais,
        fila_cancelados=fila_cancelados
    )


def calcular_distancia_metros(lat1, lon1, lat2, lon2):
    """
    Calcula distância aproximada em metros entre duas coordenadas.
    Usado para validar se o motorista está dentro do raio da base.
    """
    import math

    if lat1 is None or lon1 is None or lat2 is None or lon2 is None:
        return None

    raio_terra = 6371000
    phi1 = math.radians(float(lat1))
    phi2 = math.radians(float(lat2))
    delta_phi = math.radians(float(lat2) - float(lat1))
    delta_lambda = math.radians(float(lon2) - float(lon1))

    a = (
            math.sin(delta_phi / 2) ** 2
            + math.cos(phi1) * math.cos(phi2) * math.sin(delta_lambda / 2) ** 2
    )

    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))

    return round(raio_terra * c, 2)


def registrar_auditoria_supervisor(cursor, tipo_acao, descricao):
    """
    Grava o log de ações administrativas executadas por supervisores.
    Requer o cursor ativo para participar da mesma transação da rota.
    """
    empresa_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    ip_origem = request.remote_addr

    if not empresa_id or not usuario_id:
        return  # Proteção caso seja disparado fora de uma sessão válida

    query = """
            INSERT INTO auditoria_supervisor
                (empresa_id, usuario_id, data_acao, tipo_acao, descricao, ip_origem)
            VALUES (%s, %s, NOW(), %s, %s, %s) \
            """
    cursor.execute(query, (empresa_id, usuario_id, tipo_acao, descricao, ip_origem))


def salvar_selfie_base64(selfie_base64, empresa_id, motorista_id, escala_id, cur=None):
    """
    Salva selfie enviada em base64 dentro de static/uploads/selfies_base.
    Retorna o caminho relativo para exibição no sistema.
    """
    import base64
    import uuid

    if not selfie_base64:
        return None

    if ',' in selfie_base64:
        selfie_base64 = selfie_base64.split(',', 1)[1]

    try:
        dados = base64.b64decode(selfie_base64)
    except Exception:
        return None

    if not dados:
        return None

    pasta = os.path.join(app.root_path, 'static', 'uploads', 'selfies_base')
    os.makedirs(pasta, exist_ok=True)

    nome_arquivo = f"selfie_empresa_{empresa_id}_motorista_{motorista_id}_escala_{escala_id}_{uuid.uuid4().hex[:10]}.jpg"
    caminho_absoluto = os.path.join(pasta, nome_arquivo)

    with open(caminho_absoluto, 'wb') as arquivo:
        arquivo.write(dados)

    caminho_relativo = f"uploads/selfies_base/{nome_arquivo}"

    if cur is not None:
        return tentar_enviar_arquivo_google_drive(
            cur,
            caminho_absoluto,
            caminho_relativo,
            empresa_id=empresa_id,
            motorista_id=motorista_id,
            origem='CHECKIN_SELFIE',
            origem_id=escala_id,
            tipo_arquivo='SELFIE_CHECKIN',
            nome_original=nome_arquivo,
            mime_type='image/jpeg',
            criado_por_usuario_id=session.get('usuario_id'),
        )

    return caminho_relativo


def salvar_anexo_justificativa(arquivo, empresa_id, motorista_id, escala_id, cur=None):
    """
    Salva anexo opcional da justificativa de ausência.
    Retorna caminho relativo para exibição/consulta no sistema.
    """
    if not arquivo or not getattr(arquivo, 'filename', ''):
        return None

    nome_original = arquivo.filename or ''
    extensao = os.path.splitext(nome_original)[1].lower()
    extensoes_permitidas = {'.jpg', '.jpeg', '.png', '.pdf', '.webp'}

    if extensao not in extensoes_permitidas:
        return None

    pasta = os.path.join(app.root_path, 'static', 'uploads', 'justificativas_ausencia')
    os.makedirs(pasta, exist_ok=True)

    nome_arquivo = f"justificativa_empresa_{empresa_id}_motorista_{motorista_id}_escala_{escala_id}_{uuid.uuid4().hex[:10]}{extensao}"
    caminho_absoluto = os.path.join(pasta, nome_arquivo)
    arquivo.save(caminho_absoluto)

    caminho_relativo = f"uploads/justificativas_ausencia/{nome_arquivo}"

    if cur is not None:
        return tentar_enviar_arquivo_google_drive(
            cur,
            caminho_absoluto,
            caminho_relativo,
            empresa_id=empresa_id,
            motorista_id=motorista_id,
            origem='JUSTIFICATIVA_AUSENCIA',
            origem_id=escala_id,
            tipo_arquivo='ANEXO_JUSTIFICATIVA_AUSENCIA',
            nome_original=nome_original,
            mime_type=getattr(arquivo, 'mimetype', None),
            criado_por_usuario_id=session.get('usuario_id'),
        )

    return caminho_relativo


def gerar_codigo_qr_base(nome_base, empresa_id):
    import uuid
    base_limpa = re.sub(r'[^A-Za-z0-9]+', '', str(nome_base or 'BASE')).upper()[:10]
    return f"BASE-{empresa_id}-{base_limpa}-{uuid.uuid4().hex[:8].upper()}"



# ==========================================================
# APOIO OPERACIONAL - SCORE NA FILA DE CANCELADOS
# ==========================================================
def classificar_score_operacional(score):
    """Classificação visual do score usada em telas operacionais."""
    try:
        score_int = int(round(float(score or 0)))
    except Exception:
        score_int = 0

    if score_int >= 90:
        return {
            'nome': 'Excelente',
            'classe': 'success',
            'icone': 'fa-circle-check',
            'descricao': 'Histórico muito positivo no período.'
        }
    if score_int >= 80:
        return {
            'nome': 'Confiável',
            'classe': 'primary',
            'icone': 'fa-thumbs-up',
            'descricao': 'Histórico positivo para acompanhamento operacional.'
        }
    if score_int >= 65:
        return {
            'nome': 'Atenção',
            'classe': 'warning',
            'icone': 'fa-triangle-exclamation',
            'descricao': 'Exige acompanhamento do supervisor.'
        }
    return {
        'nome': 'Risco',
        'classe': 'danger',
        'icone': 'fa-circle-exclamation',
        'descricao': 'Histórico crítico no período.'
    }


def calcular_scores_motoristas_para_fila(empresa_id, motorista_ids, data_referencia, base_operacional_id=None):
    """
    Calcula score resumido para motoristas que estão na fila de cancelados.
    Importante: este score NÃO altera a ordem da fila; a fila continua por ordem de chegada.
    Ele serve apenas como apoio visual para ciência do supervisor.
    """
    motorista_ids = [int(mid) for mid in set(motorista_ids or []) if mid]
    if not empresa_id or not motorista_ids:
        return {}

    try:
        if hasattr(data_referencia, 'strftime'):
            data_fim = data_referencia
        else:
            data_fim = datetime.strptime(str(data_referencia)[:10], '%Y-%m-%d').date()
    except Exception:
        data_fim = date.today()

    data_inicio = data_fim.replace(day=1)
    placeholders = ','.join(['%s'] * len(motorista_ids))

    def inteiro(valor):
        try:
            return int(valor or 0)
        except Exception:
            return 0

    def limitar(valor, minimo=0, maximo=100):
        return max(minimo, min(maximo, int(round(valor))))

    mapa = {}
    for motorista_id in motorista_ids:
        mapa[motorista_id] = {
            'motorista_id': motorista_id,
            'total_escalas': 0,
            'confirmados_rota': 0,
            'cancelados_base': 0,
            'chegadas_confirmadas': 0,
            'faltas_escala': 0,
            'nao_compareceu': 0,
            'total_fila': 0,
            'rota_extra': 0,
            'dispensado': 0,
            'fila_aguardando': 0,
            'fila_falta': 0,
            'checkins_total': 0,
            'checkins_aprovados': 0,
            'checkins_bloqueados': 0,
            'bloqueios_distancia': 0,
            'bloqueios_qr': 0,
            'bloqueios_selfie': 0,
            'bloqueios_geo': 0,
            'bloqueios_base_sem_coord': 0,
            'ciencias': 0,
            'disponivel': 0,
            'ausente': 0,
            'sem_resposta': 0,
            'pontos_positivos': 0,
            'pontos_negativos': 0,
            'score': 80,
            'classificacao': classificar_score_operacional(80),
            'motivos': ['Sem movimentação crítica no período'],
            'periodo_label': f"{formatar_data_br(data_inicio)} até {formatar_data_br(data_fim)}"
        }

    con = obter_conexao()
    if con is None:
        return mapa

    cur = con.cursor(dictionary=True)
    try:
        filtro_base_em = ''
        filtro_base_f = ''
        filtro_base_aud = ''
        params_base_em = []
        params_base_f = []
        params_base_aud = []
        if base_operacional_id:
            filtro_base_em = ' AND em.base_operacional_id = %s '
            filtro_base_f = ' AND f.base_operacional_id = %s '
            filtro_base_aud = ' AND aud.base_operacional_id = %s '
            params_base_em.append(base_operacional_id)
            params_base_f.append(base_operacional_id)
            params_base_aud.append(base_operacional_id)

        cur.execute(f"""
            SELECT em.motorista_id,
                   COUNT(*) AS total_escalas,
                   SUM(CASE WHEN em.status_escala = 'Confirmado com rota' THEN 1 ELSE 0 END) AS confirmados_rota,
                   SUM(CASE WHEN em.status_escala = 'Cancelado, comparecer na base' THEN 1 ELSE 0 END) AS cancelados_base,
                   SUM(CASE WHEN em.status_presenca = 'Chegada confirmada' THEN 1 ELSE 0 END) AS chegadas_confirmadas,
                   SUM(CASE WHEN em.status_escala = 'Falta' THEN 1 ELSE 0 END) AS faltas_escala,
                   SUM(CASE WHEN em.status_presenca = 'Não compareceu' THEN 1 ELSE 0 END) AS nao_compareceu
            FROM escala_motorista em
            WHERE em.empresa_id = %s
              AND em.data_escala BETWEEN %s AND %s
              AND em.motorista_id IN ({placeholders})
              {filtro_base_em}
            GROUP BY em.motorista_id
        """, [empresa_id, data_inicio, data_fim] + motorista_ids + params_base_em)
        for row in cur.fetchall():
            alvo = mapa.get(row.get('motorista_id'))
            if alvo:
                for k in ['total_escalas', 'confirmados_rota', 'cancelados_base', 'chegadas_confirmadas', 'faltas_escala', 'nao_compareceu']:
                    alvo[k] = inteiro(row.get(k))

        cur.execute(f"""
            SELECT f.motorista_id,
                   COUNT(*) AS total_fila,
                   SUM(CASE WHEN f.status_fila = 'Atribuído para rota extra' THEN 1 ELSE 0 END) AS rota_extra,
                   SUM(CASE WHEN f.status_fila = 'Dispensado' THEN 1 ELSE 0 END) AS dispensado,
                   SUM(CASE WHEN f.status_fila = 'Aguardando rota' THEN 1 ELSE 0 END) AS fila_aguardando,
                   SUM(CASE WHEN f.status_fila = 'Falta' THEN 1 ELSE 0 END) AS fila_falta
            FROM fila_cancelados_base f
            WHERE f.empresa_id = %s
              AND f.data_fila BETWEEN %s AND %s
              AND f.motorista_id IN ({placeholders})
              {filtro_base_f}
            GROUP BY f.motorista_id
        """, [empresa_id, data_inicio, data_fim] + motorista_ids + params_base_f)
        for row in cur.fetchall():
            alvo = mapa.get(row.get('motorista_id'))
            if alvo:
                for k in ['total_fila', 'rota_extra', 'dispensado', 'fila_aguardando', 'fila_falta']:
                    alvo[k] = inteiro(row.get(k))

        cur.execute(f"""
            SELECT aud.motorista_id,
                   COUNT(*) AS checkins_total,
                   SUM(CASE WHEN aud.resultado = 'Aprovado' THEN 1 ELSE 0 END) AS checkins_aprovados,
                   SUM(CASE WHEN aud.resultado <> 'Aprovado' THEN 1 ELSE 0 END) AS checkins_bloqueados,
                   SUM(CASE WHEN aud.resultado = 'Bloqueado por distância' THEN 1 ELSE 0 END) AS bloqueios_distancia,
                   SUM(CASE WHEN aud.resultado = 'Bloqueado por QR inválido/expirado' THEN 1 ELSE 0 END) AS bloqueios_qr,
                   SUM(CASE WHEN aud.resultado = 'Bloqueado por falta de selfie' THEN 1 ELSE 0 END) AS bloqueios_selfie,
                   SUM(CASE WHEN aud.resultado = 'Bloqueado por falta de geolocalização' THEN 1 ELSE 0 END) AS bloqueios_geo,
                   SUM(CASE WHEN aud.resultado = 'Bloqueado por base sem coordenadas' THEN 1 ELSE 0 END) AS bloqueios_base_sem_coord
            FROM auditoria_checkin_base aud
            WHERE aud.empresa_id = %s
              AND DATE(aud.data_tentativa) BETWEEN %s AND %s
              AND aud.motorista_id IN ({placeholders})
              {filtro_base_aud}
            GROUP BY aud.motorista_id
        """, [empresa_id, data_inicio, data_fim] + motorista_ids + params_base_aud)
        for row in cur.fetchall():
            alvo = mapa.get(row.get('motorista_id'))
            if alvo:
                for k in ['checkins_total', 'checkins_aprovados', 'checkins_bloqueados', 'bloqueios_distancia', 'bloqueios_qr', 'bloqueios_selfie', 'bloqueios_geo', 'bloqueios_base_sem_coord']:
                    alvo[k] = inteiro(row.get(k))

        cur.execute(f"""
            SELECT c.motorista_id,
                   COUNT(*) AS ciencias
            FROM ciencia_escala_motorista c
            INNER JOIN escala_motorista em
                    ON em.id = c.escala_id
                   AND em.empresa_id = c.empresa_id
            WHERE c.empresa_id = %s
              AND em.data_escala BETWEEN %s AND %s
              AND c.motorista_id IN ({placeholders})
              {filtro_base_em}
            GROUP BY c.motorista_id
        """, [empresa_id, data_inicio, data_fim] + motorista_ids + params_base_em)
        for row in cur.fetchall():
            alvo = mapa.get(row.get('motorista_id'))
            if alvo:
                alvo['ciencias'] = inteiro(row.get('ciencias'))

        cur.execute(f"""
            SELECT dm.motorista_id,
                   SUM(CASE WHEN dm.status_disponibilidade = 'Disponível' THEN 1 ELSE 0 END) AS disponivel,
                   SUM(CASE WHEN dm.status_disponibilidade = 'Ausente' THEN 1 ELSE 0 END) AS ausente,
                   SUM(CASE WHEN dm.status_disponibilidade = 'Sem resposta' THEN 1 ELSE 0 END) AS sem_resposta
            FROM disponibilidade_motorista dm
            WHERE dm.empresa_id = %s
              AND dm.data_disponibilidade BETWEEN %s AND %s
              AND dm.motorista_id IN ({placeholders})
            GROUP BY dm.motorista_id
        """, [empresa_id, data_inicio, data_fim] + motorista_ids)
        for row in cur.fetchall():
            alvo = mapa.get(row.get('motorista_id'))
            if alvo:
                alvo['disponivel'] = inteiro(row.get('disponivel'))
                alvo['ausente'] = inteiro(row.get('ausente'))
                alvo['sem_resposta'] = inteiro(row.get('sem_resposta'))

        for item in mapa.values():
            positivos = 0
            negativos = 0
            motivos = []

            positivos += min(item['checkins_aprovados'] * 3, 12)
            positivos += min(item['rota_extra'] * 8, 16)
            positivos += min(item['confirmados_rota'] * 2, 10)
            positivos += min(item['ciencias'] * 1, 5)
            positivos += min(item['disponivel'] * 1, 6)

            negativos += item['faltas_escala'] * 20
            negativos += item['nao_compareceu'] * 20
            negativos += item['fila_falta'] * 15
            negativos += item['bloqueios_distancia'] * 15
            negativos += item['bloqueios_qr'] * 10
            negativos += (item['bloqueios_selfie'] + item['bloqueios_geo'] + item['bloqueios_base_sem_coord']) * 8
            negativos += item['ausente'] * 1
            negativos += item['sem_resposta'] * 2

            score = limitar(80 + positivos - negativos)
            item['pontos_positivos'] = positivos
            item['pontos_negativos'] = negativos
            item['score'] = score
            item['classificacao'] = classificar_score_operacional(score)

            if item['rota_extra']:
                motivos.append(f"{item['rota_extra']} rota(s) extra atribuída(s)")
            if item['checkins_aprovados']:
                motivos.append(f"{item['checkins_aprovados']} check-in(s) aprovado(s)")
            if item['faltas_escala'] or item['nao_compareceu']:
                motivos.append(f"{item['faltas_escala'] + item['nao_compareceu']} falta(s)/não comparecimento(s)")
            if item['checkins_bloqueados']:
                motivos.append(f"{item['checkins_bloqueados']} bloqueio(s) no check-in")
            if not motivos:
                motivos.append('Sem movimentação crítica no período')
            item['motivos'] = motivos[:3]

        return mapa

    except Exception as e:
        print(f"Erro ao calcular score da fila de cancelados: {e}")
        return mapa
    finally:
        cur.close()
        con.close()

# ==========================================================
# FASE 4.1.3 - CONFIRMAÇÃO DE CHEGADA + FILA DE CANCELADOS
# ==========================================================
def carregar_fila_cancelados_base(empresa_id, data_fila, base_operacional_id=None):
    con = obter_conexao()

    if con is None:
        return []

    cur = con.cursor(dictionary=True)

    try:
        sql = """
              SELECT f.id,
                     f.empresa_id,
                     f.escala_id,
                     f.motorista_id,
                     mot.nome_completo                       AS motorista_nome,
                     mot.cpf_cnpj                            AS motorista_cpf_cnpj,
                     f.base_operacional_id,
                     COALESCE(bo.nome_base, f.base_operacao) AS base_nome,
                     f.data_fila,
                     f.hora_confirmacao,
                     f.posicao_fila,
                     f.status_fila,
                     f.confirmado_por,
                     f.usuario_confirmacao_id,
                     f.latitude_confirmacao,
                     f.longitude_confirmacao,
                     f.distancia_base_metros,
                     f.geolocalizacao_validada,
                     f.qr_code_validado,
                     f.qr_token_id,
                     f.selfie_path,
                     f.observacao,
                     f.data_atualizacao,
                     em.status_escala,
                     em.status_presenca,
                     em.horario_apresentacao
              FROM fila_cancelados_base f
                       INNER JOIN pessoas mot
                                  ON mot.id = f.motorista_id
                                      AND mot.empresa_id = f.empresa_id
                       LEFT JOIN bases_operacionais bo
                                 ON bo.id = f.base_operacional_id
                                     AND bo.empresa_id = f.empresa_id
                       LEFT JOIN escala_motorista em
                                 ON em.id = f.escala_id
                                     AND em.empresa_id = f.empresa_id
              WHERE f.empresa_id = %s
                AND f.data_fila = %s \
              """
        params = [empresa_id, data_fila]

        if base_operacional_id:
            sql += " AND f.base_operacional_id = %s"
            params.append(base_operacional_id)

        sql += """
            ORDER BY
                CASE
                    WHEN f.status_fila = 'Aguardando rota' THEN 1
                    WHEN f.status_fila = 'Atribuído para rota extra' THEN 2
                    WHEN f.status_fila = 'Dispensado' THEN 3
                    ELSE 4
                END,
                f.hora_confirmacao ASC,
                f.id ASC
        """

        cur.execute(sql, params)
        filas = cur.fetchall()

        motorista_ids = [linha.get('motorista_id') for linha in filas if linha.get('motorista_id')]
        scores_fila = calcular_scores_motoristas_para_fila(
            empresa_id=empresa_id,
            motorista_ids=motorista_ids,
            data_referencia=data_fila,
            base_operacional_id=base_operacional_id
        )

        for linha in filas:
            score_info = scores_fila.get(linha.get('motorista_id')) or {}
            linha['score_motorista'] = score_info.get('score', 80)
            linha['score_classificacao'] = score_info.get('classificacao') or classificar_score_operacional(80)
            linha['score_motivos'] = score_info.get('motivos') or ['Sem movimentação crítica no período']
            linha['score_periodo_label'] = score_info.get('periodo_label', '')
            linha['score_rota_extra'] = score_info.get('rota_extra', 0)
            linha['score_faltas'] = (score_info.get('faltas_escala', 0) or 0) + (score_info.get('nao_compareceu', 0) or 0) + (score_info.get('fila_falta', 0) or 0)
            linha['score_bloqueios'] = score_info.get('checkins_bloqueados', 0)
            linha['score_checkins_aprovados'] = score_info.get('checkins_aprovados', 0)

        return filas

    except Exception as e:
        print(f"Erro ao carregar fila de cancelados: {e}")
        return []

    finally:
        cur.close()
        con.close()


def proxima_posicao_fila_cancelados(cur, empresa_id, data_fila, base_operacional_id):
    cur.execute("""
                SELECT COALESCE(MAX(posicao_fila), 0) + 1 AS proxima_posicao
                FROM fila_cancelados_base
                WHERE empresa_id = %s
                  AND data_fila = %s
                  AND (
                    (base_operacional_id = %s)
                        OR (base_operacional_id IS NULL AND %s IS NULL)
                    )
                """, (empresa_id, data_fila, base_operacional_id, base_operacional_id))

    row = cur.fetchone() or {}
    return int(row.get('proxima_posicao') or 1)


def atualizar_status_fila_cancelado(fila_id, novo_status, observacao_padrao):
    empresa_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('escala_motoristas'))

    cur = con.cursor(dictionary=True)

    try:
        if is_super_admin:
            cur.execute("""
                        SELECT f.*, em.data_escala
                        FROM fila_cancelados_base f
                                 INNER JOIN escala_motorista em ON em.id = f.escala_id AND em.empresa_id = f.empresa_id
                        WHERE f.id = %s LIMIT 1
                        """, (fila_id,))
        else:
            cur.execute("""
                        SELECT f.*, em.data_escala
                        FROM fila_cancelados_base f
                                 INNER JOIN escala_motorista em ON em.id = f.escala_id AND em.empresa_id = f.empresa_id
                        WHERE f.id = %s
                          AND f.empresa_id = %s LIMIT 1
                        """, (fila_id, empresa_id))

        fila = cur.fetchone()

        if not fila:
            flash("Registro da fila não encontrado ou sem permissão.", "danger")
            return redirect(url_for('escala_motoristas'))

        data_escala = fila.get('data_escala')
        empresa_fila_id = fila.get('empresa_id')

        cur.execute("""
                    UPDATE fila_cancelados_base
                    SET status_fila      = %s,
                        observacao       = %s,
                        usuario_acao_id  = %s,
                        data_atualizacao = NOW()
                    WHERE id = %s
                    """, (
                        novo_status,
                        observacao_padrao,
                        session.get('usuario_id'),
                        fila_id
                    ))

        if novo_status == 'Atribuído para rota extra':
            cur.execute("""
                        UPDATE escala_motorista
                        SET status_escala         = 'Confirmado com rota',
                            status_presenca       = 'Chegada confirmada',
                            observacao_supervisor = CONCAT(
                                    COALESCE(observacao_supervisor, ''),
                                    CASE WHEN COALESCE(observacao_supervisor, '') = '' THEN '' ELSE ' | ' END,
                                    'Puxado da fila de cancelados para rota extra.'
                                                    ),
                            data_atualizacao      = NOW()
                        WHERE id = %s
                          AND empresa_id = %s
                        """, (fila.get('escala_id'), empresa_fila_id))

        elif novo_status == 'Dispensado':
            cur.execute("""
                        UPDATE escala_motorista
                        SET observacao_supervisor = CONCAT(
                                COALESCE(observacao_supervisor, ''),
                                CASE WHEN COALESCE(observacao_supervisor, '') = '' THEN '' ELSE ' | ' END,
                                'Motorista compareceu à base e foi dispensado.'
                                                    ),
                            data_atualizacao      = NOW()
                        WHERE id = %s
                          AND empresa_id = %s
                        """, (fila.get('escala_id'), empresa_fila_id))

        con.commit()
        flash("Fila de cancelados atualizada com sucesso.", "success")

        return redirect(url_for(
            'escala_motoristas',
            data_escala=data_escala.strftime('%Y-%m-%d') if hasattr(data_escala, 'strftime') else str(data_escala)[:10],
            empresa_id=empresa_fila_id if is_super_admin else ''
        ))

    except Exception as e:
        con.rollback()
        print(f"Erro ao atualizar fila de cancelados: {e}")
        flash(f"Erro técnico ao atualizar fila: {e}", "danger")
        return redirect(url_for('escala_motoristas'))

    finally:
        cur.close()
        con.close()


@app.route('/operacao/fila-cancelados/<int:fila_id>/atribuir-rota', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor')
def atribuir_rota_extra_fila_cancelado(fila_id):
    return atualizar_status_fila_cancelado(
        fila_id,
        'Atribuído para rota extra',
        'Motorista puxado da fila para rota extra.'
    )


@app.route('/operacao/fila-cancelados/<int:fila_id>/dispensar', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor')
def dispensar_fila_cancelado(fila_id):
    return atualizar_status_fila_cancelado(
        fila_id,
        'Dispensado',
        'Motorista compareceu à base e foi dispensado.'
    )


# ==========================================================
# FASE 4.1.2.1 - MINHA SEMANA NO PORTAL DO MOTORISTA
# ==========================================================
def carregar_minha_semana_motorista(empresa_id, motorista_id, data_inicio, data_fim):
    """
    Carrega a visão semanal do motorista:
    - disponibilidade informada
    - escala definida pelo supervisor
    - status de presença
    - ciência da escala
    """
    con = obter_conexao()

    if con is None:
        return []

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT calendario.data_ref                                 AS data_ref,

                           COALESCE(dm.status_disponibilidade, 'Sem resposta') AS status_disponibilidade,
                           dm.observacao                                       AS observacao_disponibilidade,
                           dm.origem_lancamento                                AS origem_disponibilidade,

                           em.id                                               AS escala_id,
                           COALESCE(em.status_escala, 'Pendente')              AS status_escala,
                           COALESCE(em.status_presenca, 'Não se aplica')       AS status_presenca,
                           em.base_operacional_id,
                           em.base_operacao,
                           bo.nome_base                                        AS base_operacional_nome,
                           em.horario_apresentacao,
                           em.observacao_supervisor,
                           em.presenca_confirmada_em,
                           em.falta_automatica,
                           em.falta_marcada_em,
                           em.falta_motivo,

                           cem.data_ciencia,
                           cem.origem_ciencia,

                           fcb.id                                              AS fila_id,
                           fcb.posicao_fila,
                           fcb.status_fila,
                           fcb.hora_confirmacao                                AS hora_confirmacao_fila,

                           jam.id                                              AS justificativa_id,
                           jam.motivo                                          AS justificativa_motivo,
                           jam.observacao_motorista                            AS justificativa_observacao,
                           jam.status_justificativa,
                           jam.anexo_path                                      AS justificativa_anexo_path,
                           jam.data_envio                                      AS justificativa_data_envio,
                           jam.observacao_supervisor                           AS justificativa_observacao_supervisor

                    FROM (SELECT %s AS data_ref
                          UNION ALL
                          SELECT DATE_ADD(%s, INTERVAL 1 DAY)
                          UNION ALL
                          SELECT DATE_ADD(%s, INTERVAL 2 DAY)
                          UNION ALL
                          SELECT DATE_ADD(%s, INTERVAL 3 DAY)
                          UNION ALL
                          SELECT DATE_ADD(%s, INTERVAL 4 DAY)
                          UNION ALL
                          SELECT DATE_ADD(%s, INTERVAL 5 DAY)
                          UNION ALL
                          SELECT DATE_ADD(%s, INTERVAL 6 DAY)) calendario

                             LEFT JOIN disponibilidade_motorista dm
                                       ON dm.empresa_id = %s
                                           AND dm.motorista_id = %s
                                           AND dm.data_disponibilidade = calendario.data_ref

                             LEFT JOIN escala_motorista em
                                       ON em.empresa_id = %s
                                           AND em.motorista_id = %s
                                           AND em.data_escala = calendario.data_ref

                             LEFT JOIN bases_operacionais bo
                                       ON bo.id = em.base_operacional_id
                                           AND bo.empresa_id = em.empresa_id

                             LEFT JOIN fila_cancelados_base fcb
                                       ON fcb.empresa_id = em.empresa_id
                                           AND fcb.escala_id = em.id
                                           AND fcb.motorista_id = em.motorista_id

                             LEFT JOIN ciencia_escala_motorista cem
                                       ON cem.empresa_id = %s
                                           AND cem.motorista_id = %s
                                           AND cem.escala_id = em.id

                             LEFT JOIN justificativas_ausencia_motorista jam
                                       ON jam.empresa_id = em.empresa_id
                                           AND jam.escala_id = em.id
                                           AND jam.motorista_id = em.motorista_id

                    ORDER BY calendario.data_ref ASC
                    """, (
                        data_inicio, data_inicio, data_inicio, data_inicio, data_inicio, data_inicio, data_inicio,
                        empresa_id, motorista_id,
                        empresa_id, motorista_id,
                        empresa_id, motorista_id
                    ))

        rows = cur.fetchall()

        dias_semana_nomes = [
            'Segunda-feira',
            'Terça-feira',
            'Quarta-feira',
            'Quinta-feira',
            'Sexta-feira',
            'Sábado',
            'Domingo'
        ]

        retorno = []

        for row in rows:
            data_ref = row.get('data_ref')
            if isinstance(data_ref, str):
                data_ref_date = datetime.strptime(data_ref[:10], '%Y-%m-%d').date()
            else:
                data_ref_date = data_ref

            status_escala = row.get('status_escala') or 'Pendente'
            status_presenca = row.get('status_presenca') or 'Não se aplica'

            precisa_ciencia = status_escala not in ['Pendente']
            precisa_presenca = status_escala == 'Cancelado, comparecer na base'

            orientacao = "Aguardando definição do supervisor."

            if status_escala == 'Confirmado com rota':
                orientacao = "Você está confirmado com rota. Acompanhe as orientações operacionais e compareça conforme combinado."
            elif status_escala == 'Cancelado, comparecer na base':
                orientacao = "Você foi cancelado na rota, mas deve comparecer à base para possível encaixe avulso. A confirmação de chegada será obrigatória."
            elif status_escala == 'Cancelado definitivo':
                orientacao = "Você foi cancelado definitivamente para esta data. Não é necessário comparecer à base."
            elif status_escala == 'Reserva / Avulso':
                orientacao = "Você está como reserva/avulso. Aguarde orientação do supervisor."
            elif status_escala == 'Falta':
                orientacao = "Registro de falta para esta data. Em caso de divergência, procure o supervisor."

            retorno.append({
                'data': data_ref_date,
                'data_str': data_ref_date.strftime('%Y-%m-%d'),
                'dia_nome': dias_semana_nomes[data_ref_date.weekday()],
                'status_disponibilidade': row.get('status_disponibilidade') or 'Sem resposta',
                'observacao_disponibilidade': row.get('observacao_disponibilidade') or '',
                'origem_disponibilidade': row.get('origem_disponibilidade') or '-',
                'escala_id': row.get('escala_id'),
                'status_escala': status_escala,
                'status_presenca': status_presenca,
                'base_operacional_id': row.get('base_operacional_id'),
                'base_operacao': row.get('base_operacional_nome') or row.get('base_operacao') or '',
                'horario_apresentacao': row.get('horario_apresentacao') or '',
                'observacao_supervisor': row.get('observacao_supervisor') or '',
                'presenca_confirmada_em': row.get('presenca_confirmada_em'),
                'falta_automatica': row.get('falta_automatica') or 'N',
                'falta_marcada_em': row.get('falta_marcada_em'),
                'falta_motivo': row.get('falta_motivo') or '',
                'data_ciencia': row.get('data_ciencia'),
                'origem_ciencia': row.get('origem_ciencia') or '',
                'fila_id': row.get('fila_id'),
                'posicao_fila': row.get('posicao_fila'),
                'status_fila': row.get('status_fila') or '',
                'hora_confirmacao_fila': row.get('hora_confirmacao_fila'),
                'justificativa_id': row.get('justificativa_id'),
                'justificativa_motivo': row.get('justificativa_motivo') or '',
                'justificativa_observacao': row.get('justificativa_observacao') or '',
                'status_justificativa': row.get('status_justificativa') or '',
                'justificativa_anexo_path': row.get('justificativa_anexo_path') or '',
                'justificativa_data_envio': row.get('justificativa_data_envio'),
                'justificativa_observacao_supervisor': row.get('justificativa_observacao_supervisor') or '',
                'precisa_ciencia': precisa_ciencia,
                'precisa_presenca': precisa_presenca,
                'orientacao': orientacao
            })

        return retorno

    except Exception as e:
        print(f"Erro ao carregar minha semana do motorista: {e}")
        return []

    finally:
        cur.close()
        con.close()


@app.route('/portal-motorista/minha-semana', methods=['GET', 'POST'])
@login_required
@motorista_required
def minha_semana_motorista():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_id = session.get('empresa_id')
    motorista_id = session.get('pessoa_id')

    if not empresa_id or not motorista_id:
        flash("Sessão do motorista incompleta. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    if request.method == 'POST':
        acao = request.form.get('acao', '').strip()
        escala_id = request.form.get('escala_id', '').strip()

        latitude_confirmacao_form = request.form.get('latitude_confirmacao', '').strip()
        longitude_confirmacao_form = request.form.get('longitude_confirmacao', '').strip()
        codigo_qr_informado = request.form.get('codigo_qr_base', '').strip()
        selfie_base64 = request.form.get('selfie_base64', '').strip()

        if acao == 'justificar_ausencia' and escala_id.isdigit():
            motivo = request.form.get('motivo_ausencia', '').strip()
            observacao_motorista = request.form.get('observacao_ausencia', '').strip()
            anexo = request.files.get('anexo_justificativa')

            if not motivo:
                flash("Informe o motivo da ausência para enviar a justificativa.", "warning")
                return redirect(url_for('minha_semana_motorista'))

            if not observacao_motorista:
                flash("Descreva brevemente o que aconteceu para que o supervisor possa analisar.", "warning")
                return redirect(url_for('minha_semana_motorista'))

            con = obter_conexao()

            if con is None:
                flash("Erro de conexão com o banco de dados.", "danger")
                return redirect(url_for('minha_semana_motorista'))

            cur = con.cursor(dictionary=True)

            try:
                cur.execute("""
                            SELECT em.id,
                                   em.empresa_id,
                                   em.motorista_id,
                                   em.data_escala,
                                   em.status_escala,
                                   em.status_presenca,
                                   em.base_operacional_id,
                                   em.base_operacao,
                                   em.horario_apresentacao,
                                   bo.nome_base AS base_nome
                            FROM escala_motorista em
                            LEFT JOIN bases_operacionais bo
                                   ON bo.id = em.base_operacional_id
                                  AND bo.empresa_id = em.empresa_id
                            WHERE em.id = %s
                              AND em.empresa_id = %s
                              AND em.motorista_id = %s
                            LIMIT 1
                            """, (int(escala_id), empresa_id, motorista_id))

                escala = cur.fetchone()

                if not escala:
                    flash("Escala não encontrada para seu usuário.", "danger")
                    return redirect(url_for('minha_semana_motorista'))

                if escala.get('status_presenca') == 'Chegada confirmada':
                    flash("Sua chegada já foi confirmada. Não é necessário justificar ausência.", "info")
                    return redirect(url_for('minha_semana_motorista'))

                if escala.get('status_escala') not in ['Cancelado, comparecer na base', 'Falta', 'Reserva / Avulso'] and escala.get('status_presenca') != 'Não compareceu':
                    flash("A justificativa de ausência está disponível apenas para escalas com pendência de presença.", "warning")
                    return redirect(url_for('minha_semana_motorista'))

                anexo_path = salvar_anexo_justificativa(
                    anexo,
                    empresa_id,
                    motorista_id,
                    int(escala_id),
                    cur=cur
                )

                cur.execute("""
                            INSERT INTO justificativas_ausencia_motorista
                                (empresa_id,
                                 escala_id,
                                 motorista_id,
                                 base_operacional_id,
                                 data_escala,
                                 horario_previsto,
                                 motivo,
                                 observacao_motorista,
                                 anexo_path,
                                 status_justificativa,
                                 data_envio,
                                 created_at,
                                 updated_at)
                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s,
                                    'Pendente de análise', NOW(), NOW(), NOW())
                            ON DUPLICATE KEY UPDATE
                                motivo = VALUES(motivo),
                                observacao_motorista = VALUES(observacao_motorista),
                                anexo_path = COALESCE(VALUES(anexo_path), anexo_path),
                                status_justificativa = 'Pendente de análise',
                                observacao_supervisor = NULL,
                                usuario_analise_id = NULL,
                                data_analise = NULL,
                                data_envio = NOW(),
                                updated_at = NOW()
                            """, (
                                empresa_id,
                                int(escala_id),
                                motorista_id,
                                escala.get('base_operacional_id'),
                                escala.get('data_escala'),
                                escala.get('horario_apresentacao'),
                                motivo,
                                observacao_motorista,
                                anexo_path
                            ))

                con.commit()
                flash("Justificativa enviada com sucesso. Ela ficará pendente para análise do supervisor.", "success")

                return redirect(url_for(
                    'minha_semana_motorista',
                    semana=escala['data_escala'].strftime('%Y-%m-%d') if hasattr(escala['data_escala'], 'strftime') else str(escala['data_escala'])[:10]
                ))

            except Exception as e:
                con.rollback()
                print(f"Erro ao enviar justificativa de ausência: {e}")
                flash(f"Erro técnico ao enviar justificativa: {e}", "danger")
                return redirect(url_for('minha_semana_motorista'))

            finally:
                cur.close()
                con.close()

        if acao == 'confirmar_ciencia' and escala_id.isdigit():
            con = obter_conexao()

            if con is None:
                flash("Erro de conexão com o banco de dados.", "danger")
                return redirect(url_for('minha_semana_motorista'))

            cur = con.cursor(dictionary=True)

            try:
                cur.execute("""
                            SELECT id, data_escala, status_escala
                            FROM escala_motorista
                            WHERE id = %s
                              AND empresa_id = %s
                              AND motorista_id = %s LIMIT 1
                            """, (int(escala_id), empresa_id, motorista_id))

                escala = cur.fetchone()

                if not escala:
                    flash("Escala não encontrada para seu usuário.", "danger")
                    return redirect(url_for('minha_semana_motorista'))

                cur.execute("""
                            INSERT INTO ciencia_escala_motorista (empresa_id,
                                                                  motorista_id,
                                                                  escala_id,
                                                                  data_ciencia,
                                                                  origem_ciencia,
                                                                  usuario_id,
                                                                  created_at,
                                                                  updated_at)
                            VALUES (%s, %s, %s, NOW(), 'Motorista', %s, NOW(), NOW()) ON DUPLICATE KEY
                            UPDATE
                                data_ciencia = NOW(),
                                origem_ciencia = 'Motorista',
                                usuario_id =
                            VALUES (usuario_id), updated_at = NOW()
                            """, (
                                empresa_id,
                                motorista_id,
                                int(escala_id),
                                session.get('usuario_id')
                            ))

                con.commit()
                flash("Ciência da escala confirmada com sucesso.", "success")
                return redirect(url_for(
                    'minha_semana_motorista',
                    semana=escala['data_escala'].strftime('%Y-%m-%d') if hasattr(escala['data_escala'],
                                                                                 'strftime') else str(
                        escala['data_escala'])[:10]
                ))

            except Exception as e:
                con.rollback()
                print(f"Erro ao confirmar ciência da escala: {e}")
                flash(f"Erro técnico ao confirmar ciência: {e}", "danger")
                return redirect(url_for('minha_semana_motorista'))

            finally:
                cur.close()
                con.close()

        if acao == 'confirmar_chegada_base' and escala_id.isdigit():
            con = obter_conexao()

            if con is None:
                flash("Erro de conexão com o banco de dados.", "danger")
                return redirect(url_for('minha_semana_motorista'))

            cur = con.cursor(dictionary=True)

            try:
                cur.execute("""
                            SELECT em.id,
                                   em.empresa_id,
                                   em.motorista_id,
                                   em.data_escala,
                                   em.status_escala,
                                   em.status_presenca,
                                   em.base_operacional_id,
                                   em.base_operacao,
                                   em.horario_apresentacao,
                                   bo.nome_base AS base_nome,
                                   bo.latitude  AS base_latitude,
                                   bo.longitude AS base_longitude,
                                   bo.raio_permitido_metros,
                                   bo.codigo_qr_base
                            FROM escala_motorista em
                                     LEFT JOIN bases_operacionais bo
                                               ON bo.id = em.base_operacional_id
                                                   AND bo.empresa_id = em.empresa_id
                            WHERE em.id = %s
                              AND em.empresa_id = %s
                              AND em.motorista_id = %s LIMIT 1
                            """, (int(escala_id), empresa_id, motorista_id))

                escala = cur.fetchone()

                if not escala:
                    flash("Escala não encontrada para seu usuário.", "danger")
                    return redirect(url_for('minha_semana_motorista'))

                if escala.get('status_escala') != 'Cancelado, comparecer na base':
                    flash(
                        "A confirmação de chegada só está disponível para cancelados convocados para comparecer na base.",
                        "warning")
                    return redirect(url_for('minha_semana_motorista'))

                if escala.get('status_presenca') == 'Chegada confirmada':
                    flash("Sua chegada já está confirmada.", "info")
                    return redirect(url_for(
                        'minha_semana_motorista',
                        semana=escala['data_escala'].strftime('%Y-%m-%d') if hasattr(escala['data_escala'],
                                                                                     'strftime') else str(
                            escala['data_escala'])[:10]
                    ))

                data_fila = escala.get('data_escala')
                base_operacional_id = escala.get('base_operacional_id')
                base_operacao = escala.get('base_nome') or escala.get('base_operacao') or 'Base não informada'

                config_checkin = obter_config_checkin_empresa(empresa_id, cur=cur)

                latitude_confirmacao = normalizar_coordenada(latitude_confirmacao_form)
                longitude_confirmacao = normalizar_coordenada(longitude_confirmacao_form)
                distancia_base_metros = None
                geolocalizacao_validada = 'N'
                raio_permitido = int(escala.get('raio_permitido_metros') or 150)

                if config_checkin.get('exigir_gps_raio') == 'S':
                    if latitude_confirmacao is None or longitude_confirmacao is None:
                        registrar_auditoria_checkin_base(
                            cur,
                            empresa_id,
                            motorista_id,
                            escala_id=int(escala_id),
                            base_operacional_id=base_operacional_id,
                            codigo_qr_informado=codigo_qr_informado,
                            resultado='Bloqueado por falta de geolocalização',
                            motivo_bloqueio='Localização não capturada pelo navegador.'
                        )
                        con.commit()
                        flash(
                            "Não foi possível capturar sua localização. Ative a localização do celular e tente novamente.",
                            "danger")
                        return redirect(url_for('minha_semana_motorista'))

                    distancia_base_metros = calcular_distancia_metros(
                        latitude_confirmacao,
                        longitude_confirmacao,
                        escala.get('base_latitude'),
                        escala.get('base_longitude')
                    )

                    if distancia_base_metros is None:
                        registrar_auditoria_checkin_base(
                            cur,
                            empresa_id,
                            motorista_id,
                            escala_id=int(escala_id),
                            base_operacional_id=base_operacional_id,
                            latitude=latitude_confirmacao,
                            longitude=longitude_confirmacao,
                            codigo_qr_informado=codigo_qr_informado,
                            resultado='Bloqueado por base sem coordenadas',
                            motivo_bloqueio='Base operacional sem latitude/longitude cadastrada.'
                        )
                        con.commit()
                        flash("A base selecionada não possui latitude/longitude cadastrada. Procure o supervisor.",
                              "danger")
                        return redirect(url_for('minha_semana_motorista'))

                    if distancia_base_metros <= raio_permitido:
                        geolocalizacao_validada = 'S'
                    else:
                        registrar_auditoria_checkin_base(
                            cur,
                            empresa_id,
                            motorista_id,
                            escala_id=int(escala_id),
                            base_operacional_id=base_operacional_id,
                            latitude=latitude_confirmacao,
                            longitude=longitude_confirmacao,
                            distancia_base_metros=distancia_base_metros,
                            codigo_qr_informado=codigo_qr_informado,
                            resultado='Bloqueado por distância',
                            motivo_bloqueio=f'Distância {distancia_base_metros}m maior que o raio permitido de {raio_permitido}m.'
                        )
                        con.commit()
                        flash(
                            f"Você está fora do raio permitido da base. Distância aproximada: {distancia_base_metros}m. Raio permitido: {raio_permitido}m.",
                            "danger")
                        return redirect(url_for('minha_semana_motorista'))
                else:
                    # GPS/raio desativado pela empresa. Se o navegador enviou coordenadas, salva apenas para auditoria.
                    if latitude_confirmacao is not None and longitude_confirmacao is not None:
                        distancia_base_metros = calcular_distancia_metros(
                            latitude_confirmacao,
                            longitude_confirmacao,
                            escala.get('base_latitude'),
                            escala.get('base_longitude')
                        )
                        if distancia_base_metros is not None and distancia_base_metros <= raio_permitido:
                            geolocalizacao_validada = 'S'

                qr_code_validado = 'N'
                qr_token_id = None
                if config_checkin.get('exigir_qrcode') == 'S':
                    token_qr = validar_token_qr_dinamico(
                        cur,
                        empresa_id,
                        base_operacional_id,
                        codigo_qr_informado
                    )

                    if token_qr:
                        qr_code_validado = 'S'
                        qr_token_id = token_qr.get('id')
                    else:
                        registrar_auditoria_checkin_base(
                            cur,
                            empresa_id,
                            motorista_id,
                            escala_id=int(escala_id),
                            base_operacional_id=base_operacional_id,
                            latitude=latitude_confirmacao,
                            longitude=longitude_confirmacao,
                            distancia_base_metros=distancia_base_metros,
                            codigo_qr_informado=codigo_qr_informado,
                            resultado='Bloqueado por QR inválido/expirado',
                            motivo_bloqueio='QR Code não encontrado, expirado ou não pertence à base da escala.'
                        )
                        con.commit()
                        flash("QR Code expirado ou inválido. Leia o QR atual exibido no terminal da base.", "danger")
                        return redirect(url_for('minha_semana_motorista'))

                selfie_path = None
                if selfie_base64:
                    selfie_path = salvar_selfie_base64(
                        selfie_base64,
                        empresa_id,
                        motorista_id,
                        int(escala_id),
                        cur=cur
                    )
                elif config_checkin.get('exigir_selfie') == 'S':
                    registrar_auditoria_checkin_base(
                        cur,
                        empresa_id,
                        motorista_id,
                        escala_id=int(escala_id),
                        base_operacional_id=base_operacional_id,
                        latitude=latitude_confirmacao,
                        longitude=longitude_confirmacao,
                        distancia_base_metros=distancia_base_metros,
                        codigo_qr_informado=codigo_qr_informado,
                        qr_token_id=qr_token_id,
                        resultado='Bloqueado por falta de selfie',
                        motivo_bloqueio='Selfie obrigatória não foi enviada.'
                    )
                    con.commit()
                    flash("Selfie obrigatória não enviada. Tire a selfie e tente novamente.", "danger")
                    return redirect(url_for('minha_semana_motorista'))

                cur.execute("""
                            SELECT id, posicao_fila, status_fila
                            FROM fila_cancelados_base
                            WHERE empresa_id = %s
                              AND escala_id = %s
                              AND motorista_id = %s LIMIT 1
                            """, (empresa_id, int(escala_id), motorista_id))

                fila_existente = cur.fetchone()

                if fila_existente:
                    cur.execute("""
                                UPDATE fila_cancelados_base
                                SET status_fila             = 'Aguardando rota',
                                    latitude_confirmacao    = %s,
                                    longitude_confirmacao   = %s,
                                    distancia_base_metros   = %s,
                                    geolocalizacao_validada = %s,
                                    qr_code_validado        = %s,
                                    qr_token_id             = %s,
                                    selfie_path             = %s,
                                    observacao              = 'Chegada reconfirmada pelo portal do motorista com geolocalização, QR Code e selfie.',
                                    data_atualizacao        = NOW()
                                WHERE id = %s
                                """, (
                                    latitude_confirmacao,
                                    longitude_confirmacao,
                                    distancia_base_metros,
                                    geolocalizacao_validada,
                                    qr_code_validado,
                                    qr_token_id,
                                    selfie_path,
                                    fila_existente.get('id')
                                ))
                    posicao_fila = fila_existente.get('posicao_fila')
                else:
                    posicao_fila = proxima_posicao_fila_cancelados(
                        cur,
                        empresa_id,
                        data_fila,
                        base_operacional_id
                    )

                    cur.execute("""
                                INSERT INTO fila_cancelados_base (empresa_id,
                                                                  escala_id,
                                                                  motorista_id,
                                                                  base_operacional_id,
                                                                  base_operacao,
                                                                  data_fila,
                                                                  hora_confirmacao,
                                                                  posicao_fila,
                                                                  status_fila,
                                                                  confirmado_por,
                                                                  usuario_confirmacao_id,
                                                                  latitude_confirmacao,
                                                                  longitude_confirmacao,
                                                                  distancia_base_metros,
                                                                  geolocalizacao_validada,
                                                                  qr_code_validado,
                                                                  qr_token_id,
                                                                  selfie_path,
                                                                  observacao,
                                                                  data_criacao,
                                                                  data_atualizacao)
                                VALUES (%s, %s, %s, %s, %s, %s, NOW(), %s, 'Aguardando rota', 'Motorista', %s, %s, %s,
                                        %s, %s, %s, %s, %s, %s, NOW(), NOW())
                                """, (
                                    empresa_id,
                                    int(escala_id),
                                    motorista_id,
                                    base_operacional_id,
                                    base_operacao,
                                    data_fila,
                                    posicao_fila,
                                    session.get('usuario_id'),
                                    latitude_confirmacao,
                                    longitude_confirmacao,
                                    distancia_base_metros,
                                    geolocalizacao_validada,
                                    qr_code_validado,
                                    qr_token_id,
                                    selfie_path,
                                    'Chegada confirmada pelo portal do motorista com geolocalização, QR Code dinâmico e selfie.'
                                ))

                cur.execute("""
                            UPDATE escala_motorista
                            SET status_presenca         = 'Chegada confirmada',
                                presenca_confirmada_em  = NOW(),
                                presenca_confirmada_por = 'Motorista',
                                usuario_confirmacao_id  = %s,
                                data_atualizacao        = NOW()
                            WHERE id = %s
                              AND empresa_id = %s
                              AND motorista_id = %s
                            """, (
                                session.get('usuario_id'),
                                int(escala_id),
                                empresa_id,
                                motorista_id
                            ))

                cur.execute("""
                            INSERT INTO ciencia_escala_motorista (empresa_id,
                                                                  motorista_id,
                                                                  escala_id,
                                                                  data_ciencia,
                                                                  origem_ciencia,
                                                                  usuario_id,
                                                                  created_at,
                                                                  updated_at)
                            VALUES (%s, %s, %s, NOW(), 'Motorista', %s, NOW(), NOW()) ON DUPLICATE KEY
                            UPDATE
                                data_ciencia = NOW(),
                                origem_ciencia = 'Motorista',
                                usuario_id =
                            VALUES (usuario_id), updated_at = NOW()
                            """, (
                                empresa_id,
                                motorista_id,
                                int(escala_id),
                                session.get('usuario_id')
                            ))

                registrar_auditoria_checkin_base(
                    cur,
                    empresa_id,
                    motorista_id,
                    escala_id=int(escala_id),
                    base_operacional_id=base_operacional_id,
                    latitude=latitude_confirmacao,
                    longitude=longitude_confirmacao,
                    distancia_base_metros=distancia_base_metros,
                    codigo_qr_informado=codigo_qr_informado,
                    qr_token_id=qr_token_id,
                    resultado='Aprovado',
                    motivo_bloqueio='Check-in aprovado com GPS, QR dinâmico e selfie.',
                    selfie_path=selfie_path
                )

                con.commit()
                flash(f"Chegada confirmada com sucesso. Você entrou na fila na posição {posicao_fila}.", "success")

                return redirect(url_for(
                    'minha_semana_motorista',
                    semana=data_fila.strftime('%Y-%m-%d') if hasattr(data_fila, 'strftime') else str(data_fila)[:10]
                ))

            except Exception as e:
                con.rollback()
                print(f"Erro ao confirmar chegada na base: {e}")
                flash(f"Erro técnico ao confirmar chegada: {e}", "danger")
                return redirect(url_for('minha_semana_motorista'))

            finally:
                cur.close()
                con.close()

    data_ref_str = request.args.get('semana', '').strip()

    if data_ref_str:
        try:
            data_ref = datetime.strptime(data_ref_str[:10], '%Y-%m-%d').date()
        except Exception:
            data_ref = date.today()
    else:
        data_ref = date.today()

    data_inicio, data_fim = semana_domingo_sabado(data_ref)

    dias = carregar_minha_semana_motorista(
        empresa_id,
        motorista_id,
        data_inicio,
        data_fim
    )

    hoje = date.today()
    amanha = hoje + timedelta(days=1)

    dia_hoje = next((d for d in dias if d['data'] == hoje), None)
    dia_amanha = next((d for d in dias if d['data'] == amanha), None)

    resumo = {
        'disponiveis': sum(1 for d in dias if d['status_disponibilidade'] == 'Disponível'),
        'ausentes': sum(1 for d in dias if d['status_disponibilidade'] == 'Ausente'),
        'confirmados': sum(1 for d in dias if d['status_escala'] == 'Confirmado com rota'),
        'comparecer_base': sum(1 for d in dias if d['status_escala'] == 'Cancelado, comparecer na base'),
        'reservas': sum(1 for d in dias if d['status_escala'] == 'Reserva / Avulso'),
        'faltas': sum(1 for d in dias if d['status_escala'] == 'Falta'),
        'ciencias_pendentes': sum(1 for d in dias if d['precisa_ciencia'] and not d['data_ciencia'])
    }

    semana_anterior = data_inicio - timedelta(days=7)
    proxima_semana = data_inicio + timedelta(days=7)

    return render_template(
        'minha_semana_motorista.html',
        usuario_logado=usuario_logado,
        dias=dias,
        dia_hoje=dia_hoje,
        dia_amanha=dia_amanha,
        resumo=resumo,
        data_inicio=data_inicio,
        data_fim=data_fim,
        semana_anterior=semana_anterior,
        proxima_semana=proxima_semana
    )


# ==========================================================
# FASE 4.1.5 - TERMINAL DA BASE COM QR CODE DINÂMICO
# ==========================================================
def terminal_base_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            flash('Por favor, faça login para acessar esta página.', 'danger')
            return redirect(url_for('login'))

        if session.get('perfil_de_acesso') != 'Terminal Base':
            flash("Acesso restrito ao Terminal da Base.", "danger")
            return acesso_negado_por_perfil()

        if not session.get('base_operacional_id'):
            flash("Usuário Terminal Base sem base vinculada. Procure o administrador.", "danger")
            return redirect(url_for('logout'))

        return f(*args, **kwargs)

    return decorated_function


def gerar_token_qr_dinamico(base):
    import uuid

    agora = datetime.now()
    empresa_id = base.get('empresa_id')
    base_id = base.get('id')
    validade_minutos = int(base.get('qr_validade_minutos') or 5)

    if validade_minutos < 1:
        validade_minutos = 5

    validade_segundos = validade_minutos * 60

    con = obter_conexao()

    if con is None:
        return None

    cur = con.cursor(dictionary=True)

    try:
        # Expira tokens antigos da base.
        cur.execute("""
                    UPDATE base_qr_tokens
                    SET status_token = 'Expirado'
                    WHERE empresa_id = %s
                      AND base_operacional_id = %s
                      AND status_token = 'Ativo'
                      AND data_expiracao <= NOW()
                    """, (empresa_id, base_id))

        # Reaproveita token ativo se ainda estiver válido.
        cur.execute("""
                    SELECT id,
                           codigo_token,
                           data_geracao,
                           data_expiracao,
                           TIMESTAMPDIFF(SECOND, NOW(), data_expiracao) AS segundos_restantes
                    FROM base_qr_tokens
                    WHERE empresa_id = %s
                      AND base_operacional_id = %s
                      AND status_token = 'Ativo'
                      AND data_expiracao > NOW()
                    ORDER BY data_expiracao DESC LIMIT 1
                    """, (empresa_id, base_id))

        token = cur.fetchone()

        if token and int(token.get('segundos_restantes') or 0) > 10:
            con.commit()
            return token

        # Se estiver quase vencendo, expira e cria um novo.
        cur.execute("""
                    UPDATE base_qr_tokens
                    SET status_token = 'Expirado'
                    WHERE empresa_id = %s
                      AND base_operacional_id = %s
                      AND status_token = 'Ativo'
                    """, (empresa_id, base_id))

        base_limpa = re.sub(r'[^A-Za-z0-9]+', '', str(base.get('nome_base') or 'BASE')).upper()[:8]
        codigo_token = f"{base_limpa}-{agora.strftime('%d%m%H%M')}-{uuid.uuid4().hex[:6].upper()}"

        cur.execute("""
                    INSERT INTO base_qr_tokens (empresa_id,
                                                base_operacional_id,
                                                codigo_token,
                                                data_geracao,
                                                data_expiracao,
                                                status_token,
                                                created_at)
                    VALUES (%s, %s, %s, NOW(), DATE_ADD(NOW(), INTERVAL %s SECOND), 'Ativo', NOW())
                    """, (
                        empresa_id,
                        base_id,
                        codigo_token,
                        validade_segundos
                    ))

        token_id = cur.lastrowid

        con.commit()

        return {
            'id': token_id,
            'codigo_token': codigo_token,
            'data_geracao': agora,
            'data_expiracao': agora + timedelta(seconds=validade_segundos),
            'segundos_restantes': validade_segundos
        }

    except Exception as e:
        con.rollback()
        print(f"Erro ao gerar token QR dinâmico: {e}")
        return None

    finally:
        cur.close()
        con.close()


def validar_token_qr_dinamico(cur, empresa_id, base_operacional_id, codigo_token):
    codigo_token = (codigo_token or '').strip().upper()

    if not codigo_token:
        return None

    cur.execute("""
                SELECT id,
                       codigo_token,
                       data_geracao,
                       data_expiracao,
                       status_token
                FROM base_qr_tokens
                WHERE empresa_id = %s
                  AND base_operacional_id = %s
                  AND UPPER(codigo_token) = %s
                  AND status_token = 'Ativo'
                  AND data_expiracao > NOW() LIMIT 1
                """, (
                    empresa_id,
                    base_operacional_id,
                    codigo_token
                ))

    return cur.fetchone()


@app.route('/terminal-base/qrcode', methods=['GET'])
@login_required
@terminal_base_required
def terminal_base_qrcode():
    empresa_id = session.get('empresa_id')
    base_id = session.get('base_operacional_id')

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('logout'))

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT id,
                           empresa_id,
                           nome_base,
                           descricao,
                           endereco,
                           latitude,
                           longitude,
                           raio_permitido_metros,
                           codigo_qr_base,
                           qr_validade_minutos,
                           status_base
                    FROM bases_operacionais
                    WHERE id = %s
                      AND empresa_id = %s
                      AND status_base = 'Ativa' LIMIT 1
                    """, (base_id, empresa_id))

        base = cur.fetchone()

        if not base:
            flash("Base operacional não encontrada ou inativa.", "danger")
            return redirect(url_for('logout'))

    except Exception as e:
        print(f"Erro ao carregar terminal da base: {e}")
        flash(f"Erro técnico ao carregar terminal da base: {e}", "danger")
        return redirect(url_for('logout'))

    finally:
        cur.close()
        con.close()

    token = gerar_token_qr_dinamico(base)

    if not token:
        flash("Não foi possível gerar QR Code da base.", "danger")
        return redirect(url_for('logout'))

    segundos_restantes = int(token.get('segundos_restantes') or (int(base.get('qr_validade_minutos') or 5) * 60))

    return render_template(
        'terminal_base_qrcode.html',
        base=base,
        token=token,
        segundos_restantes=segundos_restantes
    )


@app.route('/terminal-base/qrcode/atualizar', methods=['GET'])
@login_required
@terminal_base_required
def terminal_base_qrcode_atualizar():
    empresa_id = session.get('empresa_id')
    base_id = session.get('base_operacional_id')

    con = obter_conexao()

    if con is None:
        return jsonify({
            'ok': False,
            'mensagem': 'Erro de conexão com o banco de dados.'
        }), 500

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT id,
                           empresa_id,
                           nome_base,
                           qr_validade_minutos,
                           status_base
                    FROM bases_operacionais
                    WHERE id = %s
                      AND empresa_id = %s
                      AND status_base = 'Ativa' LIMIT 1
                    """, (base_id, empresa_id))

        base = cur.fetchone()

        if not base:
            return jsonify({
                'ok': False,
                'mensagem': 'Base operacional não encontrada ou inativa.'
            }), 404

    except Exception as e:
        print(f"Erro ao carregar base para atualizar QR: {e}")
        return jsonify({
            'ok': False,
            'mensagem': f'Erro técnico: {e}'
        }), 500

    finally:
        cur.close()
        con.close()

    token = gerar_token_qr_dinamico(base)

    if not token:
        return jsonify({
            'ok': False,
            'mensagem': 'Não foi possível gerar QR Code.'
        }), 500

    segundos_restantes = int(token.get('segundos_restantes') or (int(base.get('qr_validade_minutos') or 5) * 60))

    return jsonify({
        'ok': True,
        'codigo_token': token.get('codigo_token'),
        'segundos_restantes': segundos_restantes,
        'validade_minutos': int(base.get('qr_validade_minutos') or 5),
        'data_expiracao': str(token.get('data_expiracao') or '')
    })


# ==========================================================
# FASE 4.1.6 - HISTÓRICO E AUDITORIA DE CHECK-IN DA BASE
# ==========================================================
def registrar_auditoria_checkin_base(
        cur,
        empresa_id,
        motorista_id,
        escala_id=None,
        base_operacional_id=None,
        latitude=None,
        longitude=None,
        distancia_base_metros=None,
        codigo_qr_informado=None,
        qr_token_id=None,
        resultado='Pendente',
        motivo_bloqueio=None,
        selfie_path=None
):
    """
    Registra tentativas aprovadas e bloqueadas de check-in na base.
    Usa a mesma conexão/transação da rota principal.
    """
    try:
        ip_origem = (request.remote_addr or '')[:80]
        user_agent = (request.headers.get('User-Agent') or '')[:500]

        cur.execute("""
                    INSERT INTO auditoria_checkin_base (empresa_id,
                                                        motorista_id,
                                                        escala_id,
                                                        base_operacional_id,
                                                        data_tentativa,
                                                        latitude,
                                                        longitude,
                                                        distancia_base_metros,
                                                        codigo_qr_informado,
                                                        qr_token_id,
                                                        resultado,
                                                        motivo_bloqueio,
                                                        selfie_path,
                                                        ip_origem,
                                                        user_agent,
                                                        created_at)
                    VALUES (%s, %s, %s, %s, NOW(), %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
                    """, (
                        empresa_id,
                        motorista_id,
                        escala_id,
                        base_operacional_id,
                        latitude,
                        longitude,
                        distancia_base_metros,
                        (codigo_qr_informado or '')[:120],
                        qr_token_id,
                        resultado,
                        (motivo_bloqueio or '')[:255],
                        selfie_path,
                        ip_origem,
                        user_agent
                    ))

    except Exception as e:
        print(f"Erro ao registrar auditoria de check-in: {e}")


@app.route('/operacao/auditoria-checkin-base', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor')
def auditoria_checkin_base():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    hoje = date.today()
    data_inicio_str = request.args.get('data_inicio', hoje.strftime('%Y-%m-%d')).strip()
    data_fim_str = request.args.get('data_fim', hoje.strftime('%Y-%m-%d')).strip()
    resultado_filtro = request.args.get('resultado', '').strip()
    base_id_filtro = request.args.get('base_operacional_id', '').strip()
    busca_motorista = request.args.get('motorista', '').strip()

    empresa_id_filtro = request.args.get('empresa_id', '').strip()

    if is_super_admin and empresa_id_filtro and empresa_id_filtro.isdigit():
        empresa_id = int(empresa_id_filtro)
    else:
        empresa_id = empresa_logada_id

    try:
        data_inicio = datetime.strptime(data_inicio_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_inicio = hoje
        data_inicio_str = hoje.strftime('%Y-%m-%d')

    try:
        data_fim = datetime.strptime(data_fim_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_fim = hoje
        data_fim_str = hoje.strftime('%Y-%m-%d')

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão ao carregar auditoria.", "danger")
        return redirect(url_for('dashboard'))

    cur = con.cursor(dictionary=True)

    try:
        bases_operacionais = carregar_bases_operacionais(empresa_id, apenas_ativas=False)

        sql = """
              SELECT aud.id,
                     aud.data_tentativa,
                     aud.resultado,
                     aud.motivo_bloqueio,
                     aud.latitude,
                     aud.longitude,
                     aud.distancia_base_metros,
                     aud.codigo_qr_informado,
                     aud.qr_token_id,
                     aud.selfie_path,
                     aud.ip_origem,
                     mot.nome_completo  AS motorista_nome,
                     mot.cpf_cnpj       AS motorista_cpf_cnpj,
                     bo.nome_base       AS base_nome,
                     bo.raio_permitido_metros,
                     em.data_escala,
                     em.status_escala,
                     em.status_presenca,
                     bqt.data_geracao   AS qr_data_geracao,
                     bqt.data_expiracao AS qr_data_expiracao,
                     bqt.status_token   AS qr_status_token
              FROM auditoria_checkin_base aud
                       LEFT JOIN pessoas mot
                                 ON mot.id = aud.motorista_id
                                     AND mot.empresa_id = aud.empresa_id
                       LEFT JOIN bases_operacionais bo
                                 ON bo.id = aud.base_operacional_id
                                     AND bo.empresa_id = aud.empresa_id
                       LEFT JOIN escala_motorista em
                                 ON em.id = aud.escala_id
                                     AND em.empresa_id = aud.empresa_id
                       LEFT JOIN base_qr_tokens bqt
                                 ON bqt.id = aud.qr_token_id
                                     AND bqt.empresa_id = aud.empresa_id
              WHERE aud.empresa_id = %s
                AND DATE (aud.data_tentativa) BETWEEN %s
                AND %s \
              """

        params = [empresa_id, data_inicio, data_fim]

        if resultado_filtro:
            sql += " AND aud.resultado = %s"
            params.append(resultado_filtro)

        if base_id_filtro and base_id_filtro.isdigit():
            sql += " AND aud.base_operacional_id = %s"
            params.append(int(base_id_filtro))

        if busca_motorista:
            sql += " AND mot.nome_completo LIKE %s"
            params.append(f"%{busca_motorista}%")

        sql += " ORDER BY aud.data_tentativa DESC, aud.id DESC LIMIT 500"

        cur.execute(sql, params)
        auditorias = cur.fetchall()

        resumo = {
            'total': len(auditorias),
            'aprovados': sum(1 for a in auditorias if a.get('resultado') == 'Aprovado'),
            'bloqueados': sum(1 for a in auditorias if str(a.get('resultado') or '').startswith('Bloqueado')),
            'erros': sum(1 for a in auditorias if a.get('resultado') == 'Erro técnico')
        }

        return render_template(
            'auditoria_checkin_base.html',
            usuario_logado=usuario_logado,
            auditorias=auditorias,
            bases_operacionais=bases_operacionais,
            resumo=resumo,
            data_inicio=data_inicio_str,
            data_fim=data_fim_str,
            resultado_filtro=resultado_filtro,
            base_id_filtro=base_id_filtro,
            busca_motorista=busca_motorista,
            is_super_admin=is_super_admin,
            empresa_id_filtro=empresa_id if is_super_admin else ''
        )

    except Exception as e:
        print(f"Erro ao carregar auditoria de check-in: {e}")
        flash(f"Erro técnico ao carregar auditoria: {e}", "danger")
        return redirect(url_for('escala_motoristas'))

    finally:
        cur.close()
        con.close()


@app.route('/operacao/mapa-checkins', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor')
def mapa_checkins():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    hoje = date.today()
    data_inicio_str = request.args.get('data_inicio', hoje.strftime('%Y-%m-%d')).strip()
    data_fim_str = request.args.get('data_fim', hoje.strftime('%Y-%m-%d')).strip()
    resultado_filtro = request.args.get('resultado', '').strip()
    base_id_filtro = request.args.get('base_operacional_id', '').strip()
    busca_motorista = request.args.get('motorista', '').strip()
    empresa_id_filtro = request.args.get('empresa_id', '').strip()

    if is_super_admin and empresa_id_filtro and empresa_id_filtro.isdigit():
        empresa_id = int(empresa_id_filtro)
    else:
        empresa_id = empresa_logada_id

    try:
        data_inicio = datetime.strptime(data_inicio_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_inicio = hoje
        data_inicio_str = hoje.strftime('%Y-%m-%d')

    try:
        data_fim = datetime.strptime(data_fim_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_fim = hoje
        data_fim_str = hoje.strftime('%Y-%m-%d')

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão ao carregar mapa de check-ins.", "danger")
        return redirect(url_for('dashboard'))

    cur = con.cursor(dictionary=True)

    def _float_or_none(valor):
        try:
            if valor is None or valor == '':
                return None
            return float(valor)
        except Exception:
            return None

    try:
        bases_operacionais = carregar_bases_operacionais(empresa_id, apenas_ativas=False)
        empresas = carregar_empresas_ativas() if is_super_admin else []

        sql = """
              SELECT aud.id,
                     aud.data_tentativa,
                     aud.resultado,
                     aud.motivo_bloqueio,
                     aud.latitude AS tentativa_latitude,
                     aud.longitude AS tentativa_longitude,
                     aud.distancia_base_metros,
                     aud.codigo_qr_informado,
                     aud.qr_token_id,
                     aud.selfie_path,
                     aud.ip_origem,
                     aud.user_agent,
                     mot.id AS motorista_id,
                     mot.nome_completo AS motorista_nome,
                     mot.cpf_cnpj AS motorista_cpf_cnpj,
                     bo.id AS base_id,
                     bo.nome_base AS base_nome,
                     bo.endereco AS base_endereco,
                     bo.latitude AS base_latitude,
                     bo.longitude AS base_longitude,
                     bo.raio_permitido_metros,
                     em.data_escala,
                     em.status_escala,
                     em.status_presenca
              FROM auditoria_checkin_base aud
                       LEFT JOIN pessoas mot
                                 ON mot.id = aud.motorista_id
                                     AND mot.empresa_id = aud.empresa_id
                       LEFT JOIN bases_operacionais bo
                                 ON bo.id = aud.base_operacional_id
                                     AND bo.empresa_id = aud.empresa_id
                       LEFT JOIN escala_motorista em
                                 ON em.id = aud.escala_id
                                     AND em.empresa_id = aud.empresa_id
              WHERE aud.empresa_id = %s
                AND DATE(aud.data_tentativa) BETWEEN %s AND %s
              """
        params = [empresa_id, data_inicio, data_fim]

        if resultado_filtro:
            if resultado_filtro == 'Bloqueados':
                sql += " AND aud.resultado LIKE 'Bloqueado%'"
            else:
                sql += " AND aud.resultado = %s"
                params.append(resultado_filtro)

        if base_id_filtro and base_id_filtro.isdigit():
            sql += " AND aud.base_operacional_id = %s"
            params.append(int(base_id_filtro))

        if busca_motorista:
            sql += " AND mot.nome_completo LIKE %s"
            params.append(f"%{busca_motorista}%")

        sql += " ORDER BY aud.data_tentativa DESC, aud.id DESC LIMIT 800"
        cur.execute(sql, params)
        tentativas = cur.fetchall()

        total = len(tentativas)
        aprovados = sum(1 for t in tentativas if t.get('resultado') == 'Aprovado')
        bloqueados = sum(1 for t in tentativas if str(t.get('resultado') or '').startswith('Bloqueado'))
        fora_raio = sum(1 for t in tentativas if 'distância' in str(t.get('motivo_bloqueio') or '').lower() or 'distancia' in str(t.get('motivo_bloqueio') or '').lower())
        qr_invalidos = sum(1 for t in tentativas if 'qr' in str(t.get('motivo_bloqueio') or '').lower())
        sem_geo_selfie = sum(1 for t in tentativas if any(p in str(t.get('motivo_bloqueio') or '').lower() for p in ['geolocalização', 'geolocalizacao', 'selfie']))

        resumo = {
            'total': total,
            'aprovados': aprovados,
            'bloqueados': bloqueados,
            'fora_raio': fora_raio,
            'qr_invalidos': qr_invalidos,
            'sem_geo_selfie': sem_geo_selfie
        }

        pontos_mapa = []
        bases_mapa_dict = {}

        for t in tentativas:
            tentativa_lat = _float_or_none(t.get('tentativa_latitude'))
            tentativa_lng = _float_or_none(t.get('tentativa_longitude'))
            base_lat = _float_or_none(t.get('base_latitude'))
            base_lng = _float_or_none(t.get('base_longitude'))
            raio = _float_or_none(t.get('raio_permitido_metros')) or 0
            selfie_path = t.get('selfie_path') or ''
            selfie_url = arquivo_url(selfie_path) if selfie_path else ''
            data_tentativa = t.get('data_tentativa')

            if base_lat is not None and base_lng is not None and t.get('base_id'):
                bases_mapa_dict[str(t.get('base_id'))] = {
                    'id': t.get('base_id'),
                    'nome': t.get('base_nome') or 'Base sem nome',
                    'endereco': t.get('base_endereco') or '',
                    'lat': base_lat,
                    'lng': base_lng,
                    'raio': raio
                }

            if tentativa_lat is not None and tentativa_lng is not None:
                pontos_mapa.append({
                    'id': t.get('id'),
                    'data': data_tentativa.strftime('%d/%m/%Y %H:%M') if hasattr(data_tentativa, 'strftime') else str(data_tentativa or ''),
                    'motorista': t.get('motorista_nome') or 'Motorista não identificado',
                    'base': t.get('base_nome') or 'Base não identificada',
                    'resultado': t.get('resultado') or '',
                    'motivo': t.get('motivo_bloqueio') or '',
                    'distancia': float(t.get('distancia_base_metros') or 0),
                    'lat': tentativa_lat,
                    'lng': tentativa_lng,
                    'base_lat': base_lat,
                    'base_lng': base_lng,
                    'selfie_url': selfie_url,
                    'ip': t.get('ip_origem') or ''
                })

        bases_mapa = list(bases_mapa_dict.values())

        return render_template(
            'mapa_checkins.html',
            usuario_logado=usuario_logado,
            bases_operacionais=bases_operacionais,
            empresas=empresas,
            resumo=resumo,
            tentativas=tentativas,
            pontos_mapa=pontos_mapa,
            bases_mapa=bases_mapa,
            data_inicio=data_inicio_str,
            data_fim=data_fim_str,
            resultado_filtro=resultado_filtro,
            base_id_filtro=base_id_filtro,
            busca_motorista=busca_motorista,
            is_super_admin=is_super_admin,
            empresa_id_filtro=empresa_id if is_super_admin else ''
        )

    except Exception as e:
        print(f"Erro ao carregar mapa de check-ins: {e}")
        flash(f"Erro técnico ao carregar mapa de check-ins: {e}", "danger")
        return redirect(url_for('auditoria_checkin_base'))

    finally:
        cur.close()
        con.close()


@app.route('/operacao/auditoria-checkin-base/exportar-csv', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor')
def exportar_auditoria_checkin_base_csv():
    empresa_id = session.get('empresa_id')

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    hoje = date.today()
    data_inicio_str = request.args.get('data_inicio', hoje.strftime('%Y-%m-%d')).strip()
    data_fim_str = request.args.get('data_fim', hoje.strftime('%Y-%m-%d')).strip()

    try:
        data_inicio = datetime.strptime(data_inicio_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_inicio = hoje

    try:
        data_fim = datetime.strptime(data_fim_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_fim = hoje

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão ao exportar auditoria.", "danger")
        return redirect(url_for('auditoria_checkin_base'))

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT aud.data_tentativa,
                           mot.nome_completo AS motorista_nome,
                           bo.nome_base      AS base_nome,
                           aud.resultado,
                           aud.motivo_bloqueio,
                           aud.distancia_base_metros,
                           aud.codigo_qr_informado,
                           aud.ip_origem
                    FROM auditoria_checkin_base aud
                             LEFT JOIN pessoas mot
                                       ON mot.id = aud.motorista_id
                                           AND mot.empresa_id = aud.empresa_id
                             LEFT JOIN bases_operacionais bo
                                       ON bo.id = aud.base_operacional_id
                                           AND bo.empresa_id = aud.empresa_id
                    WHERE aud.empresa_id = %s
                      AND DATE (aud.data_tentativa) BETWEEN %s
                      AND %s
                    ORDER BY aud.data_tentativa DESC
                    """, (empresa_id, data_inicio, data_fim))

        rows = cur.fetchall()

        linhas = [
            "Data tentativa;Motorista;Base;Resultado;Motivo;Distancia metros;Codigo QR;IP"
        ]

        for row in rows:
            linhas.append(
                f"{row.get('data_tentativa')};"
                f"{row.get('motorista_nome') or ''};"
                f"{row.get('base_nome') or ''};"
                f"{row.get('resultado') or ''};"
                f"{row.get('motivo_bloqueio') or ''};"
                f"{row.get('distancia_base_metros') or ''};"
                f"{row.get('codigo_qr_informado') or ''};"
                f"{row.get('ip_origem') or ''}"
            )

        conteudo = "\n".join(linhas)

        return Response(
            conteudo,
            mimetype="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": "attachment; filename=auditoria_checkin_base.csv"
            }
        )

    except Exception as e:
        print(f"Erro ao exportar auditoria: {e}")
        flash(f"Erro técnico ao exportar auditoria: {e}", "danger")
        return redirect(url_for('auditoria_checkin_base'))

    finally:
        cur.close()
        con.close()


# ==========================================================
# FASE 4.1.7 - RELATÓRIO OPERACIONAL DA ESCALA / BASE
# ==========================================================
@app.route('/operacao/relatorio-escala-base', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor')
def relatorio_operacional_escala_base():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    hoje = date.today()
    data_inicio_str = request.args.get('data_inicio', hoje.strftime('%Y-%m-%d')).strip()
    data_fim_str = request.args.get('data_fim', hoje.strftime('%Y-%m-%d')).strip()
    base_id_filtro = request.args.get('base_operacional_id', '').strip()
    empresa_id_filtro = request.args.get('empresa_id', '').strip()

    if is_super_admin and empresa_id_filtro and empresa_id_filtro.isdigit():
        empresa_id = int(empresa_id_filtro)
    else:
        empresa_id = empresa_logada_id

    try:
        data_inicio = datetime.strptime(data_inicio_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_inicio = hoje
        data_inicio_str = hoje.strftime('%Y-%m-%d')

    try:
        data_fim = datetime.strptime(data_fim_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_fim = hoje
        data_fim_str = hoje.strftime('%Y-%m-%d')

    if data_fim < data_inicio:
        data_inicio, data_fim = data_fim, data_inicio
        data_inicio_str = data_inicio.strftime('%Y-%m-%d')
        data_fim_str = data_fim.strftime('%Y-%m-%d')

    dias_periodo = max((data_fim - data_inicio).days + 1, 1)
    data_inicio_anterior = data_inicio - timedelta(days=dias_periodo)
    data_fim_anterior = data_inicio - timedelta(days=1)

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão ao carregar relatório operacional.", "danger")
        return redirect(url_for('dashboard'))

    cur = con.cursor(dictionary=True)

    def inteiro(valor):
        try:
            return int(valor or 0)
        except Exception:
            return 0

    def percentual(parte, total):
        parte = inteiro(parte)
        total = inteiro(total)
        return round((parte / total) * 100, 1) if total else 0

    def variacao_percentual(atual, anterior):
        atual = float(atual or 0)
        anterior = float(anterior or 0)
        if anterior == 0 and atual > 0:
            return 100
        if anterior == 0:
            return 0
        return round(((atual - anterior) / anterior) * 100, 1)

    try:
        bases_operacionais = carregar_bases_operacionais(empresa_id, apenas_ativas=False)

        filtro_base_sql = ""
        params_base = []
        filtro_base_fila_sql = ""
        params_base_fila = []
        filtro_base_aud_sql = ""
        params_base_aud = []

        if base_id_filtro and base_id_filtro.isdigit():
            base_id_int = int(base_id_filtro)
            filtro_base_sql = " AND em.base_operacional_id = %s "
            params_base.append(base_id_int)
            filtro_base_fila_sql = " AND f.base_operacional_id = %s "
            params_base_fila.append(base_id_int)
            filtro_base_aud_sql = " AND aud.base_operacional_id = %s "
            params_base_aud.append(base_id_int)

        def carregar_resumo_periodo(inicio, fim):
            cur.execute(f"""
                SELECT
                    COUNT(*) AS total_escalados,
                    SUM(CASE WHEN em.status_escala = 'Confirmado com rota' THEN 1 ELSE 0 END) AS confirmados_rota,
                    SUM(CASE WHEN em.status_escala = 'Cancelado, comparecer na base' THEN 1 ELSE 0 END) AS cancelados_comparecer_base,
                    SUM(CASE WHEN em.status_escala = 'Cancelado definitivo' THEN 1 ELSE 0 END) AS cancelados_definitivos,
                    SUM(CASE WHEN em.status_escala = 'Reserva / Avulso' THEN 1 ELSE 0 END) AS reservas_avulsos,
                    SUM(CASE WHEN em.status_escala = 'Falta' THEN 1 ELSE 0 END) AS faltas,
                    SUM(CASE WHEN em.status_presenca = 'Chegada confirmada' THEN 1 ELSE 0 END) AS chegadas_confirmadas,
                    SUM(CASE WHEN em.status_presenca = 'Não compareceu' THEN 1 ELSE 0 END) AS nao_compareceu,
                    SUM(CASE WHEN em.falta_automatica = 'S' THEN 1 ELSE 0 END) AS faltas_automaticas
                FROM escala_motorista em
                WHERE em.empresa_id = %s
                  AND em.data_escala BETWEEN %s AND %s
                  {filtro_base_sql}
            """, [empresa_id, inicio, fim] + params_base)
            kpis_escala = cur.fetchone() or {}

            cur.execute(f"""
                SELECT
                    COUNT(*) AS total_fila,
                    SUM(CASE WHEN f.status_fila = 'Aguardando rota' THEN 1 ELSE 0 END) AS aguardando_rota,
                    SUM(CASE WHEN f.status_fila = 'Atribuído para rota extra' THEN 1 ELSE 0 END) AS atribuidos_rota_extra,
                    SUM(CASE WHEN f.status_fila = 'Dispensado' THEN 1 ELSE 0 END) AS dispensados,
                    SUM(CASE WHEN f.status_fila = 'Falta' THEN 1 ELSE 0 END) AS faltas_fila
                FROM fila_cancelados_base f
                WHERE f.empresa_id = %s
                  AND f.data_fila BETWEEN %s AND %s
                  {filtro_base_fila_sql}
            """, [empresa_id, inicio, fim] + params_base_fila)
            kpis_fila = cur.fetchone() or {}

            cur.execute(f"""
                SELECT
                    COUNT(*) AS total_tentativas,
                    SUM(CASE WHEN aud.resultado = 'Aprovado' THEN 1 ELSE 0 END) AS checkins_aprovados,
                    SUM(CASE WHEN aud.resultado LIKE 'Bloqueado%%' THEN 1 ELSE 0 END) AS checkins_bloqueados,
                    SUM(CASE WHEN aud.resultado = 'Bloqueado por distância' THEN 1 ELSE 0 END) AS bloqueados_distancia,
                    SUM(CASE WHEN aud.resultado = 'Bloqueado por QR inválido/expirado' THEN 1 ELSE 0 END) AS bloqueados_qr,
                    SUM(CASE WHEN aud.resultado = 'Bloqueado por falta de selfie' THEN 1 ELSE 0 END) AS bloqueados_selfie,
                    SUM(CASE WHEN aud.resultado = 'Bloqueado por falta de geolocalização' THEN 1 ELSE 0 END) AS bloqueados_geo,
                    SUM(CASE WHEN aud.resultado = 'Bloqueado por base sem coordenadas' THEN 1 ELSE 0 END) AS bloqueados_base_sem_coordenada
                FROM auditoria_checkin_base aud
                WHERE aud.empresa_id = %s
                  AND DATE(aud.data_tentativa) BETWEEN %s AND %s
                  {filtro_base_aud_sql}
            """, [empresa_id, inicio, fim] + params_base_aud)
            kpis_auditoria = cur.fetchone() or {}

            resumo_tmp = {}
            resumo_tmp.update({k: inteiro(v) for k, v in kpis_escala.items()})
            resumo_tmp.update({k: inteiro(v) for k, v in kpis_fila.items()})
            resumo_tmp.update({k: inteiro(v) for k, v in kpis_auditoria.items()})
            resumo_tmp['taxa_comparecimento_base'] = percentual(
                resumo_tmp.get('chegadas_confirmadas'),
                resumo_tmp.get('cancelados_comparecer_base')
            )
            resumo_tmp['taxa_aproveitamento_fila'] = percentual(
                resumo_tmp.get('atribuidos_rota_extra'),
                resumo_tmp.get('total_fila')
            )
            resumo_tmp['taxa_bloqueio_checkin'] = percentual(
                resumo_tmp.get('checkins_bloqueados'),
                resumo_tmp.get('total_tentativas')
            )
            resumo_tmp['taxa_falta'] = percentual(
                resumo_tmp.get('faltas') + resumo_tmp.get('nao_compareceu'),
                resumo_tmp.get('total_escalados')
            )
            resumo_tmp['taxa_checkin_aprovado'] = percentual(
                resumo_tmp.get('checkins_aprovados'),
                resumo_tmp.get('total_tentativas')
            )
            return resumo_tmp

        resumo = carregar_resumo_periodo(data_inicio, data_fim)
        resumo_anterior = carregar_resumo_periodo(data_inicio_anterior, data_fim_anterior)

        for chave in [
            'total_escalados', 'confirmados_rota', 'cancelados_comparecer_base', 'chegadas_confirmadas',
            'atribuidos_rota_extra', 'dispensados', 'faltas', 'checkins_aprovados', 'checkins_bloqueados',
            'taxa_comparecimento_base', 'taxa_aproveitamento_fila', 'taxa_bloqueio_checkin'
        ]:
            resumo[f'var_{chave}'] = variacao_percentual(resumo.get(chave), resumo_anterior.get(chave))

        alertas_operacionais = []
        if resumo.get('taxa_comparecimento_base', 0) < 90 and resumo.get('cancelados_comparecer_base', 0) > 0:
            alertas_operacionais.append({
                'tipo': 'warning',
                'icone': 'fa-user-clock',
                'titulo': 'Comparecimento abaixo da meta',
                'descricao': f"Taxa de comparecimento em {resumo.get('taxa_comparecimento_base')}%. Meta sugerida: 90%."
            })
        if resumo.get('taxa_bloqueio_checkin', 0) >= 15 and resumo.get('total_tentativas', 0) > 0:
            alertas_operacionais.append({
                'tipo': 'danger',
                'icone': 'fa-shield-halved',
                'titulo': 'Bloqueios elevados no check-in',
                'descricao': f"{resumo.get('checkins_bloqueados')} bloqueios em {resumo.get('total_tentativas')} tentativas."
            })
        if resumo.get('aguardando_rota', 0) > 0:
            alertas_operacionais.append({
                'tipo': 'info',
                'icone': 'fa-people-arrows',
                'titulo': 'Motoristas aguardando decisão',
                'descricao': f"{resumo.get('aguardando_rota')} motorista(s) ainda aguardam rota ou dispensa na fila."
            })
        if resumo.get('faltas', 0) > 0 or resumo.get('nao_compareceu', 0) > 0:
            alertas_operacionais.append({
                'tipo': 'danger',
                'icone': 'fa-user-xmark',
                'titulo': 'Faltas identificadas',
                'descricao': f"{resumo.get('faltas', 0)} falta(s) na escala e {resumo.get('nao_compareceu', 0)} não comparecimento(s)."
            })
        if not alertas_operacionais:
            alertas_operacionais.append({
                'tipo': 'success',
                'icone': 'fa-circle-check',
                'titulo': 'Operação sem alerta crítico',
                'descricao': 'Nenhum desvio relevante encontrado para os filtros selecionados.'
            })

        cur.execute(f"""
            SELECT COALESCE(bo.nome_base, em.base_operacao, 'Sem base') AS base_nome,
                   COUNT(*) AS total_escalados,
                   SUM(CASE WHEN em.status_escala = 'Confirmado com rota' THEN 1 ELSE 0 END) AS confirmados_rota,
                   SUM(CASE WHEN em.status_escala = 'Cancelado, comparecer na base' THEN 1 ELSE 0 END) AS cancelados_base,
                   SUM(CASE WHEN em.status_presenca = 'Chegada confirmada' THEN 1 ELSE 0 END) AS chegadas_confirmadas,
                   SUM(CASE WHEN em.status_escala = 'Falta' THEN 1 ELSE 0 END) AS faltas,
                   SUM(CASE WHEN em.status_presenca = 'Não compareceu' THEN 1 ELSE 0 END) AS nao_compareceu
            FROM escala_motorista em
                     LEFT JOIN bases_operacionais bo
                               ON bo.id = em.base_operacional_id
                                  AND bo.empresa_id = em.empresa_id
            WHERE em.empresa_id = %s
              AND em.data_escala BETWEEN %s AND %s
              {filtro_base_sql}
            GROUP BY COALESCE(bo.nome_base, em.base_operacao, 'Sem base')
            ORDER BY total_escalados DESC
        """, [empresa_id, data_inicio, data_fim] + params_base)
        ranking_bases = cur.fetchall()
        for item in ranking_bases:
            item['taxa_comparecimento'] = percentual(item.get('chegadas_confirmadas'), item.get('cancelados_base'))
            item['taxa_falta'] = percentual(inteiro(item.get('faltas')) + inteiro(item.get('nao_compareceu')), item.get('total_escalados'))

        cur.execute(f"""
            SELECT mot.nome_completo AS motorista_nome,
                   COUNT(*) AS total
            FROM fila_cancelados_base f
                     INNER JOIN pessoas mot
                                ON mot.id = f.motorista_id
                                   AND mot.empresa_id = f.empresa_id
            WHERE f.empresa_id = %s
              AND f.data_fila BETWEEN %s AND %s
              AND f.status_fila = 'Atribuído para rota extra'
              {filtro_base_fila_sql}
            GROUP BY mot.nome_completo
            ORDER BY total DESC, mot.nome_completo ASC
            LIMIT 10
        """, [empresa_id, data_inicio, data_fim] + params_base_fila)
        ranking_aproveitados = cur.fetchall()

        cur.execute(f"""
            SELECT mot.nome_completo AS motorista_nome,
                   COUNT(*) AS total
            FROM escala_motorista em
                     INNER JOIN pessoas mot
                                ON mot.id = em.motorista_id
                                   AND mot.empresa_id = em.empresa_id
            WHERE em.empresa_id = %s
              AND em.data_escala BETWEEN %s AND %s
              AND em.status_escala = 'Falta'
              {filtro_base_sql}
            GROUP BY mot.nome_completo
            ORDER BY total DESC, mot.nome_completo ASC
            LIMIT 10
        """, [empresa_id, data_inicio, data_fim] + params_base)
        ranking_faltas = cur.fetchall()

        cur.execute(f"""
            SELECT mot.nome_completo AS motorista_nome,
                   COUNT(*) AS total
            FROM auditoria_checkin_base aud
                     INNER JOIN pessoas mot
                                ON mot.id = aud.motorista_id
                                   AND mot.empresa_id = aud.empresa_id
            WHERE aud.empresa_id = %s
              AND DATE(aud.data_tentativa) BETWEEN %s AND %s
              AND aud.resultado LIKE 'Bloqueado%%'
              {filtro_base_aud_sql}
            GROUP BY mot.nome_completo
            ORDER BY total DESC, mot.nome_completo ASC
            LIMIT 10
        """, [empresa_id, data_inicio, data_fim] + params_base_aud)
        ranking_bloqueios_motorista = cur.fetchall()

        cur.execute(f"""
            SELECT aud.resultado,
                   COALESCE(aud.motivo_bloqueio, '') AS motivo_bloqueio,
                   COUNT(*) AS total
            FROM auditoria_checkin_base aud
            WHERE aud.empresa_id = %s
              AND DATE(aud.data_tentativa) BETWEEN %s AND %s
              AND aud.resultado LIKE 'Bloqueado%%'
              {filtro_base_aud_sql}
            GROUP BY aud.resultado, COALESCE(aud.motivo_bloqueio, '')
            ORDER BY total DESC
            LIMIT 12
        """, [empresa_id, data_inicio, data_fim] + params_base_aud)
        bloqueios_por_motivo = cur.fetchall()

        cur.execute(f"""
            SELECT em.data_escala AS data_ref,
                   COUNT(*) AS total_escalados,
                   SUM(CASE WHEN em.status_escala = 'Confirmado com rota' THEN 1 ELSE 0 END) AS confirmados_rota,
                   SUM(CASE WHEN em.status_escala = 'Cancelado, comparecer na base' THEN 1 ELSE 0 END) AS cancelados_base,
                   SUM(CASE WHEN em.status_presenca = 'Chegada confirmada' THEN 1 ELSE 0 END) AS chegadas_confirmadas,
                   SUM(CASE WHEN em.status_escala = 'Falta' THEN 1 ELSE 0 END) AS faltas,
                   SUM(CASE WHEN em.status_presenca = 'Não compareceu' THEN 1 ELSE 0 END) AS nao_compareceu
            FROM escala_motorista em
            WHERE em.empresa_id = %s
              AND em.data_escala BETWEEN %s AND %s
              {filtro_base_sql}
            GROUP BY em.data_escala
            ORDER BY em.data_escala ASC
        """, [empresa_id, data_inicio, data_fim] + params_base)
        serie_diaria = cur.fetchall()
        for item in serie_diaria:
            item['data_label'] = formatar_data_br(item.get('data_ref'))

        cur.execute(f"""
            SELECT aud.data_tentativa,
                   mot.nome_completo AS motorista_nome,
                   COALESCE(bo.nome_base, 'Sem base') AS base_nome,
                   aud.resultado,
                   aud.motivo_bloqueio,
                   aud.distancia_base_metros
            FROM auditoria_checkin_base aud
                     LEFT JOIN pessoas mot
                               ON mot.id = aud.motorista_id
                                  AND mot.empresa_id = aud.empresa_id
                     LEFT JOIN bases_operacionais bo
                               ON bo.id = aud.base_operacional_id
                                  AND bo.empresa_id = aud.empresa_id
            WHERE aud.empresa_id = %s
              AND DATE(aud.data_tentativa) BETWEEN %s AND %s
              {filtro_base_aud_sql}
            ORDER BY aud.data_tentativa DESC
            LIMIT 8
        """, [empresa_id, data_inicio, data_fim] + params_base_aud)
        ultimas_tentativas = cur.fetchall()

        charts_data = {
            'serie_labels': [item.get('data_label') for item in serie_diaria],
            'serie_escalados': [inteiro(item.get('total_escalados')) for item in serie_diaria],
            'serie_chegadas': [inteiro(item.get('chegadas_confirmadas')) for item in serie_diaria],
            'serie_faltas': [inteiro(item.get('faltas')) + inteiro(item.get('nao_compareceu')) for item in serie_diaria],
            'bases_labels': [item.get('base_nome') for item in ranking_bases[:8]],
            'bases_comparecimento': [float(item.get('taxa_comparecimento') or 0) for item in ranking_bases[:8]],
            'bases_faltas': [float(item.get('taxa_falta') or 0) for item in ranking_bases[:8]],
            'bloqueios_labels': [item.get('resultado') for item in bloqueios_por_motivo[:8]],
            'bloqueios_totais': [inteiro(item.get('total')) for item in bloqueios_por_motivo[:8]],
            'status_labels': ['Aprovados', 'Bloqueados'],
            'status_totais': [inteiro(resumo.get('checkins_aprovados')), inteiro(resumo.get('checkins_bloqueados'))]
        }

        periodo_label = f"{formatar_data_br(data_inicio)} até {formatar_data_br(data_fim)}"
        base_nome_filtro = 'Todas as bases'
        if base_id_filtro and base_id_filtro.isdigit():
            for base in bases_operacionais:
                if str(base.get('id')) == str(base_id_filtro):
                    base_nome_filtro = base.get('nome_base') or 'Base selecionada'
                    break

        return render_template(
            'relatorio_operacional_escala_base.html',
            usuario_logado=usuario_logado,
            resumo=resumo,
            resumo_anterior=resumo_anterior,
            alertas_operacionais=alertas_operacionais,
            bases_operacionais=bases_operacionais,
            ranking_bases=ranking_bases,
            ranking_aproveitados=ranking_aproveitados,
            ranking_faltas=ranking_faltas,
            ranking_bloqueios_motorista=ranking_bloqueios_motorista,
            bloqueios_por_motivo=bloqueios_por_motivo,
            serie_diaria=serie_diaria,
            ultimas_tentativas=ultimas_tentativas,
            charts_data=charts_data,
            periodo_label=periodo_label,
            base_nome_filtro=base_nome_filtro,
            data_inicio=data_inicio_str,
            data_fim=data_fim_str,
            base_id_filtro=base_id_filtro,
            is_super_admin=is_super_admin,
            empresa_id_filtro=empresa_id if is_super_admin else ''
        )

    except Exception as e:
        print(f"Erro ao carregar relatório operacional: {e}")
        flash(f"Erro técnico ao carregar relatório operacional: {e}", "danger")
        return redirect(url_for('dashboard'))

    finally:
        cur.close()
        con.close()


@app.route('/operacao/relatorio-escala-base/exportar-csv', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor')
def exportar_relatorio_operacional_escala_base_csv():
    empresa_id = session.get('empresa_id')

    if not empresa_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    hoje = date.today()
    data_inicio_str = request.args.get('data_inicio', hoje.strftime('%Y-%m-%d')).strip()
    data_fim_str = request.args.get('data_fim', hoje.strftime('%Y-%m-%d')).strip()

    try:
        data_inicio = datetime.strptime(data_inicio_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_inicio = hoje

    try:
        data_fim = datetime.strptime(data_fim_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_fim = hoje

    con = obter_conexao()

    if con is None:
        flash("Erro de conexão ao exportar relatório.", "danger")
        return redirect(url_for('relatorio_operacional_escala_base'))

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
                    SELECT em.data_escala,
                           mot.nome_completo                                    AS motorista,
                           COALESCE(bo.nome_base, em.base_operacao, 'Sem base') AS base,
                           em.status_escala,
                           em.status_presenca,
                           em.presenca_confirmada_em,
                           f.status_fila,
                           f.posicao_fila,
                           f.distancia_base_metros,
                           f.geolocalizacao_validada,
                           f.qr_code_validado,
                           f.selfie_path
                    FROM escala_motorista em
                             INNER JOIN pessoas mot
                                        ON mot.id = em.motorista_id
                                            AND mot.empresa_id = em.empresa_id
                             LEFT JOIN bases_operacionais bo
                                       ON bo.id = em.base_operacional_id
                                           AND bo.empresa_id = em.empresa_id
                             LEFT JOIN fila_cancelados_base f
                                       ON f.escala_id = em.id
                                           AND f.empresa_id = em.empresa_id
                    WHERE em.empresa_id = %s
                      AND em.data_escala BETWEEN %s AND %s
                    ORDER BY em.data_escala DESC, mot.nome_completo ASC
                    """, (empresa_id, data_inicio, data_fim))

        rows = cur.fetchall()

        linhas = [
            "Data escala;Motorista;Base;Status escala;Status presenca;Chegada confirmada em;Status fila;Posicao fila;Distancia metros;GPS validado;QR validado;Selfie"
        ]

        for row in rows:
            linhas.append(
                f"{row.get('data_escala')};"
                f"{row.get('motorista') or ''};"
                f"{row.get('base') or ''};"
                f"{row.get('status_escala') or ''};"
                f"{row.get('status_presenca') or ''};"
                f"{row.get('presenca_confirmada_em') or ''};"
                f"{row.get('status_fila') or ''};"
                f"{row.get('posicao_fila') or ''};"
                f"{row.get('distancia_base_metros') or ''};"
                f"{row.get('geolocalizacao_validada') or ''};"
                f"{row.get('qr_code_validado') or ''};"
                f"{row.get('selfie_path') or ''}"
            )

        conteudo = "\n".join(linhas)

        return Response(
            conteudo,
            mimetype="text/csv; charset=utf-8",
            headers={
                "Content-Disposition": "attachment; filename=relatorio_operacional_escala_base.csv"
            }
        )

    except Exception as e:
        print(f"Erro ao exportar relatório operacional: {e}")
        flash(f"Erro técnico ao exportar relatório: {e}", "danger")
        return redirect(url_for('relatorio_operacional_escala_base'))

    finally:
        cur.close()
        con.close()


@app.route('/operacao/relatorio-escala-base/exportar-excel', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor')
def exportar_relatorio_operacional_escala_base_excel():
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    hoje = date.today()
    data_inicio_str = request.args.get('data_inicio', hoje.strftime('%Y-%m-%d')).strip()
    data_fim_str = request.args.get('data_fim', hoje.strftime('%Y-%m-%d')).strip()
    base_id_filtro = request.args.get('base_operacional_id', '').strip()
    empresa_id_filtro = request.args.get('empresa_id', '').strip()

    if is_super_admin and empresa_id_filtro and empresa_id_filtro.isdigit():
        empresa_id = int(empresa_id_filtro)
    else:
        empresa_id = empresa_logada_id

    try:
        data_inicio = datetime.strptime(data_inicio_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_inicio = hoje

    try:
        data_fim = datetime.strptime(data_fim_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_fim = hoje

    if data_fim < data_inicio:
        data_inicio, data_fim = data_fim, data_inicio

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão ao exportar Excel.", "danger")
        return redirect(url_for('relatorio_operacional_escala_base'))

    cur = con.cursor(dictionary=True)

    try:
        try:
            from io import BytesIO
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            from openpyxl.utils import get_column_letter
        except Exception:
            flash("Biblioteca openpyxl não instalada. Rode: pip install openpyxl", "danger")
            return redirect(url_for('relatorio_operacional_escala_base'))

        filtro_base_escala_sql = ""
        filtro_base_fila_sql = ""
        filtro_base_aud_sql = ""
        params_base_escala = []
        params_base_fila = []
        params_base_aud = []

        if base_id_filtro and base_id_filtro.isdigit():
            base_id_int = int(base_id_filtro)
            filtro_base_escala_sql = " AND em.base_operacional_id = %s "
            filtro_base_fila_sql = " AND f.base_operacional_id = %s "
            filtro_base_aud_sql = " AND aud.base_operacional_id = %s "
            params_base_escala.append(base_id_int)
            params_base_fila.append(base_id_int)
            params_base_aud.append(base_id_int)

        def inteiro(valor):
            try:
                return int(valor or 0)
            except Exception:
                return 0

        def percentual(parte, total):
            parte = inteiro(parte)
            total = inteiro(total)
            return round((parte / total) * 100, 2) if total else 0

        def valor_excel(valor):
            if isinstance(valor, Decimal):
                return float(valor)
            return valor

        def nome_empresa():
            cur.execute("""
                SELECT COALESCE(nome_fantasia, razao_social, CONCAT('Empresa ', id)) AS nome
                FROM empresas
                WHERE id = %s
            """, (empresa_id,))
            row = cur.fetchone() or {}
            return row.get('nome') or f"Empresa {empresa_id}"

        empresa_nome = nome_empresa()
        base_nome_filtro = 'Todas as bases'
        if base_id_filtro and base_id_filtro.isdigit():
            cur.execute("""
                SELECT nome_base
                FROM bases_operacionais
                WHERE empresa_id = %s AND id = %s
            """, (empresa_id, int(base_id_filtro)))
            row_base = cur.fetchone() or {}
            base_nome_filtro = row_base.get('nome_base') or 'Base selecionada'

        cur.execute(f"""
            SELECT
                COUNT(*) AS total_escalados,
                SUM(CASE WHEN em.status_escala = 'Confirmado com rota' THEN 1 ELSE 0 END) AS confirmados_rota,
                SUM(CASE WHEN em.status_escala = 'Cancelado, comparecer na base' THEN 1 ELSE 0 END) AS cancelados_comparecer_base,
                SUM(CASE WHEN em.status_escala = 'Cancelado definitivo' THEN 1 ELSE 0 END) AS cancelados_definitivos,
                SUM(CASE WHEN em.status_escala = 'Reserva / Avulso' THEN 1 ELSE 0 END) AS reservas_avulsos,
                SUM(CASE WHEN em.status_escala = 'Falta' THEN 1 ELSE 0 END) AS faltas,
                SUM(CASE WHEN em.status_presenca = 'Chegada confirmada' THEN 1 ELSE 0 END) AS chegadas_confirmadas,
                SUM(CASE WHEN em.status_presenca = 'Não compareceu' THEN 1 ELSE 0 END) AS nao_compareceu,
                SUM(CASE WHEN em.falta_automatica = 'S' THEN 1 ELSE 0 END) AS faltas_automaticas
            FROM escala_motorista em
            WHERE em.empresa_id = %s
              AND em.data_escala BETWEEN %s AND %s
              {filtro_base_escala_sql}
        """, [empresa_id, data_inicio, data_fim] + params_base_escala)
        kpis_escala = cur.fetchone() or {}

        cur.execute(f"""
            SELECT
                COUNT(*) AS total_fila,
                SUM(CASE WHEN f.status_fila = 'Aguardando rota' THEN 1 ELSE 0 END) AS aguardando_rota,
                SUM(CASE WHEN f.status_fila = 'Atribuído para rota extra' THEN 1 ELSE 0 END) AS atribuidos_rota_extra,
                SUM(CASE WHEN f.status_fila = 'Dispensado' THEN 1 ELSE 0 END) AS dispensados,
                SUM(CASE WHEN f.status_fila = 'Falta' THEN 1 ELSE 0 END) AS faltas_fila
            FROM fila_cancelados_base f
            WHERE f.empresa_id = %s
              AND f.data_fila BETWEEN %s AND %s
              {filtro_base_fila_sql}
        """, [empresa_id, data_inicio, data_fim] + params_base_fila)
        kpis_fila = cur.fetchone() or {}

        cur.execute(f"""
            SELECT
                COUNT(*) AS total_tentativas,
                SUM(CASE WHEN aud.resultado = 'Aprovado' THEN 1 ELSE 0 END) AS checkins_aprovados,
                SUM(CASE WHEN aud.resultado LIKE 'Bloqueado%%' THEN 1 ELSE 0 END) AS checkins_bloqueados,
                SUM(CASE WHEN aud.resultado = 'Bloqueado por distância' THEN 1 ELSE 0 END) AS bloqueados_distancia,
                SUM(CASE WHEN aud.resultado = 'Bloqueado por QR inválido/expirado' THEN 1 ELSE 0 END) AS bloqueados_qr,
                SUM(CASE WHEN aud.resultado = 'Bloqueado por falta de selfie' THEN 1 ELSE 0 END) AS bloqueados_selfie,
                SUM(CASE WHEN aud.resultado = 'Bloqueado por falta de geolocalização' THEN 1 ELSE 0 END) AS bloqueados_geo,
                SUM(CASE WHEN aud.resultado = 'Bloqueado por base sem coordenadas' THEN 1 ELSE 0 END) AS bloqueados_base_sem_coordenada
            FROM auditoria_checkin_base aud
            WHERE aud.empresa_id = %s
              AND DATE(aud.data_tentativa) BETWEEN %s AND %s
              {filtro_base_aud_sql}
        """, [empresa_id, data_inicio, data_fim] + params_base_aud)
        kpis_aud = cur.fetchone() or {}

        resumo = {}
        resumo.update({k: inteiro(v) for k, v in kpis_escala.items()})
        resumo.update({k: inteiro(v) for k, v in kpis_fila.items()})
        resumo.update({k: inteiro(v) for k, v in kpis_aud.items()})
        resumo['taxa_comparecimento_base'] = percentual(resumo.get('chegadas_confirmadas'), resumo.get('cancelados_comparecer_base'))
        resumo['taxa_aproveitamento_fila'] = percentual(resumo.get('atribuidos_rota_extra'), resumo.get('total_fila'))
        resumo['taxa_bloqueio_checkin'] = percentual(resumo.get('checkins_bloqueados'), resumo.get('total_tentativas'))
        resumo['taxa_checkin_aprovado'] = percentual(resumo.get('checkins_aprovados'), resumo.get('total_tentativas'))
        resumo['taxa_falta'] = percentual(resumo.get('faltas') + resumo.get('nao_compareceu'), resumo.get('total_escalados'))

        cur.execute(f"""
            SELECT COALESCE(bo.nome_base, em.base_operacao, 'Sem base') AS base,
                   COUNT(*) AS total_escalados,
                   SUM(CASE WHEN em.status_escala = 'Cancelado, comparecer na base' THEN 1 ELSE 0 END) AS convocados_base,
                   SUM(CASE WHEN em.status_presenca = 'Chegada confirmada' THEN 1 ELSE 0 END) AS chegadas_confirmadas,
                   SUM(CASE WHEN em.status_escala = 'Falta' OR em.status_presenca = 'Não compareceu' THEN 1 ELSE 0 END) AS faltas,
                   ROUND((SUM(CASE WHEN em.status_presenca = 'Chegada confirmada' THEN 1 ELSE 0 END) / NULLIF(SUM(CASE WHEN em.status_escala = 'Cancelado, comparecer na base' THEN 1 ELSE 0 END), 0)) * 100, 2) AS taxa_comparecimento,
                   ROUND((SUM(CASE WHEN em.status_escala = 'Falta' OR em.status_presenca = 'Não compareceu' THEN 1 ELSE 0 END) / NULLIF(COUNT(*), 0)) * 100, 2) AS taxa_falta
            FROM escala_motorista em
                     LEFT JOIN bases_operacionais bo
                               ON bo.id = em.base_operacional_id
                                  AND bo.empresa_id = em.empresa_id
            WHERE em.empresa_id = %s
              AND em.data_escala BETWEEN %s AND %s
              {filtro_base_escala_sql}
            GROUP BY base
            ORDER BY total_escalados DESC, base ASC
        """, [empresa_id, data_inicio, data_fim] + params_base_escala)
        dados_por_base = cur.fetchall()

        cur.execute(f"""
            SELECT em.data_escala,
                   mot.nome_completo AS motorista,
                   COALESCE(bo.nome_base, em.base_operacao, 'Sem base') AS base,
                   em.status_escala,
                   em.status_presenca,
                   em.horario_apresentacao,
                   em.presenca_confirmada_em,
                   em.observacao_supervisor,
                   f.status_fila,
                   f.posicao_fila,
                   f.hora_confirmacao,
                   f.distancia_base_metros,
                   f.geolocalizacao_validada,
                   f.qr_code_validado,
                   f.selfie_path
            FROM escala_motorista em
                     INNER JOIN pessoas mot
                                ON mot.id = em.motorista_id
                                   AND mot.empresa_id = em.empresa_id
                     LEFT JOIN bases_operacionais bo
                               ON bo.id = em.base_operacional_id
                                  AND bo.empresa_id = em.empresa_id
                     LEFT JOIN fila_cancelados_base f
                               ON f.escala_id = em.id
                                  AND f.empresa_id = em.empresa_id
            WHERE em.empresa_id = %s
              AND em.data_escala BETWEEN %s AND %s
              {filtro_base_escala_sql}
            ORDER BY em.data_escala DESC, mot.nome_completo ASC
        """, [empresa_id, data_inicio, data_fim] + params_base_escala)
        detalhe_escala = cur.fetchall()

        cur.execute(f"""
            SELECT aud.data_tentativa,
                   mot.nome_completo AS motorista,
                   COALESCE(bo.nome_base, 'Sem base') AS base,
                   aud.resultado,
                   aud.motivo_bloqueio,
                   aud.distancia_base_metros,
                   aud.codigo_qr_informado,
                   aud.qr_token_id,
                   aud.selfie_path,
                   aud.ip_origem,
                   aud.user_agent
            FROM auditoria_checkin_base aud
                     LEFT JOIN pessoas mot
                               ON mot.id = aud.motorista_id
                                  AND mot.empresa_id = aud.empresa_id
                     LEFT JOIN bases_operacionais bo
                               ON bo.id = aud.base_operacional_id
                                  AND bo.empresa_id = aud.empresa_id
            WHERE aud.empresa_id = %s
              AND DATE(aud.data_tentativa) BETWEEN %s AND %s
              {filtro_base_aud_sql}
            ORDER BY aud.data_tentativa DESC
        """, [empresa_id, data_inicio, data_fim] + params_base_aud)
        checkins = cur.fetchall()

        cur.execute(f"""
            SELECT aud.resultado,
                   COALESCE(NULLIF(aud.motivo_bloqueio, ''), aud.resultado) AS motivo,
                   COUNT(*) AS total
            FROM auditoria_checkin_base aud
            WHERE aud.empresa_id = %s
              AND DATE(aud.data_tentativa) BETWEEN %s AND %s
              {filtro_base_aud_sql}
            GROUP BY aud.resultado, motivo
            ORDER BY total DESC
        """, [empresa_id, data_inicio, data_fim] + params_base_aud)
        bloqueios_motivos = cur.fetchall()

        cur.execute(f"""
            SELECT mot.nome_completo AS motorista,
                   COUNT(*) AS total_rota_extra
            FROM fila_cancelados_base f
                     INNER JOIN pessoas mot
                                ON mot.id = f.motorista_id
                                   AND mot.empresa_id = f.empresa_id
            WHERE f.empresa_id = %s
              AND f.data_fila BETWEEN %s AND %s
              AND f.status_fila = 'Atribuído para rota extra'
              {filtro_base_fila_sql}
            GROUP BY f.motorista_id, mot.nome_completo
            ORDER BY total_rota_extra DESC, mot.nome_completo ASC
        """, [empresa_id, data_inicio, data_fim] + params_base_fila)
        ranking_aproveitados = cur.fetchall()

        cur.execute(f"""
            SELECT mot.nome_completo AS motorista,
                   COUNT(*) AS total_faltas
            FROM escala_motorista em
                     INNER JOIN pessoas mot
                                ON mot.id = em.motorista_id
                                   AND mot.empresa_id = em.empresa_id
            WHERE em.empresa_id = %s
              AND em.data_escala BETWEEN %s AND %s
              AND (em.status_escala = 'Falta' OR em.status_presenca = 'Não compareceu')
              {filtro_base_escala_sql}
            GROUP BY em.motorista_id, mot.nome_completo
            ORDER BY total_faltas DESC, mot.nome_completo ASC
        """, [empresa_id, data_inicio, data_fim] + params_base_escala)
        ranking_faltas = cur.fetchall()

        cur.execute(f"""
            SELECT mot.nome_completo AS motorista,
                   COUNT(*) AS total_bloqueios
            FROM auditoria_checkin_base aud
                     LEFT JOIN pessoas mot
                               ON mot.id = aud.motorista_id
                                  AND mot.empresa_id = aud.empresa_id
            WHERE aud.empresa_id = %s
              AND DATE(aud.data_tentativa) BETWEEN %s AND %s
              AND aud.resultado LIKE 'Bloqueado%%'
              {filtro_base_aud_sql}
            GROUP BY aud.motorista_id, mot.nome_completo
            ORDER BY total_bloqueios DESC, mot.nome_completo ASC
        """, [empresa_id, data_inicio, data_fim] + params_base_aud)
        ranking_bloqueios = cur.fetchall()

        cur.execute("""
            SELECT a.data_acao,
                   COALESCE(p.nome_completo, u.login, CONCAT('Usuário ', a.usuario_id)) AS usuario,
                   a.tipo_acao,
                   a.descricao,
                   a.ip_origem
            FROM auditoria_supervisor a
                     LEFT JOIN usuarios u ON u.id = a.usuario_id
                     LEFT JOIN pessoas p ON p.id = u.pessoa_id
            WHERE a.empresa_id = %s
              AND DATE(a.data_acao) BETWEEN %s AND %s
            ORDER BY a.data_acao DESC
        """, (empresa_id, data_inicio, data_fim))
        auditoria_supervisor = cur.fetchall()

        wb = Workbook()
        wb.remove(wb.active)

        cor_primaria = "1F4E79"
        cor_secundaria = "EAF2F8"
        cor_escura = "111827"
        cor_alerta = "FCE4D6"
        header_fill = PatternFill("solid", fgColor=cor_primaria)
        title_fill = PatternFill("solid", fgColor=cor_escura)
        soft_fill = PatternFill("solid", fgColor=cor_secundaria)
        alerta_fill = PatternFill("solid", fgColor=cor_alerta)
        white_font = Font(color="FFFFFF", bold=True)
        title_font = Font(color="FFFFFF", bold=True, size=14)
        header_font = Font(color="FFFFFF", bold=True)
        bold_font = Font(bold=True)
        border = Border(
            left=Side(style="thin", color="D9E2F3"),
            right=Side(style="thin", color="D9E2F3"),
            top=Side(style="thin", color="D9E2F3"),
            bottom=Side(style="thin", color="D9E2F3"),
        )

        def criar_aba(nome, titulo, headers=None):
            ws = wb.create_sheet(nome[:31])
            largura = max(len(headers or []), 6)
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=largura)
            cell = ws.cell(row=1, column=1, value=titulo)
            cell.fill = title_fill
            cell.font = title_font
            cell.alignment = Alignment(horizontal="center")
            ws.row_dimensions[1].height = 24
            if headers:
                for col, header in enumerate(headers, 1):
                    c = ws.cell(row=3, column=col, value=header)
                    c.fill = header_fill
                    c.font = header_font
                    c.alignment = Alignment(horizontal="center", vertical="center")
                    c.border = border
                ws.freeze_panes = "A4"
                ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}3"
            return ws

        def preencher_linhas(ws, headers, rows, start_row=4):
            for r_idx, row in enumerate(rows, start_row):
                for c_idx, header in enumerate(headers, 1):
                    valor = row.get(header, '') if isinstance(row, dict) else ''
                    c = ws.cell(row=r_idx, column=c_idx, value=valor_excel(valor))
                    c.border = border
                    c.alignment = Alignment(vertical="top", wrap_text=True)
                    if isinstance(valor, (datetime, date)):
                        c.number_format = "dd/mm/yyyy hh:mm" if isinstance(valor, datetime) else "dd/mm/yyyy"
                    elif isinstance(valor, (Decimal, float)):
                        c.number_format = "#,##0.00"
            return start_row + len(rows)

        def ajustar_colunas(ws, max_width=44):
            for col in ws.columns:
                col_letter = get_column_letter(col[0].column)
                maior = 10
                for cell in col[:200]:
                    if cell.value is not None:
                        maior = max(maior, min(len(str(cell.value)) + 2, max_width))
                ws.column_dimensions[col_letter].width = maior

        def rows_dict(rows, mapa):
            saida = []
            for row in rows:
                item = {}
                for header, key in mapa:
                    item[header] = row.get(key)
                saida.append(item)
            return saida

        ws = criar_aba("Resumo Geral", "Resumo Geral - Relatório Operacional Premium")
        resumo_linhas = [
            ("Empresa", empresa_nome),
            ("Período", f"{formatar_data_br(data_inicio)} até {formatar_data_br(data_fim)}"),
            ("Base", base_nome_filtro),
            ("Gerado em", datetime.now().strftime('%d/%m/%Y %H:%M')),
            ("", ""),
            ("Motoristas escalados", resumo.get('total_escalados')),
            ("Confirmados com rota", resumo.get('confirmados_rota')),
            ("Cancelados convocados para base", resumo.get('cancelados_comparecer_base')),
            ("Chegadas confirmadas", resumo.get('chegadas_confirmadas')),
            ("Não compareceu", resumo.get('nao_compareceu')),
            ("Faltas", resumo.get('faltas')),
            ("Fila total", resumo.get('total_fila')),
            ("Aguardando rota", resumo.get('aguardando_rota')),
            ("Puxados para rota extra", resumo.get('atribuidos_rota_extra')),
            ("Dispensados", resumo.get('dispensados')),
            ("Check-ins aprovados", resumo.get('checkins_aprovados')),
            ("Check-ins bloqueados", resumo.get('checkins_bloqueados')),
            ("Taxa de comparecimento na base", resumo.get('taxa_comparecimento_base') / 100),
            ("Taxa de aproveitamento da fila", resumo.get('taxa_aproveitamento_fila') / 100),
            ("Taxa de bloqueio no check-in", resumo.get('taxa_bloqueio_checkin') / 100),
            ("Taxa de falta", resumo.get('taxa_falta') / 100),
        ]
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
        ws.cell(3, 1, "Indicador").fill = header_fill
        ws.cell(3, 1).font = header_font
        ws.cell(3, 2, "Valor").fill = header_fill
        ws.cell(3, 2).font = header_font
        for idx, (indicador, valor) in enumerate(resumo_linhas, 4):
            ws.cell(idx, 1, indicador)
            ws.cell(idx, 2, valor_excel(valor))
            for col in [1, 2]:
                ws.cell(idx, col).border = border
                ws.cell(idx, col).alignment = Alignment(vertical="top", wrap_text=True)
            if indicador == "":
                ws.cell(idx, 1).fill = soft_fill
                ws.cell(idx, 2).fill = soft_fill
            if "Taxa" in indicador:
                ws.cell(idx, 2).number_format = "0.00%"
        ajustar_colunas(ws)

        headers = ["Base", "Total escalados", "Convocados base", "Chegadas confirmadas", "Faltas", "Taxa comparecimento", "Taxa falta"]
        ws = criar_aba("Por Base", "Resultado por Base", headers)
        dados = rows_dict(dados_por_base, [
            ("Base", "base"), ("Total escalados", "total_escalados"), ("Convocados base", "convocados_base"),
            ("Chegadas confirmadas", "chegadas_confirmadas"), ("Faltas", "faltas"),
            ("Taxa comparecimento", "taxa_comparecimento"), ("Taxa falta", "taxa_falta")
        ])
        preencher_linhas(ws, headers, dados)
        ajustar_colunas(ws)

        headers = ["Data escala", "Motorista", "Base", "Status escala", "Status presença", "Horário", "Chegada confirmada em", "Orientação supervisor", "Status fila", "Posição fila", "Hora confirmação", "Distância metros", "GPS validado", "QR validado", "Selfie"]
        ws = criar_aba("Escala e Fila", "Detalhe de Escala e Fila", headers)
        dados = rows_dict(detalhe_escala, [
            ("Data escala", "data_escala"), ("Motorista", "motorista"), ("Base", "base"), ("Status escala", "status_escala"),
            ("Status presença", "status_presenca"), ("Horário", "horario_apresentacao"), ("Chegada confirmada em", "presenca_confirmada_em"),
            ("Orientação supervisor", "observacao_supervisor"), ("Status fila", "status_fila"), ("Posição fila", "posicao_fila"),
            ("Hora confirmação", "hora_confirmacao"), ("Distância metros", "distancia_base_metros"), ("GPS validado", "geolocalizacao_validada"),
            ("QR validado", "qr_code_validado"), ("Selfie", "selfie_path")
        ])
        preencher_linhas(ws, headers, dados)
        ajustar_colunas(ws)

        headers = ["Data tentativa", "Motorista", "Base", "Resultado", "Motivo bloqueio", "Distância metros", "QR informado", "Token QR", "Selfie", "IP", "Dispositivo/Navegador"]
        ws = criar_aba("Check-ins", "Auditoria de Check-ins", headers)
        dados = rows_dict(checkins, [
            ("Data tentativa", "data_tentativa"), ("Motorista", "motorista"), ("Base", "base"), ("Resultado", "resultado"),
            ("Motivo bloqueio", "motivo_bloqueio"), ("Distância metros", "distancia_base_metros"), ("QR informado", "codigo_qr_informado"),
            ("Token QR", "qr_token_id"), ("Selfie", "selfie_path"), ("IP", "ip_origem"), ("Dispositivo/Navegador", "user_agent")
        ])
        preencher_linhas(ws, headers, dados)
        ajustar_colunas(ws)

        headers = ["Resultado", "Motivo", "Total"]
        ws = criar_aba("Bloqueios Motivo", "Bloqueios por Motivo", headers)
        dados = rows_dict(bloqueios_motivos, [("Resultado", "resultado"), ("Motivo", "motivo"), ("Total", "total")])
        preencher_linhas(ws, headers, dados)
        ajustar_colunas(ws)

        headers = ["Tipo", "Motorista", "Total"]
        ws = criar_aba("Ranking Motoristas", "Ranking de Motoristas", headers)
        ranking_rows = []
        for row in ranking_aproveitados:
            ranking_rows.append({"Tipo": "Mais aproveitados em rota extra", "Motorista": row.get('motorista'), "Total": row.get('total_rota_extra')})
        for row in ranking_faltas:
            ranking_rows.append({"Tipo": "Mais faltas", "Motorista": row.get('motorista'), "Total": row.get('total_faltas')})
        for row in ranking_bloqueios:
            ranking_rows.append({"Tipo": "Mais bloqueios no check-in", "Motorista": row.get('motorista'), "Total": row.get('total_bloqueios')})
        preencher_linhas(ws, headers, ranking_rows)
        ajustar_colunas(ws)

        headers = ["Data ação", "Usuário", "Tipo ação", "Descrição", "IP"]
        ws = criar_aba("Auditoria Supervisor", "Auditoria de Ações do Supervisor", headers)
        dados = rows_dict(auditoria_supervisor, [
            ("Data ação", "data_acao"), ("Usuário", "usuario"), ("Tipo ação", "tipo_acao"),
            ("Descrição", "descricao"), ("IP", "ip_origem")
        ])
        preencher_linhas(ws, headers, dados)
        ajustar_colunas(ws)

        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if cell.row >= 3 and cell.value is not None:
                        cell.border = border
                        if cell.row == 3:
                            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                        else:
                            cell.alignment = Alignment(vertical="top", wrap_text=True)
            ws.sheet_view.showGridLines = False

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        nome_arquivo = f"relatorio_operacional_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.xlsx"
        return Response(
            output.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={nome_arquivo}"}
        )

    except Exception as e:
        print(f"Erro ao exportar relatório operacional em Excel: {e}")
        flash(f"Erro técnico ao exportar Excel: {e}", "danger")
        return redirect(url_for('relatorio_operacional_escala_base'))

    finally:
        cur.close()
        con.close()



@app.route('/operacao/relatorio-escala-base/exportar-pdf', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor')
def exportar_relatorio_operacional_escala_base_pdf():
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    hoje = date.today()
    data_inicio_str = request.args.get('data_inicio', hoje.strftime('%Y-%m-%d')).strip()
    data_fim_str = request.args.get('data_fim', hoje.strftime('%Y-%m-%d')).strip()
    base_id_filtro = request.args.get('base_operacional_id', '').strip()
    empresa_id_filtro = request.args.get('empresa_id', '').strip()

    if is_super_admin and empresa_id_filtro and empresa_id_filtro.isdigit():
        empresa_id = int(empresa_id_filtro)
    else:
        empresa_id = empresa_logada_id

    try:
        data_inicio = datetime.strptime(data_inicio_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_inicio = hoje

    try:
        data_fim = datetime.strptime(data_fim_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_fim = hoje

    if data_fim < data_inicio:
        data_inicio, data_fim = data_fim, data_inicio

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão ao exportar PDF.", "danger")
        return redirect(url_for('relatorio_operacional_escala_base'))

    cur = con.cursor(dictionary=True)

    try:
        try:
            from io import BytesIO
            from reportlab.lib import colors
            from reportlab.lib.enums import TA_CENTER, TA_LEFT
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        except Exception:
            flash("Biblioteca reportlab não instalada. Rode: pip install reportlab", "danger")
            return redirect(url_for('relatorio_operacional_escala_base'))

        filtro_base_escala_sql = ""
        filtro_base_fila_sql = ""
        filtro_base_aud_sql = ""
        params_base_escala = []
        params_base_fila = []
        params_base_aud = []

        if base_id_filtro and base_id_filtro.isdigit():
            base_id_int = int(base_id_filtro)
            filtro_base_escala_sql = " AND em.base_operacional_id = %s "
            filtro_base_fila_sql = " AND f.base_operacional_id = %s "
            filtro_base_aud_sql = " AND aud.base_operacional_id = %s "
            params_base_escala.append(base_id_int)
            params_base_fila.append(base_id_int)
            params_base_aud.append(base_id_int)

        def inteiro(valor):
            try:
                return int(valor or 0)
            except Exception:
                return 0

        def percentual(parte, total):
            parte = inteiro(parte)
            total = inteiro(total)
            return round((parte / total) * 100, 1) if total else 0

        def txt(valor):
            if valor is None:
                return ""
            if isinstance(valor, (datetime, date)):
                return valor.strftime('%d/%m/%Y %H:%M') if isinstance(valor, datetime) else valor.strftime('%d/%m/%Y')
            return str(valor)

        def pct(valor):
            try:
                return f"{float(valor or 0):.1f}%".replace('.', ',')
            except Exception:
                return "0,0%"

        cur.execute("""
            SELECT COALESCE(nome_fantasia, razao_social, CONCAT('Empresa ', id)) AS nome
            FROM empresas
            WHERE id = %s
        """, (empresa_id,))
        empresa_nome = (cur.fetchone() or {}).get('nome') or f"Empresa {empresa_id}"

        base_nome_filtro = 'Todas as bases'
        if base_id_filtro and base_id_filtro.isdigit():
            cur.execute("""
                SELECT nome_base
                FROM bases_operacionais
                WHERE empresa_id = %s AND id = %s
            """, (empresa_id, int(base_id_filtro)))
            base_nome_filtro = (cur.fetchone() or {}).get('nome_base') or 'Base selecionada'

        cur.execute(f"""
            SELECT
                COUNT(*) AS total_escalados,
                SUM(CASE WHEN em.status_escala = 'Confirmado com rota' THEN 1 ELSE 0 END) AS confirmados_rota,
                SUM(CASE WHEN em.status_escala = 'Cancelado, comparecer na base' THEN 1 ELSE 0 END) AS cancelados_comparecer_base,
                SUM(CASE WHEN em.status_escala = 'Falta' THEN 1 ELSE 0 END) AS faltas,
                SUM(CASE WHEN em.status_presenca = 'Chegada confirmada' THEN 1 ELSE 0 END) AS chegadas_confirmadas,
                SUM(CASE WHEN em.status_presenca = 'Não compareceu' THEN 1 ELSE 0 END) AS nao_compareceu
            FROM escala_motorista em
            WHERE em.empresa_id = %s
              AND em.data_escala BETWEEN %s AND %s
              {filtro_base_escala_sql}
        """, [empresa_id, data_inicio, data_fim] + params_base_escala)
        kpis_escala = cur.fetchone() or {}

        cur.execute(f"""
            SELECT
                COUNT(*) AS total_fila,
                SUM(CASE WHEN f.status_fila = 'Aguardando rota' THEN 1 ELSE 0 END) AS aguardando_rota,
                SUM(CASE WHEN f.status_fila = 'Atribuído para rota extra' THEN 1 ELSE 0 END) AS atribuidos_rota_extra,
                SUM(CASE WHEN f.status_fila = 'Dispensado' THEN 1 ELSE 0 END) AS dispensados
            FROM fila_cancelados_base f
            WHERE f.empresa_id = %s
              AND f.data_fila BETWEEN %s AND %s
              {filtro_base_fila_sql}
        """, [empresa_id, data_inicio, data_fim] + params_base_fila)
        kpis_fila = cur.fetchone() or {}

        cur.execute(f"""
            SELECT
                COUNT(*) AS total_tentativas,
                SUM(CASE WHEN aud.resultado = 'Aprovado' THEN 1 ELSE 0 END) AS checkins_aprovados,
                SUM(CASE WHEN aud.resultado LIKE 'Bloqueado%%' THEN 1 ELSE 0 END) AS checkins_bloqueados
            FROM auditoria_checkin_base aud
            WHERE aud.empresa_id = %s
              AND DATE(aud.data_tentativa) BETWEEN %s AND %s
              {filtro_base_aud_sql}
        """, [empresa_id, data_inicio, data_fim] + params_base_aud)
        kpis_aud = cur.fetchone() or {}

        resumo = {}
        resumo.update({k: inteiro(v) for k, v in kpis_escala.items()})
        resumo.update({k: inteiro(v) for k, v in kpis_fila.items()})
        resumo.update({k: inteiro(v) for k, v in kpis_aud.items()})
        resumo['taxa_comparecimento_base'] = percentual(resumo.get('chegadas_confirmadas'), resumo.get('cancelados_comparecer_base'))
        resumo['taxa_aproveitamento_fila'] = percentual(resumo.get('atribuidos_rota_extra'), resumo.get('total_fila'))
        resumo['taxa_bloqueio_checkin'] = percentual(resumo.get('checkins_bloqueados'), resumo.get('total_tentativas'))
        resumo['taxa_falta'] = percentual(resumo.get('faltas') + resumo.get('nao_compareceu'), resumo.get('total_escalados'))

        alertas = []
        if resumo.get('taxa_comparecimento_base', 0) < 90 and resumo.get('cancelados_comparecer_base', 0) > 0:
            alertas.append(f"Comparecimento abaixo da meta: {pct(resumo.get('taxa_comparecimento_base'))}. Meta sugerida: 90%.")
        if resumo.get('taxa_bloqueio_checkin', 0) >= 15 and resumo.get('total_tentativas', 0) > 0:
            alertas.append(f"Bloqueios elevados: {resumo.get('checkins_bloqueados')} bloqueios em {resumo.get('total_tentativas')} tentativas.")
        if resumo.get('aguardando_rota', 0) > 0:
            alertas.append(f"Fila com pendência: {resumo.get('aguardando_rota')} motorista(s) aguardando decisão.")
        if resumo.get('faltas', 0) + resumo.get('nao_compareceu', 0) > 0:
            alertas.append(f"Faltas/não comparecimentos no período: {resumo.get('faltas', 0) + resumo.get('nao_compareceu', 0)}.")
        if not alertas:
            alertas.append("Nenhum alerta crítico identificado para o período filtrado.")

        cur.execute(f"""
            SELECT COALESCE(bo.nome_base, em.base_operacao, 'Sem base') AS base,
                   COUNT(*) AS total_escalados,
                   SUM(CASE WHEN em.status_escala = 'Cancelado, comparecer na base' THEN 1 ELSE 0 END) AS convocados_base,
                   SUM(CASE WHEN em.status_presenca = 'Chegada confirmada' THEN 1 ELSE 0 END) AS chegadas_confirmadas,
                   SUM(CASE WHEN em.status_escala = 'Falta' OR em.status_presenca = 'Não compareceu' THEN 1 ELSE 0 END) AS faltas,
                   ROUND((SUM(CASE WHEN em.status_presenca = 'Chegada confirmada' THEN 1 ELSE 0 END) / NULLIF(SUM(CASE WHEN em.status_escala = 'Cancelado, comparecer na base' THEN 1 ELSE 0 END), 0)) * 100, 1) AS taxa_comparecimento
            FROM escala_motorista em
                LEFT JOIN bases_operacionais bo ON bo.id = em.base_operacional_id AND bo.empresa_id = em.empresa_id
            WHERE em.empresa_id = %s
              AND em.data_escala BETWEEN %s AND %s
              {filtro_base_escala_sql}
            GROUP BY base
            ORDER BY total_escalados DESC, base ASC
            LIMIT 12
        """, [empresa_id, data_inicio, data_fim] + params_base_escala)
        dados_por_base = cur.fetchall() or []

        cur.execute(f"""
            SELECT aud.resultado,
                   COALESCE(aud.motivo_bloqueio, aud.resultado, 'Sem motivo') AS motivo,
                   COUNT(*) AS total
            FROM auditoria_checkin_base aud
            WHERE aud.empresa_id = %s
              AND DATE(aud.data_tentativa) BETWEEN %s AND %s
              AND aud.resultado LIKE 'Bloqueado%%'
              {filtro_base_aud_sql}
            GROUP BY aud.resultado, motivo
            ORDER BY total DESC
            LIMIT 10
        """, [empresa_id, data_inicio, data_fim] + params_base_aud)
        bloqueios_motivos = cur.fetchall() or []

        cur.execute(f"""
            SELECT mot.nome_completo AS motorista, COUNT(*) AS total_rota_extra
            FROM fila_cancelados_base f
                INNER JOIN pessoas mot ON mot.id = f.motorista_id
            WHERE f.empresa_id = %s
              AND f.data_fila BETWEEN %s AND %s
              AND f.status_fila = 'Atribuído para rota extra'
              {filtro_base_fila_sql}
            GROUP BY f.motorista_id, mot.nome_completo
            ORDER BY total_rota_extra DESC, mot.nome_completo ASC
            LIMIT 8
        """, [empresa_id, data_inicio, data_fim] + params_base_fila)
        ranking_aproveitados = cur.fetchall() or []

        cur.execute(f"""
            SELECT mot.nome_completo AS motorista,
                   COUNT(*) AS total_faltas
            FROM escala_motorista em
                INNER JOIN pessoas mot ON mot.id = em.motorista_id
            WHERE em.empresa_id = %s
              AND em.data_escala BETWEEN %s AND %s
              AND (em.status_escala = 'Falta' OR em.status_presenca = 'Não compareceu')
              {filtro_base_escala_sql}
            GROUP BY em.motorista_id, mot.nome_completo
            ORDER BY total_faltas DESC, mot.nome_completo ASC
            LIMIT 8
        """, [empresa_id, data_inicio, data_fim] + params_base_escala)
        ranking_faltas = cur.fetchall() or []

        cur.execute(f"""
            SELECT mot.nome_completo AS motorista,
                   COUNT(*) AS total_bloqueios
            FROM auditoria_checkin_base aud
                INNER JOIN pessoas mot ON mot.id = aud.motorista_id
            WHERE aud.empresa_id = %s
              AND DATE(aud.data_tentativa) BETWEEN %s AND %s
              AND aud.resultado LIKE 'Bloqueado%%'
              {filtro_base_aud_sql}
            GROUP BY aud.motorista_id, mot.nome_completo
            ORDER BY total_bloqueios DESC, mot.nome_completo ASC
            LIMIT 8
        """, [empresa_id, data_inicio, data_fim] + params_base_aud)
        ranking_bloqueios = cur.fetchall() or []

        cur.execute(f"""
            SELECT aud.data_tentativa,
                   mot.nome_completo AS motorista,
                   COALESCE(bo.nome_base, 'Sem base') AS base,
                   aud.resultado,
                   aud.motivo_bloqueio,
                   aud.distancia_base_metros
            FROM auditoria_checkin_base aud
                LEFT JOIN pessoas mot ON mot.id = aud.motorista_id
                LEFT JOIN bases_operacionais bo ON bo.id = aud.base_operacional_id
            WHERE aud.empresa_id = %s
              AND DATE(aud.data_tentativa) BETWEEN %s AND %s
              {filtro_base_aud_sql}
            ORDER BY aud.data_tentativa DESC
            LIMIT 12
        """, [empresa_id, data_inicio, data_fim] + params_base_aud)
        ultimos_checkins = cur.fetchall() or []

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(A4),
            rightMargin=1.2 * cm,
            leftMargin=1.2 * cm,
            topMargin=1.3 * cm,
            bottomMargin=1.2 * cm,
            title="Relatório Operacional Gerencial"
        )
        styles = getSampleStyleSheet()
        styles.add(ParagraphStyle(name='TituloSGR', parent=styles['Title'], fontName='Helvetica-Bold', fontSize=20, leading=24, textColor=colors.HexColor('#111827'), alignment=TA_LEFT, spaceAfter=10))
        styles.add(ParagraphStyle(name='SubtituloSGR', parent=styles['Normal'], fontSize=9.5, leading=13, textColor=colors.HexColor('#4B5563'), alignment=TA_LEFT))
        styles.add(ParagraphStyle(name='SecaoSGR', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=12, leading=15, textColor=colors.HexColor('#111827'), spaceBefore=12, spaceAfter=6))
        styles.add(ParagraphStyle(name='CardLabelSGR', parent=styles['Normal'], fontSize=7.5, leading=9, textColor=colors.HexColor('#6B7280'), alignment=TA_CENTER))
        styles.add(ParagraphStyle(name='CardValorSGR', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=13, leading=15, textColor=colors.HexColor('#111827'), alignment=TA_CENTER))
        styles.add(ParagraphStyle(name='PequenoSGR', parent=styles['Normal'], fontSize=7.5, leading=9.5, textColor=colors.HexColor('#374151')))
        styles.add(ParagraphStyle(name='HeaderTabelaSGR', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=7.5, leading=9.5, textColor=colors.white))

        def p(valor, style='PequenoSGR'):
            texto = txt(valor).replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            return Paragraph(texto, styles[style])

        elementos = []
        elementos.append(Paragraph("Relatório Operacional Gerencial", styles['TituloSGR']))
        elementos.append(Paragraph(f"Empresa: <b>{empresa_nome}</b> &nbsp;&nbsp; | &nbsp;&nbsp; Período: <b>{formatar_data_br(data_inicio)} até {formatar_data_br(data_fim)}</b> &nbsp;&nbsp; | &nbsp;&nbsp; Base: <b>{base_nome_filtro}</b>", styles['SubtituloSGR']))
        elementos.append(Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} por {session.get('usuario_nome', 'Usuário')}", styles['SubtituloSGR']))
        elementos.append(Spacer(1, 10))

        cards = [
            ("Escalados", resumo.get('total_escalados')),
            ("Confirmados rota", resumo.get('confirmados_rota')),
            ("Convocados base", resumo.get('cancelados_comparecer_base')),
            ("Chegadas", resumo.get('chegadas_confirmadas')),
            ("Faltas/NC", resumo.get('faltas', 0) + resumo.get('nao_compareceu', 0)),
            ("Rota extra", resumo.get('atribuidos_rota_extra')),
            ("Dispensados", resumo.get('dispensados')),
            ("Bloqueios", resumo.get('checkins_bloqueados')),
        ]
        card_row = []
        for label, valor in cards:
            card_row.append([p(label, 'CardLabelSGR'), p(valor, 'CardValorSGR')])
        tabela_cards = Table([card_row], colWidths=[3.1 * cm] * len(card_row))
        tabela_cards.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#F8FAFC')),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 7),
        ]))
        elementos.append(tabela_cards)
        elementos.append(Spacer(1, 8))

        taxas = [
            [p('Taxa de comparecimento'), p(pct(resumo.get('taxa_comparecimento_base')), 'CardValorSGR'), p('Chegadas confirmadas sobre convocados para base')],
            [p('Taxa de aproveitamento da fila'), p(pct(resumo.get('taxa_aproveitamento_fila')), 'CardValorSGR'), p('Motoristas puxados para rota extra sobre total na fila')],
            [p('Taxa de bloqueio no check-in'), p(pct(resumo.get('taxa_bloqueio_checkin')), 'CardValorSGR'), p('Tentativas bloqueadas sobre total de tentativas')],
            [p('Taxa de falta'), p(pct(resumo.get('taxa_falta')), 'CardValorSGR'), p('Faltas e não comparecimentos sobre escalados')],
        ]
        tabela_taxas = Table(taxas, colWidths=[6 * cm, 3 * cm, 15.5 * cm])
        tabela_taxas.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.white),
            ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
            ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#E5E7EB')),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 7),
            ('RIGHTPADDING', (0, 0), (-1, -1), 7),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        elementos.append(tabela_taxas)

        elementos.append(Paragraph("Alertas operacionais", styles['SecaoSGR']))
        for alerta in alertas:
            elementos.append(Paragraph(f"• {alerta}", styles['SubtituloSGR']))
        elementos.append(Spacer(1, 4))

        def tabela_padrao(titulo, headers, rows, col_widths=None, vazio="Nenhum registro encontrado."):
            elementos.append(Paragraph(titulo, styles['SecaoSGR']))
            if not rows:
                elementos.append(Paragraph(vazio, styles['SubtituloSGR']))
                return
            data = [[p(h, 'HeaderTabelaSGR') for h in headers]] + rows
            tbl = Table(data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#111827')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
                ('INNERGRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#E5E7EB')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F9FAFB')]),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                ('TOPPADDING', (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]))
            elementos.append(tbl)

        rows_base = [[p(r.get('base')), p(r.get('total_escalados')), p(r.get('convocados_base')), p(r.get('chegadas_confirmadas')), p(r.get('faltas')), p(pct(r.get('taxa_comparecimento')))] for r in dados_por_base]
        tabela_padrao("Resultado por base", ["Base", "Escalados", "Convocados", "Chegadas", "Faltas", "Comparecimento"], rows_base, [8 * cm, 3 * cm, 3 * cm, 3 * cm, 2.5 * cm, 4 * cm])

        rows_bloq = [[p(r.get('resultado')), p(r.get('motivo')), p(r.get('total'))] for r in bloqueios_motivos]
        tabela_padrao("Bloqueios por motivo", ["Resultado", "Motivo", "Total"], rows_bloq, [7 * cm, 14 * cm, 3 * cm])

        elementos.append(PageBreak())
        elementos.append(Paragraph("Rankings e últimas tentativas", styles['TituloSGR']))
        elementos.append(Paragraph(f"Período: {formatar_data_br(data_inicio)} até {formatar_data_br(data_fim)} | Base: {base_nome_filtro}", styles['SubtituloSGR']))

        ranking_rows = []
        for r in ranking_aproveitados:
            ranking_rows.append([p('Mais aproveitados em rota extra'), p(r.get('motorista')), p(r.get('total_rota_extra'))])
        for r in ranking_faltas:
            ranking_rows.append([p('Mais faltas'), p(r.get('motorista')), p(r.get('total_faltas'))])
        for r in ranking_bloqueios:
            ranking_rows.append([p('Mais bloqueios no check-in'), p(r.get('motorista')), p(r.get('total_bloqueios'))])
        tabela_padrao("Ranking de motoristas", ["Tipo", "Motorista", "Total"], ranking_rows, [8 * cm, 12 * cm, 3 * cm])

        rows_check = [[p(r.get('data_tentativa')), p(r.get('motorista')), p(r.get('base')), p(r.get('resultado')), p(r.get('motivo_bloqueio')), p(r.get('distancia_base_metros'))] for r in ultimos_checkins]
        tabela_padrao("Últimas tentativas de check-in", ["Data", "Motorista", "Base", "Resultado", "Motivo", "Distância"], rows_check, [4 * cm, 6 * cm, 5 * cm, 5 * cm, 6 * cm, 2.5 * cm])

        def rodape(canvas, doc_obj):
            canvas.saveState()
            canvas.setFont('Helvetica', 7)
            canvas.setFillColor(colors.HexColor('#6B7280'))
            canvas.drawString(1.2 * cm, 0.7 * cm, 'SGR Web - Relatório Operacional Gerencial')
            canvas.drawRightString(landscape(A4)[0] - 1.2 * cm, 0.7 * cm, f'Página {doc_obj.page}')
            canvas.restoreState()

        doc.build(elementos, onFirstPage=rodape, onLaterPages=rodape)
        buffer.seek(0)

        nome_arquivo = f"relatorio_operacional_{data_inicio.strftime('%Y%m%d')}_{data_fim.strftime('%Y%m%d')}.pdf"
        return Response(
            buffer.getvalue(),
            mimetype="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={nome_arquivo}"}
        )

    except Exception as e:
        print(f"Erro ao exportar relatório operacional em PDF: {e}")
        flash(f"Erro técnico ao exportar PDF: {e}", "danger")
        return redirect(url_for('relatorio_operacional_escala_base'))

    finally:
        cur.close()
        con.close()


@app.route('/operacao/auditoria-supervisor')
@login_required
@perfis_permitidos('Administrador', 'Operacional')
def visualizar_auditoria_supervisor():
    is_super = usuario_eh_super_admin_global()
    empresa_logada_id = session.get('empresa_id')

    pesquisa = request.args.get('pesquisa', '').strip()
    tipo_filtro = request.args.get('tipo_acao', '').strip()

    try:
        page = int(request.args.get('page', 1))
    except ValueError:
        page = 1
    per_page = 20
    offset = (page - 1) * per_page

    conexao = obter_conexao()
    cursor = conexao.cursor(dictionary=True)

    try:
        query_base = """
            FROM auditoria_supervisor a
            LEFT JOIN usuarios u ON a.usuario_id = u.id
            LEFT JOIN pessoas p ON u.pessoa_id = p.id
            WHERE 1=1
        """
        params = []

        if not is_super:
            query_base += " AND a.empresa_id = %s"
            params.append(empresa_logada_id)

        if tipo_filtro:
            query_base += " AND a.tipo_acao = %s"
            params.append(tipo_filtro)

        if pesquisa:
            query_base += " AND (a.descricao LIKE %s OR u.login LIKE %s OR p.nome_completo LIKE %s)"
            termo = f"%{pesquisa}%"
            params.extend([termo, termo, termo])

        cursor.execute(f"SELECT COUNT(a.id) AS total {query_base}", params)
        total_records = cursor.fetchone()['total']
        total_pages = (total_records + per_page - 1) // per_page

        query_select = f"""
            SELECT a.*, u.login, p.nome_completo AS nome_supervisor
            {query_base}
            ORDER BY a.id DESC
            LIMIT %s OFFSET %s
        """
        cursor.execute(query_select, params + [per_page, offset])
        logs = cursor.fetchall()

        filtros = {'pesquisa': pesquisa, 'tipo_acao': tipo_filtro}

        return render_template('auditoria_supervisor.html',
                               logs=logs,
                               filtros=filtros,
                               page=page,
                               total_pages=total_pages,
                               total_records=total_records)
    except Exception as e:
        print(f"Erro ao carregar auditoria do supervisor: {e}")
        flash("Erro ao carregar histórico de auditoria.", "danger")
        return redirect(url_for('dashboard'))
    finally:
        if conexao and conexao.is_connected():
            cursor.close()
            conexao.close()


# ==========================================================
# SCORE DO MOTORISTA - INTELIGÊNCIA OPERACIONAL
# ==========================================================
@app.route('/operacao/score-motoristas', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor')
def score_motoristas():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    hoje = date.today()
    primeiro_dia_mes = hoje.replace(day=1)

    data_inicio_str = request.args.get('data_inicio', primeiro_dia_mes.strftime('%Y-%m-%d')).strip()
    data_fim_str = request.args.get('data_fim', hoje.strftime('%Y-%m-%d')).strip()
    base_id_filtro = request.args.get('base_operacional_id', '').strip()
    empresa_id_filtro = request.args.get('empresa_id', '').strip()
    pesquisa = request.args.get('pesquisa', '').strip()
    classificacao_filtro = request.args.get('classificacao', '').strip()

    if is_super_admin and empresa_id_filtro and empresa_id_filtro.isdigit():
        empresa_id = int(empresa_id_filtro)
    else:
        empresa_id = empresa_logada_id

    try:
        data_inicio = datetime.strptime(data_inicio_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_inicio = primeiro_dia_mes
        data_inicio_str = primeiro_dia_mes.strftime('%Y-%m-%d')

    try:
        data_fim = datetime.strptime(data_fim_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_fim = hoje
        data_fim_str = hoje.strftime('%Y-%m-%d')

    if data_fim < data_inicio:
        data_inicio, data_fim = data_fim, data_inicio
        data_inicio_str = data_inicio.strftime('%Y-%m-%d')
        data_fim_str = data_fim.strftime('%Y-%m-%d')

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão ao carregar score dos motoristas.", "danger")
        return redirect(url_for('dashboard'))

    cur = con.cursor(dictionary=True)

    def inteiro(valor):
        try:
            return int(valor or 0)
        except Exception:
            return 0

    def limitar(valor, minimo=0, maximo=100):
        return max(minimo, min(maximo, int(round(valor))))

    def classificar(score):
        if score >= 90:
            return {
                'nome': 'Excelente',
                'classe': 'success',
                'icone': 'fa-circle-check',
                'descricao': 'Motorista com histórico muito positivo no período.'
            }
        if score >= 80:
            return {
                'nome': 'Confiável',
                'classe': 'primary',
                'icone': 'fa-thumbs-up',
                'descricao': 'Motorista indicado para escala e rota extra.'
            }
        if score >= 65:
            return {
                'nome': 'Atenção',
                'classe': 'warning',
                'icone': 'fa-triangle-exclamation',
                'descricao': 'Motorista exige acompanhamento operacional.'
            }
        return {
            'nome': 'Risco',
            'classe': 'danger',
            'icone': 'fa-circle-exclamation',
            'descricao': 'Motorista com histórico crítico no período.'
        }

    try:
        bases_operacionais = carregar_bases_operacionais(empresa_id, apenas_ativas=False)

        filtro_base_em = ''
        filtro_base_f = ''
        filtro_base_aud = ''
        params_base_em = []
        params_base_f = []
        params_base_aud = []

        if base_id_filtro and base_id_filtro.isdigit():
            base_id_int = int(base_id_filtro)
            filtro_base_em = ' AND em.base_operacional_id = %s '
            filtro_base_f = ' AND f.base_operacional_id = %s '
            filtro_base_aud = ' AND aud.base_operacional_id = %s '
            params_base_em.append(base_id_int)
            params_base_f.append(base_id_int)
            params_base_aud.append(base_id_int)

        # Base de motoristas: cadastro + motoristas que tiveram movimento no período.
        params_motoristas = [empresa_id, empresa_id, data_inicio, data_fim] + params_base_em + [empresa_id, data_inicio, data_fim] + params_base_f + [empresa_id, data_inicio, data_fim] + params_base_aud
        pesquisa_sql = ''
        if pesquisa:
            pesquisa_sql = ' AND (m.nome_completo LIKE %s OR m.apelido LIKE %s OR m.cpf_cnpj LIKE %s OR m.telefone LIKE %s) '
            termo = f"%{pesquisa}%"
            params_motoristas.extend([termo, termo, termo, termo])

        cur.execute(f"""
            SELECT DISTINCT m.id,
                   m.nome_completo,
                   m.apelido,
                   m.telefone,
                   m.cpf_cnpj,
                   m.status_cadastro
            FROM pessoas m
            WHERE m.empresa_id = %s
              AND (
                    m.tipo_cadastro = 'Motorista'
                    OR m.id IN (
                        SELECT em.motorista_id
                        FROM escala_motorista em
                        WHERE em.empresa_id = %s
                          AND em.data_escala BETWEEN %s AND %s
                          {filtro_base_em}
                    )
                    OR m.id IN (
                        SELECT f.motorista_id
                        FROM fila_cancelados_base f
                        WHERE f.empresa_id = %s
                          AND f.data_fila BETWEEN %s AND %s
                          {filtro_base_f}
                    )
                    OR m.id IN (
                        SELECT aud.motorista_id
                        FROM auditoria_checkin_base aud
                        WHERE aud.empresa_id = %s
                          AND DATE(aud.data_tentativa) BETWEEN %s AND %s
                          {filtro_base_aud}
                    )
                  )
              {pesquisa_sql}
            ORDER BY m.nome_completo ASC
        """, params_motoristas)
        motoristas = cur.fetchall()

        mapa = {}
        for mot in motoristas:
            mapa[mot['id']] = {
                'motorista_id': mot['id'],
                'nome_completo': mot.get('nome_completo') or 'Motorista sem nome',
                'apelido': mot.get('apelido') or '',
                'telefone': mot.get('telefone') or '',
                'status_cadastro': mot.get('status_cadastro') or '',
                'total_escalas': 0,
                'confirmados_rota': 0,
                'cancelados_base': 0,
                'chegadas_confirmadas': 0,
                'faltas_escala': 0,
                'nao_compareceu': 0,
                'faltas_automaticas': 0,
                'total_fila': 0,
                'rota_extra': 0,
                'dispensado': 0,
                'fila_aguardando': 0,
                'fila_falta': 0,
                'checkins_total': 0,
                'checkins_aprovados': 0,
                'checkins_bloqueados': 0,
                'bloqueios_distancia': 0,
                'bloqueios_qr': 0,
                'bloqueios_selfie': 0,
                'bloqueios_geo': 0,
                'bloqueios_base_sem_coord': 0,
                'ciencias': 0,
                'disponivel': 0,
                'ausente': 0,
                'sem_resposta': 0,
                'atividade_total': 0,
                'pontos_positivos': 0,
                'pontos_negativos': 0,
                'motivos': []
            }

        cur.execute(f"""
            SELECT em.motorista_id,
                   COUNT(*) AS total_escalas,
                   SUM(CASE WHEN em.status_escala = 'Confirmado com rota' THEN 1 ELSE 0 END) AS confirmados_rota,
                   SUM(CASE WHEN em.status_escala = 'Cancelado, comparecer na base' THEN 1 ELSE 0 END) AS cancelados_base,
                   SUM(CASE WHEN em.status_presenca = 'Chegada confirmada' THEN 1 ELSE 0 END) AS chegadas_confirmadas,
                   SUM(CASE WHEN em.status_escala = 'Falta' THEN 1 ELSE 0 END) AS faltas_escala,
                   SUM(CASE WHEN em.status_presenca = 'Não compareceu' THEN 1 ELSE 0 END) AS nao_compareceu,
                   SUM(CASE WHEN em.falta_automatica = 'S' THEN 1 ELSE 0 END) AS faltas_automaticas
            FROM escala_motorista em
            WHERE em.empresa_id = %s
              AND em.data_escala BETWEEN %s AND %s
              {filtro_base_em}
            GROUP BY em.motorista_id
        """, [empresa_id, data_inicio, data_fim] + params_base_em)
        for row in cur.fetchall():
            if row['motorista_id'] in mapa:
                alvo = mapa[row['motorista_id']]
                for k in ['total_escalas', 'confirmados_rota', 'cancelados_base', 'chegadas_confirmadas', 'faltas_escala', 'nao_compareceu', 'faltas_automaticas']:
                    alvo[k] = inteiro(row.get(k))

        cur.execute(f"""
            SELECT f.motorista_id,
                   COUNT(*) AS total_fila,
                   SUM(CASE WHEN f.status_fila = 'Atribuído para rota extra' THEN 1 ELSE 0 END) AS rota_extra,
                   SUM(CASE WHEN f.status_fila = 'Dispensado' THEN 1 ELSE 0 END) AS dispensado,
                   SUM(CASE WHEN f.status_fila = 'Aguardando rota' THEN 1 ELSE 0 END) AS fila_aguardando,
                   SUM(CASE WHEN f.status_fila = 'Falta' THEN 1 ELSE 0 END) AS fila_falta
            FROM fila_cancelados_base f
            WHERE f.empresa_id = %s
              AND f.data_fila BETWEEN %s AND %s
              {filtro_base_f}
            GROUP BY f.motorista_id
        """, [empresa_id, data_inicio, data_fim] + params_base_f)
        for row in cur.fetchall():
            if row['motorista_id'] in mapa:
                alvo = mapa[row['motorista_id']]
                for k in ['total_fila', 'rota_extra', 'dispensado', 'fila_aguardando', 'fila_falta']:
                    alvo[k] = inteiro(row.get(k))

        cur.execute(f"""
            SELECT aud.motorista_id,
                   COUNT(*) AS checkins_total,
                   SUM(CASE WHEN aud.resultado = 'Aprovado' THEN 1 ELSE 0 END) AS checkins_aprovados,
                   SUM(CASE WHEN aud.resultado LIKE 'Bloqueado%%' THEN 1 ELSE 0 END) AS checkins_bloqueados,
                   SUM(CASE WHEN aud.resultado = 'Bloqueado por distância' THEN 1 ELSE 0 END) AS bloqueios_distancia,
                   SUM(CASE WHEN aud.resultado = 'Bloqueado por QR inválido/expirado' THEN 1 ELSE 0 END) AS bloqueios_qr,
                   SUM(CASE WHEN aud.resultado = 'Bloqueado por falta de selfie' THEN 1 ELSE 0 END) AS bloqueios_selfie,
                   SUM(CASE WHEN aud.resultado = 'Bloqueado por falta de geolocalização' THEN 1 ELSE 0 END) AS bloqueios_geo,
                   SUM(CASE WHEN aud.resultado = 'Bloqueado por base sem coordenadas' THEN 1 ELSE 0 END) AS bloqueios_base_sem_coord
            FROM auditoria_checkin_base aud
            WHERE aud.empresa_id = %s
              AND DATE(aud.data_tentativa) BETWEEN %s AND %s
              {filtro_base_aud}
            GROUP BY aud.motorista_id
        """, [empresa_id, data_inicio, data_fim] + params_base_aud)
        for row in cur.fetchall():
            if row['motorista_id'] in mapa:
                alvo = mapa[row['motorista_id']]
                for k in ['checkins_total', 'checkins_aprovados', 'checkins_bloqueados', 'bloqueios_distancia', 'bloqueios_qr', 'bloqueios_selfie', 'bloqueios_geo', 'bloqueios_base_sem_coord']:
                    alvo[k] = inteiro(row.get(k))

        cur.execute(f"""
            SELECT c.motorista_id,
                   COUNT(*) AS ciencias
            FROM ciencia_escala_motorista c
            INNER JOIN escala_motorista em
                    ON em.id = c.escala_id
                   AND em.empresa_id = c.empresa_id
            WHERE c.empresa_id = %s
              AND em.data_escala BETWEEN %s AND %s
              {filtro_base_em}
            GROUP BY c.motorista_id
        """, [empresa_id, data_inicio, data_fim] + params_base_em)
        for row in cur.fetchall():
            if row['motorista_id'] in mapa:
                mapa[row['motorista_id']]['ciencias'] = inteiro(row.get('ciencias'))

        cur.execute("""
            SELECT dm.motorista_id,
                   SUM(CASE WHEN dm.status_disponibilidade = 'Disponível' THEN 1 ELSE 0 END) AS disponivel,
                   SUM(CASE WHEN dm.status_disponibilidade = 'Ausente' THEN 1 ELSE 0 END) AS ausente,
                   SUM(CASE WHEN dm.status_disponibilidade = 'Sem resposta' THEN 1 ELSE 0 END) AS sem_resposta
            FROM disponibilidade_motorista dm
            WHERE dm.empresa_id = %s
              AND dm.data_disponibilidade BETWEEN %s AND %s
            GROUP BY dm.motorista_id
        """, [empresa_id, data_inicio, data_fim])
        for row in cur.fetchall():
            if row['motorista_id'] in mapa:
                alvo = mapa[row['motorista_id']]
                alvo['disponivel'] = inteiro(row.get('disponivel'))
                alvo['ausente'] = inteiro(row.get('ausente'))
                alvo['sem_resposta'] = inteiro(row.get('sem_resposta'))

        scores = []
        for item in mapa.values():
            positivos = 0
            negativos = 0
            motivos = []

            positivos += min(item['checkins_aprovados'] * 3, 12)
            positivos += min(item['rota_extra'] * 8, 16)
            positivos += min(item['confirmados_rota'] * 2, 10)
            positivos += min(item['ciencias'] * 1, 5)
            positivos += min(item['disponivel'] * 1, 6)

            negativos += item['faltas_escala'] * 20
            negativos += item['nao_compareceu'] * 20
            negativos += item['fila_falta'] * 15
            negativos += item['bloqueios_distancia'] * 15
            negativos += item['bloqueios_qr'] * 10
            outros_bloqueios = item['bloqueios_selfie'] + item['bloqueios_geo'] + item['bloqueios_base_sem_coord']
            negativos += outros_bloqueios * 8
            negativos += item['ausente'] * 1
            negativos += item['sem_resposta'] * 2

            if item['rota_extra']:
                motivos.append(f"{item['rota_extra']} rota(s) extra atribuída(s)")
            if item['checkins_aprovados']:
                motivos.append(f"{item['checkins_aprovados']} check-in(s) aprovado(s)")
            if item['faltas_escala'] or item['nao_compareceu']:
                motivos.append(f"{item['faltas_escala'] + item['nao_compareceu']} falta(s)/não comparecimento(s)")
            if item['checkins_bloqueados']:
                motivos.append(f"{item['checkins_bloqueados']} bloqueio(s) no check-in")
            if not motivos:
                motivos.append('Sem movimentação crítica no período')

            atividade_total = (
                item['total_escalas'] + item['total_fila'] + item['checkins_total'] +
                item['ciencias'] + item['disponivel'] + item['ausente'] + item['sem_resposta']
            )
            score = limitar(80 + positivos - negativos)
            classificacao = classificar(score)

            item['pontos_positivos'] = positivos
            item['pontos_negativos'] = negativos
            item['atividade_total'] = atividade_total
            item['score'] = score
            item['classificacao'] = classificacao
            item['motivos'] = motivos[:4]
            item['taxa_checkin_aprovado'] = round((item['checkins_aprovados'] / item['checkins_total']) * 100, 1) if item['checkins_total'] else 0
            item['taxa_presenca'] = round((item['chegadas_confirmadas'] / item['cancelados_base']) * 100, 1) if item['cancelados_base'] else 0
            scores.append(item)

        if classificacao_filtro:
            scores = [s for s in scores if s['classificacao']['nome'] == classificacao_filtro]

        scores.sort(key=lambda x: (x['score'], x['atividade_total']), reverse=True)

        total_avaliados = len(scores)
        score_medio = round(sum(s['score'] for s in scores) / total_avaliados, 1) if total_avaliados else 0
        total_excelente = sum(1 for s in scores if s['classificacao']['nome'] == 'Excelente')
        total_confiavel = sum(1 for s in scores if s['classificacao']['nome'] == 'Confiável')
        total_atencao = sum(1 for s in scores if s['classificacao']['nome'] == 'Atenção')
        total_risco = sum(1 for s in scores if s['classificacao']['nome'] == 'Risco')
        total_bloqueios = sum(s['checkins_bloqueados'] for s in scores)
        total_faltas = sum((s['faltas_escala'] + s['nao_compareceu'] + s['fila_falta']) for s in scores)
        total_rota_extra = sum(s['rota_extra'] for s in scores)

        resumo_score = {
            'total_avaliados': total_avaliados,
            'score_medio': score_medio,
            'total_excelente': total_excelente,
            'total_confiavel': total_confiavel,
            'total_atencao': total_atencao,
            'total_risco': total_risco,
            'total_bloqueios': total_bloqueios,
            'total_faltas': total_faltas,
            'total_rota_extra': total_rota_extra,
            'percentual_confianca': round(((total_excelente + total_confiavel) / total_avaliados) * 100, 1) if total_avaliados else 0
        }

        ranking_top = scores[:5]
        ranking_risco = sorted(scores, key=lambda x: (x['score'], -x['atividade_total']))[:5]
        motoristas_rota_extra = sorted(scores, key=lambda x: x['rota_extra'], reverse=True)[:5]

        periodo_label = f"{formatar_data_br(data_inicio)} até {formatar_data_br(data_fim)}"
        base_nome_filtro = 'Todas as bases'
        if base_id_filtro and base_id_filtro.isdigit():
            for base in bases_operacionais:
                if str(base.get('id')) == str(base_id_filtro):
                    base_nome_filtro = base.get('nome_base') or 'Base selecionada'
                    break

        alertas_score = []
        if total_risco:
            alertas_score.append({
                'tipo': 'danger',
                'icone': 'fa-circle-exclamation',
                'titulo': 'Motoristas em risco operacional',
                'descricao': f"{total_risco} motorista(s) ficaram abaixo de 65 pontos no período."
            })
        if total_bloqueios:
            alertas_score.append({
                'tipo': 'warning',
                'icone': 'fa-shield-halved',
                'titulo': 'Bloqueios impactando score',
                'descricao': f"Foram identificados {total_bloqueios} bloqueio(s) de check-in no período."
            })
        if total_faltas:
            alertas_score.append({
                'tipo': 'danger',
                'icone': 'fa-user-xmark',
                'titulo': 'Faltas e não comparecimentos',
                'descricao': f"{total_faltas} ocorrência(s) reduziram a pontuação dos motoristas."
            })
        if not alertas_score:
            alertas_score.append({
                'tipo': 'success',
                'icone': 'fa-circle-check',
                'titulo': 'Nenhum alerta crítico',
                'descricao': 'O histórico avaliado não apresentou desvios relevantes.'
            })

        return render_template(
            'score_motoristas.html',
            usuario_logado=usuario_logado,
            scores=scores,
            resumo_score=resumo_score,
            ranking_top=ranking_top,
            ranking_risco=ranking_risco,
            motoristas_rota_extra=motoristas_rota_extra,
            alertas_score=alertas_score,
            bases_operacionais=bases_operacionais,
            periodo_label=periodo_label,
            base_nome_filtro=base_nome_filtro,
            data_inicio=data_inicio_str,
            data_fim=data_fim_str,
            base_id_filtro=base_id_filtro,
            pesquisa=pesquisa,
            classificacao_filtro=classificacao_filtro,
            is_super_admin=is_super_admin,
            empresa_id_filtro=empresa_id if is_super_admin else ''
        )

    except Exception as e:
        print(f"Erro ao carregar score dos motoristas: {e}")
        flash(f"Erro técnico ao carregar score dos motoristas: {e}", "danger")
        return redirect(url_for('dashboard'))

    finally:
        cur.close()
        con.close()


# ==========================================================
# CENTRAL DE PENDÊNCIAS DA OPERAÇÃO
# ==========================================================
@app.route('/operacao/central-pendencias', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor')
def central_pendencias_operacao():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    hoje = date.today()
    data_operacao_str = request.args.get('data_operacao', hoje.strftime('%Y-%m-%d')).strip()
    base_id_filtro = request.args.get('base_operacional_id', '').strip()
    empresa_id_filtro = request.args.get('empresa_id', '').strip()

    if is_super_admin and empresa_id_filtro and empresa_id_filtro.isdigit():
        empresa_id = int(empresa_id_filtro)
    else:
        empresa_id = empresa_logada_id

    try:
        data_operacao = datetime.strptime(data_operacao_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_operacao = hoje
        data_operacao_str = hoje.strftime('%Y-%m-%d')

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão ao carregar central de pendências.", "danger")
        return redirect(url_for('dashboard'))

    cur = con.cursor(dictionary=True)

    def inteiro(valor):
        try:
            return int(valor or 0)
        except Exception:
            return 0

    try:
        bases_operacionais = carregar_bases_operacionais(empresa_id, apenas_ativas=False)
        empresas = carregar_empresas_ativas() if is_super_admin else []

        filtro_base_em = ''
        filtro_base_f = ''
        filtro_base_aud = ''
        params_base_em = []
        params_base_f = []
        params_base_aud = []

        if base_id_filtro and base_id_filtro.isdigit():
            base_id_int = int(base_id_filtro)
            filtro_base_em = ' AND em.base_operacional_id = %s '
            filtro_base_f = ' AND f.base_operacional_id = %s '
            filtro_base_aud = ' AND aud.base_operacional_id = %s '
            params_base_em.append(base_id_int)
            params_base_f.append(base_id_int)
            params_base_aud.append(base_id_int)

        # Motoristas sem resposta na disponibilidade do dia.
        cur.execute("""
            SELECT dm.id,
                   dm.motorista_id,
                   p.nome_completo,
                   p.apelido,
                   p.telefone,
                   dm.status_disponibilidade,
                   dm.data_disponibilidade
            FROM disponibilidade_motorista dm
            LEFT JOIN pessoas p ON p.id = dm.motorista_id
            WHERE dm.empresa_id = %s
              AND dm.data_disponibilidade = %s
              AND dm.status_disponibilidade = 'Sem resposta'
            ORDER BY p.nome_completo ASC
            LIMIT 200
        """, [empresa_id, data_operacao])
        sem_resposta = cur.fetchall()

        # Escalas que ainda precisam de ciência do motorista.
        cur.execute(f"""
            SELECT em.id AS escala_id,
                   em.motorista_id,
                   p.nome_completo,
                   p.apelido,
                   p.telefone,
                   em.status_escala,
                   em.status_presenca,
                   em.horario_apresentacao,
                   em.observacao_supervisor,
                   COALESCE(b.nome_base, em.base_operacao) AS nome_base
            FROM escala_motorista em
            LEFT JOIN pessoas p ON p.id = em.motorista_id
            LEFT JOIN bases_operacionais b ON b.id = em.base_operacional_id
            WHERE em.empresa_id = %s
              AND em.data_escala = %s
              AND em.status_escala IN ('Confirmado com rota', 'Cancelado, comparecer na base', 'Reserva / Avulso')
              AND NOT EXISTS (
                    SELECT 1
                    FROM ciencia_escala_motorista c
                    WHERE c.empresa_id = em.empresa_id
                      AND c.escala_id = em.id
                      AND c.motorista_id = em.motorista_id
              )
              {filtro_base_em}
            ORDER BY em.horario_apresentacao IS NULL ASC, em.horario_apresentacao ASC, p.nome_completo ASC
            LIMIT 200
        """, [empresa_id, data_operacao] + params_base_em)
        sem_ciencia = cur.fetchall()

        # Convocados para base aguardando chegada.
        cur.execute(f"""
            SELECT em.id AS escala_id,
                   em.motorista_id,
                   p.nome_completo,
                   p.apelido,
                   p.telefone,
                   em.status_escala,
                   em.status_presenca,
                   em.horario_apresentacao,
                   em.observacao_supervisor,
                   COALESCE(b.nome_base, em.base_operacao) AS nome_base
            FROM escala_motorista em
            LEFT JOIN pessoas p ON p.id = em.motorista_id
            LEFT JOIN bases_operacionais b ON b.id = em.base_operacional_id
            WHERE em.empresa_id = %s
              AND em.data_escala = %s
              AND em.status_escala = 'Cancelado, comparecer na base'
              AND em.status_presenca = 'Aguardando chegada'
              {filtro_base_em}
            ORDER BY em.horario_apresentacao IS NULL ASC, em.horario_apresentacao ASC, p.nome_completo ASC
            LIMIT 200
        """, [empresa_id, data_operacao] + params_base_em)
        aguardando_chegada = cur.fetchall()

        # Fila aguardando decisão do supervisor.
        cur.execute(f"""
            SELECT f.id AS fila_id,
                   f.escala_id,
                   f.motorista_id,
                   p.nome_completo,
                   p.apelido,
                   p.telefone,
                   COALESCE(b.nome_base, f.base_operacao) AS nome_base,
                   f.hora_confirmacao,
                   f.posicao_fila,
                   f.status_fila,
                   f.distancia_base_metros,
                   f.geolocalizacao_validada,
                   f.qr_code_validado,
                   f.selfie_path
            FROM fila_cancelados_base f
            LEFT JOIN pessoas p ON p.id = f.motorista_id
            LEFT JOIN bases_operacionais b ON b.id = f.base_operacional_id
            WHERE f.empresa_id = %s
              AND f.data_fila = %s
              AND f.status_fila = 'Aguardando rota'
              {filtro_base_f}
            ORDER BY f.posicao_fila ASC, f.hora_confirmacao ASC
            LIMIT 200
        """, [empresa_id, data_operacao] + params_base_f)
        fila_aguardando = cur.fetchall()

        # Check-ins bloqueados no dia.
        cur.execute(f"""
            SELECT aud.id,
                   aud.motorista_id,
                   p.nome_completo,
                   p.apelido,
                   p.telefone,
                   COALESCE(b.nome_base, 'Base não identificada') AS nome_base,
                   aud.data_tentativa,
                   aud.resultado,
                   aud.motivo_bloqueio,
                   aud.distancia_base_metros,
                   aud.codigo_qr_informado
            FROM auditoria_checkin_base aud
            LEFT JOIN pessoas p ON p.id = aud.motorista_id
            LEFT JOIN bases_operacionais b ON b.id = aud.base_operacional_id
            WHERE aud.empresa_id = %s
              AND DATE(aud.data_tentativa) = %s
              AND aud.resultado LIKE 'Bloqueado%%'
              {filtro_base_aud}
            ORDER BY aud.data_tentativa DESC
            LIMIT 200
        """, [empresa_id, data_operacao] + params_base_aud)
        checkins_bloqueados = cur.fetchall()

        # Faltas e não comparecimentos.
        cur.execute(f"""
            SELECT em.id AS escala_id,
                   em.motorista_id,
                   p.nome_completo,
                   p.apelido,
                   p.telefone,
                   COALESCE(b.nome_base, em.base_operacao) AS nome_base,
                   em.status_escala,
                   em.status_presenca,
                   em.falta_automatica,
                   em.falta_marcada_em,
                   em.falta_motivo,
                   em.horario_apresentacao
            FROM escala_motorista em
            LEFT JOIN pessoas p ON p.id = em.motorista_id
            LEFT JOIN bases_operacionais b ON b.id = em.base_operacional_id
            WHERE em.empresa_id = %s
              AND em.data_escala = %s
              AND (
                    em.status_escala = 'Falta'
                    OR em.status_presenca = 'Não compareceu'
                    OR em.falta_automatica = 'S'
              )
              {filtro_base_em}
            ORDER BY em.falta_marcada_em DESC, p.nome_completo ASC
            LIMIT 200
        """, [empresa_id, data_operacao] + params_base_em)
        faltas = cur.fetchall()

        # Justificativas de ausência pendentes de análise.
        cur.execute(f"""
            SELECT jam.id AS justificativa_id,
                   jam.escala_id,
                   jam.motorista_id,
                   p.nome_completo,
                   p.apelido,
                   p.telefone,
                   COALESCE(b.nome_base, em.base_operacao) AS nome_base,
                   jam.data_escala,
                   jam.horario_previsto,
                   jam.motivo,
                   jam.observacao_motorista,
                   jam.anexo_path,
                   jam.status_justificativa,
                   jam.data_envio,
                   em.status_escala,
                   em.status_presenca,
                   em.falta_automatica,
                   em.falta_motivo
            FROM justificativas_ausencia_motorista jam
            LEFT JOIN escala_motorista em ON em.id = jam.escala_id AND em.empresa_id = jam.empresa_id
            LEFT JOIN pessoas p ON p.id = jam.motorista_id
            LEFT JOIN bases_operacionais b ON b.id = jam.base_operacional_id
            WHERE jam.empresa_id = %s
              AND jam.data_escala = %s
              AND jam.status_justificativa = 'Pendente de análise'
              {filtro_base_em.replace('em.base_operacional_id', 'jam.base_operacional_id')}
            ORDER BY jam.data_envio ASC
            LIMIT 200
        """, [empresa_id, data_operacao] + params_base_em)
        justificativas_pendentes = cur.fetchall()

        # Bases que exigem ajuste cadastral.
        cur.execute("""
            SELECT id,
                   nome_base,
                   status_base,
                   latitude,
                   longitude,
                   raio_permitido_metros
            FROM bases_operacionais
            WHERE empresa_id = %s
              AND status_base = 'Ativa'
              AND (
                    latitude IS NULL
                    OR longitude IS NULL
                    OR raio_permitido_metros IS NULL
                    OR raio_permitido_metros < 30
              )
            ORDER BY nome_base ASC
            LIMIT 100
        """, [empresa_id])
        bases_pendentes = cur.fetchall()

        # QR ativo expirado: sinaliza terminal/base que pode precisar de atenção.
        cur.execute("""
            SELECT t.id,
                   t.base_operacional_id,
                   b.nome_base,
                   t.codigo_token,
                   t.data_geracao,
                   t.data_expiracao,
                   t.status_token
            FROM base_qr_tokens t
            LEFT JOIN bases_operacionais b ON b.id = t.base_operacional_id
            WHERE t.empresa_id = %s
              AND t.status_token = 'Ativo'
              AND t.data_expiracao < NOW()
            ORDER BY t.data_expiracao DESC
            LIMIT 100
        """, [empresa_id])
        qrs_expirados = cur.fetchall()

        kpis = {
            'sem_resposta': len(sem_resposta),
            'sem_ciencia': len(sem_ciencia),
            'aguardando_chegada': len(aguardando_chegada),
            'fila_aguardando': len(fila_aguardando),
            'checkins_bloqueados': len(checkins_bloqueados),
            'faltas': len(faltas),
            'justificativas_pendentes': len(justificativas_pendentes),
            'bases_pendentes': len(bases_pendentes),
            'qrs_expirados': len(qrs_expirados),
        }
        kpis['total_pendencias'] = sum(inteiro(v) for v in kpis.values())

        alertas = []
        if kpis['fila_aguardando']:
            alertas.append({
                'tipo': 'warning',
                'icone': 'fa-people-arrows',
                'titulo': 'Motoristas aguardando decisão',
                'descricao': f"{kpis['fila_aguardando']} motorista(s) estão na fila aguardando rota extra ou dispensa."
            })
        if kpis['aguardando_chegada']:
            alertas.append({
                'tipo': 'info',
                'icone': 'fa-location-dot',
                'titulo': 'Convocados ainda não chegaram',
                'descricao': f"{kpis['aguardando_chegada']} motorista(s) convocados para base ainda estão aguardando chegada."
            })
        if kpis['checkins_bloqueados']:
            alertas.append({
                'tipo': 'danger',
                'icone': 'fa-shield-halved',
                'titulo': 'Check-ins bloqueados',
                'descricao': f"{kpis['checkins_bloqueados']} tentativa(s) bloqueada(s) no check-in da base."
            })
        if kpis['sem_ciencia']:
            alertas.append({
                'tipo': 'warning',
                'icone': 'fa-eye-slash',
                'titulo': 'Ciência pendente',
                'descricao': f"{kpis['sem_ciencia']} escala(s) ainda sem ciência do motorista."
            })
        if kpis['justificativas_pendentes']:
            alertas.append({
                'tipo': 'primary',
                'icone': 'fa-file-circle-question',
                'titulo': 'Justificativas aguardando análise',
                'descricao': f"{kpis['justificativas_pendentes']} justificativa(s) de ausência precisam de decisão do supervisor."
            })
        if kpis['bases_pendentes']:
            alertas.append({
                'tipo': 'danger',
                'icone': 'fa-map-location-dot',
                'titulo': 'Base com cadastro incompleto',
                'descricao': f"{kpis['bases_pendentes']} base(s) precisam de coordenada ou raio válido."
            })
        if not alertas:
            alertas.append({
                'tipo': 'success',
                'icone': 'fa-circle-check',
                'titulo': 'Operação sem pendência crítica',
                'descricao': 'Nenhum ponto crítico encontrado para os filtros atuais.'
            })

        base_nome_filtro = 'Todas as bases'
        if base_id_filtro and base_id_filtro.isdigit():
            for base in bases_operacionais:
                if str(base.get('id')) == str(base_id_filtro):
                    base_nome_filtro = base.get('nome_base') or 'Base selecionada'
                    break

        return render_template(
            'central_pendencias_operacao.html',
            usuario_logado=usuario_logado,
            data_operacao=data_operacao_str,
            data_operacao_obj=data_operacao,
            bases_operacionais=bases_operacionais,
            empresas=empresas,
            is_super_admin=is_super_admin,
            empresa_id_filtro=empresa_id if is_super_admin else '',
            base_id_filtro=base_id_filtro,
            base_nome_filtro=base_nome_filtro,
            kpis=kpis,
            alertas=alertas,
            sem_resposta=sem_resposta,
            sem_ciencia=sem_ciencia,
            aguardando_chegada=aguardando_chegada,
            fila_aguardando=fila_aguardando,
            checkins_bloqueados=checkins_bloqueados,
            faltas=faltas,
            justificativas_pendentes=justificativas_pendentes,
            bases_pendentes=bases_pendentes,
            qrs_expirados=qrs_expirados
        )

    except Exception as e:
        print(f"Erro ao carregar central de pendências: {e}")
        flash(f"Erro técnico ao carregar central de pendências: {e}", "danger")
        return redirect(url_for('dashboard'))

    finally:
        cur.close()
        con.close()


@app.route('/operacao/justificativas-ausencia/<int:justificativa_id>/analisar', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor')
def analisar_justificativa_ausencia(justificativa_id):
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão.", "danger")
        return redirect(url_for('logout'))

    decisao = request.form.get('decisao', '').strip()
    observacao_supervisor = request.form.get('observacao_supervisor', '').strip()
    data_operacao = request.form.get('data_operacao', '').strip()
    base_operacional_id = request.form.get('base_operacional_id', '').strip()
    empresa_id_form = request.form.get('empresa_id', '').strip()

    # Quando o Super Admin analisa uma justificativa de outra empresa,
    # precisamos usar a empresa selecionada no filtro da Central de Pendências,
    # e não a empresa da sessão. Caso contrário a consulta não encontra o registro.
    if is_super_admin and empresa_id_form and empresa_id_form.isdigit():
        empresa_id = int(empresa_id_form)
    else:
        empresa_id = empresa_logada_id

    params_retorno = {
        'data_operacao': data_operacao,
        'base_operacional_id': base_operacional_id
    }
    if is_super_admin:
        params_retorno['empresa_id'] = empresa_id

    if decisao not in ['aceitar', 'recusar', 'em_analise']:
        flash("Decisão inválida para a justificativa.", "warning")
        return redirect(url_for('central_pendencias_operacao', **params_retorno))

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão ao analisar justificativa.", "danger")
        return redirect(url_for('central_pendencias_operacao', **params_retorno))

    cur = con.cursor(dictionary=True)

    try:
        cur.execute("""
            SELECT jam.*,
                   p.nome_completo,
                   em.status_escala,
                   em.status_presenca,
                   em.falta_automatica
            FROM justificativas_ausencia_motorista jam
            LEFT JOIN pessoas p ON p.id = jam.motorista_id
            LEFT JOIN escala_motorista em ON em.id = jam.escala_id AND em.empresa_id = jam.empresa_id
            WHERE jam.id = %s
              AND jam.empresa_id = %s
            LIMIT 1
        """, (justificativa_id, empresa_id))
        justificativa = cur.fetchone()

        if not justificativa:
            flash("Justificativa não encontrada.", "warning")
            return redirect(url_for('central_pendencias_operacao', **params_retorno))

        if decisao == 'aceitar':
            novo_status = 'Aceita'
            novo_status_presenca = 'Ausência justificada'
            falta_automatica = 'N'
            falta_motivo = 'Ausência justificada pelo supervisor.'
        elif decisao == 'recusar':
            novo_status = 'Recusada'
            novo_status_presenca = 'Não compareceu'
            falta_automatica = 'S'
            falta_motivo = 'Justificativa recusada pelo supervisor.'
        else:
            novo_status = 'Em análise'
            novo_status_presenca = justificativa.get('status_presenca') or 'Não compareceu'
            falta_automatica = justificativa.get('falta_automatica') or 'N'
            falta_motivo = 'Justificativa em análise pelo supervisor.'

        cur.execute("""
            UPDATE justificativas_ausencia_motorista
            SET status_justificativa = %s,
                observacao_supervisor = %s,
                usuario_analise_id = %s,
                data_analise = NOW(),
                updated_at = NOW()
            WHERE id = %s
              AND empresa_id = %s
        """, (
            novo_status,
            observacao_supervisor,
            session.get('usuario_id'),
            justificativa_id,
            empresa_id
        ))

        cur.execute("""
            UPDATE escala_motorista
            SET status_presenca = %s,
                falta_automatica = %s,
                falta_marcada_em = COALESCE(falta_marcada_em, NOW()),
                falta_motivo = %s,
                data_atualizacao = NOW()
            WHERE id = %s
              AND empresa_id = %s
        """, (
            novo_status_presenca,
            falta_automatica,
            falta_motivo,
            justificativa.get('escala_id'),
            empresa_id
        ))

        descricao = (
            f"Justificativa de ausência #{justificativa_id} do motorista "
            f"{justificativa.get('nome_completo') or justificativa.get('motorista_id')} analisada: {novo_status}. "
            f"Observação do supervisor: {observacao_supervisor or '-'}"
        )
        registrar_auditoria_supervisor(cur, 'ANALISAR_JUSTIFICATIVA_AUSENCIA', descricao)

        con.commit()
        flash(f"Justificativa marcada como {novo_status}.", "success")

    except Exception as e:
        con.rollback()
        print(f"Erro ao analisar justificativa de ausência: {e}")
        flash(f"Erro técnico ao analisar justificativa: {e}", "danger")

    finally:
        cur.close()
        con.close()

    return redirect(url_for('central_pendencias_operacao', **params_retorno))




# ==========================================================
# HISTÓRICO COMPLETO DO MOTORISTA - DOSSIÊ OPERACIONAL
# ==========================================================
@app.route('/operacao/historico-motoristas', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor', 'Financeiro', 'Consulta')
def historico_motoristas():
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    hoje = date.today()
    primeiro_dia_mes = hoje.replace(day=1)
    data_inicio_str = request.args.get('data_inicio', primeiro_dia_mes.strftime('%Y-%m-%d')).strip()
    data_fim_str = request.args.get('data_fim', hoje.strftime('%Y-%m-%d')).strip()
    empresa_id_filtro = request.args.get('empresa_id', '').strip()
    base_id_filtro = request.args.get('base_operacional_id', '').strip()
    pesquisa = request.args.get('pesquisa', '').strip()
    classificacao_filtro = request.args.get('classificacao', '').strip()

    if is_super_admin and empresa_id_filtro and empresa_id_filtro.isdigit():
        empresa_id = int(empresa_id_filtro)
    else:
        empresa_id = empresa_logada_id

    try:
        data_inicio = datetime.strptime(data_inicio_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_inicio = primeiro_dia_mes
        data_inicio_str = data_inicio.strftime('%Y-%m-%d')

    try:
        data_fim = datetime.strptime(data_fim_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_fim = hoje
        data_fim_str = data_fim.strftime('%Y-%m-%d')

    if data_fim < data_inicio:
        data_inicio, data_fim = data_fim, data_inicio
        data_inicio_str = data_inicio.strftime('%Y-%m-%d')
        data_fim_str = data_fim.strftime('%Y-%m-%d')

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão ao carregar histórico dos motoristas.", "danger")
        return redirect(url_for('dashboard'))

    cur = con.cursor(dictionary=True)

    def inteiro(valor):
        try:
            return int(valor or 0)
        except Exception:
            return 0

    try:
        empresas = carregar_empresas_ativas() if is_super_admin else []
        bases_operacionais = carregar_bases_operacionais(empresa_id, apenas_ativas=False)

        filtro_base_em = ''
        filtro_base_f = ''
        filtro_base_aud = ''
        filtro_base_just = ''
        params_base_em = []
        params_base_f = []
        params_base_aud = []
        params_base_just = []
        if base_id_filtro and base_id_filtro.isdigit():
            base_id = int(base_id_filtro)
            filtro_base_em = ' AND em.base_operacional_id = %s '
            filtro_base_f = ' AND f.base_operacional_id = %s '
            filtro_base_aud = ' AND aud.base_operacional_id = %s '
            filtro_base_just = ' AND j.base_operacional_id = %s '
            params_base_em.append(base_id)
            params_base_f.append(base_id)
            params_base_aud.append(base_id)
            params_base_just.append(base_id)

        pesquisa_sql = ''
        params_pesquisa = []
        if pesquisa:
            pesquisa_sql = ' AND (p.nome_completo LIKE %s OR p.apelido LIKE %s OR p.cpf_cnpj LIKE %s OR p.telefone LIKE %s) '
            termo = f"%{pesquisa}%"
            params_pesquisa = [termo, termo, termo, termo]

        params = [empresa_id, empresa_id, data_inicio, data_fim] + params_base_em + [empresa_id, data_inicio, data_fim] + params_base_f + [empresa_id, data_inicio, data_fim] + params_base_aud + params_pesquisa
        cur.execute(f"""
            SELECT DISTINCT p.id, p.nome_completo, p.apelido, p.cpf_cnpj, p.telefone, p.email, p.status_cadastro
            FROM pessoas p
            WHERE p.empresa_id = %s
              AND (
                    p.tipo_cadastro = 'Motorista'
                    OR p.id IN (SELECT em.motorista_id FROM escala_motorista em WHERE em.empresa_id=%s AND em.data_escala BETWEEN %s AND %s {filtro_base_em})
                    OR p.id IN (SELECT f.motorista_id FROM fila_cancelados_base f WHERE f.empresa_id=%s AND f.data_fila BETWEEN %s AND %s {filtro_base_f})
                    OR p.id IN (SELECT aud.motorista_id FROM auditoria_checkin_base aud WHERE aud.empresa_id=%s AND DATE(aud.data_tentativa) BETWEEN %s AND %s {filtro_base_aud})
                  )
              {pesquisa_sql}
            ORDER BY p.nome_completo ASC
        """, params)
        motoristas = cur.fetchall()
        ids = [m['id'] for m in motoristas]
        mapa_score = calcular_scores_motoristas_para_fila(empresa_id, ids, data_fim, int(base_id_filtro) if base_id_filtro.isdigit() else None) if ids else {}

        mapa = {m['id']: {
            **m,
            'score': (mapa_score.get(m['id']) or {}).get('score', 80),
            'classificacao': (mapa_score.get(m['id']) or {}).get('classificacao', classificar_score_operacional(80)),
            'total_escalas': 0, 'confirmados_rota': 0, 'cancelados_base': 0, 'chegadas_confirmadas': 0,
            'faltas': 0, 'ausencias_justificadas': 0, 'rotas_extras': 0, 'dispensas': 0,
            'checkins_aprovados': 0, 'checkins_bloqueados': 0, 'justificativas_pendentes': 0
        } for m in motoristas}

        if ids:
            ph = ','.join(['%s'] * len(ids))
            cur.execute(f"""
                SELECT em.motorista_id,
                       COUNT(*) total_escalas,
                       SUM(CASE WHEN em.status_escala='Confirmado com rota' THEN 1 ELSE 0 END) confirmados_rota,
                       SUM(CASE WHEN em.status_escala='Cancelado, comparecer na base' THEN 1 ELSE 0 END) cancelados_base,
                       SUM(CASE WHEN em.status_presenca='Chegada confirmada' THEN 1 ELSE 0 END) chegadas_confirmadas,
                       SUM(CASE WHEN em.status_escala='Falta' OR em.status_presenca='Não compareceu' THEN 1 ELSE 0 END) faltas,
                       SUM(CASE WHEN em.status_presenca='Ausência justificada' THEN 1 ELSE 0 END) ausencias_justificadas
                FROM escala_motorista em
                WHERE em.empresa_id=%s AND em.data_escala BETWEEN %s AND %s {filtro_base_em} AND em.motorista_id IN ({ph})
                GROUP BY em.motorista_id
            """, [empresa_id, data_inicio, data_fim] + params_base_em + ids)
            for r in cur.fetchall():
                if r['motorista_id'] in mapa:
                    mapa[r['motorista_id']].update({k: inteiro(r.get(k)) for k in ['total_escalas','confirmados_rota','cancelados_base','chegadas_confirmadas','faltas','ausencias_justificadas']})

            cur.execute(f"""
                SELECT f.motorista_id,
                       SUM(CASE WHEN f.status_fila='Atribuído para rota extra' THEN 1 ELSE 0 END) rotas_extras,
                       SUM(CASE WHEN f.status_fila='Dispensado' THEN 1 ELSE 0 END) dispensas
                FROM fila_cancelados_base f
                WHERE f.empresa_id=%s AND f.data_fila BETWEEN %s AND %s {filtro_base_f} AND f.motorista_id IN ({ph})
                GROUP BY f.motorista_id
            """, [empresa_id, data_inicio, data_fim] + params_base_f + ids)
            for r in cur.fetchall():
                if r['motorista_id'] in mapa:
                    mapa[r['motorista_id']]['rotas_extras'] = inteiro(r.get('rotas_extras'))
                    mapa[r['motorista_id']]['dispensas'] = inteiro(r.get('dispensas'))

            cur.execute(f"""
                SELECT aud.motorista_id,
                       SUM(CASE WHEN aud.resultado='Aprovado' THEN 1 ELSE 0 END) checkins_aprovados,
                       SUM(CASE WHEN aud.resultado <> 'Aprovado' THEN 1 ELSE 0 END) checkins_bloqueados
                FROM auditoria_checkin_base aud
                WHERE aud.empresa_id=%s AND DATE(aud.data_tentativa) BETWEEN %s AND %s {filtro_base_aud} AND aud.motorista_id IN ({ph})
                GROUP BY aud.motorista_id
            """, [empresa_id, data_inicio, data_fim] + params_base_aud + ids)
            for r in cur.fetchall():
                if r['motorista_id'] in mapa:
                    mapa[r['motorista_id']]['checkins_aprovados'] = inteiro(r.get('checkins_aprovados'))
                    mapa[r['motorista_id']]['checkins_bloqueados'] = inteiro(r.get('checkins_bloqueados'))

            try:
                cur.execute(f"""
                    SELECT j.motorista_id,
                           SUM(CASE WHEN j.status_justificativa='Pendente de análise' THEN 1 ELSE 0 END) justificativas_pendentes
                    FROM justificativas_ausencia_motorista j
                    WHERE j.empresa_id=%s AND j.data_escala BETWEEN %s AND %s {filtro_base_just} AND j.motorista_id IN ({ph})
                    GROUP BY j.motorista_id
                """, [empresa_id, data_inicio, data_fim] + params_base_just + ids)
                for r in cur.fetchall():
                    if r['motorista_id'] in mapa:
                        mapa[r['motorista_id']]['justificativas_pendentes'] = inteiro(r.get('justificativas_pendentes'))
            except Exception:
                pass

        lista = list(mapa.values())
        if classificacao_filtro:
            lista = [m for m in lista if (m.get('classificacao') or {}).get('nome') == classificacao_filtro]

        total_motoristas = len(lista)
        score_medio = round(sum(int(m.get('score') or 0) for m in lista) / total_motoristas, 1) if total_motoristas else 0
        motoristas_atencao = sum(1 for m in lista if (m.get('classificacao') or {}).get('nome') == 'Atenção')
        motoristas_risco = sum(1 for m in lista if (m.get('classificacao') or {}).get('nome') == 'Risco')
        total_pendencias = sum(int(m.get('justificativas_pendentes') or 0) for m in lista)

        filtros = {
            'empresa_id': str(empresa_id),
            'data_inicio': data_inicio_str,
            'data_fim': data_fim_str,
            'base_operacional_id': base_id_filtro,
            'pesquisa': pesquisa,
            'classificacao': classificacao_filtro
        }
        return render_template('historico_motoristas.html',
                               usuario_logado=session.get('usuario_nome', 'Usuário'),
                               motoristas=lista,
                               empresas=empresas,
                               bases_operacionais=bases_operacionais,
                               filtros=filtros,
                               is_super_admin=is_super_admin,
                               total_motoristas=total_motoristas,
                               score_medio=score_medio,
                               motoristas_atencao=motoristas_atencao,
                               motoristas_risco=motoristas_risco,
                               total_pendencias=total_pendencias)
    except Exception as e:
        print(f"Erro ao carregar histórico dos motoristas: {e}")
        flash("Erro ao carregar histórico dos motoristas.", "danger")
        return redirect(url_for('dashboard'))
    finally:
        if con and con.is_connected():
            cur.close()
            con.close()


@app.route('/operacao/motoristas/<int:motorista_id>/historico', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Supervisor', 'Financeiro', 'Consulta')
def historico_motorista_detalhe(motorista_id):
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()
    hoje = date.today()
    primeiro_dia_mes = hoje.replace(day=1)
    data_inicio_str = request.args.get('data_inicio', primeiro_dia_mes.strftime('%Y-%m-%d')).strip()
    data_fim_str = request.args.get('data_fim', hoje.strftime('%Y-%m-%d')).strip()
    empresa_id_filtro = request.args.get('empresa_id', '').strip()
    if is_super_admin and empresa_id_filtro and empresa_id_filtro.isdigit():
        empresa_id = int(empresa_id_filtro)
    else:
        empresa_id = empresa_logada_id
    try:
        data_inicio = datetime.strptime(data_inicio_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_inicio = primeiro_dia_mes
    try:
        data_fim = datetime.strptime(data_fim_str[:10], '%Y-%m-%d').date()
    except Exception:
        data_fim = hoje
    if data_fim < data_inicio:
        data_inicio, data_fim = data_fim, data_inicio

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão ao carregar dossiê do motorista.", "danger")
        return redirect(url_for('historico_motoristas'))
    cur = con.cursor(dictionary=True)

    def inteiro(valor):
        try: return int(valor or 0)
        except Exception: return 0

    try:
        cur.execute("""
            SELECT p.*, e.nome_fantasia, e.razao_social
            FROM pessoas p
            LEFT JOIN empresas e ON e.id=p.empresa_id
            WHERE p.id=%s AND p.empresa_id=%s
              AND (p.tipo_cadastro='Motorista' OR p.tipo_cadastro IS NOT NULL)
        """, [motorista_id, empresa_id])
        motorista = cur.fetchone()
        if not motorista:
            flash("Motorista não encontrado para a empresa selecionada.", "warning")
            return redirect(url_for('historico_motoristas'))

        score_info = calcular_scores_motoristas_para_fila(empresa_id, [motorista_id], data_fim).get(motorista_id, {})
        score_info.setdefault('score', 80)
        score_info.setdefault('classificacao', classificar_score_operacional(80))
        score_info.setdefault('motivos', ['Sem movimentação crítica no período'])

        cur.execute("""
            SELECT COUNT(*) total_escalas,
                   SUM(CASE WHEN status_escala='Confirmado com rota' THEN 1 ELSE 0 END) confirmados_rota,
                   SUM(CASE WHEN status_escala='Cancelado, comparecer na base' THEN 1 ELSE 0 END) cancelados_base,
                   SUM(CASE WHEN status_presenca='Chegada confirmada' THEN 1 ELSE 0 END) chegadas_confirmadas,
                   SUM(CASE WHEN status_escala='Falta' OR status_presenca='Não compareceu' THEN 1 ELSE 0 END) faltas,
                   SUM(CASE WHEN status_presenca='Ausência justificada' THEN 1 ELSE 0 END) ausencias_justificadas
            FROM escala_motorista
            WHERE empresa_id=%s AND motorista_id=%s AND data_escala BETWEEN %s AND %s
        """, [empresa_id, motorista_id, data_inicio, data_fim])
        resumo = cur.fetchone() or {}
        for k in list(resumo.keys()): resumo[k] = inteiro(resumo.get(k))

        cur.execute("""
            SELECT COUNT(*) total_fila,
                   SUM(CASE WHEN status_fila='Atribuído para rota extra' THEN 1 ELSE 0 END) rotas_extras,
                   SUM(CASE WHEN status_fila='Dispensado' THEN 1 ELSE 0 END) dispensas
            FROM fila_cancelados_base
            WHERE empresa_id=%s AND motorista_id=%s AND data_fila BETWEEN %s AND %s
        """, [empresa_id, motorista_id, data_inicio, data_fim])
        fila_resumo = cur.fetchone() or {}
        resumo.update({k: inteiro(fila_resumo.get(k)) for k in ['total_fila','rotas_extras','dispensas']})

        cur.execute("""
            SELECT COUNT(*) checkins_total,
                   SUM(CASE WHEN resultado='Aprovado' THEN 1 ELSE 0 END) checkins_aprovados,
                   SUM(CASE WHEN resultado <> 'Aprovado' THEN 1 ELSE 0 END) checkins_bloqueados
            FROM auditoria_checkin_base
            WHERE empresa_id=%s AND motorista_id=%s AND DATE(data_tentativa) BETWEEN %s AND %s
        """, [empresa_id, motorista_id, data_inicio, data_fim])
        aud_resumo = cur.fetchone() or {}
        resumo.update({k: inteiro(aud_resumo.get(k)) for k in ['checkins_total','checkins_aprovados','checkins_bloqueados']})

        cur.execute("""
            SELECT em.*, b.nome_base,
                   CASE WHEN ce.id IS NULL THEN 'Não' ELSE 'Sim' END AS ciencia_registrada
            FROM escala_motorista em
            LEFT JOIN bases_operacionais b ON b.id=em.base_operacional_id
            LEFT JOIN ciencia_escala_motorista ce ON ce.escala_id=em.id AND ce.motorista_id=em.motorista_id
            WHERE em.empresa_id=%s AND em.motorista_id=%s AND em.data_escala BETWEEN %s AND %s
            ORDER BY em.data_escala DESC, em.id DESC
            LIMIT 60
        """, [empresa_id, motorista_id, data_inicio, data_fim])
        escalas = cur.fetchall()
        for e in escalas:
            e['horario_apresentacao'] = normalizar_horario_input(e.get('horario_apresentacao'))

        cur.execute("""
            SELECT f.*, b.nome_base
            FROM fila_cancelados_base f
            LEFT JOIN bases_operacionais b ON b.id=f.base_operacional_id
            WHERE f.empresa_id=%s AND f.motorista_id=%s AND f.data_fila BETWEEN %s AND %s
            ORDER BY f.hora_confirmacao DESC
            LIMIT 60
        """, [empresa_id, motorista_id, data_inicio, data_fim])
        fila = cur.fetchall()

        cur.execute("""
            SELECT aud.*, b.nome_base
            FROM auditoria_checkin_base aud
            LEFT JOIN bases_operacionais b ON b.id=aud.base_operacional_id
            WHERE aud.empresa_id=%s AND aud.motorista_id=%s AND DATE(aud.data_tentativa) BETWEEN %s AND %s
            ORDER BY aud.data_tentativa DESC
            LIMIT 80
        """, [empresa_id, motorista_id, data_inicio, data_fim])
        checkins = cur.fetchall()

        justificativas = []
        try:
            cur.execute("""
                SELECT j.*, b.nome_base, u.login AS usuario_analise_login
                FROM justificativas_ausencia_motorista j
                LEFT JOIN bases_operacionais b ON b.id=j.base_operacional_id
                LEFT JOIN usuarios u ON u.id=j.usuario_analise_id
                WHERE j.empresa_id=%s AND j.motorista_id=%s AND j.data_escala BETWEEN %s AND %s
                ORDER BY j.data_envio DESC
                LIMIT 60
            """, [empresa_id, motorista_id, data_inicio, data_fim])
            justificativas = cur.fetchall()
        except Exception:
            justificativas = []

        documentos = []
        try:
            cur.execute("""
                SELECT id, tipo_documento_pagamento, numero_nf, data_emissao, valor_total, status_nf, data_envio, data_aprovacao, data_pagamento, data_recusa, motivo_recusa
                FROM motorista_notas_fiscais
                WHERE empresa_id=%s AND motorista_id=%s
                ORDER BY data_envio DESC
                LIMIT 40
            """, [empresa_id, motorista_id])
            documentos = cur.fetchall()
        except Exception:
            documentos = []

        rotas = []
        try:
            cur.execute("""
                SELECT id, data_lancamento, identi_rota, valor_rota, valor_km, outras_despesas, tipo_rota, situacao_rota, status_motorista
                FROM rotas
                WHERE empresa_id=%s AND motorista_id=%s AND data_lancamento BETWEEN %s AND %s
                ORDER BY data_lancamento DESC, id DESC
                LIMIT 60
            """, [empresa_id, motorista_id, data_inicio, data_fim])
            rotas = cur.fetchall()
        except Exception:
            rotas = []

        # Linha do tempo unificada com os principais eventos do período.
        timeline = []
        for e in escalas:
            timeline.append({'data': e.get('data_escala'), 'tipo': 'Escala', 'icone': 'fa-calendar-check', 'classe': 'primary', 'titulo': e.get('status_escala'), 'descricao': f"Presença: {e.get('status_presenca')} | Base: {e.get('nome_base') or e.get('base_operacao') or '-'}"})
        for f in fila:
            timeline.append({'data': f.get('hora_confirmacao'), 'tipo': 'Fila', 'icone': 'fa-users-line', 'classe': 'info', 'titulo': f.get('status_fila'), 'descricao': f"Posição #{f.get('posicao_fila')} | Base: {f.get('nome_base') or f.get('base_operacao') or '-'}"})
        for c in checkins:
            classe = 'success' if c.get('resultado') == 'Aprovado' else 'danger'
            timeline.append({'data': c.get('data_tentativa'), 'tipo': 'Check-in', 'icone': 'fa-qrcode', 'classe': classe, 'titulo': c.get('resultado'), 'descricao': c.get('motivo_bloqueio') or f"Base: {c.get('nome_base') or '-'}"})
        for j in justificativas:
            timeline.append({'data': j.get('data_envio'), 'tipo': 'Justificativa', 'icone': 'fa-file-signature', 'classe': 'warning', 'titulo': j.get('status_justificativa'), 'descricao': j.get('motivo') or ''})
        timeline = sorted(timeline, key=lambda x: str(x.get('data') or ''), reverse=True)[:80]

        filtros = {'empresa_id': str(empresa_id), 'data_inicio': data_inicio.strftime('%Y-%m-%d'), 'data_fim': data_fim.strftime('%Y-%m-%d')}
        return render_template('historico_motorista_detalhe.html',
                               usuario_logado=session.get('usuario_nome', 'Usuário'),
                               motorista=motorista,
                               score_info=score_info,
                               resumo=resumo,
                               escalas=escalas,
                               fila=fila,
                               checkins=checkins,
                               justificativas=justificativas,
                               documentos=documentos,
                               rotas=rotas,
                               timeline=timeline,
                               filtros=filtros,
                               is_super_admin=is_super_admin)
    except Exception as e:
        print(f"Erro ao carregar dossiê do motorista: {e}")
        flash("Erro ao carregar dossiê do motorista.", "danger")
        return redirect(url_for('historico_motoristas'))
    finally:
        if con and con.is_connected():
            cur.close()
            con.close()




# ==========================================================
# BLOCO 4 - SOLICITAR PAGAMENTO DE NF MOTORISTA
# NF aprovada gera título financeiro automático
# ==========================================================
def _dias_vencimento_pagamento_nf_motorista():
    """Prazo padrão para vencimento do título gerado a partir da NF do motorista."""
    return 3


def _status_nf_motorista_com_pagamento_solicitado():
    return 'Pagamento solicitado'


def gerar_titulo_financeiro_por_nf_motorista(cur, nf_id, empresa_id, usuario_id, data_vencimento=None, forma_pagamento='PIX'):
    """
    Gera um título financeiro a pagar a partir de uma NF aprovada de motorista.
    Retorna (titulo_id, mensagem).
    """
    cur.execute("""
        SELECT nf.id,
               nf.empresa_id,
               nf.motorista_id,
               nf.numero_nf,
               COALESCE(nf.tipo_documento_pagamento, 'XML') AS tipo_documento_pagamento,
               nf.data_emissao,
               nf.valor_total,
               nf.valor_liquido,
               nf.status_nf,
               mot.nome_completo AS motorista_nome,
               mot.cpf_cnpj AS motorista_cpf_cnpj
        FROM motorista_notas_fiscais nf
        INNER JOIN pessoas mot
                ON mot.id = nf.motorista_id
               AND mot.empresa_id = nf.empresa_id
        WHERE nf.id = %s
          AND nf.empresa_id = %s
        LIMIT 1
    """, (nf_id, empresa_id))
    nf = cur.fetchone()

    if not nf:
        raise ValueError('Documento do motorista não encontrado ou não pertence à empresa informada.')

    if nf.get('status_nf') != 'Aprovada':
        raise ValueError(f"Somente documentos aprovados podem gerar solicitação de pagamento. Status atual: {nf.get('status_nf')}.")

    tipo_documento_pagamento = str(nf.get('tipo_documento_pagamento') or 'XML').strip().upper()
    eh_sem_nf = tipo_documento_pagamento == 'SEM_NF'
    origem_titulo = 'SEM_NF_MOTORISTA' if eh_sem_nf else 'NF_MOTORISTA'
    tipo_vinculo_documento = origem_titulo

    cur.execute("""
        SELECT id, status_titulo
        FROM titulos_financeiros
        WHERE empresa_id = %s
          AND origem = %s
          AND origem_id = %s
          AND status_titulo NOT IN ('Cancelado', 'Estornado')
        ORDER BY id DESC
        LIMIT 1
    """, (empresa_id, origem_titulo, nf_id))
    titulo_existente = cur.fetchone()

    if titulo_existente:
        raise ValueError(f"Este documento já possui título financeiro ativo vinculado: #{titulo_existente['id']} ({titulo_existente['status_titulo']}).")

    cur.execute("""
        SELECT MIN(v.id) AS id,
               v.rota_id,
               MAX(v.valor_rota) AS valor_rota,
               MAX(r.identi_rota) AS identi_rota,
               MAX(r.data_lancamento) AS data_lancamento,
               MAX(r.tipo_rota) AS tipo_rota
        FROM motorista_nf_rotas v
        INNER JOIN rotas r
                ON r.id = v.rota_id
               AND r.empresa_id = v.empresa_id
        WHERE v.motorista_nf_id = %s
          AND v.empresa_id = %s
        GROUP BY v.rota_id
        ORDER BY v.rota_id ASC
    """, (nf_id, empresa_id))
    rotas = cur.fetchall()

    if not rotas:
        raise ValueError('Não é possível solicitar pagamento: o documento não possui rotas vinculadas.')

    valor_nf = converter_decimal(nf.get('valor_liquido') or nf.get('valor_total'))
    if valor_nf <= 0:
        raise ValueError('Não é possível solicitar pagamento: valor do documento inválido.')

    hoje = date.today()
    if not data_vencimento:
        data_vencimento = (hoje + timedelta(days=_dias_vencimento_pagamento_nf_motorista())).strftime('%Y-%m-%d')

    if isinstance(data_vencimento, date):
        data_vencimento = data_vencimento.strftime('%Y-%m-%d')

    if not validar_data_iso(str(data_vencimento)):
        raise ValueError('Data de vencimento inválida para o título financeiro.')

    rotas_txt = '; '.join([str(r.get('identi_rota') or r.get('rota_id')) for r in rotas])
    numero_documento = str(nf.get('numero_nf') or (f"SEM-NF-{nf_id}" if eh_sem_nf else f"NF-{nf_id}")).strip()

    if eh_sem_nf:
        descricao = f"Prestação de serviço sem NF - {numero_documento}"
        historico = (
            f"Prestação de serviço de entrega sem emissão de NF, referente ao documento {numero_documento}, "
            f"solicitado por {nf.get('motorista_nome')}, contemplando as rotas: {rotas_txt}."
        )
        observacao_geracao = f"Título gerado automaticamente a partir da solicitação sem NF do motorista #{nf_id}."
        descricao_vinculo_documento = f"Solicitação sem NF nº {numero_documento}"
        mensagem_documento = f"documento sem NF {numero_documento}"
    else:
        descricao = f"Prestação de serviço - NFSe nº {numero_documento}"
        historico = (
            f"Prestação de serviço de entrega referente à NFSe nº {numero_documento}, "
            f"emitida por {nf.get('motorista_nome')}, contemplando as rotas: {rotas_txt}."
        )
        observacao_geracao = f"Título gerado automaticamente a partir da NF do motorista #{nf_id}."
        descricao_vinculo_documento = f"NF Motorista nº {numero_documento}"
        mensagem_documento = f"NF {numero_documento}"

    cur.execute("""
        INSERT INTO titulos_financeiros
            (empresa_id, tipo_titulo, origem, origem_id, pessoa_id, numero_documento,
             descricao, historico, valor_original, valor_desconto, valor_acrescimo,
             valor_liquido, data_emissao, data_competencia, data_vencimento,
             forma_pagamento, conta_caixa_prevista_id, status_titulo, observacao,
             usuario_criacao_id)
        VALUES
            (%s, 'PAGAR', %s, %s, %s, %s,
             %s, %s, %s, 0.00, 0.00,
             %s, CURDATE(), %s, %s,
             %s, NULL, 'Solicitado', %s,
             %s)
    """, (
        empresa_id,
        origem_titulo,
        nf_id,
        nf.get('motorista_id'),
        numero_documento,
        descricao,
        historico,
        valor_nf,
        valor_nf,
        nf.get('data_emissao') or hoje.strftime('%Y-%m-%d'),
        data_vencimento,
        forma_pagamento or 'PIX',
        observacao_geracao,
        usuario_id
    ))
    titulo_id = cur.lastrowid

    cur.execute("""
        INSERT INTO titulos_financeiros_vinculos
            (empresa_id, titulo_financeiro_id, tipo_vinculo, origem_tabela, origem_id, descricao, valor_vinculo)
        VALUES
            (%s, %s, %s, 'motorista_notas_fiscais', %s, %s, %s)
    """, (
        empresa_id,
        titulo_id,
        tipo_vinculo_documento,
        nf_id,
        descricao_vinculo_documento,
        valor_nf
    ))

    for rota in rotas:
        cur.execute("""
            INSERT INTO titulos_financeiros_vinculos
                (empresa_id, titulo_financeiro_id, tipo_vinculo, origem_tabela, origem_id, descricao, valor_vinculo)
            VALUES
                (%s, %s, 'ROTA', 'rotas', %s, %s, %s)
        """, (
            empresa_id,
            titulo_id,
            rota.get('rota_id'),
            f"Rota {rota.get('identi_rota') or rota.get('rota_id')}",
            converter_decimal(rota.get('valor_rota'))
        ))

    status_pagamento_solicitado = _status_nf_motorista_com_pagamento_solicitado()
    cur.execute("""
        UPDATE motorista_notas_fiscais
        SET status_nf = %s,
            observacao = CONCAT(
                COALESCE(observacao, ''),
                CASE WHEN COALESCE(observacao, '') = '' THEN '' ELSE '\n' END,
                'Pagamento solicitado em ',
                DATE_FORMAT(NOW(), '%d/%m/%Y %H:%i'),
                '. Título financeiro gerado: #',
                %s
            )
        WHERE id = %s
          AND empresa_id = %s
          AND status_nf = 'Aprovada'
    """, (status_pagamento_solicitado, titulo_id, nf_id, empresa_id))

    return titulo_id, f"Título financeiro #{titulo_id} gerado para o {mensagem_documento}."


@app.route('/financeiro/nfs-motoristas/<int:id>/solicitar-pagamento', methods=['POST'])
@login_required
@financeiro_nf_motorista_required
def solicitar_pagamento_nf_motorista(id):
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = int(session.get('is_super_admin') or 0) == 1

    data_vencimento = (request.form.get('data_vencimento') or '').strip()
    forma_pagamento = (request.form.get('forma_pagamento') or 'PIX').strip()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('financeiro_nfs_motoristas'))

    cur = con.cursor(dictionary=True)
    try:
        query = """
            SELECT id, empresa_id, numero_nf, status_nf
            FROM motorista_notas_fiscais
            WHERE id = %s
        """
        params = [id]
        if not is_super_admin:
            query += " AND empresa_id = %s"
            params.append(empresa_logada_id)
        query += " LIMIT 1"
        cur.execute(query, params)
        nf = cur.fetchone()

        if not nf:
            flash("Documento do motorista não encontrado ou não pertence à empresa logada.", "danger")
            return redirect(url_for('financeiro_nfs_motoristas'))

        titulo_id, msg = gerar_titulo_financeiro_por_nf_motorista(
            cur,
            nf_id=id,
            empresa_id=nf['empresa_id'],
            usuario_id=usuario_id,
            data_vencimento=data_vencimento or None,
            forma_pagamento=forma_pagamento or 'PIX'
        )
        registrar_auditoria_financeira(
            cur,
            empresa_id=nf['empresa_id'],
            usuario_id=usuario_id,
            acao='TITULO_GERADO_DOCUMENTO_MOTORISTA',
            modulo='DOCUMENTOS_MOTORISTAS',
            entidade_tipo='MOTORISTA_NOTA_FISCAL',
            entidade_id=id,
            titulo_financeiro_id=titulo_id,
            status_anterior=nf.get('status_nf'),
            status_novo='Pagamento solicitado',
            motivo='Solicitação de pagamento de documento',
            observacao=msg,
            dados_depois={'numero_nf': nf.get('numero_nf'), 'forma_pagamento': forma_pagamento or 'PIX', 'data_vencimento': data_vencimento}
        )
        con.commit()

        registrar_historico_nf_motorista(
            empresa_id=nf['empresa_id'],
            motorista_nf_id=id,
            usuario_id=usuario_id,
            status_anterior='Aprovada',
            status_novo=_status_nf_motorista_com_pagamento_solicitado(),
            motivo='Solicitação de pagamento',
            observacao=f"{msg} O título entrou em Contas a Pagar como Solicitado."
        )

        flash(f"Pagamento solicitado com sucesso. {msg}", "success")
        return redirect(url_for('detalhes_titulo_financeiro', id=titulo_id))

    except Exception as e:
        con.rollback()
        print(f"Erro ao solicitar pagamento da NF motorista {id}: {e}")
        flash(f"Erro ao solicitar pagamento: {e}", "danger")
        return redirect(url_for('detalhes_nf_motorista', id=id))

    finally:
        fechar_cursor_conexao(cur, con)


# ==========================================================
# BLOCO 3 — FINANCEIRO BASE: TÍTULOS + CONTAS CAIXA
# ==========================================================
def financeiro_base_status_titulos():
    return [
        'Aberto',
        'Solicitado',
        'Aprovado para pagamento',
        'Agendado',
        'Pago',
        'Recebido',
        'Cancelado',
        'Estornado'
    ]


def financeiro_base_origens():
    return [
        'MANUAL',
        'NF_MOTORISTA',
        'SEM_NF_MOTORISTA',
        'AJUDANTE',
        'FATURAMENTO',
        'REEMBOLSO',
        'DESPESA_OPERACIONAL',
        'OUTRO'
    ]


def financeiro_base_formas_pagamento():
    return [
        'PIX',
        'Transferência bancária',
        'Boleto',
        'Dinheiro',
        'Cartão',
        'Outro'
    ]


def financeiro_base_tipos_conta_caixa():
    return [
        'Conta corrente',
        'Conta pagamento',
        'Caixa físico',
        'Carteira digital',
        'Outro'
    ]


# ----------------------------------------------------------
# Bloco 6.0 — Parâmetros financeiros por empresa
# ----------------------------------------------------------
PARAMETROS_FINANCEIROS_PADRAO = {
    # Baixa financeira
    'baixa.exigir_conta_caixa': {
        'grupo': 'baixa', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Exigir seleção de conta caixa para baixar títulos.'
    },
    'baixa.exigir_comprovante': {
        'grupo': 'baixa', 'tipo': 'boolean', 'valor': '0',
        'descricao': 'Exigir upload de comprovante ao baixar títulos financeiros.'
    },
    'baixa.permitir_pagamento_parcial': {
        'grupo': 'baixa', 'tipo': 'boolean', 'valor': '0',
        'descricao': 'Permitir baixa com valor menor que o valor líquido do título.'
    },
    'baixa.permitir_valor_diferente': {
        'grupo': 'baixa', 'tipo': 'boolean', 'valor': '0',
        'descricao': 'Permitir baixa com valor diferente do valor líquido do título.'
    },
    'baixa.permitir_data_retroativa': {
        'grupo': 'baixa', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Permitir informar data de baixa anterior à data atual.'
    },

    # Estorno
    'estorno.permitir_estorno_baixa': {
        'grupo': 'estorno', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Permitir estornar baixas financeiras.'
    },
    'estorno.exigir_motivo': {
        'grupo': 'estorno', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Exigir motivo para estornar uma baixa financeira.'
    },
    'estorno.permitir_reabrir_titulo': {
        'grupo': 'estorno', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Permitir que o estorno reabra o título para nova baixa.'
    },
    'estorno.permitir_encerrar_estornado': {
        'grupo': 'estorno', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Permitir encerrar o título como estornado.'
    },
    'estorno.permitir_tratativa_pos_estorno': {
        'grupo': 'estorno', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Permitir tratativa pós-estorno para documentos/rotas.'
    },

    # Caixa
    'caixa.permitir_saldo_negativo': {
        'grupo': 'caixa', 'tipo': 'boolean', 'valor': '0',
        'descricao': 'Permitir baixa de contas a pagar mesmo sem saldo suficiente no caixa.'
    },
    'caixa.conta_padrao_id': {
        'grupo': 'caixa', 'tipo': 'integer', 'valor': '',
        'descricao': 'Conta caixa padrão sugerida nas baixas financeiras.'
    },
    'caixa.forma_pagamento_padrao': {
        'grupo': 'caixa', 'tipo': 'string', 'valor': 'PIX',
        'descricao': 'Forma de pagamento padrão para títulos/documentos.'
    },

    # Documentos de motorista
    'documentos.permitir_sem_nf_pf': {
        'grupo': 'documentos', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Permitir que motorista pessoa física solicite pagamento sem NF.'
    },
    'documentos.permitir_xml_nf': {
        'grupo': 'documentos', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Permitir envio de XML/NF por motoristas/prestadores.'
    },
    'documentos.permitir_reenvio_recusado': {
        'grupo': 'documentos', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Permitir reenvio de documento recusado.'
    },
    'documentos.exigir_xml_cnpj': {
        'grupo': 'documentos', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Exigir XML/NF para prestadores CNPJ.'
    },
    'documentos.permitir_reaproveitar_pos_estorno': {
        'grupo': 'documentos', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Permitir reaproveitamento de documento após estorno quando fiscalmente correto.'
    },

    # Títulos automáticos
    'titulos.gerar_automatico_documento_aprovado': {
        'grupo': 'titulos', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Gerar título financeiro automaticamente ao solicitar pagamento de documento aprovado.'
    },
    'titulos.dias_padrao_vencimento_motorista': {
        'grupo': 'titulos', 'tipo': 'integer', 'valor': '5',
        'descricao': 'Quantidade padrão de dias para vencimento de títulos de motorista.'
    },
}


def normalizar_boolean_param(valor):
    return '1' if str(valor).lower() in ['1', 'true', 'sim', 'on', 'yes'] else '0'


def parametro_bool(valor):
    return str(valor).lower() in ['1', 'true', 'sim', 'on', 'yes']


# ==========================================================
# BLOCO 12.5 - PARÂMETROS OPERACIONAIS DO MOTORISTA
# ==========================================================
PARAMETROS_OPERACIONAIS_MOTORISTA_PADRAO = {
    'motorista.horario_limite_disponibilidade': {
        'grupo': 'motorista', 'tipo': 'time', 'valor': '11:00',
        'descricao': 'Horário limite para motorista informar disponibilidade ou ausência.'
    },
    'motorista.bloquear_disponibilidade_apos_limite': {
        'grupo': 'motorista', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Bloquear alteração de disponibilidade pelo motorista após o horário limite.'
    },
    'motorista.permitir_liberacao_excepcional': {
        'grupo': 'motorista', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Permitir liberação excepcional pelo supervisor/administrador.'
    },
    'checkin.exigir_qrcode': {
        'grupo': 'checkin', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Exigir QR Code dinâmico para confirmar chegada na base.'
    },
    'checkin.exigir_selfie': {
        'grupo': 'checkin', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Exigir selfie para confirmar chegada na base.'
    },
    'checkin.exigir_gps_raio': {
        'grupo': 'checkin', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Exigir localização dentro do raio permitido da base.'
    },
    'checkin.permitir_checkin_manual_supervisor': {
        'grupo': 'checkin', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Permitir confirmação manual de chegada pelo supervisor.'
    },
    'escala.aplicar_falta_automatica_checkin': {
        'grupo': 'escala', 'tipo': 'boolean', 'valor': '1',
        'descricao': 'Aplicar falta automática se motorista convocado não fizer check-in até o horário limite.'
    },
    'escala.horario_limite_checkin': {
        'grupo': 'escala', 'tipo': 'time', 'valor': '11:00',
        'descricao': 'Horário limite para check-in antes da falta automática.'
    },
}


def normalizar_time_param(valor, padrao='11:00'):
    valor = str(valor or '').strip()[:5]
    try:
        hora, minuto = valor.split(':')
        hora = int(hora)
        minuto = int(minuto)
        if hora < 0 or hora > 23 or minuto < 0 or minuto > 59:
            return padrao
        return f'{hora:02d}:{minuto:02d}'
    except Exception:
        return padrao


def carregar_parametros_operacionais_motorista_empresa(empresa_id, cur=None):
    parametros = {chave: dados.copy() for chave, dados in PARAMETROS_OPERACIONAIS_MOTORISTA_PADRAO.items()}
    fechar_no_final = False
    con_local = None
    cur_local = cur
    try:
        if cur_local is None:
            con_local = obter_conexao()
            if con_local is None:
                return parametros
            cur_local = con_local.cursor(dictionary=True)
            fechar_no_final = True

        # Compatibilidade: se ainda houver configuração antiga, usa como base inicial.
        try:
            if tabela_existe(cur_local, 'configuracoes_disponibilidade'):
                cur_local.execute("""
                    SELECT horario_limite_edicao
                    FROM configuracoes_disponibilidade
                    WHERE empresa_id = %s LIMIT 1
                """, (empresa_id,))
                row_disp = cur_local.fetchone()
                if row_disp and row_disp.get('horario_limite_edicao'):
                    parametros['motorista.horario_limite_disponibilidade']['valor'] = str(row_disp.get('horario_limite_edicao'))[:5]
        except Exception:
            pass

        try:
            if tabela_existe(cur_local, 'configuracoes_escala_motorista'):
                cur_local.execute("""
                    SELECT horario_limite_presenca, aplicar_falta_automatica
                    FROM configuracoes_escala_motorista
                    WHERE empresa_id = %s LIMIT 1
                """, (empresa_id,))
                row_esc = cur_local.fetchone()
                if row_esc:
                    if row_esc.get('horario_limite_presenca'):
                        parametros['escala.horario_limite_checkin']['valor'] = str(row_esc.get('horario_limite_presenca'))[:5]
                    parametros['escala.aplicar_falta_automatica_checkin']['valor'] = '1' if str(row_esc.get('aplicar_falta_automatica') or 'S').upper() == 'S' else '0'
        except Exception:
            pass

        chaves = list(PARAMETROS_OPERACIONAIS_MOTORISTA_PADRAO.keys())
        cur_local.execute("""
            SELECT grupo, chave, valor, tipo, descricao
            FROM empresa_parametros
            WHERE empresa_id = %s
              AND chave IN ({})
        """.format(','.join(['%s'] * len(chaves))), tuple([empresa_id] + chaves))
        for row in cur_local.fetchall():
            chave = row.get('chave')
            if chave in parametros:
                parametros[chave]['valor'] = row.get('valor') if row.get('valor') is not None else parametros[chave].get('valor', '')
                parametros[chave]['tipo'] = row.get('tipo') or parametros[chave].get('tipo', 'string')
                parametros[chave]['descricao'] = row.get('descricao') or parametros[chave].get('descricao', '')
        return parametros
    except Exception as e:
        print(f"Erro ao carregar parâmetros operacionais da empresa {empresa_id}: {e}")
        return parametros
    finally:
        if fechar_no_final:
            fechar_cursor_conexao(cur_local, con_local)


def salvar_parametro_operacional_motorista_empresa(cur, empresa_id, chave, valor, usuario_id=None):
    base = PARAMETROS_OPERACIONAIS_MOTORISTA_PADRAO.get(chave, {})
    grupo = base.get('grupo', 'motorista')
    tipo = base.get('tipo', 'string')
    descricao = base.get('descricao', '')
    if tipo == 'boolean':
        valor = normalizar_boolean_param(valor)
    elif tipo == 'time':
        valor = normalizar_time_param(valor, base.get('valor', '11:00'))
    else:
        valor = str(valor or '').strip()

    cur.execute("""
        INSERT INTO empresa_parametros
            (empresa_id, grupo, chave, valor, tipo, descricao, grupo_financeiro, usuario_atualizacao_id, created_at, updated_at)
        VALUES
            (%s, %s, %s, %s, %s, %s, 0, %s, NOW(), NOW())
        ON DUPLICATE KEY UPDATE
            valor = VALUES(valor),
            tipo = VALUES(tipo),
            descricao = VALUES(descricao),
            grupo_financeiro = 0,
            usuario_atualizacao_id = VALUES(usuario_atualizacao_id),
            updated_at = NOW()
    """, (empresa_id, grupo, chave, valor, tipo, descricao, usuario_id))


def semear_parametros_operacionais_motorista_empresa(cur, empresa_id, usuario_id=None):
    existentes = carregar_parametros_operacionais_motorista_empresa(empresa_id, cur=cur)
    for chave, base in existentes.items():
        salvar_parametro_operacional_motorista_empresa(cur, empresa_id, chave, base.get('valor', ''), usuario_id=usuario_id)


def parametro_operacional_bool(empresa_id, chave, padrao=True, cur=None):
    parametros = carregar_parametros_operacionais_motorista_empresa(empresa_id, cur=cur)
    if chave in parametros:
        return parametro_bool(parametros[chave].get('valor'))
    return bool(padrao)


def carregar_parametros_financeiros_empresa(empresa_id, cur=None):
    """Carrega parâmetros financeiros da empresa mesclando defaults + banco."""
    parametros = {chave: dados.copy() for chave, dados in PARAMETROS_FINANCEIROS_PADRAO.items()}
    fechar_no_final = False
    con_local = None
    cur_local = cur

    try:
        if cur_local is None:
            con_local = obter_conexao()
            if con_local is None:
                return parametros
            cur_local = con_local.cursor(dictionary=True)
            fechar_no_final = True

        cur_local.execute("""
            SELECT grupo, chave, valor, tipo, descricao
            FROM empresa_parametros
            WHERE empresa_id = %s
              AND grupo_financeiro = 1
        """, (empresa_id,))
        for row in cur_local.fetchall():
            chave = row.get('chave')
            if not chave:
                continue
            if chave not in parametros:
                parametros[chave] = {
                    'grupo': row.get('grupo') or 'outros',
                    'tipo': row.get('tipo') or 'string',
                    'valor': row.get('valor') or '',
                    'descricao': row.get('descricao') or ''
                }
            else:
                parametros[chave]['valor'] = row.get('valor') if row.get('valor') is not None else parametros[chave].get('valor', '')
                parametros[chave]['tipo'] = row.get('tipo') or parametros[chave].get('tipo', 'string')
                parametros[chave]['descricao'] = row.get('descricao') or parametros[chave].get('descricao', '')
        return parametros
    except Exception as e:
        print(f"Erro ao carregar parâmetros financeiros da empresa {empresa_id}: {e}")
        return parametros
    finally:
        if fechar_no_final:
            fechar_cursor_conexao(cur_local, con_local)


def obter_parametro_empresa(empresa_id, chave, padrao=None, cur=None):
    parametros = carregar_parametros_financeiros_empresa(empresa_id, cur=cur)
    if chave in parametros:
        valor = parametros[chave].get('valor')
        tipo = parametros[chave].get('tipo')
        if tipo == 'boolean':
            return parametro_bool(valor)
        if tipo == 'integer':
            try:
                return int(valor)
            except Exception:
                try:
                    return int(padrao)
                except Exception:
                    return 0
        return valor
    return padrao


def registrar_auditoria_financeira(
    cur,
    *,
    empresa_id,
    usuario_id=None,
    acao,
    modulo='FINANCEIRO',
    entidade_tipo=None,
    entidade_id=None,
    titulo_financeiro_id=None,
    movimentacao_caixa_id=None,
    pessoa_id=None,
    status_anterior=None,
    status_novo=None,
    valor_anterior=None,
    valor_novo=None,
    motivo=None,
    observacao=None,
    dados_antes=None,
    dados_depois=None,
):
    """Registra auditoria financeira avançada sem interromper a operação principal."""
    try:
        def _json_safe(valor):
            if valor is None:
                return None
            try:
                return json.dumps(valor, ensure_ascii=False, default=str)
            except Exception:
                return str(valor)

        ip_origem = None
        user_agent = None
        try:
            ip_origem = request.headers.get('X-Forwarded-For', request.remote_addr)
            if ip_origem and ',' in ip_origem:
                ip_origem = ip_origem.split(',')[0].strip()
            user_agent = request.headers.get('User-Agent')
        except Exception:
            pass

        cur.execute("""
            INSERT INTO auditoria_financeira
                (empresa_id, usuario_id, modulo, acao, entidade_tipo, entidade_id,
                 titulo_financeiro_id, movimentacao_caixa_id, pessoa_id,
                 status_anterior, status_novo, valor_anterior, valor_novo,
                 motivo, observacao, dados_antes, dados_depois,
                 ip_origem, user_agent, created_at)
            VALUES
                (%s, %s, %s, %s, %s, %s,
                 %s, %s, %s,
                 %s, %s, %s, %s,
                 %s, %s, %s, %s,
                 %s, %s, NOW())
        """, (
            empresa_id, usuario_id, modulo, acao, entidade_tipo, entidade_id,
            titulo_financeiro_id, movimentacao_caixa_id, pessoa_id,
            status_anterior, status_novo, valor_anterior, valor_novo,
            motivo, observacao, _json_safe(dados_antes), _json_safe(dados_depois),
            ip_origem, user_agent
        ))
    except Exception as e:
        print(f"[Auditoria Financeira] Falha ao registrar auditoria: {e}")



def salvar_parametro_empresa(cur, empresa_id, chave, valor, usuario_id=None):
    base = PARAMETROS_FINANCEIROS_PADRAO.get(chave, {})
    grupo = base.get('grupo', 'outros')
    tipo = base.get('tipo', 'string')
    descricao = base.get('descricao', '')
    if tipo == 'boolean':
        valor = normalizar_boolean_param(valor)
    elif tipo == 'integer':
        valor = str(valor or '').strip()
        if valor and not valor.isdigit():
            valor = base.get('valor', '')
    else:
        valor = str(valor or '').strip()

    cur.execute("""
        INSERT INTO empresa_parametros
            (empresa_id, grupo, chave, valor, tipo, descricao, grupo_financeiro, usuario_atualizacao_id, created_at, updated_at)
        VALUES
            (%s, %s, %s, %s, %s, %s, 1, %s, NOW(), NOW())
        ON DUPLICATE KEY UPDATE
            valor = VALUES(valor),
            tipo = VALUES(tipo),
            descricao = VALUES(descricao),
            grupo_financeiro = 1,
            usuario_atualizacao_id = VALUES(usuario_atualizacao_id),
            updated_at = NOW()
    """, (empresa_id, grupo, chave, valor, tipo, descricao, usuario_id))


def semear_parametros_financeiros_empresa(cur, empresa_id, usuario_id=None):
    for chave, base in PARAMETROS_FINANCEIROS_PADRAO.items():
        salvar_parametro_empresa(cur, empresa_id, chave, base.get('valor', ''), usuario_id=usuario_id)


@app.route('/financeiro/configuracoes', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador', 'Financeiro')
def financeiro_configuracoes():
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash('Empresa não identificada na sessão. Faça login novamente.', 'danger')
        return redirect(url_for('logout'))

    con = obter_conexao()
    if con is None:
        flash('Erro de conexão com o banco de dados.', 'danger')
        return redirect(url_for('financeiro_titulos'))

    cur = con.cursor(dictionary=True)
    try:
        empresa_id_config = (request.values.get('empresa_id') or '').strip() if is_super_admin else str(empresa_logada_id)
        if not empresa_id_config or not empresa_id_config.isdigit():
            empresa_id_config = str(empresa_logada_id)
        empresa_id_config = int(empresa_id_config)

        if not is_super_admin and empresa_id_config != int(empresa_logada_id):
            flash('Você não tem permissão para alterar configurações de outra empresa.', 'danger')
            return redirect(url_for('financeiro_configuracoes'))

        cur.execute("SELECT id, nome_fantasia, razao_social FROM empresas WHERE id = %s LIMIT 1", (empresa_id_config,))
        empresa_config = cur.fetchone()
        if not empresa_config:
            flash('Empresa não encontrada para configuração.', 'danger')
            return redirect(url_for('financeiro_titulos'))

        if request.method == 'POST':
            for chave, base in PARAMETROS_FINANCEIROS_PADRAO.items():
                tipo = base.get('tipo')
                if tipo == 'boolean':
                    valor = '1' if request.form.get(chave) == '1' else '0'
                else:
                    valor = request.form.get(chave) or ''
                salvar_parametro_empresa(cur, empresa_id_config, chave, valor, usuario_id=usuario_id)

            cur.execute("""
                INSERT INTO historico_operacoes
                    (empresa_id, tipo_operacao, usuario_id, status_anterior, status_novo, motivo, observacao)
                VALUES
                    (%s, 'CONFIGURACOES_FINANCEIRAS', %s, 'Parâmetros anteriores', 'Parâmetros atualizados',
                     'Atualização de parâmetros financeiros', %s)
            """, (empresa_id_config, usuario_id, f'Configurações financeiras atualizadas para a empresa #{empresa_id_config}.'))
            registrar_auditoria_financeira(
                cur,
                empresa_id=empresa_id_config,
                usuario_id=usuario_id,
                acao='CONFIGURACAO_FINANCEIRA_ATUALIZADA',
                modulo='CONFIGURACOES_FINANCEIRAS',
                entidade_tipo='EMPRESA_PARAMETROS',
                entidade_id=empresa_id_config,
                status_anterior='Parâmetros anteriores',
                status_novo='Parâmetros atualizados',
                motivo='Atualização de parâmetros financeiros',
                observacao=f'Configurações financeiras atualizadas para a empresa #{empresa_id_config}.',
                dados_depois={chave: request.form.get(chave) for chave in PARAMETROS_FINANCEIROS_PADRAO.keys()}
            )
            con.commit()
            flash('Configurações financeiras salvas com sucesso.', 'success')
            return redirect(url_for('financeiro_configuracoes', empresa_id=empresa_id_config if is_super_admin else None))

        # Garante defaults no banco para a primeira abertura da tela.
        parametros_existentes = carregar_parametros_financeiros_empresa(empresa_id_config, cur=cur)
        if not parametros_existentes:
            semear_parametros_financeiros_empresa(cur, empresa_id_config, usuario_id=usuario_id)
            con.commit()
        parametros = carregar_parametros_financeiros_empresa(empresa_id_config, cur=cur)

        empresas = []
        if is_super_admin:
            cur.execute("SELECT id, nome_fantasia, razao_social FROM empresas ORDER BY nome_fantasia ASC")
            empresas = cur.fetchall()

        contas_caixa = carregar_contas_caixa_financeiro(empresa_id_config, is_super_admin=True, somente_ativas=True)

        grupos = {
            'baixa': 'Baixa financeira',
            'estorno': 'Estorno',
            'caixa': 'Caixa',
            'documentos': 'Documentos de motorista',
            'titulos': 'Títulos automáticos',
        }

        return render_template(
            'financeiro_configuracoes.html',
            parametros=parametros,
            grupos=grupos,
            empresa_config=empresa_config,
            empresas=empresas,
            empresa_id_config=empresa_id_config,
            is_super_admin=is_super_admin,
            contas_caixa=contas_caixa,
            formas_pagamento=financeiro_base_formas_pagamento()
        )
    except Exception as e:
        try:
            con.rollback()
        except Exception:
            pass
        print(f'Erro ao carregar/salvar configurações financeiras: {e}')
        flash('Erro técnico ao processar configurações financeiras.', 'danger')
        return redirect(url_for('financeiro_titulos'))
    finally:
        fechar_cursor_conexao(cur, con)


@app.route('/configuracoes/operacional-motorista', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador', 'Supervisor')
def configuracoes_operacionais_motorista():
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    escopo_global = usuario_eh_super_admin_ou_suporte()

    if not empresa_logada_id:
        flash('Empresa não identificada na sessão. Faça login novamente.', 'danger')
        return redirect(url_for('logout'))

    con = obter_conexao()
    if con is None:
        flash('Erro de conexão com o banco de dados.', 'danger')
        return redirect(url_for('dashboard'))

    cur = con.cursor(dictionary=True)
    try:
        empresa_id_config = (request.values.get('empresa_id') or '').strip() if escopo_global else str(empresa_logada_id)
        if not empresa_id_config or not empresa_id_config.isdigit():
            empresa_id_config = str(empresa_logada_id)
        empresa_id_config = int(empresa_id_config)

        if not escopo_global and empresa_id_config != int(empresa_logada_id):
            flash('Você não tem permissão para alterar configurações de outra empresa.', 'danger')
            return redirect(url_for('configuracoes_operacionais_motorista'))

        cur.execute("SELECT id, nome_fantasia, razao_social FROM empresas WHERE id = %s LIMIT 1", (empresa_id_config,))
        empresa_config = cur.fetchone()
        if not empresa_config:
            flash('Empresa não encontrada para configuração.', 'danger')
            return redirect(url_for('dashboard'))

        if request.method == 'POST':
            for chave, base in PARAMETROS_OPERACIONAIS_MOTORISTA_PADRAO.items():
                if base.get('tipo') == 'boolean':
                    valor = '1' if request.form.get(chave) == '1' else '0'
                else:
                    valor = request.form.get(chave) or base.get('valor', '')
                salvar_parametro_operacional_motorista_empresa(cur, empresa_id_config, chave, valor, usuario_id=usuario_id)

            # Sincroniza tabelas antigas para manter telas/rotinas já existentes compatíveis.
            horario_disp = normalizar_time_param(request.form.get('motorista.horario_limite_disponibilidade'), '11:00')
            horario_checkin = normalizar_time_param(request.form.get('escala.horario_limite_checkin'), '11:00')
            falta_auto = 'S' if request.form.get('escala.aplicar_falta_automatica_checkin') == '1' else 'N'

            try:
                if tabela_existe(cur, 'configuracoes_disponibilidade'):
                    cur.execute("""
                        UPDATE configuracoes_disponibilidade
                        SET horario_limite_edicao = %s,
                            updated_at = NOW()
                        WHERE empresa_id = %s
                    """, (horario_disp + ':00', empresa_id_config))
                    if cur.rowcount == 0:
                        cur.execute("""
                            INSERT INTO configuracoes_disponibilidade
                                (empresa_id, horario_limite_edicao, limite_dias_disponiveis_semana, permite_liberacao_setimo_dia, created_at, updated_at)
                            VALUES
                                (%s, %s, 6, 'S', NOW(), NOW())
                        """, (empresa_id_config, horario_disp + ':00'))
            except Exception as e:
                print(f"[Config Motorista] Aviso ao sincronizar configuracoes_disponibilidade: {e}")

            try:
                if tabela_existe(cur, 'configuracoes_escala_motorista'):
                    cur.execute("""
                        UPDATE configuracoes_escala_motorista
                        SET horario_limite_presenca = %s,
                            aplicar_falta_automatica = %s,
                            updated_at = NOW()
                        WHERE empresa_id = %s
                    """, (horario_checkin + ':00', falta_auto, empresa_id_config))
                    if cur.rowcount == 0:
                        cur.execute("""
                            INSERT INTO configuracoes_escala_motorista
                                (empresa_id, horario_limite_presenca, aplicar_falta_automatica, created_at, updated_at)
                            VALUES
                                (%s, %s, %s, NOW(), NOW())
                        """, (empresa_id_config, horario_checkin + ':00', falta_auto))
            except Exception as e:
                print(f"[Config Motorista] Aviso ao sincronizar configuracoes_escala_motorista: {e}")

            try:
                cur.execute("""
                    INSERT INTO historico_operacoes
                        (empresa_id, tipo_operacao, usuario_id, status_anterior, status_novo, motivo, observacao)
                    VALUES
                        (%s, 'CONFIGURACOES_OPERACIONAIS_MOTORISTA', %s, 'Parâmetros anteriores', 'Parâmetros atualizados',
                         'Atualização de parâmetros operacionais do motorista', %s)
                """, (empresa_id_config, usuario_id, f'Configurações operacionais do motorista atualizadas para a empresa #{empresa_id_config}.'))
            except Exception as e:
                print(f"[Config Motorista] Aviso ao registrar histórico: {e}")

            con.commit()
            flash('Configurações operacionais do motorista salvas com sucesso.', 'success')
            return redirect(url_for('configuracoes_operacionais_motorista', empresa_id=empresa_id_config if escopo_global else None))

        semear_parametros_operacionais_motorista_empresa(cur, empresa_id_config, usuario_id=usuario_id)
        con.commit()
        parametros = carregar_parametros_operacionais_motorista_empresa(empresa_id_config, cur=cur)

        empresas = []
        if escopo_global:
            cur.execute("SELECT id, nome_fantasia, razao_social FROM empresas ORDER BY nome_fantasia ASC, razao_social ASC")
            empresas = cur.fetchall()

        grupos = {
            'motorista': 'Disponibilidade do motorista',
            'checkin': 'Check-in em base',
            'escala': 'Falta automática'
        }

        return render_template(
            'configuracoes_operacionais_motorista.html',
            parametros=parametros,
            grupos=grupos,
            empresa_config=empresa_config,
            empresas=empresas,
            empresa_id_config=empresa_id_config,
            escopo_global=escopo_global
        )
    except Exception as e:
        try:
            con.rollback()
        except Exception:
            pass
        print(f'Erro ao carregar/salvar configurações operacionais do motorista: {e}')
        flash('Erro técnico ao processar configurações operacionais do motorista.', 'danger')
        return redirect(url_for('dashboard'))
    finally:
        fechar_cursor_conexao(cur, con)


def carregar_pessoas_financeiro(empresa_id, is_super_admin=False):
    con = obter_conexao()
    if con is None:
        return []

    cur = con.cursor(dictionary=True)
    try:
        query = """
            SELECT p.id,
                   p.empresa_id,
                   p.nome_completo,
                   p.cpf_cnpj,
                   p.tipo_cadastro,
                   e.nome_fantasia AS empresa_nome,
                   e.razao_social AS empresa_razao_social
            FROM pessoas p
            INNER JOIN empresas e ON e.id = p.empresa_id
            WHERE p.status_cadastro = 'Ativo'
        """
        params = []
        if not is_super_admin:
            query += " AND p.empresa_id = %s"
            params.append(empresa_id)
        query += " ORDER BY e.nome_fantasia ASC, p.nome_completo ASC"
        cur.execute(query, params)
        return cur.fetchall()
    except Exception as e:
        print(f"Erro ao carregar pessoas do financeiro: {e}")
        return []
    finally:
        fechar_cursor_conexao(cur, con)


def carregar_contas_caixa_financeiro(empresa_id, is_super_admin=False, somente_ativas=True):
    con = obter_conexao()
    if con is None:
        return []

    cur = con.cursor(dictionary=True)
    try:
        query = """
            SELECT c.id,
                   c.empresa_id,
                   c.nome_conta,
                   c.tipo_conta,
                   c.banco,
                   c.agencia,
                   c.numero_conta,
                   c.saldo_inicial,
                   c.status_conta,
                   e.nome_fantasia AS empresa_nome,
                   e.razao_social AS empresa_razao_social
            FROM contas_caixa c
            INNER JOIN empresas e ON e.id = c.empresa_id
            WHERE 1 = 1
        """
        params = []
        if not is_super_admin:
            query += " AND c.empresa_id = %s"
            params.append(empresa_id)
        if somente_ativas:
            query += " AND c.status_conta = 'Ativa'"
        query += " ORDER BY e.nome_fantasia ASC, c.nome_conta ASC"
        cur.execute(query, params)
        return cur.fetchall()
    except Exception as e:
        print(f"Erro ao carregar contas caixa: {e}")
        return []
    finally:
        fechar_cursor_conexao(cur, con)





# ==========================================================
# BLOCO 10 - CENTRAL DE RELATÓRIOS GERENCIAIS
# ==========================================================

def _relatorios_datas_request():
    hoje = date.today()
    primeiro_mes = hoje.replace(day=1)
    periodo = (request.args.get('periodo') or 'mes_atual').strip()
    data_inicio = (request.args.get('data_inicio') or '').strip()
    data_fim = (request.args.get('data_fim') or '').strip()

    if periodo == 'hoje':
        di, df = hoje, hoje
    elif periodo == 'mes_atual':
        di, df = primeiro_mes, hoje
    elif periodo == 'ultimos_30':
        di, df = hoje - timedelta(days=30), hoje
    elif periodo == 'personalizado' and validar_data_iso(data_inicio) and validar_data_iso(data_fim):
        di = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        df = datetime.strptime(data_fim, '%Y-%m-%d').date()
    else:
        periodo = 'mes_atual'
        di, df = primeiro_mes, hoje

    if di > df:
        di, df = df, di

    return periodo, di, df


def _relatorios_empresa_filtro(is_super_admin, empresa_logada_id):
    empresa_id_filtro = (request.args.get('empresa_id') or '').strip()
    if is_super_admin and empresa_id_filtro and empresa_id_filtro.isdigit():
        return int(empresa_id_filtro), empresa_id_filtro
    if is_super_admin and not empresa_id_filtro:
        return None, ''
    return int(empresa_logada_id), str(empresa_logada_id)


def _relatorios_financeiro_montar_filtros(empresa_consulta_id=None):
    periodo, data_inicio_dt, data_fim_dt = _relatorios_datas_request()
    tipo_titulo = (request.args.get('tipo_titulo') or '').strip().upper()
    status_titulo = (request.args.get('status_titulo') or '').strip()
    origem = (request.args.get('origem') or '').strip()
    referencia_data = (request.args.get('referencia_data') or 'vencimento').strip()
    pesquisa = (request.args.get('pesquisa') or '').strip()

    referencias = {
        'vencimento': 't.data_vencimento',
        'emissao': 't.data_emissao',
        'competencia': 't.data_competencia',
        'baixa': 't.data_baixa',
        'criacao': 'DATE(t.created_at)',
    }
    data_expr = referencias.get(referencia_data, 't.data_vencimento')
    if referencia_data not in referencias:
        referencia_data = 'vencimento'

    where = []
    params = []

    if empresa_consulta_id:
        where.append('t.empresa_id = %s')
        params.append(empresa_consulta_id)
    else:
        where.append('t.empresa_id IS NOT NULL')

    where.append(f'{data_expr} IS NOT NULL')
    where.append(f'{data_expr} BETWEEN %s AND %s')
    params.extend([data_inicio_dt, data_fim_dt])

    if tipo_titulo in ['PAGAR', 'RECEBER']:
        where.append('t.tipo_titulo = %s')
        params.append(tipo_titulo)

    if status_titulo:
        where.append('t.status_titulo = %s')
        params.append(status_titulo)

    if origem:
        where.append('t.origem = %s')
        params.append(origem)

    if pesquisa:
        like = f'%{pesquisa}%'
        where.append("""
            (
                CAST(t.id AS CHAR) LIKE %s OR t.numero_documento LIKE %s OR t.descricao LIKE %s
                OR t.historico LIKE %s OR t.observacao LIKE %s
                OR p.nome_completo LIKE %s OR p.cpf_cnpj LIKE %s OR e.nome_fantasia LIKE %s
            )
        """)
        params.extend([like, like, like, like, like, like, like, like])

    filtros = {
        'periodo': periodo,
        'data_inicio': data_inicio_dt.strftime('%Y-%m-%d'),
        'data_fim': data_fim_dt.strftime('%Y-%m-%d'),
        'tipo_titulo': tipo_titulo,
        'status_titulo': status_titulo,
        'origem': origem,
        'referencia_data': referencia_data,
        'pesquisa': pesquisa,
    }
    return ' AND '.join(where), params, filtros


@app.route('/relatorios')
@login_required
@perfis_permitidos('Administrador', 'Financeiro', 'Operacional', 'Consulta')
def relatorios_central():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash('Empresa não identificada na sessão. Faça login novamente.', 'danger')
        return redirect(url_for('logout'))

    con = obter_conexao()
    if con is None:
        flash('Erro de conexão com o banco de dados.', 'danger')
        return redirect(url_for('dashboard'))

    cur = con.cursor(dictionary=True)
    try:
        empresa_consulta_id, empresa_id_filtro = _relatorios_empresa_filtro(is_super_admin, empresa_logada_id)
        filtro_empresa_sql = 'empresa_id = %s' if empresa_consulta_id else 'empresa_id IS NOT NULL'
        filtro_empresa_params = [empresa_consulta_id] if empresa_consulta_id else []

        cards = {}
        cur.execute(f"""
            SELECT
                COUNT(*) AS qtd_titulos,
                COALESCE(SUM(valor_liquido), 0) AS valor_titulos,
                SUM(CASE WHEN status_titulo IN ('Aberto','Solicitado') THEN 1 ELSE 0 END) AS qtd_abertos
            FROM titulos_financeiros
            WHERE {filtro_empresa_sql}
        """, filtro_empresa_params)
        cards['financeiro'] = cur.fetchone() or {}

        cur.execute(f"""
            SELECT
                COUNT(*) AS qtd_rotas,
                COALESCE(SUM(
                    COALESCE(valor_rota, 0) +
                    COALESCE(valor_km, 0) +
                    COALESCE(outras_despesas, 0)
                ), 0) AS valor_rotas,
                SUM(CASE WHEN status_motorista IN ('Aguardando conferência','Divergência apontada','Liberada para NF','NF enviada','Em análise','Aprovada para pagamento') THEN 1 ELSE 0 END) AS qtd_pendentes
            FROM rotas
            WHERE {filtro_empresa_sql}
        """, filtro_empresa_params)
        cards['operacao'] = cur.fetchone() or {}

        cur.execute(f"""
            SELECT COUNT(*) AS qtd_arquivos
            FROM arquivos_sistema
            WHERE {filtro_empresa_sql}
        """, filtro_empresa_params)
        cards['arquivos'] = cur.fetchone() or {}

        cur.execute(f"""
            SELECT COUNT(*) AS qtd_auditorias
            FROM auditoria_financeira
            WHERE {filtro_empresa_sql}
        """, filtro_empresa_params)
        cards['auditoria'] = cur.fetchone() or {}

        empresas = []
        if is_super_admin:
            cur.execute('SELECT id, nome_fantasia, razao_social FROM empresas ORDER BY nome_fantasia ASC')
            empresas = cur.fetchall()

        return render_template(
            'relatorios_central.html',
            usuario_logado=usuario_logado,
            cards=cards,
            empresas=empresas,
            is_super_admin=is_super_admin,
            empresa_id_filtro=empresa_id_filtro,
        )
    except Exception as e:
        print(f"Erro ao carregar central de relatórios: {e}")
        flash(f"Erro técnico ao carregar central de relatórios: {e}", 'danger')
        return redirect(url_for('dashboard'))
    finally:
        fechar_cursor_conexao(cur, con)


@app.route('/relatorios/financeiro')
@login_required
@perfis_permitidos('Administrador', 'Financeiro', 'Operacional', 'Consulta')
def relatorios_financeiro():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash('Empresa não identificada na sessão. Faça login novamente.', 'danger')
        return redirect(url_for('logout'))

    con = obter_conexao()
    if con is None:
        flash('Erro de conexão com o banco de dados.', 'danger')
        return redirect(url_for('relatorios_central'))

    cur = con.cursor(dictionary=True)
    try:
        empresa_consulta_id, empresa_id_filtro = _relatorios_empresa_filtro(is_super_admin, empresa_logada_id)
        where_sql, params, filtros = _relatorios_financeiro_montar_filtros(empresa_consulta_id)
        filtros['empresa_id'] = empresa_id_filtro
        session['relatorios_financeiro_filtros'] = {k: v for k, v in filtros.items() if v not in [None, '']}

        cur.execute(f"""
            SELECT
                COUNT(*) AS total_registros,
                COALESCE(SUM(t.valor_liquido), 0) AS total_liquido,
                COALESCE(SUM(CASE WHEN t.tipo_titulo = 'PAGAR' THEN t.valor_liquido ELSE 0 END), 0) AS total_pagar,
                COALESCE(SUM(CASE WHEN t.tipo_titulo = 'RECEBER' THEN t.valor_liquido ELSE 0 END), 0) AS total_receber,
                COALESCE(SUM(CASE WHEN t.status_titulo IN ('Aberto','Solicitado') THEN t.valor_liquido ELSE 0 END), 0) AS total_aberto,
                COALESCE(SUM(CASE WHEN t.status_titulo IN ('Pago','Recebido') THEN COALESCE(t.valor_baixado, t.valor_liquido) ELSE 0 END), 0) AS total_baixado,
                COALESCE(SUM(CASE WHEN t.status_titulo = 'Estornado' THEN t.valor_liquido ELSE 0 END), 0) AS total_estornado,
                COALESCE(SUM(CASE WHEN t.status_titulo = 'Cancelado' THEN t.valor_liquido ELSE 0 END), 0) AS total_cancelado,
                COALESCE(SUM(CASE WHEN t.status_titulo NOT IN ('Pago','Recebido','Cancelado','Estornado') AND t.data_vencimento < CURDATE() THEN t.valor_liquido ELSE 0 END), 0) AS total_vencido,
                SUM(CASE WHEN t.status_titulo NOT IN ('Pago','Recebido','Cancelado','Estornado') AND t.data_vencimento < CURDATE() THEN 1 ELSE 0 END) AS qtd_vencidos
            FROM titulos_financeiros t
            LEFT JOIN pessoas p ON p.id = t.pessoa_id
            LEFT JOIN empresas e ON e.id = t.empresa_id
            WHERE {where_sql}
        """, params)
        resumo = cur.fetchone() or {}

        cur.execute(f"""
            SELECT
                COALESCE(AVG(t.valor_liquido), 0) AS ticket_medio,
                COALESCE(MAX(CASE WHEN t.tipo_titulo = 'PAGAR' THEN t.valor_liquido ELSE NULL END), 0) AS maior_pagamento,
                COALESCE(MAX(CASE WHEN t.tipo_titulo = 'RECEBER' THEN t.valor_liquido ELSE NULL END), 0) AS maior_recebimento,
                COALESCE(AVG(CASE WHEN t.data_baixa IS NOT NULL AND t.data_emissao IS NOT NULL THEN DATEDIFF(t.data_baixa, t.data_emissao) ELSE NULL END), 0) AS tempo_medio_baixa,
                COALESCE(SUM(CASE WHEN t.tipo_titulo = 'RECEBER' THEN t.valor_liquido ELSE -t.valor_liquido END), 0) AS saldo_periodo
            FROM titulos_financeiros t
            LEFT JOIN pessoas p ON p.id = t.pessoa_id
            LEFT JOIN empresas e ON e.id = t.empresa_id
            WHERE {where_sql}
        """, params)
        kpis_exec = cur.fetchone() or {}

        cur.execute(f"""
            SELECT
                t.id, t.empresa_id, e.nome_fantasia AS empresa_nome, t.tipo_titulo, t.origem,
                t.numero_documento, t.descricao, t.valor_liquido, t.valor_baixado,
                t.valor_original, t.valor_desconto, t.valor_acrescimo, t.comprovante_url,
                t.data_emissao, t.data_competencia, t.data_vencimento, t.data_baixa,
                t.status_titulo, t.forma_pagamento, t.created_at,
                p.nome_completo AS pessoa_nome, p.cpf_cnpj,
                cc.nome_conta AS conta_baixa
            FROM titulos_financeiros t
            LEFT JOIN pessoas p ON p.id = t.pessoa_id
            LEFT JOIN empresas e ON e.id = t.empresa_id
            LEFT JOIN contas_caixa cc ON cc.id = t.conta_caixa_baixa_id
            WHERE {where_sql}
            ORDER BY t.data_vencimento ASC, t.id DESC
            LIMIT 500
        """, params)
        titulos = cur.fetchall()

        cur.execute(f"""
            SELECT t.status_titulo, COUNT(*) AS quantidade, COALESCE(SUM(t.valor_liquido), 0) AS valor
            FROM titulos_financeiros t
            LEFT JOIN pessoas p ON p.id = t.pessoa_id
            LEFT JOIN empresas e ON e.id = t.empresa_id
            WHERE {where_sql}
            GROUP BY t.status_titulo
            ORDER BY quantidade DESC
        """, params)
        por_status = cur.fetchall()

        cur.execute(f"""
            SELECT t.origem, COUNT(*) AS quantidade, COALESCE(SUM(t.valor_liquido), 0) AS valor
            FROM titulos_financeiros t
            LEFT JOIN pessoas p ON p.id = t.pessoa_id
            LEFT JOIN empresas e ON e.id = t.empresa_id
            WHERE {where_sql}
            GROUP BY t.origem
            ORDER BY valor DESC
        """, params)
        por_origem = cur.fetchall()

        cur.execute(f"""
            SELECT p.nome_completo AS pessoa_nome, p.cpf_cnpj, COUNT(*) AS quantidade, COALESCE(SUM(t.valor_liquido), 0) AS valor
            FROM titulos_financeiros t
            LEFT JOIN pessoas p ON p.id = t.pessoa_id
            LEFT JOIN empresas e ON e.id = t.empresa_id
            WHERE {where_sql}
            GROUP BY p.id, p.nome_completo, p.cpf_cnpj
            ORDER BY valor DESC
            LIMIT 10
        """, params)
        ranking_pessoas = cur.fetchall()

        cur.execute("SELECT DISTINCT status_titulo FROM titulos_financeiros WHERE status_titulo IS NOT NULL AND status_titulo <> '' ORDER BY status_titulo ASC")
        status_opcoes = [r['status_titulo'] for r in cur.fetchall()]
        cur.execute("SELECT DISTINCT origem FROM titulos_financeiros WHERE origem IS NOT NULL AND origem <> '' ORDER BY origem ASC")
        origem_opcoes = [r['origem'] for r in cur.fetchall()]

        empresas = []
        if is_super_admin:
            cur.execute('SELECT id, nome_fantasia, razao_social FROM empresas ORDER BY nome_fantasia ASC')
            empresas = cur.fetchall()

        return render_template(
            'relatorios_financeiro.html',
            usuario_logado=usuario_logado,
            filtros=filtros,
            resumo=resumo,
            kpis_exec=kpis_exec,
            titulos=titulos,
            por_status=por_status,
            por_origem=por_origem,
            ranking_pessoas=ranking_pessoas,
            status_opcoes=status_opcoes,
            origem_opcoes=origem_opcoes,
            empresas=empresas,
            is_super_admin=is_super_admin,
        )
    except Exception as e:
        print(f"Erro ao carregar relatório financeiro: {e}")
        flash(f"Erro técnico ao carregar relatório financeiro: {e}", 'danger')
        return redirect(url_for('relatorios_central'))
    finally:
        fechar_cursor_conexao(cur, con)


def _relatorios_financeiro_buscar_exportacao():
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()
    empresa_consulta_id, empresa_id_filtro = _relatorios_empresa_filtro(is_super_admin, empresa_logada_id)
    where_sql, params, filtros = _relatorios_financeiro_montar_filtros(empresa_consulta_id)
    filtros['empresa_id'] = empresa_id_filtro

    con = obter_conexao()
    if con is None:
        raise RuntimeError('Erro de conexão com o banco de dados.')
    cur = con.cursor(dictionary=True)
    try:
        cur.execute(f"""
            SELECT
                t.id, e.nome_fantasia AS empresa, t.tipo_titulo, t.origem, t.numero_documento,
                p.nome_completo AS pessoa, p.cpf_cnpj, t.descricao, t.valor_original,
                t.valor_desconto, t.valor_acrescimo, t.valor_liquido, t.valor_baixado,
                t.data_emissao, t.data_competencia, t.data_vencimento, t.data_baixa,
                t.status_titulo, t.forma_pagamento, cc.nome_conta AS conta_baixa,
                t.observacao, t.motivo_cancelamento, t.motivo_estorno, t.created_at
            FROM titulos_financeiros t
            LEFT JOIN pessoas p ON p.id = t.pessoa_id
            LEFT JOIN empresas e ON e.id = t.empresa_id
            LEFT JOIN contas_caixa cc ON cc.id = t.conta_caixa_baixa_id
            WHERE {where_sql}
            ORDER BY t.data_vencimento ASC, t.id DESC
            LIMIT 5000
        """, params)
        return cur.fetchall(), filtros
    finally:
        fechar_cursor_conexao(cur, con)


@app.route('/relatorios/financeiro/exportar-csv')
@login_required
@perfis_permitidos('Administrador', 'Financeiro', 'Operacional', 'Consulta')
def relatorios_financeiro_exportar_csv():
    try:
        linhas, filtros = _relatorios_financeiro_buscar_exportacao()
        output = StringIO()
        writer = csv.writer(output, delimiter=';')
        writer.writerow(['ID', 'Empresa', 'Tipo', 'Origem', 'Documento', 'Pessoa', 'CPF/CNPJ', 'Descrição', 'Valor Original', 'Desconto', 'Acréscimo', 'Valor Líquido', 'Valor Baixado', 'Emissão', 'Competência', 'Vencimento', 'Baixa', 'Status', 'Forma Pgto', 'Conta Baixa', 'Observação', 'Motivo Cancelamento', 'Motivo Estorno', 'Criado em'])
        for r in linhas:
            writer.writerow([
                r.get('id'), r.get('empresa'), r.get('tipo_titulo'), r.get('origem'), r.get('numero_documento'),
                r.get('pessoa'), r.get('cpf_cnpj'), r.get('descricao'), r.get('valor_original'), r.get('valor_desconto'),
                r.get('valor_acrescimo'), r.get('valor_liquido'), r.get('valor_baixado'), r.get('data_emissao'),
                r.get('data_competencia'), r.get('data_vencimento'), r.get('data_baixa'), r.get('status_titulo'),
                r.get('forma_pagamento'), r.get('conta_baixa'), r.get('observacao'), r.get('motivo_cancelamento'),
                r.get('motivo_estorno'), r.get('created_at')
            ])
        nome = f"relatorio_financeiro_{filtros.get('data_inicio')}_a_{filtros.get('data_fim')}.csv"
        return Response(
            output.getvalue().encode('utf-8-sig'),
            mimetype='text/csv; charset=utf-8',
            headers={'Content-Disposition': f'attachment; filename={nome}'}
        )
    except Exception as e:
        print(f"Erro ao exportar relatório financeiro CSV: {e}")
        flash(f"Erro técnico ao exportar CSV: {e}", 'danger')
        return redirect(url_for('relatorios_financeiro'))


@app.route('/relatorios/financeiro/exportar-excel')
@login_required
@perfis_permitidos('Administrador', 'Financeiro', 'Operacional', 'Consulta')
def relatorios_financeiro_exportar_excel():
    try:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
        except ImportError:
            flash('Biblioteca openpyxl não instalada. Rode: pip install openpyxl', 'danger')
            return redirect(url_for('relatorios_financeiro'))

        linhas, filtros = _relatorios_financeiro_buscar_exportacao()
        wb = Workbook()
        ws = wb.active
        ws.title = 'Relatório Financeiro'

        ws['A1'] = 'SGR Web - Relatório Financeiro Gerencial'
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Período: {filtros.get('data_inicio')} até {filtros.get('data_fim')} | Referência: {filtros.get('referencia_data')}"
        ws.merge_cells('A1:X1')
        ws.merge_cells('A2:X2')

        headers = ['ID', 'Empresa', 'Tipo', 'Origem', 'Documento', 'Pessoa', 'CPF/CNPJ', 'Descrição', 'Valor Original', 'Desconto', 'Acréscimo', 'Valor Líquido', 'Valor Baixado', 'Emissão', 'Competência', 'Vencimento', 'Baixa', 'Status', 'Forma Pgto', 'Conta Baixa', 'Observação', 'Motivo Cancelamento', 'Motivo Estorno', 'Criado em']
        ws.append([])
        ws.append(headers)
        header_row = 4
        for cell in ws[header_row]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='1F2937')
            cell.alignment = Alignment(horizontal='center')

        for r in linhas:
            ws.append([
                r.get('id'), r.get('empresa'), r.get('tipo_titulo'), r.get('origem'), r.get('numero_documento'),
                r.get('pessoa'), r.get('cpf_cnpj'), r.get('descricao'), float(r.get('valor_original') or 0),
                float(r.get('valor_desconto') or 0), float(r.get('valor_acrescimo') or 0), float(r.get('valor_liquido') or 0),
                float(r.get('valor_baixado') or 0), r.get('data_emissao'), r.get('data_competencia'), r.get('data_vencimento'),
                r.get('data_baixa'), r.get('status_titulo'), r.get('forma_pagamento'), r.get('conta_baixa'),
                r.get('observacao'), r.get('motivo_cancelamento'), r.get('motivo_estorno'), r.get('created_at')
            ])

        for col in range(1, len(headers) + 1):
            letter = get_column_letter(col)
            ws.column_dimensions[letter].width = 18
        ws.column_dimensions['H'].width = 42
        ws.column_dimensions['F'].width = 32
        ws.column_dimensions['U'].width = 42
        ws.freeze_panes = 'A5'

        for row in ws.iter_rows(min_row=5, min_col=9, max_col=13):
            for cell in row:
                cell.number_format = 'R$ #,##0.00'

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        nome = f"relatorio_financeiro_{filtros.get('data_inicio')}_a_{filtros.get('data_fim')}.xlsx"
        return Response(
            bio.getvalue(),
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': f'attachment; filename={nome}'}
        )
    except Exception as e:
        print(f"Erro ao exportar relatório financeiro Excel: {e}")
        flash(f"Erro técnico ao exportar Excel: {e}", 'danger')
        return redirect(url_for('relatorios_financeiro'))


# ==========================================================
# BLOCO 7 - DASHBOARD FINANCEIRO GERENCIAL
# ==========================================================

@app.route('/financeiro/auditoria')
@login_required
@perfis_permitidos('Administrador', 'Financeiro')
def financeiro_auditoria():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash('Empresa não identificada na sessão. Faça login novamente.', 'danger')
        return redirect(url_for('logout'))

    hoje = date.today()
    data_inicio = (request.args.get('data_inicio') or (hoje - timedelta(days=30)).strftime('%Y-%m-%d')).strip()
    data_fim = (request.args.get('data_fim') or hoje.strftime('%Y-%m-%d')).strip()
    acao = (request.args.get('acao') or '').strip()
    modulo = (request.args.get('modulo') or '').strip()
    entidade_tipo = (request.args.get('entidade_tipo') or '').strip()
    usuario_id_filtro = (request.args.get('usuario_id') or '').strip()
    pesquisa = (request.args.get('pesquisa') or '').strip()
    empresa_id_filtro = (request.args.get('empresa_id') or '').strip() if is_super_admin else str(empresa_logada_id)

    if not validar_data_iso(data_inicio):
        data_inicio = (hoje - timedelta(days=30)).strftime('%Y-%m-%d')
    if not validar_data_iso(data_fim):
        data_fim = hoje.strftime('%Y-%m-%d')

    con = obter_conexao()
    if con is None:
        flash('Erro de conexão com o banco de dados.', 'danger')
        return redirect(url_for('financeiro_dashboard'))

    cur = con.cursor(dictionary=True)
    try:
        where = ["a.created_at >= %s", "a.created_at < DATE_ADD(%s, INTERVAL 1 DAY)"]
        params = [data_inicio, data_fim]

        if is_super_admin:
            if empresa_id_filtro and str(empresa_id_filtro).isdigit():
                where.append("a.empresa_id = %s")
                params.append(int(empresa_id_filtro))
        else:
            where.append("a.empresa_id = %s")
            params.append(int(empresa_logada_id))

        if acao:
            where.append("a.acao = %s")
            params.append(acao)
        if modulo:
            where.append("a.modulo = %s")
            params.append(modulo)
        if entidade_tipo:
            where.append("a.entidade_tipo = %s")
            params.append(entidade_tipo)
        if usuario_id_filtro and usuario_id_filtro.isdigit():
            where.append("a.usuario_id = %s")
            params.append(int(usuario_id_filtro))
        if pesquisa:
            like = f"%{pesquisa}%"
            where.append("""
                (
                    a.motivo LIKE %s OR a.observacao LIKE %s OR a.acao LIKE %s
                    OR a.entidade_tipo LIKE %s OR CAST(a.entidade_id AS CHAR) LIKE %s
                    OR CAST(a.titulo_financeiro_id AS CHAR) LIKE %s
                    OR p.nome_completo LIKE %s OR u.login LIKE %s
                )
            """)
            params.extend([like, like, like, like, like, like, like, like])

        where_sql = ' AND '.join(where)
        cur.execute(f"""
            SELECT
                a.*,
                e.nome_fantasia AS empresa_nome,
                u.login AS usuario_login,
                pu.nome_completo AS usuario_nome,
                p.nome_completo AS pessoa_nome
            FROM auditoria_financeira a
            LEFT JOIN empresas e ON e.id = a.empresa_id
            LEFT JOIN usuarios u ON u.id = a.usuario_id
            LEFT JOIN pessoas pu ON pu.id = u.pessoa_id
            LEFT JOIN pessoas p ON p.id = a.pessoa_id
            WHERE {where_sql}
            ORDER BY a.created_at DESC, a.id DESC
            LIMIT 500
        """, params)
        auditorias = cur.fetchall()

        cur.execute(f"""
            SELECT
                COUNT(*) AS total,
                COUNT(DISTINCT a.usuario_id) AS usuarios_distintos,
                SUM(CASE WHEN a.acao LIKE 'BAIXA%%' THEN 1 ELSE 0 END) AS baixas,
                SUM(CASE WHEN a.acao LIKE 'ESTORNO%%' THEN 1 ELSE 0 END) AS estornos,
                SUM(CASE WHEN a.acao LIKE 'CONCILIACAO%%' THEN 1 ELSE 0 END) AS conciliacoes,
                SUM(CASE WHEN a.acao LIKE 'CONFIGURACAO%%' THEN 1 ELSE 0 END) AS configuracoes
            FROM auditoria_financeira a
            LEFT JOIN usuarios u ON u.id = a.usuario_id
            LEFT JOIN pessoas p ON p.id = a.pessoa_id
            WHERE {where_sql}
        """, params)
        resumo = cur.fetchone() or {}

        cur.execute("SELECT DISTINCT acao FROM auditoria_financeira WHERE acao IS NOT NULL AND acao <> '' ORDER BY acao ASC")
        acoes = [r.get('acao') for r in cur.fetchall()]
        cur.execute("SELECT DISTINCT modulo FROM auditoria_financeira WHERE modulo IS NOT NULL AND modulo <> '' ORDER BY modulo ASC")
        modulos = [r.get('modulo') for r in cur.fetchall()]
        cur.execute("SELECT DISTINCT entidade_tipo FROM auditoria_financeira WHERE entidade_tipo IS NOT NULL AND entidade_tipo <> '' ORDER BY entidade_tipo ASC")
        entidades = [r.get('entidade_tipo') for r in cur.fetchall()]
        cur.execute("""
            SELECT
                u.id,
                u.login,
                COALESCE(p.nome_completo, u.login) AS nome
            FROM usuarios u
            LEFT JOIN pessoas p ON p.id = u.pessoa_id
            WHERE u.status_usuario = 'Ativo'
            ORDER BY COALESCE(p.nome_completo, u.login), u.login ASC
        """)
        usuarios = cur.fetchall()

        empresas = []
        if is_super_admin:
            cur.execute("SELECT id, nome_fantasia, razao_social FROM empresas ORDER BY nome_fantasia ASC")
            empresas = cur.fetchall()

        filtros = {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'acao': acao,
            'modulo': modulo,
            'entidade_tipo': entidade_tipo,
            'usuario_id': usuario_id_filtro,
            'pesquisa': pesquisa,
            'empresa_id': empresa_id_filtro,
        }

        return render_template(
            'financeiro_auditoria.html',
            usuario_logado=usuario_logado,
            auditorias=auditorias,
            resumo=resumo,
            filtros=filtros,
            acoes=acoes,
            modulos=modulos,
            entidades=entidades,
            usuarios=usuarios,
            empresas=empresas,
            is_super_admin=is_super_admin,
        )
    except Exception as e:
        print(f"Erro ao carregar auditoria financeira: {e}")
        flash(f"Erro técnico ao carregar auditoria financeira: {e}", 'danger')
        return redirect(url_for('financeiro_dashboard'))
    finally:
        fechar_cursor_conexao(cur, con)


@app.route('/financeiro/dashboard')
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro', 'Consulta')
def financeiro_dashboard():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash('Empresa não identificada na sessão. Faça login novamente.', 'danger')
        return redirect(url_for('logout'))

    hoje = date.today()
    primeiro_mes = hoje.replace(day=1)
    primeiro_mes_anterior = (primeiro_mes - timedelta(days=1)).replace(day=1)
    ultimo_mes_anterior = primeiro_mes - timedelta(days=1)

    periodo = (request.args.get('periodo') or 'mes_atual').strip()
    data_inicio = (request.args.get('data_inicio') or '').strip()
    data_fim = (request.args.get('data_fim') or '').strip()

    if periodo == 'hoje':
        data_inicio_dt = hoje
        data_fim_dt = hoje
    elif periodo == 'mes_anterior':
        data_inicio_dt = primeiro_mes_anterior
        data_fim_dt = ultimo_mes_anterior
    elif periodo == 'personalizado' and data_inicio and data_fim and validar_data_iso(data_inicio) and validar_data_iso(data_fim):
        data_inicio_dt = datetime.strptime(data_inicio, '%Y-%m-%d').date()
        data_fim_dt = datetime.strptime(data_fim, '%Y-%m-%d').date()
    else:
        periodo = 'mes_atual'
        data_inicio_dt = primeiro_mes
        data_fim_dt = hoje

    if data_inicio_dt > data_fim_dt:
        data_inicio_dt, data_fim_dt = data_fim_dt, data_inicio_dt

    tipo_titulo = (request.args.get('tipo_titulo') or '').strip()
    origem = (request.args.get('origem') or '').strip()
    conta_caixa_id = (request.args.get('conta_caixa_id') or '').strip()
    pessoa_id = (request.args.get('pessoa_id') or '').strip()
    pesquisa_pessoa = (request.args.get('pesquisa_pessoa') or '').strip()
    empresa_id_filtro = (request.args.get('empresa_id') or '').strip()

    empresa_consulta_id = empresa_logada_id
    if is_super_admin and empresa_id_filtro and empresa_id_filtro.isdigit():
        empresa_consulta_id = int(empresa_id_filtro)

    con = obter_conexao()
    if con is None:
        flash('Erro de conexão com o banco de dados.', 'danger')
        return redirect(url_for('financeiro_titulos'))

    cur = con.cursor(dictionary=True)
    try:
        filtro_empresa_params = [empresa_consulta_id]
        if is_super_admin and not (empresa_id_filtro and empresa_id_filtro.isdigit()):
            filtro_empresa_params = []

        filtros_titulos = []
        params_titulos = []
        if filtro_empresa_params:
            filtros_titulos.append('t.empresa_id = %s')
            params_titulos.extend(filtro_empresa_params)
        else:
            filtros_titulos.append('t.empresa_id IS NOT NULL')

        if tipo_titulo in ['PAGAR', 'RECEBER']:
            filtros_titulos.append('t.tipo_titulo = %s')
            params_titulos.append(tipo_titulo)
        if origem in financeiro_base_origens():
            filtros_titulos.append('t.origem = %s')
            params_titulos.append(origem)
        if pessoa_id and pessoa_id.isdigit():
            filtros_titulos.append('t.pessoa_id = %s')
            params_titulos.append(int(pessoa_id))
        if pesquisa_pessoa:
            filtros_titulos.append('(p.nome_completo LIKE %s OR p.cpf_cnpj LIKE %s OR CAST(p.id AS CHAR) LIKE %s)')
            termo = f'%{pesquisa_pessoa}%'
            params_titulos.extend([termo, termo, termo])

        where_titulos = ' AND '.join(filtros_titulos) if filtros_titulos else '1=1'

        cur.execute(f"""
            SELECT
                COALESCE(SUM(CASE WHEN t.tipo_titulo = 'PAGAR'
                    AND COALESCE(t.status_titulo, 'Aberto') NOT IN ('Pago','Recebido','Cancelado','Estornado')
                    THEN t.valor_liquido ELSE 0 END), 0) AS pagar_aberto,
                COALESCE(SUM(CASE WHEN t.tipo_titulo = 'RECEBER'
                    AND COALESCE(t.status_titulo, 'Aberto') NOT IN ('Pago','Recebido','Cancelado','Estornado')
                    THEN t.valor_liquido ELSE 0 END), 0) AS receber_aberto,
                COALESCE(SUM(CASE WHEN COALESCE(t.status_titulo, 'Aberto') NOT IN ('Pago','Recebido','Cancelado','Estornado')
                    AND t.data_vencimento IS NOT NULL AND t.data_vencimento < %s
                    THEN t.valor_liquido ELSE 0 END), 0) AS vencidos,
                COUNT(CASE WHEN COALESCE(t.status_titulo, 'Aberto') NOT IN ('Pago','Recebido','Cancelado','Estornado')
                    AND t.data_vencimento IS NOT NULL AND t.data_vencimento < %s THEN 1 END) AS qtd_vencidos,
                COALESCE(SUM(CASE WHEN t.status_titulo = 'Pago'
                    AND t.data_baixa BETWEEN %s AND %s THEN t.valor_baixado ELSE 0 END), 0) AS pagos_periodo,
                COALESCE(SUM(CASE WHEN t.status_titulo = 'Recebido'
                    AND t.data_baixa BETWEEN %s AND %s THEN t.valor_baixado ELSE 0 END), 0) AS recebidos_periodo,
                COALESCE(SUM(CASE WHEN t.status_titulo = 'Estornado'
                    AND DATE(COALESCE(t.data_estorno, t.updated_at, t.created_at)) BETWEEN %s AND %s THEN t.valor_liquido ELSE 0 END), 0) AS estornados_periodo,
                COALESCE(SUM(CASE WHEN t.status_titulo = 'Cancelado'
                    AND DATE(COALESCE(t.data_cancelamento, t.updated_at, t.created_at)) BETWEEN %s AND %s THEN t.valor_liquido ELSE 0 END), 0) AS cancelados_periodo,
                COUNT(CASE WHEN COALESCE(t.status_titulo, 'Aberto') NOT IN ('Pago','Recebido','Cancelado','Estornado') THEN 1 END) AS qtd_abertos
            FROM titulos_financeiros t
            LEFT JOIN pessoas p ON p.id = t.pessoa_id AND p.empresa_id = t.empresa_id
            WHERE {where_titulos}
        """, [hoje, hoje, data_inicio_dt, data_fim_dt, data_inicio_dt, data_fim_dt, data_inicio_dt, data_fim_dt, data_inicio_dt, data_fim_dt] + params_titulos)
        resumo_titulos = cur.fetchone() or {}

        filtros_mov = []
        params_mov = []
        if filtro_empresa_params:
            filtros_mov.append('m.empresa_id = %s')
            params_mov.extend(filtro_empresa_params)
        else:
            filtros_mov.append('m.empresa_id IS NOT NULL')
        filtros_mov.append('m.data_movimentacao BETWEEN %s AND %s')
        params_mov.extend([data_inicio_dt, data_fim_dt])
        if conta_caixa_id and conta_caixa_id.isdigit():
            filtros_mov.append('m.conta_caixa_id = %s')
            params_mov.append(int(conta_caixa_id))
        if tipo_titulo in ['PAGAR', 'RECEBER']:
            filtros_mov.append('t.tipo_titulo = %s')
            params_mov.append(tipo_titulo)
        if origem in financeiro_base_origens():
            filtros_mov.append('t.origem = %s')
            params_mov.append(origem)
        if pessoa_id and pessoa_id.isdigit():
            filtros_mov.append('t.pessoa_id = %s')
            params_mov.append(int(pessoa_id))
        if pesquisa_pessoa:
            filtros_mov.append('(p.nome_completo LIKE %s OR p.cpf_cnpj LIKE %s OR CAST(p.id AS CHAR) LIKE %s)')
            termo = f'%{pesquisa_pessoa}%'
            params_mov.extend([termo, termo, termo])

        where_mov = ' AND '.join(filtros_mov)
        cur.execute(f"""
            SELECT
                COALESCE(SUM(CASE WHEN m.tipo_movimentacao = 'ENTRADA'
                    AND COALESCE(m.status_movimentacao, 'Ativa') = 'Ativa'
                    AND COALESCE(m.estorno_de_movimentacao_id, 0) = 0
                    THEN m.valor_movimentacao ELSE 0 END), 0) AS entradas_operacionais,
                COALESCE(SUM(CASE WHEN m.tipo_movimentacao = 'SAIDA'
                    AND COALESCE(m.status_movimentacao, 'Ativa') = 'Ativa'
                    AND COALESCE(m.estorno_de_movimentacao_id, 0) = 0
                    THEN m.valor_movimentacao ELSE 0 END), 0) AS saidas_operacionais,
                COALESCE(SUM(CASE WHEN (COALESCE(m.status_movimentacao, '') = 'Estorno'
                    OR COALESCE(m.estorno_de_movimentacao_id, 0) <> 0)
                    THEN m.valor_movimentacao ELSE 0 END), 0) AS estornos,
                COALESCE(SUM(CASE WHEN (COALESCE(m.status_movimentacao, '') = 'Estorno'
                    OR COALESCE(m.estorno_de_movimentacao_id, 0) <> 0)
                    AND m.tipo_movimentacao = 'ENTRADA'
                    THEN m.valor_movimentacao ELSE 0 END), 0) AS estornos_entrada,
                COALESCE(SUM(CASE WHEN (COALESCE(m.status_movimentacao, '') = 'Estorno'
                    OR COALESCE(m.estorno_de_movimentacao_id, 0) <> 0)
                    AND m.tipo_movimentacao = 'SAIDA'
                    THEN m.valor_movimentacao ELSE 0 END), 0) AS estornos_saida,
                COALESCE(SUM(CASE WHEN COALESCE(m.status_movimentacao, '') = 'Estornada'
                    THEN m.valor_movimentacao ELSE 0 END), 0) AS movimentacoes_estornadas,
                COALESCE(SUM(CASE WHEN COALESCE(m.status_movimentacao, 'Ativa') = 'Ativa'
                    AND COALESCE(m.estorno_de_movimentacao_id, 0) = 0
                    THEN CASE WHEN m.tipo_movimentacao = 'ENTRADA' THEN m.valor_movimentacao ELSE -m.valor_movimentacao END
                    ELSE 0 END), 0) AS saldo_operacional,
                COALESCE(SUM(CASE WHEN COALESCE(m.status_movimentacao, 'Ativa') <> 'Estornada'
                    THEN CASE WHEN m.tipo_movimentacao = 'ENTRADA' THEN m.valor_movimentacao ELSE -m.valor_movimentacao END
                    ELSE 0 END), 0) AS saldo_liquido,
                COUNT(*) AS total_movimentacoes
            FROM movimentacoes_caixa m
            LEFT JOIN titulos_financeiros t ON t.id = m.titulo_financeiro_id AND t.empresa_id = m.empresa_id
            LEFT JOIN pessoas p ON p.id = t.pessoa_id AND p.empresa_id = t.empresa_id
            WHERE {where_mov}
        """, params_mov)
        resumo_caixa = cur.fetchone() or {}

        cur.execute(f"""
            SELECT COALESCE(t.origem, 'MANUAL') AS origem,
                   COUNT(*) AS quantidade,
                   COALESCE(SUM(t.valor_liquido), 0) AS total
            FROM titulos_financeiros t
            LEFT JOIN pessoas p ON p.id = t.pessoa_id AND p.empresa_id = t.empresa_id
            WHERE {where_titulos}
              AND DATE(COALESCE(t.data_emissao, t.created_at)) BETWEEN %s AND %s
            GROUP BY COALESCE(t.origem, 'MANUAL')
            ORDER BY total DESC
            LIMIT 8
        """, params_titulos + [data_inicio_dt, data_fim_dt])
        despesas_por_origem = cur.fetchall()

        cur.execute(f"""
            SELECT COALESCE(p.nome_completo, 'Sem pessoa vinculada') AS pessoa_nome,
                   COALESCE(p.cpf_cnpj, '') AS pessoa_cpf_cnpj,
                   COALESCE(p.tipo_cadastro, '') AS pessoa_tipo,
                   COUNT(*) AS quantidade,
                   COALESCE(SUM(t.valor_liquido), 0) AS total
            FROM titulos_financeiros t
            LEFT JOIN pessoas p ON p.id = t.pessoa_id AND p.empresa_id = t.empresa_id
            WHERE {where_titulos}
              AND t.tipo_titulo = 'PAGAR'
              AND COALESCE(t.status_titulo, 'Aberto') IN ('Pago','Estornado','Aberto','Solicitado')
              AND DATE(COALESCE(t.data_baixa, t.data_vencimento, t.data_emissao, t.created_at)) BETWEEN %s AND %s
            GROUP BY p.id, p.nome_completo, p.cpf_cnpj, p.tipo_cadastro
            ORDER BY total DESC
            LIMIT 10
        """, params_titulos + [data_inicio_dt, data_fim_dt])
        ranking_pessoas = cur.fetchall()

        cur.execute(f"""
            SELECT cx.nome_conta,
                   COUNT(*) AS quantidade,
                   COALESCE(SUM(CASE WHEN m.tipo_movimentacao = 'ENTRADA'
                        AND COALESCE(m.status_movimentacao, 'Ativa') = 'Ativa'
                        AND COALESCE(m.estorno_de_movimentacao_id, 0) = 0
                        THEN m.valor_movimentacao ELSE 0 END), 0) AS entradas,
                   COALESCE(SUM(CASE WHEN m.tipo_movimentacao = 'SAIDA'
                        AND COALESCE(m.status_movimentacao, 'Ativa') = 'Ativa'
                        AND COALESCE(m.estorno_de_movimentacao_id, 0) = 0
                        THEN m.valor_movimentacao ELSE 0 END), 0) AS saidas,
                   COALESCE(SUM(CASE WHEN (COALESCE(m.status_movimentacao, '') = 'Estorno'
                        OR COALESCE(m.estorno_de_movimentacao_id, 0) <> 0)
                        THEN m.valor_movimentacao ELSE 0 END), 0) AS estornos,
                   COALESCE(SUM(CASE WHEN COALESCE(m.status_movimentacao, '') = 'Estornada'
                        THEN m.valor_movimentacao ELSE 0 END), 0) AS estornadas,
                   COALESCE(SUM(CASE WHEN COALESCE(m.status_movimentacao, 'Ativa') <> 'Estornada'
                        THEN CASE WHEN m.tipo_movimentacao = 'ENTRADA' THEN m.valor_movimentacao ELSE -m.valor_movimentacao END ELSE 0 END), 0) AS saldo
            FROM movimentacoes_caixa m
            INNER JOIN contas_caixa cx ON cx.id = m.conta_caixa_id AND cx.empresa_id = m.empresa_id
            LEFT JOIN titulos_financeiros t ON t.id = m.titulo_financeiro_id AND t.empresa_id = m.empresa_id
            LEFT JOIN pessoas p ON p.id = t.pessoa_id AND p.empresa_id = t.empresa_id
            WHERE {where_mov}
            GROUP BY cx.id, cx.nome_conta
            ORDER BY ABS(COALESCE(SUM(CASE WHEN COALESCE(m.status_movimentacao, 'Ativa') <> 'Estornada'
                        THEN CASE WHEN m.tipo_movimentacao = 'ENTRADA' THEN m.valor_movimentacao ELSE -m.valor_movimentacao END ELSE 0 END), 0)) DESC,
                     quantidade DESC
            LIMIT 8
        """, params_mov)
        movimentacoes_por_conta = cur.fetchall()

        cur.execute(f"""
            SELECT t.id, t.tipo_titulo, t.numero_documento, t.descricao, t.valor_liquido,
                   t.data_vencimento, t.status_titulo, p.nome_completo AS pessoa_nome
            FROM titulos_financeiros t
            LEFT JOIN pessoas p ON p.id = t.pessoa_id AND p.empresa_id = t.empresa_id
            WHERE {where_titulos}
              AND COALESCE(t.status_titulo, 'Aberto') NOT IN ('Pago','Recebido','Cancelado','Estornado')
              AND t.data_vencimento BETWEEN %s AND %s
            ORDER BY t.data_vencimento ASC, t.valor_liquido DESC
            LIMIT 10
        """, params_titulos + [hoje, hoje + timedelta(days=7)])
        titulos_vencendo = cur.fetchall()

        cur.execute(f"""
            SELECT m.id, m.data_movimentacao, m.tipo_movimentacao, m.valor_movimentacao,
                   m.status_movimentacao, m.estorno_de_movimentacao_id, m.historico,
                   cx.nome_conta AS conta_caixa_nome, t.id AS titulo_id,
                   t.numero_documento, t.tipo_titulo, p.nome_completo AS pessoa_nome
            FROM movimentacoes_caixa m
            INNER JOIN contas_caixa cx ON cx.id = m.conta_caixa_id AND cx.empresa_id = m.empresa_id
            LEFT JOIN titulos_financeiros t ON t.id = m.titulo_financeiro_id AND t.empresa_id = m.empresa_id
            LEFT JOIN pessoas p ON p.id = t.pessoa_id AND p.empresa_id = t.empresa_id
            WHERE {where_mov}
            ORDER BY m.data_movimentacao DESC, m.id DESC
            LIMIT 10
        """, params_mov)
        ultimas_movimentacoes = cur.fetchall()

        cur.execute(f"""
            SELECT COALESCE(t.status_titulo, 'Aberto') AS status_titulo,
                   COUNT(*) AS quantidade,
                   COALESCE(SUM(t.valor_liquido), 0) AS total
            FROM titulos_financeiros t
            LEFT JOIN pessoas p ON p.id = t.pessoa_id AND p.empresa_id = t.empresa_id
            WHERE {where_titulos}
              AND DATE(COALESCE(t.data_emissao, t.created_at)) BETWEEN %s AND %s
            GROUP BY COALESCE(t.status_titulo, 'Aberto')
            ORDER BY quantidade DESC
        """, params_titulos + [data_inicio_dt, data_fim_dt])
        resumo_status = cur.fetchall()

        contas_caixa = carregar_contas_caixa_financeiro(empresa_logada_id, is_super_admin)
        empresas = []
        if is_super_admin:
            cur.execute('SELECT id, razao_social, nome_fantasia FROM empresas ORDER BY nome_fantasia ASC, razao_social ASC')
            empresas = cur.fetchall()

        filtros = {
            'periodo': periodo,
            'data_inicio': data_inicio_dt.strftime('%Y-%m-%d'),
            'data_fim': data_fim_dt.strftime('%Y-%m-%d'),
            'tipo_titulo': tipo_titulo,
            'origem': origem,
            'conta_caixa_id': conta_caixa_id,
            'pessoa_id': pessoa_id,
            'pesquisa_pessoa': pesquisa_pessoa,
            'empresa_id': empresa_id_filtro,
        }

        return render_template(
            'financeiro_dashboard.html',
            usuario_logado=usuario_logado,
            filtros=filtros,
            resumo_titulos=resumo_titulos,
            resumo_caixa=resumo_caixa,
            despesas_por_origem=despesas_por_origem,
            ranking_pessoas=ranking_pessoas,
            movimentacoes_por_conta=movimentacoes_por_conta,
            titulos_vencendo=titulos_vencendo,
            ultimas_movimentacoes=ultimas_movimentacoes,
            resumo_status=resumo_status,
            contas_caixa=contas_caixa,
            empresas=empresas,
            origens=financeiro_base_origens(),
            is_super_admin=is_super_admin,
            hoje=hoje,
        )
    except Exception as e:
        print(f'Erro ao carregar dashboard financeiro: {e}')
        flash(f'Erro técnico ao carregar dashboard financeiro: {e}', 'danger')
        return redirect(url_for('financeiro_titulos'))
    finally:
        fechar_cursor_conexao(cur, con)


@app.route('/financeiro/titulos', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro', 'Consulta')
def financeiro_titulos():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    tipo_titulo = (request.args.get('tipo_titulo') or '').strip()
    status_titulo = (request.args.get('status_titulo') or '').strip()
    origem = (request.args.get('origem') or '').strip()
    pessoa_id = (request.args.get('pessoa_id') or '').strip()
    data_inicio = (request.args.get('data_inicio') or '').strip()
    data_fim = (request.args.get('data_fim') or '').strip()
    vencimento_inicio = (request.args.get('vencimento_inicio') or '').strip()
    vencimento_fim = (request.args.get('vencimento_fim') or '').strip()
    pesquisa = (request.args.get('pesquisa') or '').strip()
    empresa_id_filtro = (request.args.get('empresa_id') or '').strip()

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('dashboard'))

    cur = con.cursor(dictionary=True)

    try:
        query = """
            SELECT t.id,
                   t.empresa_id,
                   t.tipo_titulo,
                   t.origem,
                   t.origem_id,
                   t.pessoa_id,
                   t.numero_documento,
                   t.descricao,
                   t.historico,
                   t.valor_original,
                   t.valor_desconto,
                   t.valor_acrescimo,
                   t.valor_liquido,
                   t.data_emissao,
                   t.data_competencia,
                   t.data_vencimento,
                   t.forma_pagamento,
                   t.status_titulo,
                   t.observacao,
                   t.created_at,
                   p.nome_completo AS pessoa_nome,
                   p.cpf_cnpj AS pessoa_cpf_cnpj,
                   p.tipo_cadastro AS pessoa_tipo,
                   cx.nome_conta AS conta_caixa_nome,
                   e.nome_fantasia AS empresa_nome,
                   e.razao_social AS empresa_razao_social
            FROM titulos_financeiros t
            LEFT JOIN pessoas p ON p.id = t.pessoa_id AND p.empresa_id = t.empresa_id
            LEFT JOIN contas_caixa cx ON cx.id = t.conta_caixa_prevista_id AND cx.empresa_id = t.empresa_id
            INNER JOIN empresas e ON e.id = t.empresa_id
            WHERE 1 = 1
        """
        params = []

        if is_super_admin:
            if empresa_id_filtro and empresa_id_filtro.isdigit():
                query += " AND t.empresa_id = %s"
                params.append(int(empresa_id_filtro))
        else:
            query += " AND t.empresa_id = %s"
            params.append(empresa_logada_id)

        if tipo_titulo in ['PAGAR', 'RECEBER']:
            query += " AND t.tipo_titulo = %s"
            params.append(tipo_titulo)

        if status_titulo in financeiro_base_status_titulos():
            query += " AND t.status_titulo = %s"
            params.append(status_titulo)

        if origem in financeiro_base_origens():
            query += " AND t.origem = %s"
            params.append(origem)

        if pessoa_id and pessoa_id.isdigit():
            query += " AND t.pessoa_id = %s"
            params.append(int(pessoa_id))

        if data_inicio:
            query += " AND t.data_emissao >= %s"
            params.append(data_inicio)

        if data_fim:
            query += " AND t.data_emissao <= %s"
            params.append(data_fim)

        if vencimento_inicio:
            query += " AND t.data_vencimento >= %s"
            params.append(vencimento_inicio)

        if vencimento_fim:
            query += " AND t.data_vencimento <= %s"
            params.append(vencimento_fim)

        if pesquisa:
            query += """
                AND (
                    t.numero_documento LIKE %s
                    OR t.descricao LIKE %s
                    OR t.historico LIKE %s
                    OR p.nome_completo LIKE %s
                    OR p.cpf_cnpj LIKE %s
                )
            """
            termo = f"%{pesquisa}%"
            params.extend([termo, termo, termo, termo, termo])

        query += " ORDER BY t.data_vencimento ASC, t.id DESC"
        cur.execute(query, params)
        titulos = cur.fetchall()

        resumo = {
            'pagar_aberto': Decimal('0.00'),
            'receber_aberto': Decimal('0.00'),
            'vencidos': Decimal('0.00'),
            'qtd_vencidos': 0,
            'pagos_recebidos_mes': Decimal('0.00'),
            # Total filtrado mantém todos os registros da listagem, inclusive cancelados,
            # para preservar a visão histórica do filtro aplicado.
            'total_titulos': len(titulos),
            # Indicadores gerenciais: somente títulos que ainda exigem ação.
            'qtd_abertos': 0,
            'qtd_cancelados': 0,
            'qtd_baixados': 0
        }

        hoje = date.today()
        mes_atual = hoje.strftime('%Y-%m')

        for titulo in titulos:
            status = titulo.get('status_titulo') or 'Aberto'
            tipo = titulo.get('tipo_titulo') or ''
            valor = converter_decimal(titulo.get('valor_liquido'))
            vencimento = titulo.get('data_vencimento')

            if status not in ['Pago', 'Recebido', 'Cancelado', 'Estornado']:
                resumo['qtd_abertos'] += 1

                if tipo == 'PAGAR':
                    resumo['pagar_aberto'] += valor
                elif tipo == 'RECEBER':
                    resumo['receber_aberto'] += valor

                if vencimento and vencimento < hoje:
                    resumo['vencidos'] += valor
                    resumo['qtd_vencidos'] += 1

            elif status in ['Cancelado', 'Estornado']:
                resumo['qtd_cancelados'] += 1

            elif status in ['Pago', 'Recebido']:
                resumo['qtd_baixados'] += 1
                data_emissao = titulo.get('data_emissao')
                if data_emissao and str(data_emissao)[:7] == mes_atual:
                    resumo['pagos_recebidos_mes'] += valor

        pessoas = carregar_pessoas_financeiro(empresa_logada_id, is_super_admin)
        contas_caixa = carregar_contas_caixa_financeiro(empresa_logada_id, is_super_admin)
        empresas = []
        if is_super_admin:
            cur.execute("SELECT id, razao_social, nome_fantasia FROM empresas ORDER BY nome_fantasia ASC, razao_social ASC")
            empresas = cur.fetchall()

    except Exception as e:
        print(f"Erro ao carregar títulos financeiros: {e}")
        flash(f"Erro técnico ao carregar títulos financeiros: {e}", "danger")
        titulos = []
        resumo = {
            'pagar_aberto': 0,
            'receber_aberto': 0,
            'vencidos': 0,
            'qtd_vencidos': 0,
            'pagos_recebidos_mes': 0,
            'total_titulos': 0,
            'qtd_abertos': 0,
            'qtd_cancelados': 0,
            'qtd_baixados': 0
        }
        pessoas = []
        contas_caixa = []
        empresas = []

    finally:
        fechar_cursor_conexao(cur, con)

    filtros = {
        'tipo_titulo': tipo_titulo,
        'status_titulo': status_titulo,
        'origem': origem,
        'pessoa_id': pessoa_id,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'vencimento_inicio': vencimento_inicio,
        'vencimento_fim': vencimento_fim,
        'pesquisa': pesquisa,
        'empresa_id': empresa_id_filtro
    }

    return render_template(
        'financeiro_titulos.html',
        usuario_logado=usuario_logado,
        titulos=titulos,
        resumo=resumo,
        filtros=filtros,
        pessoas=pessoas,
        contas_caixa=contas_caixa,
        empresas=empresas,
        status_titulos=financeiro_base_status_titulos(),
        origens=financeiro_base_origens(),
        formas_pagamento=financeiro_base_formas_pagamento(),
        is_super_admin=is_super_admin
    )


@app.route('/financeiro/titulos/novo', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro')
def novo_titulo_financeiro():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = usuario_eh_super_admin_global()

    if not empresa_logada_id:
        flash("Empresa não identificada na sessão. Faça login novamente.", "danger")
        return redirect(url_for('logout'))

    if request.method == 'POST':
        empresa_id = request.form.get('empresa_id') if is_super_admin else empresa_logada_id
        tipo_titulo = (request.form.get('tipo_titulo') or '').strip()
        pessoa_id = (request.form.get('pessoa_id') or '').strip()
        numero_documento = (request.form.get('numero_documento') or '').strip()
        descricao = (request.form.get('descricao') or '').strip()
        historico = (request.form.get('historico') or '').strip()
        data_emissao = (request.form.get('data_emissao') or '').strip()
        data_competencia = (request.form.get('data_competencia') or '').strip()
        data_vencimento = (request.form.get('data_vencimento') or '').strip()
        forma_pagamento = (request.form.get('forma_pagamento') or '').strip()
        conta_caixa_prevista_id = (request.form.get('conta_caixa_prevista_id') or '').strip()
        valor_original = converter_decimal(request.form.get('valor_original'))
        valor_desconto = converter_decimal(request.form.get('valor_desconto'))
        valor_acrescimo = converter_decimal(request.form.get('valor_acrescimo'))
        observacao = (request.form.get('observacao') or '').strip()

        if not empresa_id or not str(empresa_id).isdigit():
            flash("Selecione uma empresa válida.", "danger")
            return redirect(url_for('novo_titulo_financeiro'))
        empresa_id = int(empresa_id)

        if tipo_titulo not in ['PAGAR', 'RECEBER']:
            flash("Selecione se o título é Conta a Pagar ou Conta a Receber.", "danger")
            return redirect(url_for('novo_titulo_financeiro'))

        if not pessoa_id or not pessoa_id.isdigit():
            flash("Selecione a pessoa responsável pelo título.", "danger")
            return redirect(url_for('novo_titulo_financeiro'))
        pessoa_id = int(pessoa_id)

        if not numero_documento:
            flash("Informe o número do documento.", "danger")
            return redirect(url_for('novo_titulo_financeiro'))

        if not descricao:
            flash("Informe uma descrição para o título.", "danger")
            return redirect(url_for('novo_titulo_financeiro'))

        if not data_emissao or not validar_data_iso(data_emissao):
            flash("Informe uma data de emissão válida.", "danger")
            return redirect(url_for('novo_titulo_financeiro'))

        if not data_competencia:
            data_competencia = data_emissao

        if not validar_data_iso(data_competencia):
            flash("Informe uma data de competência válida.", "danger")
            return redirect(url_for('novo_titulo_financeiro'))

        if not data_vencimento or not validar_data_iso(data_vencimento):
            flash("Informe uma data de vencimento válida.", "danger")
            return redirect(url_for('novo_titulo_financeiro'))

        if valor_original <= 0:
            flash("Informe um valor maior que zero.", "danger")
            return redirect(url_for('novo_titulo_financeiro'))

        if forma_pagamento and forma_pagamento not in financeiro_base_formas_pagamento():
            flash("Forma de pagamento inválida.", "danger")
            return redirect(url_for('novo_titulo_financeiro'))

        conta_caixa_prevista_id_int = None
        if conta_caixa_prevista_id:
            if not conta_caixa_prevista_id.isdigit():
                flash("Conta caixa inválida.", "danger")
                return redirect(url_for('novo_titulo_financeiro'))
            conta_caixa_prevista_id_int = int(conta_caixa_prevista_id)

        valor_liquido = (valor_original - valor_desconto + valor_acrescimo).quantize(Decimal('0.01'))
        if valor_liquido <= 0:
            flash("O valor líquido do título precisa ser maior que zero.", "danger")
            return redirect(url_for('novo_titulo_financeiro'))

        con = obter_conexao()
        if con is None:
            flash("Erro de conexão com o banco de dados.", "danger")
            return redirect(url_for('financeiro_titulos'))

        cur = con.cursor(dictionary=True)
        try:
            cur.execute("SELECT id FROM empresas WHERE id = %s AND status_empresa = 'Ativa' LIMIT 1", (empresa_id,))
            if not cur.fetchone():
                flash("Empresa inválida ou inativa.", "danger")
                return redirect(url_for('novo_titulo_financeiro'))

            cur.execute("""
                SELECT id, nome_completo
                FROM pessoas
                WHERE id = %s
                  AND empresa_id = %s
                  AND status_cadastro = 'Ativo'
                LIMIT 1
            """, (pessoa_id, empresa_id))
            pessoa = cur.fetchone()
            if not pessoa:
                flash("Pessoa inválida ou não pertence à empresa informada.", "danger")
                return redirect(url_for('novo_titulo_financeiro'))

            if conta_caixa_prevista_id_int:
                cur.execute("""
                    SELECT id
                    FROM contas_caixa
                    WHERE id = %s
                      AND empresa_id = %s
                      AND status_conta = 'Ativa'
                    LIMIT 1
                """, (conta_caixa_prevista_id_int, empresa_id))
                if not cur.fetchone():
                    flash("Conta caixa inválida ou inativa.", "danger")
                    return redirect(url_for('novo_titulo_financeiro'))

            if not historico:
                historico = f"{descricao} - Documento {numero_documento} - {pessoa['nome_completo']}"

            cur.execute("""
                INSERT INTO titulos_financeiros
                    (empresa_id, tipo_titulo, origem, origem_id, pessoa_id, numero_documento,
                     descricao, historico, valor_original, valor_desconto, valor_acrescimo,
                     valor_liquido, data_emissao, data_competencia, data_vencimento,
                     forma_pagamento, conta_caixa_prevista_id, status_titulo, observacao,
                     usuario_criacao_id)
                VALUES
                    (%s, %s, 'MANUAL', NULL, %s, %s,
                     %s, %s, %s, %s, %s,
                     %s, %s, %s, %s,
                     %s, %s, 'Aberto', %s,
                     %s)
            """, (
                empresa_id,
                tipo_titulo,
                pessoa_id,
                numero_documento,
                descricao,
                historico,
                valor_original,
                valor_desconto,
                valor_acrescimo,
                valor_liquido,
                data_emissao,
                data_competencia,
                data_vencimento,
                forma_pagamento or None,
                conta_caixa_prevista_id_int,
                observacao or None,
                usuario_id
            ))

            titulo_id = cur.lastrowid
            registrar_auditoria_financeira(
                cur,
                empresa_id=empresa_id,
                usuario_id=usuario_id,
                acao='TITULO_MANUAL_CRIADO',
                modulo='TITULOS_FINANCEIROS',
                entidade_tipo='TITULO_FINANCEIRO',
                entidade_id=titulo_id,
                titulo_financeiro_id=titulo_id,
                pessoa_id=pessoa_id,
                status_novo='Aberto',
                valor_novo=valor_liquido,
                motivo='Criação manual de título financeiro',
                observacao=f'Título manual #{titulo_id} criado. Documento: {numero_documento}.',
                dados_depois={
                    'tipo_titulo': tipo_titulo,
                    'numero_documento': numero_documento,
                    'descricao': descricao,
                    'valor_original': str(valor_original),
                    'valor_liquido': str(valor_liquido),
                    'data_vencimento': data_vencimento,
                }
            )
            con.commit()

            flash(f"Título financeiro #{titulo_id} criado com sucesso.", "success")
            return redirect(url_for('detalhes_titulo_financeiro', id=titulo_id))

        except Exception as e:
            con.rollback()
            print(f"Erro ao criar título financeiro: {e}")
            flash(f"Erro técnico ao criar título financeiro: {e}", "danger")
            return redirect(url_for('novo_titulo_financeiro'))
        finally:
            fechar_cursor_conexao(cur, con)

    pessoas = carregar_pessoas_financeiro(empresa_logada_id, is_super_admin)
    contas_caixa = carregar_contas_caixa_financeiro(empresa_logada_id, is_super_admin)
    parametros_financeiros = carregar_parametros_financeiros_empresa(empresa_logada_id)
    conta_padrao_id = (parametros_financeiros.get('caixa.conta_padrao_id', {}) or {}).get('valor') or ''
    forma_pagamento_padrao = (parametros_financeiros.get('caixa.forma_pagamento_padrao', {}) or {}).get('valor') or 'PIX'
    empresas = []
    if is_super_admin:
        empresas = carregar_empresas_ativas()

    return render_template(
        'financeiro_titulo_form.html',
        usuario_logado=usuario_logado,
        pessoas=pessoas,
        contas_caixa=contas_caixa,
        empresas=empresas,
        formas_pagamento=financeiro_base_formas_pagamento(),
        is_super_admin=is_super_admin,
        parametros_financeiros=parametros_financeiros,
        parametro_bool=parametro_bool,
        conta_padrao_id=conta_padrao_id,
        forma_pagamento_padrao=forma_pagamento_padrao,
        hoje=date.today().strftime('%Y-%m-%d')
    )




# ----------------------------------------------------------
# Bloco 5 — Calcula saldo atual de uma conta caixa.
# Saldo = saldo inicial + entradas baixadas - saídas baixadas.
# ----------------------------------------------------------
def calcular_saldo_conta_caixa(cur, conta_caixa_id, empresa_id):
    cur.execute("""
        SELECT
            c.id,
            c.nome_conta,
            c.saldo_inicial,
            COALESCE(SUM(
                CASE
                    WHEN m.tipo_movimentacao = 'ENTRADA'
                        THEN COALESCE(m.valor_movimentacao, 0)
                    WHEN m.tipo_movimentacao = 'SAIDA'
                        THEN -COALESCE(m.valor_movimentacao, 0)
                    ELSE 0
                END
            ), 0) AS saldo_movimentado
        FROM contas_caixa c
        LEFT JOIN movimentacoes_caixa m
               ON m.conta_caixa_id = c.id
              AND m.empresa_id = c.empresa_id
        WHERE c.id = %s
          AND c.empresa_id = %s
        GROUP BY c.id, c.nome_conta, c.saldo_inicial
        LIMIT 1
    """, (conta_caixa_id, empresa_id))
    row = cur.fetchone()
    if not row:
        return None

    saldo = converter_decimal(row.get('saldo_inicial')) + converter_decimal(row.get('saldo_movimentado'))
    row['saldo_atual'] = saldo.quantize(Decimal('0.01'))
    return row


# ----------------------------------------------------------
# Bloco 5 — Salva comprovante de baixa financeira.
# Usa Google Drive quando habilitado e mantém fallback local.
# ----------------------------------------------------------
def salvar_comprovante_baixa_titulo(cur, arquivo, *, empresa_id, titulo_id, pessoa_id=None, usuario_id=None):
    if not arquivo or not arquivo.filename:
        return None

    nome_original = str(arquivo.filename or 'comprovante').strip()
    nome_seguro = nome_original.replace('\\', '_').replace('/', '_')
    nome_seguro = re.sub(r'[^a-zA-Z0-9_.-]+', '_', nome_seguro) or 'comprovante'

    pasta = os.path.join(app.root_path, 'uploads', 'comprovantes_financeiros')
    os.makedirs(pasta, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    nome_final = f"empresa_{empresa_id}_titulo_{titulo_id}_{timestamp}_{nome_seguro}"
    caminho_final = os.path.join(pasta, nome_final)
    arquivo.save(caminho_final)

    caminho_relativo = f"uploads/comprovantes_financeiros/{nome_final}"

    try:
        return tentar_enviar_arquivo_google_drive(
            cur,
            caminho_final,
            caminho_relativo,
            empresa_id=empresa_id,
            motorista_id=pessoa_id,
            origem='COMPROVANTE_FINANCEIRO',
            origem_id=titulo_id,
            tipo_arquivo='COMPROVANTE_FINANCEIRO',
            nome_original=nome_original,
            mime_type=getattr(arquivo, 'mimetype', None) or 'application/octet-stream',
            criado_por_usuario_id=usuario_id or session.get('usuario_id'),
        )
    except Exception as exc:
        print(f"[Financeiro] Falha ao enviar comprovante do título {titulo_id}: {exc}")
        return caminho_relativo


# ----------------------------------------------------------
# Bloco 5 — Atualiza documento do motorista e rotas após baixa.
# Usado quando o título nasceu de NF_MOTORISTA ou SEM_NF_MOTORISTA.
# ----------------------------------------------------------
def aplicar_baixa_em_documento_motorista_e_rotas(cur, *, titulo_id, empresa_id, usuario_id):
    """
    Bloco 5.1.2 — Sincronização robusta de baixa para documentos de motorista.

    Esta versão evita processamento rota a rota com várias consultas e não abre novas
    conexões dentro da transação principal. Ela faz a sincronização em massa:
    - Documento do motorista -> Pagamento confirmado;
    - Rotas vinculadas -> Pagamento confirmado / Quitada;
    - Histórico registrado usando a mesma conexão/cursor da baixa.
    """
    cur.execute("""
        SELECT id, empresa_id, origem, origem_id, numero_documento, status_titulo
        FROM titulos_financeiros
        WHERE id = %s
          AND empresa_id = %s
        LIMIT 1
    """, (titulo_id, empresa_id))
    titulo = cur.fetchone() or {}

    origem_titulo = str(titulo.get('origem') or '').strip()
    status_titulo = str(titulo.get('status_titulo') or '').strip()

    if origem_titulo not in ['NF_MOTORISTA', 'SEM_NF_MOTORISTA']:
        return

    if status_titulo not in ['Pago', 'Recebido']:
        return

    # ----------------------------------------------------------
    # 1. Localiza documento(s) do motorista vinculados ao título.
    # ----------------------------------------------------------
    nf_ids = []

    if titulo.get('origem_id'):
        try:
            nf_ids.append(int(titulo.get('origem_id')))
        except Exception:
            pass

    cur.execute("""
        SELECT origem_id
        FROM titulos_financeiros_vinculos
        WHERE titulo_financeiro_id = %s
          AND empresa_id = %s
          AND origem_id IS NOT NULL
          AND (
                origem_tabela = 'motorista_notas_fiscais'
                OR tipo_vinculo IN ('NF_MOTORISTA', 'SEM_NF_MOTORISTA')
              )
    """, (titulo_id, empresa_id))

    for row in cur.fetchall():
        try:
            nf_ids.append(int(row.get('origem_id')))
        except Exception:
            pass

    nf_ids = sorted(set([x for x in nf_ids if x]))

    # ----------------------------------------------------------
    # 2. Localiza rotas vinculadas ao título ou aos documentos.
    # ----------------------------------------------------------
    rota_ids = []

    cur.execute("""
        SELECT origem_id
        FROM titulos_financeiros_vinculos
        WHERE titulo_financeiro_id = %s
          AND empresa_id = %s
          AND origem_id IS NOT NULL
          AND (origem_tabela = 'rotas' OR tipo_vinculo = 'ROTA')
    """, (titulo_id, empresa_id))

    for row in cur.fetchall():
        try:
            rota_ids.append(int(row.get('origem_id')))
        except Exception:
            pass

    if nf_ids:
        placeholders_nf = ','.join(['%s'] * len(nf_ids))
        cur.execute(f"""
            SELECT DISTINCT rota_id
            FROM motorista_nf_rotas
            WHERE empresa_id = %s
              AND motorista_nf_id IN ({placeholders_nf})
              AND rota_id IS NOT NULL
        """, [empresa_id] + nf_ids)

        for row in cur.fetchall():
            try:
                rota_ids.append(int(row.get('rota_id')))
            except Exception:
                pass

    rota_ids = sorted(set([x for x in rota_ids if x]))

    # ----------------------------------------------------------
    # 3. Histórico e atualização dos documentos em lote.
    # ----------------------------------------------------------
    if nf_ids:
        placeholders_nf = ','.join(['%s'] * len(nf_ids))

        # Registra histórico antes da atualização para guardar status anterior.
        cur.execute(f"""
            INSERT INTO historico_operacoes
                (empresa_id, tipo_operacao, usuario_id, status_anterior, status_novo, motivo, observacao)
            SELECT
                empresa_id,
                'NF_MOTORISTA',
                %s,
                status_nf,
                'Pagamento confirmado',
                'Baixa financeira do título',
                CONCAT('NF Motorista ID ', id, '. Pagamento confirmado pela baixa do título financeiro #', %s, '.')
            FROM motorista_notas_fiscais
            WHERE empresa_id = %s
              AND id IN ({placeholders_nf})
              AND COALESCE(status_nf, '') NOT IN ('Pagamento confirmado', 'Recusada', 'Estornada', 'Cancelada')
        """, [usuario_id, titulo_id, empresa_id] + nf_ids)

        cur.execute(f"""
            UPDATE motorista_notas_fiscais
            SET status_nf = 'Pagamento confirmado',
                data_pagamento = COALESCE(data_pagamento, NOW()),
                usuario_pagamento_id = COALESCE(usuario_pagamento_id, %s),
                observacao = CONCAT(
                    COALESCE(observacao, ''),
                    CASE WHEN COALESCE(observacao, '') = '' THEN '' ELSE '\n' END,
                    'Pagamento confirmado em ',
                    DATE_FORMAT(NOW(), '%d/%m/%Y %H:%i'),
                    '. Título financeiro baixado: #',
                    %s
                )
            WHERE empresa_id = %s
              AND id IN ({placeholders_nf})
              AND COALESCE(status_nf, '') NOT IN ('Pagamento confirmado', 'Recusada', 'Estornada', 'Cancelada')
        """, [usuario_id, titulo_id, empresa_id] + nf_ids)

    # ----------------------------------------------------------
    # 4. Histórico e atualização das rotas em lote.
    # ----------------------------------------------------------
    if rota_ids:
        placeholders_rota = ','.join(['%s'] * len(rota_ids))

        cur.execute(f"""
            INSERT INTO historico_operacoes
                (empresa_id, tipo_operacao, rota_id, usuario_id, status_anterior, status_novo, motivo, observacao)
            SELECT
                empresa_id,
                'STATUS_MOTORISTA_ROTA',
                id,
                %s,
                status_motorista,
                'Pagamento confirmado',
                'Baixa financeira do título',
                CONCAT('Rota quitada pela baixa do título financeiro #', %s, '.')
            FROM rotas
            WHERE empresa_id = %s
              AND id IN ({placeholders_rota})
              AND COALESCE(situacao_rota, '') <> 'Cancelada'
              AND COALESCE(status_motorista, '') <> 'Pagamento confirmado'
        """, [usuario_id, titulo_id, empresa_id] + rota_ids)

        cur.execute(f"""
            UPDATE rotas
            SET status_motorista = 'Pagamento confirmado',
                situacao_rota = 'Quitada'
            WHERE empresa_id = %s
              AND id IN ({placeholders_rota})
              AND COALESCE(situacao_rota, '') <> 'Cancelada'
        """, [empresa_id] + rota_ids)


@app.route('/financeiro/titulos/<int:id>', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro', 'Consulta')
def detalhes_titulo_financeiro(id):
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('financeiro_titulos'))

    cur = con.cursor(dictionary=True)
    try:
        query = """
            SELECT t.*,
                   p.nome_completo AS pessoa_nome,
                   p.cpf_cnpj AS pessoa_cpf_cnpj,
                   p.tipo_cadastro AS pessoa_tipo,
                   cx.nome_conta AS conta_caixa_nome,
                   cxb.nome_conta AS conta_caixa_baixa_nome,
                   e.nome_fantasia AS empresa_nome,
                   e.razao_social AS empresa_razao_social,
                   u.login AS usuario_criacao_login,
                   ub.login AS usuario_baixa_login,
                   ue.login AS usuario_estorno_login
            FROM titulos_financeiros t
            LEFT JOIN pessoas p ON p.id = t.pessoa_id AND p.empresa_id = t.empresa_id
            LEFT JOIN contas_caixa cx ON cx.id = t.conta_caixa_prevista_id AND cx.empresa_id = t.empresa_id
            LEFT JOIN contas_caixa cxb ON cxb.id = t.conta_caixa_baixa_id AND cxb.empresa_id = t.empresa_id
            LEFT JOIN empresas e ON e.id = t.empresa_id
            LEFT JOIN usuarios u ON u.id = t.usuario_criacao_id
            LEFT JOIN usuarios ub ON ub.id = t.usuario_baixa_id
            LEFT JOIN usuarios ue ON ue.id = t.usuario_estorno_id
            WHERE t.id = %s
        """
        params = [id]
        if not is_super_admin:
            query += " AND t.empresa_id = %s"
            params.append(empresa_logada_id)
        query += " LIMIT 1"

        cur.execute(query, params)
        titulo = cur.fetchone()
        if not titulo:
            flash("Título financeiro não encontrado ou não pertence à empresa logada.", "danger")
            return redirect(url_for('financeiro_titulos'))

        cur.execute("""
            SELECT id, tipo_vinculo, origem_tabela, origem_id, descricao, valor_vinculo
            FROM titulos_financeiros_vinculos
            WHERE titulo_financeiro_id = %s
              AND empresa_id = %s
            ORDER BY id ASC
        """, (id, titulo['empresa_id']))
        vinculos = cur.fetchall()

        cur.execute("""
            SELECT m.*,
                   cx.nome_conta AS conta_caixa_nome,
                   u.login AS usuario_login
            FROM movimentacoes_caixa m
            INNER JOIN contas_caixa cx ON cx.id = m.conta_caixa_id AND cx.empresa_id = m.empresa_id
            LEFT JOIN usuarios u ON u.id = m.usuario_criacao_id
            WHERE m.titulo_financeiro_id = %s
              AND m.empresa_id = %s
            ORDER BY m.data_movimentacao DESC, m.id DESC
        """, (id, titulo['empresa_id']))
        movimentacoes = cur.fetchall()

        parametros_financeiros = carregar_parametros_financeiros_empresa(titulo['empresa_id'], cur=cur)
        conta_padrao_id = str(parametros_financeiros.get('caixa.conta_padrao_id', {}).get('valor') or '')
        forma_pagamento_padrao = parametros_financeiros.get('caixa.forma_pagamento_padrao', {}).get('valor') or (titulo.get('forma_pagamento') or 'PIX')

        contas_caixa = []
        if titulo.get('status_titulo') not in ['Pago', 'Recebido', 'Cancelado', 'Estornado']:
            cur.execute("""
                SELECT id, nome_conta, tipo_conta, banco, agencia, numero_conta, saldo_inicial, status_conta
                FROM contas_caixa
                WHERE empresa_id = %s
                  AND status_conta = 'Ativa'
                ORDER BY nome_conta ASC
            """, (titulo['empresa_id'],))
            contas_caixa = cur.fetchall()
            for conta in contas_caixa:
                saldo_info = calcular_saldo_conta_caixa(cur, conta['id'], titulo['empresa_id'])
                conta['saldo_atual'] = (saldo_info or {}).get('saldo_atual', Decimal('0.00'))

    except Exception as e:
        print(f"Erro ao carregar detalhes do título financeiro: {e}")
        flash(f"Erro técnico ao carregar título financeiro: {e}", "danger")
        return redirect(url_for('financeiro_titulos'))
    finally:
        fechar_cursor_conexao(cur, con)

    return render_template(
        'financeiro_titulo_detalhes.html',
        usuario_logado=usuario_logado,
        titulo=titulo,
        vinculos=vinculos,
        movimentacoes=movimentacoes,
        contas_caixa=contas_caixa,
        formas_pagamento=financeiro_base_formas_pagamento(),
        parametros_financeiros=parametros_financeiros,
        parametro_bool=parametro_bool,
        conta_padrao_id=conta_padrao_id,
        forma_pagamento_padrao=forma_pagamento_padrao,
        is_super_admin=is_super_admin,
        hoje=date.today().strftime('%Y-%m-%d')
    )

@app.route('/financeiro/titulos/<int:id>/cancelar', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro')
def cancelar_titulo_financeiro(id):
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = usuario_eh_super_admin_global()
    motivo = (request.form.get('motivo_cancelamento') or '').strip()

    if len(motivo) < 5:
        flash("Informe um motivo de cancelamento com pelo menos 5 caracteres.", "warning")
        return redirect(url_for('detalhes_titulo_financeiro', id=id))

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('financeiro_titulos'))

    cur = con.cursor(dictionary=True)
    try:
        query = "SELECT id, empresa_id, status_titulo FROM titulos_financeiros WHERE id = %s"
        params = [id]
        if not is_super_admin:
            query += " AND empresa_id = %s"
            params.append(empresa_logada_id)
        query += " LIMIT 1"
        cur.execute(query, params)
        titulo = cur.fetchone()

        if not titulo:
            flash("Título financeiro não encontrado ou não pertence à empresa logada.", "danger")
            return redirect(url_for('financeiro_titulos'))

        if titulo.get('status_titulo') in ['Pago', 'Recebido', 'Cancelado', 'Estornado']:
            flash(f"Este título não pode ser cancelado. Status atual: {titulo.get('status_titulo')}.", "warning")
            return redirect(url_for('detalhes_titulo_financeiro', id=id))

        cur.execute("""
            UPDATE titulos_financeiros
            SET status_titulo = 'Cancelado',
                motivo_cancelamento = %s,
                data_cancelamento = NOW(),
                usuario_cancelamento_id = %s,
                updated_at = NOW()
            WHERE id = %s
              AND empresa_id = %s
        """, (motivo, usuario_id, id, titulo['empresa_id']))
        registrar_auditoria_financeira(
            cur,
            empresa_id=titulo['empresa_id'],
            usuario_id=usuario_id,
            acao='TITULO_CANCELADO',
            modulo='TITULOS_FINANCEIROS',
            entidade_tipo='TITULO_FINANCEIRO',
            entidade_id=id,
            titulo_financeiro_id=id,
            status_anterior=titulo.get('status_titulo'),
            status_novo='Cancelado',
            motivo=motivo,
            observacao=f'Título financeiro #{id} cancelado.',
        )
        con.commit()
        flash("Título financeiro cancelado com sucesso.", "success")
    except Exception as e:
        con.rollback()
        print(f"Erro ao cancelar título financeiro: {e}")
        flash("Erro técnico ao cancelar título financeiro.", "danger")
    finally:
        fechar_cursor_conexao(cur, con)

    return redirect(url_for('detalhes_titulo_financeiro', id=id))




@app.route('/financeiro/titulos/<int:id>/baixar', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro')
def baixar_titulo_financeiro(id):
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = usuario_eh_super_admin_global()

    conta_caixa_id = (request.form.get('conta_caixa_id') or '').strip()
    data_pagamento = (request.form.get('data_pagamento') or '').strip()
    forma_pagamento = (request.form.get('forma_pagamento') or '').strip()
    valor_pago = converter_decimal(request.form.get('valor_pago'))
    observacao_baixa = (request.form.get('observacao_baixa') or '').strip()
    comprovante = request.files.get('comprovante')

    if not conta_caixa_id or not conta_caixa_id.isdigit():
        flash("Selecione uma conta caixa válida para baixar o título.", "danger")
        return redirect(url_for('detalhes_titulo_financeiro', id=id))
    conta_caixa_id = int(conta_caixa_id)

    if not data_pagamento or not validar_data_iso(data_pagamento):
        flash("Informe uma data de pagamento/recebimento válida.", "danger")
        return redirect(url_for('detalhes_titulo_financeiro', id=id))

    if forma_pagamento not in financeiro_base_formas_pagamento():
        flash("Selecione uma forma de pagamento válida.", "danger")
        return redirect(url_for('detalhes_titulo_financeiro', id=id))

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('financeiro_titulos'))

    cur = con.cursor(dictionary=True)
    try:
        query = """
            SELECT id, empresa_id, tipo_titulo, origem, origem_id, pessoa_id, numero_documento,
                   descricao, historico, valor_liquido, status_titulo
            FROM titulos_financeiros
            WHERE id = %s
        """
        params = [id]
        if not is_super_admin:
            query += " AND empresa_id = %s"
            params.append(empresa_logada_id)
        query += " LIMIT 1"
        cur.execute(query, params)
        titulo = cur.fetchone()

        if not titulo:
            flash("Título financeiro não encontrado ou não pertence à empresa logada.", "danger")
            return redirect(url_for('financeiro_titulos'))

        parametros_financeiros = carregar_parametros_financeiros_empresa(titulo['empresa_id'], cur=cur)
        exigir_comprovante_baixa = parametro_bool(parametros_financeiros.get('baixa.exigir_comprovante', {}).get('valor'))
        permitir_pagamento_parcial = parametro_bool(parametros_financeiros.get('baixa.permitir_pagamento_parcial', {}).get('valor'))
        permitir_valor_diferente = parametro_bool(parametros_financeiros.get('baixa.permitir_valor_diferente', {}).get('valor'))
        permitir_saldo_negativo = parametro_bool(parametros_financeiros.get('caixa.permitir_saldo_negativo', {}).get('valor'))
        permitir_data_retroativa = parametro_bool(parametros_financeiros.get('baixa.permitir_data_retroativa', {}).get('valor'))

        if not permitir_data_retroativa and data_pagamento < date.today().strftime('%Y-%m-%d'):
            flash('Baixa retroativa bloqueada pelas configurações financeiras da empresa.', 'warning')
            return redirect(url_for('detalhes_titulo_financeiro', id=id))

        if exigir_comprovante_baixa and (not comprovante or not getattr(comprovante, 'filename', '')):
            flash('Comprovante obrigatório para baixa, conforme configuração financeira da empresa.', 'warning')
            return redirect(url_for('detalhes_titulo_financeiro', id=id))

        status_atual = titulo.get('status_titulo') or 'Aberto'
        if status_atual in ['Pago', 'Recebido', 'Cancelado', 'Estornado']:
            flash(f"Este título não pode ser baixado. Status atual: {status_atual}.", "warning")
            return redirect(url_for('detalhes_titulo_financeiro', id=id))

        # Proteção contra duplicidade: se já existe movimentação ativa para este título,
        # não cria outra baixa mesmo que o status do título ainda não tenha sido atualizado.
        movimentacoes_ativas = buscar_movimentacoes_baixa_nao_estornadas(
            cur,
            titulo_id=id,
            empresa_id=titulo['empresa_id']
        )
        movimentacao_existente = movimentacoes_ativas[0] if movimentacoes_ativas else None
        if movimentacao_existente:
            flash(
                "Este título já possui uma movimentação de caixa ativa vinculada. "
                "A baixa não foi duplicada. Abra as movimentações do título para conferir.",
                "warning"
            )
            return redirect(url_for('detalhes_titulo_financeiro', id=id))

        valor_liquido = converter_decimal(titulo.get('valor_liquido'))
        if valor_pago <= 0:
            flash("Informe um valor de baixa maior que zero.", "danger")
            return redirect(url_for('detalhes_titulo_financeiro', id=id))

        if valor_pago != valor_liquido:
            if valor_pago < valor_liquido and permitir_pagamento_parcial:
                flash('Pagamento parcial registrado como baixa total ainda não está disponível nesta etapa. A configuração já foi preparada, mas a conciliação parcial será liberada em bloco próprio.', 'warning')
                return redirect(url_for('detalhes_titulo_financeiro', id=id))
            if not permitir_valor_diferente:
                flash(
                    'O valor baixado precisa ser igual ao valor líquido do título, conforme configuração financeira da empresa.',
                    'warning'
                )
                return redirect(url_for('detalhes_titulo_financeiro', id=id))

        cur.execute("""
            SELECT id, nome_conta, status_conta
            FROM contas_caixa
            WHERE id = %s
              AND empresa_id = %s
              AND status_conta = 'Ativa'
            LIMIT 1
        """, (conta_caixa_id, titulo['empresa_id']))
        conta = cur.fetchone()
        if not conta:
            flash("Conta caixa inválida, inativa ou não pertence à empresa do título.", "danger")
            return redirect(url_for('detalhes_titulo_financeiro', id=id))

        tipo_movimentacao = 'SAIDA' if titulo.get('tipo_titulo') == 'PAGAR' else 'ENTRADA'
        novo_status = 'Pago' if titulo.get('tipo_titulo') == 'PAGAR' else 'Recebido'

        if tipo_movimentacao == 'SAIDA':
            saldo_info = calcular_saldo_conta_caixa(cur, conta_caixa_id, titulo['empresa_id'])
            saldo_atual = converter_decimal((saldo_info or {}).get('saldo_atual'))
            if saldo_atual < valor_pago and not permitir_saldo_negativo:
                flash(
                    f"Baixa bloqueada: a conta caixa '{conta['nome_conta']}' possui saldo de {moeda_br(saldo_atual)}, "
                    f"menor que o valor do pagamento {moeda_br(valor_pago)}. "
                    "A empresa está configurada para não permitir caixa negativo.",
                    "danger"
                )
                return redirect(url_for('detalhes_titulo_financeiro', id=id))

        comprovante_url = salvar_comprovante_baixa_titulo(
            cur,
            comprovante,
            empresa_id=titulo['empresa_id'],
            titulo_id=id,
            pessoa_id=titulo.get('pessoa_id'),
            usuario_id=usuario_id,
        )

        historico_mov = (
            f"Baixa do título #{id} - {titulo.get('descricao') or titulo.get('numero_documento')}"
        )

        # Segunda checagem dentro do fluxo imediatamente antes de inserir,
        # reduzindo risco de duplicidade em duplo clique ou retentativa do navegador.
        if buscar_movimentacoes_baixa_nao_estornadas(
            cur,
            titulo_id=id,
            empresa_id=titulo['empresa_id']
        ):
            flash(
                "Este título já possui baixa registrada. A operação não foi duplicada.",
                "warning"
            )
            return redirect(url_for('detalhes_titulo_financeiro', id=id))

        cur.execute("""
            INSERT INTO movimentacoes_caixa
                (empresa_id, conta_caixa_id, titulo_financeiro_id, tipo_movimentacao,
                 data_movimentacao, valor_movimentacao, forma_pagamento, historico,
                 observacao, comprovante_url, status_movimentacao, usuario_criacao_id)
            VALUES
                (%s, %s, %s, %s,
                 %s, %s, %s, %s,
                 %s, %s, 'Ativa', %s)
        """, (
            titulo['empresa_id'],
            conta_caixa_id,
            id,
            tipo_movimentacao,
            data_pagamento,
            valor_pago,
            forma_pagamento,
            historico_mov,
            observacao_baixa or None,
            comprovante_url,
            usuario_id
        ))

        cur.execute("""
            UPDATE titulos_financeiros
            SET status_titulo = %s,
                conta_caixa_baixa_id = %s,
                data_baixa = %s,
                valor_baixado = %s,
                forma_pagamento = %s,
                observacao_baixa = %s,
                comprovante_url = %s,
                usuario_baixa_id = %s,
                updated_at = NOW()
            WHERE id = %s
              AND empresa_id = %s
        """, (
            novo_status,
            conta_caixa_id,
            data_pagamento,
            valor_pago,
            forma_pagamento,
            observacao_baixa or None,
            comprovante_url,
            usuario_id,
            id,
            titulo['empresa_id']
        ))

        if titulo.get('origem') in ['NF_MOTORISTA', 'SEM_NF_MOTORISTA']:
            aplicar_baixa_em_documento_motorista_e_rotas(
                cur,
                titulo_id=id,
                empresa_id=titulo['empresa_id'],
                usuario_id=usuario_id
            )

        registrar_auditoria_financeira(
            cur,
            empresa_id=titulo['empresa_id'],
            usuario_id=usuario_id,
            acao='BAIXA_TITULO',
            modulo='BAIXA_FINANCEIRA',
            entidade_tipo='TITULO_FINANCEIRO',
            entidade_id=id,
            titulo_financeiro_id=id,
            pessoa_id=titulo.get('pessoa_id'),
            status_anterior=status_atual,
            status_novo=novo_status,
            valor_anterior=titulo.get('valor_liquido'),
            valor_novo=valor_pago,
            motivo='Baixa financeira de título',
            observacao=observacao_baixa or f'Título #{id} baixado como {novo_status}.',
            dados_depois={
                'conta_caixa_id': conta_caixa_id,
                'data_pagamento': data_pagamento,
                'forma_pagamento': forma_pagamento,
                'tipo_movimentacao': tipo_movimentacao,
                'comprovante_url': comprovante_url,
            }
        )
        con.commit()
        flash(f"Título financeiro #{id} baixado com sucesso como {novo_status}.", "success")
        return redirect(url_for('detalhes_titulo_financeiro', id=id))

    except Exception as e:
        try:
            con.rollback()
        except Exception as rollback_error:
            print(f"Aviso: não foi possível executar rollback da baixa do título {id}: {rollback_error}")

        print(f"Erro ao baixar título financeiro {id}: {e}")
        flash(
            "Erro técnico ao baixar título financeiro. "
            "A operação foi interrompida com segurança; confira se o título possui movimentação antes de tentar novamente.",
            "danger"
        )
        return redirect(url_for('detalhes_titulo_financeiro', id=id))

    finally:
        fechar_cursor_conexao(cur, con)


# ----------------------------------------------------------
# Bloco 5.2 — Busca movimentações de baixa ainda não estornadas.
# Evita duplicidade e permite nova baixa após estorno com reabertura.
# ----------------------------------------------------------
def buscar_movimentacoes_baixa_nao_estornadas(cur, *, titulo_id, empresa_id):
    cur.execute("""
        SELECT m.*
        FROM movimentacoes_caixa m
        WHERE m.titulo_financeiro_id = %s
          AND m.empresa_id = %s
          AND COALESCE(m.status_movimentacao, 'Ativa') = 'Ativa'
          AND (m.estorno_de_movimentacao_id IS NULL OR m.estorno_de_movimentacao_id = 0)
          AND NOT EXISTS (
              SELECT 1
              FROM movimentacoes_caixa e
              WHERE e.empresa_id = m.empresa_id
                AND e.titulo_financeiro_id = m.titulo_financeiro_id
                AND e.estorno_de_movimentacao_id = m.id
                AND COALESCE(e.status_movimentacao, 'Ativa') IN ('Ativa', 'Estorno')
          )
        ORDER BY m.id ASC
    """, (titulo_id, empresa_id))
    return cur.fetchall()


# ----------------------------------------------------------
# Bloco 5.2 — Aplica estorno nos documentos do motorista e rotas.
# Reabre para nova baixa ou encerra como estornado, preservando histórico.
# ----------------------------------------------------------
def aplicar_estorno_em_documento_motorista_e_rotas(cur, *, titulo_id, empresa_id, usuario_id, motivo, destino, tratativa_pos_estorno='manter_bloqueadas'):
    # Documento vinculado diretamente pelo título.
    nf_ids = []
    cur.execute("""
        SELECT origem, origem_id
        FROM titulos_financeiros
        WHERE id = %s
          AND empresa_id = %s
        LIMIT 1
    """, (titulo_id, empresa_id))
    titulo_ref = cur.fetchone()
    if titulo_ref and titulo_ref.get('origem') in ['NF_MOTORISTA', 'SEM_NF_MOTORISTA'] and titulo_ref.get('origem_id'):
        nf_ids.append(int(titulo_ref.get('origem_id')))

    # Documentos vinculados pela composição do título.
    cur.execute("""
        SELECT DISTINCT origem_id
        FROM titulos_financeiros_vinculos
        WHERE titulo_financeiro_id = %s
          AND empresa_id = %s
          AND origem_tabela = 'motorista_notas_fiscais'
          AND origem_id IS NOT NULL
    """, (titulo_id, empresa_id))
    for row in cur.fetchall() or []:
        if row.get('origem_id'):
            nf_ids.append(int(row.get('origem_id')))

    nf_ids = sorted(set(nf_ids))

    rota_ids = []
    if nf_ids:
        placeholders_nf = ','.join(['%s'] * len(nf_ids))
        cur.execute(f"""
            SELECT DISTINCT rota_id
            FROM motorista_nf_rotas
            WHERE empresa_id = %s
              AND motorista_nf_id IN ({placeholders_nf})
              AND rota_id IS NOT NULL
        """, [empresa_id] + nf_ids)
        rota_ids.extend([int(r['rota_id']) for r in (cur.fetchall() or []) if r.get('rota_id')])

    # Rotas vinculadas diretamente pela composição do título.
    cur.execute("""
        SELECT DISTINCT origem_id
        FROM titulos_financeiros_vinculos
        WHERE titulo_financeiro_id = %s
          AND empresa_id = %s
          AND origem_tabela = 'rotas'
          AND origem_id IS NOT NULL
    """, (titulo_id, empresa_id))
    for row in cur.fetchall() or []:
        if row.get('origem_id'):
            rota_ids.append(int(row.get('origem_id')))

    rota_ids = sorted(set(rota_ids))

    # Bloco 5.2.1 — Tratativa pós-estorno.
    # Reabrir: corrige apenas uma baixa equivocada e permite nova baixa do mesmo título.
    # Encerrar: encerra o título antigo e define o destino do documento/rotas.
    tratativa_pos_estorno = (tratativa_pos_estorno or 'manter_bloqueadas').strip()

    if destino == 'reabrir':
        novo_status_nf = 'Pagamento solicitado'
        novo_status_rota = 'Aprovada para pagamento'
        nova_situacao_rota = 'Pendente'
        motivo_hist = 'Estorno de baixa financeira com reabertura'
        destino_legivel = 'reaberto para nova baixa'
    else:
        motivo_hist = 'Estorno de baixa financeira definitivo'
        if tratativa_pos_estorno == 'reabrir_mesmo_documento':
            novo_status_nf = 'Aprovada'
            novo_status_rota = 'Aprovada para pagamento'
            nova_situacao_rota = 'Pendente'
            destino_legivel = 'título encerrado; mesmo documento reaproveitado para nova solicitação de pagamento'
        elif tratativa_pos_estorno == 'exigir_nova_nf':
            novo_status_nf = 'Estornada'
            novo_status_rota = 'Liberada para NF'
            nova_situacao_rota = 'Pendente'
            destino_legivel = 'título encerrado; rotas liberadas exigindo novo documento/NF'
        elif tratativa_pos_estorno == 'cancelar_rotas':
            novo_status_nf = 'Estornada'
            novo_status_rota = 'Cancelada'
            nova_situacao_rota = 'Cancelada'
            destino_legivel = 'título encerrado; rotas canceladas definitivamente'
        else:
            novo_status_nf = 'Estornada'
            novo_status_rota = 'Bloqueada para correção'
            nova_situacao_rota = 'Pendente'
            destino_legivel = 'título encerrado; rotas mantidas bloqueadas para análise'

    if nf_ids:
        placeholders_nf = ','.join(['%s'] * len(nf_ids))

        cur.execute(f"""
            INSERT INTO historico_operacoes
                (empresa_id, tipo_operacao, usuario_id, status_anterior, status_novo, motivo, observacao)
            SELECT
                empresa_id,
                'NF_MOTORISTA',
                %s,
                status_nf,
                %s,
                %s,
                CONCAT('Documento de motorista estornado pelo título financeiro #', %s, '. Motivo: ', %s)
            FROM motorista_notas_fiscais
            WHERE empresa_id = %s
              AND id IN ({placeholders_nf})
        """, [usuario_id, novo_status_nf, motivo_hist, titulo_id, motivo, empresa_id] + nf_ids)

        cur.execute(f"""
            UPDATE motorista_notas_fiscais
            SET status_nf = %s,
                data_estorno_pagamento = NOW(),
                motivo_estorno_pagamento = %s,
                usuario_estorno_pagamento_id = %s,
                data_pagamento = CASE WHEN %s = 'reabrir' THEN NULL ELSE data_pagamento END,
                usuario_pagamento_id = CASE WHEN %s = 'reabrir' THEN NULL ELSE usuario_pagamento_id END,
                observacao = CONCAT(
                    COALESCE(observacao, ''),
                    CASE WHEN COALESCE(observacao, '') = '' THEN '' ELSE '\n' END,
                    'Pagamento estornado em ',
                    DATE_FORMAT(NOW(), '%d/%m/%Y %H:%i'),
                    '. Título financeiro: #',
                    %s,
                    '. Destino: ',
                    %s,
                    '. Motivo: ',
                    %s
                )
            WHERE empresa_id = %s
              AND id IN ({placeholders_nf})
        """, [novo_status_nf, motivo, usuario_id, destino, destino, titulo_id, destino_legivel, motivo, empresa_id] + nf_ids)

    if rota_ids:
        placeholders_rota = ','.join(['%s'] * len(rota_ids))

        cur.execute(f"""
            INSERT INTO historico_operacoes
                (empresa_id, tipo_operacao, rota_id, usuario_id, status_anterior, status_novo, motivo, observacao)
            SELECT
                empresa_id,
                'STATUS_MOTORISTA_ROTA',
                id,
                %s,
                status_motorista,
                %s,
                %s,
                CONCAT('Rota atualizada por estorno da baixa do título financeiro #', %s, '. Motivo: ', %s)
            FROM rotas
            WHERE empresa_id = %s
              AND id IN ({placeholders_rota})
        """, [usuario_id, novo_status_rota, motivo_hist, titulo_id, motivo, empresa_id] + rota_ids)

        cur.execute(f"""
            UPDATE rotas
            SET status_motorista = %s,
                situacao_rota = %s
            WHERE empresa_id = %s
              AND id IN ({placeholders_rota})
        """, [novo_status_rota, nova_situacao_rota, empresa_id] + rota_ids)


# ----------------------------------------------------------
# Bloco 5.2 — Estorno de baixa financeira.
# Cria movimentação inversa e reabre ou encerra o título com histórico.
# ----------------------------------------------------------
@app.route('/financeiro/titulos/<int:id>/estornar', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro')
def estornar_baixa_titulo_financeiro(id):
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = usuario_eh_super_admin_global()

    motivo = (request.form.get('motivo_estorno') or '').strip()
    data_estorno = (request.form.get('data_estorno') or '').strip()
    destino = (request.form.get('destino_estorno') or '').strip()
    tratativa_pos_estorno = (request.form.get('tratativa_pos_estorno') or 'manter_bloqueadas').strip()
    observacao_estorno = (request.form.get('observacao_estorno') or '').strip()

    if len(motivo) < 5:
        flash("Informe um motivo de estorno com pelo menos 5 caracteres.", "warning")
        return redirect(url_for('detalhes_titulo_financeiro', id=id))

    if not data_estorno or not validar_data_iso(data_estorno):
        flash("Informe uma data de estorno válida.", "danger")
        return redirect(url_for('detalhes_titulo_financeiro', id=id))

    if destino not in ['reabrir', 'encerrar']:
        flash("Selecione o destino do título após o estorno.", "danger")
        return redirect(url_for('detalhes_titulo_financeiro', id=id))

    tratativas_validas = ['manter_bloqueadas', 'reabrir_mesmo_documento', 'exigir_nova_nf', 'cancelar_rotas']
    if tratativa_pos_estorno not in tratativas_validas:
        tratativa_pos_estorno = 'manter_bloqueadas'

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('financeiro_titulos'))

    cur = con.cursor(dictionary=True)
    try:
        query = """
            SELECT id, empresa_id, tipo_titulo, origem, origem_id, pessoa_id, numero_documento,
                   descricao, historico, valor_liquido, status_titulo, data_baixa, valor_baixado
            FROM titulos_financeiros
            WHERE id = %s
        """
        params = [id]
        if not is_super_admin:
            query += " AND empresa_id = %s"
            params.append(empresa_logada_id)
        query += " LIMIT 1"
        cur.execute(query, params)
        titulo = cur.fetchone()

        if not titulo:
            flash("Título financeiro não encontrado ou não pertence à empresa logada.", "danger")
            return redirect(url_for('financeiro_titulos'))

        parametros_financeiros = carregar_parametros_financeiros_empresa(titulo['empresa_id'], cur=cur)
        if not parametro_bool(parametros_financeiros.get('estorno.permitir_estorno_baixa', {}).get('valor')):
            flash('Estorno de baixa bloqueado pelas configurações financeiras da empresa.', 'warning')
            return redirect(url_for('detalhes_titulo_financeiro', id=id))
        if destino == 'reabrir' and not parametro_bool(parametros_financeiros.get('estorno.permitir_reabrir_titulo', {}).get('valor')):
            flash('Reabrir título após estorno está bloqueado nas configurações financeiras da empresa.', 'warning')
            return redirect(url_for('detalhes_titulo_financeiro', id=id))
        if destino == 'encerrar' and not parametro_bool(parametros_financeiros.get('estorno.permitir_encerrar_estornado', {}).get('valor')):
            flash('Encerrar título como estornado está bloqueado nas configurações financeiras da empresa.', 'warning')
            return redirect(url_for('detalhes_titulo_financeiro', id=id))
        if destino == 'encerrar' and tratativa_pos_estorno != 'manter_bloqueadas' and not parametro_bool(parametros_financeiros.get('estorno.permitir_tratativa_pos_estorno', {}).get('valor')):
            flash('Tratativa pós-estorno está bloqueada nas configurações financeiras da empresa.', 'warning')
            return redirect(url_for('detalhes_titulo_financeiro', id=id))
        if tratativa_pos_estorno == 'reabrir_mesmo_documento' and not parametro_bool(parametros_financeiros.get('documentos.permitir_reaproveitar_pos_estorno', {}).get('valor')):
            flash('Reaproveitar documento após estorno está bloqueado nas configurações financeiras da empresa.', 'warning')
            return redirect(url_for('detalhes_titulo_financeiro', id=id))

        status_atual = titulo.get('status_titulo') or ''
        if status_atual not in ['Pago', 'Recebido']:
            flash(f"Somente títulos pagos ou recebidos podem ter a baixa estornada. Status atual: {status_atual}.", "warning")
            return redirect(url_for('detalhes_titulo_financeiro', id=id))

        movimentacoes = buscar_movimentacoes_baixa_nao_estornadas(
            cur,
            titulo_id=id,
            empresa_id=titulo['empresa_id']
        )
        if not movimentacoes:
            flash("Nenhuma movimentação de baixa ativa foi encontrada para estornar este título.", "warning")
            return redirect(url_for('detalhes_titulo_financeiro', id=id))

        for mov in movimentacoes:
            tipo_inverso = 'ENTRADA' if mov.get('tipo_movimentacao') == 'SAIDA' else 'SAIDA'
            historico_estorno = f"Estorno da movimentação #{mov.get('id')} / título #{id} - {titulo.get('descricao') or titulo.get('numero_documento')}"
            obs_estorno = f"Motivo: {motivo}"
            if observacao_estorno:
                obs_estorno += f" | Observação: {observacao_estorno}"

            cur.execute("""
                INSERT INTO movimentacoes_caixa
                    (empresa_id, conta_caixa_id, titulo_financeiro_id, tipo_movimentacao,
                     data_movimentacao, valor_movimentacao, forma_pagamento, historico,
                     observacao, comprovante_url, status_movimentacao, usuario_criacao_id,
                     estorno_de_movimentacao_id, motivo_estorno)
                VALUES
                    (%s, %s, %s, %s,
                     %s, %s, %s, %s,
                     %s, NULL, 'Estorno', %s,
                     %s, %s)
            """, (
                titulo['empresa_id'],
                mov.get('conta_caixa_id'),
                id,
                tipo_inverso,
                data_estorno,
                converter_decimal(mov.get('valor_movimentacao')),
                mov.get('forma_pagamento') or titulo.get('forma_pagamento'),
                historico_estorno,
                obs_estorno,
                usuario_id,
                mov.get('id'),
                motivo
            ))

            cur.execute("""
                UPDATE movimentacoes_caixa
                SET status_movimentacao = 'Estornada',
                    motivo_estorno = %s
                WHERE id = %s
                  AND empresa_id = %s
                  AND titulo_financeiro_id = %s
            """, (motivo, mov.get('id'), titulo['empresa_id'], id))

        if destino == 'reabrir':
            novo_status = 'Solicitado' if titulo.get('origem') in ['NF_MOTORISTA', 'SEM_NF_MOTORISTA'] else 'Aberto'
        else:
            novo_status = 'Estornado'

        if destino == 'encerrar':
            tratativa_legivel_titulo = {
                'manter_bloqueadas': 'manter rotas bloqueadas para análise',
                'reabrir_mesmo_documento': 'reaproveitar o mesmo documento para nova solicitação',
                'exigir_nova_nf': 'liberar rotas exigindo novo documento/NF',
                'cancelar_rotas': 'cancelar rotas definitivamente',
            }.get(tratativa_pos_estorno, 'manter rotas bloqueadas para análise')
        else:
            tratativa_legivel_titulo = 'reabrir para nova baixa'

        obs_titulo = (
            f"Baixa estornada em {datetime.now().strftime('%d/%m/%Y %H:%M')}. "
            f"Destino: {'reaberto para nova baixa' if destino == 'reabrir' else 'encerrado como estornado'}. "
            f"Tratativa: {tratativa_legivel_titulo}. "
            f"Motivo: {motivo}"
        )
        if observacao_estorno:
            obs_titulo += f". Observação: {observacao_estorno}"

        if destino == 'reabrir':
            cur.execute("""
                UPDATE titulos_financeiros
                SET status_titulo = %s,
                    conta_caixa_baixa_id = NULL,
                    data_baixa = NULL,
                    valor_baixado = NULL,
                    usuario_baixa_id = NULL,
                    data_estorno = NOW(),
                    motivo_estorno = %s,
                    usuario_estorno_id = %s,
                    destino_estorno = %s,
                    observacao_baixa = CONCAT(
                        COALESCE(observacao_baixa, ''),
                        CASE WHEN COALESCE(observacao_baixa, '') = '' THEN '' ELSE '\n' END,
                        %s
                    ),
                    updated_at = NOW()
                WHERE id = %s
                  AND empresa_id = %s
            """, (
                novo_status,
                motivo,
                usuario_id,
                destino,
                obs_titulo,
                id,
                titulo['empresa_id']
            ))
        else:
            cur.execute("""
                UPDATE titulos_financeiros
                SET status_titulo = %s,
                    data_estorno = NOW(),
                    motivo_estorno = %s,
                    usuario_estorno_id = %s,
                    destino_estorno = %s,
                    tratativa_pos_estorno_aplicada = 1,
                    tipo_tratativa_pos_estorno = %s,
                    data_tratativa_pos_estorno = NOW(),
                    usuario_tratativa_pos_estorno_id = %s,
                    motivo_tratativa_pos_estorno = %s,
                    observacao_tratativa_pos_estorno = %s,
                    observacao_baixa = CONCAT(
                        COALESCE(observacao_baixa, ''),
                        CASE WHEN COALESCE(observacao_baixa, '') = '' THEN '' ELSE '\n' END,
                        %s
                    ),
                    updated_at = NOW()
                WHERE id = %s
                  AND empresa_id = %s
            """, (
                novo_status,
                motivo,
                usuario_id,
                destino,
                tratativa_pos_estorno,
                usuario_id,
                motivo,
                observacao_estorno or None,
                obs_titulo,
                id,
                titulo['empresa_id']
            ))

        cur.execute("""
            INSERT INTO historico_operacoes
                (empresa_id, tipo_operacao, usuario_id, status_anterior, status_novo, motivo, observacao)
            VALUES
                (%s, 'TITULO_FINANCEIRO', %s, %s, %s, 'Estorno de baixa financeira', %s)
        """, (
            titulo['empresa_id'],
            usuario_id,
            status_atual,
            novo_status,
            f"Título #{id} estornado. Destino: {destino}. Tratativa: {tratativa_pos_estorno}. Motivo: {motivo}"
        ))

        if titulo.get('origem') in ['NF_MOTORISTA', 'SEM_NF_MOTORISTA']:
            aplicar_estorno_em_documento_motorista_e_rotas(
                cur,
                titulo_id=id,
                empresa_id=titulo['empresa_id'],
                usuario_id=usuario_id,
                motivo=motivo,
                destino=destino,
                tratativa_pos_estorno=tratativa_pos_estorno
            )

        registrar_auditoria_financeira(
            cur,
            empresa_id=titulo['empresa_id'],
            usuario_id=usuario_id,
            acao='ESTORNO_BAIXA_TITULO',
            modulo='ESTORNO_FINANCEIRO',
            entidade_tipo='TITULO_FINANCEIRO',
            entidade_id=id,
            titulo_financeiro_id=id,
            pessoa_id=titulo.get('pessoa_id'),
            status_anterior=status_atual,
            status_novo=novo_status,
            valor_anterior=titulo.get('valor_baixado') or titulo.get('valor_liquido'),
            valor_novo=0,
            motivo=motivo,
            observacao=f'Título #{id} estornado. Destino: {destino}. Tratativa: {tratativa_pos_estorno}.',
            dados_depois={'destino': destino, 'tratativa_pos_estorno': tratativa_pos_estorno}
        )
        con.commit()
        if destino == 'reabrir':
            flash(f"Baixa do título #{id} estornada com sucesso. O título foi reaberto para nova baixa.", "success")
        else:
            flash(f"Baixa do título #{id} estornada com sucesso. O título foi encerrado como Estornado.", "success")
        return redirect(url_for('detalhes_titulo_financeiro', id=id))

    except Exception as e:
        try:
            con.rollback()
        except Exception as rollback_error:
            print(f"Aviso: não foi possível executar rollback do estorno do título {id}: {rollback_error}")
        print(f"Erro ao estornar baixa do título financeiro {id}: {e}")
        flash("Erro técnico ao estornar baixa financeira.", "danger")
        return redirect(url_for('detalhes_titulo_financeiro', id=id))
    finally:
        fechar_cursor_conexao(cur, con)


# ----------------------------------------------------------
# Bloco 5.2.2 — Tratativa manual pós-estorno.
# Permite decidir o destino de documentos/rotas já estornados.
# ----------------------------------------------------------
@app.route('/financeiro/titulos/<int:id>/tratativa-pos-estorno', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro')
def tratar_pos_estorno_titulo_financeiro(id):
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = usuario_eh_super_admin_global()

    tratativa = (request.form.get('tratativa_pos_estorno_manual') or '').strip()
    motivo = (request.form.get('motivo_tratativa_pos_estorno') or '').strip()
    observacao = (request.form.get('observacao_tratativa_pos_estorno') or '').strip()

    tratativas_validas = {
        'manter_bloqueadas': 'Manter rotas bloqueadas para análise',
        'reabrir_mesmo_documento': 'Reaproveitar mesmo documento para nova solicitação',
        'exigir_nova_nf': 'Liberar rotas exigindo novo documento/NF',
        'cancelar_rotas': 'Cancelar rotas definitivamente',
    }

    if tratativa not in tratativas_validas:
        flash('Selecione uma tratativa válida para o pós-estorno.', 'warning')
        return redirect(url_for('detalhes_titulo_financeiro', id=id))

    if len(motivo) < 5:
        flash('Informe um motivo com pelo menos 5 caracteres para a tratativa.', 'warning')
        return redirect(url_for('detalhes_titulo_financeiro', id=id))

    con = obter_conexao()
    if con is None:
        flash('Erro de conexão com o banco de dados.', 'danger')
        return redirect(url_for('financeiro_titulos'))

    cur = con.cursor(dictionary=True)
    try:
        query = """
            SELECT id, empresa_id, origem, origem_id, status_titulo, numero_documento, descricao,
                   tratativa_pos_estorno_aplicada, tipo_tratativa_pos_estorno,
                   data_tratativa_pos_estorno, usuario_tratativa_pos_estorno_id,
                   motivo_tratativa_pos_estorno, observacao_tratativa_pos_estorno,
                   observacao_baixa
            FROM titulos_financeiros
            WHERE id = %s
        """
        params = [id]
        if not is_super_admin:
            query += " AND empresa_id = %s"
            params.append(empresa_logada_id)
        query += " LIMIT 1"

        cur.execute(query, params)
        titulo = cur.fetchone()

        if not titulo:
            flash('Título financeiro não encontrado ou não pertence à empresa logada.', 'danger')
            return redirect(url_for('financeiro_titulos'))

        if titulo.get('status_titulo') != 'Estornado':
            flash('A tratativa pós-estorno só pode ser aplicada em títulos com status Estornado.', 'warning')
            return redirect(url_for('detalhes_titulo_financeiro', id=id))

        if titulo.get('origem') not in ['NF_MOTORISTA', 'SEM_NF_MOTORISTA']:
            flash('Este título não possui documento de motorista vinculado para tratativa pós-estorno.', 'warning')
            return redirect(url_for('detalhes_titulo_financeiro', id=id))

        tratativa_ja_aplicada = bool(titulo.get('tratativa_pos_estorno_aplicada'))
        if not tratativa_ja_aplicada and titulo.get('observacao_baixa') and 'Tratativa pós-estorno aplicada' in str(titulo.get('observacao_baixa')):
            tratativa_ja_aplicada = True

        if tratativa_ja_aplicada:
            flash('Este título já possui tratativa pós-estorno aplicada. A decisão é final e não pode ser alterada por esta tela.', 'warning')
            return redirect(url_for('detalhes_titulo_financeiro', id=id))

        parametros_financeiros = carregar_parametros_financeiros_empresa(titulo['empresa_id'], cur=cur)
        if not parametro_bool(parametros_financeiros.get('estorno.permitir_tratativa_pos_estorno', {}).get('valor')):
            flash('Tratativa pós-estorno está bloqueada pelas configurações financeiras da empresa.', 'warning')
            return redirect(url_for('detalhes_titulo_financeiro', id=id))
        if tratativa == 'reabrir_mesmo_documento' and not parametro_bool(parametros_financeiros.get('documentos.permitir_reaproveitar_pos_estorno', {}).get('valor')):
            flash('Reaproveitar documento após estorno está bloqueado nas configurações financeiras da empresa.', 'warning')
            return redirect(url_for('detalhes_titulo_financeiro', id=id))

        texto_tratativa = tratativas_validas.get(tratativa, tratativa)
        obs_titulo = (
            f"Tratativa pós-estorno aplicada em {datetime.now().strftime('%d/%m/%Y %H:%M')}. "
            f"Tratativa: {texto_tratativa}. Motivo: {motivo}"
        )
        if observacao:
            obs_titulo += f". Observação: {observacao}"

        # Reutiliza a mesma regra central do estorno definitivo, mas sem gerar nova movimentação de caixa.
        aplicar_estorno_em_documento_motorista_e_rotas(
            cur,
            titulo_id=id,
            empresa_id=titulo['empresa_id'],
            usuario_id=usuario_id,
            motivo=motivo,
            destino='encerrar',
            tratativa_pos_estorno=tratativa
        )

        cur.execute("""
            UPDATE titulos_financeiros
            SET tratativa_pos_estorno_aplicada = 1,
                tipo_tratativa_pos_estorno = %s,
                data_tratativa_pos_estorno = NOW(),
                usuario_tratativa_pos_estorno_id = %s,
                motivo_tratativa_pos_estorno = %s,
                observacao_tratativa_pos_estorno = %s,
                observacao_baixa = CONCAT(
                    COALESCE(observacao_baixa, ''),
                    CASE WHEN COALESCE(observacao_baixa, '') = '' THEN '' ELSE '
' END,
                    %s
                ),
                updated_at = NOW()
            WHERE id = %s
              AND empresa_id = %s
              AND COALESCE(tratativa_pos_estorno_aplicada, 0) = 0
        """, (tratativa, usuario_id, motivo, observacao or None, obs_titulo, id, titulo['empresa_id']))

        if cur.rowcount == 0:
            con.rollback()
            flash('Este título já recebeu uma tratativa pós-estorno. A decisão anterior foi preservada.', 'warning')
            return redirect(url_for('detalhes_titulo_financeiro', id=id))

        cur.execute("""
            INSERT INTO historico_operacoes
                (empresa_id, tipo_operacao, usuario_id, status_anterior, status_novo, motivo, observacao)
            VALUES
                (%s, 'TRATATIVA_POS_ESTORNO', %s, 'Estornado', 'Estornado', %s, %s)
        """, (
            titulo['empresa_id'],
            usuario_id,
            motivo,
            f"Título #{id}. {obs_titulo}"
        ))

        registrar_auditoria_financeira(
            cur,
            empresa_id=titulo['empresa_id'],
            usuario_id=usuario_id,
            acao='TRATATIVA_POS_ESTORNO_APLICADA',
            modulo='ESTORNO_FINANCEIRO',
            entidade_tipo='TITULO_FINANCEIRO',
            entidade_id=id,
            titulo_financeiro_id=id,
            status_anterior='Estornado',
            status_novo='Estornado',
            motivo=motivo,
            observacao=f'Título #{id}. {obs_titulo}',
            dados_depois={'tratativa': tratativa, 'texto_tratativa': texto_tratativa}
        )
        con.commit()
        flash(f'Tratativa pós-estorno aplicada com sucesso: {texto_tratativa}.', 'success')
        return redirect(url_for('detalhes_titulo_financeiro', id=id))

    except Exception as e:
        try:
            con.rollback()
        except Exception as rollback_error:
            print(f"Aviso: não foi possível executar rollback da tratativa pós-estorno do título {id}: {rollback_error}")
        print(f"Erro na tratativa pós-estorno do título financeiro {id}: {e}")
        flash('Erro técnico ao aplicar tratativa pós-estorno.', 'danger')
        return redirect(url_for('detalhes_titulo_financeiro', id=id))
    finally:
        fechar_cursor_conexao(cur, con)



@app.route('/financeiro/movimentacoes-caixa', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro', 'Consulta')
def financeiro_movimentacoes_caixa():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    conta_caixa_id = (request.args.get('conta_caixa_id') or '').strip()
    tipo_movimentacao = (request.args.get('tipo_movimentacao') or '').strip()
    data_inicio = (request.args.get('data_inicio') or '').strip()
    data_fim = (request.args.get('data_fim') or '').strip()
    pesquisa = (request.args.get('pesquisa') or '').strip()
    empresa_id_filtro = (request.args.get('empresa_id') or '').strip()

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('financeiro_titulos'))

    cur = con.cursor(dictionary=True)
    try:
        query = """
            SELECT m.*,
                   cx.nome_conta AS conta_caixa_nome,
                   t.numero_documento,
                   t.descricao AS titulo_descricao,
                   t.tipo_titulo,
                   p.nome_completo AS pessoa_nome,
                   e.nome_fantasia AS empresa_nome,
                   e.razao_social AS empresa_razao_social,
                   u.login AS usuario_login
            FROM movimentacoes_caixa m
            INNER JOIN contas_caixa cx ON cx.id = m.conta_caixa_id AND cx.empresa_id = m.empresa_id
            LEFT JOIN titulos_financeiros t ON t.id = m.titulo_financeiro_id AND t.empresa_id = m.empresa_id
            LEFT JOIN pessoas p ON p.id = t.pessoa_id AND p.empresa_id = t.empresa_id
            LEFT JOIN empresas e ON e.id = m.empresa_id
            LEFT JOIN usuarios u ON u.id = m.usuario_criacao_id
            WHERE 1 = 1
        """
        params = []

        if is_super_admin:
            if empresa_id_filtro and empresa_id_filtro.isdigit():
                query += " AND m.empresa_id = %s"
                params.append(int(empresa_id_filtro))
        else:
            query += " AND m.empresa_id = %s"
            params.append(empresa_logada_id)

        if conta_caixa_id and conta_caixa_id.isdigit():
            query += " AND m.conta_caixa_id = %s"
            params.append(int(conta_caixa_id))

        if tipo_movimentacao in ['ENTRADA', 'SAIDA']:
            query += " AND m.tipo_movimentacao = %s"
            params.append(tipo_movimentacao)

        if data_inicio:
            query += " AND m.data_movimentacao >= %s"
            params.append(data_inicio)

        if data_fim:
            query += " AND m.data_movimentacao <= %s"
            params.append(data_fim)

        if pesquisa:
            query += """
                AND (
                    m.historico LIKE %s
                    OR m.observacao LIKE %s
                    OR t.numero_documento LIKE %s
                    OR t.descricao LIKE %s
                    OR p.nome_completo LIKE %s
                )
            """
            termo = f"%{pesquisa}%"
            params.extend([termo, termo, termo, termo, termo])

        query += " ORDER BY m.data_movimentacao DESC, m.id DESC"
        cur.execute(query, params)
        movimentacoes = cur.fetchall()

        resumo = {
            'entradas': Decimal('0.00'),
            'saidas': Decimal('0.00'),
            'estornos': Decimal('0.00'),
            'saldo_movimentado': Decimal('0.00'),
            'total': len(movimentacoes)
        }
        for mov in movimentacoes:
            valor = converter_decimal(mov.get('valor_movimentacao'))
            status_mov = str(mov.get('status_movimentacao') or 'Ativa')
            eh_estorno = bool(mov.get('estorno_de_movimentacao_id')) or status_mov == 'Estorno'

            # Entradas/Saídas gerenciais contam apenas movimentações operacionais ativas.
            # Estornos ficam em card separado para não parecer entrada operacional real.
            if eh_estorno:
                resumo['estornos'] += valor
            elif status_mov != 'Estornada':
                if mov.get('tipo_movimentacao') == 'ENTRADA':
                    resumo['entradas'] += valor
                elif mov.get('tipo_movimentacao') == 'SAIDA':
                    resumo['saidas'] += valor

            # Saldo movimentado é contábil: considera saída original e movimentação inversa.
            if mov.get('tipo_movimentacao') == 'ENTRADA':
                resumo['saldo_movimentado'] += valor
            elif mov.get('tipo_movimentacao') == 'SAIDA':
                resumo['saldo_movimentado'] -= valor

        contas_caixa = carregar_contas_caixa_financeiro(empresa_logada_id, is_super_admin, somente_ativas=False)
        empresas = carregar_empresas_ativas() if is_super_admin else []

    except Exception as e:
        print(f"Erro ao carregar movimentações de caixa: {e}")
        flash(f"Erro técnico ao carregar movimentações de caixa: {e}", "danger")
        movimentacoes = []
        resumo = {'entradas': 0, 'saidas': 0, 'saldo_movimentado': 0, 'total': 0}
        contas_caixa = []
        empresas = []
    finally:
        fechar_cursor_conexao(cur, con)

    filtros = {
        'conta_caixa_id': conta_caixa_id,
        'tipo_movimentacao': tipo_movimentacao,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'pesquisa': pesquisa,
        'empresa_id': empresa_id_filtro
    }

    return render_template(
        'financeiro_movimentacoes_caixa.html',
        usuario_logado=usuario_logado,
        movimentacoes=movimentacoes,
        resumo=resumo,
        contas_caixa=contas_caixa,
        empresas=empresas,
        filtros=filtros,
        is_super_admin=is_super_admin
    )



@app.route('/financeiro/conciliacao-caixa', methods=['GET'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro', 'Consulta')
def financeiro_conciliacao_caixa():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    conta_caixa_id = (request.args.get('conta_caixa_id') or '').strip()
    status_conciliacao = (request.args.get('status_conciliacao') or '').strip()
    status_movimentacao = (request.args.get('status_movimentacao') or '').strip()
    tipo_movimentacao = (request.args.get('tipo_movimentacao') or '').strip()
    data_inicio = (request.args.get('data_inicio') or '').strip()
    data_fim = (request.args.get('data_fim') or '').strip()
    pesquisa = (request.args.get('pesquisa') or '').strip()
    empresa_id_filtro = (request.args.get('empresa_id') or '').strip()

    hoje_data = date.today()
    if not data_inicio:
        data_inicio = hoje_data.replace(day=1).strftime('%Y-%m-%d')
    if not data_fim:
        data_fim = hoje_data.strftime('%Y-%m-%d')

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('financeiro_dashboard'))

    cur = con.cursor(dictionary=True)
    try:
        query = """
            SELECT m.*,
                   COALESCE(NULLIF(m.status_conciliacao, ''), 'Pendente') AS status_conciliacao_view,
                   COALESCE(NULLIF(m.status_movimentacao, ''), 'Ativa') AS status_movimentacao_view,
                   cx.nome_conta AS conta_caixa_nome,
                   t.numero_documento,
                   t.descricao AS titulo_descricao,
                   t.tipo_titulo,
                   p.nome_completo AS pessoa_nome,
                   p.cpf_cnpj AS pessoa_cpf_cnpj,
                   e.nome_fantasia AS empresa_nome,
                   e.razao_social AS empresa_razao_social,
                   u.login AS usuario_login,
                   uc.login AS usuario_conciliacao_login
            FROM movimentacoes_caixa m
            INNER JOIN contas_caixa cx ON cx.id = m.conta_caixa_id AND cx.empresa_id = m.empresa_id
            LEFT JOIN titulos_financeiros t ON t.id = m.titulo_financeiro_id AND t.empresa_id = m.empresa_id
            LEFT JOIN pessoas p ON p.id = t.pessoa_id AND p.empresa_id = t.empresa_id
            LEFT JOIN empresas e ON e.id = m.empresa_id
            LEFT JOIN usuarios u ON u.id = m.usuario_criacao_id
            LEFT JOIN usuarios uc ON uc.id = m.usuario_conciliacao_id
            WHERE 1 = 1
        """
        params = []

        if is_super_admin:
            if empresa_id_filtro and empresa_id_filtro.isdigit():
                query += " AND m.empresa_id = %s"
                params.append(int(empresa_id_filtro))
        else:
            query += " AND m.empresa_id = %s"
            params.append(empresa_logada_id)

        if conta_caixa_id and conta_caixa_id.isdigit():
            query += " AND m.conta_caixa_id = %s"
            params.append(int(conta_caixa_id))

        if tipo_movimentacao in ['ENTRADA', 'SAIDA']:
            query += " AND m.tipo_movimentacao = %s"
            params.append(tipo_movimentacao)

        if status_movimentacao in ['Ativa', 'Estornada', 'Estorno']:
            query += " AND COALESCE(NULLIF(m.status_movimentacao, ''), 'Ativa') = %s"
            params.append(status_movimentacao)

        if status_conciliacao in ['Pendente', 'Conciliada', 'Divergente', 'Nao conciliavel']:
            query += " AND COALESCE(NULLIF(m.status_conciliacao, ''), 'Pendente') = %s"
            params.append(status_conciliacao)

        if data_inicio:
            query += " AND m.data_movimentacao >= %s"
            params.append(data_inicio)

        if data_fim:
            query += " AND m.data_movimentacao <= %s"
            params.append(data_fim)

        if pesquisa:
            query += """
                AND (
                    m.historico LIKE %s
                    OR m.observacao LIKE %s
                    OR m.observacao_conciliacao LIKE %s
                    OR t.numero_documento LIKE %s
                    OR t.descricao LIKE %s
                    OR p.nome_completo LIKE %s
                    OR p.cpf_cnpj LIKE %s
                    OR CAST(m.id AS CHAR) LIKE %s
                    OR CAST(t.id AS CHAR) LIKE %s
                )
            """
            termo = f"%{pesquisa}%"
            params.extend([termo, termo, termo, termo, termo, termo, termo, termo, termo])

        query += " ORDER BY m.data_movimentacao DESC, m.id DESC"
        cur.execute(query, params)
        movimentacoes = cur.fetchall()

        resumo = {
            'registros': len(movimentacoes),
            'pendentes_qtd': 0,
            'pendentes_reais_qtd': 0,
            'pendentes_nao_operacionais_qtd': 0,
            'conciliadas_qtd': 0,
            'divergentes_qtd': 0,
            'nao_conciliaveis_qtd': 0,
            'saldo_pendente': Decimal('0.00'),
            'saldo_pendente_real': Decimal('0.00'),
            'saldo_conciliado': Decimal('0.00'),
            'saldo_divergente': Decimal('0.00'),
            'saldo_nao_conciliavel': Decimal('0.00'),
            'entradas_conciliadas': Decimal('0.00'),
            'saidas_conciliadas': Decimal('0.00'),
            'entradas_nao_conciliaveis': Decimal('0.00'),
            'saidas_nao_conciliaveis': Decimal('0.00'),
        }

        for mov in movimentacoes:
            valor = converter_decimal(mov.get('valor_movimentacao'))
            sinal = valor if mov.get('tipo_movimentacao') == 'ENTRADA' else -valor
            stc = str(mov.get('status_conciliacao_view') or 'Pendente')
            stm = str(mov.get('status_movimentacao_view') or 'Ativa')

            mov['conciliacao_bancaria_real'] = (stm == 'Ativa')

            if stc == 'Conciliada':
                resumo['conciliadas_qtd'] += 1
                resumo['saldo_conciliado'] += sinal
                if mov.get('tipo_movimentacao') == 'ENTRADA':
                    resumo['entradas_conciliadas'] += valor
                else:
                    resumo['saidas_conciliadas'] += valor
            elif stc == 'Divergente':
                resumo['divergentes_qtd'] += 1
                resumo['saldo_divergente'] += sinal
            elif stc == 'Nao conciliavel':
                resumo['nao_conciliaveis_qtd'] += 1
                resumo['saldo_nao_conciliavel'] += sinal
                if mov.get('tipo_movimentacao') == 'ENTRADA':
                    resumo['entradas_nao_conciliaveis'] += valor
                else:
                    resumo['saidas_nao_conciliaveis'] += valor
            else:
                resumo['pendentes_qtd'] += 1
                resumo['saldo_pendente'] += sinal
                if stm == 'Ativa':
                    resumo['pendentes_reais_qtd'] += 1
                    resumo['saldo_pendente_real'] += sinal
                else:
                    resumo['pendentes_nao_operacionais_qtd'] += 1

        contas_caixa = carregar_contas_caixa_financeiro(empresa_logada_id, is_super_admin, somente_ativas=False)
        empresas = carregar_empresas_ativas() if is_super_admin else []
    except Exception as e:
        print(f"Erro ao carregar conciliação de caixa: {e}")
        flash(f"Erro técnico ao carregar conciliação de caixa: {e}", "danger")
        return redirect(url_for('financeiro_dashboard'))
    finally:
        fechar_cursor_conexao(cur, con)

    filtros = {
        'conta_caixa_id': conta_caixa_id,
        'status_conciliacao': status_conciliacao,
        'status_movimentacao': status_movimentacao,
        'tipo_movimentacao': tipo_movimentacao,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'pesquisa': pesquisa,
        'empresa_id': empresa_id_filtro
    }

    return render_template(
        'financeiro_conciliacao_caixa.html',
        usuario_logado=usuario_logado,
        movimentacoes=movimentacoes,
        resumo=resumo,
        contas_caixa=contas_caixa,
        empresas=empresas,
        filtros=filtros,
        is_super_admin=is_super_admin
    )


@app.route('/financeiro/conciliacao-caixa/acao', methods=['POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro')
def financeiro_conciliacao_caixa_acao():
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = usuario_eh_super_admin_global()

    ids = request.form.getlist('movimentacao_ids')
    acao = (request.form.get('acao') or '').strip()
    observacao = (request.form.get('observacao_conciliacao') or '').strip()

    conta_caixa_id = request.form.get('filtro_conta_caixa_id') or ''
    status_conciliacao = request.form.get('filtro_status_conciliacao') or ''
    status_movimentacao = request.form.get('filtro_status_movimentacao') or ''
    tipo_movimentacao = request.form.get('filtro_tipo_movimentacao') or ''
    data_inicio = request.form.get('filtro_data_inicio') or ''
    data_fim = request.form.get('filtro_data_fim') or ''
    pesquisa = request.form.get('filtro_pesquisa') or ''
    empresa_id_filtro = request.form.get('filtro_empresa_id') or ''

    redirect_params = {
        'conta_caixa_id': conta_caixa_id,
        'status_conciliacao': status_conciliacao,
        'status_movimentacao': status_movimentacao,
        'tipo_movimentacao': tipo_movimentacao,
        'data_inicio': data_inicio,
        'data_fim': data_fim,
        'pesquisa': pesquisa,
        'empresa_id': empresa_id_filtro
    }
    redirect_params = {k: v for k, v in redirect_params.items() if v}

    ids_validos = []
    for item in ids:
        try:
            ids_validos.append(int(item))
        except Exception:
            pass

    if not ids_validos:
        flash("Selecione pelo menos uma movimentação para conciliar.", "warning")
        return redirect(url_for('financeiro_conciliacao_caixa', **redirect_params))

    mapa_status = {
        'conciliar': 'Conciliada',
        'divergente': 'Divergente',
        'pendente': 'Pendente',
        'nao_conciliavel': 'Nao conciliavel'
    }
    novo_status = mapa_status.get(acao)
    if not novo_status:
        flash("Ação de conciliação inválida.", "danger")
        return redirect(url_for('financeiro_conciliacao_caixa', **redirect_params))

    if acao in ['divergente', 'nao_conciliavel'] and not observacao:
        flash("Informe uma observação para marcar movimentação como divergente ou não conciliável.", "warning")
        return redirect(url_for('financeiro_conciliacao_caixa', **redirect_params))

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('financeiro_conciliacao_caixa', **redirect_params))

    cur = con.cursor(dictionary=True)
    try:
        placeholders = ','.join(['%s'] * len(ids_validos))
        params = list(ids_validos)
        filtro_empresa_sql = ""
        if not is_super_admin:
            filtro_empresa_sql = " AND empresa_id = %s"
            params.append(empresa_logada_id)
        elif empresa_id_filtro and str(empresa_id_filtro).isdigit():
            filtro_empresa_sql = " AND empresa_id = %s"
            params.append(int(empresa_id_filtro))

        if novo_status == 'Pendente':
            sql = f"""
                UPDATE movimentacoes_caixa
                SET status_conciliacao = 'Pendente',
                    data_conciliacao = NULL,
                    usuario_conciliacao_id = NULL,
                    observacao_conciliacao = %s
                WHERE id IN ({placeholders}) {filtro_empresa_sql}
            """
            params_update = [observacao or None] + params
        else:
            sql = f"""
                UPDATE movimentacoes_caixa
                SET status_conciliacao = %s,
                    data_conciliacao = NOW(),
                    usuario_conciliacao_id = %s,
                    observacao_conciliacao = %s
                WHERE id IN ({placeholders}) {filtro_empresa_sql}
            """
            params_update = [novo_status, usuario_id, observacao or None] + params

        cur.execute(sql, params_update)
        afetadas = cur.rowcount
        registrar_auditoria_financeira(
            cur,
            empresa_id=int(empresa_id_filtro) if is_super_admin and empresa_id_filtro and str(empresa_id_filtro).isdigit() else empresa_logada_id,
            usuario_id=usuario_id,
            acao='CONCILIACAO_CAIXA_ATUALIZADA',
            modulo='CONCILIACAO_CAIXA',
            entidade_tipo='MOVIMENTACOES_CAIXA',
            entidade_id=None,
            status_novo=novo_status,
            motivo=f'Ação de conciliação: {acao}',
            observacao=observacao or f'{afetadas} movimentação(ões) atualizada(s).',
            dados_depois={'ids': ids_validos, 'acao': acao, 'novo_status': novo_status, 'quantidade': afetadas}
        )
        con.commit()

        label = 'Não conciliável' if novo_status == 'Nao conciliavel' else novo_status
        flash(f"{afetadas} movimentação(ões) atualizada(s) para {label}.", "success")
    except Exception as e:
        try:
            con.rollback()
        except Exception:
            pass
        print(f"Erro ao aplicar conciliação de caixa: {e}")
        flash(f"Erro técnico ao aplicar conciliação: {e}", "danger")
    finally:
        fechar_cursor_conexao(cur, con)

    return redirect(url_for('financeiro_conciliacao_caixa', **redirect_params))

@app.route('/financeiro/contas-caixa/nova', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro')
def nova_conta_caixa():
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    usuario_id = session.get('usuario_id')
    is_super_admin = usuario_eh_super_admin_global()

    if request.method == 'POST':
        empresa_id = request.form.get('empresa_id') if is_super_admin else empresa_logada_id
        nome_conta = (request.form.get('nome_conta') or '').strip()
        tipo_conta = (request.form.get('tipo_conta') or '').strip()
        banco = (request.form.get('banco') or '').strip()
        agencia = (request.form.get('agencia') or '').strip()
        numero_conta = (request.form.get('numero_conta') or '').strip()
        saldo_inicial = converter_decimal(request.form.get('saldo_inicial'))
        observacao = (request.form.get('observacao') or '').strip()

        if not empresa_id or not str(empresa_id).isdigit():
            flash("Selecione uma empresa válida.", "danger")
            return redirect(url_for('nova_conta_caixa'))
        empresa_id = int(empresa_id)

        if not nome_conta:
            flash("Informe o nome da conta caixa.", "danger")
            return redirect(url_for('nova_conta_caixa'))

        if tipo_conta not in financeiro_base_tipos_conta_caixa():
            flash("Selecione um tipo de conta válido.", "danger")
            return redirect(url_for('nova_conta_caixa'))

        con = obter_conexao()
        if con is None:
            flash("Erro de conexão com o banco de dados.", "danger")
            return redirect(url_for('financeiro.financeiro_contas_caixa'))

        cur = con.cursor(dictionary=True)
        try:
            cur.execute("SELECT id FROM empresas WHERE id = %s AND status_empresa = 'Ativa' LIMIT 1", (empresa_id,))
            if not cur.fetchone():
                flash("Empresa inválida ou inativa.", "danger")
                return redirect(url_for('nova_conta_caixa'))

            cur.execute("""
                INSERT INTO contas_caixa
                    (empresa_id, nome_conta, tipo_conta, banco, agencia, numero_conta,
                     saldo_inicial, status_conta, observacao, usuario_criacao_id)
                VALUES
                    (%s, %s, %s, %s, %s, %s,
                     %s, 'Ativa', %s, %s)
            """, (
                empresa_id,
                nome_conta,
                tipo_conta,
                banco or None,
                agencia or None,
                numero_conta or None,
                saldo_inicial,
                observacao or None,
                usuario_id
            ))
            con.commit()
            flash("Conta caixa criada com sucesso.", "success")
            return redirect(url_for('financeiro.financeiro_contas_caixa'))
        except Exception as e:
            con.rollback()
            print(f"Erro ao criar conta caixa: {e}")
            flash(f"Erro técnico ao criar conta caixa: {e}", "danger")
            return redirect(url_for('nova_conta_caixa'))
        finally:
            fechar_cursor_conexao(cur, con)

    empresas = carregar_empresas_ativas() if is_super_admin else []
    return render_template(
        'financeiro_conta_caixa_form.html',
        usuario_logado=usuario_logado,
        conta=None,
        empresas=empresas,
        tipos_conta=financeiro_base_tipos_conta_caixa(),
        is_super_admin=is_super_admin
    )


@app.route('/financeiro/contas-caixa/<int:id>/editar', methods=['GET', 'POST'])
@login_required
@perfis_permitidos('Administrador', 'Operacional', 'Financeiro')
def editar_conta_caixa(id):
    usuario_logado = session.get('usuario_nome', 'Usuário')
    empresa_logada_id = session.get('empresa_id')
    is_super_admin = usuario_eh_super_admin_global()

    con = obter_conexao()
    if con is None:
        flash("Erro de conexão com o banco de dados.", "danger")
        return redirect(url_for('financeiro.financeiro_contas_caixa'))

    cur = con.cursor(dictionary=True)
    try:
        query = "SELECT * FROM contas_caixa WHERE id = %s"
        params = [id]
        if not is_super_admin:
            query += " AND empresa_id = %s"
            params.append(empresa_logada_id)
        query += " LIMIT 1"
        cur.execute(query, params)
        conta = cur.fetchone()

        if not conta:
            flash("Conta caixa não encontrada ou não pertence à empresa logada.", "danger")
            return redirect(url_for('financeiro.financeiro_contas_caixa'))

        if request.method == 'POST':
            nome_conta = (request.form.get('nome_conta') or '').strip()
            tipo_conta = (request.form.get('tipo_conta') or '').strip()
            banco = (request.form.get('banco') or '').strip()
            agencia = (request.form.get('agencia') or '').strip()
            numero_conta = (request.form.get('numero_conta') or '').strip()
            saldo_inicial = converter_decimal(request.form.get('saldo_inicial'))
            status_conta = (request.form.get('status_conta') or 'Ativa').strip()
            observacao = (request.form.get('observacao') or '').strip()

            if not nome_conta:
                flash("Informe o nome da conta caixa.", "danger")
                return redirect(url_for('editar_conta_caixa', id=id))

            if tipo_conta not in financeiro_base_tipos_conta_caixa():
                flash("Selecione um tipo de conta válido.", "danger")
                return redirect(url_for('editar_conta_caixa', id=id))

            if status_conta not in ['Ativa', 'Inativa']:
                flash("Status da conta inválido.", "danger")
                return redirect(url_for('editar_conta_caixa', id=id))

            cur.execute("""
                UPDATE contas_caixa
                SET nome_conta = %s,
                    tipo_conta = %s,
                    banco = %s,
                    agencia = %s,
                    numero_conta = %s,
                    saldo_inicial = %s,
                    status_conta = %s,
                    observacao = %s,
                    updated_at = NOW()
                WHERE id = %s
                  AND empresa_id = %s
            """, (
                nome_conta,
                tipo_conta,
                banco or None,
                agencia or None,
                numero_conta or None,
                saldo_inicial,
                status_conta,
                observacao or None,
                id,
                conta['empresa_id']
            ))
            con.commit()
            flash("Conta caixa atualizada com sucesso.", "success")
            return redirect(url_for('financeiro.financeiro_contas_caixa'))

    except Exception as e:
        con.rollback()
        print(f"Erro ao editar conta caixa: {e}")
        flash(f"Erro técnico ao editar conta caixa: {e}", "danger")
        return redirect(url_for('financeiro.financeiro_contas_caixa'))
    finally:
        fechar_cursor_conexao(cur, con)

    empresas = carregar_empresas_ativas() if is_super_admin else []
    return render_template(
        'financeiro_conta_caixa_form.html',
        usuario_logado=usuario_logado,
        conta=conta,
        empresas=empresas,
        tipos_conta=financeiro_base_tipos_conta_caixa(),
        is_super_admin=is_super_admin
    )


# ==========================================================
# BLUEPRINT PILOTO — FINANCEIRO / CONTAS CAIXA
# ==========================================================
from app_modules.financeiro import criar_financeiro_blueprint

financeiro_services = {
    "login_required": login_required,
    "perfis_permitidos": perfis_permitidos,
    "usuario_eh_super_admin_global": usuario_eh_super_admin_global,
    "carregar_contas_caixa_financeiro": carregar_contas_caixa_financeiro,
    "obter_conexao": obter_conexao,
    "calcular_saldo_conta_caixa": calcular_saldo_conta_caixa,
    "converter_decimal": converter_decimal,
    "fechar_cursor_conexao": fechar_cursor_conexao,
}

app.extensions["financeiro_services"] = financeiro_services
app.register_blueprint(criar_financeiro_blueprint(financeiro_services))


if __name__ == '__main__':
    app.run(debug=False, port=8080)
